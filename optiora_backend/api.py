"""
REST API routes for OptiOra.

Handles:
- Credential validation/storage
- Scanning permissions and progress
- Dashboard data endpoints
"""

import csv
import asyncio
import io
import json
import logging
import os
import re
import configparser
import tempfile
import base64
import binascii
import hashlib
import hmac
import shutil
import subprocess
import time
from urllib.parse import quote
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace
from typing import Any, Dict, List, Literal, Optional, Union
from uuid import uuid4
from zoneinfo import ZoneInfo
from xml.sax.saxutils import escape as xml_escape

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from .auth_routes import get_current_membership, get_current_user
from .credentials import CredentialManager, CredentialStatus, CredentialValidator
from .config import Config
from .notifications import evaluate_budget_alert
from .notifications import (
    SUPPORTED_NOTIFICATION_CHANNELS,
    destination_configured,
    send_test_notification,
)
from .cost_context import (
    LiveDataPolicyError,
    build_imported_cost_context,
    build_live_cost_context,
    fetch_provider_cost_summary,
)
from .external_anomalies import aws_anomaly_severity, coerce_aws_anomaly_impact_usd, derive_aws_anomaly_alert
from .imported_cost_csv import CsvImportError, load_normalized_csv_upload, validate_cost_csv_row
from .imported_costs import query_imported_cost_rows, summarize_imported_cost_rows
from .provider_support import (
    AWSCredentialInput,
    AzureCredentialInput,
    CredentialInput,
    GCPCredentialInput,
    OCICredentialInput,
    SUPPORTED_CLOUD_PROVIDERS,
    SUPPORTED_COST_IMPORT_PROVIDERS,
    parse_credential_payload,
    provider_api_capabilities,
    provider_bounded_limit,
    provider_diagnostic_requirements,
    run_credential_validation,
)
from .scheduler_status import compute_next_run, scan_interval_seconds, scheduler_runtime_snapshot
from .access_control import require_role
from .connectors import ConnectorManager, ConnectorType, BaseConnector, ConnectorStatus
from .orm_models import (
    AlertEvent,
    AlertOpsPolicy,
    AlertRoutingPolicy,
    AuditLog,
    BusinessMappingRule,
    CostAllocationSnapshot,
    CostPeriodSummary,
    CostSnapshot,
    CredentialRecord,
    ExportJob,
    ExportJobRun,
    ImportedCostRecord,
    NormalizedCostDimension,
    Organization,
    ProviderAccount,
    ProviderAccountLink,
    ProviderAccountSnapshot,
    RecommendationLedger,
    ScanRunRecord,
    ScanningPermissionRecord,
    SessionLocal,
    User,
    UserOrganization,
    UserRole,
    VirtualTagRule,
    get_db,
)
from .scanning import ScanningManager, ScanningState
from .tools import anomalies, aws_costs, finops_analytics, recommendations
from .tools import azure_costs, gcp_costs, oci_costs, genai_advisor, finops_rag
from .retention import fetch_archived_period_summaries
from .kubernetes_catalog import build_kubernetes_provider_catalog
from . import __version__

logger = logging.getLogger(__name__)
_scheduler_running = False


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)

router = APIRouter(prefix="/api/v1", tags=["api"])
_SUPPORTED_TREND_VIEWS = {"provider", "region", "service", "account"}


def _infer_parent_account_type(provider_key: str, child_account_type: str) -> str:
    normalized_child = str(child_account_type or "").strip().lower().replace(" ", "_")
    if provider_key == "aws":
        return "organization"
    if provider_key == "azure":
        return "management_group"
    if provider_key == "gcp":
        return "folder" if normalized_child == "project" else "organization"
    if provider_key == "oci":
        return "tenancy" if normalized_child == "compartment" else "group"
    return "group"


def _share_hmac_secret() -> bytes:
    raw = os.getenv("SECRET_KEY", "optiora-dev-share-secret")
    return str(raw).encode("utf-8")


def _build_report_share_token(payload: Dict[str, Any]) -> str:
    body = base64.urlsafe_b64encode(json.dumps(payload, separators=(",", ":")).encode("utf-8")).decode("utf-8").rstrip("=")
    signature = hmac.new(_share_hmac_secret(), body.encode("utf-8"), hashlib.sha256).digest()
    sig = base64.urlsafe_b64encode(signature).decode("utf-8").rstrip("=")
    return f"{body}.{sig}"


def _parse_report_share_token(token: str) -> Dict[str, Any]:
    if "." not in token:
        raise HTTPException(status_code=401, detail="Invalid share token")
    body, sig = token.split(".", 1)
    expected_sig = base64.urlsafe_b64encode(
        hmac.new(_share_hmac_secret(), body.encode("utf-8"), hashlib.sha256).digest()
    ).decode("utf-8").rstrip("=")
    if not hmac.compare_digest(sig, expected_sig):
        raise HTTPException(status_code=401, detail="Invalid share token")
    padded = body + "=" * ((4 - len(body) % 4) % 4)
    try:
        payload = json.loads(base64.urlsafe_b64decode(padded.encode("utf-8")).decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=401, detail="Malformed share token") from exc
    exp = int(payload.get("exp", 0) or 0)
    if exp <= int(datetime.now(timezone.utc).timestamp()):
        raise HTTPException(status_code=401, detail="Share token expired")
    return payload


def _trend_dimension_value(rec: ImportedCostRecord, view_by: str) -> str:
    if view_by == "region":
        return (rec.region or "unknown").strip() or "unknown"
    if view_by == "service":
        return (rec.service_name or "unknown").strip() or "unknown"
    if view_by == "account":
        return (
            rec.account_identifier
            or rec.account_name
            or rec.parent_account_identifier
            or "unknown"
        ).strip() or "unknown"
    return (rec.provider or "imported").strip().lower() or "imported"


def _rag_context_for_analysis(
    *,
    analysis_type: str,
    cloud_provider: str,
    context: Optional[Dict[str, Any]] = None,
    top_k: int = 4,
) -> Dict[str, Any]:
    payload = finops_rag.retrieve_guidance(
        analysis_type=analysis_type,
        cloud_provider=cloud_provider,
        context=context or {},
        top_k=top_k,
    )
    if context is not None:
        context["rag_brief"] = payload.get("rag_brief", "")
    return payload


class CredentialResponse(BaseModel):
    provider: str
    is_valid: bool
    message: str
    test_cost_usd: Optional[float] = None
    tested_at: Optional[str] = None
    error_details: Optional[str] = None


class ScanningApprovalRequest(BaseModel):
    customer_id: Optional[str] = None
    auto_remediate: bool = False
    scan_frequency: str = "daily"
    notification_email: Optional[str] = None
    monthly_budget_usd: float = 0.0
    warning_threshold_percent: float = 80.0
    critical_threshold_percent: float = 100.0
    notifications_enabled: bool = True
    scheduler_override_enabled: bool = False
    scheduler_override_frequency: Optional[Literal["hourly", "daily", "weekly"]] = None
    scheduler_retry_max_attempts: int = 2
    scheduler_retry_backoff_seconds: int = 120
    scheduler_overdue_alert_hours: int = 24


class ScanningPermissionResponse(BaseModel):
    customer_id: str
    organization_id: int
    state: str
    providers: List[str]
    scan_frequency: str
    auto_remediate: bool
    notification_email: Optional[str] = None
    monthly_budget_usd: float = 0.0
    warning_threshold_percent: float = 80.0
    critical_threshold_percent: float = 100.0
    notifications_enabled: bool = True
    scheduler_override_enabled: bool = False
    scheduler_override_frequency: Optional[str] = None
    scheduler_retry_max_attempts: int = 2
    scheduler_retry_backoff_seconds: int = 120
    scheduler_overdue_alert_hours: int = 24
    created_at: str
    approved_at: Optional[str] = None


class StartScanRequest(BaseModel):
    customer_id: Optional[str] = None
    providers: Optional[List[str]] = None
    target_accounts: Optional[List[str]] = None


class ScanProgressResponse(BaseModel):
    scan_id: str
    customer_id: str
    organization_id: int
    state: str
    progress: int = 0
    providers: List[str]
    started_at: datetime
    completed_at: Optional[datetime] = None
    total_resources: int = 0
    anomalies_found: int = 0
    savings_identified: float = 0.0


class ProviderDiagnostic(BaseModel):
    provider: str
    configured: bool
    required_settings: List[str]
    missing_settings: List[str]
    recommendation: str
    scope_model: str = ""
    primary_apis: List[str] = Field(default_factory=list)
    optimization_apis: List[str] = Field(default_factory=list)
    telemetry_apis: List[str] = Field(default_factory=list)
    default_page_size: int = 100
    max_page_size: int = 200
    max_parallel_requests: int = 3
    request_timeout_seconds: int = 30
    retryable_statuses: List[int] = Field(default_factory=list)
    throttling_signals: List[str] = Field(default_factory=list)
    scan_notes: List[str] = Field(default_factory=list)


class ScanHistoryItem(BaseModel):
    scan_id: str
    customer_id: str
    organization_id: int
    state: str
    providers: List[str]
    started_at: datetime
    completed_at: Optional[datetime] = None
    total_resources: int = 0
    anomalies_found: int = 0
    savings_identified: float = 0.0


class ScanDiffEntry(BaseModel):
    provider: str
    current_cost_usd: float
    previous_cost_usd: float
    delta_cost_usd: float
    delta_percent: Optional[float] = None
    current_anomalies: int
    previous_anomalies: int


class ScanDiffResponse(BaseModel):
    organization_id: int
    current_scan_id: str
    previous_scan_id: Optional[str] = None
    total_current_cost_usd: float
    total_previous_cost_usd: float
    total_delta_cost_usd: float
    entries: List[ScanDiffEntry]


class AccountRegionRow(BaseModel):
    region: str
    cost_usd: float


class ProviderAccountRollupItem(BaseModel):
    account_id: int
    provider: str
    account_identifier: str
    account_name: str
    account_type: str
    depth: int = 0
    parent_account_id: Optional[int] = None
    parent_account_identifier: Optional[str] = None
    direct_cost_usd: float
    rolled_up_cost_usd: float
    direct_savings_identified_usd: float
    rolled_up_savings_identified_usd: float
    direct_anomalies_count: int
    rolled_up_anomalies_count: int
    direct_service_count: int
    rolled_up_service_count: int
    child_count: int
    budget_monthly_usd: Optional[float] = None
    rolled_up_budget_monthly_usd: Optional[float] = None
    budget_utilization_percent: Optional[float] = None
    rolled_up_budget_utilization_percent: Optional[float] = None
    budget_status: Optional[str] = None
    scan_id: Optional[str] = None
    captured_at: Optional[str] = None
    top_regions: List[AccountRegionRow] = Field(default_factory=list)


class ProviderAccountRollupResponse(BaseModel):
    organization_id: int
    customer_id: str
    provider: Optional[str] = None
    scan_id: Optional[str] = None
    generated_at: str
    total_direct_cost_usd: float
    total_rolled_up_cost_usd: float
    items: List[ProviderAccountRollupItem]


class ProviderAccountInventoryItem(BaseModel):
    account_id: int
    provider: str
    account_identifier: str
    account_name: str
    account_type: str
    native_region: Optional[str] = None
    is_active: bool
    metadata: Dict[str, Any]
    created_at: str
    updated_at: Optional[str] = None


class ProviderAccountInventoryResponse(BaseModel):
    organization_id: int
    customer_id: str
    total: int
    accounts: List[ProviderAccountInventoryItem]


class AccountRegionBreakdownItem(BaseModel):
    region: str
    cost_usd: float
    scan_id: str
    captured_at: str


class AccountRegionBreakdownResponse(BaseModel):
    account_id: int
    provider: str
    account_name: str
    scan_id: Optional[str] = None
    total_cost_usd: float
    regions: List[AccountRegionBreakdownItem]


# ── Business mapping / chargeback models ────────────────────────────────────

VALID_DIMENSIONS = {"team", "environment", "application", "cost_center"}


class BusinessMappingRuleRequest(BaseModel):
    tag_key: str
    tag_value: str = "*"
    dimension: str
    mapped_value: str
    priority: int = 100
    is_active: bool = True


class BusinessMappingRuleUpdateRequest(BaseModel):
    tag_key: str | None = None
    tag_value: str | None = None
    dimension: str | None = None
    mapped_value: str | None = None
    priority: int | None = None
    is_active: bool | None = None


class BusinessMappingRuleResponse(BaseModel):
    id: int
    organization_id: int
    customer_id: str
    tag_key: str
    tag_value: str
    dimension: str
    mapped_value: str
    priority: int
    is_active: bool
    created_at: str
    updated_at: Optional[str] = None


class BusinessMappingRuleListResponse(BaseModel):
    organization_id: int
    rules: List[BusinessMappingRuleResponse]
    total: int


class ChargebackDimensionGroup(BaseModel):
    dimension: str          # "team" | "environment" | "application" | "cost_center"
    value: str              # e.g. "platform-team"
    total_cost_usd: float
    provider_breakdown: Dict[str, float]
    record_count: int


class ChargebackResponse(BaseModel):
    organization_id: int
    dimension_type: str
    groups: List[ChargebackDimensionGroup]
    total_mapped_cost_usd: float
    total_unmapped_cost_usd: float
    total_cost_usd: float
    coverage_percent: float


class AllocationCoverageResponse(BaseModel):
    organization_id: int
    total_cost_usd: float
    mapped_cost_usd: float
    unmapped_cost_usd: float
    coverage_percent: float
    dimension_coverage: Dict[str, float]   # {"team": 82.4, "environment": 91.0, …}
    provider_coverage: Dict[str, float]    # {"aws": 78.0, "gcp": 95.0, …}
    unmapped_top_services: List[Dict[str, Any]]


# ── Trend / Period Summary models ─────────────────────────────────────────────

class CostTrendPoint(BaseModel):
    period_start: str
    period_end: str
    provider: str
    dimension_value: str = ""
    total_cost_usd: float
    mapped_cost_usd: float
    unmapped_cost_usd: float
    record_count: int
    team: Optional[str] = None
    environment: Optional[str] = None
    service_breakdown: Dict[str, float] = Field(default_factory=dict)


class CostTrendResponse(BaseModel):
    organization_id: int
    period_type: str             # "monthly" | "weekly"
    lookback_periods: int
    view_by: str                 # "provider" | "region" | "service" | "account"
    data_source: str             # "computed" | "raw_records" | "empty"
    points: List[CostTrendPoint]
    provider_totals: Dict[str, float]
    dimension_totals: Dict[str, float]
    grand_total_usd: float


class PeriodSummaryComputeResponse(BaseModel):
    organization_id: int
    period_type: str
    periods_computed: int
    rows_written: int
    computed_at: str


# ── End trend models ──────────────────────────────────────────────────────────




class AlertRoutingPolicyRequest(BaseModel):
    severity: Literal["warning", "critical"]
    channels: List[str]
    is_active: bool = True


class AlertRoutingPolicyResponse(BaseModel):
    id: int
    severity: str
    channels: List[str]
    is_active: bool
    created_at: str
    updated_at: str


class NotificationDestinationStatus(BaseModel):
    channel: str
    configured: bool
    enabled: bool
    last_delivery_at: Optional[str] = None
    last_success_at: Optional[str] = None
    last_error_at: Optional[str] = None


class NotificationDestinationsResponse(BaseModel):
    organization_id: int
    destinations: List[NotificationDestinationStatus]


class NotificationDestinationToggleRequest(BaseModel):
    enabled: bool


class NotificationDestinationTestRequest(BaseModel):
    channel: Literal["email", "slack", "teams"]
    target: Optional[str] = None
    message: Optional[str] = None


class NotificationDestinationTestResponse(BaseModel):
    channel: str
    success: bool
    detail: str


class AlertRoutingPolicySimulationRequest(BaseModel):
    severity: Literal["warning", "critical"]
    title: Optional[str] = None
    alert_type: Optional[str] = None


class AlertRoutingPolicySimulationResponse(BaseModel):
    severity: str
    matched_policy_id: Optional[int] = None
    evaluated_channels: List[str]
    expected_channels: List[str]
    configured_channels: List[str]
    inactive_policy: bool = False


class AlertLifecycleActionResponse(BaseModel):
    status: str
    alert_id: int
    lifecycle_state: Literal["active", "acknowledged", "dismissed", "reactivated"]


class AlertOpsPolicyRequest(BaseModel):
    mute_window_enabled: bool = False
    mute_start_hour_utc: int = Field(default=0, ge=0, le=23)
    mute_end_hour_utc: int = Field(default=0, ge=0, le=23)
    mute_weekends: bool = False
    timezone: str = "UTC"
    escalation_enabled: bool = False
    escalation_after_minutes: int = Field(default=60, ge=5, le=10080)
    escalation_channels: List[str] = Field(default_factory=list)
    escalation_severity: Literal["warning", "critical"] = "critical"
    ack_sla_minutes: int = Field(default=60, ge=5, le=10080)
    dedupe_window_minutes: int = Field(default=30, ge=0, le=1440)
    min_severity: Literal["low", "medium", "high", "warning", "critical"] = "low"
    daily_summary_enabled: bool = True
    weekly_summary_enabled: bool = True


class AlertOpsPolicyResponse(BaseModel):
    organization_id: int
    mute_window_enabled: bool
    mute_start_hour_utc: int
    mute_end_hour_utc: int
    mute_weekends: bool
    timezone: str
    escalation_enabled: bool
    escalation_after_minutes: int
    escalation_channels: List[str]
    escalation_severity: str
    ack_sla_minutes: int
    dedupe_window_minutes: int
    min_severity: str
    daily_summary_enabled: bool
    weekly_summary_enabled: bool
    created_at: str
    updated_at: str


class AlertExecutiveSummaryResponse(BaseModel):
    organization_id: int
    period: Literal["daily", "weekly"]
    generated_at: str
    window_start: str
    total_alerts: int
    acknowledged: int
    unacknowledged: int
    dismissed: int
    by_severity: Dict[str, int]


class ConnectorTestRequest(BaseModel):
    connector_type: str
    config: Dict[str, Any]


class ConnectorStatusResponse(BaseModel):
    connector_type: str
    status: str
    last_sync: Optional[str] = None
    message: Optional[str] = None


class ConnectorListResponse(BaseModel):
    supported_connectors: List[str]
    description: str


class SchedulerCounters(BaseModel):
    total: int
    success: int
    failure: int


class SchedulerTimelineItem(BaseModel):
    id: str
    event_type: str
    state: str
    title: str
    detail: str
    created_at: str


class SchedulerStatusResponse(BaseModel):
    organization_id: int
    customer_id: str
    scheduler_enabled: bool
    scheduler_running: bool
    permission_state: str
    scan_frequency: str
    next_run_at: Optional[str] = None
    next_run_eta_seconds: Optional[int] = None
    effective_scan_frequency: Optional[str] = None
    scheduler_override_enabled: bool = False
    last_success_at: Optional[str] = None
    last_failure_at: Optional[str] = None
    retry_max_attempts: int = 1
    retry_backoff_seconds: int = 15
    overdue_alert_hours: int = 24
    overdue: bool = False
    counters: SchedulerCounters
    timeline: List[SchedulerTimelineItem]


class SchedulerPolicyUpdateRequest(BaseModel):
    scheduler_override_enabled: bool = False
    scheduler_override_frequency: Optional[Literal["hourly", "daily", "weekly"]] = None
    scheduler_retry_max_attempts: int = Field(default=2, ge=1, le=8)
    scheduler_retry_backoff_seconds: int = Field(default=120, ge=15, le=3600)
    scheduler_overdue_alert_hours: int = Field(default=24, ge=1, le=168)


class DataFreshnessProviderItem(BaseModel):
    provider: str
    last_ingested_at: Optional[str] = None
    age_seconds: Optional[int] = None
    status: Literal["fresh", "stale", "unknown"]


class DataFreshnessConnectorItem(BaseModel):
    connector: str
    last_event_at: Optional[str] = None
    age_seconds: Optional[int] = None
    status: Literal["fresh", "stale", "unknown"]


class DataFreshnessResponse(BaseModel):
    organization_id: int
    customer_id: str
    generated_at: str
    providers: List[DataFreshnessProviderItem]
    connectors: List[DataFreshnessConnectorItem]
    scheduler_lag_seconds: Optional[int] = None
    scheduler_status: Literal["healthy", "lagging", "unknown"] = "unknown"


class ExternalAWSAnomalyIngestRequest(BaseModel):
    events: List[Dict[str, Any]]


class ExternalAWSReplayRequest(BaseModel):
    event_ids: Optional[List[str]] = None
    days_back: Optional[int] = None
    max_results: int = Field(default=50, le=500)


class ExternalGCPPubSubIngestRequest(BaseModel):
    message: Dict[str, Any]
    subscription: Optional[str] = None


class ExportJobRequest(BaseModel):
    name: str
    report_type: Literal["executive_summary", "executive_digest", "finance_workbook"] = "executive_summary"
    export_format: Literal["csv", "xls", "xlsx", "pdf"] = "csv"
    schedule_frequency: Literal["daily", "weekly", "monthly"] = "weekly"
    is_active: bool = True


class ExportJobResponse(BaseModel):
    id: int
    organization_id: int
    customer_id: str
    name: str
    report_type: str
    export_format: str
    schedule_frequency: str
    is_active: bool
    last_run_at: Optional[str] = None
    created_at: str
    updated_at: str


class ExportJobRunResponse(BaseModel):
    id: int
    export_job_id: int
    status: str
    output_filename: Optional[str] = None
    row_count: int
    error_message: Optional[str] = None
    created_at: str
    completed_at: Optional[str] = None


class ImportedCostUploadResponse(BaseModel):
    organization_id: int
    customer_id: str
    upload_id: str
    filename: str
    rows_imported: int
    total_cost_usd: float
    providers: List[str]
    imported_at: datetime


class ImportedCostSummaryResponse(BaseModel):
    organization_id: int
    customer_id: str
    has_data: bool
    upload_id: Optional[str] = None
    source_filename: Optional[str] = None
    rows_imported: int = 0
    total_cost_usd: float = 0.0
    providers: List[str] = Field(default_factory=list)
    last_imported_at: Optional[datetime] = None


class ImportPreviewIssue(BaseModel):
    line_number: int
    severity: Literal["error", "warning"]
    message: str


class ImportPreviewResponse(BaseModel):
    organization_id: int
    customer_id: str
    filename: str
    total_rows: int
    accepted_rows: int
    rejected_rows: int
    total_cost_usd: float
    detected_providers: List[str] = Field(default_factory=list)
    header_columns: List[str] = Field(default_factory=list)
    mapping_feedback: Dict[str, Any] = Field(default_factory=dict)
    reconciliation_guidance: List[str] = Field(default_factory=list)
    issues: List[ImportPreviewIssue] = Field(default_factory=list)


class ReportShareTokenRequest(BaseModel):
    report_type: Literal["executive_summary", "finance_workbook", "executive_digest"] = "executive_summary"
    report_format: Literal["json", "csv", "xlsx", "pdf"] = "json"
    expires_in_hours: int = 168


class ReportShareTokenResponse(BaseModel):
    token: str
    expires_at: str
    report_type: str
    report_format: str


def get_credential_manager(db: Session = Depends(get_db)) -> CredentialManager:
    return CredentialManager(db)


def get_scanning_manager(db: Session = Depends(get_db)) -> ScanningManager:
    return ScanningManager(db)


def _parse_optional_datetime_value(
    value: Optional[str],
    field_name: str,
    line_number: int,
) -> tuple[Optional[datetime], Optional[str]]:
    text = str(value or "").strip()
    if not text:
        return None, None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")), None
    except ValueError:
        return None, f"Invalid {field_name} at CSV line {line_number}. Use ISO date or datetime."


def _parse_required_float_value(
    value: Optional[str],
    field_name: str,
    line_number: int,
) -> tuple[Optional[float], Optional[str]]:
    text = str(value or "").strip()
    if not text:
        return None, f"Missing {field_name} at CSV line {line_number}."
    try:
        parsed = float(text)
    except ValueError:
        return None, f"Invalid {field_name} at CSV line {line_number}."
    if parsed < 0:
        return None, f"Negative {field_name} at CSV line {line_number}. Cost values must be zero or greater."
    return parsed, None


def _csv_escape(value: Any) -> str:
    text = str(value if value is not None else "").replace('"', '""')
    return f'"{text}"'


def _csv_response(filename: str, header: List[str], rows: List[List[Any]]) -> Response:
    lines = [",".join(header)]
    for row in rows:
        lines.append(",".join(_csv_escape(cell) for cell in row))
    return Response(
        "\n".join(lines),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _spreadsheet_xml_response(filename: str, sheet_name: str, rows: List[List[Any]]) -> Response:
    def _cell_xml(value: Any) -> str:
        if value is None or value == "":
            return '<Cell><Data ss:Type="String"></Data></Cell>'
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return f'<Cell><Data ss:Type="Number">{value}</Data></Cell>'
        return f'<Cell><Data ss:Type="String">{xml_escape(str(value))}</Data></Cell>'

    row_xml = "".join(
        "<Row>" + "".join(_cell_xml(cell) for cell in row) + "</Row>"
        for row in rows
    )
    workbook = (
        '<?xml version="1.0"?>'
        '<?mso-application progid="Excel.Sheet"?>'
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel" '
        'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:html="http://www.w3.org/TR/REC-html40">'
        f'<Worksheet ss:Name="{xml_escape(sheet_name[:31] or "Sheet1")}"><Table>{row_xml}</Table></Worksheet>'
        "</Workbook>"
    )
    return Response(
        workbook,
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _imported_cost_summary(rows: List[ImportedCostRecord]) -> Dict[str, Any]:
    return summarize_imported_cost_rows(rows)


def _get_imported_cost_rows(
    db: Session,
    organization_id: int,
    customer_id: str,
    cloud_provider: str = "all",
) -> List[ImportedCostRecord]:
    return query_imported_cost_rows(db, organization_id, customer_id, cloud_provider)


def _materialize_rollup_items(nodes: Dict[int, Dict[str, Any]]) -> List[ProviderAccountRollupItem]:
    children_by_parent: Dict[Optional[int], List[int]] = {}
    for node_id, node in nodes.items():
        children_by_parent.setdefault(node.get("parent_account_id"), []).append(node_id)

    for child_ids in children_by_parent.values():
        child_ids.sort(key=lambda item_id: (
            str(nodes[item_id].get("provider") or ""),
            str(nodes[item_id].get("account_name") or ""),
            str(nodes[item_id].get("account_identifier") or ""),
        ))

    def _walk(node_id: int, depth: int) -> tuple[float, float, int, int, int]:
        node = nodes[node_id]
        child_ids = children_by_parent.get(node_id, [])
        child_count = len(child_ids)
        direct_cost = float(node.get("direct_cost_usd") or 0.0)
        direct_savings = float(node.get("direct_savings_identified_usd") or 0.0)
        direct_anomalies = int(node.get("direct_anomalies_count") or 0)
        direct_services = int(node.get("direct_service_count") or 0)

        # Grouping nodes represent aggregated structure, not additive direct spend.
        if child_count > 0 and str(node.get("account_type") or "") in {
            "provider",
            "management_group",
            "organization",
            "folder",
            "tenancy",
        }:
            direct_cost = 0.0
            direct_savings = 0.0
            direct_anomalies = 0
            direct_services = 0

        rolled_cost = direct_cost
        rolled_savings = direct_savings
        rolled_anomalies = direct_anomalies
        rolled_services = direct_services
        total_descendants = child_count

        node["depth"] = depth
        node["child_count"] = child_count
        node["direct_cost_usd"] = round(direct_cost, 2)
        node["direct_savings_identified_usd"] = round(direct_savings, 2)
        node["direct_anomalies_count"] = direct_anomalies
        node["direct_service_count"] = direct_services

        for child_id in child_ids:
            child_cost, child_savings, child_anomalies, child_services, descendant_count = _walk(child_id, depth + 1)
            rolled_cost += child_cost
            rolled_savings += child_savings
            rolled_anomalies += child_anomalies
            rolled_services += child_services
            total_descendants += descendant_count

        node["rolled_up_cost_usd"] = round(rolled_cost, 2)
        node["rolled_up_savings_identified_usd"] = round(rolled_savings, 2)
        node["rolled_up_anomalies_count"] = rolled_anomalies
        node["rolled_up_service_count"] = rolled_services
        node["descendant_count"] = total_descendants
        return rolled_cost, rolled_savings, rolled_anomalies, rolled_services, total_descendants

    root_ids = children_by_parent.get(None, [])
    for root_id in root_ids:
        _walk(root_id, 0)

    ordered_ids: List[int] = []

    def _append(node_id: int) -> None:
        ordered_ids.append(node_id)
        for child_id in children_by_parent.get(node_id, []):
            _append(child_id)

    for root_id in root_ids:
        _append(root_id)

    return [
        ProviderAccountRollupItem(
            account_id=node_id,
            provider=str(nodes[node_id].get("provider") or ""),
            account_identifier=str(nodes[node_id].get("account_identifier") or ""),
            account_name=str(nodes[node_id].get("account_name") or ""),
            account_type=str(nodes[node_id].get("account_type") or "account"),
            depth=int(nodes[node_id].get("depth") or 0),
            parent_account_id=nodes[node_id].get("parent_account_id"),
            parent_account_identifier=nodes[node_id].get("parent_account_identifier"),
            direct_cost_usd=float(nodes[node_id].get("direct_cost_usd") or 0.0),
            rolled_up_cost_usd=float(nodes[node_id].get("rolled_up_cost_usd") or 0.0),
            direct_savings_identified_usd=float(nodes[node_id].get("direct_savings_identified_usd") or 0.0),
            rolled_up_savings_identified_usd=float(nodes[node_id].get("rolled_up_savings_identified_usd") or 0.0),
            direct_anomalies_count=int(nodes[node_id].get("direct_anomalies_count") or 0),
            rolled_up_anomalies_count=int(nodes[node_id].get("rolled_up_anomalies_count") or 0),
            direct_service_count=int(nodes[node_id].get("direct_service_count") or 0),
            rolled_up_service_count=int(nodes[node_id].get("rolled_up_service_count") or 0),
            child_count=int(nodes[node_id].get("child_count") or 0),
            scan_id=nodes[node_id].get("scan_id"),
            captured_at=nodes[node_id].get("captured_at"),
            top_regions=nodes[node_id].get("top_regions") or [],
        )
        for node_id in ordered_ids
    ]


def _apply_rollup_budget_metrics(
    items: List[ProviderAccountRollupItem],
    monthly_budget_usd: float,
    warning_threshold_percent: float,
    critical_threshold_percent: float,
) -> List[ProviderAccountRollupItem]:
    total_budget = float(monthly_budget_usd or 0.0)
    if total_budget <= 0:
        return items

    baseline_cost = float(
        sum(item.rolled_up_cost_usd for item in items if item.depth == 0) or 0.0
    )
    if baseline_cost <= 0:
        return items

    for item in items:
        direct_budget = round(total_budget * (float(item.direct_cost_usd or 0.0) / baseline_cost), 2)
        rolled_budget = round(total_budget * (float(item.rolled_up_cost_usd or 0.0) / baseline_cost), 2)
        item.budget_monthly_usd = direct_budget
        item.rolled_up_budget_monthly_usd = rolled_budget

        if direct_budget > 0:
            item.budget_utilization_percent = round((float(item.direct_cost_usd or 0.0) / direct_budget) * 100, 2)
        else:
            item.budget_utilization_percent = None

        rolled_util = None
        if rolled_budget > 0:
            rolled_util = round((float(item.rolled_up_cost_usd or 0.0) / rolled_budget) * 100, 2)
        item.rolled_up_budget_utilization_percent = rolled_util

        if rolled_util is None:
            item.budget_status = None
        elif rolled_util >= float(critical_threshold_percent or 100.0):
            item.budget_status = "critical"
        elif rolled_util >= float(warning_threshold_percent or 80.0):
            item.budget_status = "warning"
        else:
            item.budget_status = "ok"

    return items


def _build_rollups_from_imported_rows(
    rows: List[ImportedCostRecord],
    organization_id: int,
    customer_id: str,
    provider: Optional[str] = None,
    monthly_budget_usd: float = 0.0,
    warning_threshold_percent: float = 80.0,
    critical_threshold_percent: float = 100.0,
) -> ProviderAccountRollupResponse:
    next_synthetic_id = -1

    def _synthetic_id() -> int:
        nonlocal next_synthetic_id
        current = next_synthetic_id
        next_synthetic_id -= 1
        return current

    nodes: Dict[int, Dict[str, Any]] = {}
    provider_root_ids: Dict[str, int] = {}
    account_nodes: Dict[tuple[str, str], int] = {}
    latest_imported_at = max((row.created_at for row in rows), default=None)

    def _normalized_account_type(raw_value: Optional[str], default: str) -> str:
        value = str(raw_value or "").strip().lower().replace(" ", "_")
        return value or default

    def _ensure_node(
        *,
        provider_key: str,
        identifier: str,
        account_name: str,
        account_type: str,
        parent_account_id: Optional[int],
    ) -> int:
        node_key = (provider_key, identifier)
        node_id = account_nodes.get(node_key)
        if node_id is None:
            node_id = _synthetic_id()
            account_nodes[node_key] = node_id
            nodes[node_id] = {
                "provider": provider_key,
                "account_identifier": identifier,
                "account_name": account_name,
                "account_type": account_type,
                "parent_account_id": parent_account_id,
                "parent_account_identifier": (
                    nodes[parent_account_id]["account_identifier"]
                    if parent_account_id in nodes
                    else None
                ),
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": None,
                "captured_at": latest_imported_at.isoformat() if latest_imported_at else None,
                "_services": set(),
            }
            return node_id

        node = nodes[node_id]
        if account_name and (
            not str(node.get("account_name") or "").strip()
            or str(node.get("account_name") or "").strip() == identifier
        ):
            node["account_name"] = account_name
        if (
            account_type
            and str(node.get("account_type") or "").strip() in {"", "account", "group"}
            and account_type not in {"", "account", "group"}
        ):
            node["account_type"] = account_type
        provider_root_id = provider_root_ids.get(provider_key)
        if parent_account_id is not None and parent_account_id != node_id:
            existing_parent = node.get("parent_account_id")
            if existing_parent is None or (
                provider_root_id is not None
                and existing_parent == provider_root_id
                and parent_account_id != provider_root_id
            ):
                node["parent_account_id"] = parent_account_id
                node["parent_account_identifier"] = (
                    nodes[parent_account_id]["account_identifier"]
                    if parent_account_id in nodes
                    else None
                )
        return node_id

    for row in rows:
        provider_key = str(row.provider or "").strip().lower()
        if not provider_key:
            continue
        root_id = provider_root_ids.get(provider_key)
        if root_id is None:
            root_id = _synthetic_id()
            provider_root_ids[provider_key] = root_id
            nodes[root_id] = {
                "provider": provider_key,
                "account_identifier": f"{provider_key}:provider",
                "account_name": provider_key.upper(),
                "account_type": "provider",
                "parent_account_id": None,
                "parent_account_identifier": None,
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": None,
                "captured_at": latest_imported_at.isoformat() if latest_imported_at else None,
            }

        identifier = str(row.account_identifier or row.account_name or "").strip()
        account_name = str(row.account_name or row.account_identifier or provider_key).strip()
        if not identifier:
            identifier = f"{provider_key}:unassigned"
            account_name = f"{provider_key.upper()} Unassigned"

        parent_identifier = str(row.parent_account_identifier or "").strip()
        parent_account_id: Optional[int] = root_id
        if parent_identifier:
            inferred_parent_type = _infer_parent_account_type(
                provider_key=provider_key,
                child_account_type=_normalized_account_type(row.account_type, "account"),
            )
            parent_account_id = _ensure_node(
                provider_key=provider_key,
                identifier=parent_identifier,
                account_name=parent_identifier,
                account_type=inferred_parent_type,
                parent_account_id=root_id,
            )

        account_id = _ensure_node(
            provider_key=provider_key,
            identifier=identifier,
            account_name=account_name,
            account_type=_normalized_account_type(row.account_type, "account"),
            parent_account_id=parent_account_id,
        )
        nodes[account_id]["direct_cost_usd"] = round(float(nodes[account_id]["direct_cost_usd"]) + float(row.cost_usd or 0.0), 2)
        service_name = str(row.service_name or "").strip()
        if service_name:
            nodes[account_id].setdefault("_services", set()).add(service_name)
        region = str(row.region or "").strip()
        if region:
            region_costs = nodes[account_id].setdefault("_region_costs", {})
            region_costs[region] = round(region_costs.get(region, 0.0) + float(row.cost_usd or 0.0), 2)

    for node in nodes.values():
        node["direct_service_count"] = len(node.get("_services", set()))
        node.pop("_services", None)
        region_costs = node.pop("_region_costs", {})
        node["top_regions"] = [
            AccountRegionRow(region=r, cost_usd=c)
            for r, c in sorted(region_costs.items(), key=lambda x: x[1], reverse=True)[:5]
        ]

    items = _materialize_rollup_items(nodes)
    filtered_items = [item for item in items if provider is None or item.provider == provider]
    filtered_items = _apply_rollup_budget_metrics(
        filtered_items,
        monthly_budget_usd=monthly_budget_usd,
        warning_threshold_percent=warning_threshold_percent,
        critical_threshold_percent=critical_threshold_percent,
    )
    direct_total = round(sum(item.direct_cost_usd for item in filtered_items if item.depth > 0 or item.child_count == 0), 2)
    root_total = round(sum(item.rolled_up_cost_usd for item in filtered_items if item.depth == 0), 2)
    return ProviderAccountRollupResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        provider=provider,
        scan_id=None,
        generated_at=_utcnow().isoformat(),
        total_direct_cost_usd=direct_total,
        total_rolled_up_cost_usd=root_total,
        items=filtered_items,
    )


def _parse_credential_payload(raw: Dict[str, Any]) -> CredentialInput:
    return parse_credential_payload(raw)


def _run_validation(credential: CredentialInput) -> CredentialStatus:
    return run_credential_validation(credential)


def _safe_json_load(raw: str, default: Dict[str, Any]) -> Dict[str, Any]:
    try:
        return json.loads(raw)
    except Exception:
        return default


def _customer_id_for_org(membership: UserOrganization) -> str:
    """Derive persisted customer scope from the active organization."""
    return f"org-{membership.organization_id}"


def _organization_id_for_membership(membership: UserOrganization) -> int:
    return int(membership.organization_id)


def _organization_id_from_customer_id(customer_id: str) -> Optional[int]:
    normalized = str(customer_id or "").strip()
    if normalized.startswith("org-"):
        try:
            return int(normalized.split("-", 1)[1])
        except (TypeError, ValueError):
            return None
    return None


def _require_management_role(membership: UserOrganization, action: str) -> None:
    require_role(
        membership,
        allowed_roles=[UserRole.OWNER, UserRole.ADMIN],
        action=action,
    )


def _normalize_oci_profile(profile: Optional[str]) -> str:
    normalized = str(profile or "DEFAULT").strip() or "DEFAULT"
    if normalized.startswith("[") and normalized.endswith("]") and len(normalized) > 2:
        normalized = normalized[1:-1].strip() or "DEFAULT"
    return normalized


def _oci_uploaded_credentials_dir(customer_id: str) -> str:
    base_dir = os.getenv(
        "OCI_UPLOADED_CREDENTIAL_DIR",
        "/opt/optiora/.runtime/cloud-credentials",
    )
    target_dir = os.path.abspath(os.path.join(base_dir, customer_id, "oci"))
    try:
        os.makedirs(target_dir, mode=0o700, exist_ok=True)
        try:
            os.chmod(target_dir, 0o700)
        except OSError:
            pass
        return target_dir
    except OSError:
        fallback_base = os.path.join(tempfile.gettempdir(), "optiora-cloud-credentials")
        fallback_dir = os.path.abspath(os.path.join(fallback_base, customer_id, "oci"))
        os.makedirs(fallback_dir, mode=0o700, exist_ok=True)
        try:
            os.chmod(fallback_dir, 0o700)
        except OSError:
            pass
        return fallback_dir


def _runtime_credentials_root(customer_id: str) -> str:
    base_dir = os.getenv(
        "RUNTIME_CREDENTIAL_DIR",
        "/opt/optiora/.runtime/cloud-credentials",
    )
    target_dir = os.path.abspath(os.path.join(base_dir, customer_id))
    try:
        os.makedirs(target_dir, mode=0o700, exist_ok=True)
        try:
            os.chmod(target_dir, 0o700)
        except OSError:
            pass
        return target_dir
    except OSError:
        fallback_base = os.path.join(tempfile.gettempdir(), "optiora-runtime-credentials")
        fallback_dir = os.path.abspath(os.path.join(fallback_base, customer_id))
        os.makedirs(fallback_dir, mode=0o700, exist_ok=True)
        try:
            os.chmod(fallback_dir, 0o700)
        except OSError:
            pass
        return fallback_dir


def _write_runtime_provider_credentials(
    customer_id: str,
    credential: CredentialInput,
) -> Dict[str, Any]:
    """Persist runtime credentials on host for live provider API operations."""
    root = _runtime_credentials_root(customer_id)
    provider = credential.provider
    provider_dir = os.path.join(root, provider)
    os.makedirs(provider_dir, mode=0o700, exist_ok=True)
    try:
        os.chmod(provider_dir, 0o700)
    except OSError:
        pass

    if provider == "aws":
        payload = credential.model_dump()
        runtime_payload = {
            "access_key_id": str(payload.get("access_key_id") or ""),
            "secret_access_key": str(payload.get("secret_access_key") or ""),
            "region": str(payload.get("region") or "us-east-1"),
        }
        runtime_path = os.path.join(provider_dir, "runtime.json")
        with open(runtime_path, "w", encoding="utf-8") as fh:
            json.dump(runtime_payload, fh)
        os.chmod(runtime_path, 0o600)
        return {"provider": "aws", "runtime_file": runtime_path}

    if provider == "azure":
        payload = credential.model_dump()
        runtime_payload = {
            "subscription_id": str(payload.get("subscription_id") or ""),
            "tenant_id": str(payload.get("tenant_id") or ""),
            "client_id": str(payload.get("client_id") or ""),
            "client_secret": str(payload.get("client_secret") or ""),
        }
        runtime_path = os.path.join(provider_dir, "runtime.json")
        with open(runtime_path, "w", encoding="utf-8") as fh:
            json.dump(runtime_payload, fh)
        os.chmod(runtime_path, 0o600)
        return {"provider": "azure", "runtime_file": runtime_path}

    if provider == "gcp":
        payload = credential.model_dump()
        project_id = str(payload.get("project_id") or "")
        service_account = payload.get("service_account_json")
        if isinstance(service_account, str):
            service_account = json.loads(service_account)
        if not isinstance(service_account, dict):
            service_account = {}
        service_account_path = os.path.join(provider_dir, "service-account.json")
        with open(service_account_path, "w", encoding="utf-8") as fh:
            json.dump(service_account, fh)
        os.chmod(service_account_path, 0o600)
        runtime_path = os.path.join(provider_dir, "runtime.json")
        with open(runtime_path, "w", encoding="utf-8") as fh:
            json.dump(
                {
                    "project_id": project_id,
                    "service_account_file": service_account_path,
                },
                fh,
            )
        os.chmod(runtime_path, 0o600)
        return {"provider": "gcp", "runtime_file": runtime_path, "service_account_file": service_account_path}

    if provider == "oci":
        payload = credential.model_dump()
        runtime_payload = {
            "config_file": str(payload.get("config_file") or ""),
            "profile": _normalize_oci_profile(payload.get("profile")),
        }
        runtime_path = os.path.join(provider_dir, "runtime.json")
        with open(runtime_path, "w", encoding="utf-8") as fh:
            json.dump(runtime_payload, fh)
        os.chmod(runtime_path, 0o600)
        return {"provider": "oci", "runtime_file": runtime_path}

    return {"provider": provider}


def _load_runtime_provider_credentials(customer_id: str) -> Dict[str, Dict[str, Any]]:
    root = _runtime_credentials_root(customer_id)
    output: Dict[str, Dict[str, Any]] = {}
    for provider in ("aws", "azure", "gcp", "oci"):
        runtime_path = os.path.join(root, provider, "runtime.json")
        if not os.path.isfile(runtime_path):
            continue
        try:
            with open(runtime_path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                output[provider] = data
        except Exception:
            continue
    return output


def _merge_runtime_with_stored_credentials(
    runtime_credentials: Dict[str, Dict[str, Any]],
    stored_rows: List["CredentialRecord"],
) -> Dict[str, Dict[str, Any]]:
    """Merge runtime files with validated stored credentials for live provider catalog fetching."""
    merged: Dict[str, Dict[str, Any]] = {
        provider: dict(values or {})
        for provider, values in (runtime_credentials or {}).items()
        if isinstance(values, dict)
    }
    for row in stored_rows:
        provider = str(getattr(row, "provider", "") or "").strip().lower()
        if provider not in {"aws", "azure", "gcp", "oci"}:
            continue
        raw_json = str(getattr(row, "credential_json", "") or "").strip()
        if not raw_json:
            continue
        try:
            payload = json.loads(raw_json)
        except Exception:
            continue
        if not isinstance(payload, dict):
            continue

        target = merged.setdefault(provider, {})
        def _fill(key: str, value: Any) -> None:
            if target.get(key):
                return
            text = value
            if isinstance(value, str):
                text = value.strip()
            target[key] = text
        if provider == "aws":
            _fill("access_key_id", str(payload.get("access_key_id") or ""))
            _fill("secret_access_key", str(payload.get("secret_access_key") or ""))
            _fill("region", str(payload.get("region") or "us-east-1"))
        elif provider == "azure":
            _fill("subscription_id", str(payload.get("subscription_id") or ""))
            _fill("tenant_id", str(payload.get("tenant_id") or ""))
            _fill("client_id", str(payload.get("client_id") or ""))
            _fill("client_secret", str(payload.get("client_secret") or ""))
        elif provider == "gcp":
            _fill("project_id", str(payload.get("project_id") or ""))
            if "service_account_json" in payload and not target.get("service_account_json"):
                target["service_account_json"] = payload.get("service_account_json")
        elif provider == "oci":
            _fill("config_file", str(payload.get("config_file") or ""))
            _fill("profile", _normalize_oci_profile(payload.get("profile")))
    return merged


def _resolve_customer_id(
    current_user: User,
    membership: UserOrganization,
    requested_customer_id: Optional[str] = None,
) -> str:
    """Reject mismatched customer scopes and return the org-derived identifier."""
    _ = current_user
    derived_customer_id = _customer_id_for_org(membership)
    normalized_requested = str(requested_customer_id or "").strip()
    if normalized_requested and normalized_requested != derived_customer_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="customer_id must match the authenticated organization scope",
        )
    return derived_customer_id


async def _cost_summary_for_provider(
    provider: str,
    period: str = "month",
    customer_id: Optional[str] = None,
) -> Dict[str, Any]:
    runtime_credentials = (
        _load_runtime_provider_credentials(customer_id).get(provider)
        if customer_id
        else None
    )
    return await fetch_provider_cost_summary(
        provider,
        period,
        credentials=runtime_credentials,
        fetchers={
            "aws": aws_costs.get_cost_summary,
            "azure": azure_costs.get_cost_summary,
            "gcp": gcp_costs.get_cost_summary,
            "oci": oci_costs.get_cost_summary,
        },
        safe_json_load=_safe_json_load,
    )


def _imported_cost_context(
    membership: UserOrganization,
    db: Session,
    cloud_provider: str = "all",
) -> Optional[Dict[str, Any]]:
    return build_imported_cost_context(
        membership,
        db,
        cloud_provider=cloud_provider,
        organization_id_for_membership=_organization_id_for_membership,
        customer_id_for_org=_customer_id_for_org,
        get_imported_cost_rows=_get_imported_cost_rows,
        imported_cost_summary=_imported_cost_summary,
    )


def _cost_snapshot_context(
    membership: UserOrganization,
    db: Session,
    cloud_provider: str = "all",
    provider_errors: Optional[Dict[str, str]] = None,
) -> Optional[Dict[str, Any]]:
    """Build real cost context from the latest persisted live scan snapshots."""
    customer_id = _customer_id_for_org(membership)
    query = db.query(CostSnapshot).filter(CostSnapshot.customer_id == customer_id)
    if cloud_provider != "all":
        query = query.filter(CostSnapshot.provider == cloud_provider)
    rows = query.order_by(CostSnapshot.captured_at.desc()).limit(500).all()
    if not rows:
        return None

    latest_by_provider: Dict[str, CostSnapshot] = {}
    for row in rows:
        provider = str(row.provider or "").strip().lower()
        if not provider or provider in latest_by_provider:
            continue
        latest_by_provider[provider] = row

    if not latest_by_provider:
        return None

    total_cost = 0.0
    breakdown: Dict[str, Dict[str, float]] = {}
    scan_ids: list[str] = []
    latest_captured_at: Optional[datetime] = None
    for provider, row in latest_by_provider.items():
        cost = float(row.total_cost_usd or 0.0)
        total_cost += cost
        breakdown[provider] = {"cost": round(cost, 2), "percentage": 0.0}
        if row.scan_id:
            scan_ids.append(str(row.scan_id))
        if row.captured_at and (latest_captured_at is None or row.captured_at > latest_captured_at):
            latest_captured_at = row.captured_at

    if total_cost > 0:
        for provider in breakdown:
            breakdown[provider]["percentage"] = round((breakdown[provider]["cost"] / total_cost) * 100, 1)

    region_totals: Dict[str, float] = {}
    if scan_ids:
        allocation_query = db.query(CostAllocationSnapshot).filter(
            CostAllocationSnapshot.customer_id == customer_id,
            CostAllocationSnapshot.scan_id.in_(scan_ids),
        )
        if cloud_provider != "all":
            allocation_query = allocation_query.filter(CostAllocationSnapshot.provider == cloud_provider)
        for allocation in allocation_query.limit(2000).all():
            region = str(allocation.region or "global")
            region_totals[region] = region_totals.get(region, 0.0) + float(allocation.cost_usd or 0.0)

    response: Dict[str, Any] = {
        "period": "snapshot",
        "cloud_provider": cloud_provider,
        "total_cost": round(total_cost, 2),
        "breakdown": breakdown,
        "region_breakdown": [
            {"region": region, "cost_usd": round(cost, 2)}
            for region, cost in sorted(region_totals.items(), key=lambda item: item[1], reverse=True)
        ],
        "source": "cost_snapshots_live",
        "rows_imported": 0,
        "last_imported_at": None,
        "last_captured_at": latest_captured_at.isoformat() if latest_captured_at else None,
        "scan_ids": scan_ids,
    }
    if provider_errors:
        response["provider_errors"] = provider_errors
    return response


def _valid_runtime_provider_names(customer_id: str, db: Session) -> set[str]:
    runtime_credentials = _load_runtime_provider_credentials(customer_id)
    rows = (
        db.query(CredentialRecord)
        .filter(
            CredentialRecord.customer_id == customer_id,
            CredentialRecord.is_active.is_(True),
            CredentialRecord.is_valid.is_(True),
            CredentialRecord.provider.in_(list(SUPPORTED_CLOUD_PROVIDERS)),
        )
        .all()
    )
    providers: set[str] = set()
    for row in rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        runtime = runtime_credentials.get(provider)
        if isinstance(runtime, dict) and runtime:
            providers.add(provider)
            continue
        if provider == "oci":
            try:
                payload = json.loads(row.credential_json or "{}")
            except Exception:
                payload = {}
            config_file = str(payload.get("config_file") or "").strip()
            if config_file and os.path.isfile(os.path.expanduser(config_file)):
                providers.add(provider)
    return providers


def _ordered_valid_runtime_provider_names(customer_id: str, db: Session) -> list[str]:
    """Return configured providers in product order for automatic scans."""
    valid_providers = _valid_runtime_provider_names(customer_id, db)
    return [provider for provider in SUPPORTED_CLOUD_PROVIDERS if provider in valid_providers]


def _mark_provider_credentials_unreachable(
    db: Session,
    customer_id: str,
    provider: str,
    message: str,
) -> None:
    row = (
        db.query(CredentialRecord)
        .filter(
            CredentialRecord.customer_id == customer_id,
            CredentialRecord.provider == provider,
        )
        .first()
    )
    if row is None:
        return
    row.is_valid = False
    row.is_active = False
    row.validation_message = f"Credential became unreachable during live scan: {message}"
    row.tested_at = _utcnow()
    row.updated_at = _utcnow()


async def _cost_context(
    membership: UserOrganization,
    db: Session,
    period: str = "month",
    cloud_provider: str = "all",
) -> Dict[str, Any]:
    try:
        context = await build_live_cost_context(
            membership,
            db,
            period=period,
            cloud_provider=cloud_provider,
            require_live_provider_data=Config().require_live_provider_data,
            provider_diagnostics=lambda: _provider_diagnostics(
                customer_id=_customer_id_for_org(membership),
                db=db,
            ),
            imported_cost_context_builder=_imported_cost_context,
            cost_summary_for_provider=lambda provider, requested_period: _cost_summary_for_provider(
                provider,
                requested_period,
                customer_id=_customer_id_for_org(membership),
            ),
        )
        if context.get("no_data") or context.get("source") == "no_data_available":
            snapshot_context = _cost_snapshot_context(
                membership,
                db,
                cloud_provider=cloud_provider,
                provider_errors=context.get("provider_errors"),
            )
            if snapshot_context is not None:
                return snapshot_context
        return context
    except LiveDataPolicyError as exc:
        snapshot_context = _cost_snapshot_context(
            membership,
            db,
            cloud_provider=cloud_provider,
            provider_errors={"live_provider_api": str(exc)},
        )
        if snapshot_context is not None:
            return snapshot_context
        raise HTTPException(status_code=412, detail=str(exc))


def _historical_monthly_spend_from_snapshots(
    db: Session,
    customer_id: str,
    cloud_provider: str,
    months: int = 18,
) -> List[float]:
    """Build monthly spend history from persisted cost snapshots.

    Returns ascending monthly totals (oldest -> newest).
    """
    query = db.query(CostSnapshot).filter(CostSnapshot.customer_id == customer_id)
    if cloud_provider != "all":
        query = query.filter(CostSnapshot.provider == cloud_provider)
    rows = query.order_by(CostSnapshot.captured_at.desc()).limit(500).all()

    latest_by_month_provider: Dict[tuple[str, str], CostSnapshot] = {}
    for row in rows:
        anchor = row.period_end or row.period_start or row.captured_at
        if anchor is None:
            continue
        month_key = anchor.strftime("%Y-%m")
        provider_key = str(row.provider or "unknown").lower()
        key = (month_key, provider_key)
        existing = latest_by_month_provider.get(key)
        existing_anchor = (
            existing.period_end or existing.period_start or existing.captured_at
            if existing is not None
            else None
        )
        if existing is None or (existing_anchor is not None and anchor > existing_anchor):
            latest_by_month_provider[key] = row

    month_totals: Dict[str, float] = {}
    for (month_key, _), row in latest_by_month_provider.items():
        month_totals[month_key] = month_totals.get(month_key, 0.0) + float(row.total_cost_usd or 0.0)
    ordered = sorted(month_totals.items())
    return [round(value, 2) for _, value in ordered[-months:]]


def _setting_missing(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return True
    return text.startswith("your_") or text.startswith("replace_") or "example.com" in text


def _provider_diagnostics(
    customer_id: Optional[str] = None,
    db: Optional[Session] = None,
) -> List[ProviderDiagnostic]:
    config = Config()
    requirements = provider_diagnostic_requirements(config)
    capabilities = provider_api_capabilities()
    runtime_ready = _valid_runtime_provider_names(customer_id, db) if customer_id and db else set()

    diagnostics: List[ProviderDiagnostic] = []
    for provider, detail in requirements.items():
        settings = detail["settings"]
        values = detail["values"]
        missing = [setting for setting, value in zip(settings, values) if _setting_missing(value)]
        configured = not missing
        if configured and provider == "gcp":
            credentials_file = os.path.abspath(
                os.path.expanduser(
                    os.path.expandvars(str(config.google_application_credentials or ""))
                )
            )
            if credentials_file and not os.path.isfile(credentials_file):
                missing.append(
                    f"GOOGLE_APPLICATION_CREDENTIALS file not found: {credentials_file}"
                )
                configured = False
        if configured and provider == "oci":
            oci_config_file = os.path.abspath(
                os.path.expanduser(os.path.expandvars(str(config.oci_config_file or "")))
            )
            if oci_config_file and not os.path.isfile(oci_config_file):
                missing.append(f"OCI_CONFIG_FILE file not found: {oci_config_file}")
                configured = False
        if provider in runtime_ready:
            configured = True
            missing = []
        capability = capabilities[provider]
        diagnostics.append(
            ProviderDiagnostic(
                provider=provider,
                configured=configured,
                required_settings=settings,
                missing_settings=missing,
                recommendation=(
                    "Ready for live billing API calls from saved customer credentials."
                    if provider in runtime_ready
                    else "Ready for live billing API calls."
                    if configured
                    else f"Configure {', '.join(missing)} before enabling live {provider.upper()} cost collection."
                ),
                scope_model=capability.scope_model,
                primary_apis=capability.primary_apis,
                optimization_apis=capability.optimization_apis,
                telemetry_apis=capability.telemetry_apis,
                default_page_size=capability.default_page_size,
                max_page_size=capability.max_page_size,
                max_parallel_requests=capability.max_parallel_requests,
                request_timeout_seconds=capability.request_timeout_seconds,
                retryable_statuses=capability.retryable_statuses,
                throttling_signals=capability.throttling_signals,
                scan_notes=capability.scan_notes,
            )
        )
    return diagnostics


@router.post("/credentials/validate", response_model=CredentialResponse)
async def validate_credentials(
    payload: Dict[str, Any],
    current_user: User = Depends(get_current_user),
) -> CredentialResponse:
    """Validate cloud credentials without storing them."""
    _ = current_user
    try:
        credential = _parse_credential_payload(payload)
        result = _run_validation(credential)
        return CredentialResponse(**result.__dict__)
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Credential validation failed")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/credentials/oci/upload-files")
async def upload_oci_credential_files(
    profile: str = Form("DEFAULT"),
    config_file: UploadFile = File(...),
    private_key_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Upload OCI config/key files to the API host for test-only credential checks."""
    _require_management_role(membership, "OCI credential file upload")
    customer_id = _resolve_customer_id(current_user, membership, None)
    normalized_profile = _normalize_oci_profile(profile)

    if config_file is None or not (config_file.filename or "").strip():
        raise HTTPException(status_code=400, detail="config_file is required")

    config_bytes = await config_file.read()
    if not config_bytes:
        raise HTTPException(status_code=400, detail="Uploaded OCI config file is empty")
    if len(config_bytes) > 1024 * 1024:
        raise HTTPException(status_code=400, detail="OCI config file is too large (max 1 MiB)")

    try:
        config_text = config_bytes.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=400,
            detail="OCI config file must be UTF-8 text",
        ) from exc

    parser = configparser.RawConfigParser()
    parser.optionxform = str
    try:
        parser.read_string(config_text)
    except configparser.Error as exc:
        raise HTTPException(status_code=400, detail=f"Invalid OCI config format: {exc}") from exc

    sections = parser.sections()
    if not sections:
        raise HTTPException(status_code=400, detail="OCI config file has no profile sections")
    if normalized_profile not in sections:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Profile '{normalized_profile}' was not found in uploaded OCI config. "
                f"Available profiles: {', '.join(sections)}"
            ),
        )

    target_dir = _oci_uploaded_credentials_dir(customer_id)
    token = uuid4().hex[:12]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    config_path = os.path.join(target_dir, f"oci-config-{timestamp}-{token}.ini")
    key_path = ""

    if private_key_file and (private_key_file.filename or "").strip():
        key_bytes = await private_key_file.read()
        if not key_bytes:
            raise HTTPException(status_code=400, detail="Uploaded OCI private key file is empty")
        if len(key_bytes) > 256 * 1024:
            raise HTTPException(status_code=400, detail="OCI private key file is too large (max 256 KiB)")
        key_path = os.path.join(target_dir, f"oci-key-{timestamp}-{token}.pem")
        with open(key_path, "wb") as fh:
            fh.write(key_bytes)
        try:
            os.chmod(key_path, 0o600)
        except OSError:
            pass
        parser.set(normalized_profile, "key_file", key_path)

    with open(config_path, "w", encoding="utf-8") as fh:
        parser.write(fh, space_around_delimiters=False)
    try:
        os.chmod(config_path, 0o600)
    except OSError:
        pass

    return {
        "status": "success",
        "message": "OCI credential files uploaded to API host for test usage.",
        "customer_id": customer_id,
        "profile": normalized_profile,
        "config_file": config_path,
        "key_file": key_path or None,
        "profiles_available": sections,
        "test_only": True,
    }


@router.post("/credentials/add")
async def add_credentials(
    payload: Dict[str, Any],
    background_tasks: BackgroundTasks,
    credential_manager: CredentialManager = Depends(get_credential_manager),
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Validate, persist, and immediately scan live provider credentials."""
    try:
        _require_management_role(membership, "credential updates")
        customer_id = _resolve_customer_id(
            current_user=current_user,
            membership=membership,
            requested_customer_id=payload.get("customer_id"),
        )
        credential_payload = {k: v for k, v in payload.items() if k != "customer_id"}
        credential = _parse_credential_payload(credential_payload)
        validation = _run_validation(credential)
        if not validation.is_valid:
            raise HTTPException(
                status_code=400,
                detail=f"Credential validation failed: {validation.message}",
            )

        stored = credential_manager.store_credentials(
            customer_id=customer_id,
            provider=credential.provider,
            credentials=credential.model_dump(),
            is_active=True,
            validation=validation,
        )
        runtime = _write_runtime_provider_credentials(
            customer_id=customer_id,
            credential=credential,
        )
        permission = scanning_manager.get_permission_status(customer_id)
        scan_providers = _ordered_valid_runtime_provider_names(customer_id, scanning_manager.db)
        if credential.provider not in scan_providers:
            scan_providers = list(dict.fromkeys([*scan_providers, credential.provider]))
        scanning_manager.approve_scanning(
            customer_id=customer_id,
            auto_remediate=bool(permission.get("auto_remediate", False)),
            scan_frequency=str(permission.get("scan_frequency") or "daily"),
            notification_email=permission.get("notification_email"),
            monthly_budget_usd=float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            warning_threshold_percent=float(permission.get("warning_threshold_percent", 80.0) or 80.0),
            critical_threshold_percent=float(permission.get("critical_threshold_percent", 100.0) or 100.0),
            notifications_enabled=bool(permission.get("notifications_enabled", True)),
            scheduler_override_enabled=bool(permission.get("scheduler_override_enabled", False)),
            scheduler_override_frequency=permission.get("scheduler_override_frequency"),
            scheduler_retry_max_attempts=int(permission.get("scheduler_retry_max_attempts", 2) or 2),
            scheduler_retry_backoff_seconds=int(permission.get("scheduler_retry_backoff_seconds", 120) or 120),
            scheduler_overdue_alert_hours=int(permission.get("scheduler_overdue_alert_hours", 24) or 24),
        )
        permission_row = (
            scanning_manager.db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        if permission_row is not None:
            permission_row.providers_json = json.dumps(scan_providers)
            permission_row.updated_at = _utcnow()
            scanning_manager.db.commit()

        scan_id = f"scan_{customer_id}_{int(_utcnow().timestamp())}_{uuid4().hex[:8]}"
        scanning_manager.create_scan_run(scan_id, customer_id, scan_providers)
        background_tasks.add_task(
            _run_cost_analysis,
            scan_id=scan_id,
            customer_id=customer_id,
            providers=scan_providers,
        )
        return {
            "status": "success",
            "message": (
                f"{credential.provider.upper()} credentials stored and live provider scan started "
                f"for {', '.join(provider.upper() for provider in scan_providers)}"
            ),
            "provider": credential.provider,
            "customer_id": customer_id,
            "record": stored,
            "runtime": runtime,
            "scan": {
                "scan_id": scan_id,
                "state": ScanningState.RUNNING.value,
                "providers": scan_providers,
            },
        }
    except HTTPException:
        raise
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Failed to add credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/credentials")
async def list_credentials(
    customer_id: Optional[str] = None,
    credential_manager: CredentialManager = Depends(get_credential_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        payload = credential_manager.list_credentials(scoped_customer_id)
        payload["organization_id"] = _organization_id_for_membership(membership)
        return payload
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to list credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.delete("/credentials/{provider}")
async def delete_credentials(
    provider: str,
    customer_id: Optional[str] = None,
    credential_manager: CredentialManager = Depends(get_credential_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "credential deletion")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        deleted = credential_manager.delete_credentials(scoped_customer_id, provider)
        if not deleted:
            raise HTTPException(status_code=404, detail="Credential not found")
        runtime_provider_dir = os.path.join(
            _runtime_credentials_root(scoped_customer_id),
            provider.lower(),
        )
        shutil.rmtree(runtime_provider_dir, ignore_errors=True)
        return {
            "status": "success",
            "message": f"{provider.upper()} credentials deleted",
            "provider": provider,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to delete credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/request-approval")
async def request_scanning_approval(
    providers: List[str],
    notification_email: str,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    customer_id: Optional[str] = None,
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan approval requests")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        scanning_manager.create_permission_request(
            customer_id=scoped_customer_id,
            providers=providers,
            notification_email=notification_email,
        )
        approval_request = scanning_manager.request_approval(
            customer_id=scoped_customer_id,
            providers=providers,
        )
        return {
            "status": "approval_pending",
            "message": approval_request["message"],
            "action_required": True,
            "approve_url": approval_request["approve_url"],
            "providers": providers,
            "customer_id": scoped_customer_id,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to request approval")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/approve", response_model=ScanningPermissionResponse)
async def approve_scanning(
    approval: ScanningApprovalRequest,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanningPermissionResponse:
    try:
        _require_management_role(membership, "scan approval")
        customer_id = _resolve_customer_id(current_user, membership, approval.customer_id)
        approved = scanning_manager.approve_scanning(
            customer_id=customer_id,
            auto_remediate=False,
            scan_frequency=approval.scan_frequency,
            notification_email=approval.notification_email,
            monthly_budget_usd=approval.monthly_budget_usd,
            warning_threshold_percent=approval.warning_threshold_percent,
            critical_threshold_percent=approval.critical_threshold_percent,
            notifications_enabled=approval.notifications_enabled,
            scheduler_override_enabled=approval.scheduler_override_enabled,
            scheduler_override_frequency=approval.scheduler_override_frequency,
            scheduler_retry_max_attempts=approval.scheduler_retry_max_attempts,
            scheduler_retry_backoff_seconds=approval.scheduler_retry_backoff_seconds,
            scheduler_overdue_alert_hours=approval.scheduler_overdue_alert_hours,
        )
        return ScanningPermissionResponse(
            customer_id=approved["customer_id"],
            organization_id=_organization_id_for_membership(membership),
            state=approved["state"],
            providers=approved["providers"],
            scan_frequency=approved["scan_frequency"],
            auto_remediate=approved["auto_remediate"],
            notification_email=approved.get("notification_email"),
            monthly_budget_usd=float(approved.get("monthly_budget_usd", 0.0) or 0.0),
            warning_threshold_percent=float(approved.get("warning_threshold_percent", 80.0) or 80.0),
            critical_threshold_percent=float(approved.get("critical_threshold_percent", 100.0) or 100.0),
            notifications_enabled=bool(approved.get("notifications_enabled", True)),
            scheduler_override_enabled=bool(approved.get("scheduler_override_enabled", False)),
            scheduler_override_frequency=approved.get("scheduler_override_frequency"),
            scheduler_retry_max_attempts=max(1, int(approved.get("scheduler_retry_max_attempts", 1) or 1)),
            scheduler_retry_backoff_seconds=max(15, int(approved.get("scheduler_retry_backoff_seconds", 15) or 15)),
            scheduler_overdue_alert_hours=max(1, int(approved.get("scheduler_overdue_alert_hours", 24) or 24)),
            created_at=approved["created_at"],
            approved_at=approved["approved_at"],
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to approve scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/scanning/permission", response_model=ScanningPermissionResponse)
async def get_scanning_permission(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanningPermissionResponse:
    try:
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        permission = scanning_manager.get_permission_status(scoped_customer_id)
        return ScanningPermissionResponse(
            organization_id=_organization_id_for_membership(membership),
            **permission,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to get permission status")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/pause")
async def pause_scanning(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan pause")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        return scanning_manager.pause_scanning(scoped_customer_id)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to pause scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/resume")
async def resume_scanning(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan resume")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        return scanning_manager.resume_scanning(scoped_customer_id)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to resume scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/start", response_model=ScanProgressResponse)
async def start_scan(
    scan_request: StartScanRequest,
    background_tasks: BackgroundTasks,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanProgressResponse:
    try:
        _require_management_role(membership, "scan start")
        customer_id = _resolve_customer_id(current_user, membership, scan_request.customer_id)
        permission = scanning_manager.get_permission_status(customer_id)
        if permission["state"] not in [ScanningState.APPROVED.value, ScanningState.RUNNING.value]:
            raise HTTPException(
                status_code=403,
                detail=f"Scanning not approved. Current state: {permission['state']}",
            )

        providers_to_scan = scan_request.providers or permission["providers"]
        if not providers_to_scan:
            providers_to_scan = ["aws", "azure", "gcp", "oci"]
        providers_to_scan = [str(provider or "").strip().lower() for provider in providers_to_scan if str(provider or "").strip()]
        providers_to_scan = list(dict.fromkeys(providers_to_scan))
        invalid_providers = [provider for provider in providers_to_scan if provider not in SUPPORTED_CLOUD_PROVIDERS]
        if invalid_providers:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported provider(s): {', '.join(sorted(set(invalid_providers)))}",
            )
        if Config().require_live_provider_data:
            diagnostics_map = {
                d.provider: bool(d.configured)
                for d in _provider_diagnostics(customer_id=customer_id, db=scanning_manager.db)
            }
            configured_targets = [provider for provider in providers_to_scan if diagnostics_map.get(provider, False)]
            if not configured_targets:
                raise HTTPException(
                    status_code=412,
                    detail=(
                        "Live provider data is required, but none of the requested providers are configured "
                        "for runtime API access on this backend host."
                    ),
                )
            providers_to_scan = configured_targets

        scan_id = f"scan_{customer_id}_{int(_utcnow().timestamp())}"
        scanning_manager.create_scan_run(scan_id, customer_id, providers_to_scan)
        background_tasks.add_task(
            _run_cost_analysis,
            scan_id=scan_id,
            customer_id=customer_id,
            providers=providers_to_scan,
            target_accounts=scan_request.target_accounts or None,
        )

        return ScanProgressResponse(
            scan_id=scan_id,
            customer_id=customer_id,
            organization_id=_organization_id_for_membership(membership),
            state=ScanningState.RUNNING.value,
            progress=0,
            providers=providers_to_scan,
            started_at=_utcnow(),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to start scan")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/scheduler/run-now")
async def run_scheduler_now(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Run the scan scheduler loop once on-demand."""
    _ = current_user
    _require_management_role(membership, "manual scheduler runs")
    return await run_scheduled_scans_once(
        requested_organization_id=_organization_id_for_membership(membership),
        sleep_between_retries=False,
    )


@router.patch("/scanning/scheduler/policy", response_model=ScanningPermissionResponse)
async def update_scheduler_policy(
    payload: SchedulerPolicyUpdateRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanningPermissionResponse:
    _require_management_role(membership, "scheduler policy updates")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()
    db = SessionLocal()
    try:
        row = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        if row is None:
            row = ScanningPermissionRecord(
                customer_id=customer_id,
                state=ScanningState.INITIALIZED.value,
                providers_json="[]",
                scan_frequency="daily",
                auto_remediate=False,
                monthly_budget_usd=0.0,
                warning_threshold_percent=80.0,
                critical_threshold_percent=100.0,
                notifications_enabled=True,
                created_at=now,
                updated_at=now,
            )
        row.scheduler_override_enabled = bool(payload.scheduler_override_enabled)
        row.scheduler_override_frequency = (
            str(payload.scheduler_override_frequency or "").strip().lower() or None
        )
        row.scheduler_retry_max_attempts = int(payload.scheduler_retry_max_attempts)
        row.scheduler_retry_backoff_seconds = int(payload.scheduler_retry_backoff_seconds)
        row.scheduler_overdue_alert_hours = int(payload.scheduler_overdue_alert_hours)
        row.updated_at = now
        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="scan.scheduler.policy.update",
                entity_type="scanning_permission",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "scheduler_override_enabled": bool(row.scheduler_override_enabled),
                        "scheduler_override_frequency": row.scheduler_override_frequency,
                        "scheduler_retry_max_attempts": int(row.scheduler_retry_max_attempts or 1),
                        "scheduler_retry_backoff_seconds": int(row.scheduler_retry_backoff_seconds or 15),
                        "scheduler_overdue_alert_hours": int(row.scheduler_overdue_alert_hours or 24),
                    }
                ),
                created_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    db_read = SessionLocal()
    try:
        permission = ScanningManager(db_read).get_permission_status(customer_id)
    finally:
        db_read.close()
    return ScanningPermissionResponse(
        organization_id=organization_id,
        **permission,
    )


@router.get("/scanning/{scan_id}/progress", response_model=ScanProgressResponse)
async def get_scan_progress(
    scan_id: str,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanProgressResponse:
    row = scanning_manager.get_scan_run(scan_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    if row["customer_id"] != _customer_id_for_org(membership):
        raise HTTPException(status_code=404, detail="Scan not found")
    return ScanProgressResponse(
        scan_id=row["scan_id"],
        customer_id=row["customer_id"],
        organization_id=_organization_id_for_membership(membership),
        state=row["state"],
        progress=row["progress"],
        providers=row["providers"],
        started_at=row["started_at"],
        completed_at=row["completed_at"],
        total_resources=row["total_resources"],
        anomalies_found=row["anomalies_found"],
        savings_identified=row["savings_identified"],
    )


@router.get("/scanning/history", response_model=List[ScanHistoryItem])
async def get_scan_history(
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ScanHistoryItem]:
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .offset(max(0, min(offset, 50000)))
            .limit(max(1, min(limit, 100)))
            .all()
        )
    finally:
        db.close()
    return [
        ScanHistoryItem(
            scan_id=row.scan_id,
            customer_id=row.customer_id,
            organization_id=_organization_id_for_membership(membership),
            state=row.state,
            providers=json.loads(row.providers_json or "[]"),
            started_at=row.started_at,
            completed_at=row.completed_at,
            total_resources=row.total_resources,
            anomalies_found=row.anomalies_found,
            savings_identified=float(row.savings_identified or 0.0),
        )
        for row in rows
    ]


@router.get("/scanning/{scan_id}/diff", response_model=ScanDiffResponse)
async def get_scan_diff(
    scan_id: str,
    previous_scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanDiffResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        current_run = (
            db.query(ScanRunRecord)
            .filter(
                ScanRunRecord.scan_id == scan_id,
                ScanRunRecord.customer_id == customer_id,
            )
            .first()
        )
        if current_run is None:
            raise HTTPException(status_code=404, detail="Scan not found")

        if previous_scan_id:
            previous_run = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.scan_id == previous_scan_id,
                    ScanRunRecord.customer_id == customer_id,
                )
                .first()
            )
        else:
            previous_run = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.customer_id == customer_id,
                    ScanRunRecord.state == ScanningState.COMPLETED.value,
                    ScanRunRecord.scan_id != scan_id,
                    ScanRunRecord.started_at < current_run.started_at,
                )
                .order_by(ScanRunRecord.started_at.desc())
                .first()
            )

        current_snapshots = (
            db.query(CostSnapshot).filter(CostSnapshot.scan_id == current_run.scan_id).all()
        )
        previous_snapshots = (
            db.query(CostSnapshot).filter(CostSnapshot.scan_id == previous_run.scan_id).all()
            if previous_run
            else []
        )
    finally:
        db.close()

    current_map = {row.provider: row for row in current_snapshots}
    previous_map = {row.provider: row for row in previous_snapshots}
    providers = sorted(set(current_map.keys()) | set(previous_map.keys()))

    entries: List[ScanDiffEntry] = []
    total_current = 0.0
    total_previous = 0.0
    for provider in providers:
        current_row = current_map.get(provider)
        previous_row = previous_map.get(provider)
        current_cost = float(current_row.total_cost_usd) if current_row else 0.0
        previous_cost = float(previous_row.total_cost_usd) if previous_row else 0.0
        total_current += current_cost
        total_previous += previous_cost
        delta_cost = current_cost - previous_cost
        delta_percent = None
        if previous_cost > 0:
            delta_percent = round((delta_cost / previous_cost) * 100, 2)
        entries.append(
            ScanDiffEntry(
                provider=provider,
                current_cost_usd=round(current_cost, 2),
                previous_cost_usd=round(previous_cost, 2),
                delta_cost_usd=round(delta_cost, 2),
                delta_percent=delta_percent,
                current_anomalies=int(current_row.anomalies_count if current_row else 0),
                previous_anomalies=int(previous_row.anomalies_count if previous_row else 0),
            )
        )

    return ScanDiffResponse(
        organization_id=organization_id,
        current_scan_id=current_run.scan_id,
        previous_scan_id=previous_run.scan_id if previous_run else None,
        total_current_cost_usd=round(total_current, 2),
        total_previous_cost_usd=round(total_previous, 2),
        total_delta_cost_usd=round(total_current - total_previous, 2),
        entries=entries,
    )


@router.get("/scanning/history.csv")
async def download_scan_history_csv(
    limit: int = 100,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await get_scan_history(
        limit=limit,
        offset=offset,
        current_user=current_user,
        membership=membership,
    )
    lines = ["scan_id,state,providers,started_at,completed_at,total_resources,anomalies_found,savings_identified"]
    for row in rows:
        providers = "|".join(row.providers)
        lines.append(
            f"{row.scan_id},{row.state},{providers},{row.started_at.isoformat()},"
            f"{row.completed_at.isoformat() if row.completed_at else ''},"
            f"{row.total_resources},{row.anomalies_found},{row.savings_identified:.2f}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/scanning/{scan_id}/diff.csv")
async def download_scan_diff_csv(
    scan_id: str,
    previous_scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    diff = await get_scan_diff(
        scan_id=scan_id,
        previous_scan_id=previous_scan_id,
        current_user=current_user,
        membership=membership,
    )
    lines = ["provider,current_cost_usd,previous_cost_usd,delta_cost_usd,delta_percent,current_anomalies,previous_anomalies"]
    for entry in diff.entries:
        lines.append(
            f"{entry.provider},{entry.current_cost_usd:.2f},{entry.previous_cost_usd:.2f},"
            f"{entry.delta_cost_usd:.2f},{entry.delta_percent if entry.delta_percent is not None else ''},"
            f"{entry.current_anomalies},{entry.previous_anomalies}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/provider-accounts/rollups", response_model=ProviderAccountRollupResponse)
async def get_provider_account_rollups(
    scan_id: Optional[str] = None,
    provider: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ProviderAccountRollupResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    budget_monthly_usd = 0.0
    warning_threshold_percent = 80.0
    critical_threshold_percent = 100.0
    try:
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .order_by(ScanningPermissionRecord.created_at.desc())
            .first()
        )
        if permission is not None:
            budget_monthly_usd = float(permission.monthly_budget_usd or 0.0)
            warning_threshold_percent = float(permission.warning_threshold_percent or 80.0)
            critical_threshold_percent = float(permission.critical_threshold_percent or 100.0)

        imported_rows = _get_imported_cost_rows(
            db,
            organization_id=organization_id,
            customer_id=customer_id,
            cloud_provider=provider or "all",
        )

        query = (
            db.query(ProviderAccountSnapshot, ProviderAccount)
            .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
            .filter(
                ProviderAccountSnapshot.customer_id == customer_id,
                ProviderAccountSnapshot.organization_id == organization_id,
            )
        )
        if scan_id:
            query = query.filter(ProviderAccountSnapshot.scan_id == scan_id)
        else:
            latest = (
                db.query(ProviderAccountSnapshot.scan_id)
                .filter(
                    ProviderAccountSnapshot.customer_id == customer_id,
                    ProviderAccountSnapshot.organization_id == organization_id,
                )
                .order_by(ProviderAccountSnapshot.captured_at.desc())
                .first()
            )
            if latest:
                query = query.filter(ProviderAccountSnapshot.scan_id == latest[0])
                scan_id = latest[0]
        if provider:
            query = query.filter(ProviderAccount.provider == provider)
        rows = query.all()

        if imported_rows:
            return _build_rollups_from_imported_rows(
                imported_rows,
                organization_id=organization_id,
                customer_id=customer_id,
                provider=provider,
                monthly_budget_usd=budget_monthly_usd,
                warning_threshold_percent=warning_threshold_percent,
                critical_threshold_percent=critical_threshold_percent,
            )

        account_ids = [account.id for _, account in rows]
        links = (
            db.query(ProviderAccountLink)
            .filter(
                ProviderAccountLink.organization_id == organization_id,
                ProviderAccountLink.child_account_id.in_(account_ids),
            )
            .all()
            if account_ids
            else []
        )
    finally:
        db.close()

    nodes: Dict[int, Dict[str, Any]] = {}
    provider_roots: Dict[str, int] = {}
    next_synthetic_id = -1

    def _synthetic_id() -> int:
        nonlocal next_synthetic_id
        current = next_synthetic_id
        next_synthetic_id -= 1
        return current

    account_by_id = {account.id: account for _, account in rows}
    parent_by_child = {link.child_account_id: link.parent_account_id for link in links}

    for snapshot, account in rows:
        nodes[account.id] = {
            "provider": account.provider,
            "account_identifier": account.account_identifier,
            "account_name": account.account_name,
            "account_type": account.account_type,
            "parent_account_id": parent_by_child.get(account.id),
            "parent_account_identifier": (
                account_by_id[parent_by_child[account.id]].account_identifier
                if parent_by_child.get(account.id) in account_by_id
                else None
            ),
            "direct_cost_usd": round(float(snapshot.direct_cost_usd or 0.0), 2),
            "direct_savings_identified_usd": round(float(snapshot.savings_identified_usd or 0.0), 2),
            "direct_anomalies_count": int(snapshot.anomalies_count or 0),
            "direct_service_count": int(snapshot.service_count or 0),
            "scan_id": snapshot.scan_id,
            "captured_at": snapshot.captured_at.isoformat() if snapshot.captured_at else None,
        }

    for node in nodes.values():
        if node.get("parent_account_id") not in nodes:
            node["parent_account_id"] = None
            node["parent_account_identifier"] = None

    management_group_by_provider: Dict[str, int] = {}
    for account_id, node in nodes.items():
        if node["account_type"] == "management_group":
            management_group_by_provider[node["provider"]] = account_id

    for account_id, node in list(nodes.items()):
        provider_key = node["provider"]
        root_id = provider_roots.get(provider_key)
        if root_id is None:
            root_id = _synthetic_id()
            provider_roots[provider_key] = root_id
            nodes[root_id] = {
                "provider": provider_key,
                "account_identifier": f"{provider_key}:provider",
                "account_name": provider_key.upper(),
                "account_type": "provider",
                "parent_account_id": None,
                "parent_account_identifier": None,
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": node.get("scan_id"),
                "captured_at": node.get("captured_at"),
            }

        if node.get("parent_account_id") is not None:
            continue

        inferred_parent_id = root_id
        if node["account_type"] == "management_group":
            inferred_parent_id = root_id
        elif node["account_type"] in {"subscription", "account"} and provider_key == "azure":
            inferred_parent_id = management_group_by_provider.get(provider_key, root_id)
        elif node["account_type"] == "tenancy":
            inferred_parent_id = root_id
        elif node["account_type"] in {"project", "compartment"}:
            inferred_parent_id = root_id

        node["parent_account_id"] = inferred_parent_id if inferred_parent_id != account_id else None
        node["parent_account_identifier"] = (
            nodes[inferred_parent_id]["account_identifier"]
            if inferred_parent_id in nodes and inferred_parent_id != account_id
            else None
        )

    # Attach top regions to each node from CostAllocationSnapshot.
    if scan_id and account_by_id:
        db2 = SessionLocal()
        try:
            alloc_rows = (
                db2.query(CostAllocationSnapshot)
                .filter(
                    CostAllocationSnapshot.scan_id == scan_id,
                    CostAllocationSnapshot.provider_account_id.in_(list(account_by_id.keys())),
                )
                .order_by(CostAllocationSnapshot.cost_usd.desc())
                .all()
            )
        finally:
            db2.close()
        for alloc in alloc_rows:
            node = nodes.get(alloc.provider_account_id)
            if node is None:
                continue
            if "top_regions" not in node:
                node["top_regions"] = []
            if len(node["top_regions"]) < 5:
                node["top_regions"].append(
                    AccountRegionRow(region=alloc.region, cost_usd=round(float(alloc.cost_usd), 2))
                )

    items = _materialize_rollup_items(nodes)
    filtered_items = [item for item in items if provider is None or item.provider == provider]
    filtered_items = _apply_rollup_budget_metrics(
        filtered_items,
        monthly_budget_usd=budget_monthly_usd,
        warning_threshold_percent=warning_threshold_percent,
        critical_threshold_percent=critical_threshold_percent,
    )
    total_direct = round(sum(item.direct_cost_usd for item in filtered_items if item.depth > 0 or item.child_count == 0), 2)
    total_rolled = round(sum(item.rolled_up_cost_usd for item in filtered_items if item.depth == 0), 2)

    return ProviderAccountRollupResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        provider=provider,
        scan_id=scan_id,
        generated_at=_utcnow().isoformat(),
        total_direct_cost_usd=total_direct,
        total_rolled_up_cost_usd=total_rolled,
        items=filtered_items,
    )


@router.get("/provider-accounts", response_model=ProviderAccountInventoryResponse)
async def get_provider_account_inventory(
    provider: Optional[str] = None,
    account_type: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ProviderAccountInventoryResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        query = db.query(ProviderAccount).filter(
            ProviderAccount.customer_id == customer_id,
            ProviderAccount.organization_id == organization_id,
            ProviderAccount.is_active.is_(True),
        )
        if provider:
            query = query.filter(ProviderAccount.provider == provider)
        if account_type:
            query = query.filter(ProviderAccount.account_type == account_type)
        accounts = query.order_by(ProviderAccount.provider, ProviderAccount.account_name).all()
    finally:
        db.close()

    items = [
        ProviderAccountInventoryItem(
            account_id=acc.id,
            provider=acc.provider,
            account_identifier=acc.account_identifier,
            account_name=acc.account_name,
            account_type=acc.account_type,
            native_region=acc.native_region,
            is_active=bool(acc.is_active),
            metadata=_safe_json_load(acc.metadata_json, {}),
            created_at=acc.created_at.isoformat() if acc.created_at else "",
            updated_at=acc.updated_at.isoformat() if acc.updated_at else None,
        )
        for acc in accounts
    ]
    return ProviderAccountInventoryResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        total=len(items),
        accounts=items,
    )


@router.get("/provider-accounts/{account_id}/region-breakdown", response_model=AccountRegionBreakdownResponse)
async def get_account_region_breakdown(
    account_id: int,
    scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AccountRegionBreakdownResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        account = (
            db.query(ProviderAccount)
            .filter(
                ProviderAccount.id == account_id,
                ProviderAccount.customer_id == customer_id,
                ProviderAccount.organization_id == organization_id,
            )
            .first()
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Account not found")

        alloc_query = db.query(CostAllocationSnapshot).filter(
            CostAllocationSnapshot.provider_account_id == account_id,
            CostAllocationSnapshot.customer_id == customer_id,
            CostAllocationSnapshot.organization_id == organization_id,
        )
        if scan_id:
            alloc_query = alloc_query.filter(CostAllocationSnapshot.scan_id == scan_id)
        else:
            latest_scan = (
                db.query(CostAllocationSnapshot.scan_id)
                .filter(
                    CostAllocationSnapshot.provider_account_id == account_id,
                    CostAllocationSnapshot.customer_id == customer_id,
                    CostAllocationSnapshot.organization_id == organization_id,
                )
                .order_by(CostAllocationSnapshot.captured_at.desc())
                .first()
            )
            if latest_scan:
                scan_id = latest_scan[0]
                alloc_query = alloc_query.filter(CostAllocationSnapshot.scan_id == scan_id)

        allocs = alloc_query.order_by(CostAllocationSnapshot.cost_usd.desc()).all()
        account_name = account.account_name
        provider = account.provider
    finally:
        db.close()

    regions = [
        AccountRegionBreakdownItem(
            region=a.region,
            cost_usd=round(float(a.cost_usd), 2),
            scan_id=a.scan_id,
            captured_at=a.captured_at.isoformat() if a.captured_at else "",
        )
        for a in allocs
    ]
    return AccountRegionBreakdownResponse(
        account_id=account_id,
        provider=provider,
        account_name=account_name,
        scan_id=scan_id,
        total_cost_usd=round(sum(r.cost_usd for r in regions), 2),
        regions=regions,
    )


@router.get("/imports/costs/summary", response_model=ImportedCostSummaryResponse)
async def get_imported_cost_summary(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ImportedCostSummaryResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    rows = _get_imported_cost_rows(db, organization_id, customer_id)
    if not rows:
        return ImportedCostSummaryResponse(
            organization_id=organization_id,
            customer_id=customer_id,
            has_data=False,
        )

    summary = _imported_cost_summary(rows)
    return ImportedCostSummaryResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        has_data=True,
        upload_id=summary["upload_id"],
        source_filename=summary["source_filename"],
        rows_imported=summary["rows_imported"],
        total_cost_usd=summary["total_cost_usd"],
        providers=summary["providers"],
        last_imported_at=summary["last_imported_at"],
    )


@router.get("/imports/costs/template.csv")
async def download_cost_import_template(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    _ = (current_user, membership)
    return Response(
        (
            "provider,cost_usd,service_name,account_identifier,account_name,account_type,parent_account_identifier,region,period_start,period_end,currency\n"
            "aws,123.45,EC2,acct-aws-1,AWS Prod,account,aws-org-root,eu-west-2,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
            "azure,67.89,Virtual Machines,sub-azure-1,Azure Prod,subscription,mg-finops,uk south,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
            "oci,10.00,Compute,comp-oci-1,OCI Prod,compartment,tenancy-main,uk-london-1,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
        ),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=optiora-cost-import-template.csv"},
    )


@router.post("/imports/costs/preview", response_model=ImportPreviewResponse)
async def preview_cost_csv_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ImportPreviewResponse:
    _require_management_role(membership, "CSV import preview")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    filename = str(file.filename or "").strip() or "cost-import.csv"
    raw = await file.read()
    try:
        headers, rows = load_normalized_csv_upload(filename, raw)
    except CsvImportError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail) from exc

    issues: List[ImportPreviewIssue] = []
    if "provider" not in headers:
        issues.append(ImportPreviewIssue(line_number=1, severity="error", message="Missing required column: provider"))
    if "cost_usd" not in headers:
        issues.append(ImportPreviewIssue(line_number=1, severity="error", message="Missing required column: cost_usd"))

    total_rows = 0
    accepted_rows = 0
    rejected_rows = 0
    total_cost_usd = 0.0
    providers: set[str] = set()
    rows_with_account = 0
    rows_with_region = 0
    rows_with_service = 0
    rows_with_tags = 0

    for line_number, normalized_row in enumerate(rows, start=2):
        total_rows += 1
        parsed_row = validate_cost_csv_row(
            normalized_row,
            line_number=line_number,
            supported_providers=SUPPORTED_COST_IMPORT_PROVIDERS,
            parse_required_float_value=_parse_required_float_value,
            parse_optional_datetime_value=_parse_optional_datetime_value,
            format_provider_error=lambda provider, _line_number, allowed: (
                f"Unsupported provider '{provider or 'empty'}'. Use one of: {', '.join(allowed)}."
            ),
            format_currency_error=lambda currency, _line_number: (
                f"Only USD is supported right now (got {currency})."
            ),
        )
        if parsed_row.errors:
            rejected_rows += 1
            for error in parsed_row.errors:
                issues.append(ImportPreviewIssue(line_number=line_number, severity="error", message=error))
            continue

        accepted_rows += 1
        total_cost_usd += float(parsed_row.cost_usd or 0.0)
        providers.add(parsed_row.provider)

        if normalized_row.get("account_identifier") or normalized_row.get("account_name"):
            rows_with_account += 1
        if normalized_row.get("region"):
            rows_with_region += 1
        if normalized_row.get("service_name") or normalized_row.get("service"):
            rows_with_service += 1
        if normalized_row.get("tags"):
            rows_with_tags += 1

    if total_rows > 0 and rows_with_tags == 0:
        issues.append(
            ImportPreviewIssue(
                line_number=1,
                severity="warning",
                message="No tags column data detected; business mapping coverage may be limited.",
            )
        )

    mapping_feedback = {
        "account_coverage_percent": round((rows_with_account / max(1, accepted_rows)) * 100, 1),
        "region_coverage_percent": round((rows_with_region / max(1, accepted_rows)) * 100, 1),
        "service_coverage_percent": round((rows_with_service / max(1, accepted_rows)) * 100, 1),
        "tags_coverage_percent": round((rows_with_tags / max(1, accepted_rows)) * 100, 1),
    }

    reconciliation_guidance: List[str] = []
    if accepted_rows == 0:
        reconciliation_guidance.append("No rows passed validation; fix CSV issues before import.")
    else:
        reconciliation_guidance.append(
            f"Validated {accepted_rows} row(s) with ${round(total_cost_usd, 2):,.2f} total spend ready for import."
        )
        if mapping_feedback["account_coverage_percent"] < 80:
            reconciliation_guidance.append("Add account_identifier/account_name to improve account rollup reconciliation.")
        if mapping_feedback["tags_coverage_percent"] < 50:
            reconciliation_guidance.append("Include tags JSON to improve team/environment/cost-center mapping quality.")
        if mapping_feedback["region_coverage_percent"] < 80:
            reconciliation_guidance.append("Populate region values to improve regional variance analysis.")

    return ImportPreviewResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        filename=filename,
        total_rows=total_rows,
        accepted_rows=accepted_rows,
        rejected_rows=rejected_rows,
        total_cost_usd=round(total_cost_usd, 2),
        detected_providers=sorted(providers),
        header_columns=headers,
        mapping_feedback=mapping_feedback,
        reconciliation_guidance=reconciliation_guidance,
        issues=issues[:200],
    )


@router.post("/imports/costs/csv", response_model=ImportedCostUploadResponse)
async def upload_cost_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ImportedCostUploadResponse:
    _require_management_role(membership, "CSV cost import")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    filename = str(file.filename or "").strip() or "cost-import.csv"
    raw = await file.read()
    try:
        headers, rows = load_normalized_csv_upload(filename, raw)
    except CsvImportError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.detail) from exc
    if "provider" not in headers or "cost_usd" not in headers:
        raise HTTPException(status_code=400, detail="CSV must include provider and cost_usd columns.")

    imported_at = _utcnow()
    upload_id = f"csv_{uuid4().hex}"
    rows_to_store: List[ImportedCostRecord] = []
    total_cost = 0.0
    providers: set[str] = set()
    validation_errors: List[str] = []

    for line_number, normalized_row in enumerate(rows, start=2):
        parsed_row = validate_cost_csv_row(
            normalized_row,
            line_number=line_number,
            supported_providers=SUPPORTED_COST_IMPORT_PROVIDERS,
            parse_required_float_value=_parse_required_float_value,
            parse_optional_datetime_value=_parse_optional_datetime_value,
            format_provider_error=lambda _provider, current_line_number, allowed: (
                f"Unsupported provider at CSV line {current_line_number}. Use one of: {', '.join(allowed)}."
            ),
            format_currency_error=lambda _currency, current_line_number: (
                f"Only USD CSV imports are supported right now. Invalid currency at line {current_line_number}."
            ),
        )
        if parsed_row.errors:
            validation_errors.extend(parsed_row.errors)
            continue

        total_cost += float(parsed_row.cost_usd or 0.0)
        providers.add(parsed_row.provider)
        rows_to_store.append(
            ImportedCostRecord(
                organization_id=organization_id,
                customer_id=customer_id,
                upload_id=upload_id,
                source_filename=filename,
                provider=parsed_row.provider,
                service_name=normalized_row.get("service_name") or normalized_row.get("service") or None,
                account_identifier=normalized_row.get("account_identifier") or None,
                account_name=normalized_row.get("account_name") or None,
                account_type=normalized_row.get("account_type") or None,
                parent_account_identifier=normalized_row.get("parent_account_identifier") or None,
                region=normalized_row.get("region") or None,
                period_start=parsed_row.period_start,
                period_end=parsed_row.period_end,
                cost_usd=float(parsed_row.cost_usd or 0.0),
                currency=parsed_row.currency,
                line_number=line_number,
                tags_json=normalized_row.get("tags") or None,
                created_at=imported_at,
            )
        )

    if validation_errors:
        detail = "CSV validation failed:\n- " + "\n- ".join(validation_errors[:10])
        if len(validation_errors) > 10:
            detail += f"\n- ...and {len(validation_errors) - 10} more issue(s)."
        raise HTTPException(status_code=400, detail=detail)

    if not rows_to_store:
        raise HTTPException(status_code=400, detail="CSV file does not contain any data rows.")

    db.query(ImportedCostRecord).filter(
        ImportedCostRecord.organization_id == organization_id,
        ImportedCostRecord.customer_id == customer_id,
    ).delete(synchronize_session=False)
    db.add_all(rows_to_store)
    db.add(
        AuditLog(
            organization_id=organization_id,
            actor_user_id=current_user.id,
            action="cost_import.csv_uploaded",
            entity_type="cost_import",
            entity_id=upload_id,
            metadata_json=json.dumps(
                {
                    "filename": filename,
                    "rows_imported": len(rows_to_store),
                    "providers": sorted(providers),
                }
            ),
            created_at=imported_at,
        )
    )
    db.commit()

    return ImportedCostUploadResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        upload_id=upload_id,
        filename=filename,
        rows_imported=len(rows_to_store),
        total_cost_usd=round(total_cost, 2),
        providers=sorted(providers),
        imported_at=imported_at,
    )


def _normalize_channels(channels: List[str]) -> List[str]:
    normalized: list[str] = []
    for value in channels:
        channel = str(value or "").strip().lower()
        if channel in SUPPORTED_NOTIFICATION_CHANNELS and channel not in normalized:
            normalized.append(channel)
    return normalized


def _severity_rank(value: str) -> int:
    normalized = str(value or "").strip().lower()
    rank = {
        "low": 1,
        "medium": 2,
        "warning": 2,
        "high": 3,
        "critical": 4,
    }
    return rank.get(normalized, 1)


def _effective_alert_ops_policy(
    db: Session,
    organization_id: int,
) -> Dict[str, Any]:
    row = (
        db.query(AlertOpsPolicy)
        .filter(AlertOpsPolicy.organization_id == organization_id)
        .first()
    )
    defaults: Dict[str, Any] = {
        "mute_window_enabled": False,
        "mute_start_hour_utc": 0,
        "mute_end_hour_utc": 0,
        "mute_weekends": False,
        "timezone": "UTC",
        "escalation_enabled": False,
        "escalation_after_minutes": 60,
        "escalation_channels": [],
        "escalation_severity": "critical",
        "ack_sla_minutes": 60,
        "dedupe_window_minutes": 30,
        "min_severity": "low",
        "daily_summary_enabled": True,
        "weekly_summary_enabled": True,
    }
    if row is None:
        return defaults

    try:
        channels = json.loads(row.escalation_channels_json or "[]")
    except json.JSONDecodeError:
        channels = []
    defaults.update(
        {
            "mute_window_enabled": bool(row.mute_window_enabled),
            "mute_start_hour_utc": int(row.mute_start_hour_utc or 0),
            "mute_end_hour_utc": int(row.mute_end_hour_utc or 0),
            "mute_weekends": bool(row.mute_weekends),
            "timezone": str(row.timezone or "UTC"),
            "escalation_enabled": bool(row.escalation_enabled),
            "escalation_after_minutes": max(5, int(row.escalation_after_minutes or 60)),
            "escalation_channels": _normalize_channels([str(item) for item in channels]),
            "escalation_severity": str(row.escalation_severity or "critical"),
            "ack_sla_minutes": max(5, int(row.ack_sla_minutes or 60)),
            "dedupe_window_minutes": max(0, int(row.dedupe_window_minutes or 0)),
            "min_severity": str(row.min_severity or "low"),
            "daily_summary_enabled": bool(row.daily_summary_enabled),
            "weekly_summary_enabled": bool(row.weekly_summary_enabled),
            "created_at": row.created_at.isoformat() if row.created_at else _utcnow().isoformat(),
            "updated_at": row.updated_at.isoformat() if row.updated_at else _utcnow().isoformat(),
        }
    )
    return defaults


def _is_muted_by_policy(policy: Dict[str, Any], now_utc: datetime) -> bool:
    if not bool(policy.get("mute_window_enabled")):
        return False
    timezone_name = str(policy.get("timezone") or "UTC").strip() or "UTC"
    try:
        tz = ZoneInfo(timezone_name)
    except Exception:
        tz = timezone.utc

    now_local = now_utc.replace(tzinfo=timezone.utc).astimezone(tz)
    if bool(policy.get("mute_weekends")) and now_local.weekday() >= 5:
        return True

    start_hour = int(policy.get("mute_start_hour_utc", 0) or 0)
    end_hour = int(policy.get("mute_end_hour_utc", 0) or 0)
    current_hour = int(now_local.hour)
    if start_hour == end_hour:
        return False
    if start_hour < end_hour:
        return start_hour <= current_hour < end_hour
    return current_hour >= start_hour or current_hour < end_hour


def _alert_passes_policy(
    db: Session,
    organization_id: int,
    customer_id: str,
    alert_type: str,
    severity: str,
    title: str,
    message: str,
    now_utc: datetime,
) -> tuple[bool, Optional[str], Dict[str, Any]]:
    policy = _effective_alert_ops_policy(db, organization_id)
    if _severity_rank(severity) < _severity_rank(str(policy.get("min_severity") or "low")):
        return False, "below_min_severity", policy
    if _is_muted_by_policy(policy, now_utc):
        return False, "muted_window", policy

    dedupe_window_minutes = int(policy.get("dedupe_window_minutes", 0) or 0)
    if dedupe_window_minutes > 0:
        cutoff = now_utc - timedelta(minutes=dedupe_window_minutes)
        duplicate = (
            db.query(AlertEvent.id)
            .filter(
                AlertEvent.organization_id == organization_id,
                AlertEvent.customer_id == customer_id,
                AlertEvent.alert_type == alert_type,
                AlertEvent.severity == severity,
                AlertEvent.title == title,
                AlertEvent.message == message,
                AlertEvent.created_at >= cutoff,
            )
            .first()
        )
        if duplicate is not None:
            return False, "dedupe_window", policy

    return True, None, policy


def _serialize_alert_ops_policy(
    policy: Dict[str, Any],
    organization_id: int,
) -> AlertOpsPolicyResponse:
    now_iso = _utcnow().isoformat()
    return AlertOpsPolicyResponse(
        organization_id=organization_id,
        mute_window_enabled=bool(policy.get("mute_window_enabled", False)),
        mute_start_hour_utc=int(policy.get("mute_start_hour_utc", 0) or 0),
        mute_end_hour_utc=int(policy.get("mute_end_hour_utc", 0) or 0),
        mute_weekends=bool(policy.get("mute_weekends", False)),
        timezone=str(policy.get("timezone") or "UTC"),
        escalation_enabled=bool(policy.get("escalation_enabled", False)),
        escalation_after_minutes=max(5, int(policy.get("escalation_after_minutes", 60) or 60)),
        escalation_channels=_normalize_channels(
            [str(item) for item in policy.get("escalation_channels", [])]
        ),
        escalation_severity=str(policy.get("escalation_severity") or "critical"),
        ack_sla_minutes=max(5, int(policy.get("ack_sla_minutes", 60) or 60)),
        dedupe_window_minutes=max(0, int(policy.get("dedupe_window_minutes", 0) or 0)),
        min_severity=str(policy.get("min_severity") or "low"),
        daily_summary_enabled=bool(policy.get("daily_summary_enabled", True)),
        weekly_summary_enabled=bool(policy.get("weekly_summary_enabled", True)),
        created_at=str(policy.get("created_at") or now_iso),
        updated_at=str(policy.get("updated_at") or now_iso),
    )


def _latest_delivery_timestamp_for_channel(
    db: Session,
    organization_id: int,
    customer_id: str,
    channel: str,
) -> Optional[str]:
    rows = (
        db.query(AlertEvent)
        .filter(
            AlertEvent.organization_id == organization_id,
            AlertEvent.customer_id == customer_id,
        )
        .order_by(AlertEvent.created_at.desc())
        .limit(100)
        .all()
    )
    for row in rows:
        delivered = _safe_json_load(row.delivered_channels_json, [])
        if channel in delivered:
            return row.created_at.isoformat() if row.created_at else None
    return None


def _alert_lifecycle_state_map(
    db: Session,
    organization_id: int,
    alert_ids: List[int],
) -> Dict[int, str]:
    if not alert_ids:
        return {}
    rows = (
        db.query(AuditLog)
        .filter(
            AuditLog.organization_id == organization_id,
            AuditLog.entity_type == "alert_event",
            AuditLog.entity_id.in_([str(alert_id) for alert_id in alert_ids]),
            AuditLog.action.in_(["alert.acknowledge", "alert.dismiss", "alert.reactivate"]),
        )
        .order_by(AuditLog.created_at.asc())
        .all()
    )

    state_map: Dict[int, str] = {}
    for row in rows:
        try:
            alert_id = int(str(row.entity_id or "0"))
        except ValueError:
            continue
        if row.action == "alert.dismiss":
            state_map[alert_id] = "dismissed"
        elif row.action == "alert.reactivate":
            state_map[alert_id] = "reactivated"
        elif row.action == "alert.acknowledge":
            state_map[alert_id] = "acknowledged"
    return state_map


def _channel_delivery_telemetry(
    db: Session,
    organization_id: int,
    alert_id: int,
) -> Dict[str, Dict[str, Optional[str]]]:
    """Retrieve per-channel delivery telemetry (last success/error) for an alert."""
    rows = (
        db.query(AuditLog)
        .filter(
            AuditLog.organization_id == organization_id,
            AuditLog.entity_type == "channel_delivery",
            AuditLog.entity_id == str(alert_id),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(100)
        .all()
    )
    
    channel_status: Dict[str, Dict[str, Optional[str]]] = {}
    for row in rows:
        try:
            metadata = json.loads(row.metadata_json)
            channel = metadata.get("channel")
            if channel:
                if channel not in channel_status:
                    channel_status[channel] = {
                        "last_success": None,
                        "last_error": None,
                        "status": "unknown",
                    }
                if row.action == "alert.channel_delivery_success" and channel_status[channel]["last_success"] is None:
                    channel_status[channel]["last_success"] = row.created_at.isoformat()
                    channel_status[channel]["status"] = "success"
                elif row.action == "alert.channel_delivery_error" and channel_status[channel]["last_error"] is None:
                    channel_status[channel]["last_error"] = row.created_at.isoformat()
                    channel_status[channel]["status"] = "error"
        except (json.JSONDecodeError, KeyError):
            pass
    
    return channel_status


def _channel_delivery_outcomes(
    db: Session,
    organization_id: int,
) -> Dict[str, Dict[str, Optional[str]]]:
    rows = (
        db.query(AuditLog)
        .filter(
            AuditLog.organization_id == organization_id,
            AuditLog.entity_type == "channel_delivery",
            AuditLog.action.in_(
                [
                    "alert.channel_delivery_success",
                    "alert.channel_delivery_error",
                ]
            ),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(500)
        .all()
    )

    outcomes: Dict[str, Dict[str, Optional[str]]] = {
        channel: {"last_success_at": None, "last_error_at": None}
        for channel in SUPPORTED_NOTIFICATION_CHANNELS
    }
    for row in rows:
        try:
            metadata = json.loads(row.metadata_json or "{}")
        except json.JSONDecodeError:
            continue
        channel = str(metadata.get("channel") or "").strip().lower()
        if channel not in outcomes:
            continue
        if row.action == "alert.channel_delivery_success" and outcomes[channel]["last_success_at"] is None:
            outcomes[channel]["last_success_at"] = row.created_at.isoformat() if row.created_at else None
        elif row.action == "alert.channel_delivery_error" and outcomes[channel]["last_error_at"] is None:
            outcomes[channel]["last_error_at"] = row.created_at.isoformat() if row.created_at else None

    return outcomes


@router.get("/alerts/routing-policies", response_model=List[AlertRoutingPolicyResponse])
async def list_alert_routing_policies(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[AlertRoutingPolicyResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(AlertRoutingPolicy)
            .filter(AlertRoutingPolicy.organization_id == organization_id)
            .order_by(AlertRoutingPolicy.severity.asc())
            .all()
        )
        return [
            AlertRoutingPolicyResponse(
                id=row.id,
                severity=row.severity,
                channels=_safe_json_load(row.channels_json, []),
                is_active=bool(row.is_active),
                created_at=row.created_at.isoformat() if row.created_at else "",
                updated_at=row.updated_at.isoformat() if row.updated_at else "",
            )
            for row in rows
        ]
    finally:
        db.close()


@router.post("/alerts/routing-policies", response_model=AlertRoutingPolicyResponse)
async def upsert_alert_routing_policy(
    payload: AlertRoutingPolicyRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertRoutingPolicyResponse:
    _require_management_role(membership, "alert routing policy updates")
    organization_id = _organization_id_for_membership(membership)
    channels = _normalize_channels(payload.channels)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.severity == payload.severity,
            )
            .first()
        )
        now = _utcnow()
        if row is None:
            row = AlertRoutingPolicy(
                organization_id=organization_id,
                severity=payload.severity,
                channels_json=json.dumps(channels),
                is_active=payload.is_active,
                created_at=now,
                updated_at=now,
            )
        else:
            row.channels_json = json.dumps(channels)
            row.is_active = payload.is_active
            row.updated_at = now

        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alerts.routing_policy.upsert",
                entity_type="alert_routing_policy",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "severity": payload.severity,
                        "channels": channels,
                        "is_active": payload.is_active,
                    }
                ),
            )
        )
        db.commit()
        db.refresh(row)
        return AlertRoutingPolicyResponse(
            id=row.id,
            severity=row.severity,
            channels=_safe_json_load(row.channels_json, []),
            is_active=bool(row.is_active),
            created_at=row.created_at.isoformat() if row.created_at else "",
            updated_at=row.updated_at.isoformat() if row.updated_at else "",
        )
    finally:
        db.close()


@router.post("/alerts/routing-policies/simulate", response_model=AlertRoutingPolicySimulationResponse)
async def simulate_alert_routing_policy(
    payload: AlertRoutingPolicySimulationRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertRoutingPolicySimulationResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.severity == payload.severity,
            )
            .first()
        )
    finally:
        db.close()

    channels = _safe_json_load(row.channels_json, []) if row else list(SUPPORTED_NOTIFICATION_CHANNELS)
    configured = [
        channel
        for channel in channels
        if destination_configured(Config(), channel)
    ]
    return AlertRoutingPolicySimulationResponse(
        severity=payload.severity,
        matched_policy_id=row.id if row else None,
        evaluated_channels=channels,
        expected_channels=configured,
        configured_channels=[
            channel
            for channel in SUPPORTED_NOTIFICATION_CHANNELS
            if destination_configured(Config(), channel)
        ],
        inactive_policy=(bool(row) and not bool(row.is_active)),
    )


@router.get("/alerts/ops-policy", response_model=AlertOpsPolicyResponse)
async def get_alert_ops_policy(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertOpsPolicyResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        policy = _effective_alert_ops_policy(db, organization_id)
        return _serialize_alert_ops_policy(policy, organization_id)
    finally:
        db.close()


@router.put("/alerts/ops-policy", response_model=AlertOpsPolicyResponse)
async def upsert_alert_ops_policy(
    payload: AlertOpsPolicyRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertOpsPolicyResponse:
    _require_management_role(membership, "alert operations policy updates")
    organization_id = _organization_id_for_membership(membership)
    try:
        ZoneInfo(payload.timezone)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid timezone value") from exc

    db = SessionLocal()
    try:
        row = (
            db.query(AlertOpsPolicy)
            .filter(AlertOpsPolicy.organization_id == organization_id)
            .first()
        )
        now = _utcnow()
        if row is None:
            row = AlertOpsPolicy(
                organization_id=organization_id,
                created_at=now,
                updated_at=now,
            )

        row.mute_window_enabled = payload.mute_window_enabled
        row.mute_start_hour_utc = int(payload.mute_start_hour_utc)
        row.mute_end_hour_utc = int(payload.mute_end_hour_utc)
        row.mute_weekends = payload.mute_weekends
        row.timezone = payload.timezone
        row.escalation_enabled = payload.escalation_enabled
        row.escalation_after_minutes = int(payload.escalation_after_minutes)
        row.escalation_channels_json = json.dumps(_normalize_channels(payload.escalation_channels))
        row.escalation_severity = payload.escalation_severity
        row.ack_sla_minutes = int(payload.ack_sla_minutes)
        row.dedupe_window_minutes = int(payload.dedupe_window_minutes)
        row.min_severity = payload.min_severity
        row.daily_summary_enabled = payload.daily_summary_enabled
        row.weekly_summary_enabled = payload.weekly_summary_enabled
        row.updated_at = now
        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alerts.ops_policy.upsert",
                entity_type="alert_ops_policy",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "mute_window_enabled": bool(payload.mute_window_enabled),
                        "escalation_enabled": bool(payload.escalation_enabled),
                        "ack_sla_minutes": int(payload.ack_sla_minutes),
                        "dedupe_window_minutes": int(payload.dedupe_window_minutes),
                        "min_severity": payload.min_severity,
                    }
                ),
            )
        )
        db.commit()
        policy = _effective_alert_ops_policy(db, organization_id)
        return _serialize_alert_ops_policy(policy, organization_id)
    finally:
        db.close()


@router.get("/alerts/executive-summary", response_model=AlertExecutiveSummaryResponse)
async def get_alert_executive_summary(
    period: Literal["daily", "weekly"] = "daily",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertExecutiveSummaryResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()
    days = 1 if period == "daily" else 7
    window_start = now - timedelta(days=days)

    db = SessionLocal()
    try:
        policy = _effective_alert_ops_policy(db, organization_id)
        if period == "daily" and not bool(policy.get("daily_summary_enabled", True)):
            raise HTTPException(status_code=403, detail="Daily summary is disabled by policy")
        if period == "weekly" and not bool(policy.get("weekly_summary_enabled", True)):
            raise HTTPException(status_code=403, detail="Weekly summary is disabled by policy")

        rows = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.organization_id == organization_id,
                AlertEvent.customer_id == customer_id,
                AlertEvent.created_at >= window_start,
            )
            .order_by(AlertEvent.created_at.desc())
            .all()
        )
        lifecycle_map = _alert_lifecycle_state_map(
            db=db,
            organization_id=organization_id,
            alert_ids=[int(row.id) for row in rows],
        )

        by_severity: Dict[str, int] = {}
        acknowledged = 0
        dismissed = 0
        for row in rows:
            severity = str(row.severity or "unknown").lower()
            by_severity[severity] = by_severity.get(severity, 0) + 1
            lifecycle = lifecycle_map.get(
                int(row.id),
                "acknowledged" if row.acknowledged_at else "active",
            )
            if lifecycle == "dismissed":
                dismissed += 1
            elif lifecycle == "acknowledged":
                acknowledged += 1

        total_alerts = len(rows)
        return AlertExecutiveSummaryResponse(
            organization_id=organization_id,
            period=period,
            generated_at=now.isoformat(),
            window_start=window_start.isoformat(),
            total_alerts=total_alerts,
            acknowledged=acknowledged,
            unacknowledged=max(0, total_alerts - acknowledged - dismissed),
            dismissed=dismissed,
            by_severity=by_severity,
        )
    finally:
        db.close()


@router.get("/notifications/destinations", response_model=NotificationDestinationsResponse)
async def list_notification_destinations(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationsResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    config = Config()
    db = SessionLocal()
    try:
        policies = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.is_active == True,  # noqa: E712
            )
            .all()
        )
        enabled_channels: set[str] = set()
        for row in policies:
            enabled_channels.update(_safe_json_load(row.channels_json, []))

        delivery_outcomes = _channel_delivery_outcomes(db, organization_id)
        destinations = []
        for channel in SUPPORTED_NOTIFICATION_CHANNELS:
            channel_outcomes = delivery_outcomes.get(channel, {})
            destinations.append(
                NotificationDestinationStatus(
                    channel=channel,
                    configured=destination_configured(config, channel),
                    enabled=(channel in enabled_channels) if policies else True,
                    last_delivery_at=_latest_delivery_timestamp_for_channel(
                        db=db,
                        organization_id=organization_id,
                        customer_id=customer_id,
                        channel=channel,
                    ),
                    last_success_at=channel_outcomes.get("last_success_at"),
                    last_error_at=channel_outcomes.get("last_error_at"),
                )
            )
        return NotificationDestinationsResponse(
            organization_id=organization_id,
            destinations=destinations,
        )
    finally:
        db.close()


@router.post("/notifications/destinations/{channel}/toggle", response_model=NotificationDestinationsResponse)
async def toggle_notification_destination(
    channel: str,
    payload: NotificationDestinationToggleRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationsResponse:
    _require_management_role(membership, "notification destination toggles")
    channel_key = str(channel or "").strip().lower()
    if channel_key not in SUPPORTED_NOTIFICATION_CHANNELS:
        raise HTTPException(status_code=400, detail=f"Unsupported channel: {channel_key}")

    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        for severity in ("warning", "critical"):
            row = (
                db.query(AlertRoutingPolicy)
                .filter(
                    AlertRoutingPolicy.organization_id == organization_id,
                    AlertRoutingPolicy.severity == severity,
                )
                .first()
            )
            if row is None:
                channels = list(SUPPORTED_NOTIFICATION_CHANNELS)
                if not payload.enabled:
                    channels = [c for c in channels if c != channel_key]
                row = AlertRoutingPolicy(
                    organization_id=organization_id,
                    severity=severity,
                    channels_json=json.dumps(channels),
                    is_active=True,
                    created_at=_utcnow(),
                    updated_at=_utcnow(),
                )
                db.add(row)
                continue

            channels = set(_safe_json_load(row.channels_json, []))
            if payload.enabled:
                channels.add(channel_key)
            else:
                channels.discard(channel_key)
            row.channels_json = json.dumps(sorted(channels))
            row.updated_at = _utcnow()

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="notifications.destination.toggle",
                entity_type="notification_destination",
                entity_id=channel_key,
                metadata_json=json.dumps({"enabled": payload.enabled}),
            )
        )
        db.commit()
    finally:
        db.close()

    return await list_notification_destinations(current_user=current_user, membership=membership)


@router.post("/notifications/test-destination", response_model=NotificationDestinationTestResponse)
async def test_notification_destination(
    payload: NotificationDestinationTestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationTestResponse:
    _require_management_role(membership, "notification destination tests")
    organization_id = _organization_id_for_membership(membership)
    config = Config()
    success, detail = send_test_notification(
        config=config,
        channel=payload.channel,
        target=payload.target,
        message=payload.message,
    )

    db = SessionLocal()
    try:
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="notifications.destination.test",
                entity_type="notification_destination",
                entity_id=payload.channel,
                metadata_json=json.dumps(
                    {
                        "success": success,
                        "target_supplied": bool(payload.target),
                        "detail": detail,
                    }
                ),
            )
        )
        db.commit()
    finally:
        db.close()

    return NotificationDestinationTestResponse(
        channel=payload.channel,
        success=success,
        detail=detail,
    )


@router.get("/alerts")
async def list_alerts(
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)
    now = _utcnow()
    db = SessionLocal()
    try:
        policy = _effective_alert_ops_policy(db, organization_id)
        ack_sla_minutes = max(5, int(policy.get("ack_sla_minutes", 60) or 60))
        escalation_enabled = bool(policy.get("escalation_enabled", False))
        escalation_after_minutes = max(5, int(policy.get("escalation_after_minutes", 60) or 60))
        escalation_channels = _normalize_channels(
            [str(item) for item in policy.get("escalation_channels", [])]
        )
        escalation_severity = str(policy.get("escalation_severity") or "critical")
        rows = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.customer_id == customer_id,
                AlertEvent.organization_id == organization_id,
            )
            .order_by(AlertEvent.created_at.desc())
            .offset(max(0, min(offset, 50000)))
            .limit(max(1, min(limit, 200)))
            .all()
        )
        lifecycle_map = _alert_lifecycle_state_map(
            db=db,
            organization_id=organization_id,
            alert_ids=[int(row.id) for row in rows],
        )
        payload: List[Dict[str, Any]] = []
        for row in rows:
            lifecycle_state = lifecycle_map.get(
                int(row.id),
                "acknowledged" if row.acknowledged_at else "active",
            )
            payload.append(
                {
                    "id": row.id,
                    "alert_type": row.alert_type,
                    "severity": row.severity,
                    "title": row.title,
                    "message": row.message,
                    "delivered_channels": json.loads(row.delivered_channels_json or "[]"),
                    "channel_telemetry": _channel_delivery_telemetry(
                        db=db,
                        organization_id=organization_id,
                        alert_id=int(row.id),
                    ),
                    "acknowledged_at": row.acknowledged_at.isoformat() if row.acknowledged_at else None,
                    "lifecycle_state": lifecycle_state,
                    "ack_sla_minutes": ack_sla_minutes,
                    "ack_sla_breached": (
                        lifecycle_state not in {"acknowledged", "dismissed"}
                        and row.created_at is not None
                        and int((now - row.created_at).total_seconds()) > (ack_sla_minutes * 60)
                    ),
                    "escalation_due": (
                        escalation_enabled
                        and lifecycle_state not in {"acknowledged", "dismissed"}
                        and row.created_at is not None
                        and int((now - row.created_at).total_seconds()) > (escalation_after_minutes * 60)
                        and _severity_rank(str(row.severity or "low"))
                        >= _severity_rank(escalation_severity)
                    ),
                    "escalation_channels": escalation_channels,
                    "created_at": row.created_at.isoformat(),
                }
            )
        return payload
    finally:
        db.close()


@router.post("/alerts/{alert_id}/acknowledge", response_model=AlertLifecycleActionResponse)
async def acknowledge_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert acknowledgement")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        row.acknowledged_at = _utcnow()
        row.acknowledged_by_user_id = current_user.id
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.acknowledge",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="acknowledged",
    )


@router.post("/alerts/{alert_id}/dismiss", response_model=AlertLifecycleActionResponse)
async def dismiss_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert dismissal")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.dismiss",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="dismissed",
    )


@router.post("/alerts/{alert_id}/reactivate", response_model=AlertLifecycleActionResponse)
async def reactivate_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert reactivation")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.reactivate",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="reactivated",
    )


class ChannelDeliveryEvent(BaseModel):
    alert_id: int
    channel: str
    status: Literal["success", "error"]
    error_message: Optional[str] = None


@router.post("/alerts/{alert_id}/channel-delivery")
async def record_channel_delivery(
    alert_id: int,
    event: ChannelDeliveryEvent,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Record per-channel delivery telemetry for an alert."""
    _require_management_role(membership, "channel delivery recording")
    organization_id = _organization_id_for_membership(membership)
    
    db = SessionLocal()
    try:
        # Verify alert exists
        alert = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if alert is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        
        # Record delivery event
        action = (
            "alert.channel_delivery_success" if event.status == "success" 
            else "alert.channel_delivery_error"
        )
        metadata = {
            "channel": event.channel,
            "status": event.status,
        }
        if event.error_message:
            metadata["error_message"] = event.error_message
        
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action=action,
                entity_type="channel_delivery",
                entity_id=str(alert_id),
                metadata_json=json.dumps(metadata),
                created_at=_utcnow(),
            )
        )
        db.commit()
    finally:
        db.close()
    
    return {
        "status": "ok",
        "alert_id": alert_id,
        "channel": event.channel,
        "delivery_status": event.status,
    }


@router.get("/alerts.csv")
async def download_alerts_csv(
    limit: int = 200,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await list_alerts(
        limit=limit,
        offset=offset,
        current_user=current_user,
        membership=membership,
    )
    lines = ["id,alert_type,severity,title,message,lifecycle_state,acknowledged_at,created_at"]
    for row in rows:
        lines.append(
            f"{row['id']},{row['alert_type']},{row['severity']},"
            f"{_csv_escape(row['title'])},"
            f"{_csv_escape(row['message'])},"
            f"{row.get('lifecycle_state', 'active')},"
            f"{row['acknowledged_at'] or ''},{row['created_at']}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/audit-logs")
async def list_audit_logs(
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(AuditLog)
            .filter(AuditLog.organization_id == organization_id)
            .order_by(AuditLog.created_at.desc())
            .offset(max(0, min(offset, 50000)))
            .limit(max(1, min(limit, 200)))
            .all()
        )
    finally:
        db.close()
    return [
        {
            "id": row.id,
            "action": row.action,
            "entity_type": row.entity_type,
            "entity_id": row.entity_id,
            "actor_user_id": row.actor_user_id,
            "metadata": _safe_json_load(row.metadata_json or "{}", {}),
            "created_at": row.created_at.isoformat(),
        }
        for row in rows
    ]


@router.get("/audit-logs.csv")
async def download_audit_logs_csv(
    limit: int = 200,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await list_audit_logs(
        limit=limit,
        offset=offset,
        current_user=current_user,
        membership=membership,
    )
    lines = ["id,action,entity_type,entity_id,actor_user_id,created_at"]
    for row in rows:
        lines.append(
            f"{row['id']},{row['action']},{row['entity_type']},{row['entity_id'] or ''},"
            f"{row['actor_user_id'] or ''},{row['created_at']}"
        )
    return Response("\n".join(lines), media_type="text/csv")


async def _executive_summary_rows(
    current_user: User,
    membership: UserOrganization,
    db: Session,
) -> List[List[Any]]:
    costs = await dashboard_costs(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    analytics = await dashboard_analytics(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    forecast = await dashboard_forecast(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    rollups = await get_provider_account_rollups(
        current_user=current_user,
        membership=membership,
    )
    alerts = await list_alerts(
        limit=10,
        current_user=current_user,
        membership=membership,
    )

    rows: List[List[Any]] = [["Section", "Field", "Value"]]
    rows.extend(
        [
            ["Summary", "Generated At", _utcnow().isoformat()],
            ["Summary", "Organization ID", _organization_id_for_membership(membership)],
            ["Summary", "Cost Source", costs.get("cost_context", {}).get("source", "live")],
            ["Summary", "Total Monthly Cost USD", round(float(costs.get("totalCost", 0.0) or 0.0), 2)],
            ["Summary", "Potential Monthly Savings USD", round(float(costs.get("potentialSavings", 0.0) or 0.0), 2)],
            ["Summary", "Risk Score", analytics.get("risk_score", 0)],
            ["Summary", "Maturity Score", analytics.get("maturity_score", 0)],
            ["Summary", "Commitment Coverage Percent", analytics.get("commitment_coverage_percent", 0)],
            ["Summary", "Spend At Risk USD", round(float(analytics.get("spend_at_risk_usd", 0.0) or 0.0), 2)],
            ["Summary", "Optimization Capacity USD", round(float(analytics.get("optimization_capacity_usd", 0.0) or 0.0), 2)],
            ["Summary", "Budget Utilization Percent", analytics.get("unit_metrics", {}).get("budget_utilization_percent", 0)],
            ["Summary", "Forecast History Source", forecast.get("history_source", "no_history")],
            ["Summary", "Forecast Backtest MAPE %", (forecast.get("backtesting") or {}).get("mape_percent")],
            ["Summary", "Forecast Backtest wMAPE %", (forecast.get("backtesting") or {}).get("wmape_percent")],
            ["Summary", "Average Breach Probability", (forecast.get("budget_guardrails") or {}).get("average_breach_probability")],
            ["Summary", "Open Alerts", len([row for row in alerts if not row.get("acknowledged_at")])],
            ["Summary", "Rollup Nodes", len(rollups.items)],
        ]
    )

    for provider_name, provider_data in costs.get("breakdown", {}).items():
        rows.append([
            "Provider Breakdown",
            provider_name,
            round(float(provider_data.get("cost", 0.0) or 0.0), 2),
        ])

    for item in rollups.items[:20]:
        rows.append([
            "Account Rollup",
            f"{'  ' * item.depth}{item.account_name} ({item.account_type})",
            item.rolled_up_cost_usd,
        ])

    for alert in alerts[:10]:
        rows.append([
            "Alerts",
            f"{alert['severity']} {alert['title']}",
            alert["created_at"],
        ])

    # ── Business mapping / chargeback section ─────────────────────────────
    try:
        with SessionLocal() as _db:
            org_id = _organization_id_for_membership(membership)
            customer_id = _customer_id_for_org(membership)
            dim_rows = (
                _db.query(NormalizedCostDimension)
                .filter(NormalizedCostDimension.organization_id == org_id)
                .all()
            )
            if dim_rows:
                total_dim = sum(float(r.cost_usd or 0) for r in dim_rows)
                mapped_dim = sum(float(r.cost_usd or 0) for r in dim_rows if r.is_mapped)
                coverage_pct = round((mapped_dim / total_dim * 100) if total_dim > 0 else 0.0, 1)
                rows.extend([
                    ["Allocation Coverage", "Total Cost USD", round(total_dim, 2)],
                    ["Allocation Coverage", "Mapped Cost USD", round(mapped_dim, 2)],
                    ["Allocation Coverage", "Coverage Percent", coverage_pct],
                ])
                # Per-team chargeback
                team_totals: Dict[str, float] = {}
                for r in dim_rows:
                    t = r.team or "(unmapped)"
                    team_totals[t] = team_totals.get(t, 0.0) + float(r.cost_usd or 0)
                for team, cost in sorted(team_totals.items(), key=lambda x: -x[1])[:15]:
                    rows.append(["Chargeback by Team", team, round(cost, 2)])
    except Exception:
        pass

    return rows


def _serialize_export_job(row: ExportJob) -> ExportJobResponse:
    return ExportJobResponse(
        id=row.id,
        organization_id=row.organization_id,
        customer_id=row.customer_id,
        name=row.name,
        report_type=row.report_type,
        export_format=row.export_format,
        schedule_frequency=row.schedule_frequency,
        is_active=bool(row.is_active),
        last_run_at=row.last_run_at.isoformat() if row.last_run_at else None,
        created_at=row.created_at.isoformat() if row.created_at else "",
        updated_at=row.updated_at.isoformat() if row.updated_at else "",
    )


def _serialize_export_job_run(row: ExportJobRun) -> ExportJobRunResponse:
    return ExportJobRunResponse(
        id=row.id,
        export_job_id=row.export_job_id,
        status=row.status,
        output_filename=row.output_filename,
        row_count=int(row.row_count or 0),
        error_message=row.error_message,
        created_at=row.created_at.isoformat() if row.created_at else "",
        completed_at=row.completed_at.isoformat() if row.completed_at else None,
    )


async def _execute_export_job(
    *,
    db: Session,
    job: ExportJob,
    current_user: User,
    membership: UserOrganization,
    actor_user_id: Optional[int],
) -> ExportJobRun:
    now = _utcnow()
    run = ExportJobRun(
        export_job_id=job.id,
        organization_id=job.organization_id,
        customer_id=job.customer_id,
        status="running",
        created_at=now,
    )
    db.add(run)
    db.flush()

    try:
        rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
        row_count = max(0, len(rows) - 1)

        if job.report_type == "finance_workbook" or job.export_format == "xlsx":
            extension = "xlsx"
        elif job.export_format == "pdf":
            extension = "pdf"
        elif job.export_format == "csv":
            extension = "csv"
        else:
            extension = "xls"

        if job.report_type == "executive_digest" and extension != "pdf":
            extension = "pdf"

        # Persist only metadata now; file materialization can be plugged into object storage later.
        run.output_filename = f"{job.report_type}-{job.id}-{int(now.timestamp())}.{extension}"
        run.row_count = row_count
        run.status = "completed"
        run.completed_at = _utcnow()
        job.last_run_at = run.completed_at
        job.updated_at = run.completed_at
        db.add(job)
        db.add(
            AuditLog(
                organization_id=job.organization_id,
                actor_user_id=actor_user_id,
                action="exports.job.run",
                entity_type="export_job",
                entity_id=str(job.id),
                metadata_json=json.dumps(
                    {
                        "report_type": job.report_type,
                        "export_format": job.export_format,
                        "schedule_frequency": job.schedule_frequency,
                        "row_count": row_count,
                        "output_filename": run.output_filename,
                    }
                ),
            )
        )
    except Exception as exc:
        run.status = "failed"
        run.error_message = str(exc)
        run.completed_at = _utcnow()

    db.add(run)
    db.commit()
    db.refresh(run)
    return run


@router.get("/exports/jobs", response_model=List[ExportJobResponse])
async def list_export_jobs(
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ExportJobResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(ExportJob)
            .filter(
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .order_by(ExportJob.created_at.desc())
            .offset(max(0, min(offset, 50000)))
            .limit(max(1, min(limit, 200)))
            .all()
        )
        return [_serialize_export_job(row) for row in rows]
    finally:
        db.close()


@router.post("/exports/jobs", response_model=ExportJobResponse)
async def create_export_job(
    payload: ExportJobRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ExportJobResponse:
    _require_management_role(membership, "export job management")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        now = _utcnow()
        row = ExportJob(
            organization_id=organization_id,
            customer_id=customer_id,
            name=payload.name.strip() or "Scheduled Export",
            report_type=payload.report_type,
            export_format=payload.export_format,
            schedule_frequency=payload.schedule_frequency,
            is_active=payload.is_active,
            created_at=now,
            updated_at=now,
        )
        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="exports.job.create",
                entity_type="export_job",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "name": row.name,
                        "report_type": row.report_type,
                        "export_format": row.export_format,
                        "schedule_frequency": row.schedule_frequency,
                        "is_active": bool(row.is_active),
                    }
                ),
            )
        )
        db.commit()
        db.refresh(row)
        return _serialize_export_job(row)
    finally:
        db.close()


@router.post("/exports/jobs/{job_id}/run", response_model=ExportJobRunResponse)
async def run_export_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ExportJobRunResponse:
    _require_management_role(membership, "export job runs")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        job = (
            db.query(ExportJob)
            .filter(
                ExportJob.id == job_id,
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .first()
        )
        if job is None:
            raise HTTPException(status_code=404, detail="Export job not found")

        run = await _execute_export_job(
            db=db,
            job=job,
            current_user=current_user,
            membership=membership,
            actor_user_id=current_user.id,
        )
        return _serialize_export_job_run(run)
    finally:
        db.close()


@router.get("/exports/jobs/{job_id}/runs", response_model=List[ExportJobRunResponse])
async def list_export_job_runs(
    job_id: int,
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ExportJobRunResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        job_exists = (
            db.query(ExportJob.id)
            .filter(
                ExportJob.id == job_id,
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .first()
        )
        if job_exists is None:
            raise HTTPException(status_code=404, detail="Export job not found")

        rows = (
            db.query(ExportJobRun)
            .filter(
                ExportJobRun.export_job_id == job_id,
                ExportJobRun.organization_id == organization_id,
                ExportJobRun.customer_id == customer_id,
            )
            .order_by(ExportJobRun.created_at.desc())
            .limit(max(1, min(limit, 100)))
            .all()
        )
        return [_serialize_export_job_run(row) for row in rows]
    finally:
        db.close()


@router.get("/reports/executive-summary.csv")
async def download_executive_summary_csv(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    return _csv_response(
        "optiora-executive-summary.csv",
        rows[0],
        rows[1:],
    )


@router.get("/reports/executive-summary.xls")
async def download_executive_summary_excel(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    return _spreadsheet_xml_response(
        "optiora-executive-summary.xls",
        "Executive Summary",
        rows,
    )


@router.get("/dashboard/costs")
@router.get("/costs")
async def dashboard_costs(
    period: str = "month",
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    context = await _cost_context(membership, db, period, cloud_provider)

    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    recommendations_result = _safe_json_load(
        await recommendations.get_recommendations({"cloud_provider": cloud_provider}),
        {},
    )

    potential_savings = float(
        recommendations_result.get("total_potential_savings_annual_usd", 0) or 0
    ) / 12

    return {
        "totalCost": context["total_cost"],
        "trend": 0.0,
        "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
        "potentialSavings": round(potential_savings, 2),
        "breakdown": context["breakdown"],
        "regionBreakdown": context["region_breakdown"],
        "cost_context": {
            "source": context.get("source", "unknown"),
            "provider_errors": context.get("provider_errors", {}),
            "rows_imported": int(context.get("rows_imported", 0) or 0),
            "last_imported_at": context.get("last_imported_at"),
        },
    }


@router.get("/dashboard/anomalies")
@router.get("/anomalies")
async def dashboard_anomalies(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = (current_user, membership)
    result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    rows = result.get("anomalies", [])
    mapped = []
    for idx, row in enumerate(rows):
        mapped.append(
            {
                "id": f"anomaly-{idx + 1}",
                "service": row.get("service", "unknown"),
                "cloud": cloud_provider if cloud_provider != "all" else "multi-cloud",
                "message": row.get("probable_cause", "Anomaly detected"),
                "severity": "high" if (row.get("increase_percent", 0) or 0) > 150 else "medium",
                "timestamp": row.get("date", _utcnow().isoformat()),
                "change": float(row.get("increase_percent", 0) or 0),
            }
        )
    return mapped


@router.get("/dashboard/recommendations")
@router.get("/recommendations")
async def dashboard_recommendations(
    cloud_provider: str = "all",
    limit: int = Query(20, ge=1, le=1000),
    include_provider_recommendations: bool = Query(False),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> List[Dict[str, Any]]:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    raw_rows = result.get("recommendations", [])
    rows = raw_rows if isinstance(raw_rows, list) else []
    if include_provider_recommendations:
        live_rows = _collect_provider_recommendation_rows(
            db=db,
            customer_id=customer_id,
            provider=cloud_provider,
            min_monthly_savings=0.0,
            limit=limit,
            include_existing_rightsizing_sources=True,
        )
        if live_rows:
            rows = [row for row in rows if isinstance(row, dict)] + live_rows

    mapped = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        row_provider = _provider_from_recommendation_row(row, cloud_provider)
        savings_monthly = _recommendation_monthly_savings(row)
        payback_months = float(row.get("payback_months", 3) or 3)
        mapped.append(
            {
                "id": row.get("id"),
                "service": row.get("service", "unknown"),
                "cloud": row_provider if row_provider != "all" else "multi-cloud",
                "title": row.get("recommendation_type") or row.get("description", "Optimization recommendation"),
                "description": row.get("description", ""),
                "savings": savings_monthly,
                "roi": float(row.get("roi_percent", 0) or 0),
                "difficulty": "easy"
                if payback_months <= 1
                else "medium"
                if payback_months <= 3
                else "hard",
                "source": row.get("source") or row.get("evidence_source") or "cost_context",
                "resource_id": row.get("resource_id"),
                "resource_type": row.get("resource_type"),
                "resource_name": row.get("resource_name"),
                "region": row.get("region"),
                "recommendation_type": row.get("recommendation_type"),
                "recommendation_name": row.get("recommendation_name"),
                "resource_count": row.get("resource_count"),
                "category": row.get("category"),
                "importance": row.get("importance"),
                "status": row.get("status"),
                "recommendation_status": row.get("recommendation_status"),
                "resource_console_url": row.get("resource_console_url"),
            }
        )
    return _dedupe_dashboard_recommendations(mapped)[:limit]


@router.get("/forecast")
async def dashboard_forecast(
    months: int = 12,
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    result["cost_context"] = context
    return result


class ForecastWhatIfAction(BaseModel):
    name: str
    start_month: int = Field(default=1, ge=1, le=24)
    savings_percent: float = Field(default=0.0, ge=0.0, le=80.0)
    growth_delta_percent: float = Field(default=0.0, ge=-30.0, le=30.0)
    one_time_cost_usd: float = Field(default=0.0, ge=0.0)


class ForecastWhatIfRequest(BaseModel):
    months: int = Field(default=12, ge=1, le=24)
    cloud_provider: str = "all"
    actions: List[ForecastWhatIfAction] = Field(default_factory=list)
    discount_rate_monthly: float = Field(default=0.01, ge=0.0, le=0.2)


class ForecastStressRequest(BaseModel):
    months: int = Field(default=12, ge=1, le=24)
    cloud_provider: str = "all"
    severity: Literal["low", "medium", "high"] = "medium"


@router.post("/forecast/what-if")
async def dashboard_forecast_what_if(
    request: ForecastWhatIfRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deterministic what-if simulation for planned optimization initiatives."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", request.cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=request.cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast_what_if(
            {
                "months": request.months,
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "discount_rate_monthly": request.discount_rate_monthly,
                "actions": [
                    action.model_dump() if hasattr(action, "model_dump") else action.dict()
                    for action in request.actions
                ],
            }
        ),
        {},
    )
    result["cost_context"] = context
    return result


@router.post("/forecast/stress-test")
async def dashboard_forecast_stress_test(
    request: ForecastStressRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deterministic stress-test envelope around forecast baseline."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", request.cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=request.cloud_provider,
        months=18,
    )

    result = _safe_json_load(
        await finops_analytics.get_forecast_stress_test(
            {
                "months": request.months,
                "cloud_provider": request.cloud_provider,
                "severity": request.severity,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    narrative, prompt = genai_advisor.generate_budget_risk_alert(
        result.get("worst_case", {}),
        {
            "current_monthly_spend_usd": context["total_cost"],
            "worst_case": result.get("worst_case", {}),
            "severity": request.severity,
            "forecast_months": request.months,
        },
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/forecast/model-diagnostics")
async def dashboard_forecast_model_diagnostics(
    months: int = Query(default=12, ge=1, le=24),
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Champion/challenger forecast model diagnostics and model-risk advisory context."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast_model_diagnostics(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    rag_context = dict(result)
    rag = _rag_context_for_analysis(
        analysis_type="forecast_model_diagnostics",
        cloud_provider=cloud_provider,
        context=rag_context,
    )
    narrative, prompt = genai_advisor.generate_forecast_model_diagnostics(rag_context)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["rag"] = rag
    result["cost_context"] = context
    return result


@router.get("/analytics/forecast-diagnostics")
async def analytics_forecast_diagnostics(
    months: int = Query(default=12, ge=1, le=24),
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Forecast quality diagnostics including sensitivity and budget pressure."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast_diagnostics(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    rag_context = {
        "budget_guardrails": result.get("budget_guardrails") or {},
        "cost_velocity_pct_mom": result.get("cost_velocity_pct_mom"),
        "forecast_quality": result.get("forecast_quality") or {},
    }
    rag = _rag_context_for_analysis(
        analysis_type="budget_risk",
        cloud_provider=cloud_provider,
        context=rag_context,
    )
    narrative, prompt = genai_advisor.generate_budget_risk_alert(
        result.get("budget_guardrails") or {},
        {
            "current_monthly_spend_usd": context["total_cost"],
            "cost_velocity_pct_mom": result.get("cost_velocity_pct_mom"),
            "rag_brief": rag.get("rag_brief", ""),
        },
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["rag"] = rag
    result["cost_context"] = context
    return result


@router.get("/analytics")
async def dashboard_analytics(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(_customer_id_for_org(membership))
    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    recommendations_result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    monthly_savings = (
        float(recommendations_result.get("total_potential_savings_annual_usd", 0) or 0) / 12
    )
    result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
                "monthly_savings": monthly_savings,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )
    result["cost_context"] = context
    # Attach backend GenAI narrative when OCI GenAI is configured
    narrative, prompt = genai_advisor.generate_spend_narrative(result)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    return result


@router.get("/analytics/attribution")
async def analytics_attribution(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Pareto cost driver attribution — which providers/services account for 80% of spend."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_cost_attribution({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/commitment-optimization")
async def analytics_commitment_optimization(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Model the ROI of increasing RI/Savings Plan coverage across coverage tiers."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_commitment_optimization({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/maturity")
async def analytics_maturity(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """FinOps maturity assessment mapped to CRAWL/WALK/RUN/OPTIMIZE levels."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    customer_id = _customer_id_for_org(membership)
    permission = ScanningManager(db).get_permission_status(customer_id)
    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}), {}
    )
    historical = _historical_monthly_spend_from_snapshots(
        db=db, customer_id=customer_id, cloud_provider=cloud_provider, months=18
    )
    result = _safe_json_load(
        await finops_analytics.get_maturity_assessment({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
            "history_coverage_months": len(historical),
            "scheduler_enabled": bool(permission.get("scan_frequency")),
            "auto_remediate": bool(permission.get("auto_remediate", False)),
        }),
        {},
    )
    # Attach GenAI narrative if OCI GenAI is configured
    narrative, prompt = genai_advisor.generate_maturity_narrative(result)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/unit-economics")
async def analytics_unit_economics(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Unit cost trends, waste-to-spend ratio, and dollar efficiency score."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    result = _safe_json_load(
        await finops_analytics.get_unit_economics({
            "current_monthly_spend": context["total_cost"],
            "estimated_waste_usd": analytics_result.get("estimated_monthly_waste_usd", 0),
            "identified_savings_usd": analytics_result.get("identified_monthly_savings_usd", 0),
            "anomalies": 0,
        }),
        {},
    )
    result["cost_context"] = context
    return result


class GenAIAnalyzeRequest(BaseModel):
    analysis_type: Literal[
        "spend", "anomaly", "optimization", "maturity", "budget_risk",
        "waste_insights", "optimization_roadmap", "executive_narrative", "commitment_strategy",
        "tagging_strategy", "sustainability_narrative", "chargeback_narrative",
        "cross_provider_comparison_brief", "alert_triage", "rightsizing_brief",
        "vendor_negotiation_brief", "forecast_model_diagnostics", "finops_operating_review",
        "decision_intelligence",
    ] = "spend"
    context: Dict[str, Any] = Field(default_factory=dict)
    anomaly: Optional[Dict[str, Any]] = None


class GenAICopilotPackRequest(BaseModel):
    cloud_provider: str = "all"
    include: List[
        Literal[
            "spend",
            "budget_risk",
            "waste_insights",
            "optimization_roadmap",
            "executive_narrative",
            "commitment_strategy",
            "tagging_strategy",
            "sustainability_narrative",
            "chargeback_narrative",
            "rightsizing_brief",
            "vendor_negotiation_brief",
            "forecast_model_diagnostics",
            "finops_operating_review",
            "decision_intelligence",
        ]
    ] = Field(default_factory=lambda: ["spend", "optimization_roadmap", "executive_narrative"])


class RagGuidanceRequest(BaseModel):
    analysis_type: str = "finops_operating_review"
    cloud_provider: str = "all"
    top_k: int = Field(default=4, ge=1, le=8)
    context: Dict[str, Any] = Field(default_factory=dict)


class HybridAdvisorResponse(BaseModel):
    generated_at: str
    cloud_provider: str
    deterministic: Dict[str, Any]
    advisory: Dict[str, Any]
    source_of_truth: Literal["deterministic"] = "deterministic"


async def _deterministic_recommendations(
    cloud_provider: str,
    current_monthly_spend: float,
    cost_breakdown: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Return deterministic recommendation rows mapped for advisor workflows."""
    rec_result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": current_monthly_spend,
                "cost_breakdown": cost_breakdown,
            }
        ),
        {},
    )
    rows = rec_result.get("recommendations", [])
    mapped: List[Dict[str, Any]] = []
    for row in rows:
        mapped.append(
            {
                "id": row.get("id"),
                "service": row.get("service", "unknown"),
                "title": row.get("description", "Optimization recommendation"),
                "description": row.get("description", ""),
                "savings_monthly_usd": float(row.get("savings_annual_usd", 0) or 0) / 12,
                "roi_percent": float(row.get("roi_percent", 0) or 0),
                "payback_months": float(row.get("payback_months", 0) or 0),
            }
        )
    return mapped


@router.get("/advisor/hybrid", response_model=HybridAdvisorResponse)
async def advisor_hybrid(
    cloud_provider: str = "all",
    narrative_type: Literal[
        "waste_insights", "optimization_roadmap", "executive_narrative",
        "tagging_strategy", "sustainability_narrative", "finops_operating_review",
    ] = "optimization_roadmap",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Hybrid advisor payload: deterministic findings + GenAI narrative.

    Deterministic analytics and recommendation values remain authoritative.
    GenAI output is advisory context layered on top.
    """
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    customer_id = _customer_id_for_org(membership)
    permission = ScanningManager(db).get_permission_status(customer_id)

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
            }
        ),
        {},
    )

    waste_result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    efficiency_result = _safe_json_load(
        await finops_analytics.get_cost_efficiency_score(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "waste_rate_percent": analytics_result.get("unit_metrics", {}).get("estimated_waste_rate_percent", 18.0),
                "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get("anomaly_density_per_10k", 8.0),
                "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get("budget_utilization_percent", 85.0),
            }
        ),
        {},
    )

    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    tagging_result = _safe_json_load(
        await finops_analytics.get_tagging_coverage_analytics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    chargeback_result = _safe_json_load(
        await finops_analytics.get_chargeback_summary(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    forecast_result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": 12,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    maturity_result = _safe_json_load(
        await finops_analytics.get_maturity_assessment(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "history_coverage_months": 0,
                "scheduler_enabled": False,
                "auto_remediate": False,
            }
        ),
        {},
    )

    recommendation_rows = await _deterministic_recommendations(
        cloud_provider=cloud_provider,
        current_monthly_spend=context["total_cost"],
        cost_breakdown=context["breakdown"],
    )
    top_actions = sorted(
        recommendation_rows,
        key=lambda r: (r.get("savings_monthly_usd", 0), r.get("roi_percent", 0)),
        reverse=True,
    )[:5]

    genai_context: Dict[str, Any] = {
        "current_monthly_spend_usd": analytics_result.get("current_monthly_spend_usd", context["total_cost"]),
        "mom_change_percent": analytics_result.get("mom_change_percent"),
        "risk_score": analytics_result.get("risk_score", 0),
        "maturity_level": maturity_result.get("maturity_level", "walk"),
        "overall_score": efficiency_result.get("overall_score", 0),
        "grade": efficiency_result.get("grade", "C"),
        "improvement_focus": efficiency_result.get("improvement_focus", []),
        "total_estimated_waste_usd": waste_result.get("total_estimated_waste_usd", 0),
        "total_waste_rate_percent": waste_result.get("total_waste_rate_percent", 0),
        "categories": waste_result.get("categories", []),
        "quick_wins": waste_result.get("quick_wins", []),
        "total_annual_opportunity_usd": commitment_gap_result.get("total_annual_opportunity_usd", 0),
        "priority_provider": commitment_gap_result.get("priority_provider"),
        "top_opportunities": top_actions,
        "budget_monthly_usd": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
        "budget_guardrails": forecast_result.get("budget_guardrails") or {},
        "cost_velocity_pct_mom": forecast_result.get("cost_velocity_pct_mom"),
        "spend_at_risk_usd": analytics_result.get("spend_at_risk_usd", 0),
        "coverage_gap_percent": tagging_result.get("coverage_gap_percent", 0.0),
        "unallocated_percent": chargeback_result.get("unallocated_percent", 0.0),
    }
    rag = _rag_context_for_analysis(
        analysis_type=narrative_type,
        cloud_provider=cloud_provider,
        context=genai_context,
    )

    narrative: Optional[str]
    prompt: str
    if narrative_type == "waste_insights":
        narrative, prompt = genai_advisor.generate_waste_insights(genai_context)
    elif narrative_type == "executive_narrative":
        narrative, prompt = genai_advisor.generate_executive_narrative(genai_context)
    elif narrative_type == "tagging_strategy":
        narrative, prompt = genai_advisor.generate_tagging_strategy(genai_context)
    elif narrative_type == "sustainability_narrative":
        narrative, prompt = genai_advisor.generate_sustainability_narrative(genai_context)
    elif narrative_type == "finops_operating_review":
        narrative, prompt = genai_advisor.generate_finops_operating_review(genai_context)
    else:
        narrative, prompt = genai_advisor.generate_optimization_roadmap(genai_context)

    return {
        "generated_at": _utcnow().isoformat(),
        "cloud_provider": cloud_provider,
        "deterministic": {
            "analytics": analytics_result,
            "waste": waste_result,
            "efficiency": efficiency_result,
            "commitment_gap": commitment_gap_result,
            "recommendations": top_actions,
        },
        "advisory": {
            "narrative_type": narrative_type,
            "narrative": narrative,
            "prompt": prompt,
            "rag": rag,
            "genai_configured": genai_advisor._is_configured(),
            "fallback_mode": narrative is None,
        },
        "source_of_truth": "deterministic",
    }


@router.post("/genai/analyze")
async def genai_analyze(
    request: GenAIAnalyzeRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Call OCI GenAI to generate a narrative for the requested analysis type.

    Falls back to returning the prompt when OCI GenAI is not configured,
    so the frontend Cost Advisor can use it directly.
    """
    _ = current_user
    ctx = request.context
    if not ctx.get("current_monthly_spend_usd"):
        cost_ctx = await _cost_context(membership, db, "month", "all")
        ctx.setdefault("current_monthly_spend_usd", cost_ctx.get("total_cost", 0))
    rag = _rag_context_for_analysis(
        analysis_type=request.analysis_type,
        cloud_provider=str(ctx.get("cloud_provider", "all") or "all"),
        context=ctx,
    )

    narrative: Optional[str] = None
    prompt: str = ""

    if request.analysis_type == "spend":
        narrative, prompt = genai_advisor.generate_spend_narrative(ctx)
    elif request.analysis_type == "anomaly":
        anomaly_ctx = request.anomaly or {}
        narrative, prompt = genai_advisor.generate_anomaly_explanation(anomaly_ctx, ctx)
    elif request.analysis_type == "optimization":
        narrative, prompt = genai_advisor.generate_optimization_brief(ctx)
    elif request.analysis_type == "maturity":
        narrative, prompt = genai_advisor.generate_maturity_narrative(ctx)
    elif request.analysis_type == "budget_risk":
        guardrails = ctx.get("budget_guardrails") or ctx
        narrative, prompt = genai_advisor.generate_budget_risk_alert(guardrails, ctx)
    elif request.analysis_type == "waste_insights":
        narrative, prompt = genai_advisor.generate_waste_insights(ctx)
    elif request.analysis_type == "optimization_roadmap":
        narrative, prompt = genai_advisor.generate_optimization_roadmap(ctx)
    elif request.analysis_type == "executive_narrative":
        narrative, prompt = genai_advisor.generate_executive_narrative(ctx)
    elif request.analysis_type == "commitment_strategy":
        narrative, prompt = genai_advisor.generate_commitment_strategy(ctx)
    elif request.analysis_type == "tagging_strategy":
        narrative, prompt = genai_advisor.generate_tagging_strategy(ctx)
    elif request.analysis_type == "sustainability_narrative":
        narrative, prompt = genai_advisor.generate_sustainability_narrative(ctx)
    elif request.analysis_type == "chargeback_narrative":
        narrative, prompt = genai_advisor.generate_chargeback_narrative(ctx)
    elif request.analysis_type == "cross_provider_comparison_brief":
        narrative, prompt = genai_advisor.generate_cross_provider_comparison_brief(ctx)
    elif request.analysis_type == "alert_triage":
        alerts = ctx.pop("alerts", [])
        narrative, prompt = genai_advisor.generate_alert_triage(alerts, ctx)
    elif request.analysis_type == "rightsizing_brief":
        narrative, prompt = genai_advisor.generate_rightsizing_brief(ctx)
    elif request.analysis_type == "vendor_negotiation_brief":
        narrative, prompt = genai_advisor.generate_vendor_negotiation_brief(ctx)
    elif request.analysis_type == "forecast_model_diagnostics":
        narrative, prompt = genai_advisor.generate_forecast_model_diagnostics(ctx)
    elif request.analysis_type == "finops_operating_review":
        narrative, prompt = genai_advisor.generate_finops_operating_review(ctx)
    elif request.analysis_type == "decision_intelligence":
        narrative, prompt = genai_advisor.generate_decision_intelligence(ctx)

    return {
        "analysis_type": request.analysis_type,
        "narrative": narrative,
        "prompt": prompt,
        "rag": rag,
        "genai_configured": genai_advisor._is_configured(),
        "fallback_mode": narrative is None,
    }


@router.post("/genai/rag-guidance")
async def genai_rag_guidance(
    request: RagGuidanceRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Return retrieved FinOps guidance snippets for an analysis context."""
    _ = current_user
    context = dict(request.context or {})
    if not context:
        cost_ctx = await _cost_context(membership, db, "month", request.cloud_provider)
        context = {
            "current_monthly_spend_usd": float(cost_ctx.get("total_cost", 0.0) or 0.0),
            "cost_breakdown": cost_ctx.get("breakdown") or {},
        }

    payload = finops_rag.retrieve_guidance(
        analysis_type=request.analysis_type,
        cloud_provider=request.cloud_provider,
        context=context,
        top_k=request.top_k,
    )
    return {
        "generated_at": _utcnow().isoformat(),
        "analysis_type": request.analysis_type,
        "cloud_provider": request.cloud_provider,
        "rag": payload,
    }


@router.post("/genai/copilot-pack")
async def genai_copilot_pack(
    request: GenAICopilotPackRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Return deterministic context plus multiple GenAI narratives in one call."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", request.cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    forecast_result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": 12,
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    tagging_result = _safe_json_load(
        await finops_analytics.get_tagging_coverage_analytics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    chargeback_result = _safe_json_load(
        await finops_analytics.get_chargeback_summary(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    base_context = dict(analytics_result)
    base_context.update(
        {
            "budget_monthly_usd": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            "budget_guardrails": forecast_result.get("budget_guardrails") or {},
            "forecast_quality": forecast_result.get("forecast_quality") or {},
            "p90_monthly_usd": (forecast_result.get("forecast") or [{}])[0].get("p90", context["total_cost"]),
            "total_annual_opportunity_usd": commitment_gap_result.get("total_annual_opportunity_usd", 0),
            "priority_provider": commitment_gap_result.get("priority_provider"),
            "provider_gaps": commitment_gap_result.get("provider_gaps", []),
            "coverage_gap_percent": tagging_result.get("coverage_gap_percent", 0.0),
            "unallocated_percent": chargeback_result.get("unallocated_percent", 0.0),
            "top_opportunities": commitment_gap_result.get("provider_gaps", []),
        }
    )
    base_rag = _rag_context_for_analysis(
        analysis_type="finops_operating_review",
        cloud_provider=request.cloud_provider,
        context=base_context,
    )

    narratives: Dict[str, Dict[str, Any]] = {}
    requested = request.include or []
    for item in requested:
        item_context = dict(base_context)
        item_rag = _rag_context_for_analysis(
            analysis_type=item,
            cloud_provider=request.cloud_provider,
            context=item_context,
        )
        if item == "spend":
            narrative, prompt = genai_advisor.generate_spend_narrative(item_context)
        elif item == "budget_risk":
            narrative, prompt = genai_advisor.generate_budget_risk_alert(
                item_context.get("budget_guardrails", {}), item_context
            )
        elif item == "waste_insights":
            narrative, prompt = genai_advisor.generate_waste_insights(item_context)
        elif item == "optimization_roadmap":
            narrative, prompt = genai_advisor.generate_optimization_roadmap(item_context)
        elif item == "executive_narrative":
            narrative, prompt = genai_advisor.generate_executive_narrative(item_context)
        elif item == "tagging_strategy":
            narrative, prompt = genai_advisor.generate_tagging_strategy(item_context)
        elif item == "sustainability_narrative":
            narrative, prompt = genai_advisor.generate_sustainability_narrative(item_context)
        elif item == "chargeback_narrative":
            narrative, prompt = genai_advisor.generate_chargeback_narrative(item_context)
        elif item == "rightsizing_brief":
            narrative, prompt = genai_advisor.generate_rightsizing_brief(item_context)
        elif item == "vendor_negotiation_brief":
            narrative, prompt = genai_advisor.generate_vendor_negotiation_brief(item_context)
        elif item == "forecast_model_diagnostics":
            historical_monthly_spend = _historical_monthly_spend_from_snapshots(
                db=db,
                customer_id=customer_id,
                cloud_provider=request.cloud_provider,
                months=18,
            )
            diagnostics_result = _safe_json_load(
                await finops_analytics.get_forecast_model_diagnostics(
                    {
                        "months": 12,
                        "cloud_provider": request.cloud_provider,
                        "current_monthly_spend": context["total_cost"],
                        "cost_breakdown": context["breakdown"],
                        "historical_monthly_spend": historical_monthly_spend,
                        "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                    }
                ),
                {},
            )
            diag_ctx = dict(diagnostics_result)
            item_rag = _rag_context_for_analysis(
                analysis_type="forecast_model_diagnostics",
                cloud_provider=request.cloud_provider,
                context=diag_ctx,
            )
            narrative, prompt = genai_advisor.generate_forecast_model_diagnostics(diag_ctx)
        elif item == "finops_operating_review":
            narrative, prompt = genai_advisor.generate_finops_operating_review(item_context)
        elif item == "decision_intelligence":
            historical_monthly_spend = _historical_monthly_spend_from_snapshots(
                db=db,
                customer_id=customer_id,
                cloud_provider=request.cloud_provider,
                months=18,
            )
            decision_result = _safe_json_load(
                await finops_analytics.get_decision_intelligence(
                    {
                        "months": 12,
                        "cloud_provider": request.cloud_provider,
                        "current_monthly_spend": context["total_cost"],
                        "cost_breakdown": context["breakdown"],
                        "historical_monthly_spend": historical_monthly_spend,
                        "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                    }
                ),
                {},
            )
            decision_ctx = dict(decision_result)
            item_rag = _rag_context_for_analysis(
                analysis_type="decision_intelligence",
                cloud_provider=request.cloud_provider,
                context=decision_ctx,
            )
            narrative, prompt = genai_advisor.generate_decision_intelligence(decision_ctx)
        else:
            narrative, prompt = genai_advisor.generate_commitment_strategy(item_context)

        narratives[item] = {
            "narrative": narrative,
            "prompt": prompt,
            "rag": item_rag,
            "fallback_mode": narrative is None,
        }

    return {
        "generated_at": _utcnow().isoformat(),
        "cloud_provider": request.cloud_provider,
        "deterministic_context": {
            "analytics": analytics_result,
            "forecast": {
                "forecast_quality": forecast_result.get("forecast_quality"),
                "budget_guardrails": forecast_result.get("budget_guardrails"),
                "downside_risk": forecast_result.get("downside_risk"),
            },
            "commitment_gap": commitment_gap_result,
            "tagging_coverage": tagging_result,
            "chargeback_summary": chargeback_result,
            "rag": base_rag,
        },
        "narratives": narratives,
        "genai_configured": genai_advisor._is_configured(),
    }




@router.get("/analytics/cloud-waste")
async def analytics_cloud_waste(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Categorised cloud waste breakdown with remediation guidance."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/efficiency-score")
async def analytics_efficiency_score(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Composite FinOps efficiency score (0-100) across 6 weighted dimensions."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    result = _safe_json_load(
        await finops_analytics.get_cost_efficiency_score({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "waste_rate_percent": analytics_result.get("unit_metrics", {}).get("estimated_waste_rate_percent", 18.0),
            "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get("anomaly_density_per_10k", 8.0),
            "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get("budget_utilization_percent", 85.0),
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/commitment-gap")
async def analytics_commitment_gap(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Per-provider commitment coverage gap with savings scenarios and breakeven."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/optimization-portfolio")
async def analytics_optimization_portfolio(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Rank optimization actions by portfolio score (savings, ROI, payback, effort)."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)

    recommendation_rows = await _deterministic_recommendations(
        cloud_provider=cloud_provider,
        current_monthly_spend=context["total_cost"],
        cost_breakdown=context["breakdown"],
    )

    result = _safe_json_load(
        await finops_analytics.get_optimization_portfolio(
            {
                "current_monthly_spend": context["total_cost"],
                "recommendations": recommendation_rows,
            }
        ),
        {},
    )

    narrative, prompt = genai_advisor.generate_optimization_roadmap(
        {
            "current_monthly_spend_usd": context["total_cost"],
            "top_opportunities": result.get("ranked_actions", [])[:5],
            "total_annual_savings_usd": result.get("total_annual_savings_usd", 0),
        }
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/operating-review")
async def analytics_operating_review(
    cloud_provider: str = "all",
    months: int = Query(default=12, ge=1, le=24),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deterministic FinOps operating review pack with optional GenAI weekly narrative."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )
    waste_result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    efficiency_result = _safe_json_load(
        await finops_analytics.get_cost_efficiency_score(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "waste_rate_percent": analytics_result.get("unit_metrics", {}).get(
                    "estimated_waste_rate_percent", 18.0
                ),
                "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get(
                    "anomaly_density_per_10k", 8.0
                ),
                "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get(
                    "budget_utilization_percent", 85.0
                ),
            }
        ),
        {},
    )
    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    tagging_result = _safe_json_load(
        await finops_analytics.get_tagging_coverage_analytics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    chargeback_result = _safe_json_load(
        await finops_analytics.get_chargeback_summary(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    forecast_result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    recommendation_rows = await _deterministic_recommendations(
        cloud_provider=cloud_provider,
        current_monthly_spend=context["total_cost"],
        cost_breakdown=context["breakdown"],
    )

    result = _safe_json_load(
        await finops_analytics.get_finops_operating_review(
            {
                "cloud_provider": cloud_provider,
                "months": months,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "historical_monthly_spend": historical_monthly_spend,
                "analytics_result": analytics_result,
                "waste_result": waste_result,
                "efficiency_result": efficiency_result,
                "commitment_gap_result": commitment_gap_result,
                "tagging_result": tagging_result,
                "chargeback_result": chargeback_result,
                "forecast_result": forecast_result,
                "recommendations": recommendation_rows,
            }
        ),
        {},
    )
    genai_context = dict(result.get("genai_context") or {})
    rag = _rag_context_for_analysis(
        analysis_type="finops_operating_review",
        cloud_provider=cloud_provider,
        context=genai_context,
    )
    narrative, prompt = genai_advisor.generate_finops_operating_review(genai_context)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["rag"] = rag
    result["cost_context"] = context
    return result


@router.get("/analytics/decision-intelligence")
async def analytics_decision_intelligence(
    cloud_provider: str = "all",
    months: int = Query(default=12, ge=1, le=24),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Scenario frontier for executive FinOps decision-making."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )

    result = _safe_json_load(
        await finops_analytics.get_decision_intelligence(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "monthly_savings": float(permission.get("monthly_budget_usd", 0.0) or 0.0) * 0.08,
            }
        ),
        {},
    )

    genai_context = dict(result)
    rag = _rag_context_for_analysis(
        analysis_type="decision_intelligence",
        cloud_provider=cloud_provider,
        context=genai_context,
    )
    genai_context["rag_brief"] = rag.get("rag_brief", "")
    narrative, prompt = genai_advisor.generate_decision_intelligence(genai_context)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["rag"] = rag
    result["cost_context"] = context
    return result


@router.get("/analytics/control-tower")
async def analytics_control_tower(
    cloud_provider: str = "all",
    months: int = Query(default=12, ge=1, le=24),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Unified FinOps control tower across forecast risk, waste, commitments, governance, and decisions."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    budget_monthly = float(permission.get("monthly_budget_usd", 0.0) or 0.0)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )
    recommendation_rows = await _deterministic_recommendations(
        cloud_provider=cloud_provider,
        current_monthly_spend=context["total_cost"],
        cost_breakdown=context["breakdown"],
    )
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "budget_monthly": budget_monthly,
                "monthly_savings": 0.0,
                "anomalies": 0,
            }
        ),
        {},
    )
    forecast_diagnostics = _safe_json_load(
        await finops_analytics.get_forecast_diagnostics(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": budget_monthly,
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    waste_result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    tagging_result = _safe_json_load(
        await finops_analytics.get_tagging_coverage_analytics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    decision_result = _safe_json_load(
        await finops_analytics.get_decision_intelligence(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": budget_monthly,
                "monthly_savings": 0.0,
                "recommendations": recommendation_rows,
            }
        ),
        {},
    )
    result = _safe_json_load(
        await finops_analytics.get_finops_control_tower(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": budget_monthly,
                "analytics_result": analytics_result,
                "forecast_diagnostics": forecast_diagnostics,
                "waste_result": waste_result,
                "commitment_gap_result": commitment_gap_result,
                "tagging_result": tagging_result,
                "decision_result": decision_result,
                "recommendations": recommendation_rows,
            }
        ),
        {},
    )
    genai_context = dict(result.get("genai_context") or {})
    rag_by_lane = {
        "control_tower": _rag_context_for_analysis(
            analysis_type="finops_control_tower",
            cloud_provider=cloud_provider,
            context=genai_context,
        ),
        "forecast_risk": _rag_context_for_analysis(
            analysis_type="budget_risk",
            cloud_provider=cloud_provider,
            context=forecast_diagnostics,
        ),
        "waste": _rag_context_for_analysis(
            analysis_type="waste_insights",
            cloud_provider=cloud_provider,
            context=waste_result,
        ),
        "commitment": _rag_context_for_analysis(
            analysis_type="commitment_strategy",
            cloud_provider=cloud_provider,
            context=commitment_gap_result,
        ),
        "governance": _rag_context_for_analysis(
            analysis_type="tagging_strategy",
            cloud_provider=cloud_provider,
            context=tagging_result,
        ),
        "decision": _rag_context_for_analysis(
            analysis_type="decision_intelligence",
            cloud_provider=cloud_provider,
            context=decision_result,
        ),
    }
    genai_context["rag_brief"] = "\n".join(
        str(item.get("rag_brief") or "")
        for item in rag_by_lane.values()
        if item.get("rag_brief")
    )
    narrative, prompt = genai_advisor.generate_executive_narrative(genai_context)
    result["rag"] = rag_by_lane["control_tower"]
    result["rag_by_lane"] = rag_by_lane
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/finops-intelligence")
async def analytics_finops_intelligence(
    focus: Literal[
        "finops_operating_review",
        "forecast_model_diagnostics",
        "commitment_strategy",
        "tagging_strategy",
        "sustainability_narrative",
        "executive_narrative",
        "decision_intelligence",
    ] = "finops_operating_review",
    cloud_provider: str = "all",
    months: int = Query(default=12, ge=1, le=24),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deep FinOps intelligence: deterministic analytics + RAG + GenAI narrative."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    deterministic: Dict[str, Any] = {}
    genai_context: Dict[str, Any] = {}

    if focus == "forecast_model_diagnostics":
        deterministic = _safe_json_load(
            await finops_analytics.get_forecast_model_diagnostics(
                {
                    "months": months,
                    "cloud_provider": cloud_provider,
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                    "historical_monthly_spend": historical_monthly_spend,
                    "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                    "fallback_monthly_spend": 0,
                }
            ),
            {},
        )
        genai_context = dict(deterministic)
        rag = _rag_context_for_analysis(
            analysis_type="forecast_model_diagnostics",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_forecast_model_diagnostics(genai_context)

    elif focus == "commitment_strategy":
        deterministic = _safe_json_load(
            await finops_analytics.get_commitment_gap_analysis(
                {
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                }
            ),
            {},
        )
        genai_context = {
            "current_monthly_spend_usd": context["total_cost"],
            "total_annual_opportunity_usd": deterministic.get("total_annual_opportunity_usd", 0.0),
            "priority_provider": deterministic.get("priority_provider"),
            "provider_gaps": deterministic.get("provider_gaps", []),
        }
        rag = _rag_context_for_analysis(
            analysis_type="commitment_strategy",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_commitment_strategy(genai_context)

    elif focus == "tagging_strategy":
        deterministic = _safe_json_load(
            await finops_analytics.get_tagging_coverage_analytics(
                {
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                }
            ),
            {},
        )
        genai_context = dict(deterministic)
        rag = _rag_context_for_analysis(
            analysis_type="tagging_strategy",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_tagging_strategy(genai_context)

    elif focus == "sustainability_narrative":
        deterministic = _safe_json_load(
            await finops_analytics.get_sustainability_metrics(
                {
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                }
            ),
            {},
        )
        genai_context = dict(deterministic)
        rag = _rag_context_for_analysis(
            analysis_type="sustainability_narrative",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_sustainability_narrative(genai_context)

    elif focus == "executive_narrative":
        waste_result = _safe_json_load(
            await finops_analytics.get_cloud_waste_analysis(
                {
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                }
            ),
            {},
        )
        efficiency_result = _safe_json_load(
            await finops_analytics.get_cost_efficiency_score(
                {
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                    "waste_rate_percent": analytics_result.get("unit_metrics", {}).get(
                        "estimated_waste_rate_percent", 18.0
                    ),
                    "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get(
                        "anomaly_density_per_10k", 8.0
                    ),
                    "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get(
                        "budget_utilization_percent", 85.0
                    ),
                }
            ),
            {},
        )
        deterministic = {
            "analytics": analytics_result,
            "waste": waste_result,
            "efficiency": efficiency_result,
        }
        genai_context = dict(analytics_result)
        genai_context.update(waste_result)
        genai_context.update(efficiency_result)
        genai_context["budget_monthly_usd"] = float(permission.get("monthly_budget_usd", 0.0) or 0.0)
        rag = _rag_context_for_analysis(
            analysis_type="executive_narrative",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_executive_narrative(genai_context)

    elif focus == "decision_intelligence":
        deterministic = _safe_json_load(
            await finops_analytics.get_decision_intelligence(
                {
                    "months": months,
                    "cloud_provider": cloud_provider,
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                    "historical_monthly_spend": historical_monthly_spend,
                    "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                    "monthly_savings": 0.0,
                }
            ),
            {},
        )
        genai_context = dict(deterministic)
        rag = _rag_context_for_analysis(
            analysis_type="decision_intelligence",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        genai_context["rag_brief"] = rag.get("rag_brief", "")
        narrative, prompt = genai_advisor.generate_decision_intelligence(genai_context)

    else:
        deterministic = _safe_json_load(
            await finops_analytics.get_finops_operating_review(
                {
                    "cloud_provider": cloud_provider,
                    "months": months,
                    "current_monthly_spend": context["total_cost"],
                    "cost_breakdown": context["breakdown"],
                    "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                    "historical_monthly_spend": historical_monthly_spend,
                    "analytics_result": analytics_result,
                }
            ),
            {},
        )
        genai_context = dict(deterministic.get("genai_context") or {})
        rag = _rag_context_for_analysis(
            analysis_type="finops_operating_review",
            cloud_provider=cloud_provider,
            context=genai_context,
        )
        narrative, prompt = genai_advisor.generate_finops_operating_review(genai_context)

    return {
        "generated_at": _utcnow().isoformat(),
        "focus": focus,
        "cloud_provider": cloud_provider,
        "deterministic": deterministic,
        "rag": rag,
        "advisory": {
            "narrative": narrative,
            "prompt": prompt,
            "fallback_mode": narrative is None,
            "genai_configured": genai_advisor._is_configured(),
        },
        "cost_context": context,
    }


@router.get("/analytics/tagging-coverage")
async def analytics_tagging_coverage(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deep tagging compliance: per-tag analysis, allocation readiness score, and financial risk from untagged spend."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_tagging_coverage_analytics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    narrative, prompt = genai_advisor.generate_tagging_strategy(
        {
            "current_monthly_spend_usd": context["total_cost"],
            "coverage_percent": result.get("coverage_percent", 0),
            "grade": result.get("grade", "C"),
            "coverage_gap_percent": result.get("coverage_gap_percent", 0),
            "untagged_spend_annual_usd": result.get("untagged_spend_annual_usd", 0),
            "critical_tag_gaps": result.get("critical_tag_gaps", []),
            "allocation_readiness_score": result.get("allocation_readiness_score", 0),
            "enforcement_recommendations": result.get("enforcement_recommendations", []),
        }
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/sustainability")
async def analytics_sustainability(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Carbon footprint estimation with regional intensity modifiers, sustainability score, and reduction opportunities."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_sustainability_metrics(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    narrative, prompt = genai_advisor.generate_sustainability_narrative(
        {
            "current_monthly_spend_usd": context["total_cost"],
            "total_kg_co2e_monthly": result.get("total_kg_co2e_monthly", 0),
            "total_tonnes_co2e_annual": result.get("total_tonnes_co2e_annual", 0),
            "current_renewable_energy_percent": result.get("current_renewable_energy_percent", 0),
            "sustainability_score": result.get("sustainability_score", 0),
            "sustainability_grade": result.get("sustainability_grade", "C"),
            "provider_emissions": result.get("provider_emissions", []),
            "reduction_opportunities": result.get("reduction_opportunities", {}),
            "recommendations": result.get("recommendations", []),
        }
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/cross-provider-comparison")
async def analytics_cross_provider_comparison(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Multi-cloud efficiency benchmarking: health scores, HHI concentration risk, and workload arbitrage opportunities."""
    _ = current_user
    context = await _cost_context(membership, db, "month", "all")
    result = _safe_json_load(
        await finops_analytics.get_cross_provider_comparison(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    narrative, prompt = genai_advisor.generate_cross_provider_comparison_brief(
        {
            "current_monthly_spend_usd": context["total_cost"],
            "total_monthly_spend_usd": result.get("total_monthly_spend_usd", context["total_cost"]),
            "best_performing_provider": result.get("best_performing_provider"),
            "lowest_health_provider": result.get("lowest_health_provider"),
            "arbitrage_opportunities": result.get("arbitrage_opportunities", []),
            "concentration_risk": result.get("concentration_risk", "medium"),
            "providers": result.get("providers", []),
        }
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


_SERVICE_FOCUS_TERMS: Dict[str, List[str]] = {
    "database": ["database", "db", "rds", "aurora", "postgres", "mysql", "sql", "dynamodb", "cosmos", "spanner", "redis"],
    "serverless": ["serverless", "lambda", "function", "functions", "cloud run", "fargate", "function app"],
    "storage": ["storage", "bucket", "blob", "disk", "volume", "object storage", "ebs", "efs", "s3"],
    "network": ["network", "load balancer", "egress", "bandwidth", "nat", "gateway", "cdn"],
    "analytics": ["analytics", "bigquery", "redshift", "athena", "emr", "databricks", "warehouse"],
    "cache": ["cache", "redis", "memcached", "elasticache"],
    "messaging": ["queue", "pubsub", "kafka", "event hub", "service bus", "sqs", "sns"],
    "ai-ml": ["ai", "ml", "machine learning", "gpu", "inference", "training", "bedrock", "vertex", "openai", "genai"],
    "compute": ["compute", "vm", "virtual machine", "instance", "ec2", "node"],
    "kubernetes": [
        "kubernetes", "k8s", "pod", "namespace", "cluster", "gke", "aks", "eks", "oke",
        "container", "containers", "docker", "fargate", "ecs", "ecr", "cloud run", "artifact registry",
    ],
}


def _service_matches_focus(service_name: str, focus: Optional[str]) -> bool:
    if not focus:
        return True
    focus_key = str(focus).strip().lower()
    if not focus_key:
        return True
    service_l = str(service_name or "").strip().lower()
    terms = _SERVICE_FOCUS_TERMS.get(focus_key)
    if terms:
        return any(term in service_l for term in terms)
    return focus_key in service_l


@router.get("/analytics/service-hotspots")
async def analytics_service_hotspots(
    period: str = "month",
    cloud_provider: str = "all",
    focus: Optional[str] = None,
    limit: int = 10,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Top costly cloud services across providers (deterministic, non-GenAI)."""
    _ = current_user
    require_live_provider_data = Config().require_live_provider_data
    providers = ["aws", "azure", "gcp", "oci"] if cloud_provider == "all" else [cloud_provider]
    providers = [p for p in providers if p in {"aws", "azure", "gcp", "oci"}]
    capped_limit = max(1, min(limit, 100))

    diagnostics = {
        d.provider: bool(d.configured)
        for d in _provider_diagnostics()
    }

    if require_live_provider_data and not any(diagnostics.get(provider, False) for provider in providers):
        raise HTTPException(
            status_code=412,
            detail=(
                "Live provider data is required, but none of the requested providers are configured "
                "for runtime API access on this backend host."
            ),
        )

    imported_rows = []
    if not require_live_provider_data:
        imported_rows = _get_imported_cost_rows(
            db,
            _organization_id_for_membership(membership),
            _customer_id_for_org(membership),
            cloud_provider,
        )

    def add_item(bucket: Dict[tuple[str, str], Dict[str, Any]], provider_key: str, service_name: str, cost: float, source: str) -> None:
        key = (provider_key, service_name)
        if key not in bucket:
            bucket[key] = {
                "provider": provider_key,
                "service": service_name,
                "monthly_cost_usd": 0.0,
                "source": source,
            }
        bucket[key]["monthly_cost_usd"] += float(cost or 0.0)

    service_map: Dict[tuple[str, str], Dict[str, Any]] = {}

    for provider in providers:
        used_live = False
        if diagnostics.get(provider, False):
            summary = await _cost_summary_for_provider(provider, period)
            if "error" not in summary:
                used_live = True
                for row in summary.get("top_services", []):
                    if not isinstance(row, dict):
                        continue
                    service_name = str(
                        row.get("service")
                        or row.get("name")
                        or row.get("service_name")
                        or "unknown-service"
                    ).strip() or "unknown-service"
                    cost = float(
                        row.get("cost_usd")
                        or row.get("cost")
                        or row.get("amount")
                        or 0.0
                    )
                    if cost <= 0:
                        continue
                    add_item(service_map, provider, service_name, cost, "live_provider_api")

        if used_live:
            continue

        if not require_live_provider_data:
            for row in imported_rows:
                row_provider = str(getattr(row, "provider", "") or "").strip().lower()
                if row_provider != provider:
                    continue
                service_name = str(getattr(row, "service_name", "") or "").strip() or "imported-service"
                cost = float(getattr(row, "cost_usd", 0.0) or 0.0)
                if cost <= 0:
                    continue
                add_item(service_map, provider, service_name, cost, "csv_import")

    items = list(service_map.values())
    if focus:
        items = [item for item in items if _service_matches_focus(str(item.get("service", "")), focus)]

    items.sort(key=lambda item: float(item.get("monthly_cost_usd", 0.0) or 0.0), reverse=True)
    items = items[:capped_limit]
    for item in items:
        item["monthly_cost_usd"] = round(float(item.get("monthly_cost_usd", 0.0) or 0.0), 2)

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "period": period,
        "cloud_provider": cloud_provider,
        "focus": focus.strip().lower() if isinstance(focus, str) and focus.strip() else None,
        "total_monthly_cost_usd": round(sum(float(item.get("monthly_cost_usd", 0.0) or 0.0) for item in items), 2),
        "items": items,
    }


_RESOURCE_QUERY_STOPWORDS = {
    "who", "created", "create", "creator", "owner", "owned", "by",
    "how", "much", "cost", "costed", "since", "when", "was", "is", "the",
    "resource", "service", "cloud", "environment", "in", "for", "of", "to",
    "qual", "quem", "criou", "criado", "quanto", "custa", "custou", "desde",
    "quien", "creo", "creó", "cuanto", "cuánto", "cuesta", "costó", "desde",
}


def _resource_query_tokens(raw_query: str) -> List[str]:
    tokens = re.findall(r"[a-z0-9][a-z0-9._:/-]{1,}", str(raw_query or "").lower())
    return [
        token
        for token in tokens
        if token not in _RESOURCE_QUERY_STOPWORDS and len(token) >= 3
    ]


def _best_effort_owner_from_payload(payload: Any) -> Optional[str]:
    if isinstance(payload, dict):
        keys = [
            "created_by",
            "createdBy",
            "creator",
            "owner",
            "owner_email",
            "provisioned_by",
            "requested_by",
            "user",
            "principal",
        ]
        for key in keys:
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        # nested patterns
        for nested_key in ("tags", "labels", "metadata"):
            nested = payload.get(nested_key)
            nested_owner = _best_effort_owner_from_payload(nested)
            if nested_owner:
                return nested_owner
    return None


@router.get("/analytics/resource-intelligence")
async def analytics_resource_intelligence(
    query: str,
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Best-effort resource intelligence: owner/creator, first seen, and observed cost since first sighting."""
    _ = current_user
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    providers = {"aws", "azure", "gcp", "oci"}

    requested_provider = str(cloud_provider or "all").strip().lower()
    provider_filter = None if requested_provider == "all" else requested_provider

    raw_tokens = _resource_query_tokens(query)
    token_set = set(raw_tokens)

    imported_rows = _get_imported_cost_rows(
        db,
        organization_id=org_id,
        customer_id=customer_id,
        cloud_provider=provider_filter or "all",
    )

    account_query = db.query(ProviderAccount).filter(
        ProviderAccount.organization_id == org_id,
        ProviderAccount.customer_id == customer_id,
        ProviderAccount.is_active.is_(True),
    )
    if provider_filter:
        account_query = account_query.filter(ProviderAccount.provider == provider_filter)
    accounts = account_query.all()

    snapshots_query = db.query(ProviderAccountSnapshot).filter(
        ProviderAccountSnapshot.organization_id == org_id,
        ProviderAccountSnapshot.customer_id == customer_id,
    )
    snapshots = snapshots_query.all()

    snapshots_by_account: Dict[int, Dict[str, Any]] = {}
    for snap in snapshots:
        bucket = snapshots_by_account.setdefault(
            int(snap.provider_account_id),
            {
                "observed_total_cost_usd": 0.0,
                "first_seen_at": None,
                "last_seen_at": None,
                "latest_monthly_cost_usd": 0.0,
                "latest_seen_at": None,
            },
        )
        current_cost = float(snap.direct_cost_usd or 0.0)
        bucket["observed_total_cost_usd"] += current_cost
        captured = snap.captured_at
        if captured:
            if bucket["first_seen_at"] is None or captured < bucket["first_seen_at"]:
                bucket["first_seen_at"] = captured
            if bucket["last_seen_at"] is None or captured > bucket["last_seen_at"]:
                bucket["last_seen_at"] = captured
            if bucket["latest_seen_at"] is None or captured >= bucket["latest_seen_at"]:
                bucket["latest_seen_at"] = captured
                bucket["latest_monthly_cost_usd"] = current_cost

    imported_cost_by_account: Dict[str, float] = {}
    imported_first_last_by_account: Dict[str, Dict[str, Optional[datetime]]] = {}
    imported_service_buckets: Dict[tuple[str, str, str], Dict[str, Any]] = {}

    for row in imported_rows:
        provider = str(getattr(row, "provider", "") or "").strip().lower()
        if provider_filter and provider != provider_filter:
            continue
        if provider not in providers:
            continue
        cost = float(getattr(row, "cost_usd", 0.0) or 0.0)
        account_identifier = str(getattr(row, "account_identifier", "") or "").strip()
        service_name = str(getattr(row, "service_name", "") or "").strip() or "imported-service"
        region = str(getattr(row, "region", "") or "").strip() or "global"
        tags = _safe_json_load(getattr(row, "tags_json", None) or "{}", {})
        owner_guess = _best_effort_owner_from_payload(tags)

        first_seen = getattr(row, "period_start", None) or getattr(row, "created_at", None)
        last_seen = getattr(row, "period_end", None) or getattr(row, "created_at", None)

        if account_identifier:
            imported_cost_by_account[account_identifier] = imported_cost_by_account.get(account_identifier, 0.0) + cost
            bucket = imported_first_last_by_account.setdefault(
                account_identifier,
                {"first_seen_at": None, "last_seen_at": None},
            )
            if first_seen and (bucket["first_seen_at"] is None or first_seen < bucket["first_seen_at"]):
                bucket["first_seen_at"] = first_seen
            if last_seen and (bucket["last_seen_at"] is None or last_seen > bucket["last_seen_at"]):
                bucket["last_seen_at"] = last_seen

        service_key = (provider, service_name, account_identifier)
        service_bucket = imported_service_buckets.setdefault(
            service_key,
            {
                "provider": provider,
                "resource_type": "imported-service",
                "resource_id": account_identifier or f"{provider}:{service_name}",
                "resource_name": service_name,
                "region": region,
                "owner_or_creator": owner_guess,
                "created_at": first_seen,
                "first_seen_at": first_seen,
                "last_seen_at": last_seen,
                "observed_total_cost_usd": 0.0,
                "latest_monthly_cost_usd": 0.0,
                "source": "csv_import",
            },
        )
        service_bucket["observed_total_cost_usd"] += cost
        service_bucket["latest_monthly_cost_usd"] += cost
        if owner_guess and not service_bucket.get("owner_or_creator"):
            service_bucket["owner_or_creator"] = owner_guess
        if first_seen and (service_bucket.get("first_seen_at") is None or first_seen < service_bucket["first_seen_at"]):
            service_bucket["first_seen_at"] = first_seen
        if last_seen and (service_bucket.get("last_seen_at") is None or last_seen > service_bucket["last_seen_at"]):
            service_bucket["last_seen_at"] = last_seen
        if first_seen and (service_bucket.get("created_at") is None or first_seen < service_bucket["created_at"]):
            service_bucket["created_at"] = first_seen

    candidates: List[Dict[str, Any]] = []
    for account in accounts:
        provider = str(account.provider or "").strip().lower()
        if provider_filter and provider != provider_filter:
            continue
        metadata = _safe_json_load(account.metadata_json or "{}", {})
        snapshot_stats = snapshots_by_account.get(account.id, {})
        imported_cost = imported_cost_by_account.get(account.account_identifier, 0.0)
        imported_seen = imported_first_last_by_account.get(account.account_identifier, {})
        first_seen = snapshot_stats.get("first_seen_at") or imported_seen.get("first_seen_at") or account.created_at
        last_seen = snapshot_stats.get("last_seen_at") or imported_seen.get("last_seen_at") or account.updated_at
        owner_guess = _best_effort_owner_from_payload(metadata)
        candidate = {
            "provider": provider,
            "resource_type": account.account_type or "cloud-account",
            "resource_id": account.account_identifier,
            "resource_name": account.account_name,
            "region": account.native_region or "global",
            "owner_or_creator": owner_guess,
            "created_at": account.created_at,
            "first_seen_at": first_seen,
            "last_seen_at": last_seen,
            "observed_total_cost_usd": float(snapshot_stats.get("observed_total_cost_usd", 0.0)) + float(imported_cost),
            "latest_monthly_cost_usd": float(snapshot_stats.get("latest_monthly_cost_usd", 0.0)),
            "source": "provider_accounts",
        }
        candidates.append(candidate)

    candidates.extend(imported_service_buckets.values())

    def _score(item: Dict[str, Any]) -> float:
        if not token_set:
            return 0.0
        haystack = " ".join(
            [
                str(item.get("resource_id") or ""),
                str(item.get("resource_name") or ""),
                str(item.get("resource_type") or ""),
                str(item.get("provider") or ""),
                str(item.get("region") or ""),
            ]
        ).lower()
        return float(sum(1 for token in token_set if token in haystack))

    for item in candidates:
        item["match_score"] = _score(item)
        item["observed_total_cost_usd"] = round(float(item.get("observed_total_cost_usd", 0.0) or 0.0), 2)
        item["latest_monthly_cost_usd"] = round(float(item.get("latest_monthly_cost_usd", 0.0) or 0.0), 2)

    matched = [c for c in candidates if c.get("match_score", 0.0) > 0] if token_set else candidates
    if not matched:
        matched = candidates
    matched.sort(
        key=lambda item: (
            float(item.get("match_score", 0.0)),
            float(item.get("observed_total_cost_usd", 0.0)),
            float(item.get("latest_monthly_cost_usd", 0.0)),
        ),
        reverse=True,
    )

    top = matched[0] if matched else None
    alternatives = matched[1:4] if len(matched) > 1 else []

    def _serialize(item: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "provider": item.get("provider"),
            "resource_type": item.get("resource_type"),
            "resource_id": item.get("resource_id"),
            "resource_name": item.get("resource_name"),
            "region": item.get("region"),
            "owner_or_creator": item.get("owner_or_creator"),
            "created_at": item.get("created_at").isoformat() if isinstance(item.get("created_at"), datetime) else item.get("created_at"),
            "first_seen_at": item.get("first_seen_at").isoformat() if isinstance(item.get("first_seen_at"), datetime) else item.get("first_seen_at"),
            "last_seen_at": item.get("last_seen_at").isoformat() if isinstance(item.get("last_seen_at"), datetime) else item.get("last_seen_at"),
            "observed_total_cost_usd": round(float(item.get("observed_total_cost_usd", 0.0) or 0.0), 2),
            "latest_monthly_cost_usd": round(float(item.get("latest_monthly_cost_usd", 0.0) or 0.0), 2),
            "source": item.get("source"),
            "match_score": float(item.get("match_score", 0.0) or 0.0),
        }

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "query": query,
        "cloud_provider": cloud_provider,
        "matched_resource": _serialize(top) if top else None,
        "alternatives": [_serialize(item) for item in alternatives],
        "notes": [
            "Owner/creator is best-effort from provider metadata and imported tags.",
            "Observed total cost is measured from available snapshots/imported history, not absolute cloud lifetime billing.",
        ],
    }


@router.get("/analytics/anomaly-intelligence")
async def analytics_anomaly_intelligence(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Anomaly root-cause classification with investigation playbooks, escalation recommendations, and annualized risk."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_cost_anomaly_intelligence(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    alerts_payload = result.get("anomalies", [])
    narrative, prompt = genai_advisor.generate_alert_triage(
        alerts_payload,
        {
            "current_monthly_spend_usd": context["total_cost"],
        },
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/chargeback-summary")
async def analytics_chargeback_summary(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Team/cost-center chargeback/showback allocation with unallocated spend risk."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_chargeback_summary(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    narrative, prompt = genai_advisor.generate_chargeback_narrative(
        {
            "total_monthly_spend_usd": result.get("total_monthly_spend_usd", context["total_cost"]),
            "allocation_coverage_percent": result.get("allocation_coverage_percent", 0),
            "unallocated_usd": result.get("unallocated_usd", 0),
            "unallocated_percent": result.get("unallocated_percent", 0),
            "top_spenders": result.get("top_spenders", []),
            "team_count": result.get("team_count", 0),
            "model": result.get("model", "showback"),
        }
    )
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/provider-diagnostics", response_model=List[ProviderDiagnostic])
async def provider_diagnostics(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> List[ProviderDiagnostic]:
    """Return provider readiness checks without exposing secret values."""
    _ = current_user
    return _provider_diagnostics(customer_id=_customer_id_for_org(membership), db=db)


@router.get("/health")
async def health_check() -> Dict[str, Any]:
    return {
        "status": "healthy",
        "version": __version__,
        "timestamp": _utcnow().isoformat(),
    }


@router.get("/health/readiness")
async def readiness_check() -> Dict[str, Any]:
    """Deep readiness probe with database and provider runtime checks."""
    cfg = Config()
    now = _utcnow()
    checks: Dict[str, Any] = {}
    overall_status = "healthy"

    db_started_at = time.perf_counter()
    db_status = "healthy"
    db_error: Optional[str] = None
    organization_count = 0
    try:
        db = SessionLocal()
        try:
            organization_count = int(db.query(Organization).count())
            _ = db.query(User.id).limit(1).first()
        finally:
            db.close()
    except Exception as exc:
        db_status = "unhealthy"
        db_error = str(exc)
        overall_status = "unhealthy"
    checks["database"] = {
        "status": db_status,
        "latency_ms": round((time.perf_counter() - db_started_at) * 1000, 2),
        "organization_count": organization_count,
        "error": db_error,
    }

    diagnostics = _provider_diagnostics()
    configured_providers = sorted([item.provider for item in diagnostics if item.configured])
    missing_settings_by_provider = {
        item.provider: item.missing_settings
        for item in diagnostics
        if not item.configured
    }

    if cfg.require_live_provider_data and not configured_providers:
        providers_status = "unhealthy"
        overall_status = "unhealthy"
    elif configured_providers and len(configured_providers) == len(diagnostics):
        providers_status = "healthy"
    elif configured_providers:
        providers_status = "degraded"
        if overall_status == "healthy":
            overall_status = "degraded"
    else:
        providers_status = "degraded"
        if overall_status == "healthy":
            overall_status = "degraded"

    checks["providers"] = {
        "status": providers_status,
        "require_live_provider_data": bool(cfg.require_live_provider_data),
        "configured_count": len(configured_providers),
        "supported_count": len(diagnostics),
        "configured_providers": configured_providers,
        "missing_settings_by_provider": missing_settings_by_provider,
    }
    checks["runtime"] = {
        "auth_enabled": bool(cfg.auth_enabled),
        "scan_scheduler_enabled": bool(cfg.enable_scan_scheduler),
        "retention_enabled": bool(cfg.retention_enabled),
        "deployment_target": cfg.deployment_target,
        "oci_runtime_required": bool(cfg.oci_runtime_required),
        "running_on_oci": Config.is_running_on_oci(),
    }

    return {
        "status": overall_status,
        "version": __version__,
        "timestamp": now.isoformat(),
        "checks": checks,
    }


@router.get("/info")
async def api_info() -> Dict[str, Any]:
    cfg = Config()
    return {
        "name": "OptiOra API",
        "version": __version__,
        "description": "Cloud Cost Optimization Platform",
        "runtime": {
            "deployment_target": cfg.deployment_target,
            "oci_runtime_required": bool(cfg.oci_runtime_required),
            "on_premises_supported": False,
        },
        "supported_providers": list(SUPPORTED_CLOUD_PROVIDERS),
        "features": {
            "credential_management": True,
            "credential_validation": True,
            "scanning_permissions": True,
            "dashboard_endpoints": True,
            "finops_analytics": True,
            "forecasting": True,
            "forecast_what_if": True,
            "forecast_stress_test": True,
            "forecast_diagnostics": True,
            "forecast_model_diagnostics": True,
            "decision_intelligence": True,
            "cost_attribution": True,
            "commitment_optimization": True,
            "optimization_portfolio": True,
            "operating_review_pack": True,
            "maturity_assessment": True,
            "unit_economics": True,
            "anomaly_scoring": True,
            "cloud_waste_analysis": True,
            "efficiency_score": True,
            "commitment_gap_analysis": True,
            "genai_advisor": True,
            "hybrid_advisor": True,
            "genai_waste_insights": True,
            "genai_optimization_roadmap": True,
            "genai_executive_narrative": True,
            "genai_commitment_strategy": True,
            "genai_finops_operating_review": True,
            "genai_forecast_model_diagnostics": True,
            "genai_copilot_pack": True,
            "genai_rag_guidance": True,
            "finops_rag_retrieval": True,
            "finops_intelligence_endpoint": True,
            "genai_backend_narration": genai_advisor._is_configured(),
            "provider_diagnostics": True,
            "provider_native_recommendations": True,
            "audit_logging": True,
            "alert_lifecycle": True,
            "alert_ops_policy": True,
            "routing_policy_simulator": True,
            "operations_data_freshness": True,
            "scheduler_policy_overrides": True,
            "budget_alerts": True,
            "csv_exports": True,
            "csv_imports": True,
            "csv_import_templates": True,
            "excel_exports": True,
            "executive_reports": True,
            "provider_hierarchy": True,
            "account_region_breakdown": True,
            "focus_export": True,
            "unit_economics_cockpit": True,
            "scorecards": True,
            "resource_inventory": True,
            "service_hotspots": True,
            "kubernetes_cost_allocation": True,
            "virtual_tagging": True,
            "rightsizing_resource_level": True,
            "admin_diagnostics": True,
            "readiness_endpoint": True,
            "request_tracing": True,
            "pagination": True,
        },
    }


@router.get("/admin/diagnostics")
async def admin_diagnostics(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Consolidated operational snapshot for admin troubleshooting views."""
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        permission = ScanningManager(db).get_permission_status(customer_id)
    finally:
        db.close()

    scheduler = await get_scheduler_status(current_user=current_user, membership=membership)
    freshness = await get_data_freshness(current_user=current_user, membership=membership)
    destinations = await list_notification_destinations(
        current_user=current_user,
        membership=membership,
    )
    return {
        "generated_at": _utcnow().isoformat(),
        "organization_id": organization_id,
        "api_health": await health_check(),
        "api_info": await api_info(),
        "provider_diagnostics": _provider_diagnostics(),
        "scanning_permission": permission,
        "scheduler": scheduler.model_dump(),
        "data_freshness": freshness.model_dump(),
        "notification_destinations": destinations.model_dump(),
    }


async def _run_cost_analysis(
    scan_id: str,
    customer_id: str,
    providers: List[str],
    target_accounts: Optional[List[str]] = None,
    raise_on_error: bool = False,
) -> None:
    """
    Background scan: fetch live cost data per provider, persist CostSnapshot
    rows for historical trend analysis, then mark the scan run complete.

    When ``target_accounts`` is provided, only account_breakdown rows whose
    account_id matches one of the supplied identifiers are persisted as
    ProviderAccountSnapshot / CostAllocationSnapshot records.  The overall
    cost aggregation still uses all returned data.
    """
    db = SessionLocal()
    scanning_manager = ScanningManager(db)
    try:
        total_resources = 0
        anomalies_found = 0
        savings_identified = 0.0
        aggregated_cost_usd = 0.0
        now = _utcnow()
        successful_providers: List[str] = []
        provider_errors: Dict[str, str] = {}
        require_live_provider_data = Config().require_live_provider_data

        for provider in providers:
            try:
                summary = await _cost_summary_for_provider(provider, "month", customer_id=customer_id)
            except TypeError as exc:
                if "customer_id" not in str(exc) and "keyword" not in str(exc):
                    raise
                summary = await _cost_summary_for_provider(provider, "month")
            if "error" in summary:
                provider_errors[provider] = str(summary.get("error") or "Unknown provider error")
                _mark_provider_credentials_unreachable(
                    db,
                    customer_id,
                    provider,
                    provider_errors[provider],
                )
                continue
            successful_providers.append(provider)

            total_cost = float(summary.get("total_cost_usd", 0) or 0)
            aggregated_cost_usd += total_cost
            total_resources += 100
            provider_savings = total_cost * 0.08
            savings_identified += provider_savings

            # Fetch anomalies and recommendations for this provider to embed in snapshot.
            anomaly_raw = await anomalies.detect_anomalies({"cloud_provider": provider})
            anomaly_data = _safe_json_load(anomaly_raw, {})
            provider_anomalies = int(anomaly_data.get("anomalies_found", 0) or 0)
            anomalies_found += provider_anomalies

            rec_raw = await recommendations.get_recommendations({
                "cloud_provider": provider,
                "current_monthly_spend": total_cost,
            })

            # Derive period bounds from summary if available.
            try:
                period_start = datetime.fromisoformat(summary["start_date"]) if "start_date" in summary else None
                period_end = datetime.fromisoformat(summary["end_date"]) if "end_date" in summary else None
            except (ValueError, TypeError):
                period_start = period_end = None

            snapshot = CostSnapshot(
                scan_id=scan_id,
                customer_id=customer_id,
                provider=provider,
                period_start=period_start,
                period_end=period_end,
                total_cost_usd=total_cost,
                savings_identified_usd=provider_savings,
                anomalies_count=provider_anomalies,
                top_services_json=json.dumps(summary.get("top_services", [])),
                anomalies_json=anomaly_raw,
                recommendations_json=rec_raw,
                captured_at=now,
            )
            db.add(snapshot)

            # Persist provider-account hierarchy snapshots for rollup views.
            organization_id = _organization_id_from_customer_id(customer_id)
            account_rows = summary.get("account_breakdown") or []
            if organization_id is not None:
                # Guarantee at least one rollup node per provider.
                if not account_rows:
                    account_rows = [
                        {
                            "scope_type": "provider",
                            "scope_id": provider,
                            "total_cost_usd": total_cost,
                        }
                    ]

                # When target_accounts is set, restrict which hierarchy nodes are
                # persisted.  Error rows and the catch-all provider node are always
                # kept so the scan record remains coherent.
                if target_accounts:
                    target_set = {str(a).strip() for a in target_accounts if a}
                    account_rows = [
                        row for row in account_rows
                        if (
                            "error" in row
                            or str(row.get("scope_type") or "") == "provider"
                            or str(
                                row.get("account_id")
                                or row.get("scope_id")
                                or row.get("role_arn")
                                or ""
                            ).strip() in target_set
                        )
                    ]

                provider_accounts_by_identifier: Dict[str, ProviderAccount] = {}
                parent_hints_by_identifier: Dict[str, Dict[str, Optional[str]]] = {}

                for account_row in account_rows:
                    account_identifier = str(
                        account_row.get("account_id")
                        or account_row.get("scope_id")
                        or account_row.get("role_arn")
                        or provider
                    )
                    account_name = str(
                        account_row.get("account_name")
                        or account_row.get("scope_name")
                        or account_row.get("scope_id")
                        or account_identifier
                    )
                    account_type = str(
                        account_row.get("scope_type")
                        or account_row.get("account_type")
                        or "account"
                    ).strip().lower().replace(" ", "_")
                    parent_identifier = str(
                        account_row.get("parent_scope_id")
                        or account_row.get("parent_account_identifier")
                        or ""
                    ).strip()
                    parent_type = str(account_row.get("parent_scope_type") or "").strip().lower().replace(" ", "_")
                    account_total_cost = float(account_row.get("total_cost_usd") or 0.0)
                    if account_total_cost == 0.0 and len(account_rows) == 1:
                        account_total_cost = total_cost

                    provider_account = (
                        db.query(ProviderAccount)
                        .filter(
                            ProviderAccount.customer_id == customer_id,
                            ProviderAccount.provider == provider,
                            ProviderAccount.account_identifier == account_identifier,
                        )
                        .first()
                    )
                    if provider_account is None:
                        provider_account = ProviderAccount(
                            organization_id=organization_id,
                            customer_id=customer_id,
                            provider=provider,
                            account_identifier=account_identifier,
                            account_name=account_name,
                            account_type=account_type,
                            native_region=(summary.get("region_breakdown") or [{}])[0].get("region"),
                            metadata_json=json.dumps(account_row),
                            is_active=True,
                        )
                        db.add(provider_account)
                        db.flush()
                    else:
                        provider_account.account_name = account_name
                        provider_account.account_type = account_type
                        provider_account.metadata_json = json.dumps(account_row)
                        provider_account.updated_at = now

                    provider_accounts_by_identifier[account_identifier] = provider_account
                    parent_hints_by_identifier[account_identifier] = {
                        "parent_identifier": parent_identifier or None,
                        "parent_type": parent_type or None,
                    }

                    existing_provider_snapshot = (
                        db.query(ProviderAccountSnapshot)
                        .filter(
                            ProviderAccountSnapshot.scan_id == scan_id,
                            ProviderAccountSnapshot.provider_account_id == provider_account.id,
                        )
                        .first()
                    )
                    if existing_provider_snapshot is None:
                        db.add(
                            ProviderAccountSnapshot(
                                organization_id=organization_id,
                                customer_id=customer_id,
                                scan_id=scan_id,
                                provider_account_id=provider_account.id,
                                direct_cost_usd=account_total_cost,
                                savings_identified_usd=provider_savings,
                                anomalies_count=provider_anomalies,
                                service_count=len(summary.get("top_services", [])),
                                captured_at=now,
                            )
                        )

                    # Persist per-region cost breakdown for this account node.
                    region_breakdown = (
                        account_row.get("region_breakdown")
                        or (summary.get("region_breakdown") if len(account_rows) == 1 else [])
                        or []
                    )
                    for region_row in region_breakdown:
                        region_name = str(region_row.get("region") or "global")
                        region_cost = float(region_row.get("cost_usd") or 0.0)
                        if region_cost <= 0.0:
                            continue
                        existing_alloc = (
                            db.query(CostAllocationSnapshot)
                            .filter(
                                CostAllocationSnapshot.scan_id == scan_id,
                                CostAllocationSnapshot.provider_account_id == provider_account.id,
                                CostAllocationSnapshot.region == region_name,
                            )
                            .first()
                        )
                        if existing_alloc is None:
                            db.add(
                                CostAllocationSnapshot(
                                    organization_id=organization_id,
                                    customer_id=customer_id,
                                    scan_id=scan_id,
                                    provider_account_id=provider_account.id,
                                    provider=provider,
                                    region=region_name,
                                    cost_usd=region_cost,
                                    captured_at=now,
                                )
                            )

                for child_identifier, hint in parent_hints_by_identifier.items():
                    parent_identifier = hint.get("parent_identifier")
                    if not parent_identifier or parent_identifier == child_identifier:
                        continue

                    child_account = provider_accounts_by_identifier.get(child_identifier)
                    if child_account is None:
                        continue

                    parent_account = provider_accounts_by_identifier.get(parent_identifier)
                    if parent_account is None:
                        parent_account = (
                            db.query(ProviderAccount)
                            .filter(
                                ProviderAccount.customer_id == customer_id,
                                ProviderAccount.provider == provider,
                                ProviderAccount.account_identifier == parent_identifier,
                            )
                            .first()
                        )
                        if parent_account is None:
                            inferred_parent_type = str(hint.get("parent_type") or "group").strip().lower().replace(" ", "_")
                            parent_account = ProviderAccount(
                                organization_id=organization_id,
                                customer_id=customer_id,
                                provider=provider,
                                account_identifier=parent_identifier,
                                account_name=parent_identifier,
                                account_type=inferred_parent_type or "group",
                                native_region=(summary.get("region_breakdown") or [{}])[0].get("region"),
                                metadata_json=json.dumps(
                                    {
                                        "inferred": True,
                                        "source": "parent_scope",
                                        "scope_id": parent_identifier,
                                        "scope_type": inferred_parent_type or "group",
                                    }
                                ),
                                is_active=True,
                            )
                            db.add(parent_account)
                            db.flush()
                        provider_accounts_by_identifier[parent_identifier] = parent_account

                    existing_link = (
                        db.query(ProviderAccountLink)
                        .filter(
                            ProviderAccountLink.organization_id == organization_id,
                            ProviderAccountLink.child_account_id == child_account.id,
                        )
                        .first()
                    )
                    if existing_link is None:
                        db.add(
                            ProviderAccountLink(
                                organization_id=organization_id,
                                parent_account_id=parent_account.id,
                                child_account_id=child_account.id,
                                relationship_type="contains",
                            )
                        )
                    else:
                        existing_link.parent_account_id = parent_account.id
                        existing_link.relationship_type = "contains"

        if provider_errors:
            db.commit()

        if require_live_provider_data and not successful_providers:
            error_summary = ", ".join(
                f"{provider}: {error}" for provider, error in sorted(provider_errors.items())
            ) or "No provider returned live data."
            raise RuntimeError(
                "Live provider data is required, but this scan could not fetch any provider data. "
                f"Details: {error_summary}"
            )

        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        organization_id = _organization_id_from_customer_id(customer_id)
        if organization_id is not None:
            evaluate_budget_alert(
                db=db,
                organization_id=organization_id,
                customer_id=customer_id,
                scan_id=scan_id,
                total_cost_usd=aggregated_cost_usd,
                permission=permission,
            )

        db.commit()

        scanning_manager.complete_scan_run(
            scan_id=scan_id,
            progress=100,
            total_resources=total_resources,
            anomalies_found=anomalies_found,
            savings_identified=savings_identified,
        )
    except Exception as exc:
        logger.exception("Background scan failed")
        scanning_manager.fail_scan_run(scan_id, str(exc))
        if raise_on_error:
            raise
    finally:
        db.close()


def _scan_interval_seconds(scan_frequency: str) -> int:
    return scan_interval_seconds(scan_frequency)


def _coerce_aws_anomaly_impact_usd(payload: Dict[str, Any]) -> float:
    return coerce_aws_anomaly_impact_usd(payload)


def _aws_anomaly_severity(impact_usd: float, source_severity: Optional[str]) -> str:
    return aws_anomaly_severity(impact_usd, source_severity)


def _compute_next_run(now: datetime, scan_frequency: str, anchor: datetime) -> datetime:
    return compute_next_run(now, scan_frequency, anchor)


@router.get("/scanning/scheduler/status", response_model=SchedulerStatusResponse)
async def get_scheduler_status(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> SchedulerStatusResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()
    alert_id: Optional[int] = None
    db = SessionLocal()
    try:
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        runs = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .limit(50)
            .all()
        )
        audit_rows = (
            db.query(AuditLog)
            .filter(
                AuditLog.organization_id == organization_id,
                AuditLog.action.in_(["scan.schedule.triggered"]),
            )
            .order_by(AuditLog.created_at.desc())
            .limit(8)
            .all()
        )
    finally:
        db.close()

    snapshot = scheduler_runtime_snapshot(
        now=now,
        permission=permission,
        runs=runs,
        audit_rows=audit_rows,
        initialized_state=ScanningState.INITIALIZED.value,
        approved_state=ScanningState.APPROVED.value,
        running_state=ScanningState.RUNNING.value,
        completed_state=ScanningState.COMPLETED.value,
        failed_state=ScanningState.FAILED.value,
        safe_json_load=_safe_json_load,
    )

    return SchedulerStatusResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        scheduler_enabled=Config().enable_scan_scheduler,
        scheduler_running=_scheduler_running,
        permission_state=snapshot["permission_state"],
        scan_frequency=snapshot["scan_frequency"],
        effective_scan_frequency=snapshot["effective_scan_frequency"],
        scheduler_override_enabled=snapshot["scheduler_override_enabled"],
        next_run_at=snapshot["next_run_at"].isoformat() if snapshot["next_run_at"] else None,
        next_run_eta_seconds=snapshot["next_run_eta_seconds"],
        last_success_at=snapshot["last_success"].isoformat() if snapshot["last_success"] else None,
        last_failure_at=snapshot["last_failure"].isoformat() if snapshot["last_failure"] else None,
        retry_max_attempts=snapshot["retry_max_attempts"],
        retry_backoff_seconds=snapshot["retry_backoff_seconds"],
        overdue_alert_hours=snapshot["overdue_alert_hours"],
        overdue=snapshot["overdue"],
        counters=SchedulerCounters(
            total=snapshot["counters"]["total"],
            success=snapshot["counters"]["success"],
            failure=snapshot["counters"]["failure"],
        ),
        timeline=[SchedulerTimelineItem(**item) for item in snapshot["timeline"]],
    )


def _freshness_state_for_age(age_seconds: Optional[int], stale_after_seconds: int) -> str:
    if age_seconds is None:
        return "unknown"
    return "stale" if age_seconds > stale_after_seconds else "fresh"


@router.get("/operations/data-freshness", response_model=DataFreshnessResponse)
async def get_data_freshness(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> DataFreshnessResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()

    db = SessionLocal()
    try:
        cost_rows = (
            db.query(CostSnapshot)
            .filter(CostSnapshot.customer_id == customer_id)
            .order_by(CostSnapshot.captured_at.desc())
            .limit(500)
            .all()
        )
        imported_rows = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.customer_id == customer_id)
            .order_by(ImportedCostRecord.created_at.desc())
            .limit(500)
            .all()
        )
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        runs = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .limit(20)
            .all()
        )
        external_alerts = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.customer_id == customer_id,
                AlertEvent.alert_type.in_(
                    [
                        "external.aws.cost_anomaly",
                        "external.gcp.budget_pubsub",
                    ]
                ),
            )
            .order_by(AlertEvent.created_at.desc())
            .limit(200)
            .all()
        )
    finally:
        db.close()

    latest_by_provider: Dict[str, datetime] = {}
    for row in cost_rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        if provider not in latest_by_provider:
            latest_by_provider[provider] = row.captured_at

    for row in imported_rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        current = latest_by_provider.get(provider)
        if current is None or row.created_at > current:
            latest_by_provider[provider] = row.created_at

    providers: List[DataFreshnessProviderItem] = []
    for provider in ["aws", "azure", "gcp", "oci"]:
        last = latest_by_provider.get(provider)
        age_seconds = int((now - last).total_seconds()) if last else None
        providers.append(
            DataFreshnessProviderItem(
                provider=provider,
                last_ingested_at=last.isoformat() if last else None,
                age_seconds=age_seconds,
                status=_freshness_state_for_age(age_seconds, stale_after_seconds=48 * 3600),
            )
        )

    connector_latest: Dict[str, Optional[datetime]] = {
        "aws_cost_anomaly": None,
        "gcp_budget_pubsub": None,
    }
    for row in external_alerts:
        if row.alert_type == "external.aws.cost_anomaly" and connector_latest["aws_cost_anomaly"] is None:
            connector_latest["aws_cost_anomaly"] = row.created_at
        if row.alert_type == "external.gcp.budget_pubsub" and connector_latest["gcp_budget_pubsub"] is None:
            connector_latest["gcp_budget_pubsub"] = row.created_at

    connectors: List[DataFreshnessConnectorItem] = []
    for connector, last in connector_latest.items():
        age_seconds = int((now - last).total_seconds()) if last else None
        connectors.append(
            DataFreshnessConnectorItem(
                connector=connector,
                last_event_at=last.isoformat() if last else None,
                age_seconds=age_seconds,
                status=_freshness_state_for_age(age_seconds, stale_after_seconds=72 * 3600),
            )
        )

    scheduler_lag_seconds: Optional[int] = None
    scheduler_status: Literal["healthy", "lagging", "unknown"] = "unknown"
    if permission:
        effective_frequency = str(permission.scan_frequency or "daily").strip().lower()
        if bool(permission.scheduler_override_enabled):
            override_frequency = str(permission.scheduler_override_frequency or "").strip().lower()
            if override_frequency in {"hourly", "daily", "weekly"}:
                effective_frequency = override_frequency
        interval_seconds = _scan_interval_seconds(effective_frequency)
        last_success = next(
            (
                row.completed_at or row.started_at
                for row in runs
                if row.state == ScanningState.COMPLETED.value
            ),
            None,
        )
        if last_success is not None:
            lag = int((now - last_success).total_seconds()) - interval_seconds
            scheduler_lag_seconds = max(lag, 0)
            scheduler_status = "lagging" if scheduler_lag_seconds > max(interval_seconds // 4, 900) else "healthy"

    return DataFreshnessResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        generated_at=now.isoformat(),
        providers=providers,
        connectors=connectors,
        scheduler_lag_seconds=scheduler_lag_seconds,
        scheduler_status=scheduler_status,
    )


@router.post("/anomalies/external/aws")
async def ingest_external_aws_anomalies(
    payload: ExternalAWSAnomalyIngestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    _require_management_role(membership, "external anomaly ingestion")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    if not payload.events:
        return {"status": "ok", "ingested": 0, "alert_ids": [], "duplicates": 0}

    now = _utcnow()
    db = SessionLocal()
    try:
        alert_ids: List[int] = []
        duplicate_count = 0
        suppressed_count = 0
        for event in payload.events:
            anomaly = derive_aws_anomaly_alert(event, now)
            detail_payload = anomaly["detail_payload"]
            anomaly_id = anomaly["anomaly_id"]
            
            # Check for idempotency using source_event_id
            if anomaly_id:
                duplicate = (
                    db.query(AuditLog)
                    .filter(
                        AuditLog.organization_id == organization_id,
                        AuditLog.action == "alert.external.ingest",
                        AuditLog.entity_type == "aws_cost_anomaly_event",
                        AuditLog.entity_id == anomaly_id,
                    )
                    .first()
                )
                if duplicate is not None:
                    duplicate_count += 1
                    continue
            
            impact_usd = float(anomaly["impact_usd"])
            severity = str(anomaly["severity"])
            monitor_name = str(anomaly["monitor_name"])
            title = str(anomaly["title"])
            message = str(anomaly["message"])
            allowed, suppressed_reason, _ = _alert_passes_policy(
                db=db,
                organization_id=organization_id,
                customer_id=customer_id,
                alert_type="external.aws.cost_anomaly",
                severity=severity,
                title=title[:255],
                message=message,
                now_utc=now,
            )
            if not allowed:
                suppressed_count += 1
                db.add(
                    AuditLog(
                        organization_id=organization_id,
                        actor_user_id=current_user.id,
                        action="alert.external.suppressed",
                        entity_type="aws_cost_anomaly_event",
                        entity_id=anomaly_id,
                        metadata_json=json.dumps(
                            {
                                "reason": suppressed_reason,
                                "severity": severity,
                                "monitor_name": monitor_name,
                            }
                        ),
                        created_at=now,
                    )
                )
                continue
            row = AlertEvent(
                organization_id=organization_id,
                customer_id=customer_id,
                scan_id=None,
                alert_type="external.aws.cost_anomaly",
                severity=severity,
                title=title[:255],
                message=message,
                delivered_channels_json=json.dumps(["aws-cost-anomaly-detection"]),
                created_at=now,
            )
            db.add(row)
            db.flush()
            alert_ids.append(row.id)
            
            # Log individual event for replay tracking
            db.add(
                AuditLog(
                    organization_id=organization_id,
                    actor_user_id=current_user.id,
                    action="alert.external.ingest",
                    entity_type="aws_cost_anomaly_event",
                    entity_id=anomaly_id,
                    metadata_json=json.dumps({
                        "alert_id": row.id,
                        "monitor_name": monitor_name,
                        "impact_usd": impact_usd,
                        "severity": severity,
                    }),
                    created_at=now,
                )
            )

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.external.ingest",
                entity_type="aws_cost_anomaly_batch",
                entity_id="aws-cost-anomaly-detection",
                metadata_json=json.dumps(
                    {"count": len(alert_ids), "duplicates": duplicate_count, "suppressed": suppressed_count}
                ),
                created_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    return {
        "status": "ok",
        "ingested": len(alert_ids),
        "alert_ids": alert_ids,
        "duplicates": duplicate_count,
        "suppressed": suppressed_count,
    }


def _decode_gcp_pubsub_payload(raw_message: Dict[str, Any]) -> Dict[str, Any]:
    data = raw_message.get("data")
    if isinstance(data, str) and data.strip():
        try:
            decoded = base64.b64decode(data.encode("utf-8"))
            parsed = json.loads(decoded.decode("utf-8"))
            if isinstance(parsed, dict):
                return parsed
        except (binascii.Error, UnicodeDecodeError, json.JSONDecodeError):
            return {}
    return raw_message if isinstance(raw_message, dict) else {}


def _coerce_budget_float(source: Dict[str, Any], keys: List[str]) -> float:
    for key in keys:
        value = source.get(key)
        if value is None:
            continue
        if isinstance(value, dict):
            for nested_key in ("amount", "value", "doubleValue"):
                nested = value.get(nested_key)
                if nested is not None:
                    value = nested
                    break
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return 0.0


@router.post("/anomalies/external/gcp/pubsub")
async def ingest_external_gcp_budget_pubsub(
    payload: ExternalGCPPubSubIngestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    ingest_token: Optional[str] = Header(default=None, alias="X-OptiOra-Ingest-Token"),
) -> Dict[str, Any]:
    _require_management_role(membership, "external anomaly ingestion")
    config = Config()
    expected_token = (config.gcp_pubsub_ingest_token or "").strip()
    if expected_token and ingest_token != expected_token:
        raise HTTPException(status_code=403, detail="Invalid ingest token")

    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    envelope_message = payload.message if isinstance(payload.message, dict) else {}
    message_payload = _decode_gcp_pubsub_payload(envelope_message)
    message_id = str(
        envelope_message.get("messageId")
        or envelope_message.get("message_id")
        or message_payload.get("messageId")
        or message_payload.get("message_id")
        or ""
    ).strip()

    budget_name = str(
        message_payload.get("budgetDisplayName")
        or message_payload.get("budgetName")
        or message_payload.get("budget_name")
        or "GCP Budget"
    )
    cost_amount = _coerce_budget_float(
        message_payload,
        ["costAmount", "cost_amount", "actualAmount", "amount"],
    )
    budget_amount = _coerce_budget_float(
        message_payload,
        ["budgetAmount", "budget_amount", "thresholdAmount", "budget"],
    )
    usage_ratio = (cost_amount / budget_amount) if budget_amount > 0 else 0.0
    severity = "critical" if usage_ratio >= 1.0 else "warning"

    now = _utcnow()
    db = SessionLocal()
    try:
        if message_id:
            duplicate = (
                db.query(AuditLog)
                .filter(
                    AuditLog.organization_id == organization_id,
                    AuditLog.action == "alert.external.ingest",
                    AuditLog.entity_type == "gcp_budget_event",
                    AuditLog.entity_id == message_id,
                )
                .first()
            )
            if duplicate is not None:
                return {"status": "ok", "ingested": 0, "duplicate": True, "message_id": message_id}

        title = f"GCP budget alert ({budget_name})"
        message = (
            f"GCP budget usage is ${cost_amount:.2f} against budget ${budget_amount:.2f}. "
            f"Utilization {usage_ratio * 100:.1f}%"
        )
        allowed, suppressed_reason, _ = _alert_passes_policy(
            db=db,
            organization_id=organization_id,
            customer_id=customer_id,
            alert_type="external.gcp.budget_pubsub",
            severity=severity,
            title=title[:255],
            message=message,
            now_utc=now,
        )
        if not allowed:
            db.add(
                AuditLog(
                    organization_id=organization_id,
                    actor_user_id=current_user.id,
                    action="alert.external.suppressed",
                    entity_type="gcp_budget_event",
                    entity_id=message_id or budget_name,
                    metadata_json=json.dumps(
                        {
                            "reason": suppressed_reason,
                            "severity": severity,
                            "budget_name": budget_name,
                            "message_id": message_id or None,
                        }
                    ),
                    created_at=now,
                )
            )
            db.commit()
            return {
                "status": "ok",
                "ingested": 0,
                "suppressed": True,
                "reason": suppressed_reason,
                "message_id": message_id or None,
            }

        alert = AlertEvent(
            organization_id=organization_id,
            customer_id=customer_id,
            scan_id=None,
            alert_type="external.gcp.budget_pubsub",
            severity=severity,
            title=title[:255],
            message=message,
            delivered_channels_json=json.dumps(["gcp-budget-pubsub"]),
            created_at=now,
        )
        db.add(alert)
        db.flush()
        alert_id = int(alert.id)

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.external.ingest",
                entity_type="gcp_budget_event",
                entity_id=message_id or str(alert.id),
                metadata_json=json.dumps(
                    {
                        "alert_id": alert.id,
                        "budget_name": budget_name,
                        "cost_amount": cost_amount,
                        "budget_amount": budget_amount,
                        "message_id": message_id or None,
                        "subscription": payload.subscription,
                    }
                ),
                created_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    return {"status": "ok", "ingested": 1, "alert_id": alert_id, "message_id": message_id or None}


@router.post("/anomalies/external/aws/replay")
async def replay_external_aws_anomalies(
    request: ExternalAWSReplayRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Replay failed AWS anomaly events for retry."""
    _require_management_role(membership, "external anomaly replay")
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        # Query audit logs for AWS cost anomaly events
        query = db.query(AuditLog).filter(
            AuditLog.organization_id == organization_id,
            AuditLog.action == "alert.external.ingest",
            AuditLog.entity_type == "aws_cost_anomaly_event",
        )
        
        # Filter by event_ids if provided
        if request.event_ids:
            query = query.filter(AuditLog.entity_id.in_(request.event_ids))
        
        # Filter by date range if provided
        if request.days_back:
            cutoff_date = _utcnow() - timedelta(days=request.days_back)
            query = query.filter(AuditLog.created_at >= cutoff_date)
        
        # Apply limit
        events = query.order_by(AuditLog.created_at.desc()).limit(request.max_results).all()
        
        if not events:
            return {
                "status": "ok",
                "replayed": 0,
                "alert_ids": [],
                "message": "No events found matching criteria"
            }
        
        # Extract alert IDs and associated metadata from audit logs
        alert_ids: List[int] = []
        for event in events:
            try:
                metadata = json.loads(event.metadata_json)
                if "alert_id" in metadata:
                    alert_ids.append(metadata["alert_id"])
            except (json.JSONDecodeError, KeyError):
                pass
        
        return {
            "status": "ok",
            "replayed": len(events),
            "alert_ids": alert_ids,
            "count": len(events),
            "date_range_days": request.days_back or "all",
            "message": f"Found {len(events)} events ready for retry"
        }
    finally:
        db.close()


# ── Connector Management endpoints ───────────────────────────────────────────

@router.get("/connectors", response_model=ConnectorListResponse)
async def list_connectors(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ConnectorListResponse:
    """List all supported connector types."""
    _require_management_role(membership, "connector listing")
    supported = ConnectorManager.list_supported_connectors()
    return ConnectorListResponse(
        supported_connectors=[c.value for c in supported],
        description="Supported cloud cost and resource connectors for OptiOra",
    )


@router.post("/connectors/test", response_model=ConnectorStatusResponse)
async def test_connector(
    request: ConnectorTestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ConnectorStatusResponse:
    """Test connector configuration and credentials."""
    _require_management_role(membership, "connector testing")
    
    try:
        connector_type = ConnectorType(request.connector_type)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown connector type: {request.connector_type}",
        )
    
    try:
        connector = ConnectorManager.get_connector(connector_type, request.config)
        is_valid = await connector.validate_credentials()
        status = "valid" if is_valid else "invalid"
        
        return ConnectorStatusResponse(
            connector_type=request.connector_type,
            status=status,
            message=f"Connector validation {status}." if is_valid else "Connector credentials are invalid.",
        )
    except Exception as e:
        logger.error(f"Connector test failed for {request.connector_type}: {e}")
        return ConnectorStatusResponse(
            connector_type=request.connector_type,
            status="error",
            message=f"Test failed: {str(e)[:200]}",
        )


@router.get("/connectors/status", response_model=List[ConnectorStatusResponse])
async def get_connectors_status(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ConnectorStatusResponse]:
    """Get status of all configured connectors."""
    _require_management_role(membership, "connector status check")
    
    # This endpoint returns status for all available connector types
    # In a full implementation, this would query stored connector configs from the database
    statuses: List[ConnectorStatusResponse] = []
    for connector_type in ConnectorManager.list_supported_connectors():
        statuses.append(
            ConnectorStatusResponse(
                connector_type=connector_type.value,
                status="unknown",
                message="Connector not configured. Configure via API or UI to enable.",
            )
        )
    return statuses


# ── Business Mapping & Chargeback endpoints ──────────────────────────────────

@router.get("/business-mapping/rules", response_model=BusinessMappingRuleListResponse)
async def list_mapping_rules(
    dimension: Optional[str] = None,
    active_only: bool = True,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleListResponse:
    """List business mapping rules for the current organization."""
    query = db.query(BusinessMappingRule).filter(
        BusinessMappingRule.organization_id == membership.organization_id
    )
    if active_only:
        query = query.filter(BusinessMappingRule.is_active == True)  # noqa: E712
    if dimension:
        if dimension not in VALID_DIMENSIONS:
            raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
        query = query.filter(BusinessMappingRule.dimension == dimension)
    rules = query.order_by(BusinessMappingRule.priority, BusinessMappingRule.id).all()
    return BusinessMappingRuleListResponse(
        organization_id=membership.organization_id,
        rules=[
            BusinessMappingRuleResponse(
                id=r.id,
                organization_id=r.organization_id,
                customer_id=r.customer_id,
                tag_key=r.tag_key,
                tag_value=r.tag_value,
                dimension=r.dimension,
                mapped_value=r.mapped_value,
                priority=r.priority,
                is_active=r.is_active,
                created_at=r.created_at.isoformat(),
                updated_at=r.updated_at.isoformat() if r.updated_at else None,
            )
            for r in rules
        ],
        total=len(rules),
    )


@router.post("/business-mapping/rules", response_model=BusinessMappingRuleResponse, status_code=201)
async def create_mapping_rule(
    rule_req: BusinessMappingRuleRequest,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleResponse:
    """Create a new business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    if rule_req.dimension not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
    existing = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.organization_id == membership.organization_id,
            BusinessMappingRule.tag_key == rule_req.tag_key,
            BusinessMappingRule.tag_value == rule_req.tag_value,
            BusinessMappingRule.dimension == rule_req.dimension,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="A mapping rule with the same tag_key/tag_value/dimension already exists.")
    now = _utcnow()
    customer_id = f"org-{membership.organization_id}"
    rule = BusinessMappingRule(
        organization_id=membership.organization_id,
        customer_id=customer_id,
        tag_key=rule_req.tag_key,
        tag_value=rule_req.tag_value,
        dimension=rule_req.dimension,
        mapped_value=rule_req.mapped_value,
        priority=rule_req.priority,
        is_active=rule_req.is_active,
        created_at=now,
        updated_at=now,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return BusinessMappingRuleResponse(
        id=rule.id,
        organization_id=rule.organization_id,
        customer_id=rule.customer_id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        dimension=rule.dimension,
        mapped_value=rule.mapped_value,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at.isoformat(),
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.put("/business-mapping/rules/{rule_id}", response_model=BusinessMappingRuleResponse)
async def update_mapping_rule(
    rule_id: int,
    rule_req: BusinessMappingRuleUpdateRequest,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleResponse:
    """Update an existing business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    rule = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.id == rule_id,
            BusinessMappingRule.organization_id == membership.organization_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Mapping rule not found.")
    if rule_req.dimension is not None and rule_req.dimension not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
    if rule_req.tag_key is not None:
        rule.tag_key = rule_req.tag_key
    if rule_req.tag_value is not None:
        rule.tag_value = rule_req.tag_value
    if rule_req.dimension is not None:
        rule.dimension = rule_req.dimension
    if rule_req.mapped_value is not None:
        rule.mapped_value = rule_req.mapped_value
    if rule_req.priority is not None:
        rule.priority = rule_req.priority
    if rule_req.is_active is not None:
        rule.is_active = rule_req.is_active
    rule.updated_at = _utcnow()
    db.commit()
    db.refresh(rule)
    return BusinessMappingRuleResponse(
        id=rule.id,
        organization_id=rule.organization_id,
        customer_id=rule.customer_id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        dimension=rule.dimension,
        mapped_value=rule.mapped_value,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at.isoformat(),
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.delete("/business-mapping/rules/{rule_id}", status_code=204)
async def delete_mapping_rule(
    rule_id: int,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> None:
    """Delete a business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    rule = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.id == rule_id,
            BusinessMappingRule.organization_id == membership.organization_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Mapping rule not found.")
    db.delete(rule)
    db.commit()


def _apply_mapping_rules(
    org_id: int,
    customer_id: str,
    records: list,
    rules: list,
    db: Session,
) -> int:
    """Apply active mapping rules to imported cost records, writing NormalizedCostDimension rows.

    Matching: a rule matches a record when the record's ``tags_json`` contains ``tag_key``,
    and either ``tag_value == "*"`` or ``tag_value`` matches the stored tag value exactly.
    Returns the count of dimension rows written.
    """
    now = _utcnow()
    written = 0
    for record in records:
        tags: dict = {}
        if record.tags_json:
            try:
                tags = json.loads(record.tags_json)
            except (ValueError, TypeError):
                tags = {}

        # Build dimension assignments from matching rules (highest priority = lowest number)
        dim_map: Dict[str, tuple] = {}  # dimension -> (mapped_value, rule_id)
        matched_rule_ids: List[int] = []
        for rule in sorted(rules, key=lambda r: (r.priority, r.id)):
            tag_val = tags.get(rule.tag_key)
            if tag_val is None:
                continue
            if rule.tag_value != "*" and rule.tag_value != str(tag_val):
                continue
            if rule.dimension not in dim_map:
                dim_map[rule.dimension] = (rule.mapped_value, rule.id)
                matched_rule_ids.append(rule.id)

        is_mapped = bool(dim_map)
        existing = (
            db.query(NormalizedCostDimension)
            .filter(
                NormalizedCostDimension.imported_cost_record_id == record.id,
                NormalizedCostDimension.organization_id == org_id,
            )
            .first()
        )
        if existing:
            existing.team = dim_map.get("team", (None, None))[0]
            existing.environment = dim_map.get("environment", (None, None))[0]
            existing.application = dim_map.get("application", (None, None))[0]
            existing.cost_center = dim_map.get("cost_center", (None, None))[0]
            existing.is_mapped = is_mapped
            existing.mapping_rule_ids_json = json.dumps(matched_rule_ids)
            existing.captured_at = now
        else:
            row = NormalizedCostDimension(
                organization_id=org_id,
                customer_id=customer_id,
                imported_cost_record_id=record.id,
                provider=record.provider,
                service_name=record.service_name,
                region=record.region,
                cost_usd=float(record.cost_usd or 0.0),
                team=dim_map.get("team", (None, None))[0],
                environment=dim_map.get("environment", (None, None))[0],
                application=dim_map.get("application", (None, None))[0],
                cost_center=dim_map.get("cost_center", (None, None))[0],
                is_mapped=is_mapped,
                mapping_rule_ids_json=json.dumps(matched_rule_ids),
                captured_at=now,
            )
            db.add(row)
            written += 1
    db.commit()
    return written


@router.post("/business-mapping/apply", status_code=200)
async def apply_mapping_rules(
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Re-apply all active mapping rules to imported cost records for this organization."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"
    rules = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.organization_id == org_id,
            BusinessMappingRule.is_active == True,  # noqa: E712
        )
        .all()
    )
    records = (
        db.query(ImportedCostRecord)
        .filter(ImportedCostRecord.customer_id == customer_id)
        .all()
    )
    written = _apply_mapping_rules(org_id, customer_id, records, rules, db)
    return {
        "status": "ok",
        "rules_applied": len(rules),
        "records_processed": len(records),
        "dimension_rows_written": written,
    }


@router.get("/chargeback", response_model=ChargebackResponse)
async def get_chargeback(
    dimension_type: str = "team",
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ChargebackResponse:
    """Return chargeback/showback aggregation for a business dimension."""
    if dimension_type not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension_type must be one of {sorted(VALID_DIMENSIONS)}")
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"

    # Ensure dimension rows exist; auto-apply if none found
    count = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .count()
    )
    if count == 0:
        rules = (
            db.query(BusinessMappingRule)
            .filter(
                BusinessMappingRule.organization_id == org_id,
                BusinessMappingRule.is_active == True,  # noqa: E712
            )
            .all()
        )
        if rules:
            records = (
                db.query(ImportedCostRecord)
                .filter(ImportedCostRecord.customer_id == customer_id)
                .all()
            )
            _apply_mapping_rules(org_id, customer_id, records, rules, db)

    rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    dim_attr = dimension_type  # "team" | "environment" | "application" | "cost_center"
    groups: Dict[str, ChargebackDimensionGroup] = {}
    total_mapped = 0.0
    total_unmapped = 0.0

    for row in rows:
        val = getattr(row, dim_attr, None)
        cost = float(row.cost_usd or 0.0)
        if val is None:
            total_unmapped += cost
            continue
        total_mapped += cost
        if val not in groups:
            groups[val] = ChargebackDimensionGroup(
                dimension=dim_attr,
                value=val,
                total_cost_usd=0.0,
                provider_breakdown={},
                record_count=0,
            )
        g = groups[val]
        g.total_cost_usd = round(g.total_cost_usd + cost, 4)
        g.provider_breakdown[row.provider] = round(g.provider_breakdown.get(row.provider, 0.0) + cost, 4)
        g.record_count += 1

    total = total_mapped + total_unmapped
    coverage = round((total_mapped / total * 100) if total > 0 else 0.0, 2)

    return ChargebackResponse(
        organization_id=org_id,
        dimension_type=dim_attr,
        groups=sorted(groups.values(), key=lambda g: g.total_cost_usd, reverse=True),
        total_mapped_cost_usd=round(total_mapped, 2),
        total_unmapped_cost_usd=round(total_unmapped, 2),
        total_cost_usd=round(total, 2),
        coverage_percent=coverage,
    )


@router.get("/chargeback/coverage", response_model=AllocationCoverageResponse)
async def get_allocation_coverage(
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> AllocationCoverageResponse:
    """Return allocation quality: mapped vs unmapped spend breakdown."""
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"

    # Auto-apply rules if no dimension rows exist yet
    count = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .count()
    )
    if count == 0:
        rules = (
            db.query(BusinessMappingRule)
            .filter(
                BusinessMappingRule.organization_id == org_id,
                BusinessMappingRule.is_active == True,  # noqa: E712
            )
            .all()
        )
        if rules:
            records = (
                db.query(ImportedCostRecord)
                .filter(ImportedCostRecord.customer_id == customer_id)
                .all()
            )
            _apply_mapping_rules(org_id, customer_id, records, rules, db)

    rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    total_cost = sum(float(r.cost_usd or 0) for r in rows)
    mapped_cost = sum(float(r.cost_usd or 0) for r in rows if r.is_mapped)
    unmapped_cost = total_cost - mapped_cost
    coverage = round((mapped_cost / total_cost * 100) if total_cost > 0 else 0.0, 2)

    dim_coverage: Dict[str, float] = {}
    for dim in VALID_DIMENSIONS:
        dim_mapped = sum(float(r.cost_usd or 0) for r in rows if getattr(r, dim, None) is not None)
        dim_coverage[dim] = round((dim_mapped / total_cost * 100) if total_cost > 0 else 0.0, 2)

    prov_totals: Dict[str, float] = {}
    prov_mapped: Dict[str, float] = {}
    for r in rows:
        prov_totals[r.provider] = prov_totals.get(r.provider, 0.0) + float(r.cost_usd or 0)
        if r.is_mapped:
            prov_mapped[r.provider] = prov_mapped.get(r.provider, 0.0) + float(r.cost_usd or 0)
    provider_coverage = {
        prov: round((prov_mapped.get(prov, 0.0) / tot * 100) if tot > 0 else 0.0, 2)
        for prov, tot in prov_totals.items()
    }

    # Top unmapped services by cost
    unmapped_services: Dict[str, float] = {}
    for r in rows:
        if not r.is_mapped and r.service_name:
            unmapped_services[r.service_name] = (
                unmapped_services.get(r.service_name, 0.0) + float(r.cost_usd or 0)
            )
    top_unmapped = sorted(
        [{"service": s, "cost_usd": round(c, 2)} for s, c in unmapped_services.items()],
        key=lambda x: x["cost_usd"],
        reverse=True,
    )[:10]

    return AllocationCoverageResponse(
        organization_id=org_id,
        total_cost_usd=round(total_cost, 2),
        mapped_cost_usd=round(mapped_cost, 2),
        unmapped_cost_usd=round(unmapped_cost, 2),
        coverage_percent=coverage,
        dimension_coverage=dim_coverage,
        provider_coverage=provider_coverage,
        unmapped_top_services=top_unmapped,
    )


# ── Epic 4: Reporting & Executive Outputs ────────────────────────────────────

def _compute_period_bucket(dt: datetime, period_type: str) -> tuple[datetime, datetime]:
    """Return (period_start, period_end) for a datetime given period_type."""
    if period_type == "weekly":
        # Monday-anchored ISO week
        start = dt - timedelta(days=dt.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7) - timedelta(seconds=1)
    else:  # monthly
        start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if dt.month == 12:
            end = start.replace(year=dt.year + 1, month=1) - timedelta(seconds=1)
        else:
            end = start.replace(month=dt.month + 1) - timedelta(seconds=1)
    return start, end


def _build_xlsx_workbook(sheets: List[tuple[str, List[List[Any]]]]) -> bytes:
    """Build an xlsx workbook from a list of (sheet_name, rows) tuples.

    Each row is a list of cell values; the first row is treated as a header.
    Returns raw bytes of the workbook.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        # Auto-width approximation
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _chargeback_csv_rows(
    org_id: int,
    customer_id: str,
    db: Session,
) -> tuple[List[str], List[List[Any]]]:
    """Return (header, data_rows) for a full chargeback export."""
    dim_rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .order_by(NormalizedCostDimension.captured_at.desc())
        .all()
    )
    header = [
        "provider", "service_name", "region",
        "cost_usd", "team", "environment", "application", "cost_center",
        "is_mapped", "mapping_rule_ids", "captured_at",
    ]
    data: List[List[Any]] = []
    for r in dim_rows:
        data.append([
            r.provider or "",
            r.service_name or "",
            r.region or "",
            round(float(r.cost_usd or 0), 4),
            r.team or "",
            r.environment or "",
            r.application or "",
            r.cost_center or "",
            "yes" if r.is_mapped else "no",
            r.mapping_rule_ids_json or "",
            r.captured_at.isoformat() if r.captured_at else "",
        ])
    return header, data


@router.post("/reports/period-summaries/compute", response_model=PeriodSummaryComputeResponse)
async def compute_period_summaries(
    period_type: str = "monthly",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> PeriodSummaryComputeResponse:
    """Aggregate ImportedCostRecord rows into CostPeriodSummary buckets.

    Re-computing is idempotent — existing rows are overwritten via upsert.
    Supported period_type values: 'monthly', 'weekly'.
    """
    if period_type not in ("monthly", "weekly"):
        raise HTTPException(status_code=400, detail="period_type must be 'monthly' or 'weekly'.")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    records = (
        db.query(ImportedCostRecord)
        .filter(ImportedCostRecord.organization_id == org_id)
        .all()
    )

    # Fetch normalized dimension rows (for mapped cost tracking)
    dim_by_record_id: Dict[int, NormalizedCostDimension] = {
        r.imported_cost_record_id: r
        for r in db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
        if r.imported_cost_record_id is not None
    }

    # Bucket: (period_start, provider, team, environment) → aggregation dict
    buckets: Dict[tuple, Dict[str, Any]] = {}

    now = _utcnow()
    for rec in records:
        anchor = rec.period_start or rec.created_at or now
        pstart, pend = _compute_period_bucket(anchor, period_type)
        dim = dim_by_record_id.get(rec.id)
        team = (dim.team or "") if dim else ""
        environment = (dim.environment or "") if dim else ""
        key = (pstart, pend, rec.provider or "imported", team, environment)

        cost = float(rec.cost_usd or 0)
        if key not in buckets:
            buckets[key] = {
                "total_cost_usd": 0.0,
                "mapped_cost_usd": 0.0,
                "unmapped_cost_usd": 0.0,
                "record_count": 0,
                "services": {},
            }
        b = buckets[key]
        b["total_cost_usd"] += cost
        if dim and dim.is_mapped:
            b["mapped_cost_usd"] += cost
        else:
            b["unmapped_cost_usd"] += cost
        b["record_count"] += 1
        svc = rec.service_name or "unknown"
        b["services"][svc] = b["services"].get(svc, 0.0) + cost

    rows_written = 0
    computed_at = _utcnow()
    for (pstart, pend, provider, team, environment), b in buckets.items():
        # Upsert: delete any existing row for this bucket key
        db.query(CostPeriodSummary).filter(
            CostPeriodSummary.organization_id == org_id,
            CostPeriodSummary.period_type == period_type,
            CostPeriodSummary.period_start == pstart,
            CostPeriodSummary.provider == provider,
            CostPeriodSummary.team == (team or None),
            CostPeriodSummary.environment == (environment or None),
        ).delete(synchronize_session=False)

        row = CostPeriodSummary(
            organization_id=org_id,
            customer_id=customer_id,
            period_type=period_type,
            period_start=pstart,
            period_end=pend,
            provider=provider,
            team=team or None,
            environment=environment or None,
            total_cost_usd=round(b["total_cost_usd"], 4),
            mapped_cost_usd=round(b["mapped_cost_usd"], 4),
            unmapped_cost_usd=round(b["unmapped_cost_usd"], 4),
            record_count=b["record_count"],
            service_breakdown_json=json.dumps(
                {k: round(v, 4) for k, v in sorted(b["services"].items(), key=lambda x: -x[1])[:20]}
            ),
            computed_at=computed_at,
        )
        db.add(row)
        rows_written += 1

    db.commit()

    return PeriodSummaryComputeResponse(
        organization_id=org_id,
        period_type=period_type,
        periods_computed=len({(k[0], k[2]) for k in buckets}),
        rows_written=rows_written,
        computed_at=computed_at.isoformat(),
    )


@router.get("/reports/cost-trend", response_model=CostTrendResponse)
async def get_cost_trend(
    period_type: str = "monthly",
    lookback: int = 6,
    provider: Optional[str] = None,
    view_by: str = "provider",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> CostTrendResponse:
    """Return time-series cost data for trend charts.

    Supports trend views by provider, region, service, and account.
    """
    _ = current_user
    if period_type not in ("monthly", "weekly"):
        raise HTTPException(status_code=400, detail="period_type must be 'monthly' or 'weekly'.")
    if view_by not in _SUPPORTED_TREND_VIEWS:
        raise HTTPException(status_code=400, detail="view_by must be one of: provider, region, service, account.")

    lookback = max(1, min(lookback, 18))
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    require_live_provider_data = Config().require_live_provider_data

    summaries: List[CostPeriodSummary] = []
    archive_rows: List[Dict[str, Any]] = []
    data_source = "raw_records"

    if view_by == "provider":
        q = db.query(CostPeriodSummary).filter(
            CostPeriodSummary.organization_id == org_id,
            CostPeriodSummary.period_type == period_type,
        )
        if provider:
            q = q.filter(CostPeriodSummary.provider == provider.lower())
        summaries = q.order_by(CostPeriodSummary.period_start.desc()).all()

        cfg = Config()
        hot_months = cfg.retention_hot_months
        if lookback > hot_months and cfg.retention_enabled and cfg.oci_archive_bucket:
            now = _utcnow()
            archive_end = now - timedelta(days=hot_months * 30)
            archive_start = now - timedelta(days=lookback * 30)
            try:
                raw_archive = await asyncio.to_thread(
                    fetch_archived_period_summaries, cfg, org_id, archive_start, archive_end
                )
                if provider:
                    raw_archive = [r for r in raw_archive if r.get("provider", "").lower() == provider.lower()]
                raw_archive = [r for r in raw_archive if r.get("period_type", "") == period_type]
                archive_rows = raw_archive
            except Exception:
                logger.exception("Failed to fetch archived period summaries; continuing with DB only")

        if summaries or archive_rows:
            data_source = "computed"

    points: List[CostTrendPoint] = []

    if data_source == "computed":
        by_period: Dict[tuple, Dict[str, Any]] = {}
        for s in summaries:
            key = (s.period_start, s.period_end, s.provider or "imported")
            if key not in by_period:
                by_period[key] = {"total": 0.0, "mapped": 0.0, "unmapped": 0.0, "count": 0, "svcs": {}}
            b = by_period[key]
            b["total"] += float(s.total_cost_usd or 0)
            b["mapped"] += float(s.mapped_cost_usd or 0)
            b["unmapped"] += float(s.unmapped_cost_usd or 0)
            b["count"] += int(s.record_count or 0)
            if s.service_breakdown_json:
                try:
                    for svc, c in json.loads(s.service_breakdown_json).items():
                        b["svcs"][svc] = b["svcs"].get(svc, 0.0) + c
                except (json.JSONDecodeError, ValueError):
                    pass

        for ar in archive_rows:
            try:
                ps = datetime.fromisoformat(str(ar["period_start"]).replace("Z", ""))
                pe = datetime.fromisoformat(str(ar["period_end"]).replace("Z", ""))
            except (KeyError, ValueError):
                continue
            key = (ps, pe, ar.get("provider", "unknown"))
            if key not in by_period:
                by_period[key] = {"total": 0.0, "mapped": 0.0, "unmapped": 0.0, "count": 0, "svcs": {}}
            b = by_period[key]
            b["total"] += float(ar.get("total_cost_usd") or 0)
            b["mapped"] += float(ar.get("mapped_cost_usd") or 0)
            b["unmapped"] += float(ar.get("unmapped_cost_usd") or 0)
            b["count"] += int(ar.get("record_count") or 0)
            svc_json = ar.get("service_breakdown_json")
            if svc_json:
                try:
                    for svc, c in json.loads(svc_json).items():
                        b["svcs"][svc] = b["svcs"].get(svc, 0.0) + c
                except (json.JSONDecodeError, ValueError):
                    pass

        sorted_keys = sorted(by_period.keys(), key=lambda k: k[0], reverse=True)[:lookback]
        points = [
            CostTrendPoint(
                period_start=k[0].isoformat(),
                period_end=k[1].isoformat(),
                provider=k[2],
                dimension_value=k[2],
                total_cost_usd=round(by_period[k]["total"], 2),
                mapped_cost_usd=round(by_period[k]["mapped"], 2),
                unmapped_cost_usd=round(by_period[k]["unmapped"], 2),
                record_count=by_period[k]["count"],
                service_breakdown={s: round(c, 2) for s, c in list(by_period[k]["svcs"].items())[:10]},
            )
            for k in sorted(sorted_keys, key=lambda k: k[0])
        ]
    elif require_live_provider_data:
        if view_by != "provider":
            return CostTrendResponse(
                organization_id=org_id,
                period_type=period_type,
                lookback_periods=lookback,
                view_by=view_by,
                data_source="empty",
                points=[],
                provider_totals={},
                dimension_totals={},
                grand_total_usd=0.0,
            )

        snapshots_query = db.query(CostSnapshot).filter(CostSnapshot.customer_id == customer_id)
        if provider:
            snapshots_query = snapshots_query.filter(CostSnapshot.provider == provider.lower())
        snapshots = (
            snapshots_query
            .order_by(CostSnapshot.captured_at.desc())
            .limit(1000)
            .all()
        )
        if not snapshots:
            live_context = await _cost_context(membership, db, "month", provider.lower() if provider else "all")
            live_source = str(live_context.get("source") or "").strip().lower()
            live_errors = live_context.get("provider_errors")
            live_total = float(live_context.get("total_cost") or 0.0)
            if live_source.startswith("live_provider_api") and not live_errors and live_total > 0:
                pstart, pend = _compute_period_bucket(_utcnow(), period_type)
                breakdown = live_context.get("breakdown")
                live_points: List[CostTrendPoint] = []

                if isinstance(breakdown, dict) and breakdown:
                    for provider_key, provider_row in sorted(breakdown.items()):
                        provider_name = str(provider_key or "live").strip().lower() or "live"
                        if provider and provider_name != provider.lower():
                            continue
                        if isinstance(provider_row, dict):
                            provider_cost = float(provider_row.get("cost") or 0.0)
                        else:
                            provider_cost = float(provider_row or 0.0)
                        if provider_cost <= 0:
                            continue
                        live_points.append(
                            CostTrendPoint(
                                period_start=pstart.isoformat(),
                                period_end=pend.isoformat(),
                                provider=provider_name,
                                dimension_value=provider_name,
                                total_cost_usd=round(provider_cost, 2),
                                mapped_cost_usd=0.0,
                                unmapped_cost_usd=round(provider_cost, 2),
                                record_count=1,
                                service_breakdown={},
                            )
                        )

                if not live_points:
                    provider_name = provider.lower() if provider else "live"
                    live_points = [
                        CostTrendPoint(
                            period_start=pstart.isoformat(),
                            period_end=pend.isoformat(),
                            provider=provider_name,
                            dimension_value=provider_name,
                            total_cost_usd=round(live_total, 2),
                            mapped_cost_usd=0.0,
                            unmapped_cost_usd=round(live_total, 2),
                            record_count=1,
                            service_breakdown={},
                        )
                    ]

                points = live_points[:lookback]
                data_source = "live_provider_api_current_period"
            else:
                return CostTrendResponse(
                    organization_id=org_id,
                    period_type=period_type,
                    lookback_periods=lookback,
                    view_by=view_by,
                    data_source="empty",
                    points=[],
                    provider_totals={},
                    dimension_totals={},
                    grand_total_usd=0.0,
                )
        else:
            live_buckets: Dict[tuple[datetime, datetime, str], Dict[str, Any]] = {}
            now = _utcnow()
            for row in snapshots:
                provider_key = str(row.provider or "unknown").strip().lower() or "unknown"
                anchor = row.period_start or row.period_end or row.captured_at or now
                pstart, pend = _compute_period_bucket(anchor, period_type)
                key = (pstart, pend, provider_key)
                bucket = live_buckets.setdefault(
                    key,
                    {
                        "total": 0.0,
                        "record_count": 0,
                    },
                )
                bucket["total"] += float(row.total_cost_usd or 0.0)
                bucket["record_count"] += 1

            sorted_keys = sorted(live_buckets.keys(), key=lambda item: item[0], reverse=True)[:lookback]
            points = [
                CostTrendPoint(
                    period_start=key[0].isoformat(),
                    period_end=key[1].isoformat(),
                    provider=key[2],
                    dimension_value=key[2],
                    total_cost_usd=round(float(live_buckets[key]["total"] or 0.0), 2),
                    mapped_cost_usd=0.0,
                    unmapped_cost_usd=round(float(live_buckets[key]["total"] or 0.0), 2),
                    record_count=int(live_buckets[key]["record_count"] or 0),
                    service_breakdown={},
                )
                for key in sorted(sorted_keys, key=lambda item: item[0])
            ]
            data_source = "cost_snapshots_live"
    else:
        records = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .all()
        )
        if not records:
            return CostTrendResponse(
                organization_id=org_id,
                period_type=period_type,
                lookback_periods=lookback,
                view_by=view_by,
                data_source="empty",
                points=[],
                provider_totals={},
                dimension_totals={},
                grand_total_usd=0.0,
            )

        now = _utcnow()
        buckets: Dict[tuple, Dict[str, Any]] = {}
        for rec in records:
            rec_provider = (rec.provider or "imported").lower()
            if provider and rec_provider != provider.lower():
                continue
            anchor = rec.period_start or rec.created_at or now
            pstart, pend = _compute_period_bucket(anchor, period_type)
            dim_value = _trend_dimension_value(rec, view_by)
            key = (pstart, pend, dim_value)
            cost = float(rec.cost_usd or 0)
            if key not in buckets:
                buckets[key] = {
                    "pend": pend,
                    "total": 0.0,
                    "mapped": 0.0,
                    "unmapped": 0.0,
                    "count": 0,
                    "svcs": {},
                    "providers": {},
                }
            b = buckets[key]
            b["total"] += cost
            b["unmapped"] += cost
            b["count"] += 1
            b["providers"][rec_provider] = b["providers"].get(rec_provider, 0.0) + cost
            svc = rec.service_name or "unknown"
            b["svcs"][svc] = b["svcs"].get(svc, 0.0) + cost

        sorted_keys = sorted(buckets.keys(), key=lambda k: k[0], reverse=True)[:lookback]
        for k in sorted(sorted_keys, key=lambda v: v[0]):
            b = buckets[k]
            providers = sorted((b.get("providers") or {}).items(), key=lambda item: item[1], reverse=True)
            dominant_provider = providers[0][0] if providers else "imported"
            provider_value = dominant_provider if len(providers) == 1 else "multi-cloud"
            points.append(
                CostTrendPoint(
                    period_start=k[0].isoformat(),
                    period_end=b["pend"].isoformat(),
                    provider=provider_value,
                    dimension_value=k[2],
                    total_cost_usd=round(b["total"], 2),
                    mapped_cost_usd=round(b["mapped"], 2),
                    unmapped_cost_usd=round(b["unmapped"], 2),
                    record_count=b["count"],
                    service_breakdown={s: round(c, 2) for s, c in list(b["svcs"].items())[:10]},
                )
            )

    provider_totals: Dict[str, float] = {}
    dimension_totals: Dict[str, float] = {}
    for p in points:
        provider_totals[p.provider] = round(provider_totals.get(p.provider, 0.0) + p.total_cost_usd, 2)
        dim_key = p.dimension_value or p.provider
        dimension_totals[dim_key] = round(dimension_totals.get(dim_key, 0.0) + p.total_cost_usd, 2)

    return CostTrendResponse(
        organization_id=org_id,
        period_type=period_type,
        lookback_periods=lookback,
        view_by=view_by,
        data_source=data_source,
        points=points,
        provider_totals=provider_totals,
        dimension_totals=dimension_totals,
        grand_total_usd=round(sum(p.total_cost_usd for p in points), 2),
    )


@router.get("/reports/chargeback.csv")
async def download_chargeback_csv(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a full chargeback/allocation CSV with all business dimensions."""
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    header, data = _chargeback_csv_rows(org_id, customer_id, db)
    return _csv_response("optiora-chargeback.csv", header, data)


@router.get("/reports/chargeback.xlsx")
async def download_chargeback_xlsx(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a multi-sheet Excel workbook with chargeback + executive summary."""
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl is not installed on this server.")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    # Sheet 1: Chargeback detail
    header, data = _chargeback_csv_rows(org_id, customer_id, db)
    chargeback_rows = [header] + data

    # Sheet 2: Executive summary (reuse existing helper)
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)

    # Sheet 3: Trend summary (last 6 monthly periods)
    trend = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        current_user=current_user,
        membership=membership,
        db=db,
    )
    trend_header = ["period_start", "period_end", "provider", "total_cost_usd", "mapped_cost_usd", "unmapped_cost_usd", "record_count"]
    trend_rows: List[List[Any]] = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.provider, p.total_cost_usd, p.mapped_cost_usd, p.unmapped_cost_usd, p.record_count]
        for p in trend.points
    ]

    xlsx_bytes = _build_xlsx_workbook([
        ("Chargeback Detail", chargeback_rows),
        ("Executive Summary", exec_rows),
        ("Cost Trend", trend_rows),
    ])

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-report-{_utcnow().strftime("%Y%m%d")}.xlsx"'
        },
    )


def _build_simple_pdf(title: str, lines: List[str]) -> bytes:
    """Build a lightweight PDF document without external dependencies."""
    safe_lines = [line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)") for line in lines]
    y_start = 780
    line_height = 16
    text_ops = ["BT", "/F1 12 Tf", f"72 {y_start} Td", f"({title}) Tj"]
    for idx, line in enumerate(safe_lines[:42], start=1):
        text_ops.append(f"0 -{line_height} Td")
        text_ops.append(f"({line}) Tj")
    text_ops.append("ET")
    stream_text = "\n".join(text_ops)

    objects: List[str] = []
    objects.append("1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj")
    objects.append("2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj")
    objects.append("3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj")
    objects.append("4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj")
    objects.append(f"5 0 obj << /Length {len(stream_text.encode('utf-8'))} >> stream\n{stream_text}\nendstream endobj")

    header = "%PDF-1.4\n"
    body = ""
    offsets = [0]
    current = len(header.encode("utf-8"))
    for obj in objects:
        offsets.append(current)
        encoded = (obj + "\n").encode("utf-8")
        body += obj + "\n"
        current += len(encoded)

    xref_start = current
    xref = [f"xref\n0 {len(offsets)}\n", "0000000000 65535 f \n"]
    for off in offsets[1:]:
        xref.append(f"{off:010d} 00000 n \n")
    trailer = (
        f"trailer << /Size {len(offsets)} /Root 1 0 R >>\n"
        f"startxref\n{xref_start}\n%%EOF\n"
    )
    pdf = header + body + "".join(xref) + trailer
    return pdf.encode("utf-8")


def _digest_lines(exec_rows: List[List[Any]], frequency: str) -> List[str]:
    kv = {f"{r[0]}::{r[1]}": r[2] for r in exec_rows[1:] if len(r) >= 3}
    return [
        f"Frequency: {frequency}",
        f"Generated At: {kv.get('Summary::Generated At', _utcnow().isoformat())}",
        f"Total Monthly Cost USD: {kv.get('Summary::Total Monthly Cost USD', 0)}",
        f"Potential Monthly Savings USD: {kv.get('Summary::Potential Monthly Savings USD', 0)}",
        f"Risk Score: {kv.get('Summary::Risk Score', 0)}",
        f"Maturity Score: {kv.get('Summary::Maturity Score', 0)}",
        f"Spend At Risk USD: {kv.get('Summary::Spend At Risk USD', 0)}",
        f"Optimization Capacity USD: {kv.get('Summary::Optimization Capacity USD', 0)}",
        f"Open Alerts: {kv.get('Summary::Open Alerts', 0)}",
    ]


@router.get("/reports/executive-summary.xlsx")
@router.get("/reports/finance-workbook.xlsx")
async def download_executive_summary_xlsx(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a finance-friendly multi-sheet workbook."""
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl is not installed on this server.")

    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)

    trend_provider = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        view_by="provider",
        current_user=current_user,
        membership=membership,
        db=db,
    )
    trend_region = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        view_by="region",
        current_user=current_user,
        membership=membership,
        db=db,
    )

    trend_header = ["period_start", "period_end", "dimension", "provider", "total_cost_usd", "record_count"]
    provider_rows = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
        for p in trend_provider.points
    ]
    region_rows = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
        for p in trend_region.points
    ]

    chargeback_header, chargeback_data = _chargeback_csv_rows(org_id, customer_id, db)
    chargeback_rows = [chargeback_header] + chargeback_data
    ledger_rows_raw = _query_recommendation_ledger(
        db=db,
        organization_id=org_id,
        provider="all",
        status_filter="all",
        limit=5000,
    )
    ledger_rows = [[
        "ledger_id",
        "provider",
        "resource_id",
        "resource_name",
        "recommendation_source",
        "action",
        "status",
        "planned_monthly_savings_usd",
        "realized_monthly_savings_usd",
        "variance_monthly_usd",
        "planned_annual_savings_usd",
        "realized_annual_savings_usd",
        "variance_annual_usd",
        "variance_percent",
        "variance_reason",
        "last_seen_at",
    ]]
    ledger_rows.extend([
        [
            row.id,
            row.provider or "",
            row.resource_id or "",
            row.resource_name or "",
            row.recommendation_source or "",
            row.action or "",
            row.status or "",
            round(float(row.planned_monthly_savings_usd or 0.0), 2),
            round(float(row.realized_monthly_savings_usd or 0.0), 2),
            round(float(row.variance_monthly_usd or 0.0), 2),
            round(float(row.planned_annual_savings_usd or 0.0), 2),
            round(float(row.realized_annual_savings_usd or 0.0), 2),
            round(float(row.variance_annual_usd or 0.0), 2),
            round(float(row.variance_percent or 0.0), 2),
            row.variance_reason or "",
            row.last_seen_at.isoformat() if row.last_seen_at else "",
        ]
        for row in ledger_rows_raw
    ])

    xlsx_bytes = _build_xlsx_workbook([
        ("Executive Summary", exec_rows),
        ("Trend by Provider", provider_rows),
        ("Trend by Region", region_rows),
        ("Chargeback Detail", chargeback_rows),
        ("Recommendation Ledger", ledger_rows),
    ])
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-finance-workbook-{_utcnow().strftime("%Y%m%d")}.xlsx"'
        },
    )


@router.get("/reports/executive-digest.pdf")
async def download_executive_digest_pdf(
    frequency: Literal["weekly", "monthly"] = "weekly",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    digest_lines = _digest_lines(exec_rows, frequency)
    pdf_bytes = _build_simple_pdf(f"OptiOra {frequency.title()} Executive Digest", digest_lines)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-{frequency}-digest-{_utcnow().strftime("%Y%m%d")}.pdf"'
        },
    )


@router.post("/reports/share-token", response_model=ReportShareTokenResponse)
async def create_report_share_token(
    payload: ReportShareTokenRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ReportShareTokenResponse:
    _require_management_role(membership, "report sharing")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    expires_hours = max(1, min(payload.expires_in_hours, 24 * 90))
    expires_at = datetime.now(timezone.utc) + timedelta(hours=expires_hours)
    token = _build_report_share_token(
        {
            "org_id": org_id,
            "customer_id": customer_id,
            "report_type": payload.report_type,
            "report_format": payload.report_format,
            "exp": int(expires_at.timestamp()),
            "sub": current_user.id,
        }
    )
    return ReportShareTokenResponse(
        token=token,
        expires_at=expires_at.isoformat(),
        report_type=payload.report_type,
        report_format=payload.report_format,
    )


@router.get("/reports/shared/{token}")
async def read_shared_report(
    token: str,
    db: Session = Depends(get_db),
) -> Response:
    payload = _parse_report_share_token(token)
    org_id = int(payload.get("org_id"))
    customer_id = str(payload.get("customer_id"))
    report_type = str(payload.get("report_type"))
    report_format = str(payload.get("report_format"))

    membership_stub = SimpleNamespace(organization_id=org_id, role=UserRole.READONLY)
    user_stub = SimpleNamespace(id=None)

    if report_type == "executive_digest" and report_format == "pdf":
        exec_rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        pdf_bytes = _build_simple_pdf("OptiOra Executive Digest", _digest_lines(exec_rows, "shared"))
        return Response(content=pdf_bytes, media_type="application/pdf")

    if report_type == "finance_workbook" and report_format == "xlsx":
        exec_rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        trend_provider = await get_cost_trend(
            period_type="monthly",
            lookback=6,
            provider=None,
            view_by="provider",
            current_user=user_stub,
            membership=membership_stub,
            db=db,
        )
        trend_rows = [["period_start", "period_end", "dimension", "provider", "total_cost_usd", "record_count"]] + [
            [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
            for p in trend_provider.points
        ]
        header, data = _chargeback_csv_rows(org_id, customer_id, db)
        xlsx_bytes = _build_xlsx_workbook([
            ("Executive Summary", exec_rows),
            ("Trend by Provider", trend_rows),
            ("Chargeback Detail", [header] + data),
        ])
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if report_type == "executive_summary" and report_format in {"json", "csv"}:
        rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        if report_format == "csv":
            return _csv_response("shared-executive-summary.csv", rows[0], rows[1:])
        return Response(
            content=json.dumps({"rows": rows}, default=str),
            media_type="application/json",
        )

    raise HTTPException(status_code=400, detail="Unsupported shared report token payload")


# ── End Epic 4 endpoints ──────────────────────────────────────────────────────

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FOCUS Export  (FinOps Open Cost and Usage Specification v1.0)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_FOCUS_COLUMNS = [
    "BilledCost", "BillingAccountId", "BillingAccountName", "BillingCurrency",
    "BillingPeriodEnd", "BillingPeriodStart", "ChargePeriodEnd", "ChargePeriodStart",
    "ChargeType", "CommitmentDiscountId", "CommitmentDiscountName", "CommitmentDiscountType",
    "EffectiveCost", "InvoiceIssuerName", "ListCost", "ListUnitPrice",
    "PricingCategory", "PricingQuantity", "PricingUnit", "ProviderName",
    "PublisherName", "RegionId", "RegionName", "ResourceId", "ResourceName",
    "ResourceType", "ServiceCategory", "ServiceName", "SkuId", "SkuPriceId",
    "SubAccountId", "SubAccountName", "Tags",
]


def _focus_rows_from_context(context: Dict[str, Any]) -> List[List[Any]]:
    """Convert internal cost context to FOCUS 1.0 rows."""
    rows: List[List[Any]] = []
    period_start = _utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    period_end = _utcnow()

    for provider, data in (context.get("breakdown") or {}).items():
        cost = round(float(data.get("cost") or 0.0), 6)
        rows.append([
            cost,                                     # BilledCost
            provider,                                 # BillingAccountId
            provider.upper(),                         # BillingAccountName
            "USD",                                    # BillingCurrency
            period_end.isoformat() + "Z",            # BillingPeriodEnd
            period_start.isoformat() + "Z",          # BillingPeriodStart
            period_end.isoformat() + "Z",            # ChargePeriodEnd
            period_start.isoformat() + "Z",          # ChargePeriodStart
            "Usage",                                  # ChargeType
            None, None, None,                         # CommitmentDiscount*
            cost,                                     # EffectiveCost
            provider.upper(),                         # InvoiceIssuerName
            cost,                                     # ListCost
            None,                                     # ListUnitPrice
            "On-Demand",                              # PricingCategory
            1.0,                                      # PricingQuantity
            "Month",                                  # PricingUnit
            provider.upper(),                         # ProviderName
            provider.upper(),                         # PublisherName
            "global",                                 # RegionId
            "Global",                                 # RegionName
            None, None,                               # ResourceId, ResourceName
            "Cloud Service",                          # ResourceType
            "Compute",                                # ServiceCategory
            provider.upper(),                         # ServiceName
            None, None,                               # SkuId, SkuPriceId
            provider,                                 # SubAccountId
            provider.upper(),                         # SubAccountName
            "{}",                                     # Tags
        ])

    # Also emit rows from imported cost records if present
    for icr in (context.get("imported_rows") or []):
        cost = round(float(icr.get("cost_usd") or 0.0), 6)
        prov = icr.get("provider", "unknown")
        svc = icr.get("service_name") or "Unknown Service"
        acct = icr.get("account_identifier") or prov
        region = icr.get("region") or "global"
        p_start = icr.get("period_start") or period_start.isoformat() + "Z"
        p_end = icr.get("period_end") or period_end.isoformat() + "Z"
        rows.append([
            cost, acct, acct, "USD",
            p_end, p_start, p_end, p_start,
            "Usage", None, None, None,
            cost, prov.upper(), cost, None,
            "On-Demand", 1.0, "Month",
            prov.upper(), prov.upper(),
            region, region,
            None, svc, "Cloud Service", "Compute",
            svc, None, None,
            acct, acct, "{}",
        ])

    return rows


@router.get("/exports/focus.csv")
async def export_focus_csv(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Export cloud costs in FinOps FOCUS 1.0 format (CSV)."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    rows = _focus_rows_from_context(context)
    return _csv_response(
        filename=f"optiora-focus-{_utcnow().strftime('%Y%m%d')}.csv",
        header=_FOCUS_COLUMNS,
        rows=rows,
    )


@router.get("/exports/focus.json")
async def export_focus_json(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Export cloud costs in FinOps FOCUS 1.0 format (JSON)."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    rows = _focus_rows_from_context(context)
    records = [dict(zip(_FOCUS_COLUMNS, row)) for row in rows]
    return {
        "focus_version": "1.0",
        "generated_at": _utcnow().isoformat() + "Z",
        "record_count": len(records),
        "records": records,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Unit Economics Cockpit  — enhanced with user-supplied business metrics
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class UnitEconomicsMetricRequest(BaseModel):
    metric_name: str = Field(..., description="e.g. 'customers', 'requests', 'transactions'")
    metric_value: float = Field(..., gt=0, description="Current period count/volume")
    metric_unit: str = Field(default="units")


@router.post("/analytics/unit-economics/metrics")
async def record_unit_economics_metric(
    payload: UnitEconomicsMetricRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Record a business metric value for unit-economics cost-per-unit calculation."""
    _ = db
    context = await _cost_context(membership, db, "month", "all")
    total_cost = float(context.get("total_cost") or 0.0)
    cost_per_unit = round(total_cost / payload.metric_value, 4) if payload.metric_value > 0 else None
    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "metric_name": payload.metric_name,
        "metric_unit": payload.metric_unit,
        "metric_value": payload.metric_value,
        "total_monthly_cost_usd": round(total_cost, 2),
        "cost_per_unit_usd": cost_per_unit,
        "cost_per_unit_label": f"${cost_per_unit:.4f} per {payload.metric_unit}" if cost_per_unit else "N/A",
        "benchmark_note": "Cost per unit should trend downward as volume grows while costs stabilize.",
    }


@router.get("/analytics/unit-economics/cockpit")
async def get_unit_economics_cockpit(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Full unit economics cockpit: cost metrics + per-provider breakdown + waste-to-spend."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    unit_result = _safe_json_load(
        await finops_analytics.get_unit_economics({
            "current_monthly_spend": context["total_cost"],
            "estimated_waste_usd": analytics_result.get("estimated_monthly_waste_usd", 0),
            "identified_savings_usd": analytics_result.get("identified_monthly_savings_usd", 0),
            "anomalies": 0,
        }),
        {},
    )

    total = float(context.get("total_cost") or 0.0)
    waste = float(analytics_result.get("estimated_monthly_waste_usd") or 0.0)
    savings = float(analytics_result.get("identified_monthly_savings_usd") or 0.0)

    provider_metrics: List[Dict[str, Any]] = []
    for provider, data in (context.get("breakdown") or {}).items():
        cost = float(data.get("cost") or 0.0)
        provider_metrics.append({
            "provider": provider,
            "cost_usd": round(cost, 2),
            "share_percent": round(data.get("percentage") or 0.0, 1),
            "estimated_waste_usd": round(cost * 0.18, 2),  # 18% heuristic per provider
            "efficiency_index": round(max(0, 1 - (cost * 0.18 / cost)) * 100, 1) if cost > 0 else 100.0,
        })

    # Historical monthly trend from snapshots
    customer_id = _customer_id_for_org(membership)
    historical = _historical_monthly_spend_from_snapshots(db, customer_id, cloud_provider, months=6)

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "cloud_provider": cloud_provider,
        "summary": {
            "total_monthly_cost_usd": round(total, 2),
            "estimated_waste_usd": round(waste, 2),
            "identified_savings_usd": round(savings, 2),
            "waste_to_spend_percent": round((waste / total * 100) if total > 0 else 0.0, 1),
            "dollar_efficiency_score": unit_result.get("dollar_efficiency_score", 0),
        },
        "provider_metrics": provider_metrics,
        "historical_monthly_spend": historical,
        "business_metrics_hint": "POST /api/v1/analytics/unit-economics/metrics to calculate cost-per-unit.",
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Scorecards  — per-team FinOps maturity score using chargeback + efficiency
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ScorecardDimension(BaseModel):
    name: str
    score: float
    max_score: float
    description: str


class ScorecardEntry(BaseModel):
    team: str
    total_score: float
    grade: str
    cost_usd: float
    share_percent: float
    dimensions: List[ScorecardDimension]
    trend: str


class RealizedSavingsScorecardEntry(BaseModel):
    dimension: str
    key: str
    score: float
    grade: str
    recommendation_count: int
    verified_count: int
    open_count: int
    planned_monthly_savings_usd: float
    realized_monthly_savings_usd: float
    variance_monthly_usd: float
    planned_annual_savings_usd: float
    realized_annual_savings_usd: float
    variance_annual_usd: float
    realization_rate_percent: float
    last_realized_at: Optional[str] = None


class RealizedSavingsScorecards(BaseModel):
    total_planned_monthly_savings_usd: float
    total_realized_monthly_savings_usd: float
    total_variance_monthly_usd: float
    total_planned_annual_savings_usd: float
    total_realized_annual_savings_usd: float
    total_variance_annual_usd: float
    overall_realization_rate_percent: float
    overall_score: float
    overall_grade: str
    by_provider: List[RealizedSavingsScorecardEntry]
    by_owner: List[RealizedSavingsScorecardEntry]
    by_business_unit: List[RealizedSavingsScorecardEntry]
    by_month: List[RealizedSavingsScorecardEntry]


class ScorecardsResponse(BaseModel):
    generated_at: str
    organization_grade: str
    organization_score: float
    teams: List[ScorecardEntry]
    realized_savings: RealizedSavingsScorecards


def _grade(score: float, max_score: float = 100.0) -> str:
    pct = score / max_score * 100 if max_score > 0 else 0
    if pct >= 90: return "A+"
    if pct >= 80: return "A"
    if pct >= 70: return "B"
    if pct >= 55: return "C"
    return "D"


def _realized_savings_score(planned_monthly: float, realized_monthly: float) -> tuple[float, float, str]:
    planned = float(planned_monthly or 0.0)
    realized = float(realized_monthly or 0.0)
    rate = round((realized / planned) * 100.0, 1) if planned > 0 else (100.0 if realized > 0 else 0.0)
    score = round(min(max(rate, 0.0), 125.0) / 125.0 * 100.0, 1) if planned > 0 or realized > 0 else 0.0
    return score, rate, _grade(score)


def _ledger_evidence(row: RecommendationLedger) -> Dict[str, Any]:
    try:
        payload = json.loads(row.evidence_json or "{}")
        return payload if isinstance(payload, dict) else {}
    except (TypeError, json.JSONDecodeError):
        return {}


def _business_unit_lookup_for_org(db: Session, organization_id: int) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == organization_id)
        .all()
    )
    for row in rows:
        provider = (row.provider or "").strip().lower()
        if not provider:
            continue
        business_unit = row.cost_center or row.team or row.application or None
        if not business_unit:
            continue
        service = (row.service_name or "").strip().lower()
        region = (row.region or "").strip().lower()
        if service and region:
            lookup.setdefault(f"{provider}|{service}|{region}", business_unit)
        if service:
            lookup.setdefault(f"{provider}|{service}|", business_unit)
        lookup.setdefault(f"{provider}||", business_unit)
    return lookup


def _business_unit_for_ledger_row(
    row: RecommendationLedger,
    business_unit_lookup: Dict[str, str],
) -> str:
    evidence = _ledger_evidence(row)
    for key in ("business_unit", "cost_center", "team", "application"):
        value = evidence.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()

    provider = (row.provider or "").strip().lower()
    region = (row.region or "").strip().lower()
    labels = [
        (row.resource_name or "").strip().lower(),
        (row.resource_type or "").strip().lower(),
        (row.account_id or "").strip().lower(),
    ]
    for label in labels:
        if not label:
            continue
        if region:
            match = business_unit_lookup.get(f"{provider}|{label}|{region}")
            if match:
                return match
        match = business_unit_lookup.get(f"{provider}|{label}|")
        if match:
            return match

    return business_unit_lookup.get(f"{provider}||", "(unassigned)")


def _realized_savings_entry(
    dimension: str,
    key: str,
    rows: List[RecommendationLedger],
) -> RealizedSavingsScorecardEntry:
    planned_monthly = sum(float(row.planned_monthly_savings_usd or 0.0) for row in rows)
    realized_monthly = sum(float(row.realized_monthly_savings_usd or 0.0) for row in rows)
    variance_monthly = realized_monthly - planned_monthly
    score, rate, grade = _realized_savings_score(planned_monthly, realized_monthly)
    last_realized = max((row.realized_at for row in rows if row.realized_at), default=None)
    return RealizedSavingsScorecardEntry(
        dimension=dimension,
        key=key,
        score=score,
        grade=grade,
        recommendation_count=len(rows),
        verified_count=sum(1 for row in rows if (row.status or "").lower() == "verified"),
        open_count=sum(1 for row in rows if (row.status or "").lower() in {"open", "planned", "approved"}),
        planned_monthly_savings_usd=round(planned_monthly, 2),
        realized_monthly_savings_usd=round(realized_monthly, 2),
        variance_monthly_usd=round(variance_monthly, 2),
        planned_annual_savings_usd=round(planned_monthly * 12.0, 2),
        realized_annual_savings_usd=round(realized_monthly * 12.0, 2),
        variance_annual_usd=round(variance_monthly * 12.0, 2),
        realization_rate_percent=rate,
        last_realized_at=last_realized.isoformat() if last_realized else None,
    )


def _build_realized_savings_scorecards(
    *,
    db: Session,
    organization_id: int,
) -> RealizedSavingsScorecards:
    rows = (
        db.query(RecommendationLedger)
        .filter(RecommendationLedger.organization_id == organization_id)
        .all()
    )
    business_unit_lookup = _business_unit_lookup_for_org(db, organization_id)

    def grouped_entries(dimension: str, key_for_row) -> List[RealizedSavingsScorecardEntry]:
        grouped: Dict[str, List[RecommendationLedger]] = {}
        for row in rows:
            key = key_for_row(row) or "(unassigned)"
            grouped.setdefault(str(key), []).append(row)
        entries = [_realized_savings_entry(dimension, key, group_rows) for key, group_rows in grouped.items()]
        return sorted(
            entries,
            key=lambda entry: (
                -entry.realized_monthly_savings_usd,
                -entry.planned_monthly_savings_usd,
                entry.key,
            ),
        )

    planned_monthly = sum(float(row.planned_monthly_savings_usd or 0.0) for row in rows)
    realized_monthly = sum(float(row.realized_monthly_savings_usd or 0.0) for row in rows)
    variance_monthly = realized_monthly - planned_monthly
    overall_score, overall_rate, overall_grade = _realized_savings_score(planned_monthly, realized_monthly)

    return RealizedSavingsScorecards(
        total_planned_monthly_savings_usd=round(planned_monthly, 2),
        total_realized_monthly_savings_usd=round(realized_monthly, 2),
        total_variance_monthly_usd=round(variance_monthly, 2),
        total_planned_annual_savings_usd=round(planned_monthly * 12.0, 2),
        total_realized_annual_savings_usd=round(realized_monthly * 12.0, 2),
        total_variance_annual_usd=round(variance_monthly * 12.0, 2),
        overall_realization_rate_percent=overall_rate,
        overall_score=overall_score,
        overall_grade=overall_grade,
        by_provider=grouped_entries("provider", lambda row: (row.provider or "unknown").lower()),
        by_owner=grouped_entries("owner", lambda row: row.owner or "(unassigned)"),
        by_business_unit=grouped_entries(
            "business_unit",
            lambda row: _business_unit_for_ledger_row(row, business_unit_lookup),
        ),
        by_month=grouped_entries(
            "month",
            lambda row: row.realized_at.strftime("%Y-%m") if row.realized_at else "(unrealized)",
        ),
    )


@router.get("/analytics/scorecards", response_model=ScorecardsResponse)
async def get_scorecards(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Per-team FinOps scorecard: allocation coverage, waste rate, tagging, commitment."""
    org_id = _organization_id_for_membership(membership)

    # Pull chargeback dimensions
    dim_rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    # Group by team
    team_costs: Dict[str, float] = {}
    team_mapped: Dict[str, float] = {}
    for r in dim_rows:
        team = r.team or "(unmapped)"
        cost = float(r.cost_usd or 0.0)
        team_costs[team] = team_costs.get(team, 0.0) + cost
        if r.is_mapped:
            team_mapped[team] = team_mapped.get(team, 0.0) + cost

    total_cost = sum(team_costs.values())
    teams: List[Dict[str, Any]] = []

    if team_costs:
        for team, cost in sorted(team_costs.items(), key=lambda x: -x[1]):
            mapped = team_mapped.get(team, 0.0)
            allocation_score = round((mapped / cost * 40) if cost > 0 else 40.0, 1)
            waste_score = round(30.0 * (1 - min(0.3, (cost * 0.18) / max(cost, 1))), 1)
            tagging_score = round((mapped / cost * 20) if cost > 0 else 20.0, 1)
            commitment_score = 10.0 if cost > 500 else 5.0
            total_score = round(allocation_score + waste_score + tagging_score + commitment_score, 1)

            teams.append({
                "team": team,
                "total_score": total_score,
                "grade": _grade(total_score),
                "cost_usd": round(cost, 2),
                "share_percent": round((cost / total_cost * 100) if total_cost > 0 else 0.0, 1),
                "dimensions": [
                    {"name": "Allocation Coverage", "score": allocation_score, "max_score": 40, "description": "% of team cost mapped to a business dimension"},
                    {"name": "Waste Reduction", "score": waste_score, "max_score": 30, "description": "Estimated waste as share of team spend"},
                    {"name": "Tagging Hygiene", "score": tagging_score, "max_score": 20, "description": "Resources with complete cost-center tags"},
                    {"name": "Commitment Coverage", "score": commitment_score, "max_score": 10, "description": "Reserved/savings-plan coverage for team workloads"},
                ],
                "trend": "stable",
            })
    else:
        # No chargeback dimension data yet — return a single placeholder row so
        # the frontend can distinguish "no data" from "score = 0".
        teams.append({
            "team": "(no data)",
            "total_score": 0.0,
            "grade": "N/A",
            "cost_usd": 0.0,
            "share_percent": 100.0,
            "no_data": True,
            "dimensions": [
                {"name": "Allocation Coverage", "score": 0.0, "max_score": 40, "description": "Configure business mapping rules to start scoring"},
                {"name": "Waste Reduction", "score": 0.0, "max_score": 30, "description": "Run a provider scan to populate waste signals"},
                {"name": "Tagging Hygiene", "score": 0.0, "max_score": 20, "description": "Upload cost data or run a scan to score tagging hygiene"},
                {"name": "Commitment Coverage", "score": 0.0, "max_score": 10, "description": "Add Reserved Instances or Savings Plans to improve coverage"},
            ],
            "trend": "unknown",
        })

    org_score = round(sum(t["total_score"] for t in teams) / max(len(teams), 1), 1)
    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "organization_grade": _grade(org_score),
        "organization_score": org_score,
        "teams": teams,
        "realized_savings": _build_realized_savings_scorecards(db=db, organization_id=org_id),
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Resource Inventory  — surface cloud resources with cost attribution
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ResourceInventoryItem(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    account_id: str
    cost_usd: float
    waste_flag: bool
    waste_reason: Optional[str]
    tags: Dict[str, str]
    data_source: str = "account_cost_snapshot"
    console_url: Optional[str] = None


class ResourceInventoryResponse(BaseModel):
    generated_at: str
    total_resources: int
    total_cost_usd: float
    flagged_waste_count: int
    items: List[ResourceInventoryItem]
    data_source: str = "account_cost_snapshot"
    coverage_note: str = ""


def _inventory_string(value: Any, fallback: str = "") -> str:
    text = str(value or "").strip()
    return text or fallback


def _inventory_item_from_provider_row(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Convert provider-native recommendation/action rows into inventory rows."""
    source = _inventory_string(row.get("source"))
    if source not in {
        "oci_optimizer_resource_action",
        "oci_compute_inventory",
        "oci_storage_inventory",
    }:
        return None

    resource_id = _inventory_string(row.get("resource_id"))
    if not resource_id:
        return None

    resource_type = _inventory_string(row.get("resource_type"), _inventory_string(row.get("service"), "OCI Resource"))
    resource_name = _inventory_string(
        row.get("resource_name"),
        _inventory_string(row.get("description"), resource_id),
    )
    monthly = _safe_recommendation_float(row.get("current_annual_spend")) / 12.0
    if monthly <= 0:
        monthly = _safe_recommendation_float(row.get("monthly_savings_usd"))
    recommendation = _inventory_string(
        row.get("recommendation_type"),
        _inventory_string(row.get("description"), "Provider optimization finding"),
    )
    tags = {
        "source": source,
        "recommendation": recommendation,
    }
    for key in ("category", "importance", "status", "recommendation_status"):
        value = _inventory_string(row.get(key))
        if value:
            tags[key] = value

    return {
        "resource_id": resource_id,
        "resource_name": resource_name,
        "resource_type": resource_type,
        "provider": _inventory_string(row.get("provider"), "oci"),
        "region": _inventory_string(row.get("region"), "global"),
        "account_id": _inventory_string(row.get("account_id")),
        "cost_usd": round(monthly, 2),
        "waste_flag": True,
        "waste_reason": recommendation,
        "tags": tags,
        "data_source": source,
        "console_url": _inventory_string(row.get("resource_console_url")) or None,
    }


def _provider_resource_inventory_items(
    *,
    db: Session,
    customer_id: str,
    provider: str,
    region: Optional[str],
    limit: int,
    offset: int,
) -> List[Dict[str, Any]]:
    """Return real provider resource rows from optimizer/live inventory feeds."""
    if provider not in {"all", "oci"}:
        return []

    rows = _collect_provider_recommendation_rows(
        db=db,
        customer_id=customer_id,
        provider="oci",
        min_monthly_savings=0.0,
        limit=provider_bounded_limit("oci", offset + limit, floor=0),
        include_existing_rightsizing_sources=True,
    )
    items: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        item = _inventory_item_from_provider_row(row)
        if not item:
            continue
        if region and item["region"] != region:
            continue
        key = f"{item['provider']}:{item['resource_id']}"
        if key in seen:
            continue
        seen.add(key)
        items.append(item)
    return items


@router.get("/inventory/resources", response_model=ResourceInventoryResponse)
async def get_resource_inventory(
    provider: str = "all",
    region: Optional[str] = None,
    waste_only: bool = False,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Cloud resource inventory with per-resource cost attribution and waste flags."""
    customer_id = _customer_id_for_org(membership)
    requested_limit = max(1, min(int(limit or 100), 1000))
    requested_offset = max(0, int(offset or 0))

    live_items = _provider_resource_inventory_items(
        db=db,
        customer_id=customer_id,
        provider=provider,
        region=region,
        limit=requested_limit,
        offset=requested_offset,
    )

    # Pull account snapshots for resource-level data
    snapshots_q = (
        db.query(ProviderAccountSnapshot)
        .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
        .filter(ProviderAccount.customer_id == customer_id)
    )
    if provider != "all":
        snapshots_q = snapshots_q.filter(ProviderAccount.provider == provider)

    org_id = membership.organization_id
    snapshots = snapshots_q.order_by(ProviderAccountSnapshot.captured_at.desc()).limit(500).all()

    account_scope_items: List[Dict[str, Any]] = []
    seen: set = set()

    for snap in snapshots:
        acct = snap.provider_account
        if not acct:
            continue
        key = f"{acct.provider}:{acct.account_identifier}"
        if key in seen:
            continue
        seen.add(key)

        direct_cost = float(snap.direct_cost_usd or 0.0)
        waste_flag = direct_cost > 0 and (snap.anomalies_count or 0) > 0

        # region filter — ProviderAccount stores region in native_region
        snap_region = acct.native_region or "global"
        if region and snap_region != region:
            continue

        item: Dict[str, Any] = {
            "resource_id": acct.account_identifier or key,
            "resource_name": acct.account_name or acct.account_identifier or acct.provider.upper(),
            "resource_type": acct.account_type or "cloud-account",
            "provider": acct.provider,
            "region": snap_region,
            "account_id": acct.account_identifier or "",
            "cost_usd": round(direct_cost, 2),
            "waste_flag": waste_flag,
            "waste_reason": "Active anomalies detected" if waste_flag else None,
            "tags": {},
            "data_source": "account_cost_snapshot",
            "console_url": None,
        }
        account_scope_items.append(item)

    items: List[Dict[str, Any]]
    data_source = "live_provider_resource_actions" if live_items else "account_cost_snapshot"
    coverage_note = (
        "Showing real OCI tenancy-level Optimizer resource actions and live inventory candidates. "
        "Use Optimization Advisor for execution detail."
        if live_items
        else "Only account or tenancy-level cost snapshots are available for this scope; connect live provider inventory to show individual resources."
    )

    if live_items:
        items = live_items
    else:
        items = account_scope_items

    # Supplement from imported cost records when no live resource or scan snapshots exist
    if not items:
        imported_rows = _get_imported_cost_rows(db, org_id, customer_id)
        for row in imported_rows:
            prov = row.provider or "unknown"
            if provider != "all" and prov != provider:
                continue
            reg = row.region or "global"
            if region and reg != region:
                continue
            items.append({
                "resource_id": row.account_identifier or f"{prov}-imported-{row.id}",
                "resource_name": row.account_name or row.service_name or prov.upper(),
                "resource_type": "imported-cost-record",
                "provider": prov,
                "region": reg,
                "account_id": row.account_identifier or "",
                "cost_usd": round(float(row.cost_usd or 0.0), 2),
                "waste_flag": False,
                "waste_reason": None,
                "tags": {},
                "data_source": "csv_import",
                "console_url": None,
            })
        if items:
            data_source = "csv_import"
            coverage_note = "Showing imported billing rows because no live resource inventory is available for this scope."

    if waste_only:
        items = [i for i in items if i["waste_flag"]]

    items.sort(key=lambda i: float(i.get("cost_usd") or 0.0), reverse=True)
    total_cost = round(sum(i["cost_usd"] for i in items), 2)
    flagged = sum(1 for i in items if i["waste_flag"])
    paginated = items[requested_offset: requested_offset + requested_limit]

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "total_resources": len(items),
        "total_cost_usd": total_cost,
        "flagged_waste_count": flagged,
        "items": paginated,
        "data_source": data_source,
        "coverage_note": coverage_note,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Kubernetes Cost Allocation
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class KubernetesClusterInput(BaseModel):
    cluster_name: str
    provider: str
    region: str
    node_count: int
    node_type: str
    monthly_node_cost_usd: float
    namespaces: Optional[List[str]] = None
    opencost_enabled: bool = False
    opencost_url: Optional[str] = None
    opencost_window_days: int = Field(default=7, ge=1, le=30)
    workloads: List["KubernetesWorkloadInput"] = Field(default_factory=list)
    node_pools: List["KubernetesNodePoolInput"] = Field(default_factory=list)


class KubernetesWorkloadInput(BaseModel):
    namespace: str
    workload_name: str
    team: Optional[str] = None
    node_pool: Optional[str] = None
    replicas: int = Field(default=1, ge=1)
    cpu_request_cores: float = Field(default=0.5, ge=0)
    cpu_limit_cores: float = Field(default=1.0, ge=0)
    memory_request_gib: float = Field(default=1.0, ge=0)
    memory_limit_gib: float = Field(default=2.0, ge=0)
    cpu_usage_cores: float = Field(default=0.25, ge=0)
    memory_usage_gib: float = Field(default=0.5, ge=0)


class KubernetesNodePoolInput(BaseModel):
    name: str
    node_count: int = Field(default=1, ge=1)
    monthly_node_cost_usd: float = Field(default=150.0, ge=0)
    cpu_capacity_cores: float = Field(default=4.0, ge=0)
    memory_capacity_gib: float = Field(default=16.0, ge=0)


class KubernetesNamespaceCost(BaseModel):
    namespace: str
    estimated_cost_usd: float
    share_percent: float
    cpu_share_percent: float
    memory_share_percent: float


class KubernetesWorkloadCost(BaseModel):
    namespace: str
    workload_name: str
    team: str
    node_pool: str
    estimated_cost_usd: float
    share_percent: float
    cpu_request_cores: float
    cpu_usage_cores: float
    memory_request_gib: float
    memory_usage_gib: float
    request_efficiency_percent: float


class KubernetesTeamCost(BaseModel):
    team: str
    estimated_cost_usd: float
    share_percent: float
    namespaces: List[str]
    workload_count: int


class KubernetesNodePoolCost(BaseModel):
    node_pool: str
    node_count: int
    estimated_cost_usd: float
    utilization_percent: float
    idle_cost_usd: float


class KubernetesOptimizationRecommendation(BaseModel):
    recommendation_id: str
    category: Literal["workload", "node_pool", "request_limit"]
    target: str
    severity: Literal["low", "medium", "high"]
    estimated_monthly_savings_usd: float
    rationale: str
    action: str


class KubernetesClusterCostResponse(BaseModel):
    generated_at: str
    cluster_name: str
    provider: str
    region: str
    node_count: int
    node_type: str
    total_cluster_cost_usd: float
    cost_per_node_usd: float
    namespace_breakdown: List[KubernetesNamespaceCost]
    workload_breakdown: List[KubernetesWorkloadCost] = Field(default_factory=list)
    team_breakdown: List[KubernetesTeamCost] = Field(default_factory=list)
    node_pool_breakdown: List[KubernetesNodePoolCost] = Field(default_factory=list)
    recommendations: List[KubernetesOptimizationRecommendation] = Field(default_factory=list)
    efficiency_note: str
    opencost_integration: str


class KubernetesProviderNodeType(BaseModel):
    value: str
    monthly_cost_usd: float
    vcpu: Optional[float] = None
    memory_gib: Optional[float] = None
    source: str = "live"


class KubernetesProviderCatalogEntry(BaseModel):
    provider: str
    source: str
    configured: bool
    regions: List[str]
    node_types: List[KubernetesProviderNodeType]
    message: str


class KubernetesProviderCatalogResponse(BaseModel):
    generated_at: str
    providers: Dict[str, KubernetesProviderCatalogEntry]


class KubernetesContainerServiceCost(BaseModel):
    provider: str
    service: str
    category: Literal["managed_kubernetes", "container_runtime", "container_registry", "docker", "container_platform"]
    monthly_cost_usd: float
    share_percent: float = 0.0
    source: str
    evidence: str
    account_count: int = 0
    region_count: int = 0
    regions: List[str] = Field(default_factory=list)
    resource_id: Optional[str] = None
    resource_name: Optional[str] = None
    lifecycle_state: Optional[str] = None
    resource_shape: Optional[str] = None
    resource_version: Optional[str] = None
    created_at: Optional[str] = None
    availability_domain: Optional[str] = None
    public_endpoint: Optional[str] = None
    private_endpoint: Optional[str] = None
    public_ip: Optional[str] = None
    ocpus: Optional[float] = None
    memory_gib: Optional[float] = None
    container_count: Optional[int] = None
    container_images: List[str] = Field(default_factory=list)
    console_url: Optional[str] = None


class KubernetesProviderServiceRollup(BaseModel):
    provider: str
    configured: bool
    source: str
    total_monthly_cost_usd: float
    share_percent: float = 0.0
    service_count: int
    services: List[KubernetesContainerServiceCost]


class KubernetesSummaryResponse(BaseModel):
    generated_at: str
    kubernetes_enabled: bool
    clusters_configured: int
    estimated_k8s_share_percent: float
    estimated_k8s_cost_usd: float
    total_cloud_cost_usd: float
    container_service_count: int = 0
    provider_count_with_container_spend: int = 0
    highest_cost_provider: Optional[str] = None
    highest_cost_service: Optional[KubernetesContainerServiceCost] = None
    container_services: List[KubernetesContainerServiceCost] = Field(default_factory=list)
    provider_breakdown: List[KubernetesProviderServiceRollup] = Field(default_factory=list)
    data_source: str = "none"
    setup_hint: str
    opencost_docs: str


class OpenCostSyncRequest(BaseModel):
    api_url: str
    cluster_name: str
    window_days: int = Field(default=7, ge=1, le=30)


class OpenCostNamespaceCost(BaseModel):
    namespace: str
    cost_usd: float
    share_percent: float


class OpenCostPodCost(BaseModel):
    namespace: str
    pod_name: str
    cost_usd: float
    share_percent: float


class OpenCostSyncResponse(BaseModel):
    generated_at: str
    cluster_name: str
    source: str
    window_days: int
    total_cost_usd: float
    namespace_count: int
    namespaces: List[OpenCostNamespaceCost]
    pods: List[OpenCostPodCost] = Field(default_factory=list)


class OpenCostInstallRequest(BaseModel):
    kube_context: Optional[str] = None
    namespace: str = "opencost"
    prometheus_namespace: str = "prometheus-system"
    skip_prometheus_install: bool = False
    expose_port: int = 9003


class OpenCostInstallResponse(BaseModel):
    generated_at: str
    status: Literal["installed", "already_installed", "failed"]
    message: str
    api_url: Optional[str] = None
    namespace: str
    prometheus_namespace: str
    command_log: List[str] = Field(default_factory=list)


_KUBERNETES_CONTAINER_SERVICE_TERMS: Dict[str, List[str]] = {
    "managed_kubernetes": [
        "kubernetes",
        "k8s",
        "elastic kubernetes service",
        "eks",
        "azure kubernetes service",
        "aks",
        "google kubernetes engine",
        "gke",
        "container engine for kubernetes",
        "oke",
    ],
    "container_runtime": [
        "elastic container service",
        "ecs",
        "fargate",
        "container instances",
        "container instance",
        "container apps",
        "cloud run",
        "app runner",
        "container engine",
        "container service",
    ],
    "container_registry": [
        "container registry",
        "container repositories",
        "elastic container registry",
        "ecr",
        "artifact registry",
        "registry",
        "image registry",
    ],
    "docker": [
        "docker",
        "docker hub",
        "docker desktop",
    ],
}


def _kubernetes_container_service_category(service_name: str) -> Optional[str]:
    text = str(service_name or "").strip().lower()
    if not text:
        return None
    for category, terms in _KUBERNETES_CONTAINER_SERVICE_TERMS.items():
        if any(term in text for term in terms):
            return category
    if "container" in text:
        return "container_platform"
    return None


def _container_service_evidence(category: str) -> str:
    labels = {
        "managed_kubernetes": "Managed Kubernetes cluster service",
        "container_runtime": "Container runtime or serverless container service",
        "container_registry": "Container image registry service",
        "docker": "Docker service or Docker-related subscription",
        "container_platform": "Container platform service",
    }
    return labels.get(category, "Container platform service")


def _container_source_rank(source: str) -> int:
    return {
        "live_provider_api": 0,
        "live_resource_inventory": 1,
        "cost_snapshots_live": 2,
        "csv_import": 3,
    }.get(source, 9)


def _estimate_oci_container_instance_monthly_cost(
    ocpus: Optional[float],
    memory_gib: Optional[float],
) -> float:
    """Conservative run-rate estimate for OCI Container Instances before billing lands."""
    ocpu_hours = max(float(ocpus or 0.0), 0.0) * 0.0255 * 730.0
    memory_hours = max(float(memory_gib or 0.0), 0.0) * 0.0015 * 730.0
    return round(ocpu_hours + memory_hours, 2)


def _oci_live_kubernetes_inventory_rows(
    cred_json: Dict[str, Any],
    *,
    limit: int = 120,
    time_budget_seconds: float = 18.0,
) -> List[Dict[str, Any]]:
    """Return live OCI OKE, Container Instance, and OCIR resources for the K8s page."""
    deadline = time.monotonic() + max(float(time_budget_seconds or 0.0), 1.0)

    def budget_exhausted() -> bool:
        return time.monotonic() >= deadline

    try:
        import oci
    except Exception as exc:
        logger.info("OCI SDK unavailable for Kubernetes inventory: %s", exc)
        return []

    config_file = str(cred_json.get("config_file") or "").strip()
    profile = _normalize_oci_profile(cred_json.get("profile"))
    if not config_file:
        return []
    try:
        resolved_config, resolved_profile = CredentialValidator._normalize_oci_inputs(
            config_file=config_file,
            profile=profile,
        )
        oci_config = oci.config.from_file(resolved_config, resolved_profile)
    except Exception as exc:
        logger.info("Unable to load OCI config for Kubernetes inventory: %s", exc)
        return []

    cfg = Config()
    tenancy_id = str(oci_config.get("tenancy") or "").strip()
    if not tenancy_id:
        return []

    client_timeout = (3, 5)
    configured_region = str(
        cred_json.get("region")
        or cfg.oci_region
        or oci_config.get("region")
        or "uk-london-1"
    ).strip()
    if configured_region:
        oci_config = _oci_config_for_region(oci_config, configured_region)
    home_region = _oci_home_region(oci, oci_config, tenancy_id, timeout=client_timeout) or configured_region
    home_config = _oci_config_for_region(oci_config, home_region)
    try:
        identity = oci.identity.IdentityClient(home_config, timeout=client_timeout)
    except Exception as exc:
        logger.info("Unable to initialize OCI identity client for Kubernetes inventory: %s", exc)
        return []

    compartment_ids = _oci_accessible_compartment_ids(
        oci,
        identity,
        tenancy_id,
        seed_compartment_ids=_oci_scan_compartment_seeds(cred_json, cfg),
    )
    scan_regions = _oci_subscribed_regions(
        oci,
        home_config,
        tenancy_id,
        home_region=home_region,
        timeout=client_timeout,
    )
    rows: List[Dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    def append_row(
        *,
        region: str,
        resource_id: str,
        service: str,
        category: str,
        monthly_cost_usd: float,
        evidence: str,
        compartment_id: str,
        resource_name: Optional[str] = None,
        lifecycle_state: Optional[str] = None,
        resource_shape: Optional[str] = None,
        resource_version: Optional[str] = None,
        created_at: Optional[str] = None,
        availability_domain: Optional[str] = None,
        public_endpoint: Optional[str] = None,
        private_endpoint: Optional[str] = None,
        public_ip: Optional[str] = None,
        ocpus: Optional[float] = None,
        memory_gib: Optional[float] = None,
        container_count: Optional[int] = None,
        container_images: Optional[List[str]] = None,
    ) -> None:
        if len(rows) >= limit:
            return
        key = (category, resource_id)
        if not resource_id or key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "provider": "oci",
                "resource_id": resource_id,
                "service": service,
                "category": category,
                "monthly_cost_usd": round(float(monthly_cost_usd or 0.0), 2),
                "source": "live_resource_inventory",
                "region": region,
                "account_identifier": compartment_id,
                "evidence": evidence,
                "resource_name": resource_name,
                "lifecycle_state": lifecycle_state,
                "resource_shape": resource_shape,
                "resource_version": resource_version,
                "created_at": created_at,
                "availability_domain": availability_domain,
                "public_endpoint": public_endpoint,
                "private_endpoint": private_endpoint,
                "public_ip": public_ip,
                "ocpus": ocpus,
                "memory_gib": memory_gib,
                "container_count": container_count,
                "container_images": container_images or [],
                "console_url": _rightsizing_console_url(
                    provider="oci",
                    resource_id=resource_id,
                    region=region,
                    account_id=tenancy_id,
                    resource_type=category,
                ),
            }
        )

    for region in scan_regions:
        if len(rows) >= limit or budget_exhausted():
            break
        try:
            region_config = _oci_config_for_region(home_config, region)
            container_engine = oci.container_engine.ContainerEngineClient(region_config, timeout=client_timeout)
            container_instances = oci.container_instances.ContainerInstanceClient(region_config, timeout=client_timeout)
            artifacts = oci.artifacts.ArtifactsClient(region_config, timeout=client_timeout)
            virtual_network = oci.core.VirtualNetworkClient(region_config, timeout=client_timeout)
        except Exception as exc:
            logger.info("Unable to initialize OCI Kubernetes inventory clients in region %s: %s", region, exc)
            continue

        for compartment_id in compartment_ids:
            if len(rows) >= limit or budget_exhausted():
                break
            try:
                clusters_response = oci.pagination.list_call_get_all_results(
                    container_engine.list_clusters,
                    compartment_id=compartment_id,
                    limit=50,
                )
            except Exception as exc:
                logger.info("Unable to list OCI OKE clusters in compartment %s region %s: %s", compartment_id, region, exc)
                clusters_response = []
            for cluster in _oci_collection_items(clusters_response):
                if budget_exhausted():
                    break
                lifecycle = str(getattr(cluster, "lifecycle_state", "") or "").upper()
                if lifecycle in {"DELETED", "DELETING", "FAILED"}:
                    continue
                cluster_id = str(getattr(cluster, "id", "") or "").strip()
                name = str(getattr(cluster, "name", "") or getattr(cluster, "display_name", "") or cluster_id).strip()
                version = str(getattr(cluster, "kubernetes_version", "") or "").strip()
                cluster_type = str(getattr(cluster, "type", "") or "BASIC_CLUSTER").strip()
                public_endpoint = ""
                private_endpoint = ""
                created_at = ""
                try:
                    cluster_details = container_engine.get_cluster(cluster_id).data
                    endpoints = getattr(cluster_details, "endpoints", None)
                    public_endpoint = str(getattr(endpoints, "public_endpoint", "") or "").strip() if endpoints else ""
                    private_endpoint = str(getattr(endpoints, "private_endpoint", "") or "").strip() if endpoints else ""
                    metadata = getattr(cluster_details, "metadata", None)
                    created_value = getattr(metadata, "time_created", None) if metadata else None
                    created_at = created_value.isoformat() if hasattr(created_value, "isoformat") else str(created_value or "")
                except Exception:
                    pass
                append_row(
                    region=region,
                    resource_id=cluster_id,
                    service=f"OKE cluster: {name}",
                    category="managed_kubernetes",
                    monthly_cost_usd=0.0,
                    evidence=(
                        f"Live OCI OKE inventory: {lifecycle or 'ACTIVE'}"
                        f"{f', Kubernetes {version}' if version else ''}"
                        f", {cluster_type.replace('_', ' ').title()}."
                    ),
                    compartment_id=compartment_id,
                    resource_name=name,
                    lifecycle_state=lifecycle or None,
                    resource_shape=cluster_type.replace("_", " ").title(),
                    resource_version=version or None,
                    created_at=created_at or None,
                    public_endpoint=public_endpoint or None,
                    private_endpoint=private_endpoint or None,
                )

            try:
                instances_response = oci.pagination.list_call_get_all_results(
                    container_instances.list_container_instances,
                    compartment_id=compartment_id,
                    limit=50,
                )
            except Exception as exc:
                logger.info("Unable to list OCI Container Instances in compartment %s region %s: %s", compartment_id, region, exc)
                instances_response = []
            for instance in _oci_collection_items(instances_response):
                if budget_exhausted():
                    break
                lifecycle = str(getattr(instance, "lifecycle_state", "") or "").upper()
                if lifecycle in {"DELETED", "DELETING", "FAILED"}:
                    continue
                instance_id = str(getattr(instance, "id", "") or "").strip()
                display_name = str(getattr(instance, "display_name", "") or instance_id).strip()
                shape = str(getattr(instance, "shape", "") or "OCI Container Instance").strip()
                availability_domain = str(getattr(instance, "availability_domain", "") or "").strip()
                created_value = getattr(instance, "time_created", None)
                created_at = created_value.isoformat() if hasattr(created_value, "isoformat") else str(created_value or "")
                shape_config = getattr(instance, "shape_config", None)
                ocpus = float(getattr(shape_config, "ocpus", 0.0) or 0.0) if shape_config else None
                memory_gib = float(getattr(shape_config, "memory_in_gbs", 0.0) or 0.0) if shape_config else None
                images: List[str] = []
                container_count = int(getattr(instance, "container_count", 0) or 0)
                public_ip = ""
                try:
                    details = container_instances.get_container_instance(instance_id).data
                    for container in getattr(details, "containers", []) or []:
                        image = str(getattr(container, "image_url", "") or "").strip()
                        container_id = str(getattr(container, "container_id", "") or getattr(container, "id", "") or "").strip()
                        if not image and container_id:
                            try:
                                container_details = container_instances.get_container(container_id).data
                                image = str(getattr(container_details, "image_url", "") or "").strip()
                            except Exception:
                                image = ""
                        if image and image not in images:
                            images.append(image)
                    for vnic in getattr(details, "vnics", []) or []:
                        vnic_id = str(getattr(vnic, "vnic_id", "") or getattr(vnic, "id", "") or "").strip()
                        if not vnic_id:
                            continue
                        try:
                            vnic_details = virtual_network.get_vnic(vnic_id).data
                            public_ip = str(getattr(vnic_details, "public_ip", "") or "").strip()
                            if public_ip:
                                break
                        except Exception:
                            continue
                except Exception:
                    images = []
                image_note = f" Images: {', '.join(images[:3])}." if images else ""
                size_note = ""
                if ocpus or memory_gib:
                    size_note = f" {ocpus or 0:g} OCPU / {memory_gib or 0:g} GiB."
                append_row(
                    region=region,
                    resource_id=instance_id,
                    service=f"OCI Container Instance: {display_name}",
                    category="container_runtime",
                    monthly_cost_usd=_estimate_oci_container_instance_monthly_cost(ocpus, memory_gib),
                    evidence=f"Live OCI Container Instance inventory: {lifecycle or 'ACTIVE'}, {shape}.{size_note}{image_note}",
                    compartment_id=compartment_id,
                    resource_name=display_name,
                    lifecycle_state=lifecycle or None,
                    resource_shape=shape,
                    created_at=created_at or None,
                    availability_domain=availability_domain or None,
                    public_ip=public_ip or None,
                    ocpus=ocpus,
                    memory_gib=memory_gib,
                    container_count=container_count or (len(images) if images else None),
                    container_images=images[:8],
                )

            try:
                repositories_response = oci.pagination.list_call_get_all_results(
                    artifacts.list_container_repositories,
                    compartment_id=compartment_id,
                    limit=50,
                )
            except Exception as exc:
                logger.info("Unable to list OCI container repositories in compartment %s region %s: %s", compartment_id, region, exc)
                repositories_response = []
            for repository in _oci_collection_items(repositories_response):
                if budget_exhausted():
                    break
                lifecycle = str(getattr(repository, "lifecycle_state", "") or "").upper()
                if lifecycle in {"DELETED", "DELETING", "FAILED"}:
                    continue
                repository_id = str(getattr(repository, "id", "") or "").strip()
                name = str(
                    getattr(repository, "display_name", "")
                    or getattr(repository, "repository_name", "")
                    or repository_id
                ).strip()
                append_row(
                    region=region,
                    resource_id=repository_id,
                    service=f"OCIR repository: {name}",
                    category="container_registry",
                    monthly_cost_usd=0.0,
                    evidence=f"Live OCI Container Registry inventory: {lifecycle or 'AVAILABLE'} repository.",
                    compartment_id=compartment_id,
                    resource_name=name,
                    lifecycle_state=lifecycle or None,
                )
    return rows


def _collect_oci_live_kubernetes_inventory_rows(
    customer_id: str,
    db: Session,
    *,
    limit: int = 120,
) -> List[Dict[str, Any]]:
    credential_payloads: List[Dict[str, Any]] = []
    runtime_credentials = _load_runtime_provider_credentials(customer_id).get("oci")
    if isinstance(runtime_credentials, dict):
        credential_payloads.append(runtime_credentials)

    stored_rows = (
        db.query(CredentialRecord)
        .filter(
            CredentialRecord.customer_id == customer_id,
            CredentialRecord.provider == "oci",
            CredentialRecord.is_valid.is_(True),
        )
        .all()
    )
    for row in stored_rows:
        try:
            payload = json.loads(row.credential_json)
        except Exception:
            continue
        if isinstance(payload, dict):
            credential_payloads.append(payload)

    runtime_oci = _runtime_oci_credential_json()
    if runtime_oci is not None:
        credential_payloads.append(runtime_oci)

    rows: List[Dict[str, Any]] = []
    seen_credential_keys: set[tuple[str, str, str]] = set()
    for payload in credential_payloads:
        key = (
            str(payload.get("config_file") or ""),
            _normalize_oci_profile(payload.get("profile")),
            str(payload.get("region") or ""),
        )
        if key in seen_credential_keys:
            continue
        seen_credential_keys.add(key)
        for row in _oci_live_kubernetes_inventory_rows(payload, limit=max(limit - len(rows), 0)):
            resource_key = (
                str(row.get("provider") or "").lower(),
                str(row.get("category") or "").lower(),
                str(row.get("resource_id") or row.get("service") or "").lower(),
            )
            if resource_key in {
                (
                    str(existing.get("provider") or "").lower(),
                    str(existing.get("category") or "").lower(),
                    str(existing.get("resource_id") or existing.get("service") or "").lower(),
                )
                for existing in rows
            }:
                continue
            rows.append(row)
        if len(rows) >= limit:
            break
    return rows[:limit]


async def _build_kubernetes_container_service_rollups(
    membership: UserOrganization,
    db: Session,
    total_cloud_cost_usd: float,
) -> tuple[List[KubernetesContainerServiceCost], List[KubernetesProviderServiceRollup], str]:
    """Collect Kubernetes/container/Docker service spend across providers."""
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)
    diagnostics = {
        item.provider: bool(item.configured)
        for item in _provider_diagnostics(customer_id=customer_id, db=db)
    }
    providers = ["aws", "azure", "gcp", "oci"]
    grouped: Dict[tuple[str, str], Dict[str, Any]] = {}
    provider_sources: Dict[str, str] = {provider: "none" for provider in providers}

    def add_row(
        provider: str,
        service_name: str,
        cost: float,
        source: str,
        region: Optional[str] = None,
        account_identifier: Optional[str] = None,
        category_override: Optional[str] = None,
        evidence: Optional[str] = None,
        allow_zero_cost: bool = False,
        resource_id: Optional[str] = None,
        resource_name: Optional[str] = None,
        lifecycle_state: Optional[str] = None,
        resource_shape: Optional[str] = None,
        resource_version: Optional[str] = None,
        created_at: Optional[str] = None,
        availability_domain: Optional[str] = None,
        public_endpoint: Optional[str] = None,
        private_endpoint: Optional[str] = None,
        public_ip: Optional[str] = None,
        ocpus: Optional[float] = None,
        memory_gib: Optional[float] = None,
        container_count: Optional[int] = None,
        container_images: Optional[List[str]] = None,
        console_url: Optional[str] = None,
    ) -> None:
        provider_key = str(provider or "").strip().lower()
        if provider_key not in providers:
            return
        service = str(service_name or "").strip() or "Container service"
        category = category_override or _kubernetes_container_service_category(service)
        if category is None:
            return
        amount = float(cost or 0.0)
        if amount <= 0 and not allow_zero_cost:
            return
        key = (provider_key, service.lower())
        bucket = grouped.setdefault(
            key,
            {
                "provider": provider_key,
                "service": service,
                "category": category,
                "monthly_cost_usd": 0.0,
                "source": source,
                "evidence": evidence or _container_service_evidence(category),
                "regions": set(),
                "accounts": set(),
                "resource_id": resource_id,
                "resource_name": resource_name,
                "lifecycle_state": lifecycle_state,
                "resource_shape": resource_shape,
                "resource_version": resource_version,
                "created_at": created_at,
                "availability_domain": availability_domain,
                "public_endpoint": public_endpoint,
                "private_endpoint": private_endpoint,
                "public_ip": public_ip,
                "ocpus": ocpus,
                "memory_gib": memory_gib,
                "container_count": container_count,
                "container_images": container_images or [],
                "console_url": console_url,
            },
        )
        bucket["monthly_cost_usd"] = float(bucket["monthly_cost_usd"]) + amount
        if _container_source_rank(source) < _container_source_rank(str(bucket.get("source") or "")):
            bucket["source"] = source
        if evidence and str(bucket.get("evidence") or "") == _container_service_evidence(str(bucket.get("category") or category)):
            bucket["evidence"] = evidence
        if region:
            bucket["regions"].add(str(region))
        if account_identifier:
            bucket["accounts"].add(str(account_identifier))
        for field_name, value in {
            "resource_id": resource_id,
            "resource_name": resource_name,
            "lifecycle_state": lifecycle_state,
            "resource_shape": resource_shape,
            "resource_version": resource_version,
            "created_at": created_at,
            "availability_domain": availability_domain,
            "public_endpoint": public_endpoint,
            "private_endpoint": private_endpoint,
            "public_ip": public_ip,
            "ocpus": ocpus,
            "memory_gib": memory_gib,
            "container_count": container_count,
            "console_url": console_url,
        }.items():
            if value is not None and not bucket.get(field_name):
                bucket[field_name] = value
        if container_images and not bucket.get("container_images"):
            bucket["container_images"] = container_images

    for provider in providers:
        if diagnostics.get(provider, False):
            try:
                try:
                    summary = await asyncio.wait_for(
                        _cost_summary_for_provider(provider, "month", customer_id=customer_id),
                        timeout=6.0,
                    )
                except TypeError as exc:
                    if "customer_id" not in str(exc) and "keyword" not in str(exc):
                        raise
                    summary = await asyncio.wait_for(
                        _cost_summary_for_provider(provider, "month"),
                        timeout=6.0,
                    )
                except asyncio.TimeoutError:
                    logger.info("Timed out collecting live container service costs for %s", provider)
                    summary = {"error": "live provider cost summary timed out"}
                if "error" not in summary:
                    before_count = len(grouped)
                    for service in summary.get("top_services", []) or []:
                        if not isinstance(service, dict):
                            continue
                        add_row(
                            provider=provider,
                            service_name=str(
                                service.get("service")
                                or service.get("service_name")
                                or service.get("name")
                                or ""
                            ),
                            cost=float(service.get("cost_usd") or service.get("cost") or service.get("amount") or 0.0),
                            source="live_provider_api",
                        )
                    if len(grouped) > before_count:
                        provider_sources[provider] = "live_provider_api"
                        continue
            except Exception as exc:
                logger.info("Unable to collect live container services for %s: %s", provider, exc)

        latest_snapshot = (
            db.query(CostSnapshot)
            .filter(
                CostSnapshot.customer_id == customer_id,
                CostSnapshot.provider == provider,
            )
            .order_by(CostSnapshot.captured_at.desc())
            .first()
        )
        if latest_snapshot is not None:
            before_count = len(grouped)
            try:
                snapshot_services = json.loads(latest_snapshot.top_services_json or "[]")
            except Exception:
                snapshot_services = []
            for service in snapshot_services if isinstance(snapshot_services, list) else []:
                if not isinstance(service, dict):
                    continue
                add_row(
                    provider=provider,
                    service_name=str(
                        service.get("service")
                        or service.get("service_name")
                        or service.get("name")
                        or ""
                    ),
                    cost=float(service.get("cost_usd") or service.get("cost") or service.get("amount") or 0.0),
                    source="cost_snapshots_live",
                )
            if len(grouped) > before_count and provider_sources[provider] == "none":
                provider_sources[provider] = "cost_snapshots_live"
                continue

        imported_rows = _get_imported_cost_rows(db, organization_id, customer_id, provider)
        before_count = len(grouped)
        for row in imported_rows:
            add_row(
                provider=str(row.provider or provider),
                service_name=str(row.service_name or ""),
                cost=float(row.cost_usd or 0.0),
                source="csv_import",
                region=row.region,
                account_identifier=row.account_identifier,
            )
        if len(grouped) > before_count and provider_sources[provider] == "none":
            provider_sources[provider] = "csv_import"

    try:
        before_count = len(grouped)
        for row in _collect_oci_live_kubernetes_inventory_rows(customer_id, db):
            add_row(
                provider=str(row.get("provider") or "oci"),
                service_name=str(row.get("service") or ""),
                cost=float(row.get("monthly_cost_usd") or 0.0),
                source=str(row.get("source") or "live_resource_inventory"),
                region=str(row.get("region") or ""),
                account_identifier=str(row.get("account_identifier") or ""),
                category_override=str(row.get("category") or ""),
                evidence=str(row.get("evidence") or ""),
                allow_zero_cost=True,
                resource_id=str(row.get("resource_id") or ""),
                resource_name=str(row.get("resource_name") or ""),
                lifecycle_state=str(row.get("lifecycle_state") or ""),
                resource_shape=str(row.get("resource_shape") or ""),
                resource_version=str(row.get("resource_version") or ""),
                created_at=str(row.get("created_at") or ""),
                availability_domain=str(row.get("availability_domain") or ""),
                public_endpoint=str(row.get("public_endpoint") or ""),
                private_endpoint=str(row.get("private_endpoint") or ""),
                public_ip=str(row.get("public_ip") or ""),
                ocpus=row.get("ocpus"),
                memory_gib=row.get("memory_gib"),
                container_count=row.get("container_count"),
                container_images=row.get("container_images") if isinstance(row.get("container_images"), list) else [],
                console_url=str(row.get("console_url") or ""),
            )
        if len(grouped) > before_count and provider_sources["oci"] == "none":
            provider_sources["oci"] = "live_resource_inventory"
    except Exception as exc:
        logger.info("Unable to collect OCI live Kubernetes inventory: %s", exc)

    rows: List[KubernetesContainerServiceCost] = []
    container_total = sum(float(bucket.get("monthly_cost_usd") or 0.0) for bucket in grouped.values())
    denominator = container_total or total_cloud_cost_usd or 0.0
    for bucket in grouped.values():
        cost = round(float(bucket.get("monthly_cost_usd") or 0.0), 2)
        regions = sorted(str(region) for region in bucket.get("regions", set()) if region)
        accounts = bucket.get("accounts", set())
        rows.append(
            KubernetesContainerServiceCost(
                provider=str(bucket["provider"]),
                service=str(bucket["service"]),
                category=str(bucket["category"]),  # type: ignore[arg-type]
                monthly_cost_usd=cost,
                share_percent=round((cost / denominator) * 100, 2) if denominator > 0 else 0.0,
                source=str(bucket.get("source") or "unknown"),
                evidence=str(bucket.get("evidence") or _container_service_evidence(str(bucket["category"]))),
                account_count=len(accounts),
                region_count=len(regions),
                regions=regions[:8],
                resource_id=str(bucket.get("resource_id") or "") or None,
                resource_name=str(bucket.get("resource_name") or "") or None,
                lifecycle_state=str(bucket.get("lifecycle_state") or "") or None,
                resource_shape=str(bucket.get("resource_shape") or "") or None,
                resource_version=str(bucket.get("resource_version") or "") or None,
                created_at=str(bucket.get("created_at") or "") or None,
                availability_domain=str(bucket.get("availability_domain") or "") or None,
                public_endpoint=str(bucket.get("public_endpoint") or "") or None,
                private_endpoint=str(bucket.get("private_endpoint") or "") or None,
                public_ip=str(bucket.get("public_ip") or "") or None,
                ocpus=bucket.get("ocpus") if isinstance(bucket.get("ocpus"), (int, float)) else None,
                memory_gib=bucket.get("memory_gib") if isinstance(bucket.get("memory_gib"), (int, float)) else None,
                container_count=bucket.get("container_count") if isinstance(bucket.get("container_count"), int) else None,
                container_images=[
                    str(image)
                    for image in bucket.get("container_images", [])
                    if str(image or "").strip()
                ][:8],
                console_url=str(bucket.get("console_url") or "") or None,
            )
        )
    rows.sort(key=lambda item: item.monthly_cost_usd, reverse=True)

    provider_rollups: List[KubernetesProviderServiceRollup] = []
    for provider in providers:
        provider_services = [row for row in rows if row.provider == provider]
        provider_total = round(sum(row.monthly_cost_usd for row in provider_services), 2)
        provider_rollups.append(
            KubernetesProviderServiceRollup(
                provider=provider,
                configured=diagnostics.get(provider, False),
                source=provider_sources.get(provider, "none"),
                total_monthly_cost_usd=provider_total,
                share_percent=round((provider_total / denominator) * 100, 2) if denominator > 0 else 0.0,
                service_count=len(provider_services),
                services=provider_services,
            )
        )
    provider_rollups.sort(key=lambda item: item.total_monthly_cost_usd, reverse=True)

    source = "none"
    present_sources = {row.source for row in rows}
    for candidate in ("live_provider_api", "live_resource_inventory", "cost_snapshots_live", "csv_import"):
        if candidate in present_sources:
            source = candidate
            break
    return rows, provider_rollups, source


def _default_kubernetes_workloads(namespaces: List[str]) -> List[KubernetesWorkloadInput]:
    workloads: list[KubernetesWorkloadInput] = []
    for namespace in namespaces:
        if namespace == "kube-system":
            workloads.append(
                KubernetesWorkloadInput(
                    namespace=namespace,
                    workload_name="system-daemons",
                    team="platform",
                    node_pool="system",
                    replicas=3,
                    cpu_request_cores=0.6,
                    cpu_limit_cores=1.2,
                    memory_request_gib=1.5,
                    memory_limit_gib=3.0,
                    cpu_usage_cores=0.35,
                    memory_usage_gib=1.0,
                )
            )
        elif namespace in {"monitoring", "prometheus", "observability"}:
            workloads.append(
                KubernetesWorkloadInput(
                    namespace=namespace,
                    workload_name="observability-stack",
                    team="platform",
                    node_pool="general",
                    replicas=2,
                    cpu_request_cores=1.5,
                    cpu_limit_cores=3.0,
                    memory_request_gib=3.0,
                    memory_limit_gib=6.0,
                    cpu_usage_cores=0.9,
                    memory_usage_gib=2.0,
                )
            )
        else:
            workloads.append(
                KubernetesWorkloadInput(
                    namespace=namespace,
                    workload_name=f"{namespace}-app",
                    team=namespace if namespace not in {"default", "app"} else "application",
                    node_pool="general",
                    replicas=3,
                    cpu_request_cores=2.0,
                    cpu_limit_cores=4.0,
                    memory_request_gib=4.0,
                    memory_limit_gib=8.0,
                    cpu_usage_cores=0.8,
                    memory_usage_gib=1.8,
                )
            )
    return workloads


def _build_kubernetes_deep_breakdowns(
    payload: KubernetesClusterInput,
    total_cost: float,
    namespace_breakdown: List[Dict[str, Any]],
) -> Dict[str, Any]:
    namespace_costs = {
        row["namespace"]: float(row.get("estimated_cost_usd") or 0.0)
        for row in namespace_breakdown
    }
    workloads = payload.workloads or _default_kubernetes_workloads(list(namespace_costs.keys()))
    workload_groups: Dict[str, List[KubernetesWorkloadInput]] = {}
    for workload in workloads:
        workload_groups.setdefault(workload.namespace, []).append(workload)

    workload_rows: list[dict[str, Any]] = []
    for namespace, rows in workload_groups.items():
        ns_cost = namespace_costs.get(namespace, 0.0)
        weights: list[float] = []
        for row in rows:
            cpu_weight = max(row.cpu_request_cores, row.cpu_usage_cores, 0.1)
            mem_weight = max(row.memory_request_gib / 4.0, row.memory_usage_gib / 4.0, 0.1)
            weights.append((cpu_weight + mem_weight) * max(row.replicas, 1))
        total_weight = sum(weights) or float(len(rows) or 1)
        for row, weight in zip(rows, weights):
            cost = round(ns_cost * (weight / total_weight), 2) if total_weight else 0.0
            requested = row.cpu_request_cores + (row.memory_request_gib / 4.0)
            used = row.cpu_usage_cores + (row.memory_usage_gib / 4.0)
            efficiency = round(min(100.0, (used / requested) * 100), 2) if requested > 0 else 0.0
            workload_rows.append(
                {
                    "namespace": row.namespace,
                    "workload_name": row.workload_name,
                    "team": row.team or row.namespace or "unassigned",
                    "node_pool": row.node_pool or "general",
                    "estimated_cost_usd": cost,
                    "share_percent": round((cost / total_cost) * 100, 2) if total_cost > 0 else 0.0,
                    "cpu_request_cores": round(row.cpu_request_cores, 3),
                    "cpu_usage_cores": round(row.cpu_usage_cores, 3),
                    "memory_request_gib": round(row.memory_request_gib, 3),
                    "memory_usage_gib": round(row.memory_usage_gib, 3),
                    "request_efficiency_percent": efficiency,
                }
            )

    team_totals: Dict[str, Dict[str, Any]] = {}
    for row in workload_rows:
        team = row["team"]
        entry = team_totals.setdefault(team, {"cost": 0.0, "namespaces": set(), "workloads": 0})
        entry["cost"] += float(row["estimated_cost_usd"])
        entry["namespaces"].add(row["namespace"])
        entry["workloads"] += 1
    team_rows = [
        {
            "team": team,
            "estimated_cost_usd": round(values["cost"], 2),
            "share_percent": round((values["cost"] / total_cost) * 100, 2) if total_cost > 0 else 0.0,
            "namespaces": sorted(values["namespaces"]),
            "workload_count": int(values["workloads"]),
        }
        for team, values in sorted(team_totals.items(), key=lambda item: item[1]["cost"], reverse=True)
    ]

    node_pools = payload.node_pools or [
        KubernetesNodePoolInput(
            name="general",
            node_count=payload.node_count,
            monthly_node_cost_usd=payload.monthly_node_cost_usd,
        )
    ]
    node_pool_costs = {
        pool.name: pool.node_count * pool.monthly_node_cost_usd
        for pool in node_pools
    }
    workload_cost_by_pool: Dict[str, float] = {}
    for row in workload_rows:
        workload_cost_by_pool[row["node_pool"]] = workload_cost_by_pool.get(row["node_pool"], 0.0) + float(row["estimated_cost_usd"])

    node_pool_rows: list[dict[str, Any]] = []
    for pool in node_pools:
        pool_cost = float(node_pool_costs.get(pool.name, 0.0))
        allocated = workload_cost_by_pool.get(pool.name, 0.0)
        utilization = round(min(100.0, (allocated / pool_cost) * 100), 2) if pool_cost > 0 else 0.0
        node_pool_rows.append(
            {
                "node_pool": pool.name,
                "node_count": pool.node_count,
                "estimated_cost_usd": round(pool_cost, 2),
                "utilization_percent": utilization,
                "idle_cost_usd": round(max(pool_cost - allocated, 0.0), 2),
            }
        )

    recommendations: list[dict[str, Any]] = []
    for row in workload_rows:
        if row["request_efficiency_percent"] < 55.0 and row["estimated_cost_usd"] > 0:
            recommendations.append(
                {
                    "recommendation_id": f"k8s-requests-{row['namespace']}-{row['workload_name']}",
                    "category": "request_limit",
                    "target": f"{row['namespace']}/{row['workload_name']}",
                    "severity": "medium" if row["request_efficiency_percent"] >= 30.0 else "high",
                    "estimated_monthly_savings_usd": round(row["estimated_cost_usd"] * 0.2, 2),
                    "rationale": "CPU/memory requests are materially higher than observed usage.",
                    "action": "Reduce requests and review limits after one full business cycle of metrics.",
                }
            )
    for row in node_pool_rows:
        if row["utilization_percent"] < 65.0 and row["idle_cost_usd"] > 0:
            recommendations.append(
                {
                    "recommendation_id": f"k8s-nodepool-{row['node_pool']}",
                    "category": "node_pool",
                    "target": row["node_pool"],
                    "severity": "medium" if row["utilization_percent"] >= 35.0 else "high",
                    "estimated_monthly_savings_usd": round(row["idle_cost_usd"] * 0.5, 2),
                    "rationale": "Node pool has low allocated workload cost versus monthly capacity cost.",
                    "action": "Evaluate cluster autoscaler bounds, smaller nodes, or moving workloads to shared pools.",
                }
            )

    return {
        "workload_breakdown": sorted(workload_rows, key=lambda item: item["estimated_cost_usd"], reverse=True),
        "team_breakdown": team_rows,
        "node_pool_breakdown": sorted(node_pool_rows, key=lambda item: item["estimated_cost_usd"], reverse=True),
        "recommendations": sorted(
            recommendations,
            key=lambda item: item["estimated_monthly_savings_usd"],
            reverse=True,
        )[:10],
    }


class RemediationCandidate(BaseModel):
    action_id: str
    provider: str
    resource_id: str
    action_type: Literal["downsize", "terminate", "reserve", "modernize"]
    estimated_monthly_impact_usd: float = Field(ge=0)
    risk_level: Literal["low", "medium", "high"] = "medium"
    confidence: Literal["high", "medium", "low"] = "medium"
    metadata: Dict[str, Any] = Field(default_factory=dict)


class RemediationLoopRequest(BaseModel):
    dry_run: bool = True
    max_actions_per_run: int = Field(default=10, ge=1, le=200)
    max_total_impact_usd: float = Field(default=1000.0, ge=0)
    require_approval_above_usd: float = Field(default=250.0, ge=0)
    allowed_providers: List[str] = Field(default_factory=lambda: ["aws", "azure", "gcp", "oci", "kubernetes"])
    allowed_actions: List[str] = Field(default_factory=lambda: ["downsize", "terminate", "reserve", "modernize"])
    candidates: List[RemediationCandidate] = Field(default_factory=list)


class RemediationDecision(BaseModel):
    action_id: str
    provider: str
    resource_id: str
    action_type: str
    estimated_monthly_impact_usd: float
    status: Literal["planned", "executed", "requires_approval", "skipped"]
    reason: str


class RemediationLoopResponse(BaseModel):
    generated_at: str
    dry_run: bool
    guardrails: Dict[str, Any]
    executed_count: int
    planned_count: int
    requires_approval_count: int
    skipped_count: int
    total_planned_impact_usd: float
    decisions: List[RemediationDecision]


class TagDimensionScore(BaseModel):
    dimension: str
    completeness_percent: float
    covered_cost_usd: float
    uncovered_cost_usd: float
    missing_records: int


class TagQualityScoreResponse(BaseModel):
    generated_at: str
    organization_id: int
    provider_filter: str
    data_source: str
    total_records: int
    total_cost_usd: float
    completeness_score: float
    quality_grade: str
    dimensions: List[TagDimensionScore]
    recommendations: List[str]


class DecisionRecommendationItem(BaseModel):
    recommendation_id: str
    provider: str
    category: str
    title: str
    estimated_monthly_savings_usd: float
    payback_months: float
    confidence_score: float
    urgency_score: float
    decision_score: float
    rationale: str


class DecisionRecommendationResponse(BaseModel):
    generated_at: str
    organization_id: int
    provider_filter: str
    model: str
    total_candidates: int
    top_recommendations: List[DecisionRecommendationItem]
    model_features: List[str]


class FederatedAccountCostItem(BaseModel):
    provider: str
    account_identifier: str
    account_name: str
    account_type: str
    parent_account_identifier: Optional[str] = None
    source: str
    direct_cost_usd: float
    rolled_up_cost_usd: float = 0.0
    depth: int = 0
    child_count: int = 0
    regions: Dict[str, float] = Field(default_factory=dict)


class FederationCostResponse(BaseModel):
    generated_at: str
    organization_id: int
    customer_id: str
    provider_filter: str
    total_accounts: int
    total_cost_usd: float
    provider_totals_usd: Dict[str, float]
    account_type_totals_usd: Dict[str, float] = Field(default_factory=dict)
    source_totals_usd: Dict[str, float] = Field(default_factory=dict)
    accounts: List[FederatedAccountCostItem]


class WhiteLabelConfigResponse(BaseModel):
    brand_name: str
    logo_url: Optional[str] = None
    primary_color: str = "#2563eb"
    show_powered_by: bool = True


class PartnerCustomerPortfolioItem(BaseModel):
    organization_id: int
    customer_id: str
    customer_name: str
    plan: str
    role: str
    total_cost_usd: float
    savings_identified_usd: float
    providers: List[str]
    account_count: int
    scan_count: int
    open_alert_count: int
    last_activity_at: Optional[str] = None
    health_status: Literal["healthy", "attention", "no_data"]


class PartnerCustomerPortfolioResponse(BaseModel):
    generated_at: str
    partner_mode_enabled: bool
    white_label: WhiteLabelConfigResponse
    customer_count: int
    total_cost_usd: float
    savings_identified_usd: float
    open_alert_count: int
    customers: List[PartnerCustomerPortfolioItem]


def _white_label_config() -> WhiteLabelConfigResponse:
    return WhiteLabelConfigResponse(
        brand_name=os.getenv("WHITE_LABEL_BRAND_NAME", "OptiOra"),
        logo_url=os.getenv("WHITE_LABEL_LOGO_URL", "").strip() or None,
        primary_color=os.getenv("WHITE_LABEL_PRIMARY_COLOR", "#2563eb"),
        show_powered_by=os.getenv("WHITE_LABEL_SHOW_POWERED_BY", "true").strip().lower()
        in {"1", "true", "yes"},
    )


def _normalize_opencost_url(raw: str) -> str:
    value = str(raw or "").strip()
    if not value:
        return ""
    return value.rstrip("/")


def _flatten_opencost_entries(payload: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                rows.append(item)
            elif isinstance(item, list):
                rows.extend(_flatten_opencost_entries(item))
        return rows
    if isinstance(payload, dict):
        # OpenCost can return {data: {...}} with allocation map, or {data: [..]}
        if isinstance(payload.get("data"), list):
            rows.extend(_flatten_opencost_entries(payload.get("data")))
        elif isinstance(payload.get("data"), dict):
            for value in payload.get("data", {}).values():
                if isinstance(value, dict):
                    rows.append(value)
        elif isinstance(payload.get("items"), list):
            rows.extend(_flatten_opencost_entries(payload.get("items")))
    return rows


def _opencost_row_cost_usd(row: Dict[str, Any]) -> float:
    for key in ("totalCost", "totalCostUsd", "cost", "total"):
        value = row.get(key)
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, dict):
            numeric = [float(v) for v in value.values() if isinstance(v, (int, float))]
            if numeric:
                return float(sum(numeric))
    return 0.0


def _opencost_fetch_allocations(
    api_url: str,
    start_dt: datetime,
    end_dt: datetime,
    aggregate: str,
) -> List[Dict[str, Any]]:
    import httpx

    normalized_url = _normalize_opencost_url(api_url)
    if not normalized_url:
        raise ValueError("OpenCost URL is empty")
    endpoint = f"{normalized_url}/api/v1/allocation"
    params = {
        "start": start_dt.date().isoformat(),
        "end": end_dt.date().isoformat(),
        "aggregate": aggregate,
        "step": "1d",
    }
    with httpx.Client(timeout=45.0) as client:
        resp = client.get(endpoint, params=params)
        content_type = str(resp.headers.get("content-type") or "").lower()
        if resp.status_code >= 400:
            if "text/html" in content_type:
                raise ValueError(
                    f"OpenCost endpoint returned HTTP {resp.status_code} HTML response. "
                    "Verify the URL points to OpenCost API (for example http://<host>:9003), "
                    "not the dashboard frontend route."
                )
            resp.raise_for_status()
        if "text/html" in content_type:
            preview = resp.text[:180].strip().replace("\n", " ")
            raise ValueError(
                "OpenCost endpoint returned HTML instead of JSON. "
                f"Preview: {preview}"
            )
        try:
            payload = resp.json()
        except Exception as exc:
            preview = resp.text[:180].strip().replace("\n", " ")
            raise ValueError(
                f"OpenCost response is not valid JSON. Preview: {preview}"
            ) from exc
    rows = _flatten_opencost_entries(payload)
    out: List[Dict[str, Any]] = []
    for row in rows:
        properties = row.get("properties") if isinstance(row.get("properties"), dict) else {}
        out.append(
            {
                "namespace": str(
                    properties.get("namespace")
                    or row.get("namespace")
                    or "unknown"
                ),
                "pod": str(
                    properties.get("pod")
                    or row.get("pod")
                    or ""
                ),
                "cost_usd": _opencost_row_cost_usd(row),
            }
        )
    return [item for item in out if item["cost_usd"] > 0]


def _run_command(
    args: List[str],
    command_log: List[str],
    timeout: int = 180,
) -> subprocess.CompletedProcess[str]:
    command_log.append(f"$ {' '.join(args)}")
    result = subprocess.run(
        args,
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
    )
    if result.stdout.strip():
        command_log.append(result.stdout.strip())
    if result.stderr.strip():
        command_log.append(result.stderr.strip())
    return result


@router.get("/partner/customer-portfolio", response_model=PartnerCustomerPortfolioResponse)
async def get_partner_customer_portfolio(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> PartnerCustomerPortfolioResponse:
    """Portfolio view across all organizations the current user can access."""
    _ = membership
    now = _utcnow()
    db = SessionLocal()
    try:
        memberships = (
            db.query(UserOrganization, Organization)
            .join(Organization, Organization.id == UserOrganization.organization_id)
            .filter(
                UserOrganization.user_id == current_user.id,
                Organization.is_active == True,  # noqa: E712
            )
            .order_by(Organization.name.asc())
            .all()
        )
        if not memberships:
            return PartnerCustomerPortfolioResponse(
                generated_at=now.isoformat() + "Z",
                partner_mode_enabled=os.getenv("PARTNER_MODE_ENABLED", "false").strip().lower()
                in {"1", "true", "yes"},
                white_label=_white_label_config(),
                customer_count=0,
                total_cost_usd=0.0,
                savings_identified_usd=0.0,
                open_alert_count=0,
                customers=[],
            )

        customers: list[PartnerCustomerPortfolioItem] = []
        for user_org, org in memberships:
            customer_id = f"org-{org.id}"
            imported_rows = (
                db.query(ImportedCostRecord)
                .filter(ImportedCostRecord.organization_id == org.id)
                .limit(20000)
                .all()
            )
            snapshots = (
                db.query(CostSnapshot)
                .filter(CostSnapshot.customer_id == customer_id)
                .order_by(CostSnapshot.captured_at.desc())
                .limit(200)
                .all()
            )
            providers = {
                str(row.provider or "").lower()
                for row in imported_rows
                if str(row.provider or "").strip()
            }
            providers.update(
                str(row.provider or "").lower()
                for row in snapshots
                if str(row.provider or "").strip()
            )

            imported_total = sum(float(row.cost_usd or 0.0) for row in imported_rows)
            latest_snapshot_by_provider: dict[str, CostSnapshot] = {}
            for snapshot in snapshots:
                provider_key = str(snapshot.provider or "").lower()
                if provider_key and provider_key not in latest_snapshot_by_provider:
                    latest_snapshot_by_provider[provider_key] = snapshot
            snapshot_total = sum(float(row.total_cost_usd or 0.0) for row in latest_snapshot_by_provider.values())
            snapshot_savings = sum(float(row.savings_identified_usd or 0.0) for row in latest_snapshot_by_provider.values())
            total_cost = imported_total if imported_rows else snapshot_total

            account_count = (
                db.query(ProviderAccount)
                .filter(ProviderAccount.organization_id == org.id)
                .count()
            )
            scan_count = (
                db.query(ScanRunRecord)
                .filter(ScanRunRecord.customer_id == customer_id)
                .count()
            )
            open_alert_count = (
                db.query(AlertEvent)
                .filter(
                    AlertEvent.organization_id == org.id,
                    AlertEvent.acknowledged_at.is_(None),
                )
                .count()
            )
            last_candidates = [
                *(row.created_at for row in imported_rows if row.created_at),
                *(row.captured_at for row in snapshots if row.captured_at),
                org.updated_at,
                org.created_at,
            ]
            last_activity = max([value for value in last_candidates if value], default=None)
            if total_cost <= 0 and not providers:
                health = "no_data"
            elif open_alert_count > 0:
                health = "attention"
            else:
                health = "healthy"

            customers.append(
                PartnerCustomerPortfolioItem(
                    organization_id=int(org.id),
                    customer_id=customer_id,
                    customer_name=org.name,
                    plan=str(getattr(org.plan, "value", org.plan)),
                    role=str(getattr(user_org.role, "value", user_org.role)),
                    total_cost_usd=round(total_cost, 2),
                    savings_identified_usd=round(snapshot_savings, 2),
                    providers=sorted(providers),
                    account_count=int(account_count),
                    scan_count=int(scan_count),
                    open_alert_count=int(open_alert_count),
                    last_activity_at=last_activity.isoformat() if last_activity else None,
                    health_status=health,
                )
            )

        total_cost = round(sum(item.total_cost_usd for item in customers), 2)
        savings = round(sum(item.savings_identified_usd for item in customers), 2)
        alerts = sum(item.open_alert_count for item in customers)
        return PartnerCustomerPortfolioResponse(
            generated_at=now.isoformat() + "Z",
            partner_mode_enabled=os.getenv("PARTNER_MODE_ENABLED", "false").strip().lower()
            in {"1", "true", "yes"},
            white_label=_white_label_config(),
            customer_count=len(customers),
            total_cost_usd=total_cost,
            savings_identified_usd=savings,
            open_alert_count=alerts,
            customers=sorted(customers, key=lambda item: item.total_cost_usd, reverse=True),
        )
    finally:
        db.close()


@router.post("/analytics/kubernetes/cluster-cost", response_model=KubernetesClusterCostResponse)
async def calculate_kubernetes_cluster_cost(
    payload: KubernetesClusterInput,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Estimate Kubernetes cluster cost allocation by namespace."""
    if payload.opencost_enabled and payload.opencost_url:
        try:
            start_date = _utcnow() - timedelta(days=payload.opencost_window_days)
            costs = _opencost_fetch_allocations(
                api_url=payload.opencost_url,
                start_dt=start_date,
                end_dt=_utcnow(),
                aggregate="namespace",
            )
            namespace_totals: Dict[str, float] = {}
            for point in costs:
                namespace = str(point.get("namespace") or "unknown")
                namespace_totals[namespace] = namespace_totals.get(namespace, 0.0) + float(point.get("cost_usd") or 0.0)

            total_live_cost = round(sum(namespace_totals.values()), 2)
            if total_live_cost > 0 and namespace_totals:
                namespace_breakdown = []
                for ns, ns_cost in sorted(namespace_totals.items(), key=lambda item: item[1], reverse=True):
                    share = round((ns_cost / total_live_cost) * 100, 2)
                    namespace_breakdown.append(
                        {
                            "namespace": ns,
                            "estimated_cost_usd": round(ns_cost, 2),
                            "share_percent": share,
                            "cpu_share_percent": share,
                            "memory_share_percent": share,
                        }
                    )

                deep = _build_kubernetes_deep_breakdowns(
                    payload=payload,
                    total_cost=total_live_cost,
                    namespace_breakdown=namespace_breakdown,
                )
                return {
                    "generated_at": _utcnow().isoformat() + "Z",
                    "cluster_name": payload.cluster_name,
                    "provider": payload.provider,
                    "region": payload.region,
                    "node_count": payload.node_count,
                    "node_type": payload.node_type,
                    "total_cluster_cost_usd": total_live_cost,
                    "cost_per_node_usd": round(total_live_cost / max(payload.node_count, 1), 2),
                    "namespace_breakdown": namespace_breakdown,
                    **deep,
                    "efficiency_note": "Using real OpenCost namespace allocation data.",
                    "opencost_integration": f"live:{payload.opencost_url}",
                }
        except Exception as exc:
            logger.warning("OpenCost live allocation unavailable, falling back to heuristic allocation: %s", exc)

    total_cost = round(payload.node_count * payload.monthly_node_cost_usd, 2)
    cost_per_node = payload.monthly_node_cost_usd

    namespaces = payload.namespaces or ["default", "kube-system", "monitoring", "app"]
    n = len(namespaces)
    # Even split as baseline — real allocation requires OpenCost/Prometheus metrics
    namespace_breakdown: List[Dict[str, Any]] = []
    for i, ns in enumerate(namespaces):
        # Give kube-system ~10%, monitoring ~15%, rest shared evenly
        if ns == "kube-system":
            share = 10.0
        elif ns in ("monitoring", "prometheus", "observability"):
            share = 15.0
        else:
            remaining = 75.0 / max(1, n - sum(1 for x in namespaces if x in ("kube-system", "monitoring", "prometheus", "observability")))
            share = round(remaining, 1)

        namespace_breakdown.append({
            "namespace": ns,
            "estimated_cost_usd": round(total_cost * share / 100, 2),
            "share_percent": share,
            "cpu_share_percent": share,
            "memory_share_percent": share,
        })

    deep = _build_kubernetes_deep_breakdowns(
        payload=payload,
        total_cost=total_cost,
        namespace_breakdown=namespace_breakdown,
    )
    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "cluster_name": payload.cluster_name,
        "provider": payload.provider,
        "region": payload.region,
        "node_count": payload.node_count,
        "node_type": payload.node_type,
        "total_cluster_cost_usd": total_cost,
        "cost_per_node_usd": cost_per_node,
        "namespace_breakdown": namespace_breakdown,
        **deep,
        "efficiency_note": (
            "Namespace breakdown uses proportional allocation. Connect OpenCost or Prometheus "
            "metrics to enable pod-level CPU/memory-weighted allocation."
        ),
        "opencost_integration": "POST /api/v1/analytics/kubernetes/cluster-cost with real prometheus metrics for weighted allocation.",
    }


@router.post("/analytics/kubernetes/opencost/sync", response_model=OpenCostSyncResponse)
async def sync_opencost_costs(
    payload: OpenCostSyncRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> OpenCostSyncResponse:
    """Fetch production OpenCost namespace allocation for a cluster."""
    _ = current_user
    _require_management_role(membership, "OpenCost sync")
    end_dt = _utcnow()
    start_dt = end_dt - timedelta(days=payload.window_days)
    try:
        namespace_costs = _opencost_fetch_allocations(
            api_url=payload.api_url,
            start_dt=start_dt,
            end_dt=end_dt,
            aggregate="namespace",
        )
        pod_costs = _opencost_fetch_allocations(
            api_url=payload.api_url,
            start_dt=start_dt,
            end_dt=end_dt,
            aggregate="namespace,pod",
        )
    except Exception as exc:
        detail = str(exc)
        if "Connection refused" in detail or "Failed to establish a new connection" in detail:
            detail = (
                "OpenCost sync failed: OpenCost API unreachable. "
                "Use an OpenCost URL reachable from the OptiOra API host "
                "(not a browser-only localhost endpoint)."
            )
        else:
            detail = f"OpenCost sync failed: {detail}"
        raise HTTPException(status_code=502, detail=detail) from exc
    namespace_totals: Dict[str, float] = {}
    for point in namespace_costs:
        namespace = str(point.get("namespace") or "unknown")
        namespace_totals[namespace] = namespace_totals.get(namespace, 0.0) + float(point.get("cost_usd") or 0.0)

    total_cost = round(sum(namespace_totals.values()), 2)
    rows = [
        OpenCostNamespaceCost(
            namespace=ns,
            cost_usd=round(cost, 2),
            share_percent=round((cost / total_cost) * 100, 2) if total_cost > 0 else 0.0,
        )
        for ns, cost in sorted(namespace_totals.items(), key=lambda x: x[1], reverse=True)
    ]
    pod_rows_raw: List[OpenCostPodCost] = []
    for item in pod_costs:
        ns = str(item.get("namespace") or "unknown")
        pod_name = str(item.get("pod") or "").strip() or "unknown-pod"
        cost_usd = float(item.get("cost_usd") or 0.0)
        if cost_usd <= 0:
            continue
        pod_rows_raw.append(
            OpenCostPodCost(
                namespace=ns,
                pod_name=pod_name,
                cost_usd=round(cost_usd, 2),
                share_percent=0.0,
            )
        )
    pod_total = sum(row.cost_usd for row in pod_rows_raw) or 0.0
    pod_rows = [
        OpenCostPodCost(
            namespace=row.namespace,
            pod_name=row.pod_name,
            cost_usd=row.cost_usd,
            share_percent=round((row.cost_usd / pod_total) * 100, 4) if pod_total > 0 else 0.0,
        )
        for row in sorted(pod_rows_raw, key=lambda item: item.cost_usd, reverse=True)
    ]

    return OpenCostSyncResponse(
        generated_at=_utcnow().isoformat() + "Z",
        cluster_name=payload.cluster_name,
        source=_normalize_opencost_url(payload.api_url),
        window_days=payload.window_days,
        total_cost_usd=total_cost,
        namespace_count=len(rows),
        namespaces=rows,
        pods=pod_rows[:100],
    )


@router.post("/analytics/kubernetes/opencost/auto-install", response_model=OpenCostInstallResponse)
async def auto_install_opencost(
    payload: OpenCostInstallRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> OpenCostInstallResponse:
    """Install/upgrade Prometheus + OpenCost in current Kubernetes context on API host."""
    _ = current_user
    _require_management_role(membership, "OpenCost auto-install")
    command_log: List[str] = []

    kubectl_bin = shutil.which("kubectl")
    helm_bin = shutil.which("helm")
    if not kubectl_bin or not helm_bin:
        missing = []
        if not kubectl_bin:
            missing.append("kubectl")
        if not helm_bin:
            missing.append("helm")
        return OpenCostInstallResponse(
            generated_at=_utcnow().isoformat() + "Z",
            status="failed",
            message=f"Missing required binaries on API host: {', '.join(missing)}",
            api_url=None,
            namespace=payload.namespace,
            prometheus_namespace=payload.prometheus_namespace,
            command_log=command_log,
        )

    context_args: List[str] = []
    if payload.kube_context:
        context_args = ["--context", payload.kube_context]

    # Verify cluster access
    probe = _run_command([kubectl_bin, *context_args, "get", "namespace"], command_log, timeout=60)
    if probe.returncode != 0:
        return OpenCostInstallResponse(
            generated_at=_utcnow().isoformat() + "Z",
            status="failed",
            message="kubectl cannot access a Kubernetes cluster from this API host.",
            api_url=None,
            namespace=payload.namespace,
            prometheus_namespace=payload.prometheus_namespace,
            command_log=command_log,
        )

    # Ensure helm repos
    _run_command([helm_bin, "repo", "add", "prometheus-community", "https://prometheus-community.github.io/helm-charts"], command_log, timeout=60)
    _run_command([helm_bin, "repo", "add", "opencost-charts", "https://opencost.github.io/opencost-helm-chart"], command_log, timeout=60)
    _run_command([helm_bin, "repo", "update"], command_log, timeout=120)

    if not payload.skip_prometheus_install:
        prom_cmd = [
            helm_bin,
            "upgrade",
            "--install",
            "prometheus",
            "prometheus-community/prometheus",
            "--namespace",
            payload.prometheus_namespace,
            "--create-namespace",
            "--set",
            "prometheus-pushgateway.enabled=false",
            "--set",
            "alertmanager.enabled=false",
            "-f",
            "https://raw.githubusercontent.com/opencost/opencost/develop/kubernetes/prometheus/extraScrapeConfigs.yaml",
        ]
        prom_result = _run_command(prom_cmd, command_log, timeout=420)
        if prom_result.returncode != 0:
            return OpenCostInstallResponse(
                generated_at=_utcnow().isoformat() + "Z",
                status="failed",
                message="Prometheus install/upgrade failed.",
                api_url=None,
                namespace=payload.namespace,
                prometheus_namespace=payload.prometheus_namespace,
                command_log=command_log,
            )

    oc_cmd = [
        helm_bin,
        "upgrade",
        "--install",
        "opencost",
        "opencost-charts/opencost",
        "--namespace",
        payload.namespace,
        "--create-namespace",
        "--set",
        f"opencost.prometheus.internal.namespaceName={payload.prometheus_namespace}",
        "--set",
        "opencost.prometheus.internal.serviceName=prometheus-server",
        "--set",
        "opencost.prometheus.internal.port=80",
    ]
    oc_result = _run_command(oc_cmd, command_log, timeout=420)
    if oc_result.returncode != 0:
        return OpenCostInstallResponse(
            generated_at=_utcnow().isoformat() + "Z",
            status="failed",
            message="OpenCost install/upgrade failed.",
            api_url=None,
            namespace=payload.namespace,
            prometheus_namespace=payload.prometheus_namespace,
            command_log=command_log,
        )

    svc_result = _run_command(
        [
            kubectl_bin,
            *context_args,
            "-n",
            payload.namespace,
            "get",
            "svc",
            "opencost",
            "-o",
            "jsonpath={.spec.ports[0].port}",
        ],
        command_log,
        timeout=60,
    )
    service_port = "9003"
    if svc_result.returncode == 0 and svc_result.stdout.strip():
        service_port = svc_result.stdout.strip()
    api_url = f"http://localhost:{payload.expose_port}"

    # Refresh/replace local forwarder for API host consumption.
    existing = _run_command(
        ["/bin/sh", "-lc", "pkill -f 'kubectl.*port-forward.*svc/opencost' || true"],
        command_log,
        timeout=30,
    )
    _ = existing
    port_forward_cmd = (
        f"nohup {kubectl_bin} {' '.join(context_args)} -n {payload.namespace} "
        f"port-forward svc/opencost {payload.expose_port}:{service_port} "
        f">/tmp/opencost-port-forward.log 2>&1 &"
    )
    _run_command(["/bin/sh", "-lc", port_forward_cmd], command_log, timeout=30)

    # Validate endpoint.
    try:
        _opencost_fetch_allocations(
            api_url=api_url,
            start_dt=_utcnow() - timedelta(days=1),
            end_dt=_utcnow(),
            aggregate="namespace",
        )
        status_value: Literal["installed", "already_installed", "failed"] = "installed"
        message = "OpenCost is installed and reachable from OptiOra API host."
    except Exception as exc:
        status_value = "failed"
        message = f"OpenCost install completed but API probe failed: {exc}"

    return OpenCostInstallResponse(
        generated_at=_utcnow().isoformat() + "Z",
        status=status_value,
        message=message,
        api_url=api_url if status_value != "failed" else None,
        namespace=payload.namespace,
        prometheus_namespace=payload.prometheus_namespace,
        command_log=command_log[-40:],
    )


def _quality_grade(score: float) -> str:
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    if score >= 60:
        return "D"
    return "F"


@router.get("/analytics/tag-quality", response_model=TagQualityScoreResponse)
async def get_tag_quality_score(
    provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> TagQualityScoreResponse:
    """Dimension completeness scoring engine over mapped and imported cost data."""
    _ = current_user
    org_id = membership.organization_id
    customer_id = _customer_id_for_org(membership)
    dimensions = ["team", "environment", "application", "cost_center"]

    mapped_q = db.query(NormalizedCostDimension).filter(NormalizedCostDimension.organization_id == org_id)
    if provider != "all":
        mapped_q = mapped_q.filter(NormalizedCostDimension.provider == provider)
    mapped_rows = mapped_q.all()

    dimension_data: Dict[str, Dict[str, float]] = {
        d: {"covered_cost": 0.0, "uncovered_cost": 0.0, "missing_records": 0.0}
        for d in dimensions
    }

    data_source = "normalized_dimensions"
    total_records = 0
    total_cost = 0.0

    if mapped_rows:
        total_records = len(mapped_rows)
        for row in mapped_rows:
            row_cost = float(row.cost_usd or 0.0)
            total_cost += row_cost
            for dim in dimensions:
                if getattr(row, dim, None):
                    dimension_data[dim]["covered_cost"] += row_cost
                else:
                    dimension_data[dim]["uncovered_cost"] += row_cost
                    dimension_data[dim]["missing_records"] += 1
    else:
        data_source = "imported_tags"
        imported_rows = _get_imported_cost_rows(db, org_id, customer_id, provider)
        total_records = len(imported_rows)
        synonyms = {
            "team": ["team", "owner", "squad"],
            "environment": ["environment", "env"],
            "application": ["application", "app", "service"],
            "cost_center": ["cost_center", "cost-center", "costcenter", "cc"],
        }
        for row in imported_rows:
            row_cost = float(row.cost_usd or 0.0)
            total_cost += row_cost
            tags: Dict[str, Any] = {}
            if row.tags_json:
                try:
                    tags = json.loads(row.tags_json)
                except Exception:
                    tags = {}
            lowered = {str(k).lower(): str(v) for k, v in tags.items()}
            for dim in dimensions:
                if any(lowered.get(alias) for alias in synonyms[dim]):
                    dimension_data[dim]["covered_cost"] += row_cost
                else:
                    dimension_data[dim]["uncovered_cost"] += row_cost
                    dimension_data[dim]["missing_records"] += 1

    dimension_scores: List[TagDimensionScore] = []
    completeness_values: List[float] = []
    recommendations_out: List[str] = []
    for dim in dimensions:
        covered = dimension_data[dim]["covered_cost"]
        uncovered = dimension_data[dim]["uncovered_cost"]
        completeness = round((covered / (covered + uncovered) * 100) if (covered + uncovered) > 0 else 0.0, 2)
        completeness_values.append(completeness)
        if completeness < 80:
            recommendations_out.append(
                f"Improve {dim} tagging coverage (current {completeness:.1f}%)."
            )
        dimension_scores.append(
            TagDimensionScore(
                dimension=dim,
                completeness_percent=completeness,
                covered_cost_usd=round(covered, 2),
                uncovered_cost_usd=round(uncovered, 2),
                missing_records=int(dimension_data[dim]["missing_records"]),
            )
        )

    score = round(sum(completeness_values) / len(completeness_values), 2) if completeness_values else 0.0
    return TagQualityScoreResponse(
        generated_at=_utcnow().isoformat() + "Z",
        organization_id=org_id,
        provider_filter=provider,
        data_source=data_source,
        total_records=total_records,
        total_cost_usd=round(total_cost, 2),
        completeness_score=score,
        quality_grade=_quality_grade(score),
        dimensions=dimension_scores,
        recommendations=recommendations_out,
    )


@router.get("/federation/costs", response_model=FederationCostResponse)
async def get_federated_costs(
    provider: str = "all",
    include_regions: bool = True,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> FederationCostResponse:
    """Cross-account federation across linked provider accounts with imported-cost fallback."""
    _ = current_user
    org_id = membership.organization_id
    customer_id = _customer_id_for_org(membership)

    accounts_q = db.query(ProviderAccount).filter(
        ProviderAccount.organization_id == org_id,
        ProviderAccount.customer_id == customer_id,
    )
    if provider != "all":
        accounts_q = accounts_q.filter(ProviderAccount.provider == provider)
    accounts = accounts_q.all()
    account_by_id = {a.id: a for a in accounts}

    parent_by_child: Dict[int, int] = {}
    if account_by_id:
        links = (
            db.query(ProviderAccountLink)
            .filter(
                ProviderAccountLink.organization_id == org_id,
                ProviderAccountLink.child_account_id.in_(list(account_by_id.keys())),
            )
            .all()
        )
        for link in links:
            if link.parent_account_id in account_by_id:
                parent_by_child[link.child_account_id] = link.parent_account_id

    latest_by_account: Dict[int, ProviderAccountSnapshot] = {}
    if account_by_id:
        snaps = (
            db.query(ProviderAccountSnapshot)
            .filter(
                ProviderAccountSnapshot.organization_id == org_id,
                ProviderAccountSnapshot.customer_id == customer_id,
                ProviderAccountSnapshot.provider_account_id.in_(list(account_by_id.keys())),
            )
            .order_by(ProviderAccountSnapshot.captured_at.desc())
            .all()
        )
        for snap in snaps:
            if snap.provider_account_id not in latest_by_account:
                latest_by_account[snap.provider_account_id] = snap

    region_map: Dict[int, List[AccountRegionRow]] = {}
    if include_regions and account_by_id:
        allocs = (
            db.query(CostAllocationSnapshot)
            .filter(
                CostAllocationSnapshot.organization_id == org_id,
                CostAllocationSnapshot.customer_id == customer_id,
                CostAllocationSnapshot.provider_account_id.in_(list(account_by_id.keys())),
            )
            .order_by(CostAllocationSnapshot.captured_at.desc())
            .all()
        )
        region_totals: Dict[int, Dict[str, float]] = {}
        for alloc in allocs:
            latest = latest_by_account.get(alloc.provider_account_id)
            if latest and alloc.scan_id != latest.scan_id:
                continue
            region_totals.setdefault(alloc.provider_account_id, {})
            region_totals[alloc.provider_account_id][alloc.region] = round(
                region_totals[alloc.provider_account_id].get(alloc.region, 0.0) + float(alloc.cost_usd or 0.0),
                2,
            )
        for account_id, totals in region_totals.items():
            region_map[account_id] = [
                AccountRegionRow(region=region_name, cost_usd=cost)
                for region_name, cost in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:5]
            ]

    nodes: Dict[int, Dict[str, Any]] = {}
    provider_roots: Dict[str, int] = {}
    node_id_by_identifier: Dict[tuple[str, str], int] = {}
    next_synthetic_id = -1

    def _synthetic_id() -> int:
        nonlocal next_synthetic_id
        current = next_synthetic_id
        next_synthetic_id -= 1
        return current

    def _provider_root(provider_key: str) -> int:
        provider_key = provider_key.lower()
        root_id = provider_roots.get(provider_key)
        if root_id is not None:
            return root_id
        root_id = _synthetic_id()
        provider_roots[provider_key] = root_id
        identifier = f"{provider_key}:provider"
        node_id_by_identifier[(provider_key, identifier)] = root_id
        nodes[root_id] = {
            "provider": provider_key,
            "account_identifier": identifier,
            "account_name": provider_key.upper(),
            "account_type": "provider",
            "parent_account_id": None,
            "parent_account_identifier": None,
            "direct_cost_usd": 0.0,
            "direct_savings_identified_usd": 0.0,
            "direct_anomalies_count": 0,
            "direct_service_count": 0,
            "scan_id": None,
            "captured_at": None,
            "top_regions": [],
            "source": "rollup",
        }
        return root_id

    def _ensure_synthetic_node(
        *,
        provider_key: str,
        identifier: str,
        account_name: str,
        account_type: str,
        parent_account_id: Optional[int],
        source: str,
    ) -> int:
        provider_key = provider_key.lower()
        node_key = (provider_key, identifier)
        existing_id = node_id_by_identifier.get(node_key)
        if existing_id is not None:
            node = nodes[existing_id]
            if account_name and node.get("account_name") == node.get("account_identifier"):
                node["account_name"] = account_name
            if account_type and str(node.get("account_type") or "") in {"", "group", "imported"}:
                node["account_type"] = account_type
            if parent_account_id is not None and node.get("parent_account_id") is None:
                node["parent_account_id"] = parent_account_id
                node["parent_account_identifier"] = nodes[parent_account_id]["account_identifier"]
            return existing_id

        node_id = _synthetic_id()
        node_id_by_identifier[node_key] = node_id
        nodes[node_id] = {
            "provider": provider_key,
            "account_identifier": identifier,
            "account_name": account_name or identifier,
            "account_type": account_type or "group",
            "parent_account_id": parent_account_id,
            "parent_account_identifier": (
                nodes[parent_account_id]["account_identifier"]
                if parent_account_id in nodes
                else None
            ),
            "direct_cost_usd": 0.0,
            "direct_savings_identified_usd": 0.0,
            "direct_anomalies_count": 0,
            "direct_service_count": 0,
            "scan_id": None,
            "captured_at": None,
            "top_regions": [],
            "source": source,
        }
        return node_id

    for account_id, account in account_by_id.items():
        provider_key = account.provider.lower()
        _provider_root(provider_key)
        snap = latest_by_account.get(account_id)
        cost_usd = round(float(snap.direct_cost_usd or 0.0), 2) if snap else 0.0
        parent_id = parent_by_child.get(account_id)
        nodes[account_id] = {
            "provider": provider_key,
            "account_identifier": account.account_identifier,
            "account_name": account.account_name,
            "account_type": account.account_type,
            "parent_account_id": parent_id,
            "parent_account_identifier": (
                account_by_id[parent_id].account_identifier
                if parent_id in account_by_id
                else None
            ),
            "direct_cost_usd": cost_usd,
            "direct_savings_identified_usd": round(float(snap.savings_identified_usd or 0.0), 2) if snap else 0.0,
            "direct_anomalies_count": int(snap.anomalies_count or 0) if snap else 0,
            "direct_service_count": int(snap.service_count or 0) if snap else 0,
            "scan_id": snap.scan_id if snap else None,
            "captured_at": snap.captured_at.isoformat() if snap and snap.captured_at else None,
            "top_regions": region_map.get(account_id, []),
            "source": "snapshot" if snap else "account",
        }
        node_id_by_identifier[(provider_key, account.account_identifier)] = account_id

    hierarchy_anchors: Dict[str, Dict[str, int]] = {}
    for account_id, node in nodes.items():
        if account_id < 0:
            continue
        provider_key = str(node.get("provider") or "")
        account_type = str(node.get("account_type") or "")
        if account_type in {"organization", "management_group", "folder", "tenancy"}:
            hierarchy_anchors.setdefault(provider_key, {})[account_type] = account_id

    for account_id, node in list(nodes.items()):
        if str(node.get("account_type") or "") == "provider":
            continue
        parent_id = node.get("parent_account_id")
        if parent_id in nodes and parent_id != account_id:
            continue

        provider_key = str(node.get("provider") or "").lower()
        account_type = str(node.get("account_type") or "").lower()
        root_id = _provider_root(provider_key)
        anchors = hierarchy_anchors.get(provider_key, {})
        inferred_parent_id = root_id
        if provider_key == "aws" and account_type == "account":
            inferred_parent_id = anchors.get("organization", root_id)
        elif provider_key == "azure" and account_type in {"account", "subscription"}:
            inferred_parent_id = anchors.get("management_group", root_id)
        elif provider_key == "gcp" and account_type == "project":
            inferred_parent_id = anchors.get("folder") or anchors.get("organization") or root_id
        elif provider_key == "gcp" and account_type == "folder":
            inferred_parent_id = anchors.get("organization", root_id)
        elif provider_key == "oci" and account_type == "compartment":
            inferred_parent_id = anchors.get("tenancy", root_id)

        if inferred_parent_id != account_id:
            node["parent_account_id"] = inferred_parent_id
            node["parent_account_identifier"] = nodes[inferred_parent_id]["account_identifier"]

    imported = _get_imported_cost_rows(db, org_id, customer_id, provider)
    imported_agg: Dict[str, Dict[str, Any]] = {}
    seen_account_keys = {
        f"{account.provider.lower()}:{account.account_identifier}"
        for account in account_by_id.values()
    }
    latest_imported_at = max((row.created_at for row in imported), default=None)
    for rec in imported:
        prov = (rec.provider or "unknown").lower()
        acct = rec.account_identifier or rec.account_name or f"imported-{rec.id}"
        key = f"{prov}:{acct}"
        if key in seen_account_keys:
            continue
        if key not in imported_agg:
            imported_agg[key] = {
                "provider": prov,
                "account_identifier": acct,
                "account_name": rec.account_name or acct,
                "account_type": rec.account_type or "imported",
                "parent_account_identifier": rec.parent_account_identifier or None,
                "parent_account_type": _infer_parent_account_type(prov, rec.account_type or "account"),
                "direct_cost_usd": 0.0,
                "regions": {},
            }
        if rec.parent_account_identifier and not imported_agg[key].get("parent_account_identifier"):
            imported_agg[key]["parent_account_identifier"] = rec.parent_account_identifier
            imported_agg[key]["parent_account_type"] = _infer_parent_account_type(prov, rec.account_type or "account")
        imported_agg[key]["direct_cost_usd"] += float(rec.cost_usd or 0.0)
        region_key = rec.region or "global"
        imported_agg[key]["regions"][region_key] = round(
            imported_agg[key]["regions"].get(region_key, 0.0) + float(rec.cost_usd or 0.0),
            2,
        )

    for row in imported_agg.values():
        provider_key = row["provider"]
        root_id = _provider_root(provider_key)
        parent_id = root_id
        parent_identifier = row.get("parent_account_identifier")
        if parent_identifier:
            parent_id = _ensure_synthetic_node(
                provider_key=provider_key,
                identifier=parent_identifier,
                account_name=parent_identifier,
                account_type=row.get("parent_account_type") or "group",
                parent_account_id=root_id,
                source="imported_parent",
            )
        child_id = _ensure_synthetic_node(
            provider_key=provider_key,
            identifier=row["account_identifier"],
            account_name=row["account_name"],
            account_type=row["account_type"],
            parent_account_id=parent_id,
            source="imported",
        )
        nodes[child_id]["direct_cost_usd"] = round(
            float(nodes[child_id].get("direct_cost_usd") or 0.0) + float(row["direct_cost_usd"] or 0.0),
            2,
        )
        nodes[child_id]["captured_at"] = latest_imported_at.isoformat() if latest_imported_at else None
        nodes[child_id]["top_regions"] = [
            AccountRegionRow(region=region_name, cost_usd=cost)
            for region_name, cost in sorted(row["regions"].items(), key=lambda item: item[1], reverse=True)[:5]
        ]

    items = _materialize_rollup_items(nodes)
    filtered_items = [item for item in items if provider == "all" or item.provider == provider]

    rows: List[FederatedAccountCostItem] = []
    provider_totals: Dict[str, float] = {}
    account_type_totals: Dict[str, float] = {}
    source_totals: Dict[str, float] = {}

    for item in filtered_items:
        node = nodes.get(item.account_id, {})
        source = str(node.get("source") or "account")
        regions = {
            region_row.region: region_row.cost_usd
            for region_row in item.top_regions
        } if include_regions else {}
        rows.append(
            FederatedAccountCostItem(
                provider=item.provider,
                account_identifier=item.account_identifier,
                account_name=item.account_name,
                account_type=item.account_type,
                parent_account_identifier=item.parent_account_identifier,
                source=source,
                direct_cost_usd=item.direct_cost_usd,
                rolled_up_cost_usd=item.rolled_up_cost_usd,
                depth=item.depth,
                child_count=item.child_count,
                regions=regions,
            )
        )
        account_type_totals[item.account_type] = round(
            account_type_totals.get(item.account_type, 0.0) + item.rolled_up_cost_usd,
            2,
        )
        if item.direct_cost_usd:
            source_totals[source] = round(source_totals.get(source, 0.0) + item.direct_cost_usd, 2)

    root_items = [item for item in filtered_items if item.depth == 0]
    for item in root_items:
        provider_totals[item.provider] = round(
            provider_totals.get(item.provider, 0.0) + item.rolled_up_cost_usd,
            2,
        )

    if not root_items:
        for item in filtered_items:
            provider_totals[item.provider] = round(provider_totals.get(item.provider, 0.0) + item.direct_cost_usd, 2)

    total_cost = round(
        sum(item.rolled_up_cost_usd for item in root_items)
        if root_items
        else sum(item.direct_cost_usd for item in filtered_items),
        2,
    )
    return FederationCostResponse(
        generated_at=_utcnow().isoformat() + "Z",
        organization_id=org_id,
        customer_id=customer_id,
        provider_filter=provider,
        total_accounts=len(rows),
        total_cost_usd=total_cost,
        provider_totals_usd=provider_totals,
        account_type_totals_usd=account_type_totals,
        source_totals_usd=source_totals,
        accounts=rows,
    )


@router.get("/recommendations/decision-grade", response_model=DecisionRecommendationResponse)
async def get_decision_grade_recommendations(
    provider: str = "all",
    top_n: int = Query(10, ge=1, le=50),
    min_monthly_savings: float = Query(10.0, ge=0),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> DecisionRecommendationResponse:
    """ML-enhanced (deterministic scoring) ranked optimization recommendations."""
    _ = current_user
    org_id = membership.organization_id
    context = await _cost_context(membership, db, "month", provider)
    provider_filter = provider

    raw = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": provider,
                "current_monthly_spend": context.get("total_cost", 0.0),
                "cost_breakdown": context.get("breakdown", {}),
                "min_savings_usd": min_monthly_savings * 12,
            }
        ),
        {},
    )

    scored: List[DecisionRecommendationItem] = []
    confidence_map = {"low": 0.45, "medium": 0.7, "high": 0.9}
    max_monthly_savings = max(
        [float((r.get("savings_annual_usd", 0) or 0) / 12.0) for r in raw.get("recommendations", [])] + [1.0]
    )

    for idx, rec in enumerate(raw.get("recommendations", []), start=1):
        monthly_savings = float(rec.get("savings_annual_usd", 0) or 0) / 12.0
        if monthly_savings < min_monthly_savings:
            continue
        payback = float(rec.get("payback_months", 3) or 3)
        confidence = confidence_map.get(str(rec.get("confidence", "medium")).lower(), 0.7)
        savings_signal = min(monthly_savings / max_monthly_savings, 1.0)
        urgency = 1.0 if str(rec.get("severity", "medium")).lower() == "high" else 0.65
        payback_signal = max(0.0, min(1.0, (12.0 - min(payback, 12.0)) / 12.0))
        decision_score = round((0.45 * savings_signal + 0.30 * confidence + 0.15 * urgency + 0.10 * payback_signal) * 100, 2)

        scored.append(
            DecisionRecommendationItem(
                recommendation_id=str(rec.get("id") or f"rec-{idx:03d}"),
                provider=(provider_filter if provider_filter != "all" else "multi-cloud"),
                category=str(rec.get("type") or "optimization"),
                title=str(rec.get("description") or "Optimization recommendation"),
                estimated_monthly_savings_usd=round(monthly_savings, 2),
                payback_months=round(payback, 2),
                confidence_score=round(confidence * 100, 2),
                urgency_score=round(urgency * 100, 2),
                decision_score=decision_score,
                rationale=(
                    f"Savings signal {savings_signal:.2f}, confidence {confidence:.2f}, "
                    f"urgency {urgency:.2f}, payback factor {payback_signal:.2f}."
                ),
            )
        )

    scored.sort(key=lambda item: item.decision_score, reverse=True)
    scored = scored[:top_n]
    return DecisionRecommendationResponse(
        generated_at=_utcnow().isoformat() + "Z",
        organization_id=org_id,
        provider_filter=provider_filter,
        model="ensemble_v1_deterministic",
        total_candidates=len(scored),
        top_recommendations=scored,
        model_features=[
            "normalized_monthly_savings",
            "confidence_score",
            "severity_urgency",
            "payback_signal",
        ],
    )


@router.post("/automation/remediation/loop", response_model=RemediationLoopResponse)
async def run_auto_remediation_loop(
    payload: RemediationLoopRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> RemediationLoopResponse:
    """Safe automation loop with guardrails for auto-remediation actions."""
    _require_management_role(membership, "auto-remediation")
    config = Config()
    if not payload.dry_run and not config.enable_auto_remediation:
        raise HTTPException(
            status_code=503,
            detail=(
                "Auto-remediation execution is disabled. "
                "Use dry_run=true or set ENABLE_AUTO_REMEDIATION=true."
            ),
        )

    candidates = list(payload.candidates)
    if not candidates:
        imported = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == membership.organization_id)
            .limit(30)
            .all()
        )
        synthesized = _rightsizing_from_imported_costs(imported, min_savings=5.0)
        candidates = [
            RemediationCandidate(
                action_id=f"auto-{idx:03d}",
                provider=rec.provider,
                resource_id=rec.resource_id,
                action_type=rec.action,
                estimated_monthly_impact_usd=float(rec.monthly_savings_usd),
                risk_level="high" if rec.effort == "high" else "medium" if rec.effort == "medium" else "low",
                confidence=rec.confidence,
                metadata={"generated_from": "rightsizing"},
            )
            for idx, rec in enumerate(synthesized, start=1)
        ]

    decisions: List[RemediationDecision] = []
    planned_impact = 0.0
    planned_count = 0
    executed_count = 0
    approval_count = 0
    skipped_count = 0

    for candidate in sorted(candidates, key=lambda c: c.estimated_monthly_impact_usd, reverse=True):
        if candidate.provider not in payload.allowed_providers:
            decisions.append(
                RemediationDecision(
                    action_id=candidate.action_id,
                    provider=candidate.provider,
                    resource_id=candidate.resource_id,
                    action_type=candidate.action_type,
                    estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                    status="skipped",
                    reason="Provider not allowed by guardrail.",
                )
            )
            skipped_count += 1
            continue

        if candidate.action_type not in payload.allowed_actions:
            decisions.append(
                RemediationDecision(
                    action_id=candidate.action_id,
                    provider=candidate.provider,
                    resource_id=candidate.resource_id,
                    action_type=candidate.action_type,
                    estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                    status="skipped",
                    reason="Action type not allowed by guardrail.",
                )
            )
            skipped_count += 1
            continue

        if planned_count >= payload.max_actions_per_run:
            decisions.append(
                RemediationDecision(
                    action_id=candidate.action_id,
                    provider=candidate.provider,
                    resource_id=candidate.resource_id,
                    action_type=candidate.action_type,
                    estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                    status="skipped",
                    reason="Max actions per run reached.",
                )
            )
            skipped_count += 1
            continue

        if planned_impact + candidate.estimated_monthly_impact_usd > payload.max_total_impact_usd:
            decisions.append(
                RemediationDecision(
                    action_id=candidate.action_id,
                    provider=candidate.provider,
                    resource_id=candidate.resource_id,
                    action_type=candidate.action_type,
                    estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                    status="skipped",
                    reason="Max total impact guardrail reached.",
                )
            )
            skipped_count += 1
            continue

        if candidate.risk_level == "high" or candidate.estimated_monthly_impact_usd >= payload.require_approval_above_usd:
            decisions.append(
                RemediationDecision(
                    action_id=candidate.action_id,
                    provider=candidate.provider,
                    resource_id=candidate.resource_id,
                    action_type=candidate.action_type,
                    estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                    status="requires_approval",
                    reason="Action exceeds approval threshold or is high risk.",
                )
            )
            approval_count += 1
            continue

        planned_impact += candidate.estimated_monthly_impact_usd
        planned_count += 1
        status_value: Literal["planned", "executed", "requires_approval", "skipped"] = "planned"
        reason_value = "Dry run mode: action planned but not executed."
        if not payload.dry_run:
            status_value = "executed"
            reason_value = "Executed within guardrails."
            executed_count += 1
            db.add(
                AuditLog(
                    organization_id=membership.organization_id,
                    actor_user_id=current_user.id,
                    action="auto_remediation_executed",
                    entity_type="remediation_action",
                    entity_id=candidate.action_id,
                    metadata_json=json.dumps(
                        {
                            "provider": candidate.provider,
                            "resource_id": candidate.resource_id,
                            "action_type": candidate.action_type,
                            "estimated_monthly_impact_usd": candidate.estimated_monthly_impact_usd,
                        }
                    ),
                    created_at=_utcnow(),
                )
            )
        decisions.append(
            RemediationDecision(
                action_id=candidate.action_id,
                provider=candidate.provider,
                resource_id=candidate.resource_id,
                action_type=candidate.action_type,
                estimated_monthly_impact_usd=candidate.estimated_monthly_impact_usd,
                status=status_value,
                reason=reason_value,
            )
        )

    if not payload.dry_run and executed_count > 0:
        db.commit()

    return RemediationLoopResponse(
        generated_at=_utcnow().isoformat() + "Z",
        dry_run=payload.dry_run,
        guardrails={
            "max_actions_per_run": payload.max_actions_per_run,
            "max_total_impact_usd": payload.max_total_impact_usd,
            "require_approval_above_usd": payload.require_approval_above_usd,
            "allowed_providers": payload.allowed_providers,
            "allowed_actions": payload.allowed_actions,
        },
        executed_count=executed_count,
        planned_count=planned_count,
        requires_approval_count=approval_count,
        skipped_count=skipped_count,
        total_planned_impact_usd=round(planned_impact, 2),
        decisions=decisions,
    )


@router.get("/analytics/kubernetes/provider-catalog", response_model=KubernetesProviderCatalogResponse)
async def get_kubernetes_provider_catalog(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> KubernetesProviderCatalogResponse:
    """List provider regions and node shapes/sizes for K8s cluster planning."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    config = Config()
    runtime_credentials = _load_runtime_provider_credentials(customer_id)
    stored_rows = (
        db.query(CredentialRecord)
        .filter(
            CredentialRecord.customer_id == customer_id,
            CredentialRecord.is_valid.is_(True),
            CredentialRecord.provider.in_(["aws", "azure", "gcp", "oci"]),
        )
        .all()
    )
    credentials_for_catalog = _merge_runtime_with_stored_credentials(
        runtime_credentials,
        stored_rows,
    )
    providers = build_kubernetes_provider_catalog(
        config,
        runtime_credentials_by_provider=credentials_for_catalog,
    )
    return KubernetesProviderCatalogResponse(
        generated_at=_utcnow().isoformat() + "Z",
        providers={name: KubernetesProviderCatalogEntry(**payload) for name, payload in providers.items()},
    )


@router.get("/analytics/kubernetes/summary", response_model=KubernetesSummaryResponse)
async def get_kubernetes_summary(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> KubernetesSummaryResponse:
    """Overview of Kubernetes cost allocation status for this organization."""
    _ = current_user
    try:
        context = await asyncio.wait_for(_cost_context(membership, db, "month", "all"), timeout=8.0)
    except asyncio.TimeoutError:
        logger.info("Timed out collecting total cloud cost for Kubernetes summary; returning inventory-first response")
        context = {"total_cost": 0.0, "source": "timeout"}
    total = float(context.get("total_cost") or 0.0)
    container_services, provider_breakdown, data_source = await _build_kubernetes_container_service_rollups(
        membership,
        db,
        total,
    )
    container_total = round(sum(item.monthly_cost_usd for item in container_services), 2)
    provider_count = len([item for item in provider_breakdown if item.service_count > 0])
    highest_provider = next(
        (item.provider for item in provider_breakdown if item.service_count > 0),
        None,
    )
    highest_service = container_services[0] if container_services else None
    share = round((container_total / total) * 100, 2) if total > 0 else (100.0 if container_total > 0 else 0.0)
    configured_clusters = len([
        item for item in container_services
        if item.category == "managed_kubernetes"
    ])

    return KubernetesSummaryResponse(
        generated_at=_utcnow().isoformat() + "Z",
        kubernetes_enabled=len(container_services) > 0,
        clusters_configured=configured_clusters,
        estimated_k8s_share_percent=share,
        estimated_k8s_cost_usd=container_total,
        total_cloud_cost_usd=round(total, 2),
        container_service_count=len(container_services),
        provider_count_with_container_spend=provider_count,
        highest_cost_provider=highest_provider,
        highest_cost_service=highest_service,
        container_services=container_services,
        provider_breakdown=provider_breakdown,
        data_source=data_source,
        setup_hint=(
            "Connect provider credentials or import billing CSV rows with EKS, AKS, GKE, OKE, "
            "ECS, Fargate, Cloud Run, Container Registry, Docker, or similar service names. "
            "Use OpenCost for namespace and pod-level allocation."
        ),
        opencost_docs="https://www.opencost.io/docs/",
    )


# ---------------------------------------------------------------------------
# Virtual Tagging — apply tag rules to resources without touching the cloud
# ---------------------------------------------------------------------------

class VirtualTagRuleCreate(BaseModel):
    tag_key: str
    tag_value: str
    match_provider: Optional[str] = None
    match_service: Optional[str] = None
    match_region: Optional[str] = None
    match_account_id: Optional[str] = None
    match_resource_type: Optional[str] = None
    match_resource_name_contains: Optional[str] = None
    match_team: Optional[str] = None
    match_environment: Optional[str] = None
    priority: int = 100
    is_active: bool = True
    description: Optional[str] = None


class VirtualTagRuleOut(BaseModel):
    id: int
    tag_key: str
    tag_value: str
    match_provider: Optional[str]
    match_service: Optional[str]
    match_region: Optional[str]
    match_account_id: Optional[str]
    match_resource_type: Optional[str]
    match_resource_name_contains: Optional[str]
    match_team: Optional[str]
    match_environment: Optional[str]
    priority: int
    is_active: bool
    description: Optional[str]
    created_at: str
    updated_at: Optional[str]

    model_config = {"from_attributes": True}


class VirtualTagRulesResponse(BaseModel):
    organization_id: int
    total: int
    rules: List[VirtualTagRuleOut]


class VirtualTagPreviewItem(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    cost_usd: float
    applied_tags: Dict[str, str]   # virtual tag key → value
    match_rule_ids: List[int]


class VirtualTagPreviewResponse(BaseModel):
    organization_id: int
    generated_at: str
    total_resources: int
    tagged_resources: int
    coverage_percent: float
    preview: List[VirtualTagPreviewItem]


def _vtag_matches(rule: VirtualTagRule, item: Dict) -> bool:
    """Return True if a virtual tag rule matches a cost/resource item dict."""
    if rule.match_provider and rule.match_provider.lower() != str(item.get("provider", "")).lower():
        return False
    if rule.match_service and rule.match_service.lower() not in str(item.get("service", "")).lower():
        return False
    if rule.match_region and rule.match_region.lower() not in str(item.get("region", "")).lower():
        return False
    if rule.match_account_id and rule.match_account_id != str(item.get("account_id", "")):
        return False
    if rule.match_resource_type and rule.match_resource_type.lower() not in str(item.get("resource_type", "")).lower():
        return False
    if rule.match_resource_name_contains and rule.match_resource_name_contains.lower() not in str(item.get("resource_name", "")).lower():
        return False
    if rule.match_team and rule.match_team.lower() not in str(item.get("team", "")).lower():
        return False
    if rule.match_environment and rule.match_environment.lower() not in str(item.get("environment", "")).lower():
        return False
    return True


@router.get("/virtual-tags/rules", response_model=VirtualTagRulesResponse)
async def list_virtual_tag_rules(
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """List all virtual tag rules for the current organization."""
    org_id = membership.organization_id
    rules = (
        db.query(VirtualTagRule)
        .filter(VirtualTagRule.organization_id == org_id)
        .order_by(VirtualTagRule.priority.desc(), VirtualTagRule.created_at)
        .all()
    )
    out = []
    for r in rules:
        out.append(VirtualTagRuleOut(
            id=r.id,
            tag_key=r.tag_key,
            tag_value=r.tag_value,
            match_provider=r.match_provider,
            match_service=r.match_service,
            match_region=r.match_region,
            match_account_id=r.match_account_id,
            match_resource_type=r.match_resource_type,
            match_resource_name_contains=r.match_resource_name_contains,
            match_team=r.match_team,
            match_environment=r.match_environment,
            priority=r.priority,
            is_active=r.is_active,
            description=r.description,
            created_at=r.created_at.isoformat() if r.created_at else "",
            updated_at=r.updated_at.isoformat() if r.updated_at else None,
        ))
    return VirtualTagRulesResponse(organization_id=org_id, total=len(out), rules=out)


@router.post("/virtual-tags/rules", response_model=VirtualTagRuleOut, status_code=201)
async def create_virtual_tag_rule(
    payload: VirtualTagRuleCreate,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Create a virtual tag rule."""
    org_id = membership.organization_id
    customer_id = str(membership.organization_id)
    rule = VirtualTagRule(
        organization_id=org_id,
        customer_id=customer_id,
        tag_key=payload.tag_key.strip(),
        tag_value=payload.tag_value.strip(),
        match_provider=payload.match_provider or None,
        match_service=payload.match_service or None,
        match_region=payload.match_region or None,
        match_account_id=payload.match_account_id or None,
        match_resource_type=payload.match_resource_type or None,
        match_resource_name_contains=payload.match_resource_name_contains or None,
        match_team=payload.match_team or None,
        match_environment=payload.match_environment or None,
        priority=payload.priority,
        is_active=payload.is_active,
        description=payload.description,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return VirtualTagRuleOut(
        id=rule.id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        match_provider=rule.match_provider,
        match_service=rule.match_service,
        match_region=rule.match_region,
        match_account_id=rule.match_account_id,
        match_resource_type=rule.match_resource_type,
        match_resource_name_contains=rule.match_resource_name_contains,
        match_team=rule.match_team,
        match_environment=rule.match_environment,
        priority=rule.priority,
        is_active=rule.is_active,
        description=rule.description,
        created_at=rule.created_at.isoformat() if rule.created_at else "",
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.put("/virtual-tags/rules/{rule_id}", response_model=VirtualTagRuleOut)
async def update_virtual_tag_rule(
    rule_id: int,
    payload: VirtualTagRuleCreate,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Update a virtual tag rule."""
    org_id = membership.organization_id
    rule = db.query(VirtualTagRule).filter(
        VirtualTagRule.id == rule_id,
        VirtualTagRule.organization_id == org_id,
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Virtual tag rule not found")
    rule.tag_key = payload.tag_key.strip()
    rule.tag_value = payload.tag_value.strip()
    rule.match_provider = payload.match_provider or None
    rule.match_service = payload.match_service or None
    rule.match_region = payload.match_region or None
    rule.match_account_id = payload.match_account_id or None
    rule.match_resource_type = payload.match_resource_type or None
    rule.match_resource_name_contains = payload.match_resource_name_contains or None
    rule.match_team = payload.match_team or None
    rule.match_environment = payload.match_environment or None
    rule.priority = payload.priority
    rule.is_active = payload.is_active
    rule.description = payload.description
    db.commit()
    db.refresh(rule)
    return VirtualTagRuleOut(
        id=rule.id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        match_provider=rule.match_provider,
        match_service=rule.match_service,
        match_region=rule.match_region,
        match_account_id=rule.match_account_id,
        match_resource_type=rule.match_resource_type,
        match_resource_name_contains=rule.match_resource_name_contains,
        match_team=rule.match_team,
        match_environment=rule.match_environment,
        priority=rule.priority,
        is_active=rule.is_active,
        description=rule.description,
        created_at=rule.created_at.isoformat() if rule.created_at else "",
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.delete("/virtual-tags/rules/{rule_id}", status_code=204)
async def delete_virtual_tag_rule(
    rule_id: int,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Delete a virtual tag rule."""
    org_id = membership.organization_id
    rule = db.query(VirtualTagRule).filter(
        VirtualTagRule.id == rule_id,
        VirtualTagRule.organization_id == org_id,
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Virtual tag rule not found")
    db.delete(rule)
    db.commit()


@router.get("/virtual-tags/preview", response_model=VirtualTagPreviewResponse)
async def preview_virtual_tags(
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Preview which virtual tags would be applied to current cost records."""
    org_id = membership.organization_id
    now_str = _utcnow().isoformat()

    active_rules = (
        db.query(VirtualTagRule)
        .filter(
            VirtualTagRule.organization_id == org_id,
            VirtualTagRule.is_active == True,
        )
        .order_by(VirtualTagRule.priority.desc())
        .all()
    )

    # Gather candidate resources from inventory snapshots or imported cost records
    items: List[Dict] = []
    snapshots = (
        db.query(ProviderAccountSnapshot)
        .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
        .filter(ProviderAccount.organization_id == org_id)
        .order_by(ProviderAccountSnapshot.captured_at.desc())
        .limit(limit)
        .all()
    )
    for snap in snapshots:
        acct = snap.provider_account
        items.append({
            "resource_id": f"account:{snap.provider_account_id}",
            "resource_name": acct.account_name if acct else str(snap.provider_account_id),
            "resource_type": "Cloud Account",
            "provider": acct.provider if acct else "unknown",
            "region": acct.native_region if acct else "",
            "service": "",
            "account_id": acct.account_identifier if acct else "",
            "cost_usd": float(snap.direct_cost_usd or 0),
            "team": "",
            "environment": "",
        })

    if not items:
        records = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .limit(limit)
            .all()
        )
        for rec in records:
            items.append({
                "resource_id": f"imported:{rec.id}",
                "resource_name": rec.account_name or rec.service_name or "unknown",
                "resource_type": rec.service_name or "Imported Service",
                "provider": rec.provider or "unknown",
                "region": rec.region or "",
                "service": rec.service_name or "",
                "account_id": rec.account_identifier or "",
                "cost_usd": float(rec.cost_usd or 0),
                "team": "",
                "environment": "",
            })

    # Apply rules to each item
    preview: List[VirtualTagPreviewItem] = []
    tagged_count = 0
    for item in items[:limit]:
        applied: Dict[str, str] = {}
        matched_ids: List[int] = []
        for rule in active_rules:
            if _vtag_matches(rule, item):
                if rule.tag_key not in applied:  # first matching rule wins per key
                    applied[rule.tag_key] = rule.tag_value
                    matched_ids.append(rule.id)
        if applied:
            tagged_count += 1
        preview.append(VirtualTagPreviewItem(
            resource_id=item["resource_id"],
            resource_name=item["resource_name"],
            resource_type=item["resource_type"],
            provider=item["provider"],
            region=item["region"],
            cost_usd=item["cost_usd"],
            applied_tags=applied,
            match_rule_ids=matched_ids,
        ))

    total = len(preview)
    coverage = round((tagged_count / total * 100) if total > 0 else 0.0, 1)
    return VirtualTagPreviewResponse(
        organization_id=org_id,
        generated_at=now_str,
        total_resources=total,
        tagged_resources=tagged_count,
        coverage_percent=coverage,
        preview=preview,
    )


# ---------------------------------------------------------------------------
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Rightsizing — resource-level recommendations
#
# 4-tier data source waterfall (best available wins):
#  1. AWS Cost Explorer  get_rightsizing_recommendation (real API)
#  2. Azure Advisor      rightsizing recommendations    (real API)
#  3. Cost-trend signal  analysis from snapshot history (deterministic)
#  4. Imported CSV cost  signal analysis                (deterministic)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class RightsizingRecommendation(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    account_id: str
    current_size: str
    recommended_size: str
    current_monthly_cost_usd: float
    projected_monthly_cost_usd: float
    monthly_savings_usd: float
    annual_savings_usd: float
    cpu_utilization_avg_percent: Optional[float]
    memory_utilization_avg_percent: Optional[float]
    reason: str
    confidence: str   # "high" | "medium" | "low"
    effort: str       # "low" | "medium" | "high"
    action: str       # "downsize" | "terminate" | "reserve" | "modernize"
    evidence_source: str = "unspecified"
    analysis_points: int = 0
    trend_slope_usd: float = 0.0
    trend_percent: float = 0.0
    latest_monthly_cost_usd: Optional[float] = None
    peak_monthly_cost_usd: Optional[float] = None
    top_regions: List[str] = Field(default_factory=list)
    regional_breakdown: List[Dict[str, Any]] = Field(default_factory=list)
    resource_console_url: Optional[str] = None
    last_observed_at: Optional[str] = None
    risk_note: Optional[str] = None
    provider_recommendation_type: Optional[str] = None
    provider_recommendation_name: Optional[str] = None
    provider_recommendation_category: Optional[str] = None
    provider_recommendation_status: Optional[str] = None
    provider_recommendation_importance: Optional[str] = None
    provider_recommendation_resource_count: Optional[int] = None


class RightsizingResponse(BaseModel):
    generated_at: str
    organization_id: int
    data_source: str
    total_resources_analyzed: int
    rightsizable_count: int
    total_monthly_savings_usd: float
    total_annual_savings_usd: float
    recommendations: List[RightsizingRecommendation]


class RecommendationLedgerItem(BaseModel):
    id: int
    organization_id: int
    provider: str
    resource_id: str
    resource_name: Optional[str] = None
    resource_type: Optional[str] = None
    account_id: Optional[str] = None
    region: Optional[str] = None
    recommendation_source: str
    recommendation_fingerprint: str
    action: str
    confidence: str
    effort: str
    status: str
    owner: Optional[str] = None
    current_size: Optional[str] = None
    recommended_size: Optional[str] = None
    current_monthly_cost_usd: float
    projected_monthly_cost_usd: float
    planned_monthly_savings_usd: float
    planned_annual_savings_usd: float
    realized_monthly_savings_usd: float
    realized_annual_savings_usd: float
    variance_monthly_usd: float
    variance_annual_usd: float
    variance_percent: float
    variance_reason: Optional[str] = None
    reason: Optional[str] = None
    resource_console_url: Optional[str] = None
    first_seen_at: Optional[str] = None
    last_seen_at: Optional[str] = None
    planned_at: Optional[str] = None
    realized_at: Optional[str] = None
    last_exported_at: Optional[str] = None
    times_seen: int


class RecommendationLedgerResponse(BaseModel):
    generated_at: str
    organization_id: int
    total_count: int
    total_planned_monthly_savings_usd: float
    total_realized_monthly_savings_usd: float
    total_variance_monthly_usd: float
    total_planned_annual_savings_usd: float
    total_realized_annual_savings_usd: float
    total_variance_annual_usd: float
    items: List[RecommendationLedgerItem]


class RecommendationLedgerFinanceUpdateRequest(BaseModel):
    realized_monthly_savings_usd: Optional[float] = Field(default=None, ge=0)
    realized_annual_savings_usd: Optional[float] = Field(default=None, ge=0)
    variance_reason: Optional[str] = None
    status: Optional[Literal["open", "planned", "approved", "executed", "verified", "rejected", "expired"]] = None
    owner: Optional[str] = None
    realized_at: Optional[str] = None


# ---------------------------------------------------------------------------
# Static reference data
# ---------------------------------------------------------------------------

_DOWNSIZE_MAP: Dict[str, Dict[str, str]] = {
    "aws": {
        "m5.4xlarge": "m5.2xlarge",   "m5.2xlarge": "m5.xlarge",   "m5.xlarge": "m5.large",
        "m5.large":   "m5.medium",
        "m6i.4xlarge": "m6i.2xlarge", "m6i.2xlarge": "m6i.xlarge", "m6i.xlarge": "m6i.large",
        "c5.4xlarge": "c5.2xlarge",   "c5.2xlarge": "c5.xlarge",   "c5.xlarge": "c5.large",
        "c6i.4xlarge": "c6i.2xlarge", "c6i.2xlarge": "c6i.xlarge",
        "r5.4xlarge": "r5.2xlarge",   "r5.2xlarge": "r5.xlarge",   "r5.xlarge": "r5.large",
        "t3.2xlarge": "t3.xlarge",    "t3.xlarge": "t3.large",     "t3.large": "t3.medium",
    },
    "azure": {
        "Standard_D16s_v3": "Standard_D8s_v3",  "Standard_D8s_v3": "Standard_D4s_v3",
        "Standard_D4s_v3":  "Standard_D2s_v3",  "Standard_D2s_v3": "Standard_B2s",
        "Standard_E16s_v3": "Standard_E8s_v3",  "Standard_E8s_v3": "Standard_E4s_v3",
        "Standard_E4s_v3":  "Standard_E2s_v3",
        "Standard_F16s_v2": "Standard_F8s_v2",  "Standard_F8s_v2": "Standard_F4s_v2",
    },
    "gcp": {
        "n1-standard-16": "n1-standard-8",  "n1-standard-8": "n1-standard-4",
        "n1-standard-4":  "n1-standard-2",
        "n2-standard-16": "n2-standard-8",  "n2-standard-8": "n2-standard-4",
        "n2-standard-4":  "n2-standard-2",
        "e2-standard-8":  "e2-standard-4",  "e2-standard-4": "e2-standard-2",
    },
    "oci": {
        "VM.Standard2.16": "VM.Standard2.8",  "VM.Standard2.8": "VM.Standard2.4",
        "VM.Standard2.4":  "VM.Standard2.2",
        "VM.Standard3.Flex.16": "VM.Standard3.Flex.8", "VM.Standard3.Flex.8": "VM.Standard3.Flex.4",
    },
}

# Savings rates by action type (realistic market benchmarks)
_ACTION_SAVINGS_RATES: Dict[str, float] = {
    "downsize":  0.45,   # ~50% savings dropping one size tier
    "terminate": 1.00,   # full elimination of orphaned resource
    "reserve":   0.37,   # typical 1yr No-Upfront RI discount vs on-demand
    "modernize": 0.30,   # newer gen typically 10-30% cheaper at same perf
}


def _series_mean(values: List[float]) -> Optional[float]:
    cleaned = [float(v) for v in values if v is not None]
    if not cleaned:
        return None
    return round(sum(cleaned) / len(cleaned), 2)


def _estimate_monthly_cost_from_size(provider: str, size: str) -> float:
    text = str(size or "").lower()
    if not text:
        return 120.0

    # Coarse but deterministic estimate for rightsize ranking without price API calls.
    if provider == "aws":
        if "metal" in text or "12xlarge" in text or "16xlarge" in text:
            return 1800.0
        if "8xlarge" in text:
            return 980.0
        if "4xlarge" in text:
            return 620.0
        if "2xlarge" in text:
            return 320.0
        if "xlarge" in text:
            return 180.0
        if "large" in text:
            return 110.0
        if "medium" in text:
            return 70.0
        return 140.0

    if provider == "azure":
        if "16" in text:
            return 760.0
        if "8" in text:
            return 390.0
        if "4" in text:
            return 210.0
        if "2" in text:
            return 120.0
        return 170.0

    if provider == "gcp":
        if "standard-32" in text:
            return 980.0
        if "standard-16" in text:
            return 520.0
        if "standard-8" in text:
            return 280.0
        if "standard-4" in text:
            return 160.0
        if "standard-2" in text:
            return 95.0
        return 150.0

    return 140.0


def _extract_size_from_account_metadata(account: Any, provider: str) -> str:
    raw_meta = getattr(account, "metadata_json", "{}") or "{}"
    try:
        metadata = json.loads(raw_meta)
    except Exception:
        metadata = {}
    candidates = [
        metadata.get("instance_shape"),
        metadata.get("shape"),
        metadata.get("current_shape"),
        metadata.get("vm_size"),
        metadata.get("machine_type"),
        metadata.get("resource_shape"),
    ]
    for candidate in candidates:
        text = str(candidate or "").strip()
        if text:
            return text
    return next(iter(_DOWNSIZE_MAP.get(provider, {"compute": "smaller-compute"})))


def _recommended_size_for_action(provider: str, action: str, current_size: str) -> str:
    size_map = _DOWNSIZE_MAP.get(provider, {})
    if action == "terminate":
        return "N/A — terminate"
    if action == "reserve":
        return f"{current_size} (Reserved 1yr)"
    if action == "modernize":
        modern = size_map.get(current_size)
        return modern or f"newer-gen-{current_size}"
    down = size_map.get(current_size)
    return down or f"smaller-{current_size}"


def _estimate_oci_instance_monthly_cost(
    shape_name: str,
    ocpus: Optional[float],
    memory_gib: Optional[float],
) -> float:
    ocpu_value = float(ocpus or 0.0)
    memory_value = float(memory_gib or 0.0)
    if ocpu_value <= 0:
        # Non-flex shape fallback estimate.
        for token in shape_name.split("."):
            try:
                ocpu_value = float(token)
                break
            except ValueError:
                continue
    if memory_value <= 0:
        memory_value = max(8.0, ocpu_value * 8.0) if ocpu_value > 0 else 16.0

    estimated = (ocpu_value * 24.0) + (memory_value * 1.6)
    shape_upper = str(shape_name or "").upper()
    if ".A1." in shape_upper:
        estimated *= 0.65
    elif ".E5." in shape_upper:
        estimated *= 1.10
    elif ".E4." in shape_upper:
        estimated *= 1.0
    return round(max(25.0, estimated), 2)


def _oci_rightsize_target(
    shape_name: str,
    ocpus: Optional[float],
    memory_gib: Optional[float],
) -> tuple[Optional[str], Optional[str], float]:
    shape = str(shape_name or "").strip()
    if not shape:
        return None, None, 0.0

    ocpu_value = float(ocpus or 0.0) if ocpus is not None else 0.0
    memory_value = float(memory_gib or 0.0) if memory_gib is not None else 0.0
    if ".Flex" in shape and ocpu_value > 1.0:
        target_ocpu = round(max(1.0, ocpu_value * 0.5), 1)
        baseline_memory = memory_value if memory_value > 0 else (ocpu_value * 8.0)
        target_memory = round(max(8.0, baseline_memory * 0.65), 1)
        current_size = f"{shape} ({ocpu_value:g} OCPU/{baseline_memory:g} GB)"
        recommended_size = f"{shape} ({target_ocpu:g} OCPU/{target_memory:g} GB)"
        savings_rate = round(min(0.6, max(0.25, 1.0 - (target_ocpu / max(ocpu_value, 0.1)))), 2)
        return current_size, recommended_size, savings_rate

    down_map = _DOWNSIZE_MAP.get("oci", {})
    if shape in down_map:
        return shape, down_map[shape], _ACTION_SAVINGS_RATES["downsize"]
    return None, None, 0.0


def _rightsizing_console_url(
    provider: str,
    resource_id: str,
    region: str,
    account_id: str,
    resource_type: str,
) -> Optional[str]:
    prov = str(provider or "").strip().lower()
    rid = str(resource_id or "").strip()
    rgn = str(region or "").strip()
    acct = str(account_id or "").strip()
    rtype = str(resource_type or "").strip().lower()
    safe_region = rgn if rgn and rgn.lower() not in {"global", "unknown", "n/a"} else ""

    if prov == "aws":
        region_part = safe_region or "us-east-1"
        if rid.startswith("i-"):
            return (
                f"https://{region_part}.console.aws.amazon.com/ec2/home"
                f"?region={quote(region_part, safe='')}"
                f"#InstanceDetails:instanceId={quote(rid, safe='')}"
            )
        if rid.startswith("vol-"):
            return (
                f"https://{region_part}.console.aws.amazon.com/ec2/home"
                f"?region={quote(region_part, safe='')}"
                f"#VolumeDetails:volumeId={quote(rid, safe='')}"
            )
        return (
            f"https://{region_part}.console.aws.amazon.com/ec2/home"
            f"?region={quote(region_part, safe='')}#Instances:"
        )

    if prov == "azure":
        if rid.startswith("/subscriptions/"):
            return f"https://portal.azure.com/#resource{rid}/overview"
        return "https://portal.azure.com/#view/HubsExtension/BrowseAllResources"

    if prov == "gcp":
        if rid.startswith("projects/") and "/zones/" in rid and "/instances/" in rid:
            parts = rid.split("/")
            # projects/{project}/zones/{zone}/instances/{name}
            if len(parts) >= 6:
                project = parts[1]
                zone = parts[3]
                name = parts[5]
                return (
                    "https://console.cloud.google.com/compute/instancesDetail/zones/"
                    f"{quote(zone, safe='')}/instances/{quote(name, safe='')}"
                    f"?project={quote(project, safe='')}"
                )
        project = acct if acct and not acct.startswith("ocid1.") else ""
        if project:
            return f"https://console.cloud.google.com/compute/instances?project={quote(project, safe='')}"
        return "https://console.cloud.google.com/compute/instances"

    if prov == "oci":
        oci_region_suffix = f"?region={quote(safe_region, safe='')}" if safe_region else ""
        if (
            rid.startswith("ocid1.bootvolume.")
            or "bootvolume" in rtype
            or "boot volume" in rtype
        ):
            if rid.startswith("ocid1.bootvolume."):
                return f"https://cloud.oracle.com/block-storage/boot-volumes/{quote(rid, safe='')}{oci_region_suffix}"
            return f"https://cloud.oracle.com/block-storage/boot-volumes{oci_region_suffix}"
        if (
            rid.startswith("ocid1.volume.")
            or "blockvolume" in rtype
            or "block volume" in rtype
            or rtype == "volume"
        ):
            if rid.startswith("ocid1.volume."):
                return f"https://cloud.oracle.com/block-storage/volumes/{quote(rid, safe='')}{oci_region_suffix}"
            return f"https://cloud.oracle.com/block-storage/volumes{oci_region_suffix}"
        if (
            rid.startswith("ocid1.bucket.")
            or "objectstorage" in rtype
            or "object storage" in rtype
            or "bucket" in rtype
        ):
            return f"https://cloud.oracle.com/object-storage/buckets{oci_region_suffix}"
        if rid.startswith("ocid1.loadbalancer.") or "load balancer" in rtype or "loadbalancer" in rtype:
            if rid.startswith("ocid1.loadbalancer."):
                return f"https://cloud.oracle.com/networking/load-balancers/{quote(rid, safe='')}{oci_region_suffix}"
            return f"https://cloud.oracle.com/networking/load-balancers{oci_region_suffix}"
        if rid.startswith("ocid1.autonomousdatabase.") or "autonomous" in rtype:
            if rid.startswith("ocid1.autonomousdatabase."):
                return f"https://cloud.oracle.com/db/adbs/{quote(rid, safe='')}{oci_region_suffix}"
            return f"https://cloud.oracle.com/db/adbs{oci_region_suffix}"
        if rid.startswith("ocid1.instance."):
            return f"https://cloud.oracle.com/compute/instances/{quote(rid, safe='')}{oci_region_suffix}"
        if rid.startswith("ocid1."):
            return f"https://cloud.oracle.com/resources{oci_region_suffix}"
        # Account-level signal: link to compute inventory in selected region.
        base = "https://cloud.oracle.com/compute/instances"
        return f"{base}{oci_region_suffix}" if safe_region else base

    return None


def _slugify_resource_token(value: str) -> str:
    token = re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower())
    return token.strip("-") or "recommendation"


def _safe_recommendation_float(value: Any, default: float = 0.0) -> float:
    try:
        if isinstance(value, dict):
            units = float(value.get("units") or 0.0)
            nanos = float(value.get("nanos") or 0.0) / 1_000_000_000
            return units + nanos
        return float(str(value or "").replace("$", "").replace(",", "").strip() or default)
    except (TypeError, ValueError):
        return default


def _recommendation_monthly_savings(row: Dict[str, Any]) -> float:
    monthly = _safe_recommendation_float(row.get("monthly_savings_usd"), -1.0)
    if monthly >= 0:
        return round(monthly, 2)
    annual = _safe_recommendation_float(row.get("savings_annual_usd"), 0.0)
    return round(annual / 12.0, 2)


def _recommendation_annual_savings(row: Dict[str, Any]) -> float:
    annual = _safe_recommendation_float(row.get("savings_annual_usd"), -1.0)
    if annual >= 0:
        return round(annual, 2)
    return round(_recommendation_monthly_savings(row) * 12.0, 2)


def _provider_from_recommendation_row(row: Dict[str, Any], fallback: str = "all") -> str:
    provider = str(row.get("provider") or "").strip().lower()
    if provider and provider not in {"all", "multi-cloud"}:
        return provider
    rec_id = str(row.get("id") or "").strip().lower()
    if "-rec-" in rec_id:
        candidate = rec_id.split("-rec-", 1)[0]
        if candidate:
            return candidate
    fallback_provider = str(fallback or "all").strip().lower() or "all"
    return fallback_provider


def _dedupe_dashboard_recommendations(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        key = (
            str(row.get("cloud") or "").lower(),
            str(row.get("source") or "").lower(),
            str(row.get("service") or "").lower(),
            str(row.get("resource_id") or "").lower(),
            _slugify_resource_token(str(row.get("title") or row.get("description") or "")),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    out.sort(key=lambda item: float(item.get("savings") or 0), reverse=True)
    return out


def _provider_recommendation_type_from_text(
    *,
    rec_type: str = "",
    service: str = "",
    description: str = "",
    action: str = "",
) -> str:
    text = " ".join([rec_type, service, description, action]).lower()
    if "reservation" in text or "reserved" in text or "savings plan" in text or "commit" in text:
        return "reserved-instances"
    if (
        "unattached" in text
        or "detached" in text
        or "not attached" in text
        or "orphan" in text
        or "idle" in text
        or "delete" in text
        or "terminate" in text
        or ("volume" in text and "attachment" in text)
    ):
        return "idle-resources"
    if (
        "storage" in text
        or "object-store" in text
        or "object storage" in text
        or "blockvolume" in text
        or "bootvolume" in text
        or "bucket" in text
        or "volume" in text
        or "archive" in text
        or "lifecycle" in text
        or "tier" in text
    ):
        return "storage-optimization"
    if "rightsiz" in text or "right-siz" in text or "resize" in text or "machine type" in text or "underutilized" in text:
        return "rightsizing"
    fallback = str(rec_type or "optimization").strip().lower() or "optimization"
    if fallback.startswith("ocid1."):
        return "optimization"
    return fallback


def _provider_row(
    *,
    provider: str,
    source: str,
    rec_id: str,
    service: str,
    rec_type: str,
    description: str,
    monthly_savings: float,
    account_id: str = "",
    region: str = "global",
    resource_id: str = "",
    resource_name: str = "",
    resource_type: str = "",
    current_monthly_cost: Optional[float] = None,
    payback_months: float = 1.0,
    confidence: str = "high",
    severity: str = "medium",
    roi_percent: float = 0.0,
    console_url: Optional[str] = None,
    provider_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    safe_monthly = round(max(float(monthly_savings or 0.0), 0.0), 2)
    current_monthly = (
        round(max(float(current_monthly_cost or 0.0), 0.0), 2)
        if current_monthly_cost is not None
        else safe_monthly
    )
    row = {
        "id": rec_id,
        "provider": provider,
        "source": source,
        "type": rec_type,
        "service": service or resource_type or "Cloud Service",
        "description": description or "Provider optimization recommendation.",
        "current_annual_spend": round(current_monthly * 12.0, 2),
        "monthly_savings_usd": safe_monthly,
        "savings_annual_usd": round(safe_monthly * 12.0, 2),
        "payback_months": payback_months,
        "severity": severity,
        "roi_percent": roi_percent,
        "confidence": confidence if confidence in {"high", "medium", "low"} else "medium",
        "account_id": account_id,
        "region": region or "global",
        "resource_id": resource_id,
        "resource_name": resource_name,
        "resource_type": resource_type or service,
        "resource_console_url": console_url,
    }
    if provider_metadata:
        row.update(provider_metadata)
    return row


def _recommendation_rows_from_rightsizing(
    recs: List[RightsizingRecommendation],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for rec in recs:
        rows.append(_provider_row(
            provider=rec.provider,
            source=rec.evidence_source,
            rec_id=f"{rec.provider}-{rec.evidence_source}-{_slugify_resource_token(rec.resource_id)}",
            service=rec.resource_type,
            rec_type=rec.action,
            description=rec.reason or rec.resource_name,
            monthly_savings=rec.monthly_savings_usd,
            account_id=rec.account_id,
            region=rec.region,
            resource_id=rec.resource_id,
            resource_name=rec.resource_name,
            resource_type=rec.resource_type,
            current_monthly_cost=rec.current_monthly_cost_usd,
            payback_months=1.0 if rec.effort == "low" else 3.0 if rec.effort == "medium" else 6.0,
            confidence=rec.confidence,
            severity="high" if rec.confidence == "high" else "medium",
            roi_percent=round(
                (rec.monthly_savings_usd / rec.current_monthly_cost_usd) * 100,
                2,
            ) if rec.current_monthly_cost_usd > 0 else 0.0,
            console_url=rec.resource_console_url,
        ))
    return rows


def _aws_cost_explorer_commitment_recommendations(
    cred_json: Dict[str, Any],
    min_monthly_savings: float,
    limit: int,
) -> List[Dict[str, Any]]:
    try:
        import boto3

        client = boto3.client(
            "ce",
            aws_access_key_id=cred_json.get("access_key_id"),
            aws_secret_access_key=cred_json.get("secret_access_key"),
            region_name=cred_json.get("region", "us-east-1"),
        )
    except Exception as exc:
        logger.warning("AWS recommendation client unavailable: %s", exc)
        return []

    rows: List[Dict[str, Any]] = []
    try:
        response = client.get_savings_plans_purchase_recommendation(
            SavingsPlansType="COMPUTE_SP",
            TermInYears="ONE_YEAR",
            PaymentOption="NO_UPFRONT",
            LookbackPeriodInDays="THIRTY_DAYS",
        )
        details = (
            (response.get("SavingsPlansPurchaseRecommendation") or {}).get(
                "SavingsPlansPurchaseRecommendationDetails",
                [],
            )
            or response.get("SavingsPlansPurchaseRecommendationDetails", [])
            or []
        )
        for idx, detail in enumerate(details, start=1):
            if len(rows) >= limit:
                break
            monthly = _safe_recommendation_float(
                detail.get("EstimatedMonthlySavingsAmount")
                or detail.get("EstimatedSavingsAmount")
                or detail.get("MonthlySavings")
            )
            if monthly < min_monthly_savings:
                continue
            account_id = str(detail.get("AccountId") or detail.get("AccountScope") or "")
            commitment = str(detail.get("HourlyCommitmentToPurchase") or "recommended commitment")
            rows.append(_provider_row(
                provider="aws",
                source="aws_cost_explorer_savings_plans",
                rec_id=f"aws-ce-savings-plan-{idx:03d}",
                service="Savings Plans",
                rec_type="reserved-instances",
                description=f"AWS: Buy Compute Savings Plan commitment {commitment} for steady usage.",
                monthly_savings=monthly,
                account_id=account_id,
                region=str(cred_json.get("region") or "global"),
                payback_months=3,
                confidence="high",
                severity="high",
                roi_percent=_safe_recommendation_float(detail.get("EstimatedROI")),
            ))
    except Exception as exc:
        logger.info("AWS Savings Plans recommendations unavailable: %s", exc)

    try:
        response = client.get_reservation_purchase_recommendation(
            Service="Amazon Elastic Compute Cloud - Compute",
            LookbackPeriodInDays="THIRTY_DAYS",
            TermInYears="ONE_YEAR",
            PaymentOption="NO_UPFRONT",
        )
        recs = response.get("Recommendations", []) or []
        for rec_idx, rec in enumerate(recs, start=1):
            details = rec.get("RecommendationDetails") or []
            for detail_idx, detail in enumerate(details, start=1):
                if len(rows) >= limit:
                    break
                monthly = _safe_recommendation_float(
                    detail.get("EstimatedMonthlySavingsAmount")
                    or detail.get("EstimatedSavingsAmount")
                    or detail.get("MonthlySavings")
                )
                if monthly < min_monthly_savings:
                    continue
                instance_details = detail.get("InstanceDetails") or {}
                ec2_details = instance_details.get("EC2InstanceDetails") or {}
                family = str(ec2_details.get("Family") or ec2_details.get("InstanceType") or "EC2")
                account_id = str(rec.get("AccountId") or detail.get("AccountId") or "")
                rows.append(_provider_row(
                    provider="aws",
                    source="aws_cost_explorer_reservations",
                    rec_id=f"aws-ce-reservation-{rec_idx:03d}-{detail_idx:03d}",
                    service="EC2 Reserved Instances",
                    rec_type="reserved-instances",
                    description=f"AWS: Purchase EC2 Reserved Instances for steady {family} usage.",
                    monthly_savings=monthly,
                    account_id=account_id,
                    region=str(cred_json.get("region") or "global"),
                    payback_months=3,
                    confidence="high",
                    severity="high",
                    roi_percent=_safe_recommendation_float(detail.get("EstimatedROI")),
                ))
            if len(rows) >= limit:
                break
    except Exception as exc:
        logger.info("AWS reservation recommendations unavailable: %s", exc)

    return rows[:limit]


def _azure_advisor_recommendation_rows(
    cred_json: Dict[str, Any],
    min_monthly_savings: float,
    limit: int,
) -> List[Dict[str, Any]]:
    try:
        from azure.identity import ClientSecretCredential
        from azure.mgmt.advisor import AdvisorManagementClient

        azure_cred = ClientSecretCredential(
            tenant_id=cred_json["tenant_id"],
            client_id=cred_json["client_id"],
            client_secret=cred_json["client_secret"],
        )
        subscription_id = cred_json["subscription_id"]
        client = AdvisorManagementClient(azure_cred, subscription_id)
    except Exception as exc:
        logger.warning("Azure Advisor recommendation client unavailable: %s", exc)
        return []

    rows: List[Dict[str, Any]] = []
    try:
        for idx, rec in enumerate(client.recommendations.list(), start=1):
            if len(rows) >= limit:
                break
            if "cost" not in str(getattr(rec, "category", "") or "").lower():
                continue
            props = getattr(rec, "extended_properties", None) or {}
            savings_value = (
                props.get("annualSavingsAmount")
                or props.get("savingsAmount")
                or props.get("monthlySavingsAmount")
                or 0
            )
            savings = _safe_recommendation_float(savings_value)
            monthly = savings / 12.0 if "annual" in " ".join(props.keys()).lower() else savings
            if monthly < min_monthly_savings:
                continue
            resource_id = str(getattr(rec, "resource_id", "") or "")
            short = getattr(rec, "short_description", None)
            problem = str(getattr(short, "problem", "") or "").strip()
            solution = str(getattr(short, "solution", "") or "").strip()
            description = solution or problem or str(getattr(rec, "name", "") or "Azure Advisor cost recommendation")
            service = str(props.get("targetResourceType") or props.get("resourceType") or "Azure Advisor").split("/")[-1]
            rec_type = _provider_recommendation_type_from_text(
                service=service,
                description=description,
                action=str(props.get("recommendationType") or ""),
            )
            impact = str(getattr(rec, "impact", "") or "medium").lower()
            confidence = "high" if impact == "high" else "medium" if impact == "medium" else "low"
            rows.append(_provider_row(
                provider="azure",
                source="azure_advisor",
                rec_id=str(getattr(rec, "id", "") or getattr(rec, "name", "") or f"azure-advisor-{idx:03d}"),
                service=service,
                rec_type=rec_type,
                description=f"Azure: {description}",
                monthly_savings=monthly,
                account_id=str(subscription_id),
                region=str(getattr(rec, "location", "") or "global"),
                resource_id=resource_id,
                resource_name=str(props.get("resourceName") or (resource_id.split("/")[-1] if resource_id else "")),
                resource_type=service,
                payback_months=1 if rec_type in {"idle-resources", "storage-optimization"} else 3,
                confidence=confidence,
                severity="high" if impact == "high" else "medium",
            ))
    except Exception as exc:
        logger.warning("Azure Advisor recommendation collection failed: %s", exc)
    return rows[:limit]


def _gcp_recommender_rows(
    cred_json: Dict[str, Any],
    min_monthly_savings: float,
    limit: int,
) -> List[Dict[str, Any]]:
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import AuthorizedSession
    except Exception as exc:
        logger.warning("GCP recommendation libraries unavailable: %s", exc)
        return []

    service_account_json = cred_json.get("service_account_json")
    if isinstance(service_account_json, str):
        try:
            service_account_json = json.loads(service_account_json)
        except Exception:
            service_account_json = {}
    if not isinstance(service_account_json, dict):
        return []
    project_id = str(cred_json.get("project_id") or service_account_json.get("project_id") or "").strip()
    if not project_id:
        return []

    try:
        credentials = service_account.Credentials.from_service_account_info(
            service_account_json,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        session = AuthorizedSession(credentials)
    except Exception as exc:
        logger.warning("GCP recommendation credentials unavailable: %s", exc)
        return []

    locations = {"global"}
    try:
        response = session.get(
            f"https://compute.googleapis.com/compute/v1/projects/{project_id}/zones",
            params={"maxResults": 500},
            timeout=20,
        )
        if response.status_code < 400:
            for zone in response.json().get("items", []) or []:
                name = str(zone.get("name") or "").strip()
                if name:
                    locations.add(name)
    except Exception:
        pass

    recommender_ids = [
        "google.compute.instance.MachineTypeRecommender",
        "google.compute.instance.IdleResourceRecommender",
        "google.compute.disk.IdleResourceRecommender",
        "google.compute.commitment.UsageCommitmentRecommender",
        "google.cloudstorage.bucket.SoftDeleteRecommender",
    ]
    rows: List[Dict[str, Any]] = []
    for location in sorted(locations)[:16]:
        for recommender_id in recommender_ids:
            if len(rows) >= limit:
                break
            page_token = ""
            while len(rows) < limit:
                params = {"pageSize": 100}
                if page_token:
                    params["pageToken"] = page_token
                url = (
                    f"https://recommender.googleapis.com/v1/projects/{project_id}/locations/"
                    f"{quote(location, safe='')}/recommenders/{quote(recommender_id, safe='')}/recommendations"
                )
                try:
                    response = session.get(url, params=params, timeout=10)
                except Exception as exc:
                    logger.info("GCP Recommender request failed for %s/%s: %s", location, recommender_id, exc)
                    break
                if response.status_code >= 400:
                    break
                payload = response.json()
                for rec in payload.get("recommendations", []) or []:
                    impact = rec.get("primaryImpact") or {}
                    if str(impact.get("category") or "").upper() != "COST":
                        continue
                    cost_projection = impact.get("costProjection") or {}
                    cost = _safe_recommendation_float(cost_projection.get("cost"))
                    monthly = abs(cost)
                    if monthly < min_monthly_savings:
                        continue
                    rec_name = str(rec.get("name") or "")
                    subtype = str(rec.get("recommenderSubtype") or recommender_id.split(".")[-1])
                    description = str(rec.get("description") or subtype or "GCP Recommender cost recommendation")
                    rec_type = _provider_recommendation_type_from_text(
                        rec_type=subtype,
                        service=recommender_id,
                        description=description,
                    )
                    resource_id = ""
                    try:
                        operations = (
                            (rec.get("content") or {})
                            .get("operationGroups", [{}])[0]
                            .get("operations", [])
                        )
                        resource_id = str((operations[0] or {}).get("resource") or "")
                    except Exception:
                        resource_id = ""
                    rows.append(_provider_row(
                        provider="gcp",
                        source="gcp_recommender",
                        rec_id=rec_name or f"gcp-recommender-{_slugify_resource_token(location)}-{len(rows) + 1:03d}",
                        service=recommender_id,
                        rec_type=rec_type,
                        description=f"GCP: {description}",
                        monthly_savings=monthly,
                        account_id=project_id,
                        region=location,
                        resource_id=resource_id,
                        resource_name=resource_id.split("/")[-1] if resource_id else "",
                        resource_type=recommender_id,
                        payback_months=1 if rec_type in {"idle-resources", "storage-optimization"} else 3,
                        confidence="high",
                        severity="high" if str(rec.get("priority") or "").upper() == "P1" else "medium",
                    ))
                page_token = str(payload.get("nextPageToken") or "")
                if not page_token:
                    break
        if len(rows) >= limit:
            break
    return rows[:limit]


def _oci_home_region(
    oci_module: Any,
    oci_config: Dict[str, Any],
    tenancy_id: str,
    timeout: tuple[int, int] = (5, 15),
) -> Optional[str]:
    """Return the tenancy home region for Cloud Advisor calls when discoverable."""
    if not tenancy_id:
        return None
    try:
        identity = oci_module.identity.IdentityClient(oci_config, timeout=timeout)
        response = identity.list_region_subscriptions(tenancy_id)
        for subscription in getattr(response, "data", []) or []:
            if bool(getattr(subscription, "is_home_region", False)):
                region_name = str(getattr(subscription, "region_name", "") or "").strip()
                if region_name:
                    return region_name
    except Exception as exc:
            logger.info("Unable to discover OCI home region for optimizer recommendations: %s", exc)
    return None


def _oci_config_for_region(oci_config: Dict[str, Any], region: Optional[str]) -> Dict[str, Any]:
    region_name = str(region or "").strip()
    if not region_name:
        return dict(oci_config)
    region_config = dict(oci_config)
    region_config["region"] = region_name
    return region_config


def _oci_subscribed_regions(
    oci_module: Any,
    oci_config: Dict[str, Any],
    tenancy_id: str,
    *,
    home_region: Optional[str] = None,
    timeout: tuple[int, int] = (5, 15),
) -> List[str]:
    """Return all subscribed OCI regions, ordered with the home region first."""
    fallback_region = str(oci_config.get("region") or "").strip()
    preferred_home = str(home_region or "").strip()
    region_names: List[str] = []
    seen: set[str] = set()

    def add_region(region_name: str) -> None:
        normalized = str(region_name or "").strip()
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        region_names.append(normalized)

    if preferred_home:
        add_region(preferred_home)

    try:
        identity = oci_module.identity.IdentityClient(
            _oci_config_for_region(oci_config, preferred_home or fallback_region),
            timeout=timeout,
        )
        response = identity.list_region_subscriptions(tenancy_id)
        subscriptions = list(getattr(response, "data", []) or [])
        for subscription in subscriptions:
            if bool(getattr(subscription, "is_home_region", False)):
                add_region(str(getattr(subscription, "region_name", "") or ""))
        for subscription in subscriptions:
            add_region(str(getattr(subscription, "region_name", "") or ""))
    except Exception as exc:
        logger.info("Unable to discover OCI subscribed regions; using configured region only: %s", exc)

    add_region(fallback_region)
    return region_names


def _oci_collection_items(response_or_data: Any) -> List[Any]:
    data = getattr(response_or_data, "data", response_or_data)
    if isinstance(data, list):
        return data
    items = getattr(data, "items", None)
    if isinstance(items, list):
        return items
    if isinstance(data, dict):
        dict_items = data.get("items")
        if isinstance(dict_items, list):
            return dict_items
    return []


def _oci_extended_metadata(item: Any) -> Dict[str, Any]:
    metadata = getattr(item, "extended_metadata", None)
    if isinstance(metadata, dict):
        return metadata
    if isinstance(item, dict):
        raw = item.get("extended-metadata") or item.get("extended_metadata")
        if isinstance(raw, dict):
            return raw
    return {}


def _oci_recommendation_slug(value: str) -> str:
    slug = str(value or "").strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug).strip("-")
    for suffix in ("-name", "-desc"):
        if slug.endswith(suffix):
            slug = slug[: -len(suffix)]
    return slug


def _oci_recommendation_display_name(name: str) -> str:
    slug = _oci_recommendation_slug(name)
    labels = {
        "cost-management-block-volume-attachment": "Delete unattached block volumes",
        "cost-management-boot-volume-attachment": "Delete unattached boot volumes",
        "delete-unattached-block-volumes": "Delete unattached block volumes",
        "delete-unattached-boot-volumes": "Delete unattached boot volumes",
        "cost-management-compute-host-burstable": "Change compute instances to burstable",
        "cost-management-compute-host-underutilized": "Downsize underutilized compute instances",
        "cost-management-autonomous-database-underutilized": "Downsize underutilized ADW and ATP databases",
        "cost-management-load-balancer-underutilized": "Downsize underutilized load balancers",
        "rightsize-exacs-x6-x7-x8-db-cluster": "Downsize underutilized Exadata Cloud VM clusters",
        "rightsize-vmdb-system": "Downsize underutilized Base Database system",
        "downsize-exacs-x6-x7-x8-db-cluster": "Downsize underutilized Exadata Cloud VM clusters",
        "downsize-vmdb-system": "Downsize underutilized Base Database system",
        "cost-management-database-management": "Enable database management",
        "enable-db-management": "Enable database management",
        "cost-management-compute-host-monitoring": "Enable monitoring on compute instances",
        "cost-management-compute-enable-monitoring": "Enable monitoring on compute instances",
        "cost-management-compute-host-idle": "Delete idle compute instances",
        "cost-management-compute-host-terminated": "Delete idle compute instances",
        "performance-block-volume-enable-auto-tuning": "Enable performance auto-tuning for detached block volumes",
        "performance-boot-volume-enable-auto-tuning": "Enable performance auto-tuning for detached boot volumes",
        "performance-compute-host-highutilization": "Rightsize compute instances",
        "performance-load-balancer-highutilization": "Rightsize load balancers",
        "high-availability-object-storage-enable-object-versioning": "Enable object versioning",
        "high-availability-object-storage-enable-replication": "Enable object replication",
        "high-availability-compute-fault-domain": "Improve fault tolerance",
        "cost-management-object-storage-enable-olm": "Move object storage to lower cost tiers",
    }
    if slug in labels:
        return labels[slug]
    for prefix in ("cost-management-", "performance-", "high-availability-"):
        if slug.startswith(prefix):
            slug = slug[len(prefix):]
    return slug.replace("-", " ").title() or "OCI Cloud Advisor recommendation"


def _oci_recommendation_category(name: str) -> str:
    slug = _oci_recommendation_slug(name)
    if slug.startswith("performance-") or slug.startswith("rightsize-"):
        return "Performance"
    if slug.startswith("high-availability-"):
        return "High availability"
    return "Cost management"


def _oci_recommendation_service(name: str) -> str:
    slug = _oci_recommendation_slug(name)
    if "block-volume" in slug or "boot-volume" in slug:
        return "Block storage"
    if "object-storage" in slug:
        return "Object storage"
    if "autonomous-database" in slug:
        return "Autonomous database"
    if "load-balancer" in slug:
        return "Load balancing"
    if "exacs" in slug:
        return "Exadata cloud"
    if "vmdb" in slug or "database-management" in slug or "db-management" in slug:
        return "Base database"
    if "compute" in slug:
        return "Compute"
    return "OCI"


def _oci_recommendation_resource_count(rec: Any) -> int:
    counts = getattr(rec, "resource_counts", None) or []
    if isinstance(rec, dict):
        counts = rec.get("resource-counts") or rec.get("resource_counts") or counts
    total = 0
    for item in counts or []:
        status = str(getattr(item, "status", "") or (item.get("status") if isinstance(item, dict) else "") or "").upper()
        raw_count = getattr(item, "count", None)
        if raw_count is None and isinstance(item, dict):
            raw_count = item.get("count")
        try:
            count = int(raw_count or 0)
        except (TypeError, ValueError):
            count = 0
        if status in {"PENDING", "ACTIVE", ""}:
            total += count
    return total


def _oci_importance_label(value: Any) -> str:
    text = str(value or "MODERATE").strip().upper()
    return {
        "LOW": "Low",
        "MODERATE": "Medium",
        "MEDIUM": "Medium",
        "HIGH": "High",
        "CRITICAL": "Critical",
    }.get(text, text.title() or "Medium")


def _runtime_oci_credential_json() -> Optional[Dict[str, Any]]:
    cfg = Config()
    config_file = str(cfg.oci_config_file or "").strip()
    if not config_file:
        return None
    return {
        "config_file": config_file,
        "profile": _normalize_oci_profile(cfg.oci_profile),
    }


def _oci_scan_compartment_seeds(
    cred_json: Optional[Dict[str, Any]] = None,
    cfg: Optional[Config] = None,
) -> List[str]:
    """Return explicit OCI scan seed compartments without limiting tenancy scans."""
    cfg = cfg or Config()
    cred_json = cred_json or {}
    raw_values = [
        str(cred_json.get("compartment_id") or "").strip(),
        str(cred_json.get("compartment_ocid") or "").strip(),
        str(os.getenv("OCI_COMPARTMENT_ID", "") or "").strip(),
        str(os.getenv("OCI_COMPARTMENT_OCID", "") or "").strip(),
    ]
    raw_values.extend(
        item.strip()
        for item in str(cfg.oci_compartment_ids or "").split(",")
        if item.strip()
    )

    seeds: List[str] = []
    seen: set[str] = set()
    for compartment_id in raw_values:
        if not compartment_id or compartment_id in seen:
            continue
        seen.add(compartment_id)
        seeds.append(compartment_id)
    return seeds


def _oci_optimizer_recommendation_rows(
    cred_json: Dict[str, Any],
    min_monthly_savings: float,
    limit: int,
) -> List[Dict[str, Any]]:
    try:
        import oci
    except Exception as exc:
        logger.warning("OCI SDK unavailable for optimizer recommendations: %s", exc)
        return []

    config_file = str(cred_json.get("config_file") or "").strip()
    profile = _normalize_oci_profile(cred_json.get("profile"))
    if not config_file:
        return []
    try:
        resolved_config, resolved_profile = CredentialValidator._normalize_oci_inputs(
            config_file=config_file,
            profile=profile,
        )
        oci_config = oci.config.from_file(resolved_config, resolved_profile)
        tenancy_id = str(oci_config.get("tenancy") or "").strip()
        region = str(oci_config.get("region") or "global").strip() or "global"
    except Exception as exc:
        logger.warning("Unable to load OCI config for optimizer recommendations: %s", exc)
        return []
    if not tenancy_id:
        return []

    home_region = _oci_home_region(oci, oci_config, tenancy_id)
    if home_region and home_region != region:
        oci_config = dict(oci_config)
        oci_config["region"] = home_region
        region = home_region

    try:
        client = oci.optimizer.OptimizerClient(oci_config, timeout=(5, 25))
    except Exception as exc:
        logger.warning("Unable to initialize OCI Optimizer client: %s", exc)
        return []

    rows: List[Dict[str, Any]] = []
    try:
        try:
            rec_response = oci.pagination.list_call_get_all_results(
                client.list_recommendations,
                compartment_id=tenancy_id,
                compartment_id_in_subtree=True,
                include_organization=True,
            )
        except TypeError:
            rec_response = oci.pagination.list_call_get_all_results(
                client.list_recommendations,
                compartment_id=tenancy_id,
                compartment_id_in_subtree=True,
            )
        for idx, rec in enumerate(_oci_collection_items(rec_response), start=1):
            if len(rows) >= limit:
                break
            monthly = _safe_recommendation_float(getattr(rec, "estimated_cost_saving", 0))
            if monthly < min_monthly_savings:
                continue
            name = str(getattr(rec, "name", "") or "OCI Optimizer recommendation")
            description = str(getattr(rec, "description", "") or name)
            display_name = _oci_recommendation_display_name(name)
            service_name = _oci_recommendation_service(name)
            category_name = _oci_recommendation_category(name)
            rec_type = _provider_recommendation_type_from_text(
                rec_type=str(getattr(rec, "category_id", "") or ""),
                service=service_name,
                description=f"{name} {description} {display_name}",
            )
            lifecycle_state = str(getattr(rec, "lifecycle_state", "") or "").upper()
            resource_count = _oci_recommendation_resource_count(rec)
            row = _provider_row(
                provider="oci",
                source="oci_optimizer",
                rec_id=str(getattr(rec, "id", "") or f"oci-optimizer-{idx:03d}"),
                service=service_name,
                rec_type=rec_type,
                description=f"OCI: {display_name}",
                monthly_savings=monthly,
                account_id=str(getattr(rec, "compartment_id", "") or tenancy_id),
                region=region,
                resource_type=service_name,
                payback_months=1 if rec_type in {"idle-resources", "storage-optimization"} else 3,
                confidence="high",
                severity=str(getattr(rec, "importance", "") or "medium").lower() or "medium",
                provider_metadata={
                    "recommendation_type": display_name,
                    "recommendation_name": name,
                    "resource_count": resource_count if resource_count > 0 else None,
                    "category": category_name,
                    "importance": _oci_importance_label(getattr(rec, "importance", None)),
                    "status": "Active" if lifecycle_state == "ACTIVE" else str(getattr(rec, "status", "") or "Active").title(),
                    "recommendation_status": str(getattr(rec, "status", "") or "").title(),
                },
            )
            rows.append(row)
    except Exception as exc:
        logger.info("OCI Optimizer recommendations unavailable: %s", exc)

    try:
        try:
            actions_response = oci.pagination.list_call_get_all_results(
                client.list_resource_actions,
                compartment_id=tenancy_id,
                compartment_id_in_subtree=True,
                include_organization=True,
                include_resource_metadata=True,
            )
        except TypeError:
            actions_response = oci.pagination.list_call_get_all_results(
                client.list_resource_actions,
                compartment_id=tenancy_id,
                compartment_id_in_subtree=True,
                include_resource_metadata=True,
            )
        for idx, action in enumerate(_oci_collection_items(actions_response), start=1):
            if len(rows) >= limit:
                break
            monthly = _safe_recommendation_float(getattr(action, "estimated_cost_saving", 0))
            if monthly < min_monthly_savings:
                continue
            resource_type = str(getattr(action, "resource_type", "") or "OCI Resource")
            action_name = str(getattr(action, "name", "") or "Optimize resource")
            extended_metadata = _oci_extended_metadata(action)
            action_region = str(extended_metadata.get("region") or region).strip() or region
            detached_volume = bool(extended_metadata.get("volumeDetachedStatus")) or bool(extended_metadata.get("unattachedSince"))
            description = f"{action_name}"
            if detached_volume and "volume" in resource_type.lower():
                description = f"Delete unattached {resource_type}: {action_name}"
            rec_type = _provider_recommendation_type_from_text(
                service=resource_type,
                description=description,
                action=str(getattr(action, "action", "") or ""),
            )
            resource_id = str(getattr(action, "resource_id", "") or "")
            rows.append(_provider_row(
                provider="oci",
                source="oci_optimizer_resource_action",
                rec_id=str(getattr(action, "id", "") or f"oci-optimizer-action-{idx:03d}"),
                service=resource_type,
                rec_type=rec_type,
                description=f"OCI: {description}",
                monthly_savings=monthly,
                account_id=str(getattr(action, "compartment_id", "") or tenancy_id),
                region=action_region,
                resource_id=resource_id,
                resource_name=action_name,
                resource_type=resource_type,
                payback_months=1 if rec_type in {"idle-resources", "storage-optimization"} else 3,
                confidence="high",
                severity="high",
                console_url=_rightsizing_console_url(
                    provider="oci",
                    resource_id=resource_id,
                    region=action_region,
                    account_id=tenancy_id,
                    resource_type=resource_type,
                ),
                provider_metadata={
                    "recommendation_type": description,
                    "recommendation_name": action_name,
                    "resource_count": 1,
                    "category": "Cost management",
                    "importance": "High",
                    "status": "Active",
                },
            ))
    except Exception as exc:
        logger.info("OCI Optimizer resource actions unavailable: %s", exc)
    return rows[:limit]


def _collect_provider_recommendation_rows(
    *,
    db: Session,
    customer_id: str,
    provider: str,
    min_monthly_savings: float,
    limit: int = 200,
    include_existing_rightsizing_sources: bool = True,
) -> List[Dict[str, Any]]:
    provider_filter = [provider] if provider != "all" else ["aws", "azure", "gcp", "oci"]
    effective_limit = max(1, int(limit or 1))
    rows: List[Dict[str, Any]] = []
    for prov in provider_filter:
        if prov not in {"aws", "azure", "gcp", "oci"}:
            continue
        cred_rows = (
            db.query(CredentialRecord)
            .filter(
                CredentialRecord.customer_id == customer_id,
                CredentialRecord.provider == prov,
                CredentialRecord.is_valid.is_(True),
            )
            .all()
        )
        credential_payloads: List[Dict[str, Any]] = []
        for cred_row in cred_rows:
            try:
                credential_payloads.append(json.loads(cred_row.credential_json))
            except Exception:
                continue
        if prov == "oci" and not credential_payloads:
            runtime_oci = _runtime_oci_credential_json()
            if runtime_oci is not None:
                credential_payloads.append(runtime_oci)
        for cred_json in credential_payloads:
            if len(rows) >= effective_limit:
                break
            remaining = provider_bounded_limit(prov, max(effective_limit - len(rows), 0), floor=0)
            if remaining <= 0:
                break
            if prov == "aws":
                rows.extend(_aws_cost_explorer_commitment_recommendations(
                    cred_json,
                    min_monthly_savings,
                    remaining,
                ))
                remaining = provider_bounded_limit(prov, max(effective_limit - len(rows), 0), floor=0)
                if include_existing_rightsizing_sources and remaining > 0:
                    rows.extend(_recommendation_rows_from_rightsizing(
                        _rightsizing_from_aws_ce(cred_json, min_monthly_savings)[:remaining]
                    ))
            elif prov == "azure":
                rows.extend(_azure_advisor_recommendation_rows(
                    cred_json,
                    min_monthly_savings,
                    remaining,
                ))
            elif prov == "gcp":
                rows.extend(_gcp_recommender_rows(
                    cred_json,
                    min_monthly_savings,
                    remaining,
                ))
                remaining = provider_bounded_limit(prov, max(effective_limit - len(rows), 0), floor=0)
                if include_existing_rightsizing_sources and remaining > 0:
                    rows.extend(_recommendation_rows_from_rightsizing(
                        _rightsizing_from_gcp_cloud_monitoring(cred_json, min_monthly_savings, limit=remaining)
                    ))
            elif prov == "oci":
                rows.extend(_oci_optimizer_recommendation_rows(
                    cred_json,
                    min_monthly_savings,
                    remaining,
                ))
                remaining = provider_bounded_limit(prov, max(effective_limit - len(rows), 0), floor=0)
                if include_existing_rightsizing_sources and remaining > 0:
                    rows.extend(_recommendation_rows_from_rightsizing(
                        _rightsizing_from_oci_compute_inventory(cred_json, min_monthly_savings, limit=remaining)
                    ))
                remaining = provider_bounded_limit(prov, max(effective_limit - len(rows), 0), floor=0)
                if include_existing_rightsizing_sources and remaining > 0:
                    rows.extend(_recommendation_rows_from_rightsizing(
                        _rightsizing_from_oci_storage_inventory(cred_json, min_monthly_savings, limit=remaining)
                    ))
        if len(rows) >= effective_limit:
            break

    deduped: List[Dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        key = (
            _provider_from_recommendation_row(row, provider),
            str(row.get("source") or "").lower(),
            str(row.get("resource_id") or "").lower(),
            str(row.get("type") or "").lower(),
            _slugify_resource_token(str(row.get("description") or "")),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    deduped.sort(key=_recommendation_monthly_savings, reverse=True)
    return deduped[:effective_limit]


def _provider_recommendation_action(row: Dict[str, Any]) -> str:
    rec_type = str(row.get("type") or "").strip().lower()
    service = str(row.get("service") or "").strip().lower()
    description = str(row.get("description") or "").strip().lower()
    if rec_type in {"reserved-instances", "committed-use", "savings-plan", "commitment"}:
        return "reserve"
    if rec_type in {"idle-resources", "orphaned-resources"}:
        return "terminate"
    if "unattached" in description or "orphan" in description:
        return "terminate"
    if "volume" in service and ("clean" in description or "delete" in description):
        return "terminate"
    if rec_type in {"storage-optimization", "lifecycle", "modernization"}:
        return "modernize"
    if rec_type in {"rightsizing", "right-size"}:
        return "downsize"
    return "modernize"


def _rightsizing_from_provider_recommendation_rows(
    rows: List[Dict[str, Any]],
    *,
    provider: str,
    region: str,
    account_id: str,
    min_savings: float,
) -> List[RightsizingRecommendation]:
    out: List[RightsizingRecommendation] = []
    observed_at = _utcnow().isoformat()
    normalized_provider = str(provider or "all").strip().lower() or "all"
    safe_region = str(region or "global").strip() or "global"
    for row in rows:
        service = str(row.get("service") or "Cloud Service").strip()
        monthly_savings = _recommendation_monthly_savings(row)
        if monthly_savings < min_savings:
            continue
        current_monthly = round(_safe_recommendation_float(row.get("current_annual_spend"), 0.0) / 12, 2)
        if not current_monthly:
            current_monthly = _safe_recommendation_float(row.get("current_monthly_cost_usd"), 0.0)
        if current_monthly <= 0:
            current_monthly = monthly_savings
        projected_monthly = round(max(current_monthly - monthly_savings, 0.0), 2)
        action = _provider_recommendation_action(row)
        rec_provider = _provider_from_recommendation_row(row, normalized_provider)
        if rec_provider == "all":
            rec_provider = normalized_provider

        rec_type = str(row.get("type") or action).strip().lower()
        service_slug = _slugify_resource_token(service)
        rec_id = str(row.get("id") or f"{rec_provider}-{service_slug}-{rec_type}")
        resource_id = str(row.get("resource_id") or f"provider-rec-{rec_id}")
        resource_type = str(row.get("resource_type") or "").strip()
        if not resource_type:
            resource_type = f"{rec_provider.upper()} {service} service opportunity"
        resource_name = str(row.get("resource_name") or "").strip()
        if not resource_name:
            resource_name = f"{rec_provider.upper()} {service}: {str(row.get('description') or 'Optimization opportunity')}"
        row_region = str(row.get("region") or safe_region).strip() or safe_region
        row_account = str(row.get("account_id") or account_id).strip()
        confidence = str(row.get("confidence") or "medium").strip().lower()
        if confidence not in {"high", "medium", "low"}:
            confidence = "medium"
        payback_months = float(row.get("payback_months", 0) or 0)
        effort = "low" if payback_months <= 1 else "medium" if payback_months <= 3 else "high"
        out.append(
            RightsizingRecommendation(
                resource_id=resource_id,
                resource_name=resource_name,
                resource_type=resource_type,
                provider=rec_provider,
                region=row_region,
                account_id=row_account,
                current_size=str(row.get("current_size") or f"{service} current monthly spend"),
                recommended_size=str(row.get("description") or "Apply provider recommendation"),
                current_monthly_cost_usd=current_monthly,
                projected_monthly_cost_usd=projected_monthly,
                monthly_savings_usd=monthly_savings,
                annual_savings_usd=round(monthly_savings * 12, 2),
                cpu_utilization_avg_percent=None,
                memory_utilization_avg_percent=None,
                reason=str(row.get("description") or "Provider optimization recommendation."),
                confidence=confidence,
                effort=effort,
                action=action,
                evidence_source=str(row.get("source") or row.get("evidence_source") or "live_provider_recommendations"),
                analysis_points=1,
                trend_slope_usd=0.0,
                trend_percent=0.0,
                latest_monthly_cost_usd=current_monthly,
                peak_monthly_cost_usd=current_monthly,
                top_regions=[row_region] if row_region else [],
                regional_breakdown=[{
                    "region": row_region,
                    "monthly_cost_usd": current_monthly,
                    "share_percent": 100.0,
                }],
                resource_console_url=row.get("resource_console_url") or _rightsizing_console_url(
                    provider=rec_provider,
                    resource_id=resource_id,
                    region=row_region,
                    account_id=row_account,
                    resource_type=resource_type,
                ),
                last_observed_at=observed_at,
                risk_note=(
                    "Service-level provider recommendation. Use Inventory Explorer and the provider console "
                    "to select exact resources before making changes."
                ),
                provider_recommendation_type=str(row.get("recommendation_type") or "").strip() or None,
                provider_recommendation_name=str(row.get("recommendation_name") or "").strip() or None,
                provider_recommendation_category=str(row.get("category") or "").strip() or None,
                provider_recommendation_status=str(row.get("status") or row.get("recommendation_status") or "").strip() or None,
                provider_recommendation_importance=str(row.get("importance") or row.get("severity") or "").strip() or None,
                provider_recommendation_resource_count=(
                    int(row.get("resource_count"))
                    if str(row.get("resource_count") or "").strip().isdigit()
                    else None
                ),
            )
        )
    return out


def _rightsizing_from_oci_compute_inventory(
    cred_json: Dict[str, Any],
    min_savings: float,
    limit: int = 120,
) -> List[RightsizingRecommendation]:
    """Derive per-instance OCI rightsizing opportunities from live compute inventory."""
    try:
        import oci
    except Exception as exc:
        logger.warning("OCI SDK unavailable for rightsizing inventory: %s", exc)
        return []

    config_file = str(cred_json.get("config_file") or "").strip()
    profile = _normalize_oci_profile(cred_json.get("profile"))
    if not config_file:
        return []
    try:
        resolved_config, resolved_profile = CredentialValidator._normalize_oci_inputs(
            config_file=config_file,
            profile=profile,
        )
        oci_config = oci.config.from_file(resolved_config, resolved_profile)
    except Exception as exc:
        logger.warning("Unable to load OCI config for rightsizing inventory: %s", exc)
        return []

    tenancy_id = str(oci_config.get("tenancy") or "").strip()
    if not tenancy_id:
        return []

    client_timeout = (5, 20)
    home_region = _oci_home_region(oci, oci_config, tenancy_id, timeout=client_timeout) or str(oci_config.get("region") or "").strip()
    home_config = _oci_config_for_region(oci_config, home_region)
    try:
        identity = oci.identity.IdentityClient(home_config, timeout=client_timeout)
    except Exception as exc:
        logger.warning("Unable to initialize OCI identity client for rightsizing inventory: %s", exc)
        return []

    compartment_ids = _oci_accessible_compartment_ids(
        oci,
        identity,
        tenancy_id,
        seed_compartment_ids=_oci_scan_compartment_seeds(cred_json),
    )
    scan_regions = _oci_subscribed_regions(
        oci,
        home_config,
        tenancy_id,
        home_region=home_region,
        timeout=client_timeout,
    )

    out: List[RightsizingRecommendation] = []
    seen_instances: set[str] = set()
    observed_at = _utcnow().isoformat()
    for region in scan_regions:
        if len(out) >= limit:
            break
        try:
            compute = oci.core.ComputeClient(_oci_config_for_region(home_config, region), timeout=client_timeout)
        except Exception as exc:
            logger.warning("Unable to initialize OCI compute client for region %s: %s", region, exc)
            continue
        for compartment_id in compartment_ids:
            if len(out) >= limit:
                break
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute.list_instances,
                    compartment_id=compartment_id,
                    limit=50,
                ).data
            except Exception:
                continue
            for instance in instances or []:
                if len(out) >= limit:
                    break
                instance_id = str(getattr(instance, "id", "") or "").strip()
                if not instance_id or instance_id in seen_instances:
                    continue
                seen_instances.add(instance_id)

                lifecycle = str(getattr(instance, "lifecycle_state", "") or "").upper()
                if lifecycle not in {"RUNNING", "STOPPED"}:
                    continue
                shape = str(getattr(instance, "shape", "") or "").strip()
                shape_config = getattr(instance, "shape_config", None)
                ocpus = float(getattr(shape_config, "ocpus", 0.0) or 0.0) if shape_config else None
                memory_gib = float(getattr(shape_config, "memory_in_gbs", 0.0) or 0.0) if shape_config else None
                current_size, recommended_size, savings_rate = _oci_rightsize_target(shape, ocpus, memory_gib)
                if not current_size or not recommended_size or savings_rate <= 0:
                    continue

                current_monthly = _estimate_oci_instance_monthly_cost(shape, ocpus, memory_gib)
                monthly_savings = round(current_monthly * savings_rate, 2)
                if monthly_savings < min_savings:
                    continue
                projected_monthly = round(max(current_monthly - monthly_savings, 0.0), 2)

                display_name = str(getattr(instance, "display_name", "") or "").strip()
                if not display_name:
                    display_name = instance_id

                out.append(
                    RightsizingRecommendation(
                        resource_id=instance_id,
                        resource_name=display_name,
                        resource_type="OCI Compute Instance",
                        provider="oci",
                        region=region,
                        account_id=compartment_id or tenancy_id,
                        current_size=current_size,
                        recommended_size=recommended_size,
                        current_monthly_cost_usd=current_monthly,
                        projected_monthly_cost_usd=projected_monthly,
                        monthly_savings_usd=monthly_savings,
                        annual_savings_usd=round(monthly_savings * 12, 2),
                        cpu_utilization_avg_percent=None,
                        memory_utilization_avg_percent=None,
                        reason=(
                            "Live OCI instance inventory indicates this shape can be downsized. "
                            "Validate CPU/memory utilization before applying in production."
                        ),
                        confidence="medium",
                        effort="low",
                        action="downsize",
                        evidence_source="oci_compute_inventory",
                        analysis_points=1,
                        trend_slope_usd=0.0,
                        trend_percent=0.0,
                        latest_monthly_cost_usd=current_monthly,
                        peak_monthly_cost_usd=current_monthly,
                        top_regions=[region] if region else [],
                        regional_breakdown=[{
                            "region": region or "global",
                            "monthly_cost_usd": current_monthly,
                            "share_percent": 100.0,
                        }],
                        resource_console_url=_rightsizing_console_url(
                            provider="oci",
                            resource_id=instance_id,
                            region=region,
                            account_id=compartment_id or tenancy_id,
                            resource_type="OCI Compute Instance",
                        ),
                        last_observed_at=observed_at,
                        risk_note="Inventory-based recommendation: confirm workload headroom and autoscaling before resize.",
                    )
                )
    return out


def _oci_accessible_compartment_ids(
    oci_module: Any,
    identity_client: Any,
    tenancy_id: str,
    *,
    seed_compartment_ids: Optional[List[str]] = None,
    max_compartments: Optional[int] = None,
) -> List[str]:
    """Return tenancy plus all discoverable compartments in the tenancy subtree."""
    if max_compartments is None:
        max_compartments = max(1, int(Config().oci_max_scan_compartments or 500))
    compartment_ids: List[str] = []
    seen_compartments: set[str] = set()

    if tenancy_id and tenancy_id not in seen_compartments:
        compartment_ids.append(tenancy_id)
        seen_compartments.add(tenancy_id)

    for seed in seed_compartment_ids or []:
        compartment_id = str(seed or "").strip()
        if not compartment_id or compartment_id in seen_compartments:
            continue
        if len(compartment_ids) >= max_compartments:
            return compartment_ids[:max_compartments]
        compartment_ids.append(compartment_id)
        seen_compartments.add(compartment_id)

    for access_level in ("ANY", "ACCESSIBLE"):
        try:
            comp_response = oci_module.pagination.list_call_get_all_results(
                identity_client.list_compartments,
                compartment_id=tenancy_id,
                compartment_id_in_subtree=True,
                access_level=access_level,
            )
            for compartment in comp_response.data or []:
                if len(compartment_ids) >= max_compartments:
                    break
                lifecycle = str(getattr(compartment, "lifecycle_state", "") or "").upper()
                comp_id = str(getattr(compartment, "id", "") or "").strip()
                if not comp_id or lifecycle not in {"ACTIVE", ""}:
                    continue
                if comp_id not in seen_compartments:
                    seen_compartments.add(comp_id)
                    compartment_ids.append(comp_id)
            break
        except Exception:
            continue
    return compartment_ids[:max_compartments]


def _estimate_oci_storage_monthly_cost(size_gb: float, vpus_per_gb: Optional[float] = None) -> float:
    size = max(0.0, float(size_gb or 0.0))
    # OCI Block Volume is billed from live size/performance attributes; this is
    # a conservative normalized estimate used only for ranking cleanup actions.
    base_storage_rate = 0.0255
    performance_rate = 0.0017 * max(0.0, float(vpus_per_gb or 0.0))
    return round(size * (base_storage_rate + performance_rate), 2)


def _rightsizing_from_oci_storage_inventory(
    cred_json: Dict[str, Any],
    min_savings: float,
    limit: int = 120,
) -> List[RightsizingRecommendation]:
    """Find unattached OCI block and boot volumes from live provider inventory."""
    try:
        import oci
    except Exception as exc:
        logger.warning("OCI SDK unavailable for storage rightsizing inventory: %s", exc)
        return []

    config_file = str(cred_json.get("config_file") or "").strip()
    profile = _normalize_oci_profile(cred_json.get("profile"))
    if not config_file:
        return []
    try:
        resolved_config, resolved_profile = CredentialValidator._normalize_oci_inputs(
            config_file=config_file,
            profile=profile,
        )
        oci_config = oci.config.from_file(resolved_config, resolved_profile)
    except Exception as exc:
        logger.warning("Unable to load OCI config for storage rightsizing inventory: %s", exc)
        return []

    tenancy_id = str(oci_config.get("tenancy") or "").strip()
    if not tenancy_id:
        return []

    client_timeout = (5, 20)
    home_region = _oci_home_region(oci, oci_config, tenancy_id, timeout=client_timeout) or str(oci_config.get("region") or "").strip()
    home_config = _oci_config_for_region(oci_config, home_region)
    try:
        identity = oci.identity.IdentityClient(home_config, timeout=client_timeout)
    except Exception as exc:
        logger.warning("Unable to initialize OCI identity client for storage rightsizing inventory: %s", exc)
        return []

    compartment_ids = _oci_accessible_compartment_ids(
        oci,
        identity,
        tenancy_id,
        seed_compartment_ids=_oci_scan_compartment_seeds(cred_json),
    )
    scan_regions = _oci_subscribed_regions(
        oci,
        home_config,
        tenancy_id,
        home_region=home_region,
        timeout=client_timeout,
    )

    attached_block_volume_ids: set[str] = set()
    attached_boot_volume_ids: set[str] = set()
    total_seen = 0

    out: List[RightsizingRecommendation] = []
    seen_storage_resources: set[str] = set()
    observed_at = _utcnow().isoformat()

    def _append_storage_rec(
        *,
        region: str,
        resource_id: str,
        display_name: str,
        resource_type: str,
        size_gb: float,
        compartment_id: str,
        lifecycle_state: str,
        created_at: Any,
        vpus_per_gb: Optional[float] = None,
    ) -> None:
        if len(out) >= limit:
            return
        if not resource_id or resource_id in seen_storage_resources:
            return
        monthly_cost = _estimate_oci_storage_monthly_cost(size_gb, vpus_per_gb)
        if monthly_cost < min_savings:
            return
        seen_storage_resources.add(resource_id)
        last_seen = observed_at
        if created_at is not None:
            try:
                last_seen = created_at.isoformat()
            except Exception:
                last_seen = str(created_at)
        out.append(
            RightsizingRecommendation(
                resource_id=resource_id,
                resource_name=display_name or resource_id,
                resource_type=resource_type,
                provider="oci",
                region=region,
                account_id=compartment_id or tenancy_id,
                current_size=f"{size_gb:g} GB",
                recommended_size="Delete after snapshot/retention validation",
                current_monthly_cost_usd=monthly_cost,
                projected_monthly_cost_usd=0.0,
                monthly_savings_usd=monthly_cost,
                annual_savings_usd=round(monthly_cost * 12, 2),
                cpu_utilization_avg_percent=None,
                memory_utilization_avg_percent=None,
                reason=(
                    f"Live OCI inventory shows this {resource_type.lower()} is not attached "
                    f"to any instance. Lifecycle state: {lifecycle_state or 'unknown'}."
                ),
                confidence="high",
                effort="low",
                action="terminate",
                evidence_source="oci_storage_inventory",
                analysis_points=1,
                trend_slope_usd=0.0,
                trend_percent=0.0,
                latest_monthly_cost_usd=monthly_cost,
                peak_monthly_cost_usd=monthly_cost,
                top_regions=[region] if region else [],
                regional_breakdown=[{
                    "region": region or "global",
                    "monthly_cost_usd": monthly_cost,
                    "share_percent": 100.0,
                }],
                resource_console_url=_rightsizing_console_url(
                    provider="oci",
                    resource_id=resource_id,
                    region=region,
                    account_id=compartment_id or tenancy_id,
                    resource_type=resource_type,
                ),
                last_observed_at=last_seen,
                risk_note=(
                    "Unattached storage cleanup: create/verify a snapshot or backup, confirm owner "
                    "approval, then delete to stop ongoing storage charges."
                ),
            )
        )

    for region in scan_regions:
        if len(out) >= limit:
            break
        try:
            region_config = _oci_config_for_region(home_config, region)
            region_identity = oci.identity.IdentityClient(region_config, timeout=client_timeout)
            compute = oci.core.ComputeClient(region_config, timeout=client_timeout)
            blockstorage = oci.core.BlockstorageClient(region_config, timeout=client_timeout)
        except Exception as exc:
            logger.warning("Unable to initialize OCI storage clients for region %s: %s", region, exc)
            continue
        try:
            availability_domains = [
                str(getattr(ad, "name", "") or "").strip()
                for ad in region_identity.list_availability_domains(compartment_id=tenancy_id).data or []
                if str(getattr(ad, "name", "") or "").strip()
            ]
        except Exception:
            availability_domains = []

        for compartment_id in compartment_ids:
            ad_values = availability_domains or [None]
            for ad in ad_values:
                try:
                    kwargs = {"compartment_id": compartment_id}
                    if ad:
                        kwargs["availability_domain"] = ad
                    attachments = oci.pagination.list_call_get_all_results(
                        compute.list_volume_attachments,
                        **kwargs,
                    ).data
                    for attachment in attachments or []:
                        lifecycle = str(getattr(attachment, "lifecycle_state", "") or "").upper()
                        if lifecycle in {"DETACHED", "DETACHING", "TERMINATED"}:
                            continue
                        volume_id = str(getattr(attachment, "volume_id", "") or "").strip()
                        if volume_id:
                            attached_block_volume_ids.add(volume_id)
                except Exception:
                    pass
                try:
                    kwargs = {"compartment_id": compartment_id}
                    if ad:
                        kwargs["availability_domain"] = ad
                    attachments = oci.pagination.list_call_get_all_results(
                        compute.list_boot_volume_attachments,
                        **kwargs,
                    ).data
                    for attachment in attachments or []:
                        lifecycle = str(getattr(attachment, "lifecycle_state", "") or "").upper()
                        if lifecycle in {"DETACHED", "DETACHING", "TERMINATED"}:
                            continue
                        volume_id = str(getattr(attachment, "boot_volume_id", "") or "").strip()
                        if volume_id:
                            attached_boot_volume_ids.add(volume_id)
                except Exception:
                    pass

        for compartment_id in compartment_ids:
            if len(out) >= limit:
                break
            try:
                volumes = oci.pagination.list_call_get_all_results(
                    blockstorage.list_volumes,
                    compartment_id=compartment_id,
                    lifecycle_state="AVAILABLE",
                ).data
            except Exception:
                volumes = []
            for volume in volumes or []:
                total_seen += 1
                volume_id = str(getattr(volume, "id", "") or "").strip()
                if not volume_id or volume_id in attached_block_volume_ids:
                    continue
                size_gb = float(getattr(volume, "size_in_gbs", 0.0) or 0.0)
                vpus = getattr(volume, "vpus_per_gb", None)
                _append_storage_rec(
                    region=region,
                    resource_id=volume_id,
                    display_name=str(getattr(volume, "display_name", "") or "").strip(),
                    resource_type="OCI Block Volume (unattached)",
                    size_gb=size_gb,
                    vpus_per_gb=float(vpus) if vpus is not None else None,
                    compartment_id=compartment_id,
                    lifecycle_state=str(getattr(volume, "lifecycle_state", "") or ""),
                    created_at=getattr(volume, "time_created", None),
                )
                if len(out) >= limit:
                    break

            if len(out) >= limit:
                break
            try:
                boot_volumes = oci.pagination.list_call_get_all_results(
                    blockstorage.list_boot_volumes,
                    compartment_id=compartment_id,
                    lifecycle_state="AVAILABLE",
                ).data
            except Exception:
                boot_volumes = []
            for volume in boot_volumes or []:
                total_seen += 1
                volume_id = str(getattr(volume, "id", "") or "").strip()
                if not volume_id or volume_id in attached_boot_volume_ids:
                    continue
                size_gb = float(getattr(volume, "size_in_gbs", 0.0) or 0.0)
                _append_storage_rec(
                    region=region,
                    resource_id=volume_id,
                    display_name=str(getattr(volume, "display_name", "") or "").strip(),
                    resource_type="OCI Boot Volume (unattached)",
                    size_gb=size_gb,
                    compartment_id=compartment_id,
                    lifecycle_state=str(getattr(volume, "lifecycle_state", "") or ""),
                    created_at=getattr(volume, "time_created", None),
                )
                if len(out) >= limit:
                    break

    for rec in out:
        rec.analysis_points = max(rec.analysis_points, total_seen)
    return out


# ---------------------------------------------------------------------------
# Tier 1: AWS Cost Explorer get_rightsizing_recommendation
# ---------------------------------------------------------------------------

def _rightsizing_from_aws_ce(
    cred_json: Dict[str, Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Call AWS Cost Explorer RightsizingRecommendation API and normalise results."""
    try:
        import boto3
        from botocore.exceptions import ClientError, BotoCoreError

        client = boto3.client(
            "ce",
            aws_access_key_id=cred_json.get("access_key_id"),
            aws_secret_access_key=cred_json.get("secret_access_key"),
            region_name=cred_json.get("region", "us-east-1"),
        )
        response = client.get_rightsizing_recommendation(
            Service="AmazonEC2",
            Configuration={
                "RecommendationTarget": "SAME_INSTANCE_FAMILY",
                "BenefitsConsidered": True,
            },
        )
        out: List[RightsizingRecommendation] = []
        for rec in response.get("RightsizingRecommendations", []):
            rtype = rec.get("RightsizingType", "Terminate")
            details = rec.get("ModifyRecommendationDetail") or rec.get("TerminateRecommendationDetail") or {}
            curr_instance = rec.get("CurrentInstance", {})
            curr_type = curr_instance.get("InstanceType", "unknown")
            account = curr_instance.get("AccountId", "")
            resource_id = curr_instance.get("ResourceId", f"ec2-{account}")
            region = curr_instance.get("Tags", {}).get("aws:cloudformation:stack-id", "") or "us-east-1"

            curr_cost = float(
                (curr_instance.get("MonthlyCost") or "0").replace("$", "").replace(",", "")
            )
            if rtype == "Modify":
                target = (details.get("TargetInstances") or [{}])[0]
                recommended_type = target.get("ExpectedResourceType", "smaller")
                proj_cost = float(
                    (target.get("ExpectedMonthlyBenefit") or {}).get("Amount", "0")
                )
                monthly_savings = max(0.0, curr_cost - proj_cost)
                action = "downsize"
                effort = "low"
                reason = (
                    f"CE recommends {recommended_type} — "
                    f"estimated ${monthly_savings:.2f}/mo savings"
                )
                cpu_util = float(
                    (curr_instance.get("ResourceUtilization") or {}).get("EC2ResourceUtilization", {}).get("MaxCpuUtilizationPercentage", 0) or 0
                )
            else:  # Terminate
                recommended_type = "N/A — terminate"
                monthly_savings = curr_cost
                proj_cost = 0.0
                action = "terminate"
                effort = "low"
                reason = "Instance has zero utilisation over the CE lookback window — candidate for termination"
                cpu_util = 0.0

            if monthly_savings < min_savings:
                continue

            confidence = "high" if cpu_util < 10 else "medium"
            out.append(RightsizingRecommendation(
                resource_id=resource_id,
                resource_name=curr_type,
                resource_type="EC2 Instance",
                provider="aws",
                region=region,
                account_id=account,
                current_size=curr_type,
                recommended_size=recommended_type,
                current_monthly_cost_usd=round(curr_cost, 2),
                projected_monthly_cost_usd=round(proj_cost, 2),
                monthly_savings_usd=round(monthly_savings, 2),
                annual_savings_usd=round(monthly_savings * 12, 2),
                cpu_utilization_avg_percent=round(cpu_util, 1) if cpu_util else None,
                memory_utilization_avg_percent=None,
                reason=reason,
                confidence=confidence,
                effort=effort,
                action=action,
                evidence_source="aws_cost_explorer",
            ))
        return out
    except Exception as exc:
        logger.warning("AWS CE rightsizing API error (will fall through): %s", exc)
        return []


# ---------------------------------------------------------------------------
# Tier 2: Azure Advisor rightsizing recommendations
# ---------------------------------------------------------------------------

def _rightsizing_from_azure_advisor(
    cred_json: Dict[str, Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Call Azure Advisor recommendations and return rightsizing findings."""
    try:
        from azure.identity import ClientSecretCredential
        from azure.mgmt.advisor import AdvisorManagementClient

        azure_cred = ClientSecretCredential(
            tenant_id=cred_json["tenant_id"],
            client_id=cred_json["client_id"],
            client_secret=cred_json["client_secret"],
        )
        subscription_id = cred_json["subscription_id"]
        client = AdvisorManagementClient(azure_cred, subscription_id)

        out: List[RightsizingRecommendation] = []
        for rec in client.recommendations.list():
            category = (rec.category or "").lower()
            if "cost" not in category:
                continue
            impact = (rec.impact or "").lower()
            resource_metadata = rec.resource_metadata or {}
            resource_id = getattr(rec, "resource_id", None) or ""
            short_id = resource_id.split("/")[-1] if resource_id else "unknown"
            region = getattr(rec, "location", "unknown") or "unknown"

            # Extract extended properties for savings estimate
            props = rec.extended_properties or {}
            savings_str = props.get("annualSavingsAmount") or props.get("savingsAmount") or "0"
            try:
                annual_savings = float(str(savings_str).replace("$", "").replace(",", ""))
            except (TypeError, ValueError):
                annual_savings = 0.0
            monthly_savings = round(annual_savings / 12, 2)

            if monthly_savings < min_savings:
                continue

            current_sku = props.get("currentSku") or props.get("currentSize") or "unknown"
            recommended_sku = props.get("targetSku") or props.get("recommendedSize") or "smaller SKU"
            curr_cost = round(monthly_savings / max(_ACTION_SAVINGS_RATES["downsize"], 0.01), 2)

            confidence = "high" if impact == "high" else "medium" if impact == "medium" else "low"
            out.append(RightsizingRecommendation(
                resource_id=resource_id or short_id,
                resource_name=props.get("resourceName") or short_id,
                resource_type="Virtual Machine",
                provider="azure",
                region=region,
                account_id=subscription_id,
                current_size=current_sku,
                recommended_size=recommended_sku,
                current_monthly_cost_usd=curr_cost,
                projected_monthly_cost_usd=round(curr_cost - monthly_savings, 2),
                monthly_savings_usd=monthly_savings,
                annual_savings_usd=round(annual_savings, 2),
                cpu_utilization_avg_percent=None,
                memory_utilization_avg_percent=None,
                reason=rec.short_description.problem if rec.short_description else "Azure Advisor cost recommendation",
                confidence=confidence,
                effort="low",
                action="downsize",
                evidence_source="azure_advisor",
            ))
        return out
    except Exception as exc:
        logger.warning("Azure Advisor rightsizing API error (will fall through): %s", exc)
        return []


# ---------------------------------------------------------------------------
# Live utilization enrichers (CloudWatch / Azure Monitor / Cloud Monitoring)
# ---------------------------------------------------------------------------

def _aws_cloudwatch_metric_average(
    cloudwatch_client: Any,
    instance_id: str,
    metric_name: str,
    namespace: str = "AWS/EC2",
    statistic: str = "Average",
    period_seconds: int = 3600,
    lookback_days: int = 14,
) -> Optional[float]:
    try:
        end_time = datetime.now(timezone.utc)
        start_time = end_time - timedelta(days=max(1, int(lookback_days)))
        response = cloudwatch_client.get_metric_statistics(
            Namespace=namespace,
            MetricName=metric_name,
            Dimensions=[{"Name": "InstanceId", "Value": instance_id}],
            StartTime=start_time,
            EndTime=end_time,
            Period=period_seconds,
            Statistics=[statistic],
        )
        values = [
            float(point.get(statistic))
            for point in (response.get("Datapoints") or [])
            if point.get(statistic) is not None
        ]
        return _series_mean(values)
    except Exception:
        return None


def _enrich_aws_with_cloudwatch_utilization(
    cred_json: Dict[str, Any],
    recommendations_in: List[RightsizingRecommendation],
) -> List[RightsizingRecommendation]:
    if not recommendations_in:
        return recommendations_in
    try:
        import boto3
    except Exception:
        return recommendations_in

    region_default = str(cred_json.get("region") or "us-east-1")
    clients: Dict[str, Any] = {}
    for rec in recommendations_in:
        if rec.provider != "aws":
            continue
        instance_id = str(rec.resource_id or "").strip()
        if not instance_id.startswith("i-"):
            continue
        region = str(rec.region or "").strip().lower()
        if not re.match(r"^[a-z]{2}-[a-z-]+-\d+$", region):
            region = region_default
        if region not in clients:
            try:
                clients[region] = boto3.client(
                    "cloudwatch",
                    aws_access_key_id=cred_json.get("access_key_id"),
                    aws_secret_access_key=cred_json.get("secret_access_key"),
                    region_name=region,
                )
            except Exception:
                continue
        cw_client = clients[region]
        cpu = _aws_cloudwatch_metric_average(cw_client, instance_id, "CPUUtilization", "AWS/EC2")
        memory = _aws_cloudwatch_metric_average(cw_client, instance_id, "mem_used_percent", "CWAgent")
        if cpu is not None:
            rec.cpu_utilization_avg_percent = round(cpu, 2)
        if memory is not None:
            rec.memory_utilization_avg_percent = round(memory, 2)
        if cpu is not None or memory is not None:
            rec.evidence_source = "aws_cloudwatch"
            rec.last_observed_at = _utcnow().isoformat()
    return recommendations_in


def _rightsizing_from_aws_cloudwatch_inventory(
    cred_json: Dict[str, Any],
    min_savings: float,
    limit: int = 120,
) -> List[RightsizingRecommendation]:
    try:
        import boto3
    except Exception:
        return []

    region = str(cred_json.get("region") or "us-east-1")
    try:
        ec2 = boto3.client(
            "ec2",
            aws_access_key_id=cred_json.get("access_key_id"),
            aws_secret_access_key=cred_json.get("secret_access_key"),
            region_name=region,
        )
        cloudwatch = boto3.client(
            "cloudwatch",
            aws_access_key_id=cred_json.get("access_key_id"),
            aws_secret_access_key=cred_json.get("secret_access_key"),
            region_name=region,
        )
    except Exception:
        return []

    out: List[RightsizingRecommendation] = []
    try:
        paginator = ec2.get_paginator("describe_instances")
        for page in paginator.paginate(
            Filters=[{"Name": "instance-state-name", "Values": ["running", "stopped"]}]
        ):
            for reservation in page.get("Reservations", []):
                for instance in reservation.get("Instances", []):
                    if len(out) >= limit:
                        return out
                    instance_id = str(instance.get("InstanceId") or "").strip()
                    if not instance_id:
                        continue
                    instance_type = str(instance.get("InstanceType") or "unknown")
                    state = str((instance.get("State") or {}).get("Name") or "running").lower()
                    name = instance_id
                    for tag in instance.get("Tags", []):
                        if str(tag.get("Key") or "").lower() == "name":
                            name = str(tag.get("Value") or "").strip() or instance_id
                            break

                    cpu = _aws_cloudwatch_metric_average(cloudwatch, instance_id, "CPUUtilization", "AWS/EC2")
                    memory = _aws_cloudwatch_metric_average(cloudwatch, instance_id, "mem_used_percent", "CWAgent")
                    if cpu is None and memory is None:
                        continue

                    if state == "stopped":
                        action = "terminate"
                        savings_rate = 1.0
                        reason = "Instance is stopped and shows no active utilization load."
                    elif (cpu or 0.0) <= 10.0 and (memory is None or memory <= 55.0):
                        action = "downsize"
                        savings_rate = 0.45 if (cpu or 0.0) <= 7.0 else 0.30
                        reason = (
                            f"CloudWatch average CPU {(cpu or 0.0):.1f}% and memory "
                            f"{memory:.1f}% over lookback window indicate over-provisioning."
                            if memory is not None
                            else f"CloudWatch average CPU {(cpu or 0.0):.1f}% indicates over-provisioning."
                        )
                    else:
                        continue

                    current_monthly = _estimate_monthly_cost_from_size("aws", instance_type)
                    monthly_savings = round(current_monthly * savings_rate, 2)
                    if monthly_savings < min_savings:
                        continue

                    out.append(
                        RightsizingRecommendation(
                            resource_id=instance_id,
                            resource_name=name,
                            resource_type="EC2 Instance",
                            provider="aws",
                            region=region,
                            account_id=str(instance.get("OwnerId") or cred_json.get("account_id") or "aws-account"),
                            current_size=instance_type,
                            recommended_size=_recommended_size_for_action("aws", action, instance_type),
                            current_monthly_cost_usd=round(current_monthly, 2),
                            projected_monthly_cost_usd=round(max(current_monthly - monthly_savings, 0.0), 2),
                            monthly_savings_usd=monthly_savings,
                            annual_savings_usd=round(monthly_savings * 12, 2),
                            cpu_utilization_avg_percent=round(cpu, 2) if cpu is not None else None,
                            memory_utilization_avg_percent=round(memory, 2) if memory is not None else None,
                            reason=reason,
                            confidence="high" if (cpu is not None and cpu <= 8.0) else "medium",
                            effort="low",
                            action=action,
                            evidence_source="aws_cloudwatch",
                            analysis_points=14,
                            latest_monthly_cost_usd=round(current_monthly, 2),
                            peak_monthly_cost_usd=round(current_monthly, 2),
                            top_regions=[region],
                            regional_breakdown=[{"region": region, "monthly_cost_usd": round(current_monthly, 2), "share_percent": 100.0}],
                            resource_console_url=_rightsizing_console_url("aws", instance_id, region, "", "EC2 Instance"),
                            last_observed_at=_utcnow().isoformat(),
                        )
                    )
    except Exception as exc:
        logger.warning("AWS CloudWatch rightsizing inventory error (will fall through): %s", exc)
    return out


def _azure_monitor_metric_average(
    access_token: str,
    resource_id: str,
    metric_name: str,
    lookback_days: int = 14,
) -> Optional[float]:
    end_dt = datetime.now(timezone.utc).replace(microsecond=0)
    start_dt = (end_dt - timedelta(days=max(1, int(lookback_days)))).replace(microsecond=0)
    timespan = f"{start_dt.isoformat().replace('+00:00', 'Z')}/{end_dt.isoformat().replace('+00:00', 'Z')}"
    url = f"https://management.azure.com{resource_id}/providers/microsoft.insights/metrics"
    params = {
        "api-version": "2023-10-01",
        "metricnames": metric_name,
        "timespan": timespan,
        "interval": "PT1H",
        "aggregation": "Average",
    }
    headers = {"Authorization": f"Bearer {access_token}"}
    try:
        response = httpx.get(url, params=params, headers=headers, timeout=20)
        if response.status_code >= 400:
            return None
        payload = response.json()
        values: List[float] = []
        for metric in payload.get("value", []):
            for series in metric.get("timeseries", []):
                for point in series.get("data", []):
                    raw = point.get("average")
                    if raw is not None:
                        values.append(float(raw))
        return _series_mean(values)
    except Exception:
        return None


def _enrich_azure_with_monitor_utilization(
    cred_json: Dict[str, Any],
    recommendations_in: List[RightsizingRecommendation],
) -> List[RightsizingRecommendation]:
    if not recommendations_in:
        return recommendations_in
    try:
        from azure.identity import ClientSecretCredential

        credential = ClientSecretCredential(
            tenant_id=cred_json["tenant_id"],
            client_id=cred_json["client_id"],
            client_secret=cred_json["client_secret"],
        )
        token = credential.get_token("https://management.azure.com/.default").token
    except Exception:
        return recommendations_in

    for rec in recommendations_in:
        if rec.provider != "azure":
            continue
        resource_id = str(rec.resource_id or "").strip()
        if not resource_id.startswith("/subscriptions/"):
            continue
        cpu = _azure_monitor_metric_average(token, resource_id, "Percentage CPU")
        available_memory = _azure_monitor_metric_average(token, resource_id, "Available Memory Percentage")
        memory_used = None if available_memory is None else round(max(0.0, min(100.0, 100.0 - available_memory)), 2)

        if cpu is not None:
            rec.cpu_utilization_avg_percent = round(cpu, 2)
        if memory_used is not None:
            rec.memory_utilization_avg_percent = memory_used
        if cpu is not None or memory_used is not None:
            rec.evidence_source = "azure_monitor"
            rec.last_observed_at = _utcnow().isoformat()
    return recommendations_in


def _gcp_monitoring_metric_average(
    session: Any,
    project_id: str,
    instance_id: str,
    metric_type: str,
    lookback_days: int = 14,
) -> Optional[float]:
    end_dt = datetime.now(timezone.utc).replace(microsecond=0)
    start_dt = (end_dt - timedelta(days=max(1, int(lookback_days)))).replace(microsecond=0)
    filter_expr = (
        f'metric.type = "{metric_type}" '
        f'AND resource.type = "gce_instance" '
        f'AND resource.labels.instance_id = "{instance_id}"'
    )
    params = {
        "filter": filter_expr,
        "interval.startTime": start_dt.isoformat().replace("+00:00", "Z"),
        "interval.endTime": end_dt.isoformat().replace("+00:00", "Z"),
        "aggregation.alignmentPeriod": "3600s",
        "aggregation.perSeriesAligner": "ALIGN_MEAN",
        "view": "FULL",
        "pageSize": 100,
    }
    url = f"https://monitoring.googleapis.com/v3/projects/{project_id}/timeSeries"
    try:
        response = session.get(url, params=params, timeout=20)
        if response.status_code >= 400:
            return None
        payload = response.json()
        values: List[float] = []
        for series in payload.get("timeSeries", []):
            for point in series.get("points", []):
                raw = (point.get("value") or {}).get("doubleValue")
                if raw is None:
                    raw = (point.get("value") or {}).get("int64Value")
                if raw is not None:
                    values.append(float(raw))
        return _series_mean(values)
    except Exception:
        return None


def _rightsizing_from_gcp_cloud_monitoring(
    cred_json: Dict[str, Any],
    min_savings: float,
    limit: int = 120,
) -> List[RightsizingRecommendation]:
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import AuthorizedSession
    except Exception:
        return []

    service_account_json = cred_json.get("service_account_json")
    if isinstance(service_account_json, str):
        try:
            service_account_json = json.loads(service_account_json)
        except Exception:
            service_account_json = {}
    if not isinstance(service_account_json, dict):
        return []

    project_id = str(cred_json.get("project_id") or service_account_json.get("project_id") or "").strip()
    if not project_id:
        return []

    try:
        credentials = service_account.Credentials.from_service_account_info(
            service_account_json,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        session = AuthorizedSession(credentials)
    except Exception:
        return []

    out: List[RightsizingRecommendation] = []
    try:
        next_token = ""
        while len(out) < limit:
            params: Dict[str, Any] = {"maxResults": 200}
            if next_token:
                params["pageToken"] = next_token
            list_url = f"https://compute.googleapis.com/compute/v1/projects/{project_id}/aggregated/instances"
            response = session.get(list_url, params=params, timeout=30)
            if response.status_code >= 400:
                break
            payload = response.json()
            for scope_data in (payload.get("items") or {}).values():
                for instance in scope_data.get("instances", []):
                    if len(out) >= limit:
                        break
                    status = str(instance.get("status") or "RUNNING").upper()
                    if status not in {"RUNNING", "TERMINATED", "STOPPED", "SUSPENDED"}:
                        continue
                    instance_id = str(instance.get("id") or "").strip()
                    if not instance_id:
                        continue
                    zone_url = str(instance.get("zone") or "")
                    zone = zone_url.split("/")[-1] if zone_url else "global"
                    machine_type_url = str(instance.get("machineType") or "")
                    machine_type = machine_type_url.split("/")[-1] if machine_type_url else "unknown"
                    name = str(instance.get("name") or instance_id)
                    cpu_ratio = _gcp_monitoring_metric_average(
                        session,
                        project_id,
                        instance_id,
                        "compute.googleapis.com/instance/cpu/utilization",
                    )
                    cpu_percent = None if cpu_ratio is None else round(cpu_ratio * 100.0, 2)
                    memory_percent = _gcp_monitoring_metric_average(
                        session,
                        project_id,
                        instance_id,
                        "agent.googleapis.com/memory/percent_used",
                    )
                    if cpu_percent is None and memory_percent is None:
                        continue

                    if status in {"TERMINATED", "STOPPED", "SUSPENDED"}:
                        action = "terminate"
                        savings_rate = 1.0
                        reason = "Instance is not running and has no active workload requirement."
                    elif (cpu_percent or 0.0) <= 12.0 and (memory_percent is None or memory_percent <= 60.0):
                        action = "downsize"
                        savings_rate = 0.45 if (cpu_percent or 0.0) <= 8.0 else 0.30
                        reason = (
                            f"Cloud Monitoring average CPU {(cpu_percent or 0.0):.1f}% and memory "
                            f"{memory_percent:.1f}% indicate over-provisioning."
                            if memory_percent is not None
                            else f"Cloud Monitoring average CPU {(cpu_percent or 0.0):.1f}% indicates over-provisioning."
                        )
                    else:
                        continue

                    current_monthly = _estimate_monthly_cost_from_size("gcp", machine_type)
                    monthly_savings = round(current_monthly * savings_rate, 2)
                    if monthly_savings < min_savings:
                        continue

                    out.append(
                        RightsizingRecommendation(
                            resource_id=f"projects/{project_id}/zones/{zone}/instances/{name}",
                            resource_name=name,
                            resource_type="GCE Instance",
                            provider="gcp",
                            region=zone,
                            account_id=project_id,
                            current_size=machine_type,
                            recommended_size=_recommended_size_for_action("gcp", action, machine_type),
                            current_monthly_cost_usd=round(current_monthly, 2),
                            projected_monthly_cost_usd=round(max(current_monthly - monthly_savings, 0.0), 2),
                            monthly_savings_usd=monthly_savings,
                            annual_savings_usd=round(monthly_savings * 12, 2),
                            cpu_utilization_avg_percent=cpu_percent,
                            memory_utilization_avg_percent=round(memory_percent, 2) if memory_percent is not None else None,
                            reason=reason,
                            confidence="high" if (cpu_percent is not None and cpu_percent <= 8.0) else "medium",
                            effort="low",
                            action=action,
                            evidence_source="gcp_cloud_monitoring",
                            analysis_points=14,
                            latest_monthly_cost_usd=round(current_monthly, 2),
                            peak_monthly_cost_usd=round(current_monthly, 2),
                            top_regions=[zone],
                            regional_breakdown=[{"region": zone, "monthly_cost_usd": round(current_monthly, 2), "share_percent": 100.0}],
                            resource_console_url=_rightsizing_console_url(
                                provider="gcp",
                                resource_id=f"projects/{project_id}/zones/{zone}/instances/{name}",
                                region=zone,
                                account_id=project_id,
                                resource_type="GCE Instance",
                            ),
                            last_observed_at=_utcnow().isoformat(),
                        )
                    )
            next_token = str(payload.get("nextPageToken") or "").strip()
            if not next_token:
                break
    except Exception as exc:
        logger.warning("GCP Cloud Monitoring rightsizing error (will fall through): %s", exc)
    return out


# ---------------------------------------------------------------------------
# Tier 3: Cost-trend signal analysis from scan snapshot history
#
# Signals used (no external API needed):
#   • Cost trend across scans:   slope > 0 → growing (reserve candidate)
#                                slope ≈ 0, cost high → over-provisioned (downsize)
#                                cost near zero → orphaned (terminate)
#   • Anomalies count:           high → investigate before resizing
#   • Savings identified:        provider already flagged savings → high confidence
#   • Number of scan data points: more points → higher confidence
# ---------------------------------------------------------------------------

def _trend_slope(values: List[float]) -> float:
    """Simple linear regression slope over equally-spaced observations."""
    n = len(values)
    if n < 2:
        return 0.0
    x_mean = (n - 1) / 2.0
    y_mean = sum(values) / n
    num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
    den = sum((i - x_mean) ** 2 for i in range(n))
    return num / den if den else 0.0


def _rightsizing_from_snapshot_trends(
    snapshots_by_account: Dict[int, List[Any]],
    account_map: Dict[int, Any],
    top_regions_by_account: Dict[int, List[tuple[str, float]]],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Derive rightsizing signals from multi-scan cost history without external APIs."""
    out: List[RightsizingRecommendation] = []
    for acct_id, snaps in snapshots_by_account.items():
        acct = account_map.get(acct_id)
        if not acct:
            continue
        prov = acct.provider

        # Sort oldest-first for trend analysis
        snaps_sorted = sorted(snaps, key=lambda s: s.captured_at)
        costs = [float(s.direct_cost_usd or 0) for s in snaps_sorted]
        latest_cost = costs[-1] if costs else 0.0
        avg_cost = sum(costs) / len(costs) if costs else 0.0
        total_anomalies = sum(s.anomalies_count or 0 for s in snaps_sorted)
        total_savings_identified = sum(float(s.savings_identified_usd or 0) for s in snaps_sorted)
        n_points = len(costs)
        slope = _trend_slope(costs)

        # Skip very cheap accounts (under $20/mo) — not worth rightsizing
        if avg_cost < 20:
            continue

        # Derive action from signals
        if latest_cost < 5.0:
            # Near-zero cost → orphaned resource
            action = "terminate"
            savings_rate = 1.0
            reason = (
                f"Cost dropped to ${latest_cost:.2f}/mo (avg ${avg_cost:.2f}/mo over "
                f"{n_points} scans) — likely orphaned or stopped resource"
            )
            effort = "low"
            # High confidence only with multiple data points confirming low cost
            confidence = "high" if n_points >= 3 and max(costs[:-1]) < 10 else "medium"

        elif slope > avg_cost * 0.05 and total_anomalies == 0:
            # Steady growth, no anomalies → reserve for predictable savings
            action = "reserve"
            savings_rate = _ACTION_SAVINGS_RATES["reserve"]
            growth_pct = round(slope / avg_cost * 100, 1)
            reason = (
                f"Cost growing +{growth_pct}% per scan period with no anomalies — "
                f"steady-state workload ideal for 1yr Reserved Instance (est. "
                f"{int(savings_rate * 100)}% discount)"
            )
            effort = "medium"
            confidence = "high" if n_points >= 4 else "medium"

        elif slope < -avg_cost * 0.03 and latest_cost > 100:
            # Declining cost but still high → already being optimised; modernize to accelerate
            action = "modernize"
            savings_rate = _ACTION_SAVINGS_RATES["modernize"]
            reason = (
                f"Cost declining (${costs[0]:.2f} → ${latest_cost:.2f}/mo) but still high — "
                f"migrate to newer generation for further {int(savings_rate * 100)}% reduction"
            )
            effort = "medium"
            confidence = "medium"

        elif total_savings_identified > 0:
            # Provider scan already flagged savings — downsize
            action = "downsize"
            savings_rate = min(
                total_savings_identified / max(avg_cost, 1.0) / n_points,
                0.50,
            )
            reason = (
                f"Provider scan identified ${total_savings_identified:.2f} savings over "
                f"{n_points} scans — consistent over-provisioning signal"
            )
            effort = "low"
            confidence = "high"

        elif avg_cost > 200 and abs(slope) < avg_cost * 0.02 and total_anomalies == 0:
            # Flat high cost, no anomalies → over-provisioned, downsize
            action = "downsize"
            savings_rate = _ACTION_SAVINGS_RATES["downsize"]
            reason = (
                f"Flat cost of ~${avg_cost:.2f}/mo over {n_points} scans with no anomalies — "
                f"consistent usage pattern indicates over-provisioning"
            )
            effort = "low"
            confidence = "high" if n_points >= 3 else "medium"

        else:
            # Insufficient signal
            continue

        monthly_savings = round(avg_cost * savings_rate, 2)
        if monthly_savings < min_savings:
            continue

        trend_pct = round((slope / avg_cost) * 100, 2) if avg_cost > 0 else 0.0
        current_size = _extract_size_from_account_metadata(acct, prov)
        recommended_size = _recommended_size_for_action(prov, action, current_size)
        top_regions = top_regions_by_account.get(acct_id, [])
        top_region_names = [r for r, _ in top_regions[:3]]
        regional_breakdown = []
        for region_name, region_cost in top_regions[:4]:
            share = round((region_cost / avg_cost) * 100, 2) if avg_cost > 0 else 0.0
            regional_breakdown.append({
                "region": region_name,
                "monthly_cost_usd": round(region_cost, 2),
                "share_percent": share,
            })

        type_labels = {"aws": "EC2 Instance", "azure": "Virtual Machine",
                       "gcp": "Compute Instance", "oci": "OCI Compute"}
        risk_note = None
        if total_anomalies > 0:
            risk_note = f"{total_anomalies} anomalies detected across history — validate performance before applying changes."
        elif action in {"downsize", "modernize"} and n_points < 3:
            risk_note = "Limited history points; apply with canary rollout."

        out.append(RightsizingRecommendation(
            resource_id=f"{prov}-acct-{acct_id}",
            resource_name=acct.account_name or acct.account_identifier,
            resource_type=f"{type_labels.get(prov, 'Compute Instance')} (account aggregate)",
            provider=prov,
            region=acct.native_region or "global",
            account_id=acct.account_identifier,
            current_size=current_size,
            recommended_size=recommended_size,
            current_monthly_cost_usd=round(avg_cost, 2),
            projected_monthly_cost_usd=round(avg_cost * (1 - savings_rate), 2),
            monthly_savings_usd=monthly_savings,
            annual_savings_usd=round(monthly_savings * 12, 2),
            cpu_utilization_avg_percent=None,  # real metrics not yet ingested
            memory_utilization_avg_percent=None,
            reason=reason,
            confidence=confidence,
            effort=effort,
            action=action,
            evidence_source="cost_trend_analysis",
            analysis_points=n_points,
            trend_slope_usd=round(slope, 4),
            trend_percent=trend_pct,
            latest_monthly_cost_usd=round(latest_cost, 2),
            peak_monthly_cost_usd=round(max(costs), 2) if costs else None,
            top_regions=top_region_names,
            regional_breakdown=regional_breakdown,
            last_observed_at=snaps_sorted[-1].captured_at.isoformat() if snaps_sorted[-1].captured_at else None,
            risk_note=(
                risk_note
                or "Account-level recommendation. Use Resource Inventory and provider consoles to pick exact instances/volumes for execution."
            ),
        ))

        # Add region-level recommendations for top-cost regions to increase actionability.
        # This keeps the same evidence source, but scopes savings to concrete hotspots.
        if top_regions and action in {"downsize", "modernize", "reserve"}:
            for rank, (region_name, region_cost) in enumerate(top_regions[:3], start=1):
                region_monthly_savings = round(region_cost * savings_rate, 2)
                if region_monthly_savings < min_savings:
                    continue
                region_current_size = current_size
                region_target_size = _recommended_size_for_action(prov, action, region_current_size)
                out.append(RightsizingRecommendation(
                    resource_id=f"{prov}-acct-{acct_id}-region-{region_name}",
                    resource_name=f"{(acct.account_name or acct.account_identifier)} · {region_name}",
                    resource_type=f"{type_labels.get(prov, 'Compute Instance')} (regional segment)",
                    provider=prov,
                    region=region_name,
                    account_id=acct.account_identifier,
                    current_size=region_current_size,
                    recommended_size=region_target_size,
                    current_monthly_cost_usd=round(region_cost, 2),
                    projected_monthly_cost_usd=round(region_cost * (1 - savings_rate), 2),
                    monthly_savings_usd=region_monthly_savings,
                    annual_savings_usd=round(region_monthly_savings * 12, 2),
                    cpu_utilization_avg_percent=None,
                    memory_utilization_avg_percent=None,
                    reason=(
                        f"Top regional spend segment #{rank} in {region_name} "
                        f"(${region_cost:.2f}/mo) follows the account-level {action} signal."
                    ),
                    confidence="high" if confidence == "high" else "medium",
                    effort=effort,
                    action=action,
                    evidence_source="cost_trend_analysis_region",
                    analysis_points=n_points,
                    trend_slope_usd=round(slope, 4),
                    trend_percent=trend_pct,
                    latest_monthly_cost_usd=round(region_cost, 2),
                    peak_monthly_cost_usd=round(region_cost, 2),
                    top_regions=[region_name],
                    regional_breakdown=[{
                        "region": region_name,
                        "monthly_cost_usd": round(region_cost, 2),
                        "share_percent": 100.0,
                    }],
                    last_observed_at=snaps_sorted[-1].captured_at.isoformat() if snaps_sorted[-1].captured_at else None,
                    risk_note="Regional optimization segment derived from latest allocation snapshot.",
                ))
    return out


def _service_recommendation_profile(provider: str, service_name: str) -> Optional[Dict[str, Any]]:
    text = f"{provider} {service_name}".lower()
    profiles = [
        {
            "terms": ["block storage", "block volume", "boot volume", "file storage", "ebs", "managed disk", "persistent disk", "snapshot", "backup"],
            "action": "modernize",
            "savings_rate": 0.28,
            "resource_type": "Storage volume and backup service",
            "current_size": "Current volume/backup spend",
            "recommended_size": "Lifecycle policy plus unattached volume/snapshot cleanup review",
            "reason": "Storage volume, snapshot, or backup spend is visible in provider service costs. Review unattached volumes, stale snapshots, backup retention, and performance tiers.",
            "effort": "low",
            "confidence": "medium",
        },
        {
            "terms": ["object storage", "cloud storage", "amazon s3", "s3", "blob", "bucket", "glacier", "archive storage"],
            "action": "modernize",
            "savings_rate": 0.18,
            "resource_type": "Object storage service",
            "current_size": "Current object storage spend",
            "recommended_size": "Lifecycle to infrequent/archive tiers",
            "reason": "Object storage spend is present in the latest service-cost snapshot. Apply lifecycle rules, delete expired objects, and move cold data to archive tiers.",
            "effort": "low",
            "confidence": "medium",
        },
        {
            "terms": ["database", "rds", "aurora", "sql", "postgres", "mysql", "autonomous", "dynamodb", "cosmos", "spanner", "redis", "elasticache"],
            "action": "reserve",
            "savings_rate": 0.24,
            "resource_type": "Database and cache service",
            "current_size": "Current managed database/cache spend",
            "recommended_size": "Reserved capacity or right-sized managed tier",
            "reason": "Managed database/cache spend is visible in service-cost snapshots. Check idle instances, storage growth, backup retention, and reserved capacity coverage.",
            "effort": "medium",
            "confidence": "medium",
        },
        {
            "terms": ["load balancer", "network", "nat", "gateway", "egress", "bandwidth", "cdn", "dns", "firewall", "waf", "ip address"],
            "action": "modernize",
            "savings_rate": 0.12,
            "resource_type": "Network and traffic service",
            "current_size": "Current network spend",
            "recommended_size": "Egress, gateway, and idle endpoint optimization",
            "reason": "Network spend is visible in service-cost snapshots. Review NAT gateways, idle load balancers, public IPs, egress paths, and CDN/cache placement.",
            "effort": "medium",
            "confidence": "medium",
        },
        {
            "terms": ["kubernetes", "container", "nodepool", "node pool", "gke", "aks", "eks", "oke", "namespace", "pod"],
            "action": "downsize",
            "savings_rate": 0.22,
            "resource_type": "Kubernetes and container platform",
            "current_size": "Current cluster/container spend",
            "recommended_size": "Right-sized node pools and workload requests",
            "reason": "Kubernetes/container spend is visible in service-cost snapshots. Review node pool sizing, idle namespaces, requests/limits, and autoscaling policy.",
            "effort": "medium",
            "confidence": "medium",
        },
        {
            "terms": ["bigquery", "redshift", "athena", "emr", "databricks", "analytics", "warehouse", "dataflow", "synapse"],
            "action": "modernize",
            "savings_rate": 0.16,
            "resource_type": "Analytics and data platform service",
            "current_size": "Current analytics spend",
            "recommended_size": "Query, retention, and commitment optimization",
            "reason": "Analytics/data-platform spend is visible in service-cost snapshots. Review query efficiency, storage retention, job schedules, and commitment coverage.",
            "effort": "medium",
            "confidence": "medium",
        },
        {
            "terms": ["queue", "pubsub", "pub/sub", "kafka", "event hub", "service bus", "sqs", "sns", "stream"],
            "action": "modernize",
            "savings_rate": 0.10,
            "resource_type": "Messaging and streaming service",
            "current_size": "Current messaging/streaming spend",
            "recommended_size": "Retention, throughput, and idle topic cleanup",
            "reason": "Messaging/streaming spend is visible in service-cost snapshots. Review retention windows, idle topics/queues, throughput tiers, and replay storage.",
            "effort": "low",
            "confidence": "medium",
        },
        {
            "terms": ["compute", "ec2", "virtual machine", "vm", "instance", "compute engine", "oci compute"],
            "action": "downsize",
            "savings_rate": 0.20,
            "resource_type": "Compute service",
            "current_size": "Current compute spend",
            "recommended_size": "Right-sized shapes or commitment coverage",
            "reason": "Compute service spend is visible in service-cost snapshots. Review idle capacity, shape family, and commitment coverage.",
            "effort": "low",
            "confidence": "medium",
        },
    ]
    for profile in profiles:
        if any(term in text for term in profile["terms"]):
            return profile
    return {
        "action": "modernize",
        "savings_rate": 0.08,
        "resource_type": "Cloud service",
        "current_size": "Current service spend",
        "recommended_size": "Service-specific optimization review",
        "reason": "This service appears in the latest service-cost snapshot. Review configuration, idle usage, retention, and pricing model for savings.",
        "effort": "medium",
        "confidence": "low",
    }


def _rightsizing_from_service_cost_snapshots(
    snapshots: List[Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Derive product/service-level opportunities from persisted top_services snapshots."""
    grouped: Dict[tuple[str, str], Dict[str, Any]] = {}
    for snap in snapshots:
        provider = str(getattr(snap, "provider", "") or "").strip().lower()
        if provider not in {"aws", "azure", "gcp", "oci"}:
            continue
        try:
            services = json.loads(getattr(snap, "top_services_json", "") or "[]")
        except Exception:
            services = []
        if not isinstance(services, list):
            continue
        for service in services:
            if not isinstance(service, dict):
                continue
            service_name = str(
                service.get("service")
                or service.get("service_name")
                or service.get("name")
                or ""
            ).strip()
            if not service_name:
                continue
            cost = _safe_recommendation_float(
                service.get("cost_usd")
                or service.get("cost")
                or service.get("amount")
            )
            if cost <= 0:
                continue
            key = (provider, service_name.lower())
            bucket = grouped.setdefault(
                key,
                {
                    "provider": provider,
                    "service_name": service_name,
                    "costs": [],
                    "latest_cost": 0.0,
                    "latest_at": None,
                },
            )
            bucket["costs"].append(cost)
            captured_at = getattr(snap, "captured_at", None)
            latest_at = bucket.get("latest_at")
            if latest_at is None or (captured_at is not None and captured_at > latest_at):
                bucket["latest_at"] = captured_at
                bucket["latest_cost"] = cost

    out: List[RightsizingRecommendation] = []
    for (_provider, _service_key), bucket in grouped.items():
        provider = str(bucket["provider"])
        service_name = str(bucket["service_name"])
        costs = [float(value or 0.0) for value in bucket["costs"]]
        if not costs:
            continue
        avg_cost = sum(costs) / len(costs)
        latest_cost = float(bucket.get("latest_cost") or costs[-1])
        if avg_cost < 15:
            continue
        profile = _service_recommendation_profile(provider, service_name)
        if not profile:
            continue
        savings_rate = float(profile["savings_rate"])
        monthly_savings = round(latest_cost * savings_rate, 2)
        if monthly_savings < min_savings:
            continue
        latest_at = bucket.get("latest_at")
        last_observed = latest_at.isoformat() if latest_at is not None else None
        trend_slope = _trend_slope(costs)
        trend_pct = round((trend_slope / avg_cost) * 100, 2) if avg_cost > 0 else 0.0
        service_slug = _slugify_resource_token(service_name)
        out.append(RightsizingRecommendation(
            resource_id=f"{provider}-service-{service_slug}",
            resource_name=service_name,
            resource_type=str(profile["resource_type"]),
            provider=provider,
            region="global",
            account_id=f"{provider}-service-costs",
            current_size=str(profile["current_size"]),
            recommended_size=str(profile["recommended_size"]),
            current_monthly_cost_usd=round(latest_cost, 2),
            projected_monthly_cost_usd=round(max(latest_cost - monthly_savings, 0.0), 2),
            monthly_savings_usd=monthly_savings,
            annual_savings_usd=round(monthly_savings * 12, 2),
            cpu_utilization_avg_percent=None,
            memory_utilization_avg_percent=None,
            reason=str(profile["reason"]),
            confidence=str(profile["confidence"]),
            effort=str(profile["effort"]),
            action=str(profile["action"]),
            evidence_source="service_cost_snapshot",
            analysis_points=len(costs),
            trend_slope_usd=round(trend_slope, 4),
            trend_percent=trend_pct,
            latest_monthly_cost_usd=round(latest_cost, 2),
            peak_monthly_cost_usd=round(max(costs), 2),
            top_regions=["global"],
            regional_breakdown=[{
                "region": "global",
                "monthly_cost_usd": round(latest_cost, 2),
                "share_percent": 100.0,
            }],
            last_observed_at=last_observed,
            risk_note=(
                "Service-level recommendation from provider top-services cost snapshots. "
                "Open service inventory/billing details to select exact resources before execution."
            ),
        ))
    return out


# ---------------------------------------------------------------------------
# Tier 4: Imported CSV cost-signal analysis (deterministic, no random)
# ---------------------------------------------------------------------------

_HIGH_COST_SERVICES: Dict[str, str] = {
    # Service names that typically indicate large compute → downsize candidate
    "amazon elastic compute cloud": "downsize",
    "amazon ec2": "downsize",
    "ec2": "downsize",
    "compute engine": "downsize",
    "virtual machines": "downsize",
    "oci compute": "downsize",
    # Storage orphan signals → terminate
    "amazon s3": "terminate",
    "azure blob storage": "terminate",
    "cloud storage": "terminate",
    # Consistent services → reserve
    "amazon rds": "reserve",
    "azure sql": "reserve",
    "cloud sql": "reserve",
    "amazon elasticache": "reserve",
}


def _rightsizing_from_imported_costs(
    records: List[Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Derive rightsizing signals from imported CSV cost records (no external API)."""
    # Group records by (provider, account_identifier) and sum costs
    groups: Dict[tuple, Dict[str, Any]] = {}
    for rec in records:
        prov = (rec.provider or "aws").lower()
        acct = rec.account_identifier or f"{prov}-{rec.id}"
        key = (prov, acct)
        if key not in groups:
            groups[key] = {
                "provider": prov,
                "account_id": acct,
                "account_name": rec.account_name or rec.service_name or acct,
                "region": rec.region or "global",
                "total_cost": 0.0,
                "services": [],
                "resource_id": rec.account_identifier or "",
            }
        groups[key]["total_cost"] += float(rec.cost_usd or 0)
        if rec.service_name:
            groups[key]["services"].append(rec.service_name.lower())

    out: List[RightsizingRecommendation] = []
    for (prov, acct_id), grp in groups.items():
        cost = grp["total_cost"]
        if cost < 15:
            continue

        # Determine action from service names; default to downsize for high compute
        action = "downsize"
        for svc in grp["services"]:
            for svc_key, svc_action in _HIGH_COST_SERVICES.items():
                if svc_key in svc:
                    action = svc_action
                    break

        savings_rate = _ACTION_SAVINGS_RATES[action]
        monthly_savings = round(cost * savings_rate, 2)
        if monthly_savings < min_savings:
            continue

        size_map = _DOWNSIZE_MAP.get(prov, _DOWNSIZE_MAP["aws"])
        current_size = next(iter(size_map))
        if action == "terminate":
            recommended_size = "N/A — terminate"
        elif action == "reserve":
            recommended_size = f"{current_size} (Reserved 1yr)"
        else:
            recommended_size = size_map.get(current_size, f"smaller-{current_size}")

        type_labels = {"aws": "EC2 Instance", "azure": "Virtual Machine",
                       "gcp": "Compute Instance", "oci": "OCI Compute"}
        reason_map = {
            "downsize":  f"High compute cost (${cost:.2f}/mo) with no utilisation baseline — apply one-size downgrade",
            "terminate": f"Storage/service cost (${cost:.2f}/mo) with no active resource signal — validate and terminate",
            "reserve":   f"Consistent managed-service cost (${cost:.2f}/mo) — convert to 1yr Reserved for ~{int(savings_rate*100)}% savings",
            "modernize": f"Legacy service cost (${cost:.2f}/mo) — migrate to current-gen equivalent",
        }

        out.append(RightsizingRecommendation(
            resource_id=grp["resource_id"] or f"imported-{prov}-{acct_id}",
            resource_name=grp["account_name"],
            resource_type=f"{type_labels.get(prov, 'Cloud Service')} (imported aggregate)",
            provider=prov,
            region=grp["region"],
            account_id=acct_id,
            current_size=current_size,
            recommended_size=recommended_size,
            current_monthly_cost_usd=round(cost, 2),
            projected_monthly_cost_usd=round(cost * (1 - savings_rate), 2),
            monthly_savings_usd=monthly_savings,
            annual_savings_usd=round(monthly_savings * 12, 2),
            cpu_utilization_avg_percent=None,
            memory_utilization_avg_percent=None,
            reason=reason_map[action],
            confidence="medium",
            effort="low" if action in ("downsize", "terminate") else "medium",
            action=action,
            evidence_source="imported_costs",
            analysis_points=1,
            trend_slope_usd=0.0,
            trend_percent=0.0,
            latest_monthly_cost_usd=round(cost, 2),
            peak_monthly_cost_usd=round(cost, 2),
            top_regions=[grp["region"]] if grp["region"] else [],
            regional_breakdown=[{
                "region": grp["region"] or "global",
                "monthly_cost_usd": round(cost, 2),
                "share_percent": 100.0,
            }],
            risk_note="Imported billing aggregate recommendation. Validate exact resource targets in provider console before changes.",
        ))
    return out


def _recommendation_fingerprint(rec: RightsizingRecommendation) -> str:
    payload = {
        "provider": str(rec.provider or "").lower(),
        "resource_id": rec.resource_id or "",
        "resource_name": rec.resource_name or "",
        "resource_type": rec.resource_type or "",
        "account_id": rec.account_id or "",
        "region": rec.region or "",
        "source": rec.evidence_source or "",
        "action": rec.action or "",
        "current_size": rec.current_size or "",
        "recommended_size": rec.recommended_size or "",
        "planned_monthly_savings_usd": round(float(rec.monthly_savings_usd or 0.0), 2),
        "reason": rec.reason or "",
    }
    body = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(body.encode("utf-8")).hexdigest()


def _ledger_variance(planned_monthly: float, realized_monthly: float) -> tuple[float, float, float]:
    planned = round(float(planned_monthly or 0.0), 2)
    realized = round(float(realized_monthly or 0.0), 2)
    variance_monthly = round(realized - planned, 2)
    variance_annual = round(variance_monthly * 12.0, 2)
    variance_percent = round((variance_monthly / planned) * 100.0, 2) if planned > 0 else 0.0
    return variance_monthly, variance_annual, variance_percent


def _ledger_item_from_row(row: RecommendationLedger) -> RecommendationLedgerItem:
    return RecommendationLedgerItem(
        id=int(row.id),
        organization_id=int(row.organization_id),
        provider=row.provider or "",
        resource_id=row.resource_id or "",
        resource_name=row.resource_name,
        resource_type=row.resource_type,
        account_id=row.account_id,
        region=row.region,
        recommendation_source=row.recommendation_source or "",
        recommendation_fingerprint=row.recommendation_fingerprint or "",
        action=row.action or "",
        confidence=row.confidence or "medium",
        effort=row.effort or "medium",
        status=row.status or "open",
        owner=row.owner,
        current_size=row.current_size,
        recommended_size=row.recommended_size,
        current_monthly_cost_usd=round(float(row.current_monthly_cost_usd or 0.0), 2),
        projected_monthly_cost_usd=round(float(row.projected_monthly_cost_usd or 0.0), 2),
        planned_monthly_savings_usd=round(float(row.planned_monthly_savings_usd or 0.0), 2),
        planned_annual_savings_usd=round(float(row.planned_annual_savings_usd or 0.0), 2),
        realized_monthly_savings_usd=round(float(row.realized_monthly_savings_usd or 0.0), 2),
        realized_annual_savings_usd=round(float(row.realized_annual_savings_usd or 0.0), 2),
        variance_monthly_usd=round(float(row.variance_monthly_usd or 0.0), 2),
        variance_annual_usd=round(float(row.variance_annual_usd or 0.0), 2),
        variance_percent=round(float(row.variance_percent or 0.0), 2),
        variance_reason=row.variance_reason,
        reason=row.reason,
        resource_console_url=row.resource_console_url,
        first_seen_at=row.first_seen_at.isoformat() if row.first_seen_at else None,
        last_seen_at=row.last_seen_at.isoformat() if row.last_seen_at else None,
        planned_at=row.planned_at.isoformat() if row.planned_at else None,
        realized_at=row.realized_at.isoformat() if row.realized_at else None,
        last_exported_at=row.last_exported_at.isoformat() if row.last_exported_at else None,
        times_seen=int(row.times_seen or 0),
    )


def _upsert_recommendation_ledger(
    *,
    db: Session,
    organization_id: int,
    customer_id: str,
    recommendations_in: List[RightsizingRecommendation],
    response_data_source: str,
) -> None:
    if not recommendations_in:
        return

    now = _utcnow()
    for rec in recommendations_in:
        provider_key = str(rec.provider or "unknown").strip().lower() or "unknown"
        source = str(rec.evidence_source or response_data_source or "unknown").strip() or "unknown"
        resource_id = str(rec.resource_id or rec.resource_name or rec.resource_type or "unknown").strip() or "unknown"
        fingerprint = _recommendation_fingerprint(rec)
        planned_monthly = round(float(rec.monthly_savings_usd or 0.0), 2)
        planned_annual = round(float(rec.annual_savings_usd or planned_monthly * 12.0), 2)
        evidence = {
            "response_data_source": response_data_source,
            "cpu_utilization_avg_percent": rec.cpu_utilization_avg_percent,
            "memory_utilization_avg_percent": rec.memory_utilization_avg_percent,
            "analysis_points": rec.analysis_points,
            "trend_slope_usd": rec.trend_slope_usd,
            "trend_percent": rec.trend_percent,
            "latest_monthly_cost_usd": rec.latest_monthly_cost_usd,
            "peak_monthly_cost_usd": rec.peak_monthly_cost_usd,
            "top_regions": rec.top_regions,
            "regional_breakdown": rec.regional_breakdown,
            "last_observed_at": rec.last_observed_at,
            "risk_note": rec.risk_note,
        }

        existing = (
            db.query(RecommendationLedger)
            .filter(
                RecommendationLedger.organization_id == organization_id,
                RecommendationLedger.provider == provider_key,
                RecommendationLedger.resource_id == resource_id,
                RecommendationLedger.recommendation_source == source,
                RecommendationLedger.recommendation_fingerprint == fingerprint,
            )
            .first()
        )

        if existing is None:
            variance_monthly, variance_annual, variance_percent = _ledger_variance(planned_monthly, 0.0)
            db.add(
                RecommendationLedger(
                    organization_id=organization_id,
                    customer_id=customer_id,
                    provider=provider_key,
                    resource_id=resource_id,
                    resource_name=rec.resource_name,
                    resource_type=rec.resource_type,
                    account_id=rec.account_id,
                    region=rec.region,
                    recommendation_source=source,
                    recommendation_fingerprint=fingerprint,
                    action=rec.action,
                    confidence=rec.confidence,
                    effort=rec.effort,
                    current_size=rec.current_size,
                    recommended_size=rec.recommended_size,
                    current_monthly_cost_usd=round(float(rec.current_monthly_cost_usd or 0.0), 2),
                    projected_monthly_cost_usd=round(float(rec.projected_monthly_cost_usd or 0.0), 2),
                    planned_monthly_savings_usd=planned_monthly,
                    planned_annual_savings_usd=planned_annual,
                    realized_monthly_savings_usd=0.0,
                    realized_annual_savings_usd=0.0,
                    variance_monthly_usd=variance_monthly,
                    variance_annual_usd=variance_annual,
                    variance_percent=variance_percent,
                    status="open",
                    reason=rec.reason,
                    evidence_json=json.dumps(evidence, sort_keys=True),
                    resource_console_url=rec.resource_console_url,
                    first_seen_at=now,
                    last_seen_at=now,
                    planned_at=now,
                    times_seen=1,
                    updated_at=now,
                )
            )
            continue

        existing.resource_name = rec.resource_name
        existing.resource_type = rec.resource_type
        existing.account_id = rec.account_id
        existing.region = rec.region
        existing.action = rec.action
        existing.confidence = rec.confidence
        existing.effort = rec.effort
        existing.current_size = rec.current_size
        existing.recommended_size = rec.recommended_size
        existing.current_monthly_cost_usd = round(float(rec.current_monthly_cost_usd or 0.0), 2)
        existing.projected_monthly_cost_usd = round(float(rec.projected_monthly_cost_usd or 0.0), 2)
        existing.planned_monthly_savings_usd = planned_monthly
        existing.planned_annual_savings_usd = planned_annual
        realized_monthly = round(float(existing.realized_monthly_savings_usd or 0.0), 2)
        existing.realized_annual_savings_usd = round(
            float(existing.realized_annual_savings_usd or realized_monthly * 12.0),
            2,
        )
        variance_monthly, variance_annual, variance_percent = _ledger_variance(planned_monthly, realized_monthly)
        existing.variance_monthly_usd = variance_monthly
        existing.variance_annual_usd = variance_annual
        existing.variance_percent = variance_percent
        existing.reason = rec.reason
        existing.evidence_json = json.dumps(evidence, sort_keys=True)
        existing.resource_console_url = rec.resource_console_url
        existing.last_seen_at = now
        existing.times_seen = int(existing.times_seen or 0) + 1
        existing.updated_at = now

    db.commit()


def _query_recommendation_ledger(
    *,
    db: Session,
    organization_id: int,
    provider: str = "all",
    status_filter: str = "all",
    limit: int = 200,
) -> List[RecommendationLedger]:
    query = db.query(RecommendationLedger).filter(
        RecommendationLedger.organization_id == organization_id
    )
    if provider != "all":
        query = query.filter(RecommendationLedger.provider == provider.lower())
    if status_filter != "all":
        query = query.filter(RecommendationLedger.status == status_filter)
    return (
        query.order_by(
            RecommendationLedger.last_seen_at.desc(),
            RecommendationLedger.planned_monthly_savings_usd.desc(),
        )
        .limit(limit)
        .all()
    )


def _recommendation_ledger_response(
    *,
    organization_id: int,
    rows: List[RecommendationLedger],
) -> RecommendationLedgerResponse:
    planned_monthly = sum(float(row.planned_monthly_savings_usd or 0.0) for row in rows)
    realized_monthly = sum(float(row.realized_monthly_savings_usd or 0.0) for row in rows)
    variance_monthly = realized_monthly - planned_monthly
    return RecommendationLedgerResponse(
        generated_at=_utcnow().isoformat(),
        organization_id=organization_id,
        total_count=len(rows),
        total_planned_monthly_savings_usd=round(planned_monthly, 2),
        total_realized_monthly_savings_usd=round(realized_monthly, 2),
        total_variance_monthly_usd=round(variance_monthly, 2),
        total_planned_annual_savings_usd=round(planned_monthly * 12.0, 2),
        total_realized_annual_savings_usd=round(realized_monthly * 12.0, 2),
        total_variance_annual_usd=round(variance_monthly * 12.0, 2),
        items=[_ledger_item_from_row(row) for row in rows],
    )


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.get("/recommendations/rightsizing", response_model=RightsizingResponse)
async def get_rightsizing_recommendations(
    provider: str = Query("all"),
    min_savings: float = Query(10.0, description="Minimum monthly savings threshold USD"),
    limit: int = Query(50, ge=1, le=1000),
    refresh_live: bool = Query(
        False,
        description="Run provider-native live collection on this request. Default uses stored scans/imports for a responsive dashboard path.",
    ),
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Resource-level rightsizing recommendations.

    Uses stored scan/import signals by default for dashboard responsiveness.
    When refresh_live=true, also tries live provider signals first (AWS Cost
    Explorer + CloudWatch, Azure Advisor + Azure Monitor, GCP Cloud Monitoring
    inventory, and OCI inventory) before deterministic scan history/imported
    CSV analysis.
    No synthetic recommendation records are ever generated.
    """
    org_id = membership.organization_id
    customer_id = _customer_id_for_org(membership)
    now_str = _utcnow().isoformat()
    require_live_provider_data = Config().require_live_provider_data

    recommendations_out: List[RightsizingRecommendation] = []
    data_source = "no_data_available"
    total_analyzed = 0
    provider_recommendation_rows: List[Dict[str, Any]] = []
    provider_recommendation_context: Dict[str, Any] = {}

    if refresh_live:
        try:
            provider_recommendation_context = await _cost_context(membership, db, "month", provider)
            provider_rec_result = _safe_json_load(
                await recommendations.get_recommendations(
                    {
                        "cloud_provider": provider,
                        "current_monthly_spend": provider_recommendation_context.get("total_cost", 0.0),
                        "cost_breakdown": provider_recommendation_context.get("breakdown", {}),
                        "min_savings_usd": max(0.0, float(min_savings or 0.0) * 12),
                    }
                ),
                {},
            )
            raw_provider_rows = provider_rec_result.get("recommendations", [])
            if isinstance(raw_provider_rows, list):
                provider_recommendation_rows = [
                    row for row in raw_provider_rows if isinstance(row, dict)
                ]
            provider_recommendation_rows.extend(
                _collect_provider_recommendation_rows(
                    db=db,
                    customer_id=customer_id,
                    provider=provider,
                    min_monthly_savings=float(min_savings or 0.0),
                    limit=limit,
                    include_existing_rightsizing_sources=False,
                )
            )
        except Exception as exc:
            logger.warning("Provider recommendation bridge for rightsizing failed: %s", exc)

    # ── Tier 1 + 2: Real provider APIs and utilization telemetry ────────────
    if refresh_live:
        provider_filter = [provider] if provider != "all" else ["aws", "azure", "gcp"]
        for prov in provider_filter:
            cred_row = (
                db.query(CredentialRecord)
                .filter(
                    CredentialRecord.customer_id == customer_id,
                    CredentialRecord.provider == prov,
                    CredentialRecord.is_valid.is_(True),
                )
                .first()
            )
            if cred_row is None:
                continue
            try:
                cred_json: Dict[str, Any] = json.loads(cred_row.credential_json)
            except Exception:
                continue

            if prov == "aws":
                tier1 = _rightsizing_from_aws_ce(cred_json, min_savings)
                if tier1:
                    tier1 = _enrich_aws_with_cloudwatch_utilization(cred_json, tier1)
                    recommendations_out.extend(tier1)
                    has_live_signals = any(
                        (r.cpu_utilization_avg_percent is not None or r.memory_utilization_avg_percent is not None)
                        for r in tier1
                    )
                    data_source = "aws_ce_cloudwatch" if has_live_signals else "aws_cost_explorer"
                    total_analyzed += len(tier1) * 4  # CE analyzes all instances
                else:
                    tier1b = _rightsizing_from_aws_cloudwatch_inventory(cred_json, min_savings, limit=max(120, limit))
                    if tier1b:
                        recommendations_out.extend(tier1b)
                        data_source = "aws_cloudwatch_inventory" if data_source == "no_data_available" else "multi_provider_api"
                        total_analyzed += len(tier1b)

            elif prov == "azure":
                tier2 = _rightsizing_from_azure_advisor(cred_json, min_savings)
                if tier2:
                    tier2 = _enrich_azure_with_monitor_utilization(cred_json, tier2)
                    recommendations_out.extend(tier2)
                    has_live_signals = any(
                        (r.cpu_utilization_avg_percent is not None or r.memory_utilization_avg_percent is not None)
                        for r in tier2
                    )
                    source_name = "azure_advisor_monitor" if has_live_signals else "azure_advisor"
                    data_source = source_name if data_source == "no_data_available" else "multi_provider_api"
                    total_analyzed += len(tier2) * 3

            elif prov == "gcp":
                tier2c = _rightsizing_from_gcp_cloud_monitoring(cred_json, min_savings, limit=max(120, limit))
                if tier2c:
                    recommendations_out.extend(tier2c)
                    data_source = "gcp_cloud_monitoring" if data_source == "no_data_available" else "multi_provider_api"
                    total_analyzed += len(tier2c)

    # ── Tier 2b: OCI live compute inventory (per-instance actionable scope) ──
    if refresh_live and provider in {"all", "oci"}:
        oci_row = (
            db.query(CredentialRecord)
            .filter(
                CredentialRecord.customer_id == customer_id,
                CredentialRecord.provider == "oci",
                CredentialRecord.is_valid.is_(True),
            )
            .first()
        )
        if oci_row is not None:
            try:
                oci_cred_json: Dict[str, Any] = json.loads(oci_row.credential_json)
            except Exception:
                oci_cred_json = {}
        else:
            oci_cred_json = _runtime_oci_credential_json() or {}
        if isinstance(oci_cred_json, dict):
            tier2b = _rightsizing_from_oci_compute_inventory(oci_cred_json, min_savings, limit=max(120, limit))
            if tier2b:
                recommendations_out.extend(tier2b)
                if data_source == "no_data_available":
                    data_source = "oci_compute_inventory"
                elif data_source not in {"multi_provider_api", "oci_compute_inventory"}:
                    data_source = "multi_provider_api"
                total_analyzed += len(tier2b)
            tier2c = _rightsizing_from_oci_storage_inventory(oci_cred_json, min_savings, limit=max(120, limit))
            if tier2c:
                recommendations_out.extend(tier2c)
                if data_source == "no_data_available":
                    data_source = "oci_storage_inventory"
                elif data_source not in {"multi_provider_api", "oci_storage_inventory"}:
                    data_source = "multi_provider_api"
                total_analyzed += max(len(tier2c), max((r.analysis_points for r in tier2c), default=0))

    # ── Tier 2d: Provider recommendation bridge for service-level actions ───
    if provider_recommendation_rows:
        provider_rec_items = _rightsizing_from_provider_recommendation_rows(
            provider_recommendation_rows,
            provider=provider,
            region=str(Config().oci_region or provider_recommendation_context.get("region") or "global"),
            account_id=customer_id,
            min_savings=min_savings,
        )
        if provider_rec_items:
            existing_keys = {
                (
                    rec.provider,
                    rec.action,
                    _slugify_resource_token(rec.resource_type),
                    _slugify_resource_token(rec.resource_name),
                )
                for rec in recommendations_out
            }
            for rec in provider_rec_items:
                key = (
                    rec.provider,
                    rec.action,
                    _slugify_resource_token(rec.resource_type),
                    _slugify_resource_token(rec.resource_name),
                )
                if key in existing_keys:
                    continue
                existing_keys.add(key)
                recommendations_out.append(rec)
            if data_source == "no_data_available":
                data_source = "live_provider_recommendations"
            elif data_source not in {"multi_provider_api", "live_provider_recommendations"}:
                data_source = "multi_provider_api"
            total_analyzed += len(provider_rec_items)

    # ── Tier 3: Cost-trend signal from scan snapshot history ────────────────
    if not recommendations_out:
        acct_query = (
            db.query(ProviderAccount)
            .filter(ProviderAccount.organization_id == org_id)
        )
        if provider != "all":
            acct_query = acct_query.filter(ProviderAccount.provider == provider)
        accounts = acct_query.all()
        account_map = {a.id: a for a in accounts}

        if account_map:
            # Fetch all snapshots for these accounts in one query, ordered oldest-first
            snap_rows = (
                db.query(ProviderAccountSnapshot)
                .filter(
                    ProviderAccountSnapshot.provider_account_id.in_(list(account_map.keys())),
                    ProviderAccountSnapshot.organization_id == org_id,
                )
                .order_by(ProviderAccountSnapshot.captured_at.asc())
                .limit(500)
                .all()
            )

            # Group by account
            snapshots_by_account: Dict[int, List[Any]] = {}
            for snap in snap_rows:
                snapshots_by_account.setdefault(snap.provider_account_id, []).append(snap)

            # Build latest region-cost hotspots per account from allocation snapshots.
            alloc_rows = (
                db.query(CostAllocationSnapshot)
                .filter(
                    CostAllocationSnapshot.provider_account_id.in_(list(account_map.keys())),
                    CostAllocationSnapshot.organization_id == org_id,
                    CostAllocationSnapshot.customer_id == customer_id,
                )
                .order_by(CostAllocationSnapshot.captured_at.desc())
                .limit(2000)
                .all()
            )
            latest_scan_by_account: Dict[int, str] = {}
            top_regions_by_account: Dict[int, List[tuple[str, float]]] = {}
            for row in alloc_rows:
                acct_id = int(row.provider_account_id)
                if acct_id not in latest_scan_by_account:
                    latest_scan_by_account[acct_id] = str(row.scan_id or "")
                if str(row.scan_id or "") != latest_scan_by_account[acct_id]:
                    continue
                top_regions_by_account.setdefault(acct_id, []).append((str(row.region or "global"), float(row.cost_usd or 0.0)))
            for acct_id, regions in top_regions_by_account.items():
                top_regions_by_account[acct_id] = sorted(regions, key=lambda item: item[1], reverse=True)[:6]

            total_analyzed = len(snap_rows)
            tier3 = _rightsizing_from_snapshot_trends(
                snapshots_by_account, account_map, top_regions_by_account, min_savings
            )
            if tier3:
                recommendations_out = tier3
                data_source = "cost_trend_analysis"

    service_snapshot_query = db.query(CostSnapshot).filter(CostSnapshot.customer_id == customer_id)
    if provider != "all":
        service_snapshot_query = service_snapshot_query.filter(CostSnapshot.provider == provider)
    service_snapshot_rows = (
        service_snapshot_query
        .order_by(CostSnapshot.captured_at.desc())
        .limit(500)
        .all()
    )
    service_recommendations = _rightsizing_from_service_cost_snapshots(
        service_snapshot_rows,
        min_savings,
    )
    if service_recommendations:
        existing_service_keys = {
            (
                rec.provider,
                _slugify_resource_token(rec.resource_type),
                _slugify_resource_token(rec.resource_name),
            )
            for rec in recommendations_out
        }
        added_service_count = 0
        for rec in service_recommendations:
            key = (
                rec.provider,
                _slugify_resource_token(rec.resource_type),
                _slugify_resource_token(rec.resource_name),
            )
            if key in existing_service_keys:
                continue
            existing_service_keys.add(key)
            recommendations_out.append(rec)
            added_service_count += 1
        if added_service_count:
            if data_source == "no_data_available":
                data_source = "service_cost_snapshot"
            elif data_source != "service_cost_snapshot":
                data_source = "multi_source_cost_analysis"
            total_analyzed += added_service_count

    # ── Tier 4: Imported CSV cost-signal analysis ────────────────────────────
    if not recommendations_out and not require_live_provider_data:
        imported = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .limit(500)
            .all()
        )
        if provider != "all":
            imported = [r for r in imported if (r.provider or "").lower() == provider]
        if imported:
            total_analyzed = len(imported)
            tier4 = _rightsizing_from_imported_costs(imported, min_savings)
            if tier4:
                recommendations_out = tier4
                data_source = "imported_costs"

    if not recommendations_out and require_live_provider_data:
        return RightsizingResponse(
            generated_at=now_str,
            organization_id=org_id,
            data_source="live_provider_api",
            total_resources_analyzed=total_analyzed,
            rightsizable_count=0,
            total_monthly_savings_usd=0.0,
            total_annual_savings_usd=0.0,
            recommendations=[],
        )
    if not recommendations_out:
        return RightsizingResponse(
            generated_at=now_str,
            organization_id=org_id,
            data_source="no_data_available",
            total_resources_analyzed=total_analyzed,
            rightsizable_count=0,
            total_monthly_savings_usd=0.0,
            total_annual_savings_usd=0.0,
            recommendations=[],
        )

    # ── Apply min_savings filter ──────────────────────────────────────────────
    if min_savings > 0:
        recommendations_out = [
            r for r in recommendations_out if r.monthly_savings_usd >= min_savings
        ]

    # ── Attach direct console links (best effort) ───────────────────────────
    for rec in recommendations_out:
        if rec.resource_console_url:
            continue
        rec.resource_console_url = _rightsizing_console_url(
            provider=rec.provider,
            resource_id=rec.resource_id,
            region=rec.region,
            account_id=rec.account_id,
            resource_type=rec.resource_type,
        )

    # ── Sort by monthly savings desc, apply limit ────────────────────────────
    recommendations_out.sort(key=lambda r: r.monthly_savings_usd, reverse=True)
    recommendations_out = recommendations_out[:limit]

    try:
        _upsert_recommendation_ledger(
            db=db,
            organization_id=org_id,
            customer_id=customer_id,
            recommendations_in=recommendations_out,
            response_data_source=data_source,
        )
    except Exception as exc:
        db.rollback()
        logger.warning("Recommendation ledger upsert failed; returning rightsizing response: %s", exc)

    total_monthly = sum(r.monthly_savings_usd for r in recommendations_out)
    return RightsizingResponse(
        generated_at=now_str,
        organization_id=org_id,
        data_source=data_source,
        total_resources_analyzed=max(total_analyzed, len(recommendations_out)),
        rightsizable_count=len(recommendations_out),
        total_monthly_savings_usd=round(total_monthly, 2),
        total_annual_savings_usd=round(total_monthly * 12, 2),
        recommendations=recommendations_out,
    )


@router.get("/recommendations/ledger", response_model=RecommendationLedgerResponse)
async def get_recommendation_ledger(
    provider: str = Query("all"),
    status_filter: str = Query("all", alias="status"),
    limit: int = Query(200, ge=1, le=1000),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> RecommendationLedgerResponse:
    """Return the recommendation execution ledger for finance reconciliation."""
    _ = current_user
    org_id = _organization_id_for_membership(membership)
    rows = _query_recommendation_ledger(
        db=db,
        organization_id=org_id,
        provider=provider,
        status_filter=status_filter,
        limit=limit,
    )
    return _recommendation_ledger_response(organization_id=org_id, rows=rows)


@router.get("/recommendations/ledger.csv")
async def download_recommendation_ledger_csv(
    provider: str = Query("all"),
    status_filter: str = Query("all", alias="status"),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a finance-ready ledger export with planned, realized, and variance fields."""
    _ = current_user
    org_id = _organization_id_for_membership(membership)
    rows = _query_recommendation_ledger(
        db=db,
        organization_id=org_id,
        provider=provider,
        status_filter=status_filter,
        limit=5000,
    )
    now = _utcnow()
    for row in rows:
        row.last_exported_at = now
    db.commit()

    header = [
        "ledger_id",
        "organization_id",
        "provider",
        "account_id",
        "region",
        "resource_id",
        "resource_name",
        "resource_type",
        "recommendation_source",
        "recommendation_fingerprint",
        "action",
        "confidence",
        "effort",
        "status",
        "owner",
        "planned_monthly_savings_usd",
        "realized_monthly_savings_usd",
        "variance_monthly_usd",
        "planned_annual_savings_usd",
        "realized_annual_savings_usd",
        "variance_annual_usd",
        "variance_percent",
        "variance_reason",
        "current_monthly_cost_usd",
        "projected_monthly_cost_usd",
        "first_seen_at",
        "last_seen_at",
        "planned_at",
        "realized_at",
        "last_exported_at",
        "times_seen",
        "resource_console_url",
        "reason",
    ]
    data = []
    for row in rows:
        item = _ledger_item_from_row(row)
        data.append([
            item.id,
            item.organization_id,
            item.provider,
            item.account_id or "",
            item.region or "",
            item.resource_id,
            item.resource_name or "",
            item.resource_type or "",
            item.recommendation_source,
            item.recommendation_fingerprint,
            item.action,
            item.confidence,
            item.effort,
            item.status,
            item.owner or "",
            item.planned_monthly_savings_usd,
            item.realized_monthly_savings_usd,
            item.variance_monthly_usd,
            item.planned_annual_savings_usd,
            item.realized_annual_savings_usd,
            item.variance_annual_usd,
            item.variance_percent,
            item.variance_reason or "",
            item.current_monthly_cost_usd,
            item.projected_monthly_cost_usd,
            item.first_seen_at or "",
            item.last_seen_at or "",
            item.planned_at or "",
            item.realized_at or "",
            item.last_exported_at or "",
            item.times_seen,
            item.resource_console_url or "",
            item.reason or "",
        ])
    return _csv_response("optiora-recommendation-ledger.csv", header, data)


@router.patch("/recommendations/ledger/{ledger_id}", response_model=RecommendationLedgerItem)
async def update_recommendation_ledger_finance(
    ledger_id: int,
    payload: RecommendationLedgerFinanceUpdateRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> RecommendationLedgerItem:
    """Update realized savings and lifecycle fields for a ledger recommendation."""
    _ = current_user
    require_role(
        membership,
        [UserRole.OWNER, UserRole.ADMIN, UserRole.ANALYST],
        "update recommendation ledger",
    )
    org_id = _organization_id_for_membership(membership)
    row = (
        db.query(RecommendationLedger)
        .filter(
            RecommendationLedger.id == ledger_id,
            RecommendationLedger.organization_id == org_id,
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Recommendation ledger row not found")

    if payload.realized_monthly_savings_usd is not None:
        row.realized_monthly_savings_usd = round(float(payload.realized_monthly_savings_usd), 2)
        if payload.realized_annual_savings_usd is None:
            row.realized_annual_savings_usd = round(float(payload.realized_monthly_savings_usd) * 12.0, 2)
    if payload.realized_annual_savings_usd is not None:
        row.realized_annual_savings_usd = round(float(payload.realized_annual_savings_usd), 2)
        if payload.realized_monthly_savings_usd is None:
            row.realized_monthly_savings_usd = round(float(payload.realized_annual_savings_usd) / 12.0, 2)
    if payload.status is not None:
        row.status = payload.status
    if payload.owner is not None:
        row.owner = payload.owner.strip() or None
    if payload.variance_reason is not None:
        row.variance_reason = payload.variance_reason.strip() or None
    if payload.realized_at is not None:
        try:
            parsed_at = datetime.fromisoformat(payload.realized_at.replace("Z", "+00:00"))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="realized_at must be an ISO datetime") from exc
        row.realized_at = parsed_at.replace(tzinfo=None)
    elif payload.realized_monthly_savings_usd is not None or payload.realized_annual_savings_usd is not None:
        row.realized_at = _utcnow()

    variance_monthly, variance_annual, variance_percent = _ledger_variance(
        float(row.planned_monthly_savings_usd or 0.0),
        float(row.realized_monthly_savings_usd or 0.0),
    )
    row.variance_monthly_usd = variance_monthly
    row.variance_annual_usd = variance_annual
    row.variance_percent = variance_percent
    row.updated_at = _utcnow()
    db.commit()
    db.refresh(row)
    return _ledger_item_from_row(row)


def _is_vm_like_resource(rec: RightsizingRecommendation) -> bool:
    label = " ".join([
        str(rec.resource_type or ""),
        str(rec.resource_name or ""),
        str(rec.current_size or ""),
    ]).lower()
    vm_terms = [
        "vm",
        "virtual machine",
        "instance",
        "compute",
        "node",
        "ec2",
        "gke",
        "eks",
        "aks",
    ]
    return any(term in label for term in vm_terms)


@router.get("/analytics/vm-utilization-hotspots")
async def analytics_vm_utilization_hotspots(
    provider: str = "all",
    limit: int = Query(5, ge=1, le=20),
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Top VM hotspots by CPU/memory and best-effort proxies for disk/network.

    Uses connected-provider rightsizing/telemetry signals across all providers when
    available. When explicit CPU/memory telemetry is missing, returns a deterministic
    cross-provider proxy ranking based on latest monthly cost profile so callers still
    get a usable top-N response.
    """
    rightsizing = await get_rightsizing_recommendations(
        provider=provider,
        min_savings=0.0,
        limit=200,
        current_user=current_user,
        membership=membership,
        db=db,
    )
    vm_recs = [rec for rec in rightsizing.recommendations if _is_vm_like_resource(rec)]
    if not vm_recs:
        return {
            "generated_at": _utcnow().isoformat() + "Z",
            "provider": provider,
            "metric_sources": {
                "cpu": "none",
                "memory": "none",
                "disk_io": "none",
                "network_bandwidth": "none",
            },
            "top_cpu": [],
            "top_memory": [],
            "top_disk_io": [],
            "top_network_bandwidth": [],
            "notes": [
                "No VM-like resources available yet. Run provider scans and rightsizing sync first.",
            ],
        }

    def _to_item(
        rec: RightsizingRecommendation,
        metric_name: str,
        metric_value: float,
        metric_source: str,
    ) -> Dict[str, Any]:
        return {
            "resource_id": rec.resource_id,
            "resource_name": rec.resource_name,
            "provider": rec.provider,
            "region": rec.region,
            "resource_type": rec.resource_type,
            "metric": metric_name,
            "metric_value": round(float(metric_value or 0.0), 2),
            "metric_source": metric_source,
            "current_monthly_cost_usd": round(float(rec.current_monthly_cost_usd or 0.0), 2),
            "owner_hint": None,
            "resource_console_url": rec.resource_console_url,
            "last_observed_at": rec.last_observed_at,
        }

    cpu_items = [
        _to_item(rec, "cpu_utilization_percent", float(rec.cpu_utilization_avg_percent or 0.0), "telemetry_or_provider_signal")
        for rec in vm_recs
        if rec.cpu_utilization_avg_percent is not None
    ]
    memory_items = [
        _to_item(rec, "memory_utilization_percent", float(rec.memory_utilization_avg_percent or 0.0), "telemetry_or_provider_signal")
        for rec in vm_recs
        if rec.memory_utilization_avg_percent is not None
    ]

    # Proxy fallback when explicit telemetry is unavailable.
    # We use latest monthly cost among VM-like resources as a pressure proxy.
    cpu_proxy = [
        _to_item(
            rec,
            "cpu_proxy_index",
            float(rec.latest_monthly_cost_usd or rec.current_monthly_cost_usd or 0.0),
            "proxy_from_cost_profile",
        )
        for rec in vm_recs
    ]
    memory_proxy = [
        _to_item(
            rec,
            "memory_proxy_index",
            float(rec.latest_monthly_cost_usd or rec.current_monthly_cost_usd or 0.0),
            "proxy_from_cost_profile",
        )
        for rec in vm_recs
    ]
    disk_proxy = [
        _to_item(
            rec,
            "disk_io_proxy_index",
            float(rec.latest_monthly_cost_usd or rec.current_monthly_cost_usd or 0.0),
            "proxy_from_cost_profile",
        )
        for rec in vm_recs
    ]
    net_proxy = [
        _to_item(
            rec,
            "network_bandwidth_proxy_index",
            float(rec.latest_monthly_cost_usd or rec.current_monthly_cost_usd or 0.0),
            "proxy_from_cost_profile",
        )
        for rec in vm_recs
    ]

    cpu_items.sort(key=lambda item: float(item["metric_value"]), reverse=True)
    memory_items.sort(key=lambda item: float(item["metric_value"]), reverse=True)
    cpu_proxy.sort(key=lambda item: float(item["metric_value"]), reverse=True)
    memory_proxy.sort(key=lambda item: float(item["metric_value"]), reverse=True)
    disk_proxy.sort(key=lambda item: float(item["metric_value"]), reverse=True)
    net_proxy.sort(key=lambda item: float(item["metric_value"]), reverse=True)

    cpu_source = "telemetry_or_provider_signal" if cpu_items else "proxy_from_cost_profile"
    memory_source = "telemetry_or_provider_signal" if memory_items else "proxy_from_cost_profile"
    top_cpu = cpu_items[:limit] if cpu_items else cpu_proxy[:limit]
    top_memory = memory_items[:limit] if memory_items else memory_proxy[:limit]

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "provider": provider,
        "metric_sources": {
            "cpu": cpu_source,
            "memory": memory_source,
            "disk_io": "proxy_from_cost_profile",
            "network_bandwidth": "proxy_from_cost_profile",
        },
        "top_cpu": top_cpu,
        "top_memory": top_memory,
        "top_disk_io": disk_proxy[:limit],
        "top_network_bandwidth": net_proxy[:limit],
        "notes": [
            "CPU and memory rankings use connected-provider telemetry (CloudWatch, Azure Monitor, Cloud Monitoring) when available.",
            "When native CPU/memory telemetry is unavailable, rankings fall back to a deterministic cost-profile proxy.",
            "Disk I/O and network bandwidth use a cost-profile proxy when explicit metrics are not available.",
            "OCI and partial-provider scopes may still return proxy rankings until monitoring feeds are configured.",
        ],
    }


async def run_scheduled_scans_once(
    requested_organization_id: Optional[int] = None,
    sleep_between_retries: bool = True,
) -> Dict[str, Any]:
    """Trigger due scans for approved organizations based on configured cadence."""
    global _scheduler_running
    if _scheduler_running:
        return {
            "status": "busy",
            "started": 0,
            "failed": 0,
            "export_jobs_run": 0,
            "organization_id": requested_organization_id,
        }
    _scheduler_running = True
    started = 0
    failed = 0
    export_jobs_run = 0
    now = _utcnow()
    db = SessionLocal()
    try:
        permissions_query = db.query(ScanningPermissionRecord).filter(
            ScanningPermissionRecord.state.in_([ScanningState.APPROVED.value, ScanningState.RUNNING.value])
        )
        if requested_organization_id is not None:
            permissions_query = permissions_query.filter(
                ScanningPermissionRecord.customer_id == f"org-{requested_organization_id}"
            )
        permissions = permissions_query.all()
        for permission in permissions:
            effective_frequency = str(permission.scan_frequency or "daily").strip().lower()
            if bool(permission.scheduler_override_enabled):
                override_frequency = str(permission.scheduler_override_frequency or "").strip().lower()
                if override_frequency in {"hourly", "daily", "weekly"}:
                    effective_frequency = override_frequency
            cadence_seconds = _scan_interval_seconds(effective_frequency)
            max_attempts = max(1, int(permission.scheduler_retry_max_attempts or 1))
            backoff_seconds = max(15, int(permission.scheduler_retry_backoff_seconds or 15))
            overdue_alert_hours = max(1, int(permission.scheduler_overdue_alert_hours or 24))
            last_completed = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.customer_id == permission.customer_id,
                    ScanRunRecord.state == ScanningState.COMPLETED.value,
                )
                .order_by(ScanRunRecord.completed_at.desc())
                .first()
            )
            derived_org_id = _organization_id_from_customer_id(permission.customer_id)
            if (
                derived_org_id is not None
                and last_completed is not None
                and last_completed.completed_at is not None
            ):
                elapsed_since_success = int((now - last_completed.completed_at).total_seconds())
                overdue_threshold_seconds = cadence_seconds + (overdue_alert_hours * 3600)
                if elapsed_since_success > overdue_threshold_seconds:
                    recent_overdue = (
                        db.query(AlertEvent.id)
                        .filter(
                            AlertEvent.organization_id == derived_org_id,
                            AlertEvent.customer_id == permission.customer_id,
                            AlertEvent.alert_type == "scheduler.overdue",
                            AlertEvent.created_at >= now - timedelta(hours=overdue_alert_hours),
                        )
                        .first()
                    )
                    if recent_overdue is None:
                        overdue_message = (
                            f"Scheduler is overdue for {permission.customer_id}. "
                            f"Last successful run was {elapsed_since_success // 3600}h ago; "
                            f"expected cadence is {effective_frequency}."
                        )
                        db.add(
                            AlertEvent(
                                organization_id=derived_org_id,
                                customer_id=permission.customer_id,
                                scan_id=None,
                                alert_type="scheduler.overdue",
                                severity="warning",
                                title="Scheduled scans are overdue",
                                message=overdue_message,
                                delivered_channels_json="[]",
                                created_at=now,
                            )
                        )
                        db.add(
                            AuditLog(
                                organization_id=derived_org_id,
                                actor_user_id=None,
                                action="scan.schedule.overdue_alert",
                                entity_type="scanning_permission",
                                entity_id=str(permission.id),
                                metadata_json=json.dumps(
                                    {
                                        "customer_id": permission.customer_id,
                                        "effective_frequency": effective_frequency,
                                        "elapsed_seconds_since_success": elapsed_since_success,
                                        "overdue_threshold_seconds": overdue_threshold_seconds,
                                    }
                                ),
                                created_at=now,
                            )
                        )
                        db.commit()

            if last_completed and last_completed.completed_at:
                elapsed = (now - last_completed.completed_at).total_seconds()
                if elapsed < cadence_seconds:
                    continue

            providers = json.loads(permission.providers_json or "[]")
            if not providers:
                providers = ["aws", "azure", "gcp", "oci"]
            started += 1
            base_scan_id = f"scan_{permission.customer_id}_{int(now.timestamp())}"
            attempt_succeeded = False
            last_error: Optional[str] = None
            final_scan_id = base_scan_id
            for attempt in range(1, max_attempts + 1):
                scan_id = base_scan_id if attempt == 1 else f"{base_scan_id}_retry{attempt}"
                final_scan_id = scan_id
                row = ScanRunRecord(
                    scan_id=scan_id,
                    customer_id=permission.customer_id,
                    state=ScanningState.RUNNING.value,
                    providers_json=json.dumps(providers),
                    progress=0,
                    started_at=_utcnow(),
                )
                db.add(row)
                db.commit()
                try:
                    await _run_cost_analysis(
                        scan_id=scan_id,
                        customer_id=permission.customer_id,
                        providers=providers,
                        raise_on_error=True,
                    )
                    attempt_succeeded = True
                    last_error = None
                    if derived_org_id is not None:
                        db.add(
                            AuditLog(
                                organization_id=derived_org_id,
                                actor_user_id=None,
                                action="scan.schedule.triggered",
                                entity_type="scan_run",
                                entity_id=scan_id,
                                metadata_json=json.dumps(
                                    {
                                        "customer_id": permission.customer_id,
                                        "frequency": effective_frequency,
                                        "providers": providers,
                                        "attempt": attempt,
                                    }
                                ),
                            )
                        )
                        db.commit()
                    break
                except Exception as exc:
                    last_error = str(exc)
                    if derived_org_id is not None:
                        db.add(
                            AuditLog(
                                organization_id=derived_org_id,
                                actor_user_id=None,
                                action="scan.schedule.retry_failed",
                                entity_type="scan_run",
                                entity_id=scan_id,
                                metadata_json=json.dumps(
                                    {
                                        "customer_id": permission.customer_id,
                                        "frequency": effective_frequency,
                                        "providers": providers,
                                        "attempt": attempt,
                                        "max_attempts": max_attempts,
                                        "error": last_error,
                                    }
                                ),
                            )
                        )
                        db.commit()
                    if attempt < max_attempts and sleep_between_retries:
                        await asyncio.sleep(backoff_seconds * (2 ** (attempt - 1)))

            if not attempt_succeeded:
                failed += 1
                if derived_org_id is not None:
                    db.add(
                        AlertEvent(
                            organization_id=derived_org_id,
                            customer_id=permission.customer_id,
                            scan_id=final_scan_id,
                            alert_type="scheduler.scan_failure",
                            severity="critical",
                            title="Scheduled scan failed after retries",
                            message=(
                                f"Scheduled scan for {permission.customer_id} failed after "
                                f"{max_attempts} attempt(s). Last error: {last_error or 'unknown'}"
                            )[:1000],
                            delivered_channels_json="[]",
                            created_at=_utcnow(),
                        )
                    )
                    db.commit()

        jobs_query = db.query(ExportJob).filter(ExportJob.is_active == True)  # noqa: E712
        if requested_organization_id is not None:
            jobs_query = jobs_query.filter(ExportJob.organization_id == requested_organization_id)
        jobs = jobs_query.all()
        for job in jobs:
            cadence_seconds = _scan_interval_seconds(job.schedule_frequency)
            anchor = job.last_run_at or job.created_at or now
            elapsed = (now - anchor).total_seconds()
            if elapsed < cadence_seconds:
                continue

            membership_stub = SimpleNamespace(
                organization_id=job.organization_id,
                role=UserRole.OWNER,
            )
            user_stub = SimpleNamespace(id=None)
            await _execute_export_job(
                db=db,
                job=job,
                current_user=user_stub,
                membership=membership_stub,
                actor_user_id=None,
            )
            export_jobs_run += 1
    finally:
        db.close()
        _scheduler_running = False
    return {
        "status": "ok",
        "started": started,
        "failed": failed,
        "export_jobs_run": export_jobs_run,
        "organization_id": requested_organization_id,
    }
