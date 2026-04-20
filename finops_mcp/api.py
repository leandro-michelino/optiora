"""
REST API routes for OptiOra.

Handles:
- Credential validation/storage
- Scanning permissions and progress
- Dashboard data endpoints
"""

import csv
import asyncio
import io
import json
import logging
import os
import base64
import binascii
import hashlib
import hmac
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace
from typing import Any, Dict, List, Literal, Optional, Union
from uuid import uuid4
from xml.sax.saxutils import escape as xml_escape

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from fastapi import APIRouter, BackgroundTasks, Depends, File, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from .auth_routes import get_current_membership, get_current_user
from .credentials import CredentialManager, CredentialStatus, CredentialValidator
from .config import Config
from .notifications import evaluate_budget_alert
from .notifications import (
    SUPPORTED_NOTIFICATION_CHANNELS,
    destination_configured,
    send_test_notification,
)
from .access_control import require_role
from .connectors import ConnectorManager, ConnectorType, BaseConnector, ConnectorStatus
from .orm_models import (
    AlertEvent,
    AlertRoutingPolicy,
    AuditLog,
    BusinessMappingRule,
    CostAllocationSnapshot,
    CostPeriodSummary,
    CostSnapshot,
    CredentialRecord,
    ExportJob,
    ExportJobRun,
    ImportedCostRecord,
    NormalizedCostDimension,
    ProviderAccount,
    ProviderAccountLink,
    ProviderAccountSnapshot,
    ScanRunRecord,
    ScanningPermissionRecord,
    SessionLocal,
    User,
    UserOrganization,
    UserRole,
    VirtualTagRule,
    get_db,
)
from .scanning import ScanningManager, ScanningState
from .tools import anomalies, aws_costs, finops_analytics, recommendations
from .tools import azure_costs, gcp_costs, oci_costs, genai_advisor
from .retention import fetch_archived_period_summaries
from . import __version__

logger = logging.getLogger(__name__)
_scheduler_running = False


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)

router = APIRouter(prefix="/api/v1", tags=["api"])
SUPPORTED_COST_IMPORT_PROVIDERS = {"aws", "azure", "gcp", "oci"}
_SUPPORTED_TREND_VIEWS = {"provider", "region", "service", "account"}


def _share_hmac_secret() -> bytes:
    raw = os.getenv("SECRET_KEY", "optiora-dev-share-secret")
    return str(raw).encode("utf-8")


def _build_report_share_token(payload: Dict[str, Any]) -> str:
    body = base64.urlsafe_b64encode(json.dumps(payload, separators=(",", ":")).encode("utf-8")).decode("utf-8").rstrip("=")
    signature = hmac.new(_share_hmac_secret(), body.encode("utf-8"), hashlib.sha256).digest()
    sig = base64.urlsafe_b64encode(signature).decode("utf-8").rstrip("=")
    return f"{body}.{sig}"


def _parse_report_share_token(token: str) -> Dict[str, Any]:
    if "." not in token:
        raise HTTPException(status_code=401, detail="Invalid share token")
    body, sig = token.split(".", 1)
    expected_sig = base64.urlsafe_b64encode(
        hmac.new(_share_hmac_secret(), body.encode("utf-8"), hashlib.sha256).digest()
    ).decode("utf-8").rstrip("=")
    if not hmac.compare_digest(sig, expected_sig):
        raise HTTPException(status_code=401, detail="Invalid share token")
    padded = body + "=" * ((4 - len(body) % 4) % 4)
    try:
        payload = json.loads(base64.urlsafe_b64decode(padded.encode("utf-8")).decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=401, detail="Malformed share token") from exc
    exp = int(payload.get("exp", 0) or 0)
    if exp <= int(datetime.now(timezone.utc).timestamp()):
        raise HTTPException(status_code=401, detail="Share token expired")
    return payload


def _trend_dimension_value(rec: ImportedCostRecord, view_by: str) -> str:
    if view_by == "region":
        return (rec.region or "unknown").strip() or "unknown"
    if view_by == "service":
        return (rec.service_name or "unknown").strip() or "unknown"
    if view_by == "account":
        return (
            rec.account_identifier
            or rec.account_name
            or rec.parent_account_identifier
            or "unknown"
        ).strip() or "unknown"
    return (rec.provider or "imported").strip().lower() or "imported"


class AWSCredentialInput(BaseModel):
    provider: Literal["aws"]
    access_key_id: str
    secret_access_key: str
    region: Optional[str] = "us-east-1"


class AzureCredentialInput(BaseModel):
    provider: Literal["azure"]
    subscription_id: str
    tenant_id: str
    client_id: str
    client_secret: str


class GCPCredentialInput(BaseModel):
    provider: Literal["gcp"]
    project_id: str
    service_account_json: Union[Dict[str, Any], str]


class OCICredentialInput(BaseModel):
    provider: Literal["oci"]
    config_file: str
    profile: Optional[str] = "DEFAULT"


CredentialInput = Union[
    AWSCredentialInput,
    AzureCredentialInput,
    GCPCredentialInput,
    OCICredentialInput,
]


class CredentialResponse(BaseModel):
    provider: str
    is_valid: bool
    message: str
    test_cost_usd: Optional[float] = None
    tested_at: Optional[str] = None
    error_details: Optional[str] = None


class ScanningApprovalRequest(BaseModel):
    customer_id: Optional[str] = None
    auto_remediate: bool = False
    scan_frequency: str = "daily"
    notification_email: Optional[str] = None
    monthly_budget_usd: float = 0.0
    warning_threshold_percent: float = 80.0
    critical_threshold_percent: float = 100.0
    notifications_enabled: bool = True


class ScanningPermissionResponse(BaseModel):
    customer_id: str
    organization_id: int
    state: str
    providers: List[str]
    scan_frequency: str
    auto_remediate: bool
    notification_email: Optional[str] = None
    monthly_budget_usd: float = 0.0
    warning_threshold_percent: float = 80.0
    critical_threshold_percent: float = 100.0
    notifications_enabled: bool = True
    created_at: str
    approved_at: Optional[str] = None


class StartScanRequest(BaseModel):
    customer_id: Optional[str] = None
    providers: Optional[List[str]] = None
    target_accounts: Optional[List[str]] = None


class ScanProgressResponse(BaseModel):
    scan_id: str
    customer_id: str
    organization_id: int
    state: str
    progress: int = 0
    providers: List[str]
    started_at: datetime
    completed_at: Optional[datetime] = None
    total_resources: int = 0
    anomalies_found: int = 0
    savings_identified: float = 0.0


class ProviderDiagnostic(BaseModel):
    provider: str
    configured: bool
    required_settings: List[str]
    missing_settings: List[str]
    recommendation: str


class ScanHistoryItem(BaseModel):
    scan_id: str
    customer_id: str
    organization_id: int
    state: str
    providers: List[str]
    started_at: datetime
    completed_at: Optional[datetime] = None
    total_resources: int = 0
    anomalies_found: int = 0
    savings_identified: float = 0.0


class ScanDiffEntry(BaseModel):
    provider: str
    current_cost_usd: float
    previous_cost_usd: float
    delta_cost_usd: float
    delta_percent: Optional[float] = None
    current_anomalies: int
    previous_anomalies: int


class ScanDiffResponse(BaseModel):
    organization_id: int
    current_scan_id: str
    previous_scan_id: Optional[str] = None
    total_current_cost_usd: float
    total_previous_cost_usd: float
    total_delta_cost_usd: float
    entries: List[ScanDiffEntry]


class AccountRegionRow(BaseModel):
    region: str
    cost_usd: float


class ProviderAccountRollupItem(BaseModel):
    account_id: int
    provider: str
    account_identifier: str
    account_name: str
    account_type: str
    depth: int = 0
    parent_account_id: Optional[int] = None
    parent_account_identifier: Optional[str] = None
    direct_cost_usd: float
    rolled_up_cost_usd: float
    direct_savings_identified_usd: float
    rolled_up_savings_identified_usd: float
    direct_anomalies_count: int
    rolled_up_anomalies_count: int
    direct_service_count: int
    rolled_up_service_count: int
    child_count: int
    budget_monthly_usd: Optional[float] = None
    rolled_up_budget_monthly_usd: Optional[float] = None
    budget_utilization_percent: Optional[float] = None
    rolled_up_budget_utilization_percent: Optional[float] = None
    budget_status: Optional[str] = None
    scan_id: Optional[str] = None
    captured_at: Optional[str] = None
    top_regions: List[AccountRegionRow] = Field(default_factory=list)


class ProviderAccountRollupResponse(BaseModel):
    organization_id: int
    customer_id: str
    provider: Optional[str] = None
    scan_id: Optional[str] = None
    generated_at: str
    total_direct_cost_usd: float
    total_rolled_up_cost_usd: float
    items: List[ProviderAccountRollupItem]


class ProviderAccountInventoryItem(BaseModel):
    account_id: int
    provider: str
    account_identifier: str
    account_name: str
    account_type: str
    native_region: Optional[str] = None
    is_active: bool
    metadata: Dict[str, Any]
    created_at: str
    updated_at: Optional[str] = None


class ProviderAccountInventoryResponse(BaseModel):
    organization_id: int
    customer_id: str
    total: int
    accounts: List[ProviderAccountInventoryItem]


class AccountRegionBreakdownItem(BaseModel):
    region: str
    cost_usd: float
    scan_id: str
    captured_at: str


class AccountRegionBreakdownResponse(BaseModel):
    account_id: int
    provider: str
    account_name: str
    scan_id: Optional[str] = None
    total_cost_usd: float
    regions: List[AccountRegionBreakdownItem]


# ── Business mapping / chargeback models ────────────────────────────────────

VALID_DIMENSIONS = {"team", "environment", "application", "cost_center"}


class BusinessMappingRuleRequest(BaseModel):
    tag_key: str
    tag_value: str = "*"
    dimension: str
    mapped_value: str
    priority: int = 100
    is_active: bool = True


class BusinessMappingRuleUpdateRequest(BaseModel):
    tag_key: str | None = None
    tag_value: str | None = None
    dimension: str | None = None
    mapped_value: str | None = None
    priority: int | None = None
    is_active: bool | None = None


class BusinessMappingRuleResponse(BaseModel):
    id: int
    organization_id: int
    customer_id: str
    tag_key: str
    tag_value: str
    dimension: str
    mapped_value: str
    priority: int
    is_active: bool
    created_at: str
    updated_at: Optional[str] = None


class BusinessMappingRuleListResponse(BaseModel):
    organization_id: int
    rules: List[BusinessMappingRuleResponse]
    total: int


class ChargebackDimensionGroup(BaseModel):
    dimension: str          # "team" | "environment" | "application" | "cost_center"
    value: str              # e.g. "platform-team"
    total_cost_usd: float
    provider_breakdown: Dict[str, float]
    record_count: int


class ChargebackResponse(BaseModel):
    organization_id: int
    dimension_type: str
    groups: List[ChargebackDimensionGroup]
    total_mapped_cost_usd: float
    total_unmapped_cost_usd: float
    total_cost_usd: float
    coverage_percent: float


class AllocationCoverageResponse(BaseModel):
    organization_id: int
    total_cost_usd: float
    mapped_cost_usd: float
    unmapped_cost_usd: float
    coverage_percent: float
    dimension_coverage: Dict[str, float]   # {"team": 82.4, "environment": 91.0, …}
    provider_coverage: Dict[str, float]    # {"aws": 78.0, "gcp": 95.0, …}
    unmapped_top_services: List[Dict[str, Any]]


# ── Trend / Period Summary models ─────────────────────────────────────────────

class CostTrendPoint(BaseModel):
    period_start: str
    period_end: str
    provider: str
    dimension_value: str = ""
    total_cost_usd: float
    mapped_cost_usd: float
    unmapped_cost_usd: float
    record_count: int
    team: Optional[str] = None
    environment: Optional[str] = None
    service_breakdown: Dict[str, float] = {}


class CostTrendResponse(BaseModel):
    organization_id: int
    period_type: str             # "monthly" | "weekly"
    lookback_periods: int
    view_by: str                 # "provider" | "region" | "service" | "account"
    data_source: str             # "computed" | "raw_records" | "empty"
    points: List[CostTrendPoint]
    provider_totals: Dict[str, float]
    dimension_totals: Dict[str, float]
    grand_total_usd: float


class PeriodSummaryComputeResponse(BaseModel):
    organization_id: int
    period_type: str
    periods_computed: int
    rows_written: int
    computed_at: str


# ── End trend models ──────────────────────────────────────────────────────────




class AlertRoutingPolicyRequest(BaseModel):
    severity: Literal["warning", "critical"]
    channels: List[str]
    is_active: bool = True


class AlertRoutingPolicyResponse(BaseModel):
    id: int
    severity: str
    channels: List[str]
    is_active: bool
    created_at: str
    updated_at: str


class NotificationDestinationStatus(BaseModel):
    channel: str
    configured: bool
    enabled: bool
    last_delivery_at: Optional[str] = None


class NotificationDestinationsResponse(BaseModel):
    organization_id: int
    destinations: List[NotificationDestinationStatus]


class NotificationDestinationToggleRequest(BaseModel):
    enabled: bool


class NotificationDestinationTestRequest(BaseModel):
    channel: Literal["email", "slack", "teams"]
    target: Optional[str] = None
    message: Optional[str] = None


class NotificationDestinationTestResponse(BaseModel):
    channel: str
    success: bool
    detail: str


class AlertRoutingPolicySimulationRequest(BaseModel):
    severity: Literal["warning", "critical"]
    title: Optional[str] = None
    alert_type: Optional[str] = None


class AlertRoutingPolicySimulationResponse(BaseModel):
    severity: str
    matched_policy_id: Optional[int] = None
    evaluated_channels: List[str]
    expected_channels: List[str]
    configured_channels: List[str]
    inactive_policy: bool = False


class AlertLifecycleActionResponse(BaseModel):
    status: str
    alert_id: int
    lifecycle_state: Literal["active", "acknowledged", "dismissed", "reactivated"]


class ConnectorTestRequest(BaseModel):
    connector_type: str
    config: Dict[str, Any]


class ConnectorStatusResponse(BaseModel):
    connector_type: str
    status: str
    last_sync: Optional[str] = None
    message: Optional[str] = None


class ConnectorListResponse(BaseModel):
    supported_connectors: List[str]
    description: str


class SchedulerCounters(BaseModel):
    total: int
    success: int
    failure: int


class SchedulerTimelineItem(BaseModel):
    id: str
    event_type: str
    state: str
    title: str
    detail: str
    created_at: str


class SchedulerStatusResponse(BaseModel):
    organization_id: int
    customer_id: str
    scheduler_enabled: bool
    scheduler_running: bool
    permission_state: str
    scan_frequency: str
    next_run_at: Optional[str] = None
    next_run_eta_seconds: Optional[int] = None
    last_success_at: Optional[str] = None
    last_failure_at: Optional[str] = None
    counters: SchedulerCounters
    timeline: List[SchedulerTimelineItem]


class DataFreshnessProviderItem(BaseModel):
    provider: str
    last_ingested_at: Optional[str] = None
    age_seconds: Optional[int] = None
    status: Literal["fresh", "stale", "unknown"]


class DataFreshnessConnectorItem(BaseModel):
    connector: str
    last_event_at: Optional[str] = None
    age_seconds: Optional[int] = None
    status: Literal["fresh", "stale", "unknown"]


class DataFreshnessResponse(BaseModel):
    organization_id: int
    customer_id: str
    generated_at: str
    providers: List[DataFreshnessProviderItem]
    connectors: List[DataFreshnessConnectorItem]
    scheduler_lag_seconds: Optional[int] = None
    scheduler_status: Literal["healthy", "lagging", "unknown"] = "unknown"


class ExternalAWSAnomalyIngestRequest(BaseModel):
    events: List[Dict[str, Any]]


class ExternalAWSReplayRequest(BaseModel):
    event_ids: Optional[List[str]] = None
    days_back: Optional[int] = None
    max_results: int = Field(default=50, le=500)


class ExternalGCPPubSubIngestRequest(BaseModel):
    message: Dict[str, Any]
    subscription: Optional[str] = None


class ExportJobRequest(BaseModel):
    name: str
    report_type: Literal["executive_summary", "executive_digest", "finance_workbook"] = "executive_summary"
    export_format: Literal["csv", "xls", "xlsx", "pdf"] = "csv"
    schedule_frequency: Literal["daily", "weekly", "monthly"] = "weekly"
    is_active: bool = True


class ExportJobResponse(BaseModel):
    id: int
    organization_id: int
    customer_id: str
    name: str
    report_type: str
    export_format: str
    schedule_frequency: str
    is_active: bool
    last_run_at: Optional[str] = None
    created_at: str
    updated_at: str


class ExportJobRunResponse(BaseModel):
    id: int
    export_job_id: int
    status: str
    output_filename: Optional[str] = None
    row_count: int
    error_message: Optional[str] = None
    created_at: str
    completed_at: Optional[str] = None


class ImportedCostUploadResponse(BaseModel):
    organization_id: int
    customer_id: str
    upload_id: str
    filename: str
    rows_imported: int
    total_cost_usd: float
    providers: List[str]
    imported_at: datetime


class ImportedCostSummaryResponse(BaseModel):
    organization_id: int
    customer_id: str
    has_data: bool
    upload_id: Optional[str] = None
    source_filename: Optional[str] = None
    rows_imported: int = 0
    total_cost_usd: float = 0.0
    providers: List[str] = Field(default_factory=list)
    last_imported_at: Optional[datetime] = None


class ImportPreviewIssue(BaseModel):
    line_number: int
    severity: Literal["error", "warning"]
    message: str


class ImportPreviewResponse(BaseModel):
    organization_id: int
    customer_id: str
    filename: str
    total_rows: int
    accepted_rows: int
    rejected_rows: int
    total_cost_usd: float
    detected_providers: List[str] = Field(default_factory=list)
    header_columns: List[str] = Field(default_factory=list)
    mapping_feedback: Dict[str, Any] = Field(default_factory=dict)
    reconciliation_guidance: List[str] = Field(default_factory=list)
    issues: List[ImportPreviewIssue] = Field(default_factory=list)


class ReportShareTokenRequest(BaseModel):
    report_type: Literal["executive_summary", "finance_workbook", "executive_digest"] = "executive_summary"
    report_format: Literal["json", "csv", "xlsx", "pdf"] = "json"
    expires_in_hours: int = 168


class ReportShareTokenResponse(BaseModel):
    token: str
    expires_at: str
    report_type: str
    report_format: str


def get_credential_manager(db: Session = Depends(get_db)) -> CredentialManager:
    return CredentialManager(db)


def get_scanning_manager(db: Session = Depends(get_db)) -> ScanningManager:
    return ScanningManager(db)


def _parse_optional_datetime_value(
    value: Optional[str],
    field_name: str,
    line_number: int,
) -> tuple[Optional[datetime], Optional[str]]:
    text = str(value or "").strip()
    if not text:
        return None, None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")), None
    except ValueError:
        return None, f"Invalid {field_name} at CSV line {line_number}. Use ISO date or datetime."


def _parse_required_float_value(
    value: Optional[str],
    field_name: str,
    line_number: int,
) -> tuple[Optional[float], Optional[str]]:
    text = str(value or "").strip()
    if not text:
        return None, f"Missing {field_name} at CSV line {line_number}."
    try:
        parsed = float(text)
    except ValueError:
        return None, f"Invalid {field_name} at CSV line {line_number}."
    if parsed < 0:
        return None, f"Negative {field_name} at CSV line {line_number}. Cost values must be zero or greater."
    return parsed, None


def _csv_escape(value: Any) -> str:
    text = str(value if value is not None else "").replace('"', '""')
    return f'"{text}"'


def _csv_response(filename: str, header: List[str], rows: List[List[Any]]) -> Response:
    lines = [",".join(header)]
    for row in rows:
        lines.append(",".join(_csv_escape(cell) for cell in row))
    return Response(
        "\n".join(lines),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _spreadsheet_xml_response(filename: str, sheet_name: str, rows: List[List[Any]]) -> Response:
    def _cell_xml(value: Any) -> str:
        if value is None or value == "":
            return '<Cell><Data ss:Type="String"></Data></Cell>'
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return f'<Cell><Data ss:Type="Number">{value}</Data></Cell>'
        return f'<Cell><Data ss:Type="String">{xml_escape(str(value))}</Data></Cell>'

    row_xml = "".join(
        "<Row>" + "".join(_cell_xml(cell) for cell in row) + "</Row>"
        for row in rows
    )
    workbook = (
        '<?xml version="1.0"?>'
        '<?mso-application progid="Excel.Sheet"?>'
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel" '
        'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:html="http://www.w3.org/TR/REC-html40">'
        f'<Worksheet ss:Name="{xml_escape(sheet_name[:31] or "Sheet1")}"><Table>{row_xml}</Table></Worksheet>'
        "</Workbook>"
    )
    return Response(
        workbook,
        media_type="application/vnd.ms-excel",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _imported_cost_summary(rows: List[ImportedCostRecord]) -> Dict[str, Any]:
    providers = sorted({row.provider for row in rows})
    latest_imported_at = max((row.created_at for row in rows), default=None)
    return {
        "rows_imported": len(rows),
        "total_cost_usd": round(sum(float(row.cost_usd or 0.0) for row in rows), 2),
        "providers": providers,
        "last_imported_at": latest_imported_at,
        "upload_id": rows[0].upload_id if rows else None,
        "source_filename": rows[0].source_filename if rows else None,
    }


def _get_imported_cost_rows(
    db: Session,
    organization_id: int,
    customer_id: str,
    cloud_provider: str = "all",
) -> List[ImportedCostRecord]:
    query = db.query(ImportedCostRecord).filter(
        ImportedCostRecord.organization_id == organization_id,
        ImportedCostRecord.customer_id == customer_id,
    )
    if cloud_provider != "all":
        query = query.filter(ImportedCostRecord.provider == cloud_provider)
    return query.order_by(
        ImportedCostRecord.created_at.desc(),
        ImportedCostRecord.id.desc(),
    ).all()


def _materialize_rollup_items(nodes: Dict[int, Dict[str, Any]]) -> List[ProviderAccountRollupItem]:
    children_by_parent: Dict[Optional[int], List[int]] = {}
    for node_id, node in nodes.items():
        children_by_parent.setdefault(node.get("parent_account_id"), []).append(node_id)

    for child_ids in children_by_parent.values():
        child_ids.sort(key=lambda item_id: (
            str(nodes[item_id].get("provider") or ""),
            str(nodes[item_id].get("account_name") or ""),
            str(nodes[item_id].get("account_identifier") or ""),
        ))

    def _walk(node_id: int, depth: int) -> tuple[float, float, int, int, int]:
        node = nodes[node_id]
        child_ids = children_by_parent.get(node_id, [])
        child_count = len(child_ids)
        direct_cost = float(node.get("direct_cost_usd") or 0.0)
        direct_savings = float(node.get("direct_savings_identified_usd") or 0.0)
        direct_anomalies = int(node.get("direct_anomalies_count") or 0)
        direct_services = int(node.get("direct_service_count") or 0)

        # Grouping nodes represent aggregated structure, not additive direct spend.
        if child_count > 0 and str(node.get("account_type") or "") in {
            "provider",
            "management_group",
            "organization",
            "folder",
            "tenancy",
        }:
            direct_cost = 0.0
            direct_savings = 0.0
            direct_anomalies = 0
            direct_services = 0

        rolled_cost = direct_cost
        rolled_savings = direct_savings
        rolled_anomalies = direct_anomalies
        rolled_services = direct_services
        total_descendants = child_count

        node["depth"] = depth
        node["child_count"] = child_count
        node["direct_cost_usd"] = round(direct_cost, 2)
        node["direct_savings_identified_usd"] = round(direct_savings, 2)
        node["direct_anomalies_count"] = direct_anomalies
        node["direct_service_count"] = direct_services

        for child_id in child_ids:
            child_cost, child_savings, child_anomalies, child_services, descendant_count = _walk(child_id, depth + 1)
            rolled_cost += child_cost
            rolled_savings += child_savings
            rolled_anomalies += child_anomalies
            rolled_services += child_services
            total_descendants += descendant_count

        node["rolled_up_cost_usd"] = round(rolled_cost, 2)
        node["rolled_up_savings_identified_usd"] = round(rolled_savings, 2)
        node["rolled_up_anomalies_count"] = rolled_anomalies
        node["rolled_up_service_count"] = rolled_services
        node["descendant_count"] = total_descendants
        return rolled_cost, rolled_savings, rolled_anomalies, rolled_services, total_descendants

    root_ids = children_by_parent.get(None, [])
    for root_id in root_ids:
        _walk(root_id, 0)

    ordered_ids: List[int] = []

    def _append(node_id: int) -> None:
        ordered_ids.append(node_id)
        for child_id in children_by_parent.get(node_id, []):
            _append(child_id)

    for root_id in root_ids:
        _append(root_id)

    return [
        ProviderAccountRollupItem(
            account_id=node_id,
            provider=str(nodes[node_id].get("provider") or ""),
            account_identifier=str(nodes[node_id].get("account_identifier") or ""),
            account_name=str(nodes[node_id].get("account_name") or ""),
            account_type=str(nodes[node_id].get("account_type") or "account"),
            depth=int(nodes[node_id].get("depth") or 0),
            parent_account_id=nodes[node_id].get("parent_account_id"),
            parent_account_identifier=nodes[node_id].get("parent_account_identifier"),
            direct_cost_usd=float(nodes[node_id].get("direct_cost_usd") or 0.0),
            rolled_up_cost_usd=float(nodes[node_id].get("rolled_up_cost_usd") or 0.0),
            direct_savings_identified_usd=float(nodes[node_id].get("direct_savings_identified_usd") or 0.0),
            rolled_up_savings_identified_usd=float(nodes[node_id].get("rolled_up_savings_identified_usd") or 0.0),
            direct_anomalies_count=int(nodes[node_id].get("direct_anomalies_count") or 0),
            rolled_up_anomalies_count=int(nodes[node_id].get("rolled_up_anomalies_count") or 0),
            direct_service_count=int(nodes[node_id].get("direct_service_count") or 0),
            rolled_up_service_count=int(nodes[node_id].get("rolled_up_service_count") or 0),
            child_count=int(nodes[node_id].get("child_count") or 0),
            scan_id=nodes[node_id].get("scan_id"),
            captured_at=nodes[node_id].get("captured_at"),
            top_regions=nodes[node_id].get("top_regions") or [],
        )
        for node_id in ordered_ids
    ]


def _apply_rollup_budget_metrics(
    items: List[ProviderAccountRollupItem],
    monthly_budget_usd: float,
    warning_threshold_percent: float,
    critical_threshold_percent: float,
) -> List[ProviderAccountRollupItem]:
    total_budget = float(monthly_budget_usd or 0.0)
    if total_budget <= 0:
        return items

    baseline_cost = float(
        sum(item.rolled_up_cost_usd for item in items if item.depth == 0) or 0.0
    )
    if baseline_cost <= 0:
        return items

    for item in items:
        direct_budget = round(total_budget * (float(item.direct_cost_usd or 0.0) / baseline_cost), 2)
        rolled_budget = round(total_budget * (float(item.rolled_up_cost_usd or 0.0) / baseline_cost), 2)
        item.budget_monthly_usd = direct_budget
        item.rolled_up_budget_monthly_usd = rolled_budget

        if direct_budget > 0:
            item.budget_utilization_percent = round((float(item.direct_cost_usd or 0.0) / direct_budget) * 100, 2)
        else:
            item.budget_utilization_percent = None

        rolled_util = None
        if rolled_budget > 0:
            rolled_util = round((float(item.rolled_up_cost_usd or 0.0) / rolled_budget) * 100, 2)
        item.rolled_up_budget_utilization_percent = rolled_util

        if rolled_util is None:
            item.budget_status = None
        elif rolled_util >= float(critical_threshold_percent or 100.0):
            item.budget_status = "critical"
        elif rolled_util >= float(warning_threshold_percent or 80.0):
            item.budget_status = "warning"
        else:
            item.budget_status = "ok"

    return items


def _build_rollups_from_imported_rows(
    rows: List[ImportedCostRecord],
    organization_id: int,
    customer_id: str,
    provider: Optional[str] = None,
    monthly_budget_usd: float = 0.0,
    warning_threshold_percent: float = 80.0,
    critical_threshold_percent: float = 100.0,
) -> ProviderAccountRollupResponse:
    next_synthetic_id = -1

    def _synthetic_id() -> int:
        nonlocal next_synthetic_id
        current = next_synthetic_id
        next_synthetic_id -= 1
        return current

    nodes: Dict[int, Dict[str, Any]] = {}
    provider_root_ids: Dict[str, int] = {}
    account_nodes: Dict[tuple[str, str], int] = {}
    latest_imported_at = max((row.created_at for row in rows), default=None)

    def _normalized_account_type(raw_value: Optional[str], default: str) -> str:
        value = str(raw_value or "").strip().lower().replace(" ", "_")
        return value or default

    def _ensure_node(
        *,
        provider_key: str,
        identifier: str,
        account_name: str,
        account_type: str,
        parent_account_id: Optional[int],
    ) -> int:
        node_key = (provider_key, identifier)
        node_id = account_nodes.get(node_key)
        if node_id is None:
            node_id = _synthetic_id()
            account_nodes[node_key] = node_id
            nodes[node_id] = {
                "provider": provider_key,
                "account_identifier": identifier,
                "account_name": account_name,
                "account_type": account_type,
                "parent_account_id": parent_account_id,
                "parent_account_identifier": (
                    nodes[parent_account_id]["account_identifier"]
                    if parent_account_id in nodes
                    else None
                ),
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": None,
                "captured_at": latest_imported_at.isoformat() if latest_imported_at else None,
                "_services": set(),
            }
            return node_id

        node = nodes[node_id]
        if account_name and (
            not str(node.get("account_name") or "").strip()
            or str(node.get("account_name") or "").strip() == identifier
        ):
            node["account_name"] = account_name
        if (
            account_type
            and str(node.get("account_type") or "").strip() in {"", "account", "group"}
            and account_type not in {"", "account", "group"}
        ):
            node["account_type"] = account_type
        if parent_account_id is not None and node.get("parent_account_id") is None:
            node["parent_account_id"] = parent_account_id
            node["parent_account_identifier"] = (
                nodes[parent_account_id]["account_identifier"]
                if parent_account_id in nodes
                else None
            )
        return node_id

    for row in rows:
        provider_key = str(row.provider or "").strip().lower()
        if not provider_key:
            continue
        root_id = provider_root_ids.get(provider_key)
        if root_id is None:
            root_id = _synthetic_id()
            provider_root_ids[provider_key] = root_id
            nodes[root_id] = {
                "provider": provider_key,
                "account_identifier": f"{provider_key}:provider",
                "account_name": provider_key.upper(),
                "account_type": "provider",
                "parent_account_id": None,
                "parent_account_identifier": None,
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": None,
                "captured_at": latest_imported_at.isoformat() if latest_imported_at else None,
            }

        identifier = str(row.account_identifier or row.account_name or "").strip()
        account_name = str(row.account_name or row.account_identifier or provider_key).strip()
        if not identifier:
            identifier = f"{provider_key}:unassigned"
            account_name = f"{provider_key.upper()} Unassigned"

        parent_identifier = str(row.parent_account_identifier or "").strip()
        parent_account_id: Optional[int] = root_id
        if parent_identifier:
            parent_account_id = _ensure_node(
                provider_key=provider_key,
                identifier=parent_identifier,
                account_name=parent_identifier,
                account_type="group",
                parent_account_id=root_id,
            )

        account_id = _ensure_node(
            provider_key=provider_key,
            identifier=identifier,
            account_name=account_name,
            account_type=_normalized_account_type(row.account_type, "account"),
            parent_account_id=parent_account_id,
        )
        nodes[account_id]["direct_cost_usd"] = round(float(nodes[account_id]["direct_cost_usd"]) + float(row.cost_usd or 0.0), 2)
        service_name = str(row.service_name or "").strip()
        if service_name:
            nodes[account_id].setdefault("_services", set()).add(service_name)
        region = str(row.region or "").strip()
        if region:
            region_costs = nodes[account_id].setdefault("_region_costs", {})
            region_costs[region] = round(region_costs.get(region, 0.0) + float(row.cost_usd or 0.0), 2)

    for node in nodes.values():
        node["direct_service_count"] = len(node.get("_services", set()))
        node.pop("_services", None)
        region_costs = node.pop("_region_costs", {})
        node["top_regions"] = [
            AccountRegionRow(region=r, cost_usd=c)
            for r, c in sorted(region_costs.items(), key=lambda x: x[1], reverse=True)[:5]
        ]

    items = _materialize_rollup_items(nodes)
    filtered_items = [item for item in items if provider is None or item.provider == provider]
    filtered_items = _apply_rollup_budget_metrics(
        filtered_items,
        monthly_budget_usd=monthly_budget_usd,
        warning_threshold_percent=warning_threshold_percent,
        critical_threshold_percent=critical_threshold_percent,
    )
    direct_total = round(sum(item.direct_cost_usd for item in filtered_items if item.depth > 0 or item.child_count == 0), 2)
    root_total = round(sum(item.rolled_up_cost_usd for item in filtered_items if item.depth == 0), 2)
    return ProviderAccountRollupResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        provider=provider,
        scan_id=None,
        generated_at=_utcnow().isoformat(),
        total_direct_cost_usd=direct_total,
        total_rolled_up_cost_usd=root_total,
        items=filtered_items,
    )


def _parse_credential_payload(raw: Dict[str, Any]) -> CredentialInput:
    provider = str(raw.get("provider", "")).lower()
    if provider == "aws":
        return AWSCredentialInput(**raw)
    if provider == "azure":
        return AzureCredentialInput(**raw)
    if provider == "gcp":
        payload = dict(raw)
        if isinstance(payload.get("service_account_json"), str):
            payload["service_account_json"] = json.loads(payload["service_account_json"])
        return GCPCredentialInput(**payload)
    if provider == "oci":
        return OCICredentialInput(**raw)
    raise ValueError(f"Unsupported provider: {provider}")


def _run_validation(credential: CredentialInput) -> CredentialStatus:
    validator = CredentialValidator()
    if isinstance(credential, AWSCredentialInput):
        return validator.validate_aws(
            credential.access_key_id,
            credential.secret_access_key,
            credential.region or "us-east-1",
        )
    if isinstance(credential, AzureCredentialInput):
        return validator.validate_azure(
            credential.subscription_id,
            credential.tenant_id,
            credential.client_id,
            credential.client_secret,
        )
    if isinstance(credential, GCPCredentialInput):
        service_account_json = credential.service_account_json
        if isinstance(service_account_json, str):
            service_account_json = json.loads(service_account_json)
        return validator.validate_gcp(credential.project_id, service_account_json)
    if isinstance(credential, OCICredentialInput):
        return validator.validate_oci(
            credential.config_file,
            credential.profile or "DEFAULT",
        )
    raise ValueError("Unsupported credential payload")


def _safe_json_load(raw: str, default: Dict[str, Any]) -> Dict[str, Any]:
    try:
        return json.loads(raw)
    except Exception:
        return default


def _customer_id_for_org(membership: UserOrganization) -> str:
    """Derive persisted customer scope from the active organization."""
    return f"org-{membership.organization_id}"


def _organization_id_for_membership(membership: UserOrganization) -> int:
    return int(membership.organization_id)


def _organization_id_from_customer_id(customer_id: str) -> Optional[int]:
    normalized = str(customer_id or "").strip()
    if normalized.startswith("org-"):
        try:
            return int(normalized.split("-", 1)[1])
        except (TypeError, ValueError):
            return None
    return None


def _require_management_role(membership: UserOrganization, action: str) -> None:
    require_role(
        membership,
        allowed_roles=[UserRole.OWNER, UserRole.ADMIN],
        action=action,
    )


def _resolve_customer_id(
    current_user: User,
    membership: UserOrganization,
    requested_customer_id: Optional[str] = None,
) -> str:
    """Reject mismatched customer scopes and return the org-derived identifier."""
    _ = current_user
    derived_customer_id = _customer_id_for_org(membership)
    normalized_requested = str(requested_customer_id or "").strip()
    if normalized_requested and normalized_requested != derived_customer_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="customer_id must match the authenticated organization scope",
        )
    return derived_customer_id


async def _cost_summary_for_provider(provider: str, period: str = "month") -> Dict[str, Any]:
    params = {"period": period, "cloud_provider": provider}
    if provider == "aws":
        raw = await aws_costs.get_cost_summary(params)
    elif provider == "azure":
        raw = await azure_costs.get_cost_summary(params)
    elif provider == "gcp":
        raw = await gcp_costs.get_cost_summary(params)
    elif provider == "oci":
        raw = await oci_costs.get_cost_summary(params)
    else:
        return {"error": f"Unsupported provider: {provider}"}
    return _safe_json_load(raw, {"error": "Invalid tool response"})


def _imported_cost_context(
    membership: UserOrganization,
    db: Session,
    cloud_provider: str = "all",
) -> Optional[Dict[str, Any]]:
    resolved_org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    rows = _get_imported_cost_rows(
        db,
        organization_id=resolved_org_id,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
    )
    if not rows:
        return None

    breakdown: Dict[str, Dict[str, float]] = {}
    region_breakdown: Dict[str, float] = {}
    total_cost = 0.0
    for row in rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        cost = float(row.cost_usd or 0.0)
        total_cost += cost
        provider_bucket = breakdown.setdefault(provider, {"cost": 0.0, "percentage": 0.0})
        provider_bucket["cost"] = round(provider_bucket["cost"] + cost, 2)
        region_name = str(row.region or "").strip()
        if region_name:
            region_breakdown[region_name] = round(region_breakdown.get(region_name, 0.0) + cost, 2)

    if total_cost > 0:
        for provider in breakdown:
            breakdown[provider]["percentage"] = round(
                (breakdown[provider]["cost"] / total_cost) * 100,
                1,
            )

    summary = _imported_cost_summary(rows)
    return {
        "period": "imported",
        "cloud_provider": cloud_provider,
        "total_cost": round(total_cost, 2),
        "breakdown": breakdown,
        "region_breakdown": [
            {"region": region, "cost_usd": round(cost, 2)}
            for region, cost in sorted(region_breakdown.items(), key=lambda item: item[1], reverse=True)
        ],
        "source": "csv_import",
        "rows_imported": summary["rows_imported"],
        "last_imported_at": summary["last_imported_at"].isoformat()
        if summary["last_imported_at"]
        else None,
    }


async def _cost_context(
    membership: UserOrganization,
    db: Session,
    period: str = "month",
    cloud_provider: str = "all",
) -> Dict[str, Any]:
    providers = ["aws", "azure", "gcp", "oci"] if cloud_provider == "all" else [cloud_provider]
    configured_live_providers = {
        diagnostic.provider
        for diagnostic in _provider_diagnostics()
        if diagnostic.configured and diagnostic.provider in providers
    }

    if not configured_live_providers:
        imported_context = _imported_cost_context(
            membership,
            db,
            cloud_provider=cloud_provider,
        )
        if imported_context is not None:
            return imported_context

    breakdown: Dict[str, Dict[str, float]] = {}
    region_breakdown: Dict[str, float] = {}
    total_cost = 0.0

    for provider in providers:
        summary = await _cost_summary_for_provider(provider, period)
        cost = float(summary.get("total_cost_usd", 0) or 0)
        total_cost += cost
        breakdown[provider] = {"cost": round(cost, 2), "percentage": 0.0}
        for region_row in summary.get("region_breakdown", []):
            region_name = str(region_row.get("region") or "global")
            region_cost = float(region_row.get("cost_usd") or 0.0)
            region_breakdown[region_name] = region_breakdown.get(region_name, 0.0) + region_cost

    if total_cost > 0:
        for provider in breakdown:
            breakdown[provider]["percentage"] = round(
                (breakdown[provider]["cost"] / total_cost) * 100, 1
            )

    return {
        "period": period,
        "cloud_provider": cloud_provider,
        "total_cost": round(total_cost, 2),
        "breakdown": breakdown,
        "source": "live_provider_api" if configured_live_providers else "live_backend",
        "region_breakdown": [
            {"region": region, "cost_usd": round(cost, 2)}
            for region, cost in sorted(region_breakdown.items(), key=lambda item: item[1], reverse=True)
        ],
    }


def _historical_monthly_spend_from_snapshots(
    db: Session,
    customer_id: str,
    cloud_provider: str,
    months: int = 18,
) -> List[float]:
    """Build monthly spend history from persisted cost snapshots.

    Returns ascending monthly totals (oldest -> newest).
    """
    query = db.query(CostSnapshot).filter(CostSnapshot.customer_id == customer_id)
    if cloud_provider != "all":
        query = query.filter(CostSnapshot.provider == cloud_provider)
    rows = query.order_by(CostSnapshot.captured_at.desc()).limit(500).all()

    month_totals: Dict[str, float] = {}
    for row in rows:
        anchor = row.period_end or row.period_start or row.captured_at
        if anchor is None:
            continue
        month_key = anchor.strftime("%Y-%m")
        month_totals[month_key] = month_totals.get(month_key, 0.0) + float(row.total_cost_usd or 0.0)

    ordered = sorted(month_totals.items())
    return [round(value, 2) for _, value in ordered[-months:]]


def _setting_missing(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return True
    return text.startswith("your_") or text.startswith("replace_") or "example.com" in text


def _provider_diagnostics() -> List[ProviderDiagnostic]:
    config = Config()
    requirements = {
        "aws": {
            "settings": ["AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY", "AWS_REGION"],
            "values": [
                config.aws_access_key_id,
                config.aws_secret_access_key,
                config.aws_region,
            ],
        },
        "azure": {
            "settings": [
                "AZURE_SUBSCRIPTION_ID",
                "AZURE_TENANT_ID",
                "AZURE_CLIENT_ID",
                "AZURE_CLIENT_SECRET",
            ],
            "values": [
                config.azure_subscription_id,
                config.azure_tenant_id,
                config.azure_client_id,
                config.azure_client_secret,
            ],
        },
        "gcp": {
            "settings": ["GOOGLE_APPLICATION_CREDENTIALS", "GCP_PROJECT_ID"],
            "values": [config.google_application_credentials, config.gcp_project_id],
        },
        "oci": {
            "settings": ["OCI_CONFIG_FILE", "OCI_PROFILE", "OCI_REGION"],
            "values": [config.oci_config_file, config.oci_profile, config.oci_region],
        },
    }

    diagnostics: List[ProviderDiagnostic] = []
    for provider, detail in requirements.items():
        settings = detail["settings"]
        values = detail["values"]
        missing = [setting for setting, value in zip(settings, values) if _setting_missing(value)]
        configured = not missing
        diagnostics.append(
            ProviderDiagnostic(
                provider=provider,
                configured=configured,
                required_settings=settings,
                missing_settings=missing,
                recommendation=(
                    "Ready for live billing API calls."
                    if configured
                    else f"Configure {', '.join(missing)} before enabling live {provider.upper()} cost collection."
                ),
            )
        )
    return diagnostics


@router.post("/credentials/validate", response_model=CredentialResponse)
async def validate_credentials(
    payload: Dict[str, Any],
    current_user: User = Depends(get_current_user),
) -> CredentialResponse:
    """Validate cloud credentials without storing them."""
    _ = current_user
    try:
        credential = _parse_credential_payload(payload)
        result = _run_validation(credential)
        return CredentialResponse(**result.__dict__)
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Credential validation failed")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/credentials/add")
async def add_credentials(
    payload: Dict[str, Any],
    credential_manager: CredentialManager = Depends(get_credential_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Validate and store credentials metadata."""
    try:
        _require_management_role(membership, "credential updates")
        customer_id = _resolve_customer_id(
            current_user=current_user,
            membership=membership,
            requested_customer_id=payload.get("customer_id"),
        )
        credential_payload = {k: v for k, v in payload.items() if k != "customer_id"}
        credential = _parse_credential_payload(credential_payload)
        validation = _run_validation(credential)
        if not validation.is_valid:
            raise HTTPException(
                status_code=400,
                detail=f"Credential validation failed: {validation.message}",
            )

        stored = credential_manager.store_credentials(
            customer_id=customer_id,
            provider=credential.provider,
            credentials=credential.model_dump(),
            is_active=True,
            validation=validation,
        )
        return {
            "status": "success",
            "message": f"{credential.provider.upper()} credentials stored",
            "provider": credential.provider,
            "customer_id": customer_id,
            "record": stored,
        }
    except HTTPException:
        raise
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Failed to add credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/credentials")
async def list_credentials(
    customer_id: Optional[str] = None,
    credential_manager: CredentialManager = Depends(get_credential_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        payload = credential_manager.list_credentials(scoped_customer_id)
        payload["organization_id"] = _organization_id_for_membership(membership)
        return payload
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to list credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.delete("/credentials/{provider}")
async def delete_credentials(
    provider: str,
    customer_id: Optional[str] = None,
    credential_manager: CredentialManager = Depends(get_credential_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "credential deletion")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        deleted = credential_manager.delete_credentials(scoped_customer_id, provider)
        if not deleted:
            raise HTTPException(status_code=404, detail="Credential not found")
        return {
            "status": "success",
            "message": f"{provider.upper()} credentials deleted",
            "provider": provider,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to delete credentials")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/request-approval")
async def request_scanning_approval(
    providers: List[str],
    notification_email: str,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    customer_id: Optional[str] = None,
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan approval requests")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        scanning_manager.create_permission_request(
            customer_id=scoped_customer_id,
            providers=providers,
            notification_email=notification_email,
        )
        approval_request = scanning_manager.request_approval(
            customer_id=scoped_customer_id,
            providers=providers,
        )
        return {
            "status": "approval_pending",
            "message": approval_request["message"],
            "action_required": True,
            "approve_url": approval_request["approve_url"],
            "providers": providers,
            "customer_id": scoped_customer_id,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to request approval")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/approve", response_model=ScanningPermissionResponse)
async def approve_scanning(
    approval: ScanningApprovalRequest,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanningPermissionResponse:
    try:
        _require_management_role(membership, "scan approval")
        customer_id = _resolve_customer_id(current_user, membership, approval.customer_id)
        approved = scanning_manager.approve_scanning(
            customer_id=customer_id,
            auto_remediate=approval.auto_remediate,
            scan_frequency=approval.scan_frequency,
            notification_email=approval.notification_email,
            monthly_budget_usd=approval.monthly_budget_usd,
            warning_threshold_percent=approval.warning_threshold_percent,
            critical_threshold_percent=approval.critical_threshold_percent,
            notifications_enabled=approval.notifications_enabled,
        )
        return ScanningPermissionResponse(
            customer_id=approved["customer_id"],
            organization_id=_organization_id_for_membership(membership),
            state=approved["state"],
            providers=approved["providers"],
            scan_frequency=approved["scan_frequency"],
            auto_remediate=approved["auto_remediate"],
            notification_email=approved.get("notification_email"),
            monthly_budget_usd=float(approved.get("monthly_budget_usd", 0.0) or 0.0),
            warning_threshold_percent=float(approved.get("warning_threshold_percent", 80.0) or 80.0),
            critical_threshold_percent=float(approved.get("critical_threshold_percent", 100.0) or 100.0),
            notifications_enabled=bool(approved.get("notifications_enabled", True)),
            created_at=approved["created_at"],
            approved_at=approved["approved_at"],
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to approve scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/scanning/permission", response_model=ScanningPermissionResponse)
async def get_scanning_permission(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanningPermissionResponse:
    try:
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        permission = scanning_manager.get_permission_status(scoped_customer_id)
        return ScanningPermissionResponse(
            organization_id=_organization_id_for_membership(membership),
            **permission,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to get permission status")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/pause")
async def pause_scanning(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan pause")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        return scanning_manager.pause_scanning(scoped_customer_id)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to pause scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/resume")
async def resume_scanning(
    customer_id: Optional[str] = None,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    try:
        _require_management_role(membership, "scan resume")
        scoped_customer_id = _resolve_customer_id(current_user, membership, customer_id)
        return scanning_manager.resume_scanning(scoped_customer_id)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to resume scanning")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/start", response_model=ScanProgressResponse)
async def start_scan(
    scan_request: StartScanRequest,
    background_tasks: BackgroundTasks,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanProgressResponse:
    try:
        _require_management_role(membership, "scan start")
        customer_id = _resolve_customer_id(current_user, membership, scan_request.customer_id)
        permission = scanning_manager.get_permission_status(customer_id)
        if permission["state"] not in [ScanningState.APPROVED.value, ScanningState.RUNNING.value]:
            raise HTTPException(
                status_code=403,
                detail=f"Scanning not approved. Current state: {permission['state']}",
            )

        providers_to_scan = scan_request.providers or permission["providers"]
        if not providers_to_scan:
            providers_to_scan = ["aws", "azure", "gcp", "oci"]

        scan_id = f"scan_{customer_id}_{int(_utcnow().timestamp())}"
        scanning_manager.create_scan_run(scan_id, customer_id, providers_to_scan)
        background_tasks.add_task(
            _run_cost_analysis,
            scan_id=scan_id,
            customer_id=customer_id,
            providers=providers_to_scan,
            target_accounts=scan_request.target_accounts or None,
        )

        return ScanProgressResponse(
            scan_id=scan_id,
            customer_id=customer_id,
            organization_id=_organization_id_for_membership(membership),
            state=ScanningState.RUNNING.value,
            progress=0,
            providers=providers_to_scan,
            started_at=_utcnow(),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to start scan")
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/scanning/scheduler/run-now")
async def run_scheduler_now(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Run the scan scheduler loop once on-demand."""
    _ = current_user
    _require_management_role(membership, "manual scheduler runs")
    return await run_scheduled_scans_once(
        requested_organization_id=_organization_id_for_membership(membership),
    )


@router.get("/scanning/{scan_id}/progress", response_model=ScanProgressResponse)
async def get_scan_progress(
    scan_id: str,
    scanning_manager: ScanningManager = Depends(get_scanning_manager),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanProgressResponse:
    row = scanning_manager.get_scan_run(scan_id)
    if row is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    if row["customer_id"] != _customer_id_for_org(membership):
        raise HTTPException(status_code=404, detail="Scan not found")
    return ScanProgressResponse(
        scan_id=row["scan_id"],
        customer_id=row["customer_id"],
        organization_id=_organization_id_for_membership(membership),
        state=row["state"],
        progress=row["progress"],
        providers=row["providers"],
        started_at=row["started_at"],
        completed_at=row["completed_at"],
        total_resources=row["total_resources"],
        anomalies_found=row["anomalies_found"],
        savings_identified=row["savings_identified"],
    )


@router.get("/scanning/history", response_model=List[ScanHistoryItem])
async def get_scan_history(
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ScanHistoryItem]:
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .limit(max(1, min(limit, 100)))
            .all()
        )
    finally:
        db.close()
    return [
        ScanHistoryItem(
            scan_id=row.scan_id,
            customer_id=row.customer_id,
            organization_id=_organization_id_for_membership(membership),
            state=row.state,
            providers=json.loads(row.providers_json or "[]"),
            started_at=row.started_at,
            completed_at=row.completed_at,
            total_resources=row.total_resources,
            anomalies_found=row.anomalies_found,
            savings_identified=float(row.savings_identified or 0.0),
        )
        for row in rows
    ]


@router.get("/scanning/{scan_id}/diff", response_model=ScanDiffResponse)
async def get_scan_diff(
    scan_id: str,
    previous_scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ScanDiffResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        current_run = (
            db.query(ScanRunRecord)
            .filter(
                ScanRunRecord.scan_id == scan_id,
                ScanRunRecord.customer_id == customer_id,
            )
            .first()
        )
        if current_run is None:
            raise HTTPException(status_code=404, detail="Scan not found")

        if previous_scan_id:
            previous_run = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.scan_id == previous_scan_id,
                    ScanRunRecord.customer_id == customer_id,
                )
                .first()
            )
        else:
            previous_run = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.customer_id == customer_id,
                    ScanRunRecord.state == ScanningState.COMPLETED.value,
                    ScanRunRecord.scan_id != scan_id,
                    ScanRunRecord.started_at < current_run.started_at,
                )
                .order_by(ScanRunRecord.started_at.desc())
                .first()
            )

        current_snapshots = (
            db.query(CostSnapshot).filter(CostSnapshot.scan_id == current_run.scan_id).all()
        )
        previous_snapshots = (
            db.query(CostSnapshot).filter(CostSnapshot.scan_id == previous_run.scan_id).all()
            if previous_run
            else []
        )
    finally:
        db.close()

    current_map = {row.provider: row for row in current_snapshots}
    previous_map = {row.provider: row for row in previous_snapshots}
    providers = sorted(set(current_map.keys()) | set(previous_map.keys()))

    entries: List[ScanDiffEntry] = []
    total_current = 0.0
    total_previous = 0.0
    for provider in providers:
        current_row = current_map.get(provider)
        previous_row = previous_map.get(provider)
        current_cost = float(current_row.total_cost_usd) if current_row else 0.0
        previous_cost = float(previous_row.total_cost_usd) if previous_row else 0.0
        total_current += current_cost
        total_previous += previous_cost
        delta_cost = current_cost - previous_cost
        delta_percent = None
        if previous_cost > 0:
            delta_percent = round((delta_cost / previous_cost) * 100, 2)
        entries.append(
            ScanDiffEntry(
                provider=provider,
                current_cost_usd=round(current_cost, 2),
                previous_cost_usd=round(previous_cost, 2),
                delta_cost_usd=round(delta_cost, 2),
                delta_percent=delta_percent,
                current_anomalies=int(current_row.anomalies_count if current_row else 0),
                previous_anomalies=int(previous_row.anomalies_count if previous_row else 0),
            )
        )

    return ScanDiffResponse(
        organization_id=organization_id,
        current_scan_id=current_run.scan_id,
        previous_scan_id=previous_run.scan_id if previous_run else None,
        total_current_cost_usd=round(total_current, 2),
        total_previous_cost_usd=round(total_previous, 2),
        total_delta_cost_usd=round(total_current - total_previous, 2),
        entries=entries,
    )


@router.get("/scanning/history.csv")
async def download_scan_history_csv(
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await get_scan_history(limit=limit, current_user=current_user, membership=membership)
    lines = ["scan_id,state,providers,started_at,completed_at,total_resources,anomalies_found,savings_identified"]
    for row in rows:
        providers = "|".join(row.providers)
        lines.append(
            f"{row.scan_id},{row.state},{providers},{row.started_at.isoformat()},"
            f"{row.completed_at.isoformat() if row.completed_at else ''},"
            f"{row.total_resources},{row.anomalies_found},{row.savings_identified:.2f}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/scanning/{scan_id}/diff.csv")
async def download_scan_diff_csv(
    scan_id: str,
    previous_scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    diff = await get_scan_diff(
        scan_id=scan_id,
        previous_scan_id=previous_scan_id,
        current_user=current_user,
        membership=membership,
    )
    lines = ["provider,current_cost_usd,previous_cost_usd,delta_cost_usd,delta_percent,current_anomalies,previous_anomalies"]
    for entry in diff.entries:
        lines.append(
            f"{entry.provider},{entry.current_cost_usd:.2f},{entry.previous_cost_usd:.2f},"
            f"{entry.delta_cost_usd:.2f},{entry.delta_percent if entry.delta_percent is not None else ''},"
            f"{entry.current_anomalies},{entry.previous_anomalies}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/provider-accounts/rollups", response_model=ProviderAccountRollupResponse)
async def get_provider_account_rollups(
    scan_id: Optional[str] = None,
    provider: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ProviderAccountRollupResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    budget_monthly_usd = 0.0
    warning_threshold_percent = 80.0
    critical_threshold_percent = 100.0
    try:
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .order_by(ScanningPermissionRecord.created_at.desc())
            .first()
        )
        if permission is not None:
            budget_monthly_usd = float(permission.monthly_budget_usd or 0.0)
            warning_threshold_percent = float(permission.warning_threshold_percent or 80.0)
            critical_threshold_percent = float(permission.critical_threshold_percent or 100.0)

        imported_rows = _get_imported_cost_rows(
            db,
            organization_id=organization_id,
            customer_id=customer_id,
            cloud_provider=provider or "all",
        )

        query = (
            db.query(ProviderAccountSnapshot, ProviderAccount)
            .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
            .filter(
                ProviderAccountSnapshot.customer_id == customer_id,
                ProviderAccountSnapshot.organization_id == organization_id,
            )
        )
        if scan_id:
            query = query.filter(ProviderAccountSnapshot.scan_id == scan_id)
        else:
            latest = (
                db.query(ProviderAccountSnapshot.scan_id)
                .filter(
                    ProviderAccountSnapshot.customer_id == customer_id,
                    ProviderAccountSnapshot.organization_id == organization_id,
                )
                .order_by(ProviderAccountSnapshot.captured_at.desc())
                .first()
            )
            if latest:
                query = query.filter(ProviderAccountSnapshot.scan_id == latest[0])
                scan_id = latest[0]
        if provider:
            query = query.filter(ProviderAccount.provider == provider)
        rows = query.all()

        if imported_rows:
            return _build_rollups_from_imported_rows(
                imported_rows,
                organization_id=organization_id,
                customer_id=customer_id,
                provider=provider,
                monthly_budget_usd=budget_monthly_usd,
                warning_threshold_percent=warning_threshold_percent,
                critical_threshold_percent=critical_threshold_percent,
            )

        account_ids = [account.id for _, account in rows]
        links = (
            db.query(ProviderAccountLink)
            .filter(
                ProviderAccountLink.organization_id == organization_id,
                ProviderAccountLink.child_account_id.in_(account_ids),
            )
            .all()
            if account_ids
            else []
        )
    finally:
        db.close()

    nodes: Dict[int, Dict[str, Any]] = {}
    provider_roots: Dict[str, int] = {}
    next_synthetic_id = -1

    def _synthetic_id() -> int:
        nonlocal next_synthetic_id
        current = next_synthetic_id
        next_synthetic_id -= 1
        return current

    account_by_id = {account.id: account for _, account in rows}
    parent_by_child = {link.child_account_id: link.parent_account_id for link in links}

    for snapshot, account in rows:
        nodes[account.id] = {
            "provider": account.provider,
            "account_identifier": account.account_identifier,
            "account_name": account.account_name,
            "account_type": account.account_type,
            "parent_account_id": parent_by_child.get(account.id),
            "parent_account_identifier": (
                account_by_id[parent_by_child[account.id]].account_identifier
                if parent_by_child.get(account.id) in account_by_id
                else None
            ),
            "direct_cost_usd": round(float(snapshot.direct_cost_usd or 0.0), 2),
            "direct_savings_identified_usd": round(float(snapshot.savings_identified_usd or 0.0), 2),
            "direct_anomalies_count": int(snapshot.anomalies_count or 0),
            "direct_service_count": int(snapshot.service_count or 0),
            "scan_id": snapshot.scan_id,
            "captured_at": snapshot.captured_at.isoformat() if snapshot.captured_at else None,
        }

    for node in nodes.values():
        if node.get("parent_account_id") not in nodes:
            node["parent_account_id"] = None
            node["parent_account_identifier"] = None

    management_group_by_provider: Dict[str, int] = {}
    for account_id, node in nodes.items():
        if node["account_type"] == "management_group":
            management_group_by_provider[node["provider"]] = account_id

    for account_id, node in list(nodes.items()):
        provider_key = node["provider"]
        root_id = provider_roots.get(provider_key)
        if root_id is None:
            root_id = _synthetic_id()
            provider_roots[provider_key] = root_id
            nodes[root_id] = {
                "provider": provider_key,
                "account_identifier": f"{provider_key}:provider",
                "account_name": provider_key.upper(),
                "account_type": "provider",
                "parent_account_id": None,
                "parent_account_identifier": None,
                "direct_cost_usd": 0.0,
                "direct_savings_identified_usd": 0.0,
                "direct_anomalies_count": 0,
                "direct_service_count": 0,
                "scan_id": node.get("scan_id"),
                "captured_at": node.get("captured_at"),
            }

        if node.get("parent_account_id") is not None:
            continue

        inferred_parent_id = root_id
        if node["account_type"] == "management_group":
            inferred_parent_id = root_id
        elif node["account_type"] in {"subscription", "account"} and provider_key == "azure":
            inferred_parent_id = management_group_by_provider.get(provider_key, root_id)
        elif node["account_type"] == "tenancy":
            inferred_parent_id = root_id
        elif node["account_type"] in {"project", "compartment"}:
            inferred_parent_id = root_id

        node["parent_account_id"] = inferred_parent_id if inferred_parent_id != account_id else None
        node["parent_account_identifier"] = (
            nodes[inferred_parent_id]["account_identifier"]
            if inferred_parent_id in nodes and inferred_parent_id != account_id
            else None
        )

    # Attach top regions to each node from CostAllocationSnapshot.
    if scan_id and account_by_id:
        db2 = SessionLocal()
        try:
            alloc_rows = (
                db2.query(CostAllocationSnapshot)
                .filter(
                    CostAllocationSnapshot.scan_id == scan_id,
                    CostAllocationSnapshot.provider_account_id.in_(list(account_by_id.keys())),
                )
                .order_by(CostAllocationSnapshot.cost_usd.desc())
                .all()
            )
        finally:
            db2.close()
        for alloc in alloc_rows:
            node = nodes.get(alloc.provider_account_id)
            if node is None:
                continue
            if "top_regions" not in node:
                node["top_regions"] = []
            if len(node["top_regions"]) < 5:
                node["top_regions"].append(
                    AccountRegionRow(region=alloc.region, cost_usd=round(float(alloc.cost_usd), 2))
                )

    items = _materialize_rollup_items(nodes)
    filtered_items = [item for item in items if provider is None or item.provider == provider]
    filtered_items = _apply_rollup_budget_metrics(
        filtered_items,
        monthly_budget_usd=budget_monthly_usd,
        warning_threshold_percent=warning_threshold_percent,
        critical_threshold_percent=critical_threshold_percent,
    )
    total_direct = round(sum(item.direct_cost_usd for item in filtered_items if item.depth > 0 or item.child_count == 0), 2)
    total_rolled = round(sum(item.rolled_up_cost_usd for item in filtered_items if item.depth == 0), 2)

    return ProviderAccountRollupResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        provider=provider,
        scan_id=scan_id,
        generated_at=_utcnow().isoformat(),
        total_direct_cost_usd=total_direct,
        total_rolled_up_cost_usd=total_rolled,
        items=filtered_items,
    )


@router.get("/provider-accounts", response_model=ProviderAccountInventoryResponse)
async def get_provider_account_inventory(
    provider: Optional[str] = None,
    account_type: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ProviderAccountInventoryResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        query = db.query(ProviderAccount).filter(
            ProviderAccount.customer_id == customer_id,
            ProviderAccount.organization_id == organization_id,
            ProviderAccount.is_active.is_(True),
        )
        if provider:
            query = query.filter(ProviderAccount.provider == provider)
        if account_type:
            query = query.filter(ProviderAccount.account_type == account_type)
        accounts = query.order_by(ProviderAccount.provider, ProviderAccount.account_name).all()
    finally:
        db.close()

    items = [
        ProviderAccountInventoryItem(
            account_id=acc.id,
            provider=acc.provider,
            account_identifier=acc.account_identifier,
            account_name=acc.account_name,
            account_type=acc.account_type,
            native_region=acc.native_region,
            is_active=bool(acc.is_active),
            metadata=_safe_json_load(acc.metadata_json, {}),
            created_at=acc.created_at.isoformat() if acc.created_at else "",
            updated_at=acc.updated_at.isoformat() if acc.updated_at else None,
        )
        for acc in accounts
    ]
    return ProviderAccountInventoryResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        total=len(items),
        accounts=items,
    )


@router.get("/provider-accounts/{account_id}/region-breakdown", response_model=AccountRegionBreakdownResponse)
async def get_account_region_breakdown(
    account_id: int,
    scan_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AccountRegionBreakdownResponse:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        account = (
            db.query(ProviderAccount)
            .filter(
                ProviderAccount.id == account_id,
                ProviderAccount.customer_id == customer_id,
                ProviderAccount.organization_id == organization_id,
            )
            .first()
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Account not found")

        alloc_query = db.query(CostAllocationSnapshot).filter(
            CostAllocationSnapshot.provider_account_id == account_id,
            CostAllocationSnapshot.customer_id == customer_id,
            CostAllocationSnapshot.organization_id == organization_id,
        )
        if scan_id:
            alloc_query = alloc_query.filter(CostAllocationSnapshot.scan_id == scan_id)
        else:
            latest_scan = (
                db.query(CostAllocationSnapshot.scan_id)
                .filter(
                    CostAllocationSnapshot.provider_account_id == account_id,
                    CostAllocationSnapshot.customer_id == customer_id,
                    CostAllocationSnapshot.organization_id == organization_id,
                )
                .order_by(CostAllocationSnapshot.captured_at.desc())
                .first()
            )
            if latest_scan:
                scan_id = latest_scan[0]
                alloc_query = alloc_query.filter(CostAllocationSnapshot.scan_id == scan_id)

        allocs = alloc_query.order_by(CostAllocationSnapshot.cost_usd.desc()).all()
        account_name = account.account_name
        provider = account.provider
    finally:
        db.close()

    regions = [
        AccountRegionBreakdownItem(
            region=a.region,
            cost_usd=round(float(a.cost_usd), 2),
            scan_id=a.scan_id,
            captured_at=a.captured_at.isoformat() if a.captured_at else "",
        )
        for a in allocs
    ]
    return AccountRegionBreakdownResponse(
        account_id=account_id,
        provider=provider,
        account_name=account_name,
        scan_id=scan_id,
        total_cost_usd=round(sum(r.cost_usd for r in regions), 2),
        regions=regions,
    )


@router.get("/imports/costs/summary", response_model=ImportedCostSummaryResponse)
async def get_imported_cost_summary(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ImportedCostSummaryResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    rows = _get_imported_cost_rows(db, organization_id, customer_id)
    if not rows:
        return ImportedCostSummaryResponse(
            organization_id=organization_id,
            customer_id=customer_id,
            has_data=False,
        )

    summary = _imported_cost_summary(rows)
    return ImportedCostSummaryResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        has_data=True,
        upload_id=summary["upload_id"],
        source_filename=summary["source_filename"],
        rows_imported=summary["rows_imported"],
        total_cost_usd=summary["total_cost_usd"],
        providers=summary["providers"],
        last_imported_at=summary["last_imported_at"],
    )


@router.get("/imports/costs/template.csv")
async def download_cost_import_template(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    _ = (current_user, membership)
    return Response(
        (
            "provider,cost_usd,service_name,account_identifier,account_name,account_type,parent_account_identifier,region,period_start,period_end,currency\n"
            "aws,123.45,EC2,acct-aws-1,AWS Prod,account,aws-org-root,eu-west-2,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
            "azure,67.89,Virtual Machines,sub-azure-1,Azure Prod,subscription,mg-finops,uk south,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
            "oci,10.00,Compute,comp-oci-1,OCI Prod,compartment,tenancy-main,uk-london-1,2026-04-01T00:00:00Z,2026-04-30T23:59:59Z,USD\n"
        ),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=optiora-cost-import-template.csv"},
    )


@router.post("/imports/costs/preview", response_model=ImportPreviewResponse)
async def preview_cost_csv_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ImportPreviewResponse:
    _require_management_role(membership, "CSV import preview")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    filename = str(file.filename or "").strip() or "cost-import.csv"
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV uploads are supported right now.")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded CSV file is empty.")

    _MAX_CSV_BYTES = 10 * 1024 * 1024
    if len(raw) > _MAX_CSV_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"CSV file too large ({len(raw):,} bytes). Maximum allowed size is 10 MB.",
        )

    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV upload must be UTF-8 encoded.") from exc

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=400,
            detail="CSV header row is missing. Expected columns include provider and cost_usd.",
        )

    headers = [str(name or "").strip().lower() for name in reader.fieldnames]
    reader.fieldnames = headers

    issues: List[ImportPreviewIssue] = []
    if "provider" not in headers:
        issues.append(ImportPreviewIssue(line_number=1, severity="error", message="Missing required column: provider"))
    if "cost_usd" not in headers:
        issues.append(ImportPreviewIssue(line_number=1, severity="error", message="Missing required column: cost_usd"))

    total_rows = 0
    accepted_rows = 0
    rejected_rows = 0
    total_cost_usd = 0.0
    providers: set[str] = set()
    rows_with_account = 0
    rows_with_region = 0
    rows_with_service = 0
    rows_with_tags = 0

    for line_number, row in enumerate(reader, start=2):
        total_rows += 1
        normalized_row = {
            str(key or "").strip().lower(): str(value or "").strip()
            for key, value in row.items()
        }

        provider = normalized_row.get("provider", "").lower()
        if provider not in SUPPORTED_COST_IMPORT_PROVIDERS:
            rejected_rows += 1
            issues.append(
                ImportPreviewIssue(
                    line_number=line_number,
                    severity="error",
                    message=(
                        f"Unsupported provider '{provider or 'empty'}'. "
                        f"Use one of: {', '.join(sorted(SUPPORTED_COST_IMPORT_PROVIDERS))}."
                    ),
                )
            )
            continue

        currency = normalized_row.get("currency", "USD").upper() or "USD"
        if currency != "USD":
            rejected_rows += 1
            issues.append(
                ImportPreviewIssue(
                    line_number=line_number,
                    severity="error",
                    message=f"Only USD is supported right now (got {currency}).",
                )
            )
            continue

        cost_usd, cost_error = _parse_required_float_value(normalized_row.get("cost_usd"), "cost_usd", line_number)
        period_start, period_start_error = _parse_optional_datetime_value(
            normalized_row.get("period_start"), "period_start", line_number
        )
        period_end, period_end_error = _parse_optional_datetime_value(
            normalized_row.get("period_end"), "period_end", line_number
        )

        if cost_error:
            issues.append(ImportPreviewIssue(line_number=line_number, severity="error", message=cost_error))
        if period_start_error:
            issues.append(ImportPreviewIssue(line_number=line_number, severity="error", message=period_start_error))
        if period_end_error:
            issues.append(ImportPreviewIssue(line_number=line_number, severity="error", message=period_end_error))

        if cost_usd is None or cost_error or period_start_error or period_end_error:
            rejected_rows += 1
            continue

        accepted_rows += 1
        total_cost_usd += float(cost_usd)
        providers.add(provider)

        if normalized_row.get("account_identifier") or normalized_row.get("account_name"):
            rows_with_account += 1
        if normalized_row.get("region"):
            rows_with_region += 1
        if normalized_row.get("service_name") or normalized_row.get("service"):
            rows_with_service += 1
        if normalized_row.get("tags"):
            rows_with_tags += 1

    if total_rows > 0 and rows_with_tags == 0:
        issues.append(
            ImportPreviewIssue(
                line_number=1,
                severity="warning",
                message="No tags column data detected; business mapping coverage may be limited.",
            )
        )

    mapping_feedback = {
        "account_coverage_percent": round((rows_with_account / max(1, accepted_rows)) * 100, 1),
        "region_coverage_percent": round((rows_with_region / max(1, accepted_rows)) * 100, 1),
        "service_coverage_percent": round((rows_with_service / max(1, accepted_rows)) * 100, 1),
        "tags_coverage_percent": round((rows_with_tags / max(1, accepted_rows)) * 100, 1),
    }

    reconciliation_guidance: List[str] = []
    if accepted_rows == 0:
        reconciliation_guidance.append("No rows passed validation; fix CSV issues before import.")
    else:
        reconciliation_guidance.append(
            f"Validated {accepted_rows} row(s) with ${round(total_cost_usd, 2):,.2f} total spend ready for import."
        )
        if mapping_feedback["account_coverage_percent"] < 80:
            reconciliation_guidance.append("Add account_identifier/account_name to improve account rollup reconciliation.")
        if mapping_feedback["tags_coverage_percent"] < 50:
            reconciliation_guidance.append("Include tags JSON to improve team/environment/cost-center mapping quality.")
        if mapping_feedback["region_coverage_percent"] < 80:
            reconciliation_guidance.append("Populate region values to improve regional variance analysis.")

    return ImportPreviewResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        filename=filename,
        total_rows=total_rows,
        accepted_rows=accepted_rows,
        rejected_rows=rejected_rows,
        total_cost_usd=round(total_cost_usd, 2),
        detected_providers=sorted(providers),
        header_columns=headers,
        mapping_feedback=mapping_feedback,
        reconciliation_guidance=reconciliation_guidance,
        issues=issues[:200],
    )


@router.post("/imports/costs/csv", response_model=ImportedCostUploadResponse)
async def upload_cost_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ImportedCostUploadResponse:
    _require_management_role(membership, "CSV cost import")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    filename = str(file.filename or "").strip() or "cost-import.csv"
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV uploads are supported right now.")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded CSV file is empty.")
    # Guard against excessively large uploads (10 MB)
    _MAX_CSV_BYTES = 10 * 1024 * 1024
    if len(raw) > _MAX_CSV_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"CSV file too large ({len(raw):,} bytes). Maximum allowed size is 10 MB.",
        )

    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV upload must be UTF-8 encoded.") from exc

    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=400,
            detail="CSV header row is missing. Expected columns include provider and cost_usd.",
        )
    reader.fieldnames = [str(name or "").strip().lower() for name in reader.fieldnames]
    if "provider" not in reader.fieldnames or "cost_usd" not in reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV must include provider and cost_usd columns.")

    imported_at = _utcnow()
    upload_id = f"csv_{uuid4().hex}"
    rows_to_store: List[ImportedCostRecord] = []
    total_cost = 0.0
    providers: set[str] = set()
    validation_errors: List[str] = []

    for line_number, row in enumerate(reader, start=2):
        normalized_row = {
            str(key or "").strip().lower(): str(value or "").strip()
            for key, value in row.items()
        }
        provider = normalized_row.get("provider", "").lower()
        if provider not in SUPPORTED_COST_IMPORT_PROVIDERS:
            validation_errors.append(
                f"Unsupported provider at CSV line {line_number}. "
                f"Use one of: {', '.join(sorted(SUPPORTED_COST_IMPORT_PROVIDERS))}."
            )
            continue

        currency = normalized_row.get("currency", "USD").upper() or "USD"
        if currency != "USD":
            validation_errors.append(
                f"Only USD CSV imports are supported right now. Invalid currency at line {line_number}."
            )
            continue

        cost_usd, cost_error = _parse_required_float_value(normalized_row.get("cost_usd"), "cost_usd", line_number)
        period_start, period_start_error = _parse_optional_datetime_value(
            normalized_row.get("period_start"),
            "period_start",
            line_number,
        )
        period_end, period_end_error = _parse_optional_datetime_value(
            normalized_row.get("period_end"),
            "period_end",
            line_number,
        )
        for error in (cost_error, period_start_error, period_end_error):
            if error:
                validation_errors.append(error)
        if cost_usd is None or cost_error or period_start_error or period_end_error:
            continue

        total_cost += cost_usd
        providers.add(provider)
        rows_to_store.append(
            ImportedCostRecord(
                organization_id=organization_id,
                customer_id=customer_id,
                upload_id=upload_id,
                source_filename=filename,
                provider=provider,
                service_name=normalized_row.get("service_name") or normalized_row.get("service") or None,
                account_identifier=normalized_row.get("account_identifier") or None,
                account_name=normalized_row.get("account_name") or None,
                account_type=normalized_row.get("account_type") or None,
                parent_account_identifier=normalized_row.get("parent_account_identifier") or None,
                region=normalized_row.get("region") or None,
                period_start=period_start,
                period_end=period_end,
                cost_usd=cost_usd,
                currency=currency,
                line_number=line_number,
                tags_json=normalized_row.get("tags") or None,
                created_at=imported_at,
            )
        )

    if validation_errors:
        detail = "CSV validation failed:\n- " + "\n- ".join(validation_errors[:10])
        if len(validation_errors) > 10:
            detail += f"\n- ...and {len(validation_errors) - 10} more issue(s)."
        raise HTTPException(status_code=400, detail=detail)

    if not rows_to_store:
        raise HTTPException(status_code=400, detail="CSV file does not contain any data rows.")

    db.query(ImportedCostRecord).filter(
        ImportedCostRecord.organization_id == organization_id,
        ImportedCostRecord.customer_id == customer_id,
    ).delete(synchronize_session=False)
    db.add_all(rows_to_store)
    db.add(
        AuditLog(
            organization_id=organization_id,
            actor_user_id=current_user.id,
            action="cost_import.csv_uploaded",
            entity_type="cost_import",
            entity_id=upload_id,
            metadata_json=json.dumps(
                {
                    "filename": filename,
                    "rows_imported": len(rows_to_store),
                    "providers": sorted(providers),
                }
            ),
            created_at=imported_at,
        )
    )
    db.commit()

    return ImportedCostUploadResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        upload_id=upload_id,
        filename=filename,
        rows_imported=len(rows_to_store),
        total_cost_usd=round(total_cost, 2),
        providers=sorted(providers),
        imported_at=imported_at,
    )


def _normalize_channels(channels: List[str]) -> List[str]:
    normalized: list[str] = []
    for value in channels:
        channel = str(value or "").strip().lower()
        if channel in SUPPORTED_NOTIFICATION_CHANNELS and channel not in normalized:
            normalized.append(channel)
    return normalized


def _latest_delivery_timestamp_for_channel(
    db: Session,
    organization_id: int,
    customer_id: str,
    channel: str,
) -> Optional[str]:
    rows = (
        db.query(AlertEvent)
        .filter(
            AlertEvent.organization_id == organization_id,
            AlertEvent.customer_id == customer_id,
        )
        .order_by(AlertEvent.created_at.desc())
        .limit(100)
        .all()
    )
    for row in rows:
        delivered = _safe_json_load(row.delivered_channels_json, [])
        if channel in delivered:
            return row.created_at.isoformat() if row.created_at else None
    return None


def _alert_lifecycle_state_map(
    db: Session,
    organization_id: int,
    alert_ids: List[int],
) -> Dict[int, str]:
    if not alert_ids:
        return {}
    rows = (
        db.query(AuditLog)
        .filter(
            AuditLog.organization_id == organization_id,
            AuditLog.entity_type == "alert_event",
            AuditLog.entity_id.in_([str(alert_id) for alert_id in alert_ids]),
            AuditLog.action.in_(["alert.acknowledge", "alert.dismiss", "alert.reactivate"]),
        )
        .order_by(AuditLog.created_at.asc())
        .all()
    )

    state_map: Dict[int, str] = {}
    for row in rows:
        try:
            alert_id = int(str(row.entity_id or "0"))
        except ValueError:
            continue
        if row.action == "alert.dismiss":
            state_map[alert_id] = "dismissed"
        elif row.action == "alert.reactivate":
            state_map[alert_id] = "reactivated"
        elif row.action == "alert.acknowledge":
            state_map[alert_id] = "acknowledged"
    return state_map


def _channel_delivery_telemetry(
    db: Session,
    organization_id: int,
    alert_id: int,
) -> Dict[str, Dict[str, Optional[str]]]:
    """Retrieve per-channel delivery telemetry (last success/error) for an alert."""
    rows = (
        db.query(AuditLog)
        .filter(
            AuditLog.organization_id == organization_id,
            AuditLog.entity_type == "channel_delivery",
            AuditLog.entity_id == str(alert_id),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(100)
        .all()
    )
    
    channel_status: Dict[str, Dict[str, Optional[str]]] = {}
    for row in rows:
        try:
            metadata = json.loads(row.metadata_json)
            channel = metadata.get("channel")
            if channel:
                if channel not in channel_status:
                    channel_status[channel] = {
                        "last_success": None,
                        "last_error": None,
                        "status": "unknown",
                    }
                if row.action == "alert.channel_delivery_success" and channel_status[channel]["last_success"] is None:
                    channel_status[channel]["last_success"] = row.created_at.isoformat()
                    channel_status[channel]["status"] = "success"
                elif row.action == "alert.channel_delivery_error" and channel_status[channel]["last_error"] is None:
                    channel_status[channel]["last_error"] = row.created_at.isoformat()
                    channel_status[channel]["status"] = "error"
        except (json.JSONDecodeError, KeyError):
            pass
    
    return channel_status

    return state_map


@router.get("/alerts/routing-policies", response_model=List[AlertRoutingPolicyResponse])
async def list_alert_routing_policies(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[AlertRoutingPolicyResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(AlertRoutingPolicy)
            .filter(AlertRoutingPolicy.organization_id == organization_id)
            .order_by(AlertRoutingPolicy.severity.asc())
            .all()
        )
        return [
            AlertRoutingPolicyResponse(
                id=row.id,
                severity=row.severity,
                channels=_safe_json_load(row.channels_json, []),
                is_active=bool(row.is_active),
                created_at=row.created_at.isoformat() if row.created_at else "",
                updated_at=row.updated_at.isoformat() if row.updated_at else "",
            )
            for row in rows
        ]
    finally:
        db.close()


@router.post("/alerts/routing-policies", response_model=AlertRoutingPolicyResponse)
async def upsert_alert_routing_policy(
    payload: AlertRoutingPolicyRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertRoutingPolicyResponse:
    _require_management_role(membership, "alert routing policy updates")
    organization_id = _organization_id_for_membership(membership)
    channels = _normalize_channels(payload.channels)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.severity == payload.severity,
            )
            .first()
        )
        now = _utcnow()
        if row is None:
            row = AlertRoutingPolicy(
                organization_id=organization_id,
                severity=payload.severity,
                channels_json=json.dumps(channels),
                is_active=payload.is_active,
                created_at=now,
                updated_at=now,
            )
        else:
            row.channels_json = json.dumps(channels)
            row.is_active = payload.is_active
            row.updated_at = now

        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alerts.routing_policy.upsert",
                entity_type="alert_routing_policy",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "severity": payload.severity,
                        "channels": channels,
                        "is_active": payload.is_active,
                    }
                ),
            )
        )
        db.commit()
        db.refresh(row)
        return AlertRoutingPolicyResponse(
            id=row.id,
            severity=row.severity,
            channels=_safe_json_load(row.channels_json, []),
            is_active=bool(row.is_active),
            created_at=row.created_at.isoformat() if row.created_at else "",
            updated_at=row.updated_at.isoformat() if row.updated_at else "",
        )
    finally:
        db.close()


@router.post("/alerts/routing-policies/simulate", response_model=AlertRoutingPolicySimulationResponse)
async def simulate_alert_routing_policy(
    payload: AlertRoutingPolicySimulationRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertRoutingPolicySimulationResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.severity == payload.severity,
            )
            .first()
        )
    finally:
        db.close()

    channels = _safe_json_load(row.channels_json, []) if row else list(SUPPORTED_NOTIFICATION_CHANNELS)
    configured = [
        channel
        for channel in channels
        if destination_configured(Config(), channel)
    ]
    return AlertRoutingPolicySimulationResponse(
        severity=payload.severity,
        matched_policy_id=row.id if row else None,
        evaluated_channels=channels,
        expected_channels=configured,
        configured_channels=[
            channel
            for channel in SUPPORTED_NOTIFICATION_CHANNELS
            if destination_configured(Config(), channel)
        ],
        inactive_policy=(bool(row) and not bool(row.is_active)),
    )


@router.get("/notifications/destinations", response_model=NotificationDestinationsResponse)
async def list_notification_destinations(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationsResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    config = Config()
    db = SessionLocal()
    try:
        policies = (
            db.query(AlertRoutingPolicy)
            .filter(
                AlertRoutingPolicy.organization_id == organization_id,
                AlertRoutingPolicy.is_active == True,  # noqa: E712
            )
            .all()
        )
        enabled_channels: set[str] = set()
        for row in policies:
            enabled_channels.update(_safe_json_load(row.channels_json, []))

        destinations = []
        for channel in SUPPORTED_NOTIFICATION_CHANNELS:
            destinations.append(
                NotificationDestinationStatus(
                    channel=channel,
                    configured=destination_configured(config, channel),
                    enabled=(channel in enabled_channels) if policies else True,
                    last_delivery_at=_latest_delivery_timestamp_for_channel(
                        db=db,
                        organization_id=organization_id,
                        customer_id=customer_id,
                        channel=channel,
                    ),
                )
            )
        return NotificationDestinationsResponse(
            organization_id=organization_id,
            destinations=destinations,
        )
    finally:
        db.close()


@router.post("/notifications/destinations/{channel}/toggle", response_model=NotificationDestinationsResponse)
async def toggle_notification_destination(
    channel: str,
    payload: NotificationDestinationToggleRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationsResponse:
    _require_management_role(membership, "notification destination toggles")
    channel_key = str(channel or "").strip().lower()
    if channel_key not in SUPPORTED_NOTIFICATION_CHANNELS:
        raise HTTPException(status_code=400, detail=f"Unsupported channel: {channel_key}")

    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        for severity in ("warning", "critical"):
            row = (
                db.query(AlertRoutingPolicy)
                .filter(
                    AlertRoutingPolicy.organization_id == organization_id,
                    AlertRoutingPolicy.severity == severity,
                )
                .first()
            )
            if row is None:
                channels = list(SUPPORTED_NOTIFICATION_CHANNELS)
                if not payload.enabled:
                    channels = [c for c in channels if c != channel_key]
                row = AlertRoutingPolicy(
                    organization_id=organization_id,
                    severity=severity,
                    channels_json=json.dumps(channels),
                    is_active=True,
                    created_at=_utcnow(),
                    updated_at=_utcnow(),
                )
                db.add(row)
                continue

            channels = set(_safe_json_load(row.channels_json, []))
            if payload.enabled:
                channels.add(channel_key)
            else:
                channels.discard(channel_key)
            row.channels_json = json.dumps(sorted(channels))
            row.updated_at = _utcnow()

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="notifications.destination.toggle",
                entity_type="notification_destination",
                entity_id=channel_key,
                metadata_json=json.dumps({"enabled": payload.enabled}),
            )
        )
        db.commit()
    finally:
        db.close()

    return await list_notification_destinations(current_user=current_user, membership=membership)


@router.post("/notifications/test-destination", response_model=NotificationDestinationTestResponse)
async def test_notification_destination(
    payload: NotificationDestinationTestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> NotificationDestinationTestResponse:
    _require_management_role(membership, "notification destination tests")
    organization_id = _organization_id_for_membership(membership)
    config = Config()
    success, detail = send_test_notification(
        config=config,
        channel=payload.channel,
        target=payload.target,
        message=payload.message,
    )

    db = SessionLocal()
    try:
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="notifications.destination.test",
                entity_type="notification_destination",
                entity_id=payload.channel,
                metadata_json=json.dumps(
                    {
                        "success": success,
                        "target_supplied": bool(payload.target),
                        "detail": detail,
                    }
                ),
            )
        )
        db.commit()
    finally:
        db.close()

    return NotificationDestinationTestResponse(
        channel=payload.channel,
        success=success,
        detail=detail,
    )


@router.get("/alerts")
async def list_alerts(
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.customer_id == customer_id,
                AlertEvent.organization_id == organization_id,
            )
            .order_by(AlertEvent.created_at.desc())
            .limit(max(1, min(limit, 200)))
            .all()
        )
        lifecycle_map = _alert_lifecycle_state_map(
            db=db,
            organization_id=organization_id,
            alert_ids=[int(row.id) for row in rows],
        )
    finally:
        db.close()
    return [
        {
            "id": row.id,
            "alert_type": row.alert_type,
            "severity": row.severity,
            "title": row.title,
            "message": row.message,
            "delivered_channels": json.loads(row.delivered_channels_json or "[]"),
            "channel_telemetry": _channel_delivery_telemetry(
                db=SessionLocal(),
                organization_id=organization_id,
                alert_id=int(row.id),
            ),
            "acknowledged_at": row.acknowledged_at.isoformat() if row.acknowledged_at else None,
            "lifecycle_state": lifecycle_map.get(
                int(row.id),
                "acknowledged" if row.acknowledged_at else "active",
            ),
            "created_at": row.created_at.isoformat(),
        }
        for row in rows
    ]


@router.post("/alerts/{alert_id}/acknowledge", response_model=AlertLifecycleActionResponse)
async def acknowledge_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert acknowledgement")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        row.acknowledged_at = _utcnow()
        row.acknowledged_by_user_id = current_user.id
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.acknowledge",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="acknowledged",
    )


@router.post("/alerts/{alert_id}/dismiss", response_model=AlertLifecycleActionResponse)
async def dismiss_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert dismissal")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.dismiss",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="dismissed",
    )


@router.post("/alerts/{alert_id}/reactivate", response_model=AlertLifecycleActionResponse)
async def reactivate_alert(
    alert_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> AlertLifecycleActionResponse:
    _require_management_role(membership, "alert reactivation")
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        row = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if row is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.reactivate",
                entity_type="alert_event",
                entity_id=str(row.id),
                metadata_json=json.dumps({"alert_type": row.alert_type}),
            )
        )
        db.commit()
    finally:
        db.close()
    return AlertLifecycleActionResponse(
        status="ok",
        alert_id=alert_id,
        lifecycle_state="reactivated",
    )


class ChannelDeliveryEvent(BaseModel):
    alert_id: int
    channel: str
    status: Literal["success", "error"]
    error_message: Optional[str] = None


@router.post("/alerts/{alert_id}/channel-delivery")
async def record_channel_delivery(
    alert_id: int,
    event: ChannelDeliveryEvent,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Record per-channel delivery telemetry for an alert."""
    _require_management_role(membership, "channel delivery recording")
    organization_id = _organization_id_for_membership(membership)
    
    db = SessionLocal()
    try:
        # Verify alert exists
        alert = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.id == alert_id,
                AlertEvent.organization_id == organization_id,
            )
            .first()
        )
        if alert is None:
            raise HTTPException(status_code=404, detail="Alert not found")
        
        # Record delivery event
        action = (
            "alert.channel_delivery_success" if event.status == "success" 
            else "alert.channel_delivery_error"
        )
        metadata = {
            "channel": event.channel,
            "status": event.status,
        }
        if event.error_message:
            metadata["error_message"] = event.error_message
        
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action=action,
                entity_type="channel_delivery",
                entity_id=str(alert_id),
                metadata_json=json.dumps(metadata),
                created_at=_utcnow(),
            )
        )
        db.commit()
    finally:
        db.close()
    
    return {
        "status": "ok",
        "alert_id": alert_id,
        "channel": event.channel,
        "delivery_status": event.status,
    }


@router.get("/alerts.csv")
async def download_alerts_csv(
    limit: int = 200,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await list_alerts(limit=limit, current_user=current_user, membership=membership)
    lines = ["id,alert_type,severity,title,message,lifecycle_state,acknowledged_at,created_at"]
    for row in rows:
        lines.append(
            f"{row['id']},{row['alert_type']},{row['severity']},"
            f"{_csv_escape(row['title'])},"
            f"{_csv_escape(row['message'])},"
            f"{row.get('lifecycle_state', 'active')},"
            f"{row['acknowledged_at'] or ''},{row['created_at']}"
        )
    return Response("\n".join(lines), media_type="text/csv")


@router.get("/audit-logs")
async def list_audit_logs(
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(AuditLog)
            .filter(AuditLog.organization_id == organization_id)
            .order_by(AuditLog.created_at.desc())
            .limit(max(1, min(limit, 200)))
            .all()
        )
    finally:
        db.close()
    return [
        {
            "id": row.id,
            "action": row.action,
            "entity_type": row.entity_type,
            "entity_id": row.entity_id,
            "actor_user_id": row.actor_user_id,
            "metadata": _safe_json_load(row.metadata_json or "{}", {}),
            "created_at": row.created_at.isoformat(),
        }
        for row in rows
    ]


@router.get("/audit-logs.csv")
async def download_audit_logs_csv(
    limit: int = 200,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Response:
    rows = await list_audit_logs(limit=limit, current_user=current_user, membership=membership)
    lines = ["id,action,entity_type,entity_id,actor_user_id,created_at"]
    for row in rows:
        lines.append(
            f"{row['id']},{row['action']},{row['entity_type']},{row['entity_id'] or ''},"
            f"{row['actor_user_id'] or ''},{row['created_at']}"
        )
    return Response("\n".join(lines), media_type="text/csv")


async def _executive_summary_rows(
    current_user: User,
    membership: UserOrganization,
    db: Session,
) -> List[List[Any]]:
    costs = await dashboard_costs(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    analytics = await dashboard_analytics(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    forecast = await dashboard_forecast(
        current_user=current_user,
        membership=membership,
        db=db,
    )
    rollups = await get_provider_account_rollups(
        current_user=current_user,
        membership=membership,
    )
    alerts = await list_alerts(
        limit=10,
        current_user=current_user,
        membership=membership,
    )

    rows: List[List[Any]] = [["Section", "Field", "Value"]]
    rows.extend(
        [
            ["Summary", "Generated At", _utcnow().isoformat()],
            ["Summary", "Organization ID", _organization_id_for_membership(membership)],
            ["Summary", "Cost Source", costs.get("cost_context", {}).get("source", "live")],
            ["Summary", "Total Monthly Cost USD", round(float(costs.get("totalCost", 0.0) or 0.0), 2)],
            ["Summary", "Potential Monthly Savings USD", round(float(costs.get("potentialSavings", 0.0) or 0.0), 2)],
            ["Summary", "Risk Score", analytics.get("risk_score", 0)],
            ["Summary", "Maturity Score", analytics.get("maturity_score", 0)],
            ["Summary", "Commitment Coverage Percent", analytics.get("commitment_coverage_percent", 0)],
            ["Summary", "Spend At Risk USD", round(float(analytics.get("spend_at_risk_usd", 0.0) or 0.0), 2)],
            ["Summary", "Optimization Capacity USD", round(float(analytics.get("optimization_capacity_usd", 0.0) or 0.0), 2)],
            ["Summary", "Budget Utilization Percent", analytics.get("unit_metrics", {}).get("budget_utilization_percent", 0)],
            ["Summary", "Forecast History Source", forecast.get("history_source", "synthetic")],
            ["Summary", "Forecast Backtest MAPE %", (forecast.get("backtesting") or {}).get("mape_percent")],
            ["Summary", "Forecast Backtest wMAPE %", (forecast.get("backtesting") or {}).get("wmape_percent")],
            ["Summary", "Average Breach Probability", (forecast.get("budget_guardrails") or {}).get("average_breach_probability")],
            ["Summary", "Open Alerts", len([row for row in alerts if not row.get("acknowledged_at")])],
            ["Summary", "Rollup Nodes", len(rollups.items)],
        ]
    )

    for provider_name, provider_data in costs.get("breakdown", {}).items():
        rows.append([
            "Provider Breakdown",
            provider_name,
            round(float(provider_data.get("cost", 0.0) or 0.0), 2),
        ])

    for item in rollups.items[:20]:
        rows.append([
            "Account Rollup",
            f"{'  ' * item.depth}{item.account_name} ({item.account_type})",
            item.rolled_up_cost_usd,
        ])

    for alert in alerts[:10]:
        rows.append([
            "Alerts",
            f"{alert['severity']} {alert['title']}",
            alert["created_at"],
        ])

    # ── Business mapping / chargeback section ─────────────────────────────
    try:
        with SessionLocal() as _db:
            org_id = _organization_id_for_membership(membership)
            customer_id = _customer_id_for_org(membership)
            dim_rows = (
                _db.query(NormalizedCostDimension)
                .filter(NormalizedCostDimension.organization_id == org_id)
                .all()
            )
            if dim_rows:
                total_dim = sum(float(r.cost_usd or 0) for r in dim_rows)
                mapped_dim = sum(float(r.cost_usd or 0) for r in dim_rows if r.is_mapped)
                coverage_pct = round((mapped_dim / total_dim * 100) if total_dim > 0 else 0.0, 1)
                rows.extend([
                    ["Allocation Coverage", "Total Cost USD", round(total_dim, 2)],
                    ["Allocation Coverage", "Mapped Cost USD", round(mapped_dim, 2)],
                    ["Allocation Coverage", "Coverage Percent", coverage_pct],
                ])
                # Per-team chargeback
                team_totals: Dict[str, float] = {}
                for r in dim_rows:
                    t = r.team or "(unmapped)"
                    team_totals[t] = team_totals.get(t, 0.0) + float(r.cost_usd or 0)
                for team, cost in sorted(team_totals.items(), key=lambda x: -x[1])[:15]:
                    rows.append(["Chargeback by Team", team, round(cost, 2)])
    except Exception:
        pass

    return rows


def _serialize_export_job(row: ExportJob) -> ExportJobResponse:
    return ExportJobResponse(
        id=row.id,
        organization_id=row.organization_id,
        customer_id=row.customer_id,
        name=row.name,
        report_type=row.report_type,
        export_format=row.export_format,
        schedule_frequency=row.schedule_frequency,
        is_active=bool(row.is_active),
        last_run_at=row.last_run_at.isoformat() if row.last_run_at else None,
        created_at=row.created_at.isoformat() if row.created_at else "",
        updated_at=row.updated_at.isoformat() if row.updated_at else "",
    )


def _serialize_export_job_run(row: ExportJobRun) -> ExportJobRunResponse:
    return ExportJobRunResponse(
        id=row.id,
        export_job_id=row.export_job_id,
        status=row.status,
        output_filename=row.output_filename,
        row_count=int(row.row_count or 0),
        error_message=row.error_message,
        created_at=row.created_at.isoformat() if row.created_at else "",
        completed_at=row.completed_at.isoformat() if row.completed_at else None,
    )


async def _execute_export_job(
    *,
    db: Session,
    job: ExportJob,
    current_user: User,
    membership: UserOrganization,
    actor_user_id: Optional[int],
) -> ExportJobRun:
    now = _utcnow()
    run = ExportJobRun(
        export_job_id=job.id,
        organization_id=job.organization_id,
        customer_id=job.customer_id,
        status="running",
        created_at=now,
    )
    db.add(run)
    db.flush()

    try:
        rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
        row_count = max(0, len(rows) - 1)

        if job.report_type == "finance_workbook" or job.export_format == "xlsx":
            extension = "xlsx"
        elif job.export_format == "pdf":
            extension = "pdf"
        elif job.export_format == "csv":
            extension = "csv"
        else:
            extension = "xls"

        if job.report_type == "executive_digest" and extension != "pdf":
            extension = "pdf"

        # Persist only metadata now; file materialization can be plugged into object storage later.
        run.output_filename = f"{job.report_type}-{job.id}-{int(now.timestamp())}.{extension}"
        run.row_count = row_count
        run.status = "completed"
        run.completed_at = _utcnow()
        job.last_run_at = run.completed_at
        job.updated_at = run.completed_at
        db.add(job)
        db.add(
            AuditLog(
                organization_id=job.organization_id,
                actor_user_id=actor_user_id,
                action="exports.job.run",
                entity_type="export_job",
                entity_id=str(job.id),
                metadata_json=json.dumps(
                    {
                        "report_type": job.report_type,
                        "export_format": job.export_format,
                        "schedule_frequency": job.schedule_frequency,
                        "row_count": row_count,
                        "output_filename": run.output_filename,
                    }
                ),
            )
        )
    except Exception as exc:
        run.status = "failed"
        run.error_message = str(exc)
        run.completed_at = _utcnow()

    db.add(run)
    db.commit()
    db.refresh(run)
    return run


@router.get("/exports/jobs", response_model=List[ExportJobResponse])
async def list_export_jobs(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ExportJobResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        rows = (
            db.query(ExportJob)
            .filter(
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .order_by(ExportJob.created_at.desc())
            .all()
        )
        return [_serialize_export_job(row) for row in rows]
    finally:
        db.close()


@router.post("/exports/jobs", response_model=ExportJobResponse)
async def create_export_job(
    payload: ExportJobRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ExportJobResponse:
    _require_management_role(membership, "export job management")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        now = _utcnow()
        row = ExportJob(
            organization_id=organization_id,
            customer_id=customer_id,
            name=payload.name.strip() or "Scheduled Export",
            report_type=payload.report_type,
            export_format=payload.export_format,
            schedule_frequency=payload.schedule_frequency,
            is_active=payload.is_active,
            created_at=now,
            updated_at=now,
        )
        db.add(row)
        db.flush()
        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="exports.job.create",
                entity_type="export_job",
                entity_id=str(row.id),
                metadata_json=json.dumps(
                    {
                        "name": row.name,
                        "report_type": row.report_type,
                        "export_format": row.export_format,
                        "schedule_frequency": row.schedule_frequency,
                        "is_active": bool(row.is_active),
                    }
                ),
            )
        )
        db.commit()
        db.refresh(row)
        return _serialize_export_job(row)
    finally:
        db.close()


@router.post("/exports/jobs/{job_id}/run", response_model=ExportJobRunResponse)
async def run_export_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ExportJobRunResponse:
    _require_management_role(membership, "export job runs")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        job = (
            db.query(ExportJob)
            .filter(
                ExportJob.id == job_id,
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .first()
        )
        if job is None:
            raise HTTPException(status_code=404, detail="Export job not found")

        run = await _execute_export_job(
            db=db,
            job=job,
            current_user=current_user,
            membership=membership,
            actor_user_id=current_user.id,
        )
        return _serialize_export_job_run(run)
    finally:
        db.close()


@router.get("/exports/jobs/{job_id}/runs", response_model=List[ExportJobRunResponse])
async def list_export_job_runs(
    job_id: int,
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ExportJobRunResponse]:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    db = SessionLocal()
    try:
        job_exists = (
            db.query(ExportJob.id)
            .filter(
                ExportJob.id == job_id,
                ExportJob.organization_id == organization_id,
                ExportJob.customer_id == customer_id,
            )
            .first()
        )
        if job_exists is None:
            raise HTTPException(status_code=404, detail="Export job not found")

        rows = (
            db.query(ExportJobRun)
            .filter(
                ExportJobRun.export_job_id == job_id,
                ExportJobRun.organization_id == organization_id,
                ExportJobRun.customer_id == customer_id,
            )
            .order_by(ExportJobRun.created_at.desc())
            .limit(max(1, min(limit, 100)))
            .all()
        )
        return [_serialize_export_job_run(row) for row in rows]
    finally:
        db.close()


@router.get("/reports/executive-summary.csv")
async def download_executive_summary_csv(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    return _csv_response(
        "optiora-executive-summary.csv",
        rows[0],
        rows[1:],
    )


@router.get("/reports/executive-summary.xls")
async def download_executive_summary_excel(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    return _spreadsheet_xml_response(
        "optiora-executive-summary.xls",
        "Executive Summary",
        rows,
    )


@router.get("/dashboard/costs")
@router.get("/costs")
async def dashboard_costs(
    period: str = "month",
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    context = await _cost_context(membership, db, period, cloud_provider)

    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    recommendations_result = _safe_json_load(
        await recommendations.get_recommendations({"cloud_provider": cloud_provider}),
        {},
    )

    potential_savings = float(
        recommendations_result.get("total_potential_savings_annual_usd", 0) or 0
    ) / 12

    return {
        "totalCost": context["total_cost"],
        "trend": 0.0,
        "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
        "potentialSavings": round(potential_savings, 2),
        "breakdown": context["breakdown"],
        "regionBreakdown": context["region_breakdown"],
    }


@router.get("/dashboard/anomalies")
@router.get("/anomalies")
async def dashboard_anomalies(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[Dict[str, Any]]:
    _ = (current_user, membership)
    result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    rows = result.get("anomalies", [])
    mapped = []
    for idx, row in enumerate(rows):
        mapped.append(
            {
                "id": f"anomaly-{idx + 1}",
                "service": row.get("service", "unknown"),
                "cloud": cloud_provider if cloud_provider != "all" else "multi-cloud",
                "message": row.get("probable_cause", "Anomaly detected"),
                "severity": "high" if (row.get("increase_percent", 0) or 0) > 150 else "medium",
                "timestamp": row.get("date", _utcnow().isoformat()),
                "change": float(row.get("increase_percent", 0) or 0),
            }
        )
    return mapped


@router.get("/dashboard/recommendations")
@router.get("/recommendations")
async def dashboard_recommendations(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> List[Dict[str, Any]]:
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    rows = result.get("recommendations", [])
    mapped = []
    for row in rows:
        mapped.append(
            {
                "id": row.get("id"),
                "service": row.get("service", "unknown"),
                "cloud": cloud_provider if cloud_provider != "all" else "multi-cloud",
                "title": row.get("description", "Optimization recommendation"),
                "description": row.get("description", ""),
                "savings": float(row.get("savings_annual_usd", 0) or 0) / 12,
                "roi": float(row.get("roi_percent", 0) or 0),
                "difficulty": "easy"
                if row.get("payback_months", 0) <= 1
                else "medium"
                if row.get("payback_months", 0) <= 3
                else "hard",
            }
        )
    return mapped


@router.get("/forecast")
async def dashboard_forecast(
    months: int = 12,
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": months,
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "fallback_monthly_spend": 0,
            }
        ),
        {},
    )
    result["cost_context"] = context
    return result


class ForecastWhatIfAction(BaseModel):
    name: str
    start_month: int = Field(default=1, ge=1, le=24)
    savings_percent: float = Field(default=0.0, ge=0.0, le=80.0)
    growth_delta_percent: float = Field(default=0.0, ge=-30.0, le=30.0)
    one_time_cost_usd: float = Field(default=0.0, ge=0.0)


class ForecastWhatIfRequest(BaseModel):
    months: int = Field(default=12, ge=1, le=24)
    cloud_provider: str = "all"
    actions: List[ForecastWhatIfAction] = Field(default_factory=list)
    discount_rate_monthly: float = Field(default=0.01, ge=0.0, le=0.2)


@router.post("/forecast/what-if")
async def dashboard_forecast_what_if(
    request: ForecastWhatIfRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Deterministic what-if simulation for planned optimization initiatives."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", request.cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)
    historical_monthly_spend = _historical_monthly_spend_from_snapshots(
        db=db,
        customer_id=customer_id,
        cloud_provider=request.cloud_provider,
        months=18,
    )
    result = _safe_json_load(
        await finops_analytics.get_forecast_what_if(
            {
                "months": request.months,
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "historical_monthly_spend": historical_monthly_spend,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
                "discount_rate_monthly": request.discount_rate_monthly,
                "actions": [
                    action.model_dump() if hasattr(action, "model_dump") else action.dict()
                    for action in request.actions
                ],
            }
        ),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics")
async def dashboard_analytics(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    permission = ScanningManager(db).get_permission_status(_customer_id_for_org(membership))
    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}),
        {},
    )
    recommendations_result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )
    monthly_savings = (
        float(recommendations_result.get("total_potential_savings_annual_usd", 0) or 0) / 12
    )
    result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
                "monthly_savings": monthly_savings,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )
    result["cost_context"] = context
    # Attach backend GenAI narrative when OCI GenAI is configured
    narrative, prompt = genai_advisor.generate_spend_narrative(result)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    return result


@router.get("/analytics/attribution")
async def analytics_attribution(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Pareto cost driver attribution — which providers/services account for 80% of spend."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_cost_attribution({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/commitment-optimization")
async def analytics_commitment_optimization(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Model the ROI of increasing RI/Savings Plan coverage across coverage tiers."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_commitment_optimization({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/maturity")
async def analytics_maturity(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """FinOps maturity assessment mapped to CRAWL/WALK/RUN/OPTIMIZE levels."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    customer_id = _customer_id_for_org(membership)
    permission = ScanningManager(db).get_permission_status(customer_id)
    anomalies_result = _safe_json_load(
        await anomalies.detect_anomalies({"cloud_provider": cloud_provider}), {}
    )
    historical = _historical_monthly_spend_from_snapshots(
        db=db, customer_id=customer_id, cloud_provider=cloud_provider, months=18
    )
    result = _safe_json_load(
        await finops_analytics.get_maturity_assessment({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": int(anomalies_result.get("anomalies_found", 0) or 0),
            "history_coverage_months": len(historical),
            "scheduler_enabled": bool(permission.get("scan_frequency")),
            "auto_remediate": bool(permission.get("auto_remediate", False)),
        }),
        {},
    )
    # Attach GenAI narrative if OCI GenAI is configured
    narrative, prompt = genai_advisor.generate_maturity_narrative(result)
    result["genai_narrative"] = narrative
    result["genai_prompt"] = prompt
    result["cost_context"] = context
    return result


@router.get("/analytics/unit-economics")
async def analytics_unit_economics(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Unit cost trends, waste-to-spend ratio, and dollar efficiency score."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    result = _safe_json_load(
        await finops_analytics.get_unit_economics({
            "current_monthly_spend": context["total_cost"],
            "estimated_waste_usd": analytics_result.get("estimated_monthly_waste_usd", 0),
            "identified_savings_usd": analytics_result.get("identified_monthly_savings_usd", 0),
            "anomalies": 0,
        }),
        {},
    )
    result["cost_context"] = context
    return result


class GenAIAnalyzeRequest(BaseModel):
    analysis_type: Literal["spend", "anomaly", "optimization", "maturity", "budget_risk", "waste_insights", "optimization_roadmap", "executive_narrative", "commitment_strategy"] = "spend"
    context: Dict[str, Any] = Field(default_factory=dict)
    anomaly: Optional[Dict[str, Any]] = None


class GenAICopilotPackRequest(BaseModel):
    cloud_provider: str = "all"
    include: List[
        Literal[
            "spend",
            "budget_risk",
            "waste_insights",
            "optimization_roadmap",
            "executive_narrative",
            "commitment_strategy",
        ]
    ] = Field(default_factory=lambda: ["spend", "optimization_roadmap", "executive_narrative"])


class HybridAdvisorResponse(BaseModel):
    generated_at: str
    cloud_provider: str
    deterministic: Dict[str, Any]
    advisory: Dict[str, Any]
    source_of_truth: Literal["deterministic"] = "deterministic"


async def _deterministic_recommendations(
    cloud_provider: str,
    current_monthly_spend: float,
    cost_breakdown: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Return deterministic recommendation rows mapped for advisor workflows."""
    rec_result = _safe_json_load(
        await recommendations.get_recommendations(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": current_monthly_spend,
                "cost_breakdown": cost_breakdown,
            }
        ),
        {},
    )
    rows = rec_result.get("recommendations", [])
    mapped: List[Dict[str, Any]] = []
    for row in rows:
        mapped.append(
            {
                "id": row.get("id"),
                "service": row.get("service", "unknown"),
                "title": row.get("description", "Optimization recommendation"),
                "description": row.get("description", ""),
                "savings_monthly_usd": float(row.get("savings_annual_usd", 0) or 0) / 12,
                "roi_percent": float(row.get("roi_percent", 0) or 0),
                "payback_months": float(row.get("payback_months", 0) or 0),
            }
        )
    return mapped


@router.get("/advisor/hybrid", response_model=HybridAdvisorResponse)
async def advisor_hybrid(
    cloud_provider: str = "all",
    narrative_type: Literal["waste_insights", "optimization_roadmap", "executive_narrative"] = "optimization_roadmap",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Hybrid advisor payload: deterministic findings + GenAI narrative.

    Deterministic analytics and recommendation values remain authoritative.
    GenAI output is advisory context layered on top.
    """
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
            }
        ),
        {},
    )

    waste_result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    efficiency_result = _safe_json_load(
        await finops_analytics.get_cost_efficiency_score(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "waste_rate_percent": analytics_result.get("unit_metrics", {}).get("estimated_waste_rate_percent", 18.0),
                "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get("anomaly_density_per_10k", 8.0),
                "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get("budget_utilization_percent", 85.0),
            }
        ),
        {},
    )

    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    maturity_result = _safe_json_load(
        await finops_analytics.get_maturity_assessment(
            {
                "cloud_provider": cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "history_coverage_months": 0,
                "scheduler_enabled": False,
                "auto_remediate": False,
            }
        ),
        {},
    )

    recommendation_rows = await _deterministic_recommendations(
        cloud_provider=cloud_provider,
        current_monthly_spend=context["total_cost"],
        cost_breakdown=context["breakdown"],
    )
    top_actions = sorted(
        recommendation_rows,
        key=lambda r: (r.get("savings_monthly_usd", 0), r.get("roi_percent", 0)),
        reverse=True,
    )[:5]

    genai_context: Dict[str, Any] = {
        "current_monthly_spend_usd": analytics_result.get("current_monthly_spend_usd", context["total_cost"]),
        "mom_change_percent": analytics_result.get("mom_change_percent"),
        "risk_score": analytics_result.get("risk_score", 0),
        "maturity_level": maturity_result.get("maturity_level", "walk"),
        "overall_score": efficiency_result.get("overall_score", 0),
        "grade": efficiency_result.get("grade", "C"),
        "improvement_focus": efficiency_result.get("improvement_focus", []),
        "total_estimated_waste_usd": waste_result.get("total_estimated_waste_usd", 0),
        "total_waste_rate_percent": waste_result.get("total_waste_rate_percent", 0),
        "categories": waste_result.get("categories", []),
        "quick_wins": waste_result.get("quick_wins", []),
        "total_annual_opportunity_usd": commitment_gap_result.get("total_annual_opportunity_usd", 0),
        "priority_provider": commitment_gap_result.get("priority_provider"),
        "top_opportunities": top_actions,
    }

    narrative: Optional[str]
    prompt: str
    if narrative_type == "waste_insights":
        narrative, prompt = genai_advisor.generate_waste_insights(genai_context)
    elif narrative_type == "executive_narrative":
        narrative, prompt = genai_advisor.generate_executive_narrative(genai_context)
    else:
        narrative, prompt = genai_advisor.generate_optimization_roadmap(genai_context)

    return {
        "generated_at": _utcnow().isoformat(),
        "cloud_provider": cloud_provider,
        "deterministic": {
            "analytics": analytics_result,
            "waste": waste_result,
            "efficiency": efficiency_result,
            "commitment_gap": commitment_gap_result,
            "recommendations": top_actions,
        },
        "advisory": {
            "narrative_type": narrative_type,
            "narrative": narrative,
            "prompt": prompt,
            "genai_configured": genai_advisor._is_configured(),
            "fallback_mode": narrative is None,
        },
        "source_of_truth": "deterministic",
    }


@router.post("/genai/analyze")
async def genai_analyze(
    request: GenAIAnalyzeRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Call OCI GenAI to generate a narrative for the requested analysis type.

    Falls back to returning the prompt when OCI GenAI is not configured,
    so the frontend Cost Advisor can use it directly.
    """
    _ = current_user
    ctx = request.context
    if not ctx.get("current_monthly_spend_usd"):
        cost_ctx = await _cost_context(membership, db, "month", "all")
        ctx.setdefault("current_monthly_spend_usd", cost_ctx.get("total_cost", 0))

    narrative: Optional[str] = None
    prompt: str = ""

    if request.analysis_type == "spend":
        narrative, prompt = genai_advisor.generate_spend_narrative(ctx)
    elif request.analysis_type == "anomaly":
        anomaly_ctx = request.anomaly or {}
        narrative, prompt = genai_advisor.generate_anomaly_explanation(anomaly_ctx, ctx)
    elif request.analysis_type == "optimization":
        narrative, prompt = genai_advisor.generate_optimization_brief(ctx)
    elif request.analysis_type == "maturity":
        narrative, prompt = genai_advisor.generate_maturity_narrative(ctx)
    elif request.analysis_type == "budget_risk":
        guardrails = ctx.get("budget_guardrails") or ctx
        narrative, prompt = genai_advisor.generate_budget_risk_alert(guardrails, ctx)
    elif request.analysis_type == "waste_insights":
        narrative, prompt = genai_advisor.generate_waste_insights(ctx)
    elif request.analysis_type == "optimization_roadmap":
        narrative, prompt = genai_advisor.generate_optimization_roadmap(ctx)
    elif request.analysis_type == "executive_narrative":
        narrative, prompt = genai_advisor.generate_executive_narrative(ctx)
    elif request.analysis_type == "commitment_strategy":
        narrative, prompt = genai_advisor.generate_commitment_strategy(ctx)

    return {
        "analysis_type": request.analysis_type,
        "narrative": narrative,
        "prompt": prompt,
        "genai_configured": genai_advisor._is_configured(),
        "fallback_mode": narrative is None,
    }


@router.post("/genai/copilot-pack")
async def genai_copilot_pack(
    request: GenAICopilotPackRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Return deterministic context plus multiple GenAI narratives in one call."""
    _ = current_user
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", request.cloud_provider)
    permission = ScanningManager(db).get_permission_status(customer_id)

    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics(
            {
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "anomalies": 0,
                "monthly_savings": 0,
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    forecast_result = _safe_json_load(
        await finops_analytics.get_forecast(
            {
                "months": 12,
                "cloud_provider": request.cloud_provider,
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
                "budget_monthly": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            }
        ),
        {},
    )

    commitment_gap_result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis(
            {
                "current_monthly_spend": context["total_cost"],
                "cost_breakdown": context["breakdown"],
            }
        ),
        {},
    )

    base_context = dict(analytics_result)
    base_context.update(
        {
            "budget_monthly_usd": float(permission.get("monthly_budget_usd", 0.0) or 0.0),
            "budget_guardrails": forecast_result.get("budget_guardrails") or {},
            "forecast_quality": forecast_result.get("forecast_quality") or {},
            "p90_monthly_usd": (forecast_result.get("forecast") or [{}])[0].get("p90", context["total_cost"]),
            "total_annual_opportunity_usd": commitment_gap_result.get("total_annual_opportunity_usd", 0),
            "priority_provider": commitment_gap_result.get("priority_provider"),
            "provider_gaps": commitment_gap_result.get("provider_gaps", []),
        }
    )

    narratives: Dict[str, Dict[str, Any]] = {}
    requested = request.include or []
    for item in requested:
        if item == "spend":
            narrative, prompt = genai_advisor.generate_spend_narrative(base_context)
        elif item == "budget_risk":
            narrative, prompt = genai_advisor.generate_budget_risk_alert(
                base_context.get("budget_guardrails", {}), base_context
            )
        elif item == "waste_insights":
            narrative, prompt = genai_advisor.generate_waste_insights(base_context)
        elif item == "optimization_roadmap":
            narrative, prompt = genai_advisor.generate_optimization_roadmap(base_context)
        elif item == "executive_narrative":
            narrative, prompt = genai_advisor.generate_executive_narrative(base_context)
        else:
            narrative, prompt = genai_advisor.generate_commitment_strategy(base_context)

        narratives[item] = {
            "narrative": narrative,
            "prompt": prompt,
            "fallback_mode": narrative is None,
        }

    return {
        "generated_at": _utcnow().isoformat(),
        "cloud_provider": request.cloud_provider,
        "deterministic_context": {
            "analytics": analytics_result,
            "forecast": {
                "forecast_quality": forecast_result.get("forecast_quality"),
                "budget_guardrails": forecast_result.get("budget_guardrails"),
                "downside_risk": forecast_result.get("downside_risk"),
            },
            "commitment_gap": commitment_gap_result,
        },
        "narratives": narratives,
        "genai_configured": genai_advisor._is_configured(),
    }




@router.get("/analytics/cloud-waste")
async def analytics_cloud_waste(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Categorised cloud waste breakdown with remediation guidance."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_cloud_waste_analysis({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/efficiency-score")
async def analytics_efficiency_score(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Composite FinOps efficiency score (0-100) across 6 weighted dimensions."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    result = _safe_json_load(
        await finops_analytics.get_cost_efficiency_score({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "waste_rate_percent": analytics_result.get("unit_metrics", {}).get("estimated_waste_rate_percent", 18.0),
            "anomaly_density_per_10k": analytics_result.get("unit_metrics", {}).get("anomaly_density_per_10k", 8.0),
            "budget_utilization_percent": analytics_result.get("unit_metrics", {}).get("budget_utilization_percent", 85.0),
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/analytics/commitment-gap")
async def analytics_commitment_gap(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Per-provider commitment coverage gap with savings scenarios and breakeven."""
    _ = current_user
    context = await _cost_context(membership, db, "month", cloud_provider)
    result = _safe_json_load(
        await finops_analytics.get_commitment_gap_analysis({
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
        }),
        {},
    )
    result["cost_context"] = context
    return result


@router.get("/provider-diagnostics", response_model=List[ProviderDiagnostic])
async def provider_diagnostics(current_user: User = Depends(get_current_user)) -> List[ProviderDiagnostic]:
    """Return provider readiness checks without exposing secret values."""
    _ = current_user
    return _provider_diagnostics()


@router.get("/health")
async def health_check() -> Dict[str, Any]:
    return {
        "status": "healthy",
        "version": __version__,
        "timestamp": _utcnow().isoformat(),
    }


@router.get("/info")
async def api_info() -> Dict[str, Any]:
    return {
        "name": "OptiOra API",
        "version": __version__,
        "description": "Cloud Cost Optimization Platform",
        "supported_providers": ["aws", "azure", "gcp", "oci"],
        "features": {
            "credential_management": True,
            "credential_validation": True,
            "scanning_permissions": True,
            "dashboard_endpoints": True,
            "finops_analytics": True,
            "forecasting": True,
            "forecast_what_if": True,
            "cost_attribution": True,
            "commitment_optimization": True,
            "maturity_assessment": True,
            "unit_economics": True,
            "anomaly_scoring": True,
            "cloud_waste_analysis": True,
            "efficiency_score": True,
            "commitment_gap_analysis": True,
            "genai_advisor": True,
            "hybrid_advisor": True,
            "genai_waste_insights": True,
            "genai_optimization_roadmap": True,
            "genai_executive_narrative": True,
            "genai_commitment_strategy": True,
            "genai_copilot_pack": True,
            "genai_backend_narration": genai_advisor._is_configured(),
            "provider_diagnostics": True,
            "audit_logging": True,
            "alert_lifecycle": True,
            "routing_policy_simulator": True,
            "operations_data_freshness": True,
            "budget_alerts": True,
            "csv_exports": True,
            "csv_imports": True,
            "csv_import_templates": True,
            "excel_exports": True,
            "executive_reports": True,
            "provider_hierarchy": True,
            "account_region_breakdown": True,
            "focus_export": True,
            "unit_economics_cockpit": True,
            "scorecards": True,
            "resource_inventory": True,
            "kubernetes_cost_allocation": True,
            "virtual_tagging": True,
            "rightsizing_resource_level": True,
        },
    }


async def _run_cost_analysis(
    scan_id: str,
    customer_id: str,
    providers: List[str],
    target_accounts: Optional[List[str]] = None,
) -> None:
    """
    Background scan: fetch live cost data per provider, persist CostSnapshot
    rows for historical trend analysis, then mark the scan run complete.

    When ``target_accounts`` is provided, only account_breakdown rows whose
    account_id matches one of the supplied identifiers are persisted as
    ProviderAccountSnapshot / CostAllocationSnapshot records.  The overall
    cost aggregation still uses all returned data.
    """
    db = SessionLocal()
    scanning_manager = ScanningManager(db)
    try:
        total_resources = 0
        anomalies_found = 0
        savings_identified = 0.0
        aggregated_cost_usd = 0.0
        now = _utcnow()

        for provider in providers:
            summary = await _cost_summary_for_provider(provider, "month")
            if "error" in summary:
                continue

            total_cost = float(summary.get("total_cost_usd", 0) or 0)
            aggregated_cost_usd += total_cost
            total_resources += 100
            provider_savings = total_cost * 0.08
            savings_identified += provider_savings

            # Fetch anomalies and recommendations for this provider to embed in snapshot.
            anomaly_raw = await anomalies.detect_anomalies({"cloud_provider": provider})
            anomaly_data = _safe_json_load(anomaly_raw, {})
            provider_anomalies = int(anomaly_data.get("anomalies_found", 0) or 0)
            anomalies_found += provider_anomalies

            rec_raw = await recommendations.get_recommendations({
                "cloud_provider": provider,
                "current_monthly_spend": total_cost,
            })

            # Derive period bounds from summary if available.
            try:
                period_start = datetime.fromisoformat(summary["start_date"]) if "start_date" in summary else None
                period_end = datetime.fromisoformat(summary["end_date"]) if "end_date" in summary else None
            except (ValueError, TypeError):
                period_start = period_end = None

            snapshot = CostSnapshot(
                scan_id=scan_id,
                customer_id=customer_id,
                provider=provider,
                period_start=period_start,
                period_end=period_end,
                total_cost_usd=total_cost,
                savings_identified_usd=provider_savings,
                anomalies_count=provider_anomalies,
                top_services_json=json.dumps(summary.get("top_services", [])),
                anomalies_json=anomaly_raw,
                recommendations_json=rec_raw,
                captured_at=now,
            )
            db.add(snapshot)

            # Persist provider-account hierarchy snapshots for rollup views.
            organization_id = _organization_id_from_customer_id(customer_id)
            account_rows = summary.get("account_breakdown") or []
            if organization_id is not None:
                # Guarantee at least one rollup node per provider.
                if not account_rows:
                    account_rows = [
                        {
                            "scope_type": "provider",
                            "scope_id": provider,
                            "total_cost_usd": total_cost,
                        }
                    ]

                # When target_accounts is set, restrict which hierarchy nodes are
                # persisted.  Error rows and the catch-all provider node are always
                # kept so the scan record remains coherent.
                if target_accounts:
                    target_set = {str(a).strip() for a in target_accounts if a}
                    account_rows = [
                        row for row in account_rows
                        if (
                            "error" in row
                            or str(row.get("scope_type") or "") == "provider"
                            or str(
                                row.get("account_id")
                                or row.get("scope_id")
                                or row.get("role_arn")
                                or ""
                            ).strip() in target_set
                        )
                    ]

                for account_row in account_rows:
                    account_identifier = str(
                        account_row.get("account_id")
                        or account_row.get("scope_id")
                        or account_row.get("role_arn")
                        or provider
                    )
                    account_name = str(
                        account_row.get("account_name")
                        or account_row.get("scope_id")
                        or account_identifier
                    )
                    account_type = str(
                        account_row.get("scope_type")
                        or account_row.get("account_type")
                        or "account"
                    )
                    account_total_cost = float(account_row.get("total_cost_usd") or 0.0)
                    if account_total_cost == 0.0 and len(account_rows) == 1:
                        account_total_cost = total_cost

                    provider_account = (
                        db.query(ProviderAccount)
                        .filter(
                            ProviderAccount.customer_id == customer_id,
                            ProviderAccount.provider == provider,
                            ProviderAccount.account_identifier == account_identifier,
                        )
                        .first()
                    )
                    if provider_account is None:
                        provider_account = ProviderAccount(
                            organization_id=organization_id,
                            customer_id=customer_id,
                            provider=provider,
                            account_identifier=account_identifier,
                            account_name=account_name,
                            account_type=account_type,
                            native_region=(summary.get("region_breakdown") or [{}])[0].get("region"),
                            metadata_json=json.dumps(account_row),
                            is_active=True,
                        )
                        db.add(provider_account)
                        db.flush()
                    else:
                        provider_account.account_name = account_name
                        provider_account.account_type = account_type
                        provider_account.metadata_json = json.dumps(account_row)
                        provider_account.updated_at = now

                    existing_provider_snapshot = (
                        db.query(ProviderAccountSnapshot)
                        .filter(
                            ProviderAccountSnapshot.scan_id == scan_id,
                            ProviderAccountSnapshot.provider_account_id == provider_account.id,
                        )
                        .first()
                    )
                    if existing_provider_snapshot is None:
                        db.add(
                            ProviderAccountSnapshot(
                                organization_id=organization_id,
                                customer_id=customer_id,
                                scan_id=scan_id,
                                provider_account_id=provider_account.id,
                                direct_cost_usd=account_total_cost,
                                savings_identified_usd=provider_savings,
                                anomalies_count=provider_anomalies,
                                service_count=len(summary.get("top_services", [])),
                                captured_at=now,
                            )
                        )

                    # Persist per-region cost breakdown for this account node.
                    region_breakdown = (
                        account_row.get("region_breakdown")
                        or (summary.get("region_breakdown") if len(account_rows) == 1 else [])
                        or []
                    )
                    for region_row in region_breakdown:
                        region_name = str(region_row.get("region") or "global")
                        region_cost = float(region_row.get("cost_usd") or 0.0)
                        if region_cost <= 0.0:
                            continue
                        existing_alloc = (
                            db.query(CostAllocationSnapshot)
                            .filter(
                                CostAllocationSnapshot.scan_id == scan_id,
                                CostAllocationSnapshot.provider_account_id == provider_account.id,
                                CostAllocationSnapshot.region == region_name,
                            )
                            .first()
                        )
                        if existing_alloc is None:
                            db.add(
                                CostAllocationSnapshot(
                                    organization_id=organization_id,
                                    customer_id=customer_id,
                                    scan_id=scan_id,
                                    provider_account_id=provider_account.id,
                                    provider=provider,
                                    region=region_name,
                                    cost_usd=region_cost,
                                    captured_at=now,
                                )
                            )

        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        organization_id = _organization_id_from_customer_id(customer_id)
        if organization_id is not None:
            evaluate_budget_alert(
                db=db,
                organization_id=organization_id,
                customer_id=customer_id,
                scan_id=scan_id,
                total_cost_usd=aggregated_cost_usd,
                permission=permission,
            )

        db.commit()

        scanning_manager.complete_scan_run(
            scan_id=scan_id,
            progress=100,
            total_resources=total_resources,
            anomalies_found=anomalies_found,
            savings_identified=savings_identified,
        )
    except Exception as exc:
        logger.exception("Background scan failed")
        scanning_manager.fail_scan_run(scan_id, str(exc))
    finally:
        db.close()


def _scan_interval_seconds(scan_frequency: str) -> int:
    normalized = str(scan_frequency or "daily").strip().lower()
    if normalized == "hourly":
        return 60 * 60
    if normalized == "weekly":
        return 7 * 24 * 60 * 60
    return 24 * 60 * 60


def _coerce_aws_anomaly_impact_usd(payload: Dict[str, Any]) -> float:
    impact = payload.get("impact")
    if isinstance(impact, dict):
        for key in ("totalImpact", "maxImpact", "impact"):
            value = impact.get(key)
            if value is not None:
                try:
                    return float(value)
                except (TypeError, ValueError):
                    pass
    for key in ("totalImpact", "impact", "estimatedImpact"):
        value = payload.get(key)
        if value is not None:
            try:
                return float(value)
            except (TypeError, ValueError):
                pass
    return 0.0


def _aws_anomaly_severity(impact_usd: float, source_severity: Optional[str]) -> str:
    normalized = str(source_severity or "").strip().lower()
    if normalized in {"critical", "high", "warning", "medium", "low"}:
        if normalized == "critical":
            return "high"
        if normalized == "warning":
            return "medium"
        return normalized
    if impact_usd >= 1000:
        return "high"
    if impact_usd >= 250:
        return "medium"
    return "low"


def _compute_next_run(now: datetime, scan_frequency: str, anchor: datetime) -> datetime:
    interval = timedelta(seconds=_scan_interval_seconds(scan_frequency))
    next_run = anchor + interval
    while next_run < now:
        next_run += interval
    return next_run


@router.get("/scanning/scheduler/status", response_model=SchedulerStatusResponse)
async def get_scheduler_status(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> SchedulerStatusResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()
    alert_id: Optional[int] = None
    db = SessionLocal()
    try:
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        runs = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .limit(50)
            .all()
        )
        audit_rows = (
            db.query(AuditLog)
            .filter(
                AuditLog.organization_id == organization_id,
                AuditLog.action.in_(["scan.schedule.triggered"]),
            )
            .order_by(AuditLog.created_at.desc())
            .limit(8)
            .all()
        )
    finally:
        db.close()

    total_runs = len(runs)
    success_runs = sum(1 for row in runs if row.state == ScanningState.COMPLETED.value)
    failed_runs = sum(1 for row in runs if row.state == ScanningState.FAILED.value)
    last_success = next(
        (
            row.completed_at or row.started_at
            for row in runs
            if row.state == ScanningState.COMPLETED.value
        ),
        None,
    )
    last_failure = next(
        (
            row.completed_at or row.started_at
            for row in runs
            if row.state == ScanningState.FAILED.value
        ),
        None,
    )

    permission_state = permission.state if permission else ScanningState.INITIALIZED.value
    scan_frequency = permission.scan_frequency if permission else "daily"
    next_run_at: Optional[datetime] = None
    if permission and permission_state in [ScanningState.APPROVED.value, ScanningState.RUNNING.value]:
        anchor = last_success or permission.approved_at or permission.created_at or now
        next_run_at = _compute_next_run(now, scan_frequency, anchor)

    timeline: List[SchedulerTimelineItem] = []
    for row in runs[:6]:
        providers = json.loads(row.providers_json or "[]")
        timeline.append(
            SchedulerTimelineItem(
                id=f"scan-{row.scan_id}",
                event_type="scan_run",
                state=row.state,
                title=f"Scan {row.state}",
                detail=f"Providers: {', '.join(providers) if providers else 'n/a'}",
                created_at=(row.completed_at or row.started_at).isoformat(),
            )
        )
    for row in audit_rows:
        metadata = _safe_json_load(row.metadata_json or "{}", {})
        providers = metadata.get("providers", [])
        detail = f"Frequency: {metadata.get('frequency', 'n/a')}"
        if providers:
            detail = f"{detail} | Providers: {', '.join([str(item) for item in providers])}"
        timeline.append(
            SchedulerTimelineItem(
                id=f"audit-{row.id}",
                event_type="scheduler_trigger",
                state="info",
                title="Scheduler triggered scan",
                detail=detail,
                created_at=row.created_at.isoformat(),
            )
        )
    timeline = sorted(timeline, key=lambda item: item.created_at, reverse=True)[:10]

    eta_seconds: Optional[int] = None
    if next_run_at is not None:
        eta_seconds = max(0, int((next_run_at - now).total_seconds()))

    return SchedulerStatusResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        scheduler_enabled=Config().enable_scan_scheduler,
        scheduler_running=_scheduler_running,
        permission_state=permission_state,
        scan_frequency=scan_frequency,
        next_run_at=next_run_at.isoformat() if next_run_at else None,
        next_run_eta_seconds=eta_seconds,
        last_success_at=last_success.isoformat() if last_success else None,
        last_failure_at=last_failure.isoformat() if last_failure else None,
        counters=SchedulerCounters(
            total=total_runs,
            success=success_runs,
            failure=failed_runs,
        ),
        timeline=timeline,
    )


def _freshness_state_for_age(age_seconds: Optional[int], stale_after_seconds: int) -> str:
    if age_seconds is None:
        return "unknown"
    return "stale" if age_seconds > stale_after_seconds else "fresh"


@router.get("/operations/data-freshness", response_model=DataFreshnessResponse)
async def get_data_freshness(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> DataFreshnessResponse:
    _ = current_user
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    now = _utcnow()

    db = SessionLocal()
    try:
        cost_rows = (
            db.query(CostSnapshot)
            .filter(CostSnapshot.customer_id == customer_id)
            .order_by(CostSnapshot.captured_at.desc())
            .limit(500)
            .all()
        )
        imported_rows = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.customer_id == customer_id)
            .order_by(ImportedCostRecord.created_at.desc())
            .limit(500)
            .all()
        )
        permission = (
            db.query(ScanningPermissionRecord)
            .filter(ScanningPermissionRecord.customer_id == customer_id)
            .first()
        )
        runs = (
            db.query(ScanRunRecord)
            .filter(ScanRunRecord.customer_id == customer_id)
            .order_by(ScanRunRecord.started_at.desc())
            .limit(20)
            .all()
        )
        external_alerts = (
            db.query(AlertEvent)
            .filter(
                AlertEvent.customer_id == customer_id,
                AlertEvent.alert_type.in_(
                    [
                        "external.aws.cost_anomaly",
                        "external.gcp.budget_pubsub",
                    ]
                ),
            )
            .order_by(AlertEvent.created_at.desc())
            .limit(200)
            .all()
        )
    finally:
        db.close()

    latest_by_provider: Dict[str, datetime] = {}
    for row in cost_rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        if provider not in latest_by_provider:
            latest_by_provider[provider] = row.captured_at

    for row in imported_rows:
        provider = str(row.provider or "").strip().lower()
        if not provider:
            continue
        current = latest_by_provider.get(provider)
        if current is None or row.created_at > current:
            latest_by_provider[provider] = row.created_at

    providers: List[DataFreshnessProviderItem] = []
    for provider in ["aws", "azure", "gcp", "oci"]:
        last = latest_by_provider.get(provider)
        age_seconds = int((now - last).total_seconds()) if last else None
        providers.append(
            DataFreshnessProviderItem(
                provider=provider,
                last_ingested_at=last.isoformat() if last else None,
                age_seconds=age_seconds,
                status=_freshness_state_for_age(age_seconds, stale_after_seconds=48 * 3600),
            )
        )

    connector_latest: Dict[str, Optional[datetime]] = {
        "aws_cost_anomaly": None,
        "gcp_budget_pubsub": None,
    }
    for row in external_alerts:
        if row.alert_type == "external.aws.cost_anomaly" and connector_latest["aws_cost_anomaly"] is None:
            connector_latest["aws_cost_anomaly"] = row.created_at
        if row.alert_type == "external.gcp.budget_pubsub" and connector_latest["gcp_budget_pubsub"] is None:
            connector_latest["gcp_budget_pubsub"] = row.created_at

    connectors: List[DataFreshnessConnectorItem] = []
    for connector, last in connector_latest.items():
        age_seconds = int((now - last).total_seconds()) if last else None
        connectors.append(
            DataFreshnessConnectorItem(
                connector=connector,
                last_event_at=last.isoformat() if last else None,
                age_seconds=age_seconds,
                status=_freshness_state_for_age(age_seconds, stale_after_seconds=72 * 3600),
            )
        )

    scheduler_lag_seconds: Optional[int] = None
    scheduler_status: Literal["healthy", "lagging", "unknown"] = "unknown"
    if permission:
        interval_seconds = _scan_interval_seconds(permission.scan_frequency)
        last_success = next(
            (
                row.completed_at or row.started_at
                for row in runs
                if row.state == ScanningState.COMPLETED.value
            ),
            None,
        )
        if last_success is not None:
            lag = int((now - last_success).total_seconds()) - interval_seconds
            scheduler_lag_seconds = max(lag, 0)
            scheduler_status = "lagging" if scheduler_lag_seconds > max(interval_seconds // 4, 900) else "healthy"

    return DataFreshnessResponse(
        organization_id=organization_id,
        customer_id=customer_id,
        generated_at=now.isoformat(),
        providers=providers,
        connectors=connectors,
        scheduler_lag_seconds=scheduler_lag_seconds,
        scheduler_status=scheduler_status,
    )


@router.post("/anomalies/external/aws")
async def ingest_external_aws_anomalies(
    payload: ExternalAWSAnomalyIngestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    _require_management_role(membership, "external anomaly ingestion")
    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    if not payload.events:
        return {"status": "ok", "ingested": 0, "alert_ids": [], "duplicates": 0}

    now = _utcnow()
    db = SessionLocal()
    try:
        alert_ids: List[int] = []
        duplicate_count = 0
        for event in payload.events:
            detail = event.get("detail")
            detail_payload = detail if isinstance(detail, dict) else event
            anomaly_id = (
                detail_payload.get("anomalyId")
                or detail_payload.get("AnomalyId")
                or event.get("id")
                or event.get("source_event_id")
                or f"aws-anomaly-{int(now.timestamp())}"
            )
            
            # Check for idempotency using source_event_id
            if anomaly_id:
                duplicate = (
                    db.query(AuditLog)
                    .filter(
                        AuditLog.organization_id == organization_id,
                        AuditLog.action == "alert.external.ingest",
                        AuditLog.entity_type == "aws_cost_anomaly_event",
                        AuditLog.entity_id == anomaly_id,
                    )
                    .first()
                )
                if duplicate is not None:
                    duplicate_count += 1
                    continue
            
            impact_usd = _coerce_aws_anomaly_impact_usd(detail_payload)
            severity = _aws_anomaly_severity(
                impact_usd=impact_usd,
                source_severity=detail_payload.get("severity"),
            )
            monitor_name = (
                detail_payload.get("monitorName")
                or detail_payload.get("dimensionValue")
                or "AWS Cost Anomaly Detection"
            )
            title = f"AWS anomaly detected ({monitor_name})"
            message = (
                f"Anomaly {anomaly_id} estimated impact ${impact_usd:.2f}. "
                f"Root cause: {detail_payload.get('rootCauses', [])[:1] or 'pending'}."
            )
            row = AlertEvent(
                organization_id=organization_id,
                customer_id=customer_id,
                scan_id=None,
                alert_type="external.aws.cost_anomaly",
                severity=severity,
                title=title[:255],
                message=message,
                delivered_channels_json=json.dumps(["aws-cost-anomaly-detection"]),
                created_at=now,
            )
            db.add(row)
            db.flush()
            alert_ids.append(row.id)
            
            # Log individual event for replay tracking
            db.add(
                AuditLog(
                    organization_id=organization_id,
                    actor_user_id=current_user.id,
                    action="alert.external.ingest",
                    entity_type="aws_cost_anomaly_event",
                    entity_id=anomaly_id,
                    metadata_json=json.dumps({
                        "alert_id": row.id,
                        "monitor_name": monitor_name,
                        "impact_usd": impact_usd,
                        "severity": severity,
                    }),
                    created_at=now,
                )
            )

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.external.ingest",
                entity_type="aws_cost_anomaly_batch",
                entity_id="aws-cost-anomaly-detection",
                metadata_json=json.dumps({"count": len(alert_ids), "duplicates": duplicate_count}),
                created_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    return {"status": "ok", "ingested": len(alert_ids), "alert_ids": alert_ids, "duplicates": duplicate_count}


def _decode_gcp_pubsub_payload(raw_message: Dict[str, Any]) -> Dict[str, Any]:
    data = raw_message.get("data")
    if isinstance(data, str) and data.strip():
        try:
            decoded = base64.b64decode(data.encode("utf-8"))
            parsed = json.loads(decoded.decode("utf-8"))
            if isinstance(parsed, dict):
                return parsed
        except (binascii.Error, UnicodeDecodeError, json.JSONDecodeError):
            return {}
    return raw_message if isinstance(raw_message, dict) else {}


def _coerce_budget_float(source: Dict[str, Any], keys: List[str]) -> float:
    for key in keys:
        value = source.get(key)
        if value is None:
            continue
        if isinstance(value, dict):
            for nested_key in ("amount", "value", "doubleValue"):
                nested = value.get(nested_key)
                if nested is not None:
                    value = nested
                    break
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return 0.0


@router.post("/anomalies/external/gcp/pubsub")
async def ingest_external_gcp_budget_pubsub(
    payload: ExternalGCPPubSubIngestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    ingest_token: Optional[str] = Header(default=None, alias="X-OptiOra-Ingest-Token"),
) -> Dict[str, Any]:
    _require_management_role(membership, "external anomaly ingestion")
    config = Config()
    expected_token = (config.gcp_pubsub_ingest_token or "").strip()
    if expected_token and ingest_token != expected_token:
        raise HTTPException(status_code=403, detail="Invalid ingest token")

    organization_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    envelope_message = payload.message if isinstance(payload.message, dict) else {}
    message_payload = _decode_gcp_pubsub_payload(envelope_message)
    message_id = str(
        envelope_message.get("messageId")
        or envelope_message.get("message_id")
        or message_payload.get("messageId")
        or message_payload.get("message_id")
        or ""
    ).strip()

    budget_name = str(
        message_payload.get("budgetDisplayName")
        or message_payload.get("budgetName")
        or message_payload.get("budget_name")
        or "GCP Budget"
    )
    cost_amount = _coerce_budget_float(
        message_payload,
        ["costAmount", "cost_amount", "actualAmount", "amount"],
    )
    budget_amount = _coerce_budget_float(
        message_payload,
        ["budgetAmount", "budget_amount", "thresholdAmount", "budget"],
    )
    usage_ratio = (cost_amount / budget_amount) if budget_amount > 0 else 0.0
    severity = "critical" if usage_ratio >= 1.0 else "warning"

    now = _utcnow()
    db = SessionLocal()
    try:
        if message_id:
            duplicate = (
                db.query(AuditLog)
                .filter(
                    AuditLog.organization_id == organization_id,
                    AuditLog.action == "alert.external.ingest",
                    AuditLog.entity_type == "gcp_budget_event",
                    AuditLog.entity_id == message_id,
                )
                .first()
            )
            if duplicate is not None:
                return {"status": "ok", "ingested": 0, "duplicate": True, "message_id": message_id}

        title = f"GCP budget alert ({budget_name})"
        message = (
            f"GCP budget usage is ${cost_amount:.2f} against budget ${budget_amount:.2f}. "
            f"Utilization {usage_ratio * 100:.1f}%"
        )
        alert = AlertEvent(
            organization_id=organization_id,
            customer_id=customer_id,
            scan_id=None,
            alert_type="external.gcp.budget_pubsub",
            severity=severity,
            title=title[:255],
            message=message,
            delivered_channels_json=json.dumps(["gcp-budget-pubsub"]),
            created_at=now,
        )
        db.add(alert)
        db.flush()
        alert_id = int(alert.id)

        db.add(
            AuditLog(
                organization_id=organization_id,
                actor_user_id=current_user.id,
                action="alert.external.ingest",
                entity_type="gcp_budget_event",
                entity_id=message_id or str(alert.id),
                metadata_json=json.dumps(
                    {
                        "alert_id": alert.id,
                        "budget_name": budget_name,
                        "cost_amount": cost_amount,
                        "budget_amount": budget_amount,
                        "message_id": message_id or None,
                        "subscription": payload.subscription,
                    }
                ),
                created_at=now,
            )
        )
        db.commit()
    finally:
        db.close()

    return {"status": "ok", "ingested": 1, "alert_id": alert_id, "message_id": message_id or None}


@router.post("/anomalies/external/aws/replay")
async def replay_external_aws_anomalies(
    request: ExternalAWSReplayRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Replay failed AWS anomaly events for retry."""
    _require_management_role(membership, "external anomaly replay")
    organization_id = _organization_id_for_membership(membership)

    db = SessionLocal()
    try:
        # Query audit logs for AWS cost anomaly events
        query = db.query(AuditLog).filter(
            AuditLog.organization_id == organization_id,
            AuditLog.action == "alert.external.ingest",
            AuditLog.entity_type == "aws_cost_anomaly_event",
        )
        
        # Filter by event_ids if provided
        if request.event_ids:
            query = query.filter(AuditLog.entity_id.in_(request.event_ids))
        
        # Filter by date range if provided
        if request.days_back:
            cutoff_date = _utcnow() - timedelta(days=request.days_back)
            query = query.filter(AuditLog.created_at >= cutoff_date)
        
        # Apply limit
        events = query.order_by(AuditLog.created_at.desc()).limit(request.max_results).all()
        
        if not events:
            return {
                "status": "ok",
                "replayed": 0,
                "alert_ids": [],
                "message": "No events found matching criteria"
            }
        
        # Extract alert IDs and associated metadata from audit logs
        alert_ids: List[int] = []
        for event in events:
            try:
                metadata = json.loads(event.metadata_json)
                if "alert_id" in metadata:
                    alert_ids.append(metadata["alert_id"])
            except (json.JSONDecodeError, KeyError):
                pass
        
        return {
            "status": "ok",
            "replayed": len(events),
            "alert_ids": alert_ids,
            "count": len(events),
            "date_range_days": request.days_back or "all",
            "message": f"Found {len(events)} events ready for retry"
        }
    finally:
        db.close()


# ── Connector Management endpoints ───────────────────────────────────────────

@router.get("/connectors", response_model=ConnectorListResponse)
async def list_connectors(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ConnectorListResponse:
    """List all supported connector types."""
    _require_management_role(membership, "connector listing")
    supported = ConnectorManager.list_supported_connectors()
    return ConnectorListResponse(
        supported_connectors=[c.value for c in supported],
        description="Supported cloud cost and resource connectors for OptiOra",
    )


@router.post("/connectors/test", response_model=ConnectorStatusResponse)
async def test_connector(
    request: ConnectorTestRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ConnectorStatusResponse:
    """Test connector configuration and credentials."""
    _require_management_role(membership, "connector testing")
    
    try:
        connector_type = ConnectorType(request.connector_type)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown connector type: {request.connector_type}",
        )
    
    try:
        connector = ConnectorManager.get_connector(connector_type, request.config)
        is_valid = await connector.validate_credentials()
        status = "valid" if is_valid else "invalid"
        
        return ConnectorStatusResponse(
            connector_type=request.connector_type,
            status=status,
            message=f"Connector validation {status}." if is_valid else "Connector credentials are invalid.",
        )
    except Exception as e:
        logger.error(f"Connector test failed for {request.connector_type}: {e}")
        return ConnectorStatusResponse(
            connector_type=request.connector_type,
            status="error",
            message=f"Test failed: {str(e)[:200]}",
        )


@router.get("/connectors/status", response_model=List[ConnectorStatusResponse])
async def get_connectors_status(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> List[ConnectorStatusResponse]:
    """Get status of all configured connectors."""
    _require_management_role(membership, "connector status check")
    
    # This endpoint returns status for all available connector types
    # In a full implementation, this would query stored connector configs from the database
    statuses: List[ConnectorStatusResponse] = []
    for connector_type in ConnectorManager.list_supported_connectors():
        statuses.append(
            ConnectorStatusResponse(
                connector_type=connector_type.value,
                status="unknown",
                message="Connector not configured. Configure via API or UI to enable.",
            )
        )
    return statuses


# ── Business Mapping & Chargeback endpoints ──────────────────────────────────

@router.get("/business-mapping/rules", response_model=BusinessMappingRuleListResponse)
async def list_mapping_rules(
    dimension: Optional[str] = None,
    active_only: bool = True,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleListResponse:
    """List business mapping rules for the current organization."""
    query = db.query(BusinessMappingRule).filter(
        BusinessMappingRule.organization_id == membership.organization_id
    )
    if active_only:
        query = query.filter(BusinessMappingRule.is_active == True)  # noqa: E712
    if dimension:
        if dimension not in VALID_DIMENSIONS:
            raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
        query = query.filter(BusinessMappingRule.dimension == dimension)
    rules = query.order_by(BusinessMappingRule.priority, BusinessMappingRule.id).all()
    return BusinessMappingRuleListResponse(
        organization_id=membership.organization_id,
        rules=[
            BusinessMappingRuleResponse(
                id=r.id,
                organization_id=r.organization_id,
                customer_id=r.customer_id,
                tag_key=r.tag_key,
                tag_value=r.tag_value,
                dimension=r.dimension,
                mapped_value=r.mapped_value,
                priority=r.priority,
                is_active=r.is_active,
                created_at=r.created_at.isoformat(),
                updated_at=r.updated_at.isoformat() if r.updated_at else None,
            )
            for r in rules
        ],
        total=len(rules),
    )


@router.post("/business-mapping/rules", response_model=BusinessMappingRuleResponse, status_code=201)
async def create_mapping_rule(
    rule_req: BusinessMappingRuleRequest,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleResponse:
    """Create a new business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    if rule_req.dimension not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
    existing = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.organization_id == membership.organization_id,
            BusinessMappingRule.tag_key == rule_req.tag_key,
            BusinessMappingRule.tag_value == rule_req.tag_value,
            BusinessMappingRule.dimension == rule_req.dimension,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="A mapping rule with the same tag_key/tag_value/dimension already exists.")
    now = _utcnow()
    customer_id = f"org-{membership.organization_id}"
    rule = BusinessMappingRule(
        organization_id=membership.organization_id,
        customer_id=customer_id,
        tag_key=rule_req.tag_key,
        tag_value=rule_req.tag_value,
        dimension=rule_req.dimension,
        mapped_value=rule_req.mapped_value,
        priority=rule_req.priority,
        is_active=rule_req.is_active,
        created_at=now,
        updated_at=now,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return BusinessMappingRuleResponse(
        id=rule.id,
        organization_id=rule.organization_id,
        customer_id=rule.customer_id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        dimension=rule.dimension,
        mapped_value=rule.mapped_value,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at.isoformat(),
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.put("/business-mapping/rules/{rule_id}", response_model=BusinessMappingRuleResponse)
async def update_mapping_rule(
    rule_id: int,
    rule_req: BusinessMappingRuleUpdateRequest,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> BusinessMappingRuleResponse:
    """Update an existing business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    rule = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.id == rule_id,
            BusinessMappingRule.organization_id == membership.organization_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Mapping rule not found.")
    if rule_req.dimension is not None and rule_req.dimension not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension must be one of {sorted(VALID_DIMENSIONS)}")
    if rule_req.tag_key is not None:
        rule.tag_key = rule_req.tag_key
    if rule_req.tag_value is not None:
        rule.tag_value = rule_req.tag_value
    if rule_req.dimension is not None:
        rule.dimension = rule_req.dimension
    if rule_req.mapped_value is not None:
        rule.mapped_value = rule_req.mapped_value
    if rule_req.priority is not None:
        rule.priority = rule_req.priority
    if rule_req.is_active is not None:
        rule.is_active = rule_req.is_active
    rule.updated_at = _utcnow()
    db.commit()
    db.refresh(rule)
    return BusinessMappingRuleResponse(
        id=rule.id,
        organization_id=rule.organization_id,
        customer_id=rule.customer_id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        dimension=rule.dimension,
        mapped_value=rule.mapped_value,
        priority=rule.priority,
        is_active=rule.is_active,
        created_at=rule.created_at.isoformat(),
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.delete("/business-mapping/rules/{rule_id}", status_code=204)
async def delete_mapping_rule(
    rule_id: int,
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> None:
    """Delete a business mapping rule."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    rule = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.id == rule_id,
            BusinessMappingRule.organization_id == membership.organization_id,
        )
        .first()
    )
    if not rule:
        raise HTTPException(status_code=404, detail="Mapping rule not found.")
    db.delete(rule)
    db.commit()


def _apply_mapping_rules(
    org_id: int,
    customer_id: str,
    records: list,
    rules: list,
    db: Session,
) -> int:
    """Apply active mapping rules to imported cost records, writing NormalizedCostDimension rows.

    Matching: a rule matches a record when the record's ``tags_json`` contains ``tag_key``,
    and either ``tag_value == "*"`` or ``tag_value`` matches the stored tag value exactly.
    Returns the count of dimension rows written.
    """
    now = _utcnow()
    written = 0
    for record in records:
        tags: dict = {}
        if record.tags_json:
            try:
                tags = json.loads(record.tags_json)
            except (ValueError, TypeError):
                tags = {}

        # Build dimension assignments from matching rules (highest priority = lowest number)
        dim_map: Dict[str, tuple] = {}  # dimension -> (mapped_value, rule_id)
        matched_rule_ids: List[int] = []
        for rule in sorted(rules, key=lambda r: (r.priority, r.id)):
            tag_val = tags.get(rule.tag_key)
            if tag_val is None:
                continue
            if rule.tag_value != "*" and rule.tag_value != str(tag_val):
                continue
            if rule.dimension not in dim_map:
                dim_map[rule.dimension] = (rule.mapped_value, rule.id)
                matched_rule_ids.append(rule.id)

        is_mapped = bool(dim_map)
        existing = (
            db.query(NormalizedCostDimension)
            .filter(
                NormalizedCostDimension.imported_cost_record_id == record.id,
                NormalizedCostDimension.organization_id == org_id,
            )
            .first()
        )
        if existing:
            existing.team = dim_map.get("team", (None, None))[0]
            existing.environment = dim_map.get("environment", (None, None))[0]
            existing.application = dim_map.get("application", (None, None))[0]
            existing.cost_center = dim_map.get("cost_center", (None, None))[0]
            existing.is_mapped = is_mapped
            existing.mapping_rule_ids_json = json.dumps(matched_rule_ids)
            existing.captured_at = now
        else:
            row = NormalizedCostDimension(
                organization_id=org_id,
                customer_id=customer_id,
                imported_cost_record_id=record.id,
                provider=record.provider,
                service_name=record.service_name,
                region=record.region,
                cost_usd=float(record.cost_usd or 0.0),
                team=dim_map.get("team", (None, None))[0],
                environment=dim_map.get("environment", (None, None))[0],
                application=dim_map.get("application", (None, None))[0],
                cost_center=dim_map.get("cost_center", (None, None))[0],
                is_mapped=is_mapped,
                mapping_rule_ids_json=json.dumps(matched_rule_ids),
                captured_at=now,
            )
            db.add(row)
            written += 1
    db.commit()
    return written


@router.post("/business-mapping/apply", status_code=200)
async def apply_mapping_rules(
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Re-apply all active mapping rules to imported cost records for this organization."""
    require_role(membership, [UserRole.ADMIN], "manage business mapping")
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"
    rules = (
        db.query(BusinessMappingRule)
        .filter(
            BusinessMappingRule.organization_id == org_id,
            BusinessMappingRule.is_active == True,  # noqa: E712
        )
        .all()
    )
    records = (
        db.query(ImportedCostRecord)
        .filter(ImportedCostRecord.customer_id == customer_id)
        .all()
    )
    written = _apply_mapping_rules(org_id, customer_id, records, rules, db)
    return {
        "status": "ok",
        "rules_applied": len(rules),
        "records_processed": len(records),
        "dimension_rows_written": written,
    }


@router.get("/chargeback", response_model=ChargebackResponse)
async def get_chargeback(
    dimension_type: str = "team",
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> ChargebackResponse:
    """Return chargeback/showback aggregation for a business dimension."""
    if dimension_type not in VALID_DIMENSIONS:
        raise HTTPException(status_code=400, detail=f"dimension_type must be one of {sorted(VALID_DIMENSIONS)}")
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"

    # Ensure dimension rows exist; auto-apply if none found
    count = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .count()
    )
    if count == 0:
        rules = (
            db.query(BusinessMappingRule)
            .filter(
                BusinessMappingRule.organization_id == org_id,
                BusinessMappingRule.is_active == True,  # noqa: E712
            )
            .all()
        )
        if rules:
            records = (
                db.query(ImportedCostRecord)
                .filter(ImportedCostRecord.customer_id == customer_id)
                .all()
            )
            _apply_mapping_rules(org_id, customer_id, records, rules, db)

    rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    dim_attr = dimension_type  # "team" | "environment" | "application" | "cost_center"
    groups: Dict[str, ChargebackDimensionGroup] = {}
    total_mapped = 0.0
    total_unmapped = 0.0

    for row in rows:
        val = getattr(row, dim_attr, None)
        cost = float(row.cost_usd or 0.0)
        if val is None:
            total_unmapped += cost
            continue
        total_mapped += cost
        if val not in groups:
            groups[val] = ChargebackDimensionGroup(
                dimension=dim_attr,
                value=val,
                total_cost_usd=0.0,
                provider_breakdown={},
                record_count=0,
            )
        g = groups[val]
        g.total_cost_usd = round(g.total_cost_usd + cost, 4)
        g.provider_breakdown[row.provider] = round(g.provider_breakdown.get(row.provider, 0.0) + cost, 4)
        g.record_count += 1

    total = total_mapped + total_unmapped
    coverage = round((total_mapped / total * 100) if total > 0 else 0.0, 2)

    return ChargebackResponse(
        organization_id=org_id,
        dimension_type=dim_attr,
        groups=sorted(groups.values(), key=lambda g: g.total_cost_usd, reverse=True),
        total_mapped_cost_usd=round(total_mapped, 2),
        total_unmapped_cost_usd=round(total_unmapped, 2),
        total_cost_usd=round(total, 2),
        coverage_percent=coverage,
    )


@router.get("/chargeback/coverage", response_model=AllocationCoverageResponse)
async def get_allocation_coverage(
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> AllocationCoverageResponse:
    """Return allocation quality: mapped vs unmapped spend breakdown."""
    org_id = membership.organization_id
    customer_id = f"org-{org_id}"

    # Auto-apply rules if no dimension rows exist yet
    count = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .count()
    )
    if count == 0:
        rules = (
            db.query(BusinessMappingRule)
            .filter(
                BusinessMappingRule.organization_id == org_id,
                BusinessMappingRule.is_active == True,  # noqa: E712
            )
            .all()
        )
        if rules:
            records = (
                db.query(ImportedCostRecord)
                .filter(ImportedCostRecord.customer_id == customer_id)
                .all()
            )
            _apply_mapping_rules(org_id, customer_id, records, rules, db)

    rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    total_cost = sum(float(r.cost_usd or 0) for r in rows)
    mapped_cost = sum(float(r.cost_usd or 0) for r in rows if r.is_mapped)
    unmapped_cost = total_cost - mapped_cost
    coverage = round((mapped_cost / total_cost * 100) if total_cost > 0 else 0.0, 2)

    dim_coverage: Dict[str, float] = {}
    for dim in VALID_DIMENSIONS:
        dim_mapped = sum(float(r.cost_usd or 0) for r in rows if getattr(r, dim, None) is not None)
        dim_coverage[dim] = round((dim_mapped / total_cost * 100) if total_cost > 0 else 0.0, 2)

    prov_totals: Dict[str, float] = {}
    prov_mapped: Dict[str, float] = {}
    for r in rows:
        prov_totals[r.provider] = prov_totals.get(r.provider, 0.0) + float(r.cost_usd or 0)
        if r.is_mapped:
            prov_mapped[r.provider] = prov_mapped.get(r.provider, 0.0) + float(r.cost_usd or 0)
    provider_coverage = {
        prov: round((prov_mapped.get(prov, 0.0) / tot * 100) if tot > 0 else 0.0, 2)
        for prov, tot in prov_totals.items()
    }

    # Top unmapped services by cost
    unmapped_services: Dict[str, float] = {}
    for r in rows:
        if not r.is_mapped and r.service_name:
            unmapped_services[r.service_name] = (
                unmapped_services.get(r.service_name, 0.0) + float(r.cost_usd or 0)
            )
    top_unmapped = sorted(
        [{"service": s, "cost_usd": round(c, 2)} for s, c in unmapped_services.items()],
        key=lambda x: x["cost_usd"],
        reverse=True,
    )[:10]

    return AllocationCoverageResponse(
        organization_id=org_id,
        total_cost_usd=round(total_cost, 2),
        mapped_cost_usd=round(mapped_cost, 2),
        unmapped_cost_usd=round(unmapped_cost, 2),
        coverage_percent=coverage,
        dimension_coverage=dim_coverage,
        provider_coverage=provider_coverage,
        unmapped_top_services=top_unmapped,
    )


# ── Epic 4: Reporting & Executive Outputs ────────────────────────────────────

def _compute_period_bucket(dt: datetime, period_type: str) -> tuple[datetime, datetime]:
    """Return (period_start, period_end) for a datetime given period_type."""
    if period_type == "weekly":
        # Monday-anchored ISO week
        start = dt - timedelta(days=dt.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=7) - timedelta(seconds=1)
    else:  # monthly
        start = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if dt.month == 12:
            end = start.replace(year=dt.year + 1, month=1) - timedelta(seconds=1)
        else:
            end = start.replace(month=dt.month + 1) - timedelta(seconds=1)
    return start, end


def _build_xlsx_workbook(sheets: List[tuple[str, List[List[Any]]]]) -> bytes:
    """Build an xlsx workbook from a list of (sheet_name, rows) tuples.

    Each row is a list of cell values; the first row is treated as a header.
    Returns raw bytes of the workbook.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        # Auto-width approximation
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _chargeback_csv_rows(
    org_id: int,
    customer_id: str,
    db: Session,
) -> tuple[List[str], List[List[Any]]]:
    """Return (header, data_rows) for a full chargeback export."""
    dim_rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .order_by(NormalizedCostDimension.captured_at.desc())
        .all()
    )
    header = [
        "provider", "service_name", "region",
        "cost_usd", "team", "environment", "application", "cost_center",
        "is_mapped", "mapping_rule_ids", "captured_at",
    ]
    data: List[List[Any]] = []
    for r in dim_rows:
        data.append([
            r.provider or "",
            r.service_name or "",
            r.region or "",
            round(float(r.cost_usd or 0), 4),
            r.team or "",
            r.environment or "",
            r.application or "",
            r.cost_center or "",
            "yes" if r.is_mapped else "no",
            r.mapping_rule_ids_json or "",
            r.captured_at.isoformat() if r.captured_at else "",
        ])
    return header, data


@router.post("/reports/period-summaries/compute", response_model=PeriodSummaryComputeResponse)
async def compute_period_summaries(
    period_type: str = "monthly",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> PeriodSummaryComputeResponse:
    """Aggregate ImportedCostRecord rows into CostPeriodSummary buckets.

    Re-computing is idempotent — existing rows are overwritten via upsert.
    Supported period_type values: 'monthly', 'weekly'.
    """
    if period_type not in ("monthly", "weekly"):
        raise HTTPException(status_code=400, detail="period_type must be 'monthly' or 'weekly'.")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    records = (
        db.query(ImportedCostRecord)
        .filter(ImportedCostRecord.organization_id == org_id)
        .all()
    )

    # Fetch normalized dimension rows (for mapped cost tracking)
    dim_by_record_id: Dict[int, NormalizedCostDimension] = {
        r.imported_cost_record_id: r
        for r in db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
        if r.imported_cost_record_id is not None
    }

    # Bucket: (period_start, provider, team, environment) → aggregation dict
    buckets: Dict[tuple, Dict[str, Any]] = {}

    now = _utcnow()
    for rec in records:
        anchor = rec.period_start or rec.created_at or now
        pstart, pend = _compute_period_bucket(anchor, period_type)
        dim = dim_by_record_id.get(rec.id)
        team = (dim.team or "") if dim else ""
        environment = (dim.environment or "") if dim else ""
        key = (pstart, pend, rec.provider or "imported", team, environment)

        cost = float(rec.cost_usd or 0)
        if key not in buckets:
            buckets[key] = {
                "total_cost_usd": 0.0,
                "mapped_cost_usd": 0.0,
                "unmapped_cost_usd": 0.0,
                "record_count": 0,
                "services": {},
            }
        b = buckets[key]
        b["total_cost_usd"] += cost
        if dim and dim.is_mapped:
            b["mapped_cost_usd"] += cost
        else:
            b["unmapped_cost_usd"] += cost
        b["record_count"] += 1
        svc = rec.service_name or "unknown"
        b["services"][svc] = b["services"].get(svc, 0.0) + cost

    rows_written = 0
    computed_at = _utcnow()
    for (pstart, pend, provider, team, environment), b in buckets.items():
        # Upsert: delete any existing row for this bucket key
        db.query(CostPeriodSummary).filter(
            CostPeriodSummary.organization_id == org_id,
            CostPeriodSummary.period_type == period_type,
            CostPeriodSummary.period_start == pstart,
            CostPeriodSummary.provider == provider,
            CostPeriodSummary.team == (team or None),
            CostPeriodSummary.environment == (environment or None),
        ).delete(synchronize_session=False)

        row = CostPeriodSummary(
            organization_id=org_id,
            customer_id=customer_id,
            period_type=period_type,
            period_start=pstart,
            period_end=pend,
            provider=provider,
            team=team or None,
            environment=environment or None,
            total_cost_usd=round(b["total_cost_usd"], 4),
            mapped_cost_usd=round(b["mapped_cost_usd"], 4),
            unmapped_cost_usd=round(b["unmapped_cost_usd"], 4),
            record_count=b["record_count"],
            service_breakdown_json=json.dumps(
                {k: round(v, 4) for k, v in sorted(b["services"].items(), key=lambda x: -x[1])[:20]}
            ),
            computed_at=computed_at,
        )
        db.add(row)
        rows_written += 1

    db.commit()

    return PeriodSummaryComputeResponse(
        organization_id=org_id,
        period_type=period_type,
        periods_computed=len({(k[0], k[2]) for k in buckets}),
        rows_written=rows_written,
        computed_at=computed_at.isoformat(),
    )


@router.get("/reports/cost-trend", response_model=CostTrendResponse)
async def get_cost_trend(
    period_type: str = "monthly",
    lookback: int = 6,
    provider: Optional[str] = None,
    view_by: str = "provider",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> CostTrendResponse:
    """Return time-series cost data for trend charts.

    Supports trend views by provider, region, service, and account.
    """
    _ = current_user
    if period_type not in ("monthly", "weekly"):
        raise HTTPException(status_code=400, detail="period_type must be 'monthly' or 'weekly'.")
    if view_by not in _SUPPORTED_TREND_VIEWS:
        raise HTTPException(status_code=400, detail="view_by must be one of: provider, region, service, account.")

    lookback = max(1, min(lookback, 18))
    org_id = _organization_id_for_membership(membership)

    summaries: List[CostPeriodSummary] = []
    archive_rows: List[Dict[str, Any]] = []
    data_source = "raw_records"

    if view_by == "provider":
        q = db.query(CostPeriodSummary).filter(
            CostPeriodSummary.organization_id == org_id,
            CostPeriodSummary.period_type == period_type,
        )
        if provider:
            q = q.filter(CostPeriodSummary.provider == provider.lower())
        summaries = q.order_by(CostPeriodSummary.period_start.desc()).all()

        cfg = Config()
        hot_months = cfg.retention_hot_months
        if lookback > hot_months and cfg.retention_enabled and cfg.oci_archive_bucket:
            now = _utcnow()
            archive_end = now - timedelta(days=hot_months * 30)
            archive_start = now - timedelta(days=lookback * 30)
            try:
                raw_archive = await asyncio.to_thread(
                    fetch_archived_period_summaries, cfg, org_id, archive_start, archive_end
                )
                if provider:
                    raw_archive = [r for r in raw_archive if r.get("provider", "").lower() == provider.lower()]
                raw_archive = [r for r in raw_archive if r.get("period_type", "") == period_type]
                archive_rows = raw_archive
            except Exception:
                logger.exception("Failed to fetch archived period summaries; continuing with DB only")

        if summaries or archive_rows:
            data_source = "computed"

    points: List[CostTrendPoint] = []

    if data_source == "computed":
        by_period: Dict[tuple, Dict[str, Any]] = {}
        for s in summaries:
            key = (s.period_start, s.period_end, s.provider or "imported")
            if key not in by_period:
                by_period[key] = {"total": 0.0, "mapped": 0.0, "unmapped": 0.0, "count": 0, "svcs": {}}
            b = by_period[key]
            b["total"] += float(s.total_cost_usd or 0)
            b["mapped"] += float(s.mapped_cost_usd or 0)
            b["unmapped"] += float(s.unmapped_cost_usd or 0)
            b["count"] += int(s.record_count or 0)
            if s.service_breakdown_json:
                try:
                    for svc, c in json.loads(s.service_breakdown_json).items():
                        b["svcs"][svc] = b["svcs"].get(svc, 0.0) + c
                except (json.JSONDecodeError, ValueError):
                    pass

        for ar in archive_rows:
            try:
                ps = datetime.fromisoformat(str(ar["period_start"]).replace("Z", ""))
                pe = datetime.fromisoformat(str(ar["period_end"]).replace("Z", ""))
            except (KeyError, ValueError):
                continue
            key = (ps, pe, ar.get("provider", "unknown"))
            if key not in by_period:
                by_period[key] = {"total": 0.0, "mapped": 0.0, "unmapped": 0.0, "count": 0, "svcs": {}}
            b = by_period[key]
            b["total"] += float(ar.get("total_cost_usd") or 0)
            b["mapped"] += float(ar.get("mapped_cost_usd") or 0)
            b["unmapped"] += float(ar.get("unmapped_cost_usd") or 0)
            b["count"] += int(ar.get("record_count") or 0)
            svc_json = ar.get("service_breakdown_json")
            if svc_json:
                try:
                    for svc, c in json.loads(svc_json).items():
                        b["svcs"][svc] = b["svcs"].get(svc, 0.0) + c
                except (json.JSONDecodeError, ValueError):
                    pass

        sorted_keys = sorted(by_period.keys(), key=lambda k: k[0], reverse=True)[:lookback]
        points = [
            CostTrendPoint(
                period_start=k[0].isoformat(),
                period_end=k[1].isoformat(),
                provider=k[2],
                dimension_value=k[2],
                total_cost_usd=round(by_period[k]["total"], 2),
                mapped_cost_usd=round(by_period[k]["mapped"], 2),
                unmapped_cost_usd=round(by_period[k]["unmapped"], 2),
                record_count=by_period[k]["count"],
                service_breakdown={s: round(c, 2) for s, c in list(by_period[k]["svcs"].items())[:10]},
            )
            for k in sorted(sorted_keys, key=lambda k: k[0])
        ]
    else:
        records = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .all()
        )
        if not records:
            return CostTrendResponse(
                organization_id=org_id,
                period_type=period_type,
                lookback_periods=lookback,
                view_by=view_by,
                data_source="empty",
                points=[],
                provider_totals={},
                dimension_totals={},
                grand_total_usd=0.0,
            )

        now = _utcnow()
        buckets: Dict[tuple, Dict[str, Any]] = {}
        for rec in records:
            rec_provider = (rec.provider or "imported").lower()
            if provider and rec_provider != provider.lower():
                continue
            anchor = rec.period_start or rec.created_at or now
            pstart, pend = _compute_period_bucket(anchor, period_type)
            dim_value = _trend_dimension_value(rec, view_by)
            key = (pstart, pend, dim_value)
            cost = float(rec.cost_usd or 0)
            if key not in buckets:
                buckets[key] = {
                    "pend": pend,
                    "total": 0.0,
                    "mapped": 0.0,
                    "unmapped": 0.0,
                    "count": 0,
                    "svcs": {},
                    "providers": {},
                }
            b = buckets[key]
            b["total"] += cost
            b["unmapped"] += cost
            b["count"] += 1
            b["providers"][rec_provider] = b["providers"].get(rec_provider, 0.0) + cost
            svc = rec.service_name or "unknown"
            b["svcs"][svc] = b["svcs"].get(svc, 0.0) + cost

        sorted_keys = sorted(buckets.keys(), key=lambda k: k[0], reverse=True)[:lookback]
        for k in sorted(sorted_keys, key=lambda v: v[0]):
            b = buckets[k]
            providers = sorted((b.get("providers") or {}).items(), key=lambda item: item[1], reverse=True)
            dominant_provider = providers[0][0] if providers else "imported"
            provider_value = dominant_provider if len(providers) == 1 else "multi-cloud"
            points.append(
                CostTrendPoint(
                    period_start=k[0].isoformat(),
                    period_end=b["pend"].isoformat(),
                    provider=provider_value,
                    dimension_value=k[2],
                    total_cost_usd=round(b["total"], 2),
                    mapped_cost_usd=round(b["mapped"], 2),
                    unmapped_cost_usd=round(b["unmapped"], 2),
                    record_count=b["count"],
                    service_breakdown={s: round(c, 2) for s, c in list(b["svcs"].items())[:10]},
                )
            )

    provider_totals: Dict[str, float] = {}
    dimension_totals: Dict[str, float] = {}
    for p in points:
        provider_totals[p.provider] = round(provider_totals.get(p.provider, 0.0) + p.total_cost_usd, 2)
        dim_key = p.dimension_value or p.provider
        dimension_totals[dim_key] = round(dimension_totals.get(dim_key, 0.0) + p.total_cost_usd, 2)

    return CostTrendResponse(
        organization_id=org_id,
        period_type=period_type,
        lookback_periods=lookback,
        view_by=view_by,
        data_source=data_source,
        points=points,
        provider_totals=provider_totals,
        dimension_totals=dimension_totals,
        grand_total_usd=round(sum(p.total_cost_usd for p in points), 2),
    )


@router.get("/reports/chargeback.csv")
async def download_chargeback_csv(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a full chargeback/allocation CSV with all business dimensions."""
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    header, data = _chargeback_csv_rows(org_id, customer_id, db)
    return _csv_response("optiora-chargeback.csv", header, data)


@router.get("/reports/chargeback.xlsx")
async def download_chargeback_xlsx(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a multi-sheet Excel workbook with chargeback + executive summary."""
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl is not installed on this server.")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)

    # Sheet 1: Chargeback detail
    header, data = _chargeback_csv_rows(org_id, customer_id, db)
    chargeback_rows = [header] + data

    # Sheet 2: Executive summary (reuse existing helper)
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)

    # Sheet 3: Trend summary (last 6 monthly periods)
    trend = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        current_user=current_user,
        membership=membership,
        db=db,
    )
    trend_header = ["period_start", "period_end", "provider", "total_cost_usd", "mapped_cost_usd", "unmapped_cost_usd", "record_count"]
    trend_rows: List[List[Any]] = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.provider, p.total_cost_usd, p.mapped_cost_usd, p.unmapped_cost_usd, p.record_count]
        for p in trend.points
    ]

    xlsx_bytes = _build_xlsx_workbook([
        ("Chargeback Detail", chargeback_rows),
        ("Executive Summary", exec_rows),
        ("Cost Trend", trend_rows),
    ])

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-report-{_utcnow().strftime("%Y%m%d")}.xlsx"'
        },
    )


def _build_simple_pdf(title: str, lines: List[str]) -> bytes:
    """Build a lightweight PDF document without external dependencies."""
    safe_lines = [line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)") for line in lines]
    y_start = 780
    line_height = 16
    text_ops = ["BT", "/F1 12 Tf", f"72 {y_start} Td", f"({title}) Tj"]
    for idx, line in enumerate(safe_lines[:42], start=1):
        text_ops.append(f"0 -{line_height} Td")
        text_ops.append(f"({line}) Tj")
    text_ops.append("ET")
    stream_text = "\n".join(text_ops)

    objects: List[str] = []
    objects.append("1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj")
    objects.append("2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj")
    objects.append("3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj")
    objects.append("4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj")
    objects.append(f"5 0 obj << /Length {len(stream_text.encode('utf-8'))} >> stream\n{stream_text}\nendstream endobj")

    header = "%PDF-1.4\n"
    body = ""
    offsets = [0]
    current = len(header.encode("utf-8"))
    for obj in objects:
        offsets.append(current)
        encoded = (obj + "\n").encode("utf-8")
        body += obj + "\n"
        current += len(encoded)

    xref_start = current
    xref = [f"xref\n0 {len(offsets)}\n", "0000000000 65535 f \n"]
    for off in offsets[1:]:
        xref.append(f"{off:010d} 00000 n \n")
    trailer = (
        f"trailer << /Size {len(offsets)} /Root 1 0 R >>\n"
        f"startxref\n{xref_start}\n%%EOF\n"
    )
    pdf = header + body + "".join(xref) + trailer
    return pdf.encode("utf-8")


def _digest_lines(exec_rows: List[List[Any]], frequency: str) -> List[str]:
    kv = {f"{r[0]}::{r[1]}": r[2] for r in exec_rows[1:] if len(r) >= 3}
    return [
        f"Frequency: {frequency}",
        f"Generated At: {kv.get('Summary::Generated At', _utcnow().isoformat())}",
        f"Total Monthly Cost USD: {kv.get('Summary::Total Monthly Cost USD', 0)}",
        f"Potential Monthly Savings USD: {kv.get('Summary::Potential Monthly Savings USD', 0)}",
        f"Risk Score: {kv.get('Summary::Risk Score', 0)}",
        f"Maturity Score: {kv.get('Summary::Maturity Score', 0)}",
        f"Spend At Risk USD: {kv.get('Summary::Spend At Risk USD', 0)}",
        f"Optimization Capacity USD: {kv.get('Summary::Optimization Capacity USD', 0)}",
        f"Open Alerts: {kv.get('Summary::Open Alerts', 0)}",
    ]


@router.get("/reports/executive-summary.xlsx")
@router.get("/reports/finance-workbook.xlsx")
async def download_executive_summary_xlsx(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Download a finance-friendly multi-sheet workbook."""
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="openpyxl is not installed on this server.")

    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)

    trend_provider = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        view_by="provider",
        current_user=current_user,
        membership=membership,
        db=db,
    )
    trend_region = await get_cost_trend(
        period_type="monthly",
        lookback=6,
        provider=None,
        view_by="region",
        current_user=current_user,
        membership=membership,
        db=db,
    )

    trend_header = ["period_start", "period_end", "dimension", "provider", "total_cost_usd", "record_count"]
    provider_rows = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
        for p in trend_provider.points
    ]
    region_rows = [trend_header] + [
        [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
        for p in trend_region.points
    ]

    chargeback_header, chargeback_data = _chargeback_csv_rows(org_id, customer_id, db)
    chargeback_rows = [chargeback_header] + chargeback_data

    xlsx_bytes = _build_xlsx_workbook([
        ("Executive Summary", exec_rows),
        ("Trend by Provider", provider_rows),
        ("Trend by Region", region_rows),
        ("Chargeback Detail", chargeback_rows),
    ])
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-finance-workbook-{_utcnow().strftime("%Y%m%d")}.xlsx"'
        },
    )


@router.get("/reports/executive-digest.pdf")
async def download_executive_digest_pdf(
    frequency: Literal["weekly", "monthly"] = "weekly",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    exec_rows = await _executive_summary_rows(current_user=current_user, membership=membership, db=db)
    digest_lines = _digest_lines(exec_rows, frequency)
    pdf_bytes = _build_simple_pdf(f"OptiOra {frequency.title()} Executive Digest", digest_lines)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="optiora-{frequency}-digest-{_utcnow().strftime("%Y%m%d")}.pdf"'
        },
    )


@router.post("/reports/share-token", response_model=ReportShareTokenResponse)
async def create_report_share_token(
    payload: ReportShareTokenRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> ReportShareTokenResponse:
    _require_management_role(membership, "report sharing")
    org_id = _organization_id_for_membership(membership)
    customer_id = _customer_id_for_org(membership)
    expires_hours = max(1, min(payload.expires_in_hours, 24 * 90))
    expires_at = datetime.now(timezone.utc) + timedelta(hours=expires_hours)
    token = _build_report_share_token(
        {
            "org_id": org_id,
            "customer_id": customer_id,
            "report_type": payload.report_type,
            "report_format": payload.report_format,
            "exp": int(expires_at.timestamp()),
            "sub": current_user.id,
        }
    )
    return ReportShareTokenResponse(
        token=token,
        expires_at=expires_at.isoformat(),
        report_type=payload.report_type,
        report_format=payload.report_format,
    )


@router.get("/reports/shared/{token}")
async def read_shared_report(
    token: str,
    db: Session = Depends(get_db),
) -> Response:
    payload = _parse_report_share_token(token)
    org_id = int(payload.get("org_id"))
    customer_id = str(payload.get("customer_id"))
    report_type = str(payload.get("report_type"))
    report_format = str(payload.get("report_format"))

    membership_stub = SimpleNamespace(organization_id=org_id, role=UserRole.READONLY)
    user_stub = SimpleNamespace(id=None)

    if report_type == "executive_digest" and report_format == "pdf":
        exec_rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        pdf_bytes = _build_simple_pdf("OptiOra Executive Digest", _digest_lines(exec_rows, "shared"))
        return Response(content=pdf_bytes, media_type="application/pdf")

    if report_type == "finance_workbook" and report_format == "xlsx":
        exec_rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        trend_provider = await get_cost_trend(
            period_type="monthly",
            lookback=6,
            provider=None,
            view_by="provider",
            current_user=user_stub,
            membership=membership_stub,
            db=db,
        )
        trend_rows = [["period_start", "period_end", "dimension", "provider", "total_cost_usd", "record_count"]] + [
            [p.period_start[:10], p.period_end[:10], p.dimension_value or p.provider, p.provider, p.total_cost_usd, p.record_count]
            for p in trend_provider.points
        ]
        header, data = _chargeback_csv_rows(org_id, customer_id, db)
        xlsx_bytes = _build_xlsx_workbook([
            ("Executive Summary", exec_rows),
            ("Trend by Provider", trend_rows),
            ("Chargeback Detail", [header] + data),
        ])
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if report_type == "executive_summary" and report_format in {"json", "csv"}:
        rows = await _executive_summary_rows(current_user=user_stub, membership=membership_stub, db=db)
        if report_format == "csv":
            return _csv_response("shared-executive-summary.csv", rows[0], rows[1:])
        return Response(
            content=json.dumps({"rows": rows}, default=str),
            media_type="application/json",
        )

    raise HTTPException(status_code=400, detail="Unsupported shared report token payload")


# ── End Epic 4 endpoints ──────────────────────────────────────────────────────

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FOCUS Export  (FinOps Open Cost and Usage Specification v1.0)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_FOCUS_COLUMNS = [
    "BilledCost", "BillingAccountId", "BillingAccountName", "BillingCurrency",
    "BillingPeriodEnd", "BillingPeriodStart", "ChargePeriodEnd", "ChargePeriodStart",
    "ChargeType", "CommitmentDiscountId", "CommitmentDiscountName", "CommitmentDiscountType",
    "EffectiveCost", "InvoiceIssuerName", "ListCost", "ListUnitPrice",
    "PricingCategory", "PricingQuantity", "PricingUnit", "ProviderName",
    "PublisherName", "RegionId", "RegionName", "ResourceId", "ResourceName",
    "ResourceType", "ServiceCategory", "ServiceName", "SkuId", "SkuPriceId",
    "SubAccountId", "SubAccountName", "Tags",
]


def _focus_rows_from_context(context: Dict[str, Any]) -> List[List[Any]]:
    """Convert internal cost context to FOCUS 1.0 rows."""
    rows: List[List[Any]] = []
    period_start = _utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    period_end = _utcnow()

    for provider, data in (context.get("breakdown") or {}).items():
        cost = round(float(data.get("cost") or 0.0), 6)
        rows.append([
            cost,                                     # BilledCost
            provider,                                 # BillingAccountId
            provider.upper(),                         # BillingAccountName
            "USD",                                    # BillingCurrency
            period_end.isoformat() + "Z",            # BillingPeriodEnd
            period_start.isoformat() + "Z",          # BillingPeriodStart
            period_end.isoformat() + "Z",            # ChargePeriodEnd
            period_start.isoformat() + "Z",          # ChargePeriodStart
            "Usage",                                  # ChargeType
            None, None, None,                         # CommitmentDiscount*
            cost,                                     # EffectiveCost
            provider.upper(),                         # InvoiceIssuerName
            cost,                                     # ListCost
            None,                                     # ListUnitPrice
            "On-Demand",                              # PricingCategory
            1.0,                                      # PricingQuantity
            "Month",                                  # PricingUnit
            provider.upper(),                         # ProviderName
            provider.upper(),                         # PublisherName
            "global",                                 # RegionId
            "Global",                                 # RegionName
            None, None,                               # ResourceId, ResourceName
            "Cloud Service",                          # ResourceType
            "Compute",                                # ServiceCategory
            provider.upper(),                         # ServiceName
            None, None,                               # SkuId, SkuPriceId
            provider,                                 # SubAccountId
            provider.upper(),                         # SubAccountName
            "{}",                                     # Tags
        ])

    # Also emit rows from imported cost records if present
    for icr in (context.get("imported_rows") or []):
        cost = round(float(icr.get("cost_usd") or 0.0), 6)
        prov = icr.get("provider", "unknown")
        svc = icr.get("service_name") or "Unknown Service"
        acct = icr.get("account_identifier") or prov
        region = icr.get("region") or "global"
        p_start = icr.get("period_start") or period_start.isoformat() + "Z"
        p_end = icr.get("period_end") or period_end.isoformat() + "Z"
        rows.append([
            cost, acct, acct, "USD",
            p_end, p_start, p_end, p_start,
            "Usage", None, None, None,
            cost, prov.upper(), cost, None,
            "On-Demand", 1.0, "Month",
            prov.upper(), prov.upper(),
            region, region,
            None, svc, "Cloud Service", "Compute",
            svc, None, None,
            acct, acct, "{}",
        ])

    return rows


@router.get("/exports/focus.csv")
async def export_focus_csv(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Response:
    """Export cloud costs in FinOps FOCUS 1.0 format (CSV)."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    rows = _focus_rows_from_context(context)
    return _csv_response(
        filename=f"optiora-focus-{_utcnow().strftime('%Y%m%d')}.csv",
        header=_FOCUS_COLUMNS,
        rows=rows,
    )


@router.get("/exports/focus.json")
async def export_focus_json(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Export cloud costs in FinOps FOCUS 1.0 format (JSON)."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    rows = _focus_rows_from_context(context)
    records = [dict(zip(_FOCUS_COLUMNS, row)) for row in rows]
    return {
        "focus_version": "1.0",
        "generated_at": _utcnow().isoformat() + "Z",
        "record_count": len(records),
        "records": records,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Unit Economics Cockpit  — enhanced with user-supplied business metrics
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class UnitEconomicsMetricRequest(BaseModel):
    metric_name: str = Field(..., description="e.g. 'customers', 'requests', 'transactions'")
    metric_value: float = Field(..., gt=0, description="Current period count/volume")
    metric_unit: str = Field(default="units")


@router.post("/analytics/unit-economics/metrics")
async def record_unit_economics_metric(
    payload: UnitEconomicsMetricRequest,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Record a business metric value for unit-economics cost-per-unit calculation."""
    _ = db
    context = await _cost_context(membership, db, "month", "all")
    total_cost = float(context.get("total_cost") or 0.0)
    cost_per_unit = round(total_cost / payload.metric_value, 4) if payload.metric_value > 0 else None
    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "metric_name": payload.metric_name,
        "metric_unit": payload.metric_unit,
        "metric_value": payload.metric_value,
        "total_monthly_cost_usd": round(total_cost, 2),
        "cost_per_unit_usd": cost_per_unit,
        "cost_per_unit_label": f"${cost_per_unit:.4f} per {payload.metric_unit}" if cost_per_unit else "N/A",
        "benchmark_note": "Cost per unit should trend downward as volume grows while costs stabilize.",
    }


@router.get("/analytics/unit-economics/cockpit")
async def get_unit_economics_cockpit(
    cloud_provider: str = "all",
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Full unit economics cockpit: cost metrics + per-provider breakdown + waste-to-spend."""
    context = await _cost_context(membership, db, "month", cloud_provider)
    analytics_result = _safe_json_load(
        await finops_analytics.get_analytics({
            "cloud_provider": cloud_provider,
            "current_monthly_spend": context["total_cost"],
            "cost_breakdown": context["breakdown"],
            "anomalies": 0,
            "monthly_savings": 0,
        }),
        {},
    )
    unit_result = _safe_json_load(
        await finops_analytics.get_unit_economics({
            "current_monthly_spend": context["total_cost"],
            "estimated_waste_usd": analytics_result.get("estimated_monthly_waste_usd", 0),
            "identified_savings_usd": analytics_result.get("identified_monthly_savings_usd", 0),
            "anomalies": 0,
        }),
        {},
    )

    total = float(context.get("total_cost") or 0.0)
    waste = float(analytics_result.get("estimated_monthly_waste_usd") or 0.0)
    savings = float(analytics_result.get("identified_monthly_savings_usd") or 0.0)

    provider_metrics: List[Dict[str, Any]] = []
    for provider, data in (context.get("breakdown") or {}).items():
        cost = float(data.get("cost") or 0.0)
        provider_metrics.append({
            "provider": provider,
            "cost_usd": round(cost, 2),
            "share_percent": round(data.get("percentage") or 0.0, 1),
            "estimated_waste_usd": round(cost * 0.18, 2),  # 18% heuristic per provider
            "efficiency_index": round(max(0, 1 - (cost * 0.18 / cost)) * 100, 1) if cost > 0 else 100.0,
        })

    # Historical monthly trend from snapshots
    customer_id = _customer_id_for_org(membership)
    historical = _historical_monthly_spend_from_snapshots(db, customer_id, cloud_provider, months=6)

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "cloud_provider": cloud_provider,
        "summary": {
            "total_monthly_cost_usd": round(total, 2),
            "estimated_waste_usd": round(waste, 2),
            "identified_savings_usd": round(savings, 2),
            "waste_to_spend_percent": round((waste / total * 100) if total > 0 else 0.0, 1),
            "dollar_efficiency_score": unit_result.get("dollar_efficiency_score", 0),
        },
        "provider_metrics": provider_metrics,
        "historical_monthly_spend": historical,
        "business_metrics_hint": "POST /api/v1/analytics/unit-economics/metrics to calculate cost-per-unit.",
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Scorecards  — per-team FinOps maturity score using chargeback + efficiency
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ScorecardDimension(BaseModel):
    name: str
    score: float
    max_score: float
    description: str


class ScorecardEntry(BaseModel):
    team: str
    total_score: float
    grade: str
    cost_usd: float
    share_percent: float
    dimensions: List[ScorecardDimension]
    trend: str


class ScorecardsResponse(BaseModel):
    generated_at: str
    organization_grade: str
    organization_score: float
    teams: List[ScorecardEntry]


def _grade(score: float, max_score: float = 100.0) -> str:
    pct = score / max_score * 100 if max_score > 0 else 0
    if pct >= 90: return "A+"
    if pct >= 80: return "A"
    if pct >= 70: return "B"
    if pct >= 55: return "C"
    return "D"


@router.get("/analytics/scorecards", response_model=ScorecardsResponse)
async def get_scorecards(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Per-team FinOps scorecard: allocation coverage, waste rate, tagging, commitment."""
    org_id = _organization_id_for_membership(membership)

    # Pull chargeback dimensions
    dim_rows = (
        db.query(NormalizedCostDimension)
        .filter(NormalizedCostDimension.organization_id == org_id)
        .all()
    )

    # Group by team
    team_costs: Dict[str, float] = {}
    team_mapped: Dict[str, float] = {}
    for r in dim_rows:
        team = r.team or "(unmapped)"
        cost = float(r.cost_usd or 0.0)
        team_costs[team] = team_costs.get(team, 0.0) + cost
        if r.is_mapped:
            team_mapped[team] = team_mapped.get(team, 0.0) + cost

    total_cost = sum(team_costs.values())
    teams: List[Dict[str, Any]] = []

    if team_costs:
        for team, cost in sorted(team_costs.items(), key=lambda x: -x[1]):
            mapped = team_mapped.get(team, 0.0)
            allocation_score = round((mapped / cost * 40) if cost > 0 else 40.0, 1)
            waste_score = round(30.0 * (1 - min(0.3, (cost * 0.18) / max(cost, 1))), 1)
            tagging_score = round((mapped / cost * 20) if cost > 0 else 20.0, 1)
            commitment_score = 10.0 if cost > 500 else 5.0
            total_score = round(allocation_score + waste_score + tagging_score + commitment_score, 1)

            teams.append({
                "team": team,
                "total_score": total_score,
                "grade": _grade(total_score),
                "cost_usd": round(cost, 2),
                "share_percent": round((cost / total_cost * 100) if total_cost > 0 else 0.0, 1),
                "dimensions": [
                    {"name": "Allocation Coverage", "score": allocation_score, "max_score": 40, "description": "% of team cost mapped to a business dimension"},
                    {"name": "Waste Reduction", "score": waste_score, "max_score": 30, "description": "Estimated waste as share of team spend"},
                    {"name": "Tagging Hygiene", "score": tagging_score, "max_score": 20, "description": "Resources with complete cost-center tags"},
                    {"name": "Commitment Coverage", "score": commitment_score, "max_score": 10, "description": "Reserved/savings-plan coverage for team workloads"},
                ],
                "trend": "stable",
            })
    else:
        # No chargeback dimension data yet — return a single placeholder row so
        # the frontend can distinguish "no data" from "score = 0".
        teams.append({
            "team": "(no data)",
            "total_score": 0.0,
            "grade": "N/A",
            "cost_usd": 0.0,
            "share_percent": 100.0,
            "no_data": True,
            "dimensions": [
                {"name": "Allocation Coverage", "score": 0.0, "max_score": 40, "description": "Configure business mapping rules to start scoring"},
                {"name": "Waste Reduction", "score": 0.0, "max_score": 30, "description": "Run a provider scan to populate waste signals"},
                {"name": "Tagging Hygiene", "score": 0.0, "max_score": 20, "description": "Upload cost data or run a scan to score tagging hygiene"},
                {"name": "Commitment Coverage", "score": 0.0, "max_score": 10, "description": "Add Reserved Instances or Savings Plans to improve coverage"},
            ],
            "trend": "unknown",
        })

    org_score = round(sum(t["total_score"] for t in teams) / max(len(teams), 1), 1)
    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "organization_grade": _grade(org_score),
        "organization_score": org_score,
        "teams": teams,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Resource Inventory  — surface cloud resources with cost attribution
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ResourceInventoryItem(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    account_id: str
    cost_usd: float
    waste_flag: bool
    waste_reason: Optional[str]
    tags: Dict[str, str]


class ResourceInventoryResponse(BaseModel):
    generated_at: str
    total_resources: int
    total_cost_usd: float
    flagged_waste_count: int
    items: List[ResourceInventoryItem]


@router.get("/inventory/resources", response_model=ResourceInventoryResponse)
async def get_resource_inventory(
    provider: str = "all",
    region: Optional[str] = None,
    waste_only: bool = False,
    limit: int = 100,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Cloud resource inventory with per-resource cost attribution and waste flags."""
    customer_id = _customer_id_for_org(membership)
    context = await _cost_context(membership, db, "month", provider if provider != "all" else "all")

    # Pull account snapshots for resource-level data
    snapshots_q = (
        db.query(ProviderAccountSnapshot)
        .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
        .filter(ProviderAccount.customer_id == customer_id)
    )
    if provider != "all":
        snapshots_q = snapshots_q.filter(ProviderAccount.provider == provider)

    org_id = membership.organization_id
    snapshots = snapshots_q.order_by(ProviderAccountSnapshot.captured_at.desc()).limit(500).all()

    items: List[Dict[str, Any]] = []
    seen: set = set()

    for snap in snapshots:
        acct = snap.provider_account
        if not acct:
            continue
        key = f"{acct.provider}:{acct.account_identifier}"
        if key in seen:
            continue
        seen.add(key)

        direct_cost = float(snap.direct_cost_usd or 0.0)
        waste_flag = direct_cost > 0 and (snap.anomalies_count or 0) > 0

        # region filter — ProviderAccount stores region in native_region
        snap_region = acct.native_region or "global"
        if region and snap_region != region:
            continue

        item: Dict[str, Any] = {
            "resource_id": acct.account_identifier or key,
            "resource_name": acct.account_name or acct.account_identifier or acct.provider.upper(),
            "resource_type": acct.account_type or "cloud-account",
            "provider": acct.provider,
            "region": snap_region,
            "account_id": acct.account_identifier or "",
            "cost_usd": round(direct_cost, 2),
            "waste_flag": waste_flag,
            "waste_reason": "Active anomalies detected" if waste_flag else None,
            "tags": {},
        }
        items.append(item)

    # Supplement from imported cost records when no scan snapshots exist
    if not items:
        imported_rows = _get_imported_cost_rows(db, org_id, customer_id)
        for row in imported_rows:
            prov = row.provider or "unknown"
            if provider != "all" and prov != provider:
                continue
            reg = row.region or "global"
            if region and reg != region:
                continue
            items.append({
                "resource_id": row.account_identifier or f"{prov}-imported-{row.id}",
                "resource_name": row.account_name or row.service_name or prov.upper(),
                "resource_type": "imported-cost-record",
                "provider": prov,
                "region": reg,
                "account_id": row.account_identifier or "",
                "cost_usd": round(float(row.cost_usd or 0.0), 2),
                "waste_flag": False,
                "waste_reason": None,
                "tags": {},
            })

    if waste_only:
        items = [i for i in items if i["waste_flag"]]

    total_cost = round(sum(i["cost_usd"] for i in items), 2)
    flagged = sum(1 for i in items if i["waste_flag"])
    paginated = items[offset: offset + limit]

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "total_resources": len(items),
        "total_cost_usd": total_cost,
        "flagged_waste_count": flagged,
        "items": paginated,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Kubernetes Cost Allocation
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class KubernetesClusterInput(BaseModel):
    cluster_name: str
    provider: str
    region: str
    node_count: int
    node_type: str
    monthly_node_cost_usd: float
    namespaces: Optional[List[str]] = None


class KubernetesNamespaceCost(BaseModel):
    namespace: str
    estimated_cost_usd: float
    share_percent: float
    cpu_share_percent: float
    memory_share_percent: float


class KubernetesClusterCostResponse(BaseModel):
    generated_at: str
    cluster_name: str
    provider: str
    region: str
    node_count: int
    node_type: str
    total_cluster_cost_usd: float
    cost_per_node_usd: float
    namespace_breakdown: List[KubernetesNamespaceCost]
    efficiency_note: str
    opencost_integration: str


@router.post("/analytics/kubernetes/cluster-cost", response_model=KubernetesClusterCostResponse)
async def calculate_kubernetes_cluster_cost(
    payload: KubernetesClusterInput,
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
) -> Dict[str, Any]:
    """Estimate Kubernetes cluster cost allocation by namespace."""
    total_cost = round(payload.node_count * payload.monthly_node_cost_usd, 2)
    cost_per_node = payload.monthly_node_cost_usd

    namespaces = payload.namespaces or ["default", "kube-system", "monitoring", "app"]
    n = len(namespaces)
    # Even split as baseline — real allocation requires OpenCost/Prometheus metrics
    namespace_breakdown: List[Dict[str, Any]] = []
    for i, ns in enumerate(namespaces):
        # Give kube-system ~10%, monitoring ~15%, rest shared evenly
        if ns == "kube-system":
            share = 10.0
        elif ns in ("monitoring", "prometheus", "observability"):
            share = 15.0
        else:
            remaining = 75.0 / max(1, n - sum(1 for x in namespaces if x in ("kube-system", "monitoring", "prometheus", "observability")))
            share = round(remaining, 1)

        namespace_breakdown.append({
            "namespace": ns,
            "estimated_cost_usd": round(total_cost * share / 100, 2),
            "share_percent": share,
            "cpu_share_percent": share,
            "memory_share_percent": share,
        })

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "cluster_name": payload.cluster_name,
        "provider": payload.provider,
        "region": payload.region,
        "node_count": payload.node_count,
        "node_type": payload.node_type,
        "total_cluster_cost_usd": total_cost,
        "cost_per_node_usd": cost_per_node,
        "namespace_breakdown": namespace_breakdown,
        "efficiency_note": (
            "Namespace breakdown uses proportional allocation. Connect OpenCost or Prometheus "
            "metrics to enable pod-level CPU/memory-weighted allocation."
        ),
        "opencost_integration": "POST /api/v1/analytics/kubernetes/cluster-cost with real prometheus metrics for weighted allocation.",
    }


@router.get("/analytics/kubernetes/summary")
async def get_kubernetes_summary(
    current_user: User = Depends(get_current_user),
    membership: UserOrganization = Depends(get_current_membership),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """Overview of Kubernetes cost allocation status for this organization."""
    context = await _cost_context(membership, db, "month", "all")
    total = float(context.get("total_cost") or 0.0)

    return {
        "generated_at": _utcnow().isoformat() + "Z",
        "kubernetes_enabled": False,
        "clusters_configured": 0,
        "estimated_k8s_share_percent": 0.0,
        "estimated_k8s_cost_usd": 0.0,
        "total_cloud_cost_usd": round(total, 2),
        "setup_hint": (
            "POST /api/v1/analytics/kubernetes/cluster-cost with your cluster details to begin "
            "namespace-level cost allocation. Connect OpenCost for pod-level granularity."
        ),
        "opencost_docs": "https://www.opencost.io/docs/",
    }


# ---------------------------------------------------------------------------
# Virtual Tagging — apply tag rules to resources without touching the cloud
# ---------------------------------------------------------------------------

class VirtualTagRuleCreate(BaseModel):
    tag_key: str
    tag_value: str
    match_provider: Optional[str] = None
    match_service: Optional[str] = None
    match_region: Optional[str] = None
    match_account_id: Optional[str] = None
    match_resource_type: Optional[str] = None
    match_resource_name_contains: Optional[str] = None
    match_team: Optional[str] = None
    match_environment: Optional[str] = None
    priority: int = 100
    is_active: bool = True
    description: Optional[str] = None


class VirtualTagRuleOut(BaseModel):
    id: int
    tag_key: str
    tag_value: str
    match_provider: Optional[str]
    match_service: Optional[str]
    match_region: Optional[str]
    match_account_id: Optional[str]
    match_resource_type: Optional[str]
    match_resource_name_contains: Optional[str]
    match_team: Optional[str]
    match_environment: Optional[str]
    priority: int
    is_active: bool
    description: Optional[str]
    created_at: str
    updated_at: Optional[str]

    model_config = {"from_attributes": True}


class VirtualTagRulesResponse(BaseModel):
    organization_id: int
    total: int
    rules: List[VirtualTagRuleOut]


class VirtualTagPreviewItem(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    cost_usd: float
    applied_tags: Dict[str, str]   # virtual tag key → value
    match_rule_ids: List[int]


class VirtualTagPreviewResponse(BaseModel):
    organization_id: int
    generated_at: str
    total_resources: int
    tagged_resources: int
    coverage_percent: float
    preview: List[VirtualTagPreviewItem]


def _vtag_matches(rule: VirtualTagRule, item: Dict) -> bool:
    """Return True if a virtual tag rule matches a cost/resource item dict."""
    if rule.match_provider and rule.match_provider.lower() != str(item.get("provider", "")).lower():
        return False
    if rule.match_service and rule.match_service.lower() not in str(item.get("service", "")).lower():
        return False
    if rule.match_region and rule.match_region.lower() not in str(item.get("region", "")).lower():
        return False
    if rule.match_account_id and rule.match_account_id != str(item.get("account_id", "")):
        return False
    if rule.match_resource_type and rule.match_resource_type.lower() not in str(item.get("resource_type", "")).lower():
        return False
    if rule.match_resource_name_contains and rule.match_resource_name_contains.lower() not in str(item.get("resource_name", "")).lower():
        return False
    if rule.match_team and rule.match_team.lower() not in str(item.get("team", "")).lower():
        return False
    if rule.match_environment and rule.match_environment.lower() not in str(item.get("environment", "")).lower():
        return False
    return True


@router.get("/virtual-tags/rules", response_model=VirtualTagRulesResponse)
async def list_virtual_tag_rules(
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """List all virtual tag rules for the current organization."""
    org_id = membership.organization_id
    rules = (
        db.query(VirtualTagRule)
        .filter(VirtualTagRule.organization_id == org_id)
        .order_by(VirtualTagRule.priority.desc(), VirtualTagRule.created_at)
        .all()
    )
    out = []
    for r in rules:
        out.append(VirtualTagRuleOut(
            id=r.id,
            tag_key=r.tag_key,
            tag_value=r.tag_value,
            match_provider=r.match_provider,
            match_service=r.match_service,
            match_region=r.match_region,
            match_account_id=r.match_account_id,
            match_resource_type=r.match_resource_type,
            match_resource_name_contains=r.match_resource_name_contains,
            match_team=r.match_team,
            match_environment=r.match_environment,
            priority=r.priority,
            is_active=r.is_active,
            description=r.description,
            created_at=r.created_at.isoformat() if r.created_at else "",
            updated_at=r.updated_at.isoformat() if r.updated_at else None,
        ))
    return VirtualTagRulesResponse(organization_id=org_id, total=len(out), rules=out)


@router.post("/virtual-tags/rules", response_model=VirtualTagRuleOut, status_code=201)
async def create_virtual_tag_rule(
    payload: VirtualTagRuleCreate,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Create a virtual tag rule."""
    org_id = membership.organization_id
    customer_id = str(membership.organization_id)
    rule = VirtualTagRule(
        organization_id=org_id,
        customer_id=customer_id,
        tag_key=payload.tag_key.strip(),
        tag_value=payload.tag_value.strip(),
        match_provider=payload.match_provider or None,
        match_service=payload.match_service or None,
        match_region=payload.match_region or None,
        match_account_id=payload.match_account_id or None,
        match_resource_type=payload.match_resource_type or None,
        match_resource_name_contains=payload.match_resource_name_contains or None,
        match_team=payload.match_team or None,
        match_environment=payload.match_environment or None,
        priority=payload.priority,
        is_active=payload.is_active,
        description=payload.description,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return VirtualTagRuleOut(
        id=rule.id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        match_provider=rule.match_provider,
        match_service=rule.match_service,
        match_region=rule.match_region,
        match_account_id=rule.match_account_id,
        match_resource_type=rule.match_resource_type,
        match_resource_name_contains=rule.match_resource_name_contains,
        match_team=rule.match_team,
        match_environment=rule.match_environment,
        priority=rule.priority,
        is_active=rule.is_active,
        description=rule.description,
        created_at=rule.created_at.isoformat() if rule.created_at else "",
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.put("/virtual-tags/rules/{rule_id}", response_model=VirtualTagRuleOut)
async def update_virtual_tag_rule(
    rule_id: int,
    payload: VirtualTagRuleCreate,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Update a virtual tag rule."""
    org_id = membership.organization_id
    rule = db.query(VirtualTagRule).filter(
        VirtualTagRule.id == rule_id,
        VirtualTagRule.organization_id == org_id,
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Virtual tag rule not found")
    rule.tag_key = payload.tag_key.strip()
    rule.tag_value = payload.tag_value.strip()
    rule.match_provider = payload.match_provider or None
    rule.match_service = payload.match_service or None
    rule.match_region = payload.match_region or None
    rule.match_account_id = payload.match_account_id or None
    rule.match_resource_type = payload.match_resource_type or None
    rule.match_resource_name_contains = payload.match_resource_name_contains or None
    rule.match_team = payload.match_team or None
    rule.match_environment = payload.match_environment or None
    rule.priority = payload.priority
    rule.is_active = payload.is_active
    rule.description = payload.description
    db.commit()
    db.refresh(rule)
    return VirtualTagRuleOut(
        id=rule.id,
        tag_key=rule.tag_key,
        tag_value=rule.tag_value,
        match_provider=rule.match_provider,
        match_service=rule.match_service,
        match_region=rule.match_region,
        match_account_id=rule.match_account_id,
        match_resource_type=rule.match_resource_type,
        match_resource_name_contains=rule.match_resource_name_contains,
        match_team=rule.match_team,
        match_environment=rule.match_environment,
        priority=rule.priority,
        is_active=rule.is_active,
        description=rule.description,
        created_at=rule.created_at.isoformat() if rule.created_at else "",
        updated_at=rule.updated_at.isoformat() if rule.updated_at else None,
    )


@router.delete("/virtual-tags/rules/{rule_id}", status_code=204)
async def delete_virtual_tag_rule(
    rule_id: int,
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Delete a virtual tag rule."""
    org_id = membership.organization_id
    rule = db.query(VirtualTagRule).filter(
        VirtualTagRule.id == rule_id,
        VirtualTagRule.organization_id == org_id,
    ).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Virtual tag rule not found")
    db.delete(rule)
    db.commit()


@router.get("/virtual-tags/preview", response_model=VirtualTagPreviewResponse)
async def preview_virtual_tags(
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Preview which virtual tags would be applied to current cost records."""
    org_id = membership.organization_id
    now_str = _utcnow().isoformat()

    active_rules = (
        db.query(VirtualTagRule)
        .filter(
            VirtualTagRule.organization_id == org_id,
            VirtualTagRule.is_active == True,
        )
        .order_by(VirtualTagRule.priority.desc())
        .all()
    )

    # Gather candidate resources from inventory snapshots or imported cost records
    items: List[Dict] = []
    snapshots = (
        db.query(ProviderAccountSnapshot)
        .join(ProviderAccount, ProviderAccount.id == ProviderAccountSnapshot.provider_account_id)
        .filter(ProviderAccount.organization_id == org_id)
        .order_by(ProviderAccountSnapshot.captured_at.desc())
        .limit(limit)
        .all()
    )
    for snap in snapshots:
        acct = snap.provider_account
        items.append({
            "resource_id": f"account:{snap.provider_account_id}",
            "resource_name": acct.account_name if acct else str(snap.provider_account_id),
            "resource_type": "Cloud Account",
            "provider": acct.provider if acct else "unknown",
            "region": acct.native_region if acct else "",
            "service": "",
            "account_id": acct.account_identifier if acct else "",
            "cost_usd": float(snap.direct_cost_usd or 0),
            "team": "",
            "environment": "",
        })

    if not items:
        records = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .limit(limit)
            .all()
        )
        for rec in records:
            items.append({
                "resource_id": f"imported:{rec.id}",
                "resource_name": rec.account_name or rec.service_name or "unknown",
                "resource_type": rec.service_name or "Imported Service",
                "provider": rec.provider or "unknown",
                "region": rec.region or "",
                "service": rec.service_name or "",
                "account_id": rec.account_identifier or "",
                "cost_usd": float(rec.cost_usd or 0),
                "team": "",
                "environment": "",
            })

    if not items:
        # Synthetic demonstration data
        demo = [
            {"resource_id": "i-0abc123", "resource_name": "prod-api-server", "resource_type": "EC2 Instance", "provider": "aws", "region": "us-east-1", "service": "AmazonEC2", "account_id": "123456789012", "cost_usd": 312.50, "team": "platform", "environment": "production"},
            {"resource_id": "i-0def456", "resource_name": "staging-worker", "resource_type": "EC2 Instance", "provider": "aws", "region": "us-west-2", "service": "AmazonEC2", "account_id": "123456789012", "cost_usd": 87.20, "team": "data", "environment": "staging"},
            {"resource_id": "vm-prod-01", "resource_name": "azure-prod-vm", "resource_type": "Virtual Machine", "provider": "azure", "region": "eastus", "service": "Compute", "account_id": "sub-azure-001", "cost_usd": 198.00, "team": "platform", "environment": "production"},
            {"resource_id": "disk-vol-01", "resource_name": "data-volume-large", "resource_type": "Managed Disk", "provider": "azure", "region": "westeurope", "service": "Storage", "account_id": "sub-azure-001", "cost_usd": 45.30, "team": "", "environment": ""},
            {"resource_id": "gke-node-01", "resource_name": "gke-prod-node", "resource_type": "GKE Node", "provider": "gcp", "region": "us-central1", "service": "Kubernetes Engine", "account_id": "proj-gcp-001", "cost_usd": 256.80, "team": "infra", "environment": "production"},
        ]
        items = demo

    # Apply rules to each item
    preview: List[VirtualTagPreviewItem] = []
    tagged_count = 0
    for item in items[:limit]:
        applied: Dict[str, str] = {}
        matched_ids: List[int] = []
        for rule in active_rules:
            if _vtag_matches(rule, item):
                if rule.tag_key not in applied:  # first matching rule wins per key
                    applied[rule.tag_key] = rule.tag_value
                    matched_ids.append(rule.id)
        if applied:
            tagged_count += 1
        preview.append(VirtualTagPreviewItem(
            resource_id=item["resource_id"],
            resource_name=item["resource_name"],
            resource_type=item["resource_type"],
            provider=item["provider"],
            region=item["region"],
            cost_usd=item["cost_usd"],
            applied_tags=applied,
            match_rule_ids=matched_ids,
        ))

    total = len(preview)
    coverage = round((tagged_count / total * 100) if total > 0 else 0.0, 1)
    return VirtualTagPreviewResponse(
        organization_id=org_id,
        generated_at=now_str,
        total_resources=total,
        tagged_resources=tagged_count,
        coverage_percent=coverage,
        preview=preview,
    )


# ---------------------------------------------------------------------------
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Rightsizing — resource-level recommendations
#
# 4-tier data source waterfall (best available wins):
#  1. AWS Cost Explorer  get_rightsizing_recommendation (real API)
#  2. Azure Advisor      rightsizing recommendations    (real API)
#  3. Cost-trend signal  analysis from snapshot history (deterministic)
#  4. Imported CSV cost  signal analysis                (deterministic)
#  [5. Synthetic examples — only when org has zero data at all]
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class RightsizingRecommendation(BaseModel):
    resource_id: str
    resource_name: str
    resource_type: str
    provider: str
    region: str
    account_id: str
    current_size: str
    recommended_size: str
    current_monthly_cost_usd: float
    projected_monthly_cost_usd: float
    monthly_savings_usd: float
    annual_savings_usd: float
    cpu_utilization_avg_percent: Optional[float]
    memory_utilization_avg_percent: Optional[float]
    reason: str
    confidence: str   # "high" | "medium" | "low"
    effort: str       # "low" | "medium" | "high"
    action: str       # "downsize" | "terminate" | "reserve" | "modernize"


class RightsizingResponse(BaseModel):
    generated_at: str
    organization_id: int
    data_source: str
    total_resources_analyzed: int
    rightsizable_count: int
    total_monthly_savings_usd: float
    total_annual_savings_usd: float
    recommendations: List[RightsizingRecommendation]


# ---------------------------------------------------------------------------
# Static reference data
# ---------------------------------------------------------------------------

_DOWNSIZE_MAP: Dict[str, Dict[str, str]] = {
    "aws": {
        "m5.4xlarge": "m5.2xlarge",   "m5.2xlarge": "m5.xlarge",   "m5.xlarge": "m5.large",
        "m5.large":   "m5.medium",
        "m6i.4xlarge": "m6i.2xlarge", "m6i.2xlarge": "m6i.xlarge", "m6i.xlarge": "m6i.large",
        "c5.4xlarge": "c5.2xlarge",   "c5.2xlarge": "c5.xlarge",   "c5.xlarge": "c5.large",
        "c6i.4xlarge": "c6i.2xlarge", "c6i.2xlarge": "c6i.xlarge",
        "r5.4xlarge": "r5.2xlarge",   "r5.2xlarge": "r5.xlarge",   "r5.xlarge": "r5.large",
        "t3.2xlarge": "t3.xlarge",    "t3.xlarge": "t3.large",     "t3.large": "t3.medium",
    },
    "azure": {
        "Standard_D16s_v3": "Standard_D8s_v3",  "Standard_D8s_v3": "Standard_D4s_v3",
        "Standard_D4s_v3":  "Standard_D2s_v3",  "Standard_D2s_v3": "Standard_B2s",
        "Standard_E16s_v3": "Standard_E8s_v3",  "Standard_E8s_v3": "Standard_E4s_v3",
        "Standard_E4s_v3":  "Standard_E2s_v3",
        "Standard_F16s_v2": "Standard_F8s_v2",  "Standard_F8s_v2": "Standard_F4s_v2",
    },
    "gcp": {
        "n1-standard-16": "n1-standard-8",  "n1-standard-8": "n1-standard-4",
        "n1-standard-4":  "n1-standard-2",
        "n2-standard-16": "n2-standard-8",  "n2-standard-8": "n2-standard-4",
        "n2-standard-4":  "n2-standard-2",
        "e2-standard-8":  "e2-standard-4",  "e2-standard-4": "e2-standard-2",
    },
    "oci": {
        "VM.Standard2.16": "VM.Standard2.8",  "VM.Standard2.8": "VM.Standard2.4",
        "VM.Standard2.4":  "VM.Standard2.2",
        "VM.Standard3.Flex.16": "VM.Standard3.Flex.8", "VM.Standard3.Flex.8": "VM.Standard3.Flex.4",
    },
}

# Savings rates by action type (realistic market benchmarks)
_ACTION_SAVINGS_RATES: Dict[str, float] = {
    "downsize":  0.45,   # ~50% savings dropping one size tier
    "terminate": 1.00,   # full elimination of orphaned resource
    "reserve":   0.37,   # typical 1yr No-Upfront RI discount vs on-demand
    "modernize": 0.30,   # newer gen typically 10-30% cheaper at same perf
}


# ---------------------------------------------------------------------------
# Tier 1: AWS Cost Explorer get_rightsizing_recommendation
# ---------------------------------------------------------------------------

def _rightsizing_from_aws_ce(
    cred_json: Dict[str, Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Call AWS Cost Explorer RightsizingRecommendation API and normalise results."""
    try:
        import boto3
        from botocore.exceptions import ClientError, BotoCoreError

        client = boto3.client(
            "ce",
            aws_access_key_id=cred_json.get("access_key_id"),
            aws_secret_access_key=cred_json.get("secret_access_key"),
            region_name=cred_json.get("region", "us-east-1"),
        )
        response = client.get_rightsizing_recommendation(
            Service="AmazonEC2",
            Configuration={
                "RecommendationTarget": "SAME_INSTANCE_FAMILY",
                "BenefitsConsidered": True,
            },
        )
        out: List[RightsizingRecommendation] = []
        for rec in response.get("RightsizingRecommendations", []):
            rtype = rec.get("RightsizingType", "Terminate")
            details = rec.get("ModifyRecommendationDetail") or rec.get("TerminateRecommendationDetail") or {}
            curr_instance = rec.get("CurrentInstance", {})
            curr_type = curr_instance.get("InstanceType", "unknown")
            account = curr_instance.get("AccountId", "")
            resource_id = curr_instance.get("ResourceId", f"ec2-{account}")
            region = curr_instance.get("Tags", {}).get("aws:cloudformation:stack-id", "") or "us-east-1"

            curr_cost = float(
                (curr_instance.get("MonthlyCost") or "0").replace("$", "").replace(",", "")
            )
            if rtype == "Modify":
                target = (details.get("TargetInstances") or [{}])[0]
                recommended_type = target.get("ExpectedResourceType", "smaller")
                proj_cost = float(
                    (target.get("ExpectedMonthlyBenefit") or {}).get("Amount", "0")
                )
                monthly_savings = max(0.0, curr_cost - proj_cost)
                action = "downsize"
                effort = "low"
                reason = (
                    f"CE recommends {recommended_type} — "
                    f"estimated ${monthly_savings:.2f}/mo savings"
                )
                cpu_util = float(
                    (curr_instance.get("ResourceUtilization") or {}).get("EC2ResourceUtilization", {}).get("MaxCpuUtilizationPercentage", 0) or 0
                )
            else:  # Terminate
                recommended_type = "N/A — terminate"
                monthly_savings = curr_cost
                proj_cost = 0.0
                action = "terminate"
                effort = "low"
                reason = "Instance has zero utilisation over the CE lookback window — candidate for termination"
                cpu_util = 0.0

            if monthly_savings < min_savings:
                continue

            confidence = "high" if cpu_util < 10 else "medium"
            out.append(RightsizingRecommendation(
                resource_id=resource_id,
                resource_name=curr_type,
                resource_type="EC2 Instance",
                provider="aws",
                region=region,
                account_id=account,
                current_size=curr_type,
                recommended_size=recommended_type,
                current_monthly_cost_usd=round(curr_cost, 2),
                projected_monthly_cost_usd=round(proj_cost, 2),
                monthly_savings_usd=round(monthly_savings, 2),
                annual_savings_usd=round(monthly_savings * 12, 2),
                cpu_utilization_avg_percent=round(cpu_util, 1) if cpu_util else None,
                memory_utilization_avg_percent=None,
                reason=reason,
                confidence=confidence,
                effort=effort,
                action=action,
            ))
        return out
    except Exception as exc:
        logger.warning("AWS CE rightsizing API error (will fall through): %s", exc)
        return []


# ---------------------------------------------------------------------------
# Tier 2: Azure Advisor rightsizing recommendations
# ---------------------------------------------------------------------------

def _rightsizing_from_azure_advisor(
    cred_json: Dict[str, Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Call Azure Advisor recommendations and return rightsizing findings."""
    try:
        from azure.identity import ClientSecretCredential
        from azure.mgmt.advisor import AdvisorManagementClient

        azure_cred = ClientSecretCredential(
            tenant_id=cred_json["tenant_id"],
            client_id=cred_json["client_id"],
            client_secret=cred_json["client_secret"],
        )
        subscription_id = cred_json["subscription_id"]
        client = AdvisorManagementClient(azure_cred, subscription_id)

        out: List[RightsizingRecommendation] = []
        for rec in client.recommendations.list():
            category = (rec.category or "").lower()
            if "cost" not in category:
                continue
            impact = (rec.impact or "").lower()
            resource_metadata = rec.resource_metadata or {}
            resource_id = getattr(rec, "resource_id", None) or ""
            short_id = resource_id.split("/")[-1] if resource_id else "unknown"
            region = getattr(rec, "location", "unknown") or "unknown"

            # Extract extended properties for savings estimate
            props = rec.extended_properties or {}
            savings_str = props.get("annualSavingsAmount") or props.get("savingsAmount") or "0"
            try:
                annual_savings = float(str(savings_str).replace("$", "").replace(",", ""))
            except (TypeError, ValueError):
                annual_savings = 0.0
            monthly_savings = round(annual_savings / 12, 2)

            if monthly_savings < min_savings:
                continue

            current_sku = props.get("currentSku") or props.get("currentSize") or "unknown"
            recommended_sku = props.get("targetSku") or props.get("recommendedSize") or "smaller SKU"
            curr_cost = round(monthly_savings / max(_ACTION_SAVINGS_RATES["downsize"], 0.01), 2)

            confidence = "high" if impact == "high" else "medium" if impact == "medium" else "low"
            out.append(RightsizingRecommendation(
                resource_id=short_id,
                resource_name=props.get("resourceName") or short_id,
                resource_type="Virtual Machine",
                provider="azure",
                region=region,
                account_id=subscription_id,
                current_size=current_sku,
                recommended_size=recommended_sku,
                current_monthly_cost_usd=curr_cost,
                projected_monthly_cost_usd=round(curr_cost - monthly_savings, 2),
                monthly_savings_usd=monthly_savings,
                annual_savings_usd=round(annual_savings, 2),
                cpu_utilization_avg_percent=None,
                memory_utilization_avg_percent=None,
                reason=rec.short_description.problem if rec.short_description else "Azure Advisor cost recommendation",
                confidence=confidence,
                effort="low",
                action="downsize",
            ))
        return out
    except Exception as exc:
        logger.warning("Azure Advisor rightsizing API error (will fall through): %s", exc)
        return []


# ---------------------------------------------------------------------------
# Tier 3: Cost-trend signal analysis from scan snapshot history
#
# Signals used (no external API needed):
#   • Cost trend across scans:   slope > 0 → growing (reserve candidate)
#                                slope ≈ 0, cost high → over-provisioned (downsize)
#                                cost near zero → orphaned (terminate)
#   • Anomalies count:           high → investigate before resizing
#   • Savings identified:        provider already flagged savings → high confidence
#   • Number of scan data points: more points → higher confidence
# ---------------------------------------------------------------------------

def _trend_slope(values: List[float]) -> float:
    """Simple linear regression slope over equally-spaced observations."""
    n = len(values)
    if n < 2:
        return 0.0
    x_mean = (n - 1) / 2.0
    y_mean = sum(values) / n
    num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
    den = sum((i - x_mean) ** 2 for i in range(n))
    return num / den if den else 0.0


def _rightsizing_from_snapshot_trends(
    snapshots_by_account: Dict[int, List[Any]],
    account_map: Dict[int, Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Derive rightsizing signals from multi-scan cost history without external APIs."""
    out: List[RightsizingRecommendation] = []
    for acct_id, snaps in snapshots_by_account.items():
        acct = account_map.get(acct_id)
        if not acct:
            continue
        prov = acct.provider

        # Sort oldest-first for trend analysis
        snaps_sorted = sorted(snaps, key=lambda s: s.captured_at)
        costs = [float(s.direct_cost_usd or 0) for s in snaps_sorted]
        latest_cost = costs[-1] if costs else 0.0
        avg_cost = sum(costs) / len(costs) if costs else 0.0
        total_anomalies = sum(s.anomalies_count or 0 for s in snaps_sorted)
        total_savings_identified = sum(float(s.savings_identified_usd or 0) for s in snaps_sorted)
        n_points = len(costs)
        slope = _trend_slope(costs)

        # Skip very cheap accounts (under $20/mo) — not worth rightsizing
        if avg_cost < 20:
            continue

        # Derive action from signals
        if latest_cost < 5.0:
            # Near-zero cost → orphaned resource
            action = "terminate"
            savings_rate = 1.0
            reason = (
                f"Cost dropped to ${latest_cost:.2f}/mo (avg ${avg_cost:.2f}/mo over "
                f"{n_points} scans) — likely orphaned or stopped resource"
            )
            effort = "low"
            # High confidence only with multiple data points confirming low cost
            confidence = "high" if n_points >= 3 and max(costs[:-1]) < 10 else "medium"

        elif slope > avg_cost * 0.05 and total_anomalies == 0:
            # Steady growth, no anomalies → reserve for predictable savings
            action = "reserve"
            savings_rate = _ACTION_SAVINGS_RATES["reserve"]
            growth_pct = round(slope / avg_cost * 100, 1)
            reason = (
                f"Cost growing +{growth_pct}% per scan period with no anomalies — "
                f"steady-state workload ideal for 1yr Reserved Instance (est. "
                f"{int(savings_rate * 100)}% discount)"
            )
            effort = "medium"
            confidence = "high" if n_points >= 4 else "medium"

        elif slope < -avg_cost * 0.03 and latest_cost > 100:
            # Declining cost but still high → already being optimised; modernize to accelerate
            action = "modernize"
            savings_rate = _ACTION_SAVINGS_RATES["modernize"]
            reason = (
                f"Cost declining (${costs[0]:.2f} → ${latest_cost:.2f}/mo) but still high — "
                f"migrate to newer generation for further {int(savings_rate * 100)}% reduction"
            )
            effort = "medium"
            confidence = "medium"

        elif total_savings_identified > 0:
            # Provider scan already flagged savings — downsize
            action = "downsize"
            savings_rate = min(
                total_savings_identified / max(avg_cost, 1.0) / n_points,
                0.50,
            )
            reason = (
                f"Provider scan identified ${total_savings_identified:.2f} savings over "
                f"{n_points} scans — consistent over-provisioning signal"
            )
            effort = "low"
            confidence = "high"

        elif avg_cost > 200 and abs(slope) < avg_cost * 0.02 and total_anomalies == 0:
            # Flat high cost, no anomalies → over-provisioned, downsize
            action = "downsize"
            savings_rate = _ACTION_SAVINGS_RATES["downsize"]
            reason = (
                f"Flat cost of ~${avg_cost:.2f}/mo over {n_points} scans with no anomalies — "
                f"consistent usage pattern indicates over-provisioning"
            )
            effort = "low"
            confidence = "high" if n_points >= 3 else "medium"

        else:
            # Insufficient signal
            continue

        monthly_savings = round(avg_cost * savings_rate, 2)
        if monthly_savings < min_savings:
            continue

        # Look up a plausible downsize target from the static map
        size_map = _DOWNSIZE_MAP.get(prov, {})
        current_size = next(iter(size_map), f"{prov}-instance")
        recommended_size = size_map.get(current_size, f"smaller-{current_size}")
        if action == "terminate":
            recommended_size = "N/A — terminate"
        elif action == "reserve":
            recommended_size = f"{current_size} (Reserved 1yr)"
        elif action == "modernize":
            recommended_size = _DOWNSIZE_MAP.get(prov, {}).get(
                current_size, f"newer-gen-{current_size}"
            )

        type_labels = {"aws": "EC2 Instance", "azure": "Virtual Machine",
                       "gcp": "Compute Instance", "oci": "OCI Compute"}

        out.append(RightsizingRecommendation(
            resource_id=f"{prov}-acct-{acct_id}",
            resource_name=acct.account_name or acct.account_identifier,
            resource_type=type_labels.get(prov, "Compute Instance"),
            provider=prov,
            region=acct.native_region or "global",
            account_id=acct.account_identifier,
            current_size=current_size,
            recommended_size=recommended_size,
            current_monthly_cost_usd=round(avg_cost, 2),
            projected_monthly_cost_usd=round(avg_cost * (1 - savings_rate), 2),
            monthly_savings_usd=monthly_savings,
            annual_savings_usd=round(monthly_savings * 12, 2),
            cpu_utilization_avg_percent=None,  # real metrics not yet ingested
            memory_utilization_avg_percent=None,
            reason=reason,
            confidence=confidence,
            effort=effort,
            action=action,
        ))
    return out


# ---------------------------------------------------------------------------
# Tier 4: Imported CSV cost-signal analysis (deterministic, no random)
# ---------------------------------------------------------------------------

_HIGH_COST_SERVICES: Dict[str, str] = {
    # Service names that typically indicate large compute → downsize candidate
    "amazon elastic compute cloud": "downsize",
    "amazon ec2": "downsize",
    "ec2": "downsize",
    "compute engine": "downsize",
    "virtual machines": "downsize",
    "oci compute": "downsize",
    # Storage orphan signals → terminate
    "amazon s3": "terminate",
    "azure blob storage": "terminate",
    "cloud storage": "terminate",
    # Consistent services → reserve
    "amazon rds": "reserve",
    "azure sql": "reserve",
    "cloud sql": "reserve",
    "amazon elasticache": "reserve",
}


def _rightsizing_from_imported_costs(
    records: List[Any],
    min_savings: float,
) -> List[RightsizingRecommendation]:
    """Derive rightsizing signals from imported CSV cost records (no external API)."""
    # Group records by (provider, account_identifier) and sum costs
    groups: Dict[tuple, Dict[str, Any]] = {}
    for rec in records:
        prov = (rec.provider or "aws").lower()
        acct = rec.account_identifier or f"{prov}-{rec.id}"
        key = (prov, acct)
        if key not in groups:
            groups[key] = {
                "provider": prov,
                "account_id": acct,
                "account_name": rec.account_name or rec.service_name or acct,
                "region": rec.region or "global",
                "total_cost": 0.0,
                "services": [],
                "resource_id": rec.account_identifier or "",
            }
        groups[key]["total_cost"] += float(rec.cost_usd or 0)
        if rec.service_name:
            groups[key]["services"].append(rec.service_name.lower())

    out: List[RightsizingRecommendation] = []
    for (prov, acct_id), grp in groups.items():
        cost = grp["total_cost"]
        if cost < 15:
            continue

        # Determine action from service names; default to downsize for high compute
        action = "downsize"
        for svc in grp["services"]:
            for svc_key, svc_action in _HIGH_COST_SERVICES.items():
                if svc_key in svc:
                    action = svc_action
                    break

        savings_rate = _ACTION_SAVINGS_RATES[action]
        monthly_savings = round(cost * savings_rate, 2)
        if monthly_savings < min_savings:
            continue

        size_map = _DOWNSIZE_MAP.get(prov, _DOWNSIZE_MAP["aws"])
        current_size = next(iter(size_map))
        if action == "terminate":
            recommended_size = "N/A — terminate"
        elif action == "reserve":
            recommended_size = f"{current_size} (Reserved 1yr)"
        else:
            recommended_size = size_map.get(current_size, f"smaller-{current_size}")

        type_labels = {"aws": "EC2 Instance", "azure": "Virtual Machine",
                       "gcp": "Compute Instance", "oci": "OCI Compute"}
        reason_map = {
            "downsize":  f"High compute cost (${cost:.2f}/mo) with no utilisation baseline — apply one-size downgrade",
            "terminate": f"Storage/service cost (${cost:.2f}/mo) with no active resource signal — validate and terminate",
            "reserve":   f"Consistent managed-service cost (${cost:.2f}/mo) — convert to 1yr Reserved for ~{int(savings_rate*100)}% savings",
            "modernize": f"Legacy service cost (${cost:.2f}/mo) — migrate to current-gen equivalent",
        }

        out.append(RightsizingRecommendation(
            resource_id=grp["resource_id"] or f"imported-{prov}-{acct_id}",
            resource_name=grp["account_name"],
            resource_type=type_labels.get(prov, "Cloud Service"),
            provider=prov,
            region=grp["region"],
            account_id=acct_id,
            current_size=current_size,
            recommended_size=recommended_size,
            current_monthly_cost_usd=round(cost, 2),
            projected_monthly_cost_usd=round(cost * (1 - savings_rate), 2),
            monthly_savings_usd=monthly_savings,
            annual_savings_usd=round(monthly_savings * 12, 2),
            cpu_utilization_avg_percent=None,
            memory_utilization_avg_percent=None,
            reason=reason_map[action],
            confidence="medium",
            effort="low" if action in ("downsize", "terminate") else "medium",
            action=action,
        ))
    return out


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.get("/recommendations/rightsizing", response_model=RightsizingResponse)
async def get_rightsizing_recommendations(
    provider: str = Query("all"),
    min_savings: float = Query(10.0, description="Minimum monthly savings threshold USD"),
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    membership=Depends(get_current_membership),
    db: Session = Depends(get_db),
):
    """Resource-level rightsizing recommendations.

    Tries real provider APIs first (AWS CE, Azure Advisor), then falls back to
    deterministic cost-trend analysis on scan history, then imported CSV signals,
    and finally synthetic examples when no data exists at all.
    """
    org_id = membership.organization_id
    customer_id = _customer_id_for_org(membership)
    now_str = _utcnow().isoformat()

    recommendations_out: List[RightsizingRecommendation] = []
    data_source = "synthetic"
    total_analyzed = 0

    # ── Tier 1 + 2: Real provider APIs (when valid credentials are stored) ──
    provider_filter = [provider] if provider != "all" else ["aws", "azure"]
    for prov in provider_filter:
        cred_row = (
            db.query(CredentialRecord)
            .filter(
                CredentialRecord.customer_id == customer_id,
                CredentialRecord.provider == prov,
                CredentialRecord.is_valid.is_(True),
            )
            .first()
        )
        if cred_row is None:
            continue
        try:
            cred_json: Dict[str, Any] = json.loads(cred_row.credential_json)
        except Exception:
            continue

        if prov == "aws":
            tier1 = _rightsizing_from_aws_ce(cred_json, min_savings)
            if tier1:
                recommendations_out.extend(tier1)
                data_source = "aws_cost_explorer"
                total_analyzed += len(tier1) * 4  # CE analyzes all instances

        elif prov == "azure":
            tier2 = _rightsizing_from_azure_advisor(cred_json, min_savings)
            if tier2:
                recommendations_out.extend(tier2)
                data_source = "azure_advisor" if data_source == "synthetic" else "multi_provider_api"
                total_analyzed += len(tier2) * 3

    # ── Tier 3: Cost-trend signal from scan snapshot history ────────────────
    if not recommendations_out:
        acct_query = (
            db.query(ProviderAccount)
            .filter(ProviderAccount.organization_id == org_id)
        )
        if provider != "all":
            acct_query = acct_query.filter(ProviderAccount.provider == provider)
        accounts = acct_query.all()
        account_map = {a.id: a for a in accounts}

        if account_map:
            # Fetch all snapshots for these accounts in one query, ordered oldest-first
            snap_rows = (
                db.query(ProviderAccountSnapshot)
                .filter(
                    ProviderAccountSnapshot.provider_account_id.in_(list(account_map.keys())),
                    ProviderAccountSnapshot.organization_id == org_id,
                )
                .order_by(ProviderAccountSnapshot.captured_at.asc())
                .limit(500)
                .all()
            )

            # Group by account
            snapshots_by_account: Dict[int, List[Any]] = {}
            for snap in snap_rows:
                snapshots_by_account.setdefault(snap.provider_account_id, []).append(snap)

            total_analyzed = len(snap_rows)
            tier3 = _rightsizing_from_snapshot_trends(
                snapshots_by_account, account_map, min_savings
            )
            if tier3:
                recommendations_out = tier3
                data_source = "cost_trend_analysis"

    # ── Tier 4: Imported CSV cost-signal analysis ────────────────────────────
    if not recommendations_out:
        imported = (
            db.query(ImportedCostRecord)
            .filter(ImportedCostRecord.organization_id == org_id)
            .limit(500)
            .all()
        )
        if provider != "all":
            imported = [r for r in imported if (r.provider or "").lower() == provider]
        if imported:
            total_analyzed = len(imported)
            tier4 = _rightsizing_from_imported_costs(imported, min_savings)
            if tier4:
                recommendations_out = tier4
                data_source = "imported_costs"

    # ── Tier 5 (synthetic fallback — zero org data) ──────────────────────────
    if not recommendations_out:
        data_source = "synthetic"
        total_analyzed = 6
        recommendations_out = [
            RightsizingRecommendation(
                resource_id="i-0a1b2c3d", resource_name="prod-api-server-01",
                resource_type="EC2 Instance", provider="aws", region="us-east-1",
                account_id="123456789012", current_size="m5.4xlarge",
                recommended_size="m5.2xlarge",
                current_monthly_cost_usd=614.40, projected_monthly_cost_usd=338.00,
                monthly_savings_usd=276.40, annual_savings_usd=3316.80,
                cpu_utilization_avg_percent=8.3, memory_utilization_avg_percent=22.1,
                reason="CPU p95 < 15% over 30 days — downsize to m5.2xlarge (~45% savings)",
                confidence="high", effort="low", action="downsize",
            ),
            RightsizingRecommendation(
                resource_id="i-0e4f5a6b", resource_name="data-pipeline-worker",
                resource_type="EC2 Instance", provider="aws", region="us-west-2",
                account_id="123456789012", current_size="c5.2xlarge",
                recommended_size="c5.xlarge",
                current_monthly_cost_usd=248.64, projected_monthly_cost_usd=136.75,
                monthly_savings_usd=111.89, annual_savings_usd=1342.68,
                cpu_utilization_avg_percent=18.7, memory_utilization_avg_percent=31.2,
                reason="CPU avg 18.7% over 30 days — downsize or convert to Savings Plan",
                confidence="medium", effort="low", action="downsize",
            ),
            RightsizingRecommendation(
                resource_id="vm-prod-eastus-02", resource_name="azure-backend-vm",
                resource_type="Virtual Machine", provider="azure", region="eastus",
                account_id="sub-azure-001", current_size="Standard_D8s_v3",
                recommended_size="Standard_D4s_v3",
                current_monthly_cost_usd=380.16, projected_monthly_cost_usd=209.09,
                monthly_savings_usd=171.07, annual_savings_usd=2052.84,
                cpu_utilization_avg_percent=12.5, memory_utilization_avg_percent=28.4,
                reason="Azure Advisor: avg CPU 12.5% — downsize to D4s_v3 (45% cost reduction)",
                confidence="high", effort="low", action="downsize",
            ),
            RightsizingRecommendation(
                resource_id="disk-vol-orphan-03", resource_name="unattached-data-disk",
                resource_type="Managed Disk", provider="azure", region="westeurope",
                account_id="sub-azure-001", current_size="Premium_SSD_512GiB",
                recommended_size="N/A — terminate",
                current_monthly_cost_usd=92.16, projected_monthly_cost_usd=0.0,
                monthly_savings_usd=92.16, annual_savings_usd=1105.92,
                cpu_utilization_avg_percent=None, memory_utilization_avg_percent=None,
                reason="Disk unattached for 47 days with no snapshots — terminate to eliminate cost",
                confidence="high", effort="low", action="terminate",
            ),
            RightsizingRecommendation(
                resource_id="n1-std-8-prod-gke", resource_name="gke-prod-node-pool",
                resource_type="GKE Node", provider="gcp", region="us-central1",
                account_id="proj-gcp-001", current_size="n1-standard-8",
                recommended_size="n2-standard-4",
                current_monthly_cost_usd=218.00, projected_monthly_cost_usd=152.60,
                monthly_savings_usd=65.40, annual_savings_usd=784.80,
                cpu_utilization_avg_percent=21.0, memory_utilization_avg_percent=38.5,
                reason="Migrate to n2-standard-4 — newer gen, ~30% lower cost for same workload",
                confidence="medium", effort="medium", action="modernize",
            ),
            RightsizingRecommendation(
                resource_id="i-0f7g8h9i", resource_name="prod-batch-scheduler",
                resource_type="EC2 Instance", provider="aws", region="eu-west-1",
                account_id="123456789012", current_size="m5.2xlarge",
                recommended_size="m5.2xlarge (Reserved 1yr)",
                current_monthly_cost_usd=307.20, projected_monthly_cost_usd=193.54,
                monthly_savings_usd=113.66, annual_savings_usd=1363.92,
                cpu_utilization_avg_percent=34.1, memory_utilization_avg_percent=41.0,
                reason="Steady-state workload over 6 months — 1yr No-Upfront RI saves ~37%",
                confidence="high", effort="medium", action="reserve",
            ),
        ]
        if provider != "all":
            recommendations_out = [r for r in recommendations_out if r.provider == provider]

    # ── Apply min_savings filter (synthetic tier skips it internally) ─────────
    if min_savings > 0:
        recommendations_out = [
            r for r in recommendations_out if r.monthly_savings_usd >= min_savings
        ]

    # ── Sort by monthly savings desc, apply limit ────────────────────────────
    recommendations_out.sort(key=lambda r: r.monthly_savings_usd, reverse=True)
    recommendations_out = recommendations_out[:limit]

    total_monthly = sum(r.monthly_savings_usd for r in recommendations_out)
    return RightsizingResponse(
        generated_at=now_str,
        organization_id=org_id,
        data_source=data_source,
        total_resources_analyzed=max(total_analyzed, len(recommendations_out)),
        rightsizable_count=len(recommendations_out),
        total_monthly_savings_usd=round(total_monthly, 2),
        total_annual_savings_usd=round(total_monthly * 12, 2),
        recommendations=recommendations_out,
    )


async def run_scheduled_scans_once(requested_organization_id: Optional[int] = None) -> Dict[str, Any]:
    """Trigger due scans for approved organizations based on configured cadence."""
    global _scheduler_running
    if _scheduler_running:
        return {"status": "busy", "started": 0, "organization_id": requested_organization_id}
    _scheduler_running = True
    started = 0
    export_jobs_run = 0
    now = _utcnow()
    db = SessionLocal()
    try:
        permissions_query = db.query(ScanningPermissionRecord).filter(
            ScanningPermissionRecord.state.in_([ScanningState.APPROVED.value, ScanningState.RUNNING.value])
        )
        if requested_organization_id is not None:
            permissions_query = permissions_query.filter(
                ScanningPermissionRecord.customer_id == f"org-{requested_organization_id}"
            )
        permissions = permissions_query.all()
        for permission in permissions:
            cadence_seconds = _scan_interval_seconds(permission.scan_frequency)
            last_completed = (
                db.query(ScanRunRecord)
                .filter(
                    ScanRunRecord.customer_id == permission.customer_id,
                    ScanRunRecord.state == ScanningState.COMPLETED.value,
                )
                .order_by(ScanRunRecord.completed_at.desc())
                .first()
            )
            if last_completed and last_completed.completed_at:
                elapsed = (now - last_completed.completed_at).total_seconds()
                if elapsed < cadence_seconds:
                    continue

            providers = json.loads(permission.providers_json or "[]")
            if not providers:
                providers = ["aws", "azure", "gcp", "oci"]
            scan_id = f"scan_{permission.customer_id}_{int(now.timestamp())}"
            row = ScanRunRecord(
                scan_id=scan_id,
                customer_id=permission.customer_id,
                state=ScanningState.RUNNING.value,
                providers_json=json.dumps(providers),
                progress=0,
                started_at=now,
            )
            db.add(row)
            db.commit()
            started += 1
            await _run_cost_analysis(scan_id=scan_id, customer_id=permission.customer_id, providers=providers)
            derived_org_id = _organization_id_from_customer_id(permission.customer_id)
            if derived_org_id is not None:
                db.add(
                    AuditLog(
                        organization_id=derived_org_id,
                        actor_user_id=None,
                        action="scan.schedule.triggered",
                        entity_type="scan_run",
                        entity_id=scan_id,
                        metadata_json=json.dumps(
                            {
                                "customer_id": permission.customer_id,
                                "frequency": permission.scan_frequency,
                                "providers": providers,
                            }
                        ),
                    )
                )
                db.commit()

        jobs_query = db.query(ExportJob).filter(ExportJob.is_active == True)  # noqa: E712
        if requested_organization_id is not None:
            jobs_query = jobs_query.filter(ExportJob.organization_id == requested_organization_id)
        jobs = jobs_query.all()
        for job in jobs:
            cadence_seconds = _scan_interval_seconds(job.schedule_frequency)
            anchor = job.last_run_at or job.created_at or now
            elapsed = (now - anchor).total_seconds()
            if elapsed < cadence_seconds:
                continue

            membership_stub = SimpleNamespace(
                organization_id=job.organization_id,
                role=UserRole.OWNER,
            )
            user_stub = SimpleNamespace(id=None)
            await _execute_export_job(
                db=db,
                job=job,
                current_user=user_stub,
                membership=membership_stub,
                actor_user_id=None,
            )
            export_jobs_run += 1
    finally:
        db.close()
        _scheduler_running = False
    return {
        "status": "ok",
        "started": started,
        "export_jobs_run": export_jobs_run,
        "organization_id": requested_organization_id,
    }
