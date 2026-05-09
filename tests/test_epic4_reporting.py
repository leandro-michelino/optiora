"""Epic 4: Reporting and Executive Outputs tests.

Covers:
- cost_period_summaries table presence and schema
- POST /api/v1/reports/period-summaries/compute
- GET /api/v1/reports/cost-trend (computed + raw-record fallback + empty)
- GET /api/v1/reports/chargeback.csv column headers
- GET /api/v1/reports/chargeback.xlsx (if openpyxl available)
- GET /api/v1/reports/executive-summary.xlsx (if openpyxl available)
- Enhanced executive summary CSV now includes Allocation Coverage section
"""

import io
import json
import os
import tempfile
import unittest
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

TEST_DB = os.path.join(tempfile.gettempdir(), "optiora_epic4_test.db")
os.environ.setdefault("DATABASE_URL", f"sqlite:///{TEST_DB}")
os.environ["ENABLE_AUTH"] = "true"
os.environ["SECRET_KEY"] = "test-epic4-secret"
os.environ["PASSWORD_RESET_RETURN_TOKEN"] = "true"

try:
    from fastapi.testclient import TestClient
    from sqlalchemy import inspect as sa_inspect

    from finops_mcp.app import app
    from finops_mcp.api import CostTrendPoint
    from finops_mcp.orm_models import (
        Base,
        CostPeriodSummary,
        ImportedCostRecord,
        NormalizedCostDimension,
        Organization,
        SessionLocal,
        User,
        UserOrganization,
        UserRole,
        ensure_public_workspace,
        engine,
    )
except ImportError as exc:
    raise unittest.SkipTest(f"Backend dependencies not installed: {exc}") from exc

try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _setup_db() -> None:
    engine.dispose()
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def _teardown_db() -> None:
    Base.metadata.drop_all(bind=engine)
    engine.dispose()


def _get_token(client: TestClient, email: str, password: str) -> str:
    resp = client.post("/auth/login", json={"email": email, "password": password})
    assert resp.status_code == 200, resp.text
    return resp.json()["access_token"]


def _register_admin(
    client: TestClient,
    email: str = "admin4@test.com",
    password: str = "Pass1234!",
) -> str:
    resp = client.post("/auth/register", json={"email": email, "password": password})
    assert resp.status_code in (200, 201, 409), resp.text
    token = _get_token(client, email, password)
    with SessionLocal() as db:
        u = db.query(User).filter_by(email=email).first()
        if u:
            uo = db.query(UserOrganization).filter_by(user_id=u.id).first()
            if uo:
                uo.role = UserRole.ADMIN
                db.commit()
    return token


IMPORT_CSV = """\
provider,service,region,cost_usd,period_start,tags
aws,EC2,us-east-1,300.00,2024-01-01,"{""team"":""platform""}"
aws,S3,us-east-1,80.00,2024-01-01,"{""team"":""data""}"
aws,RDS,us-east-1,120.00,2024-02-01,"{""team"":""platform""}"
gcp,BigQuery,us-central1,60.00,2024-02-01,"{""team"":""data""}"
"""


def _upload_csv(client: TestClient, token: str) -> None:
    resp = client.post(
        "/api/v1/imports/costs/csv",
        files={"file": ("costs.csv", IMPORT_CSV.encode(), "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code in (200, 201), resp.text


# ─────────────────────────────────────────────────────────────────────────────
# Test classes
# ─────────────────────────────────────────────────────────────────────────────

class CostPeriodSummarySchemaTest(unittest.TestCase):
    """Verify cost_period_summaries table is created with required columns."""

    def setUp(self):
        _setup_db()

    def tearDown(self):
        _teardown_db()

    def test_table_exists(self):
        inspector = sa_inspect(engine)
        self.assertIn("cost_period_summaries", inspector.get_table_names())

    def test_required_columns(self):
        inspector = sa_inspect(engine)
        columns = {c["name"] for c in inspector.get_columns("cost_period_summaries")}
        for col in (
            "id",
            "organization_id",
            "period_type",
            "period_start",
            "period_end",
            "provider",
            "total_cost_usd",
            "mapped_cost_usd",
            "unmapped_cost_usd",
            "record_count",
            "service_breakdown_json",
            "computed_at",
        ):
            self.assertIn(col, columns, f"Missing column: {col}")


class PeriodSummaryComputeTest(unittest.TestCase):
    """Test POST /api/v1/reports/period-summaries/compute."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_compute_returns_200(self):
        resp = self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=monthly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)

    def test_compute_response_shape(self):
        resp = self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=monthly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertIn("organization_id", data)
        self.assertIn("period_type", data)
        self.assertIn("periods_computed", data)
        self.assertIn("rows_written", data)
        self.assertIn("computed_at", data)
        self.assertGreater(data["rows_written"], 0)

    def test_compute_rows_persisted(self):
        resp = self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=monthly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        with SessionLocal() as db:
            count = db.query(CostPeriodSummary).count()
        self.assertGreater(count, 0)

    def test_compute_invalid_period_type(self):
        resp = self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=yearly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 400, resp.text)

    def test_compute_idempotent(self):
        """Running compute twice should not duplicate rows."""
        for _ in range(2):
            self.client.post(
                "/api/v1/reports/period-summaries/compute?period_type=monthly",
                headers={"Authorization": f"Bearer {self.token}"},
            )
        with SessionLocal() as db:
            rows1 = db.query(CostPeriodSummary).count()
        self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=monthly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        with SessionLocal() as db:
            rows2 = db.query(CostPeriodSummary).count()
        self.assertEqual(rows1, rows2, "Idempotent compute must not grow row count")


class CostTrendEndpointTest(unittest.TestCase):
    """Test GET /api/v1/reports/cost-trend."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_trend_empty_without_computed_rows(self):
        """Before compute, trend falls back to raw records or returns empty."""
        resp = self.client.get(
            "/api/v1/reports/cost-trend?period_type=monthly&lookback=6",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertIn("data_source", data)
        self.assertIn(data["data_source"], ("raw_records", "empty", "computed"))

    def test_live_required_trend_uses_current_live_cost_context_without_snapshots(self):
        live_required_config = SimpleNamespace(
            require_live_provider_data=True,
            retention_hot_months=3,
            retention_enabled=False,
            oci_archive_bucket="",
        )
        live_context = {
            "source": "live_provider_api",
            "total_cost": 42.5,
            "breakdown": {"oci": {"cost": 42.5, "percentage": 100.0}},
            "provider_errors": {},
        }

        with patch("finops_mcp.api.Config", return_value=live_required_config):
            with patch("finops_mcp.api._cost_context", new=AsyncMock(return_value=live_context)):
                resp = self.client.get(
                    "/api/v1/reports/cost-trend?period_type=monthly&lookback=3",
                    headers={"Authorization": f"Bearer {self.token}"},
                )

        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertEqual(data["data_source"], "live_provider_api_current_period")
        self.assertEqual(data["grand_total_usd"], 42.5)
        self.assertEqual(data["provider_totals"], {"oci": 42.5})
        self.assertEqual(len(data["points"]), 1)
        self.assertEqual(data["points"][0]["provider"], "oci")

    def test_trend_after_compute(self):
        """After compute, data_source should be 'computed'."""
        self.client.post(
            "/api/v1/reports/period-summaries/compute?period_type=monthly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        resp = self.client.get(
            "/api/v1/reports/cost-trend?period_type=monthly&lookback=6",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertEqual(data["data_source"], "computed")
        self.assertIsInstance(data["points"], list)
        self.assertGreater(len(data["points"]), 0)

    def test_trend_response_shape(self):
        resp = self.client.get(
            "/api/v1/reports/cost-trend?period_type=monthly&lookback=6",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        for key in ("organization_id", "period_type", "lookback_periods", "data_source", "points", "provider_totals", "grand_total_usd"):
            self.assertIn(key, data)
        if data["points"]:
            point = data["points"][0]
            for k in ("period_start", "period_end", "provider", "total_cost_usd", "mapped_cost_usd", "unmapped_cost_usd", "record_count"):
                self.assertIn(k, point)

    def test_trend_invalid_period_type(self):
        resp = self.client.get(
            "/api/v1/reports/cost-trend?period_type=quarterly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 400, resp.text)

    def test_trend_view_by_region(self):
        resp = self.client.get(
            "/api/v1/reports/cost-trend?period_type=monthly&lookback=6&view_by=region",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertEqual(data.get("view_by"), "region")
        self.assertIn("dimension_totals", data)


class ImportPreviewTest(unittest.TestCase):
    """Test CSV import validation preview with mapping feedback."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_preview_returns_mapping_feedback(self):
        resp = self.client.post(
            "/api/v1/imports/costs/preview",
            files={"file": ("costs.csv", IMPORT_CSV.encode(), "text/csv")},
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        data = resp.json()
        self.assertIn("mapping_feedback", data)
        self.assertIn("reconciliation_guidance", data)
        self.assertIn("issues", data)
        self.assertEqual(data.get("total_rows"), data.get("accepted_rows") + data.get("rejected_rows"))


class CostTrendPointModelTest(unittest.TestCase):
    def test_service_breakdown_defaults_are_isolated(self):
        left = CostTrendPoint(
            period_start="2026-05-01",
            period_end="2026-05-31",
            provider="aws",
            total_cost_usd=10.0,
            mapped_cost_usd=8.0,
            unmapped_cost_usd=2.0,
            record_count=1,
        )
        right = CostTrendPoint(
            period_start="2026-05-01",
            period_end="2026-05-31",
            provider="gcp",
            total_cost_usd=20.0,
            mapped_cost_usd=18.0,
            unmapped_cost_usd=2.0,
            record_count=1,
        )

        left.service_breakdown["compute"] = 10.0

        self.assertEqual(right.service_breakdown, {})


class ChargebackCsvExportTest(unittest.TestCase):
    """Test GET /api/v1/reports/chargeback.csv."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_csv_returns_200(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)

    def test_csv_content_type(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertIn("text/csv", resp.headers.get("content-type", ""))

    def test_csv_header_columns(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)
        first_line = resp.text.splitlines()[0]
        for col in ("provider", "service_name", "region", "cost_usd", "team", "environment", "is_mapped"):
            self.assertIn(col, first_line, f"Missing column '{col}' in chargeback CSV header")


class ChargebackXlsxExportTest(unittest.TestCase):
    """Test GET /api/v1/reports/chargeback.xlsx — skipped if openpyxl not available."""

    @classmethod
    def setUpClass(cls):
        if not _OPENPYXL:
            raise unittest.SkipTest("openpyxl not installed")
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)

    @classmethod
    def tearDownClass(cls):
        if _OPENPYXL:
            _teardown_db()

    def test_xlsx_returns_200(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text[:200])

    def test_xlsx_content_type(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        ct = resp.headers.get("content-type", "")
        self.assertIn("spreadsheet", ct, f"Expected XLSX content-type, got: {ct}")

    def test_xlsx_is_valid_workbook(self):
        resp = self.client.get(
            "/api/v1/reports/chargeback.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertGreater(len(wb.sheetnames), 0)
        self.assertIn("Chargeback Detail", wb.sheetnames)
        self.assertIn("Executive Summary", wb.sheetnames)


class ExecutiveSummaryXlsxTest(unittest.TestCase):
    """Test GET /api/v1/reports/executive-summary.xlsx — skipped if openpyxl not available."""

    @classmethod
    def setUpClass(cls):
        if not _OPENPYXL:
            raise unittest.SkipTest("openpyxl not installed")
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)

    @classmethod
    def tearDownClass(cls):
        if _OPENPYXL:
            _teardown_db()

    def test_xlsx_returns_200(self):
        resp = self.client.get(
            "/api/v1/reports/executive-summary.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text[:200])

    def test_xlsx_valid_workbook(self):
        resp = self.client.get(
            "/api/v1/reports/executive-summary.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertIn("Executive Summary", wb.sheetnames)
        self.assertIn("Trend by Provider", wb.sheetnames)
        self.assertIn("Trend by Region", wb.sheetnames)

    def test_xlsx_has_header_row(self):
        resp = self.client.get(
            "/api/v1/reports/executive-summary.xlsx",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb["Executive Summary"]
        first_row = [ws.cell(1, c).value for c in range(1, 4)]
        self.assertEqual(first_row, ["Section", "Field", "Value"])


class PdfDigestAndSharingTest(unittest.TestCase):
    """Test weekly/monthly PDF digest and tokenized read-only report sharing."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_pdf_digest_download(self):
        resp = self.client.get(
            "/api/v1/reports/executive-digest.pdf?frequency=weekly",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text[:200])
        self.assertIn("application/pdf", resp.headers.get("content-type", ""))

    def test_tokenized_read_only_share(self):
        create = self.client.post(
            "/api/v1/reports/share-token",
            json={"report_type": "executive_summary", "report_format": "json", "expires_in_hours": 2},
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(create.status_code, 200, create.text)
        token = create.json()["token"]
        shared = self.client.get(f"/api/v1/reports/shared/{token}")
        self.assertEqual(shared.status_code, 200, shared.text)
        self.assertIn("application/json", shared.headers.get("content-type", ""))


class EnhancedExecutiveSummaryCsvTest(unittest.TestCase):
    """Verify the enhanced executive summary CSV includes chargeback/allocation rows."""

    @classmethod
    def setUpClass(cls):
        _setup_db()
        cls.client = TestClient(app, raise_server_exceptions=False)
        cls.token = _register_admin(cls.client)
        _upload_csv(cls.client, cls.token)
        # Apply mapping rules so allocation rows appear
        cls.client.post(
            "/api/v1/business-mapping/apply",
            headers={"Authorization": f"Bearer {cls.token}"},
        )

    @classmethod
    def tearDownClass(cls):
        _teardown_db()

    def test_csv_returns_200(self):
        resp = self.client.get(
            "/api/v1/reports/executive-summary.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertEqual(resp.status_code, 200, resp.text)

    def test_csv_has_summary_section(self):
        resp = self.client.get(
            "/api/v1/reports/executive-summary.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertIn("Summary", resp.text)

    def test_csv_has_allocation_section_when_dims_exist(self):
        """If NormalizedCostDimension rows were written, Allocation Coverage section appears."""
        with SessionLocal() as db:
            has_dims = db.query(NormalizedCostDimension).count() > 0
        if not has_dims:
            self.skipTest("No NormalizedCostDimension rows — allocation section not expected")
        resp = self.client.get(
            "/api/v1/reports/executive-summary.csv",
            headers={"Authorization": f"Bearer {self.token}"},
        )
        self.assertIn("Allocation Coverage", resp.text)


if __name__ == "__main__":
    unittest.main()
